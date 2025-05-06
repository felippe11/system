import os
import uuid
import csv
import pytz
from flask import Response
from datetime import datetime
import logging
import pandas as pd
import qrcode
import requests
import mercadopago
from flask import abort
from sqlalchemy.exc import IntegrityError
from flask import send_from_directory
from models import Ministrante
from models import Cliente
from flask import (Flask, Blueprint, render_template, redirect, url_for, flash,
                   request, jsonify, send_file, session)
from extensions import socketio
from models import TrabalhoCientifico
from flask_socketio import emit, join_room, leave_room
from utils import formatar_brasilia
from sqlalchemy import or_
from PIL import Image, ImageDraw, ImageFont





# routes.py (no início do arquivo)
from models import CampoPersonalizadoCadastro

from flask_login import login_user, logout_user, login_required, current_user
from flask_login import LoginManager
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from models import RespostaFormulario, RespostaCampo
from utils import enviar_email
from utils import gerar_certificado_personalizado
from datetime import datetime
from flask_mail import Message
from models import Evento, EventoInscricaoTipo, LoteInscricao, LoteTipoInscricao
from flask import current_app
from models import FormularioTemplate
from models import CampoFormulario
from models import Formulario
from models import CampoFormularioTemplate
from models import CertificadoTemplate
from utils import gerar_certificados_pdf

# Adicione esta linha na seção de importações no topo do arquivo routes.py
from sqlalchemy import and_, func, or_
from dateutil import parser  # NOVA IMPORTAÇÃO
from models import (
    ConfiguracaoAgendamento, 
    SalaVisitacao, 
    HorarioVisitacao, 
    AgendamentoVisita, 
    AlunoVisitante, 
    ProfessorBloqueado,
    RegraInscricaoEvento
)

from models import Evento
from flask import request, jsonify, flash, redirect, url_for
from flask import Response, send_file, request
from io import StringIO, BytesIO
import csv
from openpyxl import Workbook
from models import Patrocinador



# Extensões e modelos (utilize sempre o mesmo ponto de importação para o db)
from extensions import db, login_manager
from models import (Usuario, Oficina, Inscricao, OficinaDia, Checkin,
                    Configuracao, Feedback, Ministrante, RelatorioOficina, MaterialOficina, ConfiguracaoCliente, FeedbackCampo, RespostaFormulario, RespostaCampo, InscricaoTipo)
from utils import obter_estados, obter_cidades, gerar_qr_code, gerar_qr_code_inscricao, gerar_comprovante_pdf, gerar_etiquetas_pdf  # Funções auxiliares
from reportlab.lib.units import inch
from reportlab.platypus import Image

from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image, PageBreak
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from models import LinkCadastro
from flask import Blueprint, request, render_template, redirect, url_for, flash, jsonify
from models import Formulario, CampoFormulario
from extensions import db
from flask_login import login_required, current_user
from collections import defaultdict
from datetime import datetime
from markupsafe import Markup
from jinja2 import Environment, FileSystemLoader


# ReportLab para PDFs
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, landscape, A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, LongTable
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.utils import ImageReader
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics

sdk = mercadopago.SDK(os.getenv("MERCADOPAGO_ACCESS_TOKEN"))



# Registrar a fonte personalizada
pdfmetrics.registerFont(TTFont("AlexBrush", "AlexBrush-Regular.ttf"))



# Configurações da aplicação e Blueprint
app = Flask(__name__)
app.config["UPLOAD_FOLDER"] = "uploads"
ALLOWED_EXTENSIONS = {"xlsx"}

# Inicialize o LoginManager com o app e defina a rota de login.
login_manager.init_app(app)
login_manager.login_view = 'routes.login'  # Essa é a rota que será usada para login

routes = Blueprint("routes", __name__)

# routes.py, logo após criar o Blueprint
@routes.before_request
def bloquear_usuarios_pendentes():

    # Todas as restrições de pagamento foram removidas
    # Agora todos os usuários podem acessar todas as funcionalidades, independentemente do status de pagamento
    # Permitir que participantes com pagamento pendente tenham acesso a todas as funcionalidades.
    # Código de bloqueio removido para permitir acesso irrestrito.

    return

       
def register_routes(app):
    app.register_blueprint(routes)


# ===========================
#        ROTAS GERAIS
# ===========================
@routes.route('/')
def home():
    try:
        # Buscar todos os eventos ativos e públicos, ordenados pela data de início
        eventos_destaque = Evento.query.filter(
            Evento.status == 'ativo',
            Evento.publico == True,
            Evento.data_inicio >= datetime.now()  # apenas eventos futuros
        ).order_by(Evento.data_inicio.asc()).all()
        
        # Converter os eventos para um formato que pode ser enviado ao template
        eventos_json = []
        for evento in eventos_destaque:
            evento_data = {
                'id': evento.id,
                'nome': evento.nome,
                'descricao': evento.descricao,
                'data_inicio': evento.data_inicio.strftime('%d/%m/%Y') if evento.data_inicio else 'Data a definir',
                'data_fim': evento.data_fim.strftime('%d/%m/%Y') if evento.data_fim else '',
                'localizacao': evento.localizacao or 'Local a definir',
                'banner_url': evento.banner_url or url_for('static', filename='images/event-placeholder.jpg'),
                'preco_base': 0  # valor padrão
            }
            
            # Verifica se há tipos de inscrição e pega o menor preço
            if evento.tipos_inscricao and len(evento.tipos_inscricao) > 0:
                evento_data['preco_base'] = min(tipo.preco for tipo in evento.tipos_inscricao)
            
            eventos_json.append(evento_data)
        
        return render_template('index.html', eventos_destaque=eventos_json)
    
    except Exception as e:
        # Log do erro (implemente seu sistema de logs)
        print(f"Erro na rota home: {str(e)}")
        # Retorna uma lista vazia em caso de erro
        return render_template('index.html', eventos_destaque=[])

@routes.route('/api/eventos/destaque')
def get_eventos_destaque():
    try:
        eventos = Evento.query.filter(
            Evento.data_inicio >= datetime.now(),
            Evento.status == 'ativo',
            Evento.publico == True  # adicionado filtro de eventos públicos
        ).order_by(Evento.data_inicio.asc()).limit(5).all()
        
        eventos_json = []
        for evento in eventos:
            evento_data = {
                'id': evento.id,
                'nome': evento.nome,
                'descricao': evento.descricao,
                'data_inicio': evento.data_inicio.strftime('%d/%m/%Y') if evento.data_inicio else 'Data a definir',
                'data_fim': evento.data_fim.strftime('%d/%m/%Y') if evento.data_fim else '',
                'localizacao': evento.localizacao or 'Local a definir',
                'banner_url': evento.banner_url or url_for('static', filename='images/event-placeholder.jpg'),
                'preco_base': 0  # valor padrão
            }
            
            if evento.tipos_inscricao and len(evento.tipos_inscricao) > 0:
                evento_data['preco_base'] = min(tipo.preco for tipo in evento.tipos_inscricao)
            
            eventos_json.append(evento_data)
        
        return jsonify(eventos_json)
    
    except Exception as e:
        print(f"Erro em get_eventos_destaque: {str(e)}")
        return jsonify([])

@routes.route('/eventos')
def listar_eventos():
    try:
        # Lista completa de eventos
        eventos = Evento.query.filter(
            Evento.data_inicio >= datetime.now(),
            Evento.status == 'ativo',
            Evento.publico == True  # adicionado filtro de eventos públicos
        ).order_by(Evento.data_inicio.asc()).all()
        
        # Processa os eventos para o template
        eventos_processed = []
        for evento in eventos:
            evento_dict = {
                'id': evento.id,
                'nome': evento.nome,
                # adicione outros campos necessários para o template eventos.html
                'data_inicio': evento.data_inicio.strftime('%d/%m/%Y') if evento.data_inicio else 'Data a definir',
                'localizacao': evento.localizacao or 'Local a definir',
                'preco_base': 0
            }
            
            if evento.tipos_inscricao and len(evento.tipos_inscricao) > 0:
                evento_dict['preco_base'] = min(tipo.preco for tipo in evento.tipos_inscricao)
            
            eventos_processed.append(evento_dict)
        
        return render_template('eventos.html', eventos=eventos_processed)
    
    except Exception as e:
        print(f"Erro em listar_eventos: {str(e)}")
        return render_template('eventos.html', eventos=[])
    
@routes.route('/api/placeholder/<int:width>/<int:height>')
def placeholder_image(width, height):
    try:
        # Crie uma imagem placeholder
        img = Image.new('RGB', (width, height), color=(200, 200, 200))
        d = ImageDraw.Draw(img)
        
        # Tente carregar a fonte Arial ou use a fonte padrão se não estiver disponível
        try:
            font = ImageFont.truetype("arial.ttf", 30)
        except IOError:
            # Fallback para uma fonte que provavelmente existe no sistema
            try:
                font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 30)
            except IOError:
                # Se nenhuma fonte específica estiver disponível, use a fonte padrão
                font = ImageFont.load_default()
        
        # Desenhe o texto na imagem
        text = f"{width} x {height}"
        text_width, text_height = d.textsize(text, font=font) if hasattr(d, 'textsize') else (100, 30)
        d.text(((width-text_width)//2, (height-text_height)//2), text, fill=(80, 80, 80), font=font)
        
        # Salve a imagem em memória e a envie como resposta
        img_io = io.BytesIO()
        img.save(img_io, 'JPEG', quality=70)
        img_io.seek(0)
        return send_file(img_io, mimetype='image/jpeg')
    except Exception as e:
        # Em caso de erro, retorne uma imagem padrão muito simples
        app.logger.error(f"Erro ao gerar imagem placeholder: {str(e)}")
        img = Image.new('RGB', (width, height), color=(200, 200, 200))
        img_io = io.BytesIO()
        img.save(img_io, 'JPEG', quality=70)
        img_io.seek(0)
        return send_file(img_io, mimetype='image/jpeg')

@routes.route('/enviar_proposta', methods=['POST'])
def enviar_proposta():
    nome = request.form.get('nome')
    email = request.form.get('email')
    tipo_evento = request.form.get('tipo_evento')
    descricao = request.form.get('descricao')

    if not all([nome, email, tipo_evento, descricao]):
        flash('Por favor, preencha todos os campos.', 'danger')
        return redirect(url_for('routes.home'))

    nova_proposta = Proposta(
        nome=nome,
        email=email,
        tipo_evento=tipo_evento,
        descricao=descricao
    )

    try:
        db.session.add(nova_proposta)
        db.session.commit()
        flash('Proposta enviada com sucesso! Entraremos em contato em breve.', 'success')
    except Exception as e:
        db.session.rollback()
        flash('Erro ao enviar proposta. Por favor, tente novamente.', 'danger')

    return render_template(
        'index.html',
        nome=nome,
        email=email,
        tipo_evento=tipo_evento,
        descricao=descricao
    )



# ===========================
#    CADASTRO DE PARTICIPANTE
# ===========================
@routes.route('/cadastro_participante', methods=['GET', 'POST'])
@routes.route('/inscricao/<path:identifier>', methods=['GET', 'POST'])  # Aceita slug ou token
def cadastro_participante(identifier: str | None = None):
    """Fluxo completo de cadastro de participante com tratamento robusto de lotes e pagamento Mercado Pago.
    
    1. Valida link (token ou slug)
    2. Verifica lote vigente com validação robusta
    3. Renderiza formulário (GET)
    4. Cria Usuario, Inscricao (status **pending**)
    5. Se gratuito → approved → login
       Se pago → gera preference → redireciona init_point
    """
    # ────────────────────────────── Imports locais ─────────────────────────────
    from collections import defaultdict
    from datetime import datetime
    from dateutil import parser
    import os, mercadopago, logging
    
    # Configuração de logging
    logging.basicConfig(level=logging.DEBUG)
    logger = logging.getLogger(__name__)
    
    # ────────────────────────────── Verificação do SDK MP ─────────────────────────────
    # Verificação do token do Mercado Pago
    mp_token = os.getenv("MERCADOPAGO_ACCESS_TOKEN")
    if not mp_token:
        logger.error("Token do Mercado Pago não configurado")
        flash("Erro no sistema de pagamento", "danger")
        return redirect(url_for("routes.login"))

    sdk = mercadopago.SDK(mp_token)
    
    # ────────────────────────────── Validação do link ─────────────────────────────
    token = request.args.get("token") if not identifier else identifier
    link = None
    if token:
        link = LinkCadastro.query.filter(
            (LinkCadastro.token == token) | (LinkCadastro.slug_customizado == token)
        ).first()
    
    if not link:
        flash("Link de cadastro inválido ou expirado", "danger")
        return redirect(url_for("routes.login"))

    cliente_id = link.cliente_id
    evento = Evento.query.get(link.evento_id) if link.evento_id else None

    # ──────────────────── VERIFICAÇÃO DE LOTE VIGENTE (VERSÃO ROBUSTA) ────────────────
    def lote_disponivel(lote: LoteInscricao) -> bool:
        """Verifica se um lote ainda está disponível para inscrição"""
        try:
            now = datetime.utcnow()
            
            # Verificação de datas
            if lote.data_inicio and lote.data_fim:
                inicio = lote.data_inicio if isinstance(lote.data_inicio, datetime) else parser.parse(lote.data_inicio)
                fim = lote.data_fim if isinstance(lote.data_fim, datetime) else parser.parse(lote.data_fim)
                if not (inicio <= now <= fim):
                    return False
                    
            # Verificação de quantidade
            if lote.qtd_maxima is not None:
                count = db.session.query(func.count(Inscricao.id)).filter(
                    Inscricao.lote_id == lote.id,
                    Inscricao.status_pagamento.in_(["approved", "pending"])
                ).scalar()
                if count >= lote.qtd_maxima:
                    return False
                    
            return True
            
        except Exception as e:
            logger.error(f"Erro ao verificar disponibilidade do lote {lote.id}: {str(e)}")
            return False
    
    lote_vigente = None
    lotes_ativos = []
    if evento and evento.habilitar_lotes:
        lotes_ativos = LoteInscricao.query.filter_by(
            evento_id=evento.id, 
            ativo=True
        ).order_by(LoteInscricao.ordem).all()
        
        # Debug de status dos lotes
        for lote in lotes_ativos:
            count = db.session.query(func.count(Inscricao.id)).filter(
                Inscricao.evento_id == evento.id,
                Inscricao.lote_id == lote.id,
                Inscricao.status_pagamento.in_(["approved", "pending"])
            ).scalar()
            
            logger.debug(f"Lote: {lote.nome}, Máximo: {lote.qtd_maxima}, Contagem: {count}, Ativo: {lote.ativo}")
            
            # Verificação robusta do lote
            try:
                # Contagem precisa de inscrições válidas
                valid_date = True
                if lote.data_inicio and lote.data_fim:
                    inicio = lote.data_inicio if isinstance(lote.data_inicio, datetime) else parser.parse(lote.data_inicio)
                    fim = lote.data_fim if isinstance(lote.data_fim, datetime) else parser.parse(lote.data_fim)
                    valid_date = inicio <= datetime.utcnow() <= fim
                    if not valid_date:
                        logger.debug(f"  - Inválido por data: início={inicio}, fim={fim}, agora={datetime.utcnow()}")

                # Verificação de vagas
                valid_qtd = True
                if lote.qtd_maxima is not None:
                    valid_qtd = count < lote.qtd_maxima
                    if not valid_qtd:
                        logger.debug(f"  - Inválido por quantidade: {count}/{lote.qtd_maxima}")

                if valid_date and valid_qtd:
                    lote_vigente = lote
                    logger.debug(f"  => Lote vigente: {lote.nome}")
                    break

            except Exception as e:
                logger.error(f"Erro ao verificar lote {lote.id}: {str(e)}")
                continue

        if not lote_vigente and lotes_ativos:
            logger.warning("Nenhum lote válido encontrado!")
            flash("Nenhum lote disponível no momento", "warning")
            return redirect(url_for("routes.login"))

    # ──────────────────── PREPARAÇÃO DE DADOS PARA O TEMPLATE ────────────────
    oficinas = Oficina.query.filter_by(evento_id=evento.id).all() if evento else []
    
    # Carrega ministrantes
    ministrantes = (
        Ministrante.query.join(Oficina).filter(Oficina.evento_id == evento.id).distinct().all() if evento else []
    )
    
    # Agrupamento de oficinas por dia
    grouped_oficinas = defaultdict(list)
    for oficina in oficinas:
        for dia in oficina.dias:
            data_str = dia.data.strftime("%d/%m/%Y")
            grouped_oficinas[data_str].append({
                "oficina": oficina,
                "titulo": oficina.titulo,
                "descricao": oficina.descricao,
                "ministrante": oficina.ministrante_obj,
                "horario_inicio": dia.horario_inicio,
                "horario_fim": dia.horario_fim,
            })
    
    # Ordenação das oficinas por horário e das datas
    for lst in grouped_oficinas.values():
        lst.sort(key=lambda x: x["horario_inicio"])
    sorted_keys = sorted(grouped_oficinas.keys(), key=lambda d: datetime.strptime(d, "%d/%m/%Y"))
    
    # Garantir que todas as oficinas estejam devidamente processadas
    oficinas_formatadas = []
    for idx, oficina in enumerate(oficinas):
        # Determinar as informações do evento para a oficina
        if oficina.evento_id:
            # Obter informações do evento diretamente
            evento_obj = Evento.query.get(oficina.evento_id)
            if evento_obj:
                evento_nome = evento_obj.nome
                # Determinar status do evento
                data_atual = datetime.now().date()
                ja_ocorreu = False
                eh_futuro = False
                
                if evento_obj.data_fim:
                    if hasattr(evento_obj.data_fim, 'date'):
                        data_fim = evento_obj.data_fim.date()
                    else:
                        data_fim = evento_obj.data_fim
                    ja_ocorreu = data_fim < data_atual
                    
                if evento_obj.data_inicio:
                    if hasattr(evento_obj.data_inicio, 'date'):
                        data_inicio = evento_obj.data_inicio.date()
                    else:
                        data_inicio = evento_obj.data_inicio
                    eh_futuro = data_inicio > data_atual
                
                evento_status = 'Futuro' if eh_futuro else ('Encerrado' if ja_ocorreu else 'Atual')
            else:
                evento_nome = "Evento Desconhecido"
                evento_status = "Atual"
        else:
            evento_nome = "Atividades Gerais"
            evento_status = "Atuais"
            
        oficinas_formatadas.append({
            'id': oficina.id,
            'titulo': oficina.titulo,
            'descricao': oficina.descricao,
            'evento_id': oficina.evento_id if oficina.evento_id else 'sem_evento',
            'evento_nome': evento_nome,
            'evento_status': evento_status,
        })
    
    # Patrocinadores e campos de cadastro
    patrocinadores = Patrocinador.query.filter_by(evento_id=evento.id).all() if evento else []
    campos_personalizados = CampoPersonalizadoCadastro.query.filter_by(cliente_id=cliente_id).all()
    
    # Garante lista completa de ministrantes (principal + extras)
    all_ministrantes = set()
    for of in oficinas:
        if of.ministrante_obj:
            all_ministrantes.add(of.ministrante_obj)
        if hasattr(of, 'ministrantes_associados'):
            all_ministrantes.update(of.ministrantes_associados)

    # ──────────────────── TRATAMENTO DO FORMULÁRIO (POST) ────────────────
    if request.method == "POST":
        # Debug de todos os campos do formulário
        logger.debug(f"Dados do formulário: {request.form}")
        
        # Dados básicos
        nome = request.form.get("nome")
        cpf = request.form.get("cpf")
        email = request.form.get("email")
        senha = request.form.get("senha")
        formacao = request.form.get("formacao")
        
        # Verifica se está usando lotes ou tipo de inscrição direto
        lote_tipo_inscricao_id = request.form.get("lote_tipo_inscricao_id")
        tipo_inscricao_id = request.form.get("tipo_inscricao_id")
        lote_id = request.form.get("lote_id")
        
        logger.debug(f"IDs recebidos: lote_tipo_inscricao_id={lote_tipo_inscricao_id}, tipo_inscricao_id={tipo_inscricao_id}, lote_id={lote_id}")
        
        estados_str = ",".join(request.form.getlist("estados[]") or [])
        cidades_str = ",".join(request.form.getlist("cidades[]") or [])

        # Validação mínima
        if not all([nome, cpf, email, senha, formacao]):
            flash("Todos os campos obrigatórios devem ser preenchidos!", "danger")
            return render_template(
                "cadastro_participante.html",
                token=link.token,
                evento=evento,
                sorted_keys=sorted_keys,
                ministrantes=ministrantes,
                grouped_oficinas=grouped_oficinas,
                lote_vigente=lote_vigente,
                patrocinadores=patrocinadores,
                campos_personalizados=campos_personalizados
            )

        if Usuario.query.filter_by(email=email).first():
            flash("Este e-mail já está cadastrado!", "danger")
            return redirect(request.url)
        if Usuario.query.filter_by(cpf=cpf).first():
            flash("CPF já está sendo usado por outro usuário!", "danger")
            return redirect(request.url)

        # ────────────────────────────── VERIFICAÇÃO MAIS ROBUSTA DO LOTE ─────────────────────────────
        # Verificar se o lote informado ainda é válido (se estiver usando lotes)
        if lote_id and evento and evento.habilitar_lotes:
            # Verificar o lote novamente para ter certeza
            lotes_validos = []
            lote_atual = None
            
            # Percorrer todos os lotes para encontrar o lote vigente atual
            lotes = LoteInscricao.query.filter_by(evento_id=evento.id, ativo=True).order_by(LoteInscricao.ordem).all()
            for l in lotes:
                # Verificação explícita da validade
                if lote_disponivel(l):
                    lotes_validos.append(l)
                    
                if l.id == int(lote_id):
                    lote_atual = l
            
            # Se o lote enviado não for mais válido
            if not lote_atual or lote_atual not in lotes_validos:
                logger.warning(f"Lote {lote_id} não é mais válido!")
                
                if lotes_validos:
                    # Usar o primeiro lote válido
                    lote_vigente = lotes_validos[0]
                    lote_id = lote_vigente.id
                    
                    # Atualizar o lote_tipo_inscricao_id se possível
                    if tipo_inscricao_id:
                        novo_lote_tipo = LoteTipoInscricao.query.filter_by(
                            lote_id=lote_vigente.id,
                            tipo_inscricao_id=tipo_inscricao_id
                        ).first()
                        
                        if novo_lote_tipo:
                            lote_tipo_inscricao_id = novo_lote_tipo.id
                            logger.info(f"Tipo de inscrição atualizado para o novo lote: {lote_vigente.nome}")
                        else:
                            # Usar o primeiro tipo disponível no novo lote
                            primeiro_lote_tipo = LoteTipoInscricao.query.filter_by(
                                lote_id=lote_vigente.id
                            ).first()
                            
                            if primeiro_lote_tipo:
                                lote_tipo_inscricao_id = primeiro_lote_tipo.id
                                tipo_inscricao_id = primeiro_lote_tipo.tipo_inscricao_id
                                logger.info(f"Usando primeiro tipo disponível no novo lote: {lote_vigente.nome}")
                            else:
                                flash("Não foi possível encontrar um tipo de inscrição válido.", "danger")
                                return redirect(request.url)
                    
                    flash(f"O lote foi atualizado para '{lote_vigente.nome}' pois o lote anterior não está mais disponível.", "info")
                else:
                    flash("O lote selecionado não está mais disponível", "danger")
                    return redirect(request.url)

        # Cria usuário
        novo_usuario = Usuario(
            nome=nome,
            cpf=cpf,
            email=email,
            senha=generate_password_hash(senha),
            formacao=formacao,
            tipo="participante",
            estados=estados_str,
            cidades=cidades_str,
            cliente_id=cliente_id,
            tipo_inscricao_id=tipo_inscricao_id,
            evento_id=link.evento_id,
        )
        db.session.add(novo_usuario)
        db.session.flush()  # garante id

        # Campos personalizados
        for campo in campos_personalizados:
            valor = request.form.get(f"campo_{campo.id}")
            if campo.obrigatorio and not valor:
                db.session.rollback()
                flash(f"O campo '{campo.nome}' é obrigatório.", "danger")
                return redirect(request.url)
            try:
                db.session.add(
                    RespostaCampo(
                        resposta_formulario_id=novo_usuario.id,
                        campo_id=campo.id,
                        valor=valor or "",
                    )
                )
            except Exception as e:
                logger.error(f"Erro ao salvar campo personalizado: {str(e)}")
                pass  # continua mesmo se der erro nos campos personalizados

        # Cria inscrição pendente
        inscricao = Inscricao(
            usuario_id=novo_usuario.id,
            evento_id=evento.id if evento else None,
            cliente_id=cliente_id,
            status_pagamento="pending",
        )
        
        # Adiciona o lote_id se estiver disponível
        if lote_id:
            inscricao.lote_id = lote_id
            logger.debug(f"Definindo lote_id da inscrição: {lote_id}")
            
        db.session.add(inscricao)
        db.session.flush()  # Garante que a inscrição tenha um ID

        # Determina o preço baseado no tipo de inscrição usado
        preco = 0
        titulo_inscricao = "Inscrição"
        
        # Se estiver usando lotes
        if lote_tipo_inscricao_id and evento and evento.habilitar_lotes:
            logger.debug(f"Usando preço do lote_tipo_inscricao_id={lote_tipo_inscricao_id}")
            lote_tipo = LoteTipoInscricao.query.get(lote_tipo_inscricao_id)
            
            if lote_tipo:
                preco = float(lote_tipo.preco)
                tipo_inscricao = EventoInscricaoTipo.query.get(lote_tipo.tipo_inscricao_id)
                if tipo_inscricao:
                    titulo_inscricao = f"Inscrição – {evento.nome} – {tipo_inscricao.nome} ({lote_tipo.lote.nome})"
                    # Atualizamos também o tipo_inscricao_id do usuário
                    novo_usuario.tipo_inscricao_id = tipo_inscricao.id
        # Se estiver usando tipo de inscrição direto
        elif tipo_inscricao_id:
            logger.debug(f"Usando preço do tipo_inscricao_id={tipo_inscricao_id}")
            tipo_inscricao = EventoInscricaoTipo.query.get(tipo_inscricao_id)
            if tipo_inscricao:
                preco = float(tipo_inscricao.preco)
                titulo_inscricao = f"Inscrição – {evento.nome} – {tipo_inscricao.nome}"
        
        # Verifica se inscrição é gratuita pelo evento
        if evento and evento.inscricao_gratuita:
            logger.debug("Evento com inscrição gratuita, ignorando preço")
            preco = 0

        logger.debug(f"Preço calculado: {preco}, Título: {titulo_inscricao}")

        # Preferência de pagamento se for pago
        if preco > 0:
            logger.debug("Criando preferência de pagamento no Mercado Pago")
            # Preparar URLs para o Mercado Pago usando URL absoluta
            success_url = url_for("routes.pagamento_sucesso", _external=True)
            failure_url = url_for("routes.pagamento_falha", _external=True)
            pending_url = url_for("routes.pagamento_pendente", _external=True)
            webhook_url = url_for("routes.webhook_mp", _external=True)
            
            logger.debug(f"URLs de callback: Success={success_url}, Failure={failure_url}, Pending={pending_url}")
            logger.debug(f"Webhook URL: {webhook_url}")
            
            preference_data = {
                "items": [
                    {
                        "id": str(lote_tipo_inscricao_id or tipo_inscricao_id or inscricao.id),
                        "title": titulo_inscricao,
                        "quantity": 1,
                        "currency_id": "BRL",
                        "unit_price": preco,
                    }
                ],
                "payer": {"email": novo_usuario.email, "name": novo_usuario.nome},
                "external_reference": str(inscricao.id),
                "back_urls": {
                    "success": success_url,
                    "failure": failure_url,
                    "pending": pending_url,
                },
                "notification_url": webhook_url,
                "auto_return": "approved",
            }
            
            logger.debug(f"Dados da preferência: {preference_data}")
            
            try:
                pref = sdk.preference().create(preference_data)
                logger.debug(f"Resposta do Mercado Pago: {pref}")
                
                if "response" in pref and "init_point" in pref["response"]:
                    # Salvar o ID do pagamento na inscrição
                    if "id" in pref["response"]:
                        inscricao.payment_id = pref["response"]["id"]
                    
                    db.session.commit()
                    
                    # Extrair a URL completa do init_point
                    mp_url = pref["response"]["init_point"]
                    logger.debug(f"URL do Mercado Pago: {mp_url}")
                    
                    # Redirecionar diretamente para a URL do Mercado Pago
                    return redirect(mp_url)
                else:
                    # Erro na criação da preferência
                    logger.error(f"Erro na resposta do Mercado Pago: {pref}")
                    db.session.rollback()
                    flash("Erro ao processar pagamento: resposta inválida do Mercado Pago.", "danger")
                    return redirect(request.url)
            except Exception as e:
                logger.exception(f"Exceção ao criar preferência: {str(e)}")
                db.session.rollback()
                flash(f"Erro ao processar pagamento: {str(e)}", "danger")
                return redirect(request.url)

        # Gratuito – aprova direto
        logger.debug("Inscrição gratuita - aprovando diretamente")
        inscricao.status_pagamento = "approved"
        db.session.commit()
        flash("Cadastro realizado com sucesso!", "success")
        return redirect(url_for("routes.login"))

    # ──────────────────── PREPARAÇÃO DO RESPONSE (GET) ────────────────
    # Estatísticas do lote vigente
    lote_stats = None
    if lote_vigente:
        count = db.session.query(func.count(Inscricao.id)).filter(
            Inscricao.evento_id == evento.id,
            Inscricao.lote_id == lote_vigente.id,
            Inscricao.status_pagamento.in_(["approved", "pending"])
        ).scalar()
        
        vagas_disponiveis = lote_vigente.qtd_maxima - count if lote_vigente.qtd_maxima else "ilimitado"
        
        lote_stats = {
            "nome": lote_vigente.nome,
            "vagas_totais": lote_vigente.qtd_maxima or "ilimitado",
            "vagas_usadas": count,
            "vagas_disponiveis": vagas_disponiveis,
            "data_inicio": lote_vigente.data_inicio.strftime("%d/%m/%Y") if lote_vigente.data_inicio else "Imediato",
            "data_fim": lote_vigente.data_fim.strftime("%d/%m/%Y") if lote_vigente.data_fim else "Não definido"
        }
        logger.debug(f"Estatísticas do lote: {lote_stats}")
        
         
    # Renderiza o template com os dados necessários
    return render_template(
        "cadastro_participante.html",
        token=link.token,
        evento=evento,
        sorted_keys=sorted_keys,
        grouped_oficinas=grouped_oficinas,
        ministrantes=list(all_ministrantes) if all_ministrantes else ministrantes,
        patrocinadores=patrocinadores,
        campos_personalizados=campos_personalizados,
        lote_vigente=lote_vigente,
        lote_stats=lote_stats,
        lotes_ativos=lotes_ativos,
        tipos_inscricao=EventoInscricaoTipo.query.filter_by(evento_id=evento.id).all() if evento else []
    )
   
    
@routes.route('/api/lote_vigente/<int:evento_id>', methods=['GET'])
def api_lote_vigente(evento_id):
    """API para verificar o lote vigente atual com tratamento robusto de datas"""
    from datetime import datetime
    import logging
    from dateutil import parser
    
    logger = logging.getLogger(__name__)
    now = datetime.utcnow()
    
    try:
        lotes = LoteInscricao.query.filter_by(evento_id=evento_id, ativo=True).order_by(LoteInscricao.ordem).all()
        lote_vigente = None
        
        for lote in lotes:
            # Contagem de inscrições no lote
            count = Inscricao.query.filter_by(
                evento_id=evento_id, 
                lote_id=lote.id
            ).filter(
                Inscricao.status_pagamento.in_(["approved", "pending"])
            ).count()
            
            # Verificação explícita da validade
            is_valid = True
            
            # Verificação robusta de datas
            if lote.data_inicio and lote.data_fim:
                data_inicio = lote.data_inicio if isinstance(lote.data_inicio, datetime) else parser.parse(lote.data_inicio)
                data_fim = lote.data_fim if isinstance(lote.data_fim, datetime) else parser.parse(lote.data_fim)
                
                if now < data_inicio or now > data_fim:
                    is_valid = False
            
            # Verificação por quantidade
            if lote.qtd_maxima is not None and count >= lote.qtd_maxima:
                is_valid = False
                
            if is_valid:
                lote_vigente = lote
                break
                
        if lote_vigente:
            count = Inscricao.query.filter_by(
                evento_id=evento_id, 
                lote_id=lote_vigente.id
            ).filter(
                Inscricao.status_pagamento.in_(["approved", "pending"])
            ).count()
            
            return jsonify({
                'success': True,
                'lote_id': lote_vigente.id,
                'nome': lote_vigente.nome,
                'vagas_totais': lote_vigente.qtd_maxima,
                'vagas_usadas': count,
                'vagas_disponiveis': lote_vigente.qtd_maxima - count if lote_vigente.qtd_maxima else "ilimitado",
                'data_inicio': lote_vigente.data_inicio.isoformat() if lote_vigente.data_inicio else None,
                'data_fim': lote_vigente.data_fim.isoformat() if lote_vigente.data_fim else None
            })
        else:
            return jsonify({
                'success': False,
                'message': 'Nenhum lote válido encontrado'
            })
            
    except Exception as e:
        logger.error(f"Erro na API lote_vigente: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Erro ao verificar lote vigente: {str(e)}'
        })
        
# ===========================
#      EDITAR PARTICIPANTE
# ===========================

@routes.route('/editar_participante', methods=['GET', 'POST'])
@routes.route('/editar_participante/<int:usuario_id>/<int:oficina_id>', methods=['GET', 'POST'])
@login_required
def editar_participante(usuario_id=None, oficina_id=None):
    # Caso o cliente esteja editando um usuário
    if usuario_id:
        if not hasattr(current_user, 'tipo') or current_user.tipo != 'cliente':
            flash('Acesso negado!', 'danger')
            return redirect(url_for('routes.dashboard'))

        usuario = Usuario.query.get_or_404(usuario_id)
        oficina = Oficina.query.get_or_404(oficina_id)
    else:
        # Participante editando a si mesmo
        if not hasattr(current_user, 'tipo') or current_user.tipo != 'participante':
            flash('Acesso negado!', 'danger')
            return redirect(url_for('routes.dashboard'))
        usuario = current_user
        oficina = None  # Não necessário nesse caso

    if request.method == 'POST':
        usuario.nome = request.form.get('nome')
        usuario.cpf = request.form.get('cpf')
        usuario.email = request.form.get('email')
        usuario.formacao = request.form.get('formacao')
        usuario.estados = ','.join(request.form.getlist('estados[]') or [])
        usuario.cidades = ','.join(request.form.getlist('cidades[]') or [])

        nova_senha = request.form.get('senha')
        if nova_senha:
            usuario.senha = generate_password_hash(nova_senha)

        try:
            db.session.commit()
            flash("Perfil atualizado com sucesso!", "success")
            if usuario_id:
                return redirect(url_for('routes.editar_participante', usuario_id=usuario.id, oficina_id=oficina_id))
            return redirect(url_for('routes.dashboard_participante'))
        except Exception as e:
            db.session.rollback()
            flash("Erro ao atualizar o perfil: " + str(e), "danger")

    return render_template('editar_participante.html', usuario=usuario, oficina=oficina)


@routes.route('/cliente/checkin_manual/<int:usuario_id>/<int:oficina_id>', methods=['POST'])
@login_required
def checkin_manual(usuario_id, oficina_id):
    checkin_existente = Checkin.query.filter_by(usuario_id=usuario_id, oficina_id=oficina_id).first()
    if checkin_existente:
        flash('Participante já realizou check-in.', 'warning')
        return redirect(request.referrer or url_for('routes.gerenciar_inscricoes'))

    oficina = Oficina.query.get_or_404(oficina_id)
    
    checkin = Checkin(
        usuario_id=usuario_id,
        oficina_id=oficina_id,
        palavra_chave="manual",
        cliente_id=oficina.cliente_id,
        evento_id=oficina.evento_id
    )
    db.session.add(checkin)
    db.session.commit()

    flash('Check-in manual registrado com sucesso!', 'success')
    return redirect(request.referrer or url_for('routes.gerenciar_inscricoes'))



# ===========================
#      GESTÃO DE USUÁRIOS
# ===========================
@login_manager.user_loader
def load_user(user_id):
    user_type = session.get('user_type')
    if user_type == 'ministrante':
        return Ministrante.query.get(int(user_id))
    elif user_type in ['admin', 'participante']:
        return Usuario.query.get(int(user_id))
    elif user_type == 'cliente':
        from models import Cliente
        return Cliente.query.get(int(user_id))
    # Fallback: tenta buscar em Usuario e Ministrante
    user = Usuario.query.get(int(user_id))
    if user:
        return user
    return Ministrante.query.get(int(user_id))

@routes.route('/login', methods=['GET', 'POST'], endpoint='login')
def login():
    if request.method == 'POST':
        email = request.form['email']
        senha = request.form['senha']

        usuario = Usuario.query.filter_by(email=email).first()
        if isinstance(usuario, Cliente) and not usuario.ativo:
            flash('Sua conta está desativada. Contate o administrador.', 'danger')
            return render_template('login.html')
        
        if not usuario:
            usuario = Ministrante.query.filter_by(email=email).first()
        if not usuario:
            usuario = Cliente.query.filter_by(email=email).first()

        if not usuario:
            flash('E-mail ou senha incorretos!', 'danger')
            return render_template('login.html')
        
        from flask_login import logout_user

        if isinstance(usuario, Cliente) and not usuario.ativo:
            logout_user()  # Isso garante que o cliente seja deslogado se estiver inativo
            flash('Sua conta está desativada. Contate o administrador.', 'danger')
            return render_template('login.html')

        # Verificação correta e simplificada da senha
        senha_correta = check_password_hash(usuario.senha, senha)

        if senha_correta:
            login_user(usuario)

            # Definir corretamente o tipo de usuário na sessão
            if isinstance(usuario, Cliente):
                session['user_type'] = 'cliente'
            elif isinstance(usuario, Ministrante):
                session['user_type'] = 'ministrante'
            else:
                session['user_type'] = usuario.tipo

            flash('Login realizado com sucesso!', 'success')

            # Redirecionamento correto e simplificado
            destino = {
                'admin': 'routes.dashboard',
                'cliente': 'routes.dashboard_cliente',
                'participante': 'routes.dashboard_participante',
                'ministrante': 'routes.dashboard_ministrante',
                'professor': 'routes.dashboard_professor'
            }.get(session.get('user_type'), 'routes.dashboard')

            return redirect(url_for(destino))

        else:
            flash('E-mail ou senha incorretos!', 'danger')

    return render_template('login.html')

@routes.route('/dashboard_professor')
@login_required
def dashboard_professor():
    if current_user.tipo != 'professor':
        return redirect(url_for('routes.dashboard'))
    
    # Buscando os agendamentos do professor logado
    agendamentos_professor = AgendamentoVisita.query.filter_by(professor_id=current_user.id).all()

    return render_template(
        'dashboard_professor.html', 
        agendamentos=agendamentos_professor
    )

@routes.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Logout realizado com sucesso!', 'info')
    return redirect(url_for('routes.home'))


# ===========================
#  DASHBOARD ADMIN & PARTICIPANTE
# ===========================
@routes.route('/dashboard')
@login_required
def dashboard():
    from sqlalchemy import func, or_

    # Verificações iniciais
    is_admin = (current_user.tipo == 'admin')
    is_cliente = (current_user.tipo == 'cliente')
    is_professor = (current_user.tipo == 'professor')

    propostas = Proposta.query.all()
    eventos = Evento.query.filter_by(cliente_id=current_user.id).all() if is_cliente else Evento.query.all()
    oficinas = Oficina.query.filter_by(cliente_id=current_user.id).all() if is_cliente else Oficina.query.all()
    participantes = Usuario.query.filter_by(tipo='participante').all()
    ministrantes = Ministrante.query.all()
    relatorios = RelatorioOficina.query.order_by(RelatorioOficina.enviado_em.desc()).all()
    inscritos = Inscricao.query.filter((Inscricao.cliente_id == current_user.id) | (Inscricao.cliente_id.is_(None))).all() if is_cliente else Inscricao.query.all()
    
    # Estatísticas básicas
    total_oficinas = len(oficinas)
    total_vagas = sum(
        of.vagas if of.tipo_inscricao == 'com_inscricao_com_limite' else len(of.inscritos)
        for of in oficinas
    )
    total_inscricoes = len(inscritos)
    percentual_adesao = (total_inscricoes / total_vagas) * 100 if total_vagas > 0 else 0
    eventos_ativos = Evento.query.filter_by(cliente_id=current_user.id).all() if is_cliente else Evento.query.all()
    total_eventos = len(eventos_ativos)

    # Check-ins QR
    checkins_via_qr = Checkin.query.outerjoin(Checkin.oficina)\
        .outerjoin(Checkin.usuario)\
        .filter(
            Checkin.palavra_chave.in_(['QR-AUTO', 'QR-EVENTO']),
            or_(
                Usuario.cliente_id == current_user.id,
                Oficina.cliente_id == current_user.id,
                Checkin.cliente_id == current_user.id
            ) if is_cliente else True
        ).order_by(Checkin.data_hora.desc()).all()

    # Estatísticas por oficina
    oficinas_estatisticas = []
    oficinas_com_inscritos = []
    for of in oficinas:
        inscritos_oficina = Inscricao.query.filter_by(oficina_id=of.id).all()
        perc_ocupacao = (len(inscritos_oficina) / of.vagas) * 100 if of.vagas > 0 else 0

        oficinas_estatisticas.append({
            'id': of.id,
            'titulo': of.titulo,
            'vagas': of.vagas,
            'inscritos': len(inscritos_oficina),
            'ocupacao': perc_ocupacao
        })

        inscritos_info = [{
            'id': u.id, 'nome': u.nome, 'cpf': u.cpf, 'email': u.email, 'formacao': u.formacao
        } for i in inscritos_oficina if (u := Usuario.query.get(i.usuario_id))]

        oficinas_com_inscritos.append({
            'id': of.id, 'titulo': of.titulo, 'descricao': of.descricao,
            'ministrantes': [m.nome for m in of.ministrantes_associados] if of.ministrantes_associados else [],
            'vagas': of.vagas, 'carga_horaria': of.carga_horaria,
            'dias': OficinaDia.query.filter_by(oficina_id=of.id).all(),
            'inscritos': inscritos_info
        })

    # Configuração do cliente
    config_cliente = obter_configuracao_do_cliente(current_user.id) if is_cliente else None

    # Dados extras para cliente (agendamento de visitas)
    hoje = datetime.utcnow().date()
    agendamentos_hoje = proximos_agendamentos = []
    agendamentos_totais = agendamentos_confirmados = agendamentos_realizados = agendamentos_cancelados = total_visitantes = ocupacao_media = 0

    if is_cliente:
        base_agendamento = db.session.query(AgendamentoVisita.id).join(HorarioVisitacao).join(Evento).filter(Evento.cliente_id == current_user.id)

        agendamentos_totais = base_agendamento.count()
        agendamentos_confirmados = base_agendamento.filter(AgendamentoVisita.status == 'confirmado').count()
        agendamentos_realizados = base_agendamento.filter(AgendamentoVisita.status == 'realizado').count()
        agendamentos_cancelados = base_agendamento.filter(AgendamentoVisita.status == 'cancelado').count()
        total_visitantes = db.session.query(func.sum(AgendamentoVisita.quantidade_alunos))\
            .join(HorarioVisitacao).join(Evento)\
            .filter(Evento.cliente_id == current_user.id,
                    AgendamentoVisita.status.in_(['confirmado', 'realizado'])).scalar() or 0

        agendamentos_hoje = AgendamentoVisita.query.join(HorarioVisitacao).join(Evento)\
            .filter(Evento.cliente_id == current_user.id,
                    HorarioVisitacao.data == hoje,
                    AgendamentoVisita.status == 'confirmado').all()

        data_limite = hoje + timedelta(days=7)
        proximos_agendamentos = AgendamentoVisita.query.join(HorarioVisitacao).join(Evento)\
            .filter(Evento.cliente_id == current_user.id,
                    HorarioVisitacao.data > hoje,
                    HorarioVisitacao.data <= data_limite,
                    AgendamentoVisita.status == 'confirmado').all()

        ocupacao = db.session.query(
            func.sum(HorarioVisitacao.capacidade_total - HorarioVisitacao.vagas_disponiveis).label('ocupadas'),
            func.sum(HorarioVisitacao.capacidade_total).label('total')
        ).join(Evento).filter(Evento.cliente_id == current_user.id, HorarioVisitacao.data >= hoje).first()

        if ocupacao and ocupacao.total and ocupacao.total > 0:
            ocupacao_media = (ocupacao.ocupadas / ocupacao.total) * 100

    return render_template(
        'dashboard_admin.html' if is_admin else
        'dashboard_cliente.html' if is_cliente else
        'dashboard_professor.html' if is_professor else
        'dashboard_participante.html',

        usuario=current_user,
        participantes=participantes,
        eventos=eventos,
        oficinas=oficinas_com_inscritos,
        total_oficinas=total_oficinas,
        total_vagas=total_vagas,
        total_inscricoes=total_inscricoes,
        percentual_adesao=percentual_adesao,
        oficinas_estatisticas=oficinas_estatisticas,
        checkins_via_qr=checkins_via_qr,
        ministrantes=ministrantes,
        relatorios=relatorios,
        inscritos=inscritos,
        propostas=propostas,
        config_cliente=config_cliente,
        agendamentos_totais=agendamentos_totais,
        agendamentos_confirmados=agendamentos_confirmados,
        agendamentos_realizados=agendamentos_realizados,
        agendamentos_cancelados=agendamentos_cancelados,
        total_visitantes=total_visitantes,
        agendamentos_hoje=agendamentos_hoje,
        proximos_agendamentos=proximos_agendamentos,
        ocupacao_media=ocupacao_media,
        total_eventos=total_eventos
    )


@routes.route('/dashboard_participante')
@login_required
def dashboard_participante():
    print("\n======= INICIANDO DASHBOARD PARTICIPANTE =======")
    print(f"DEBUG [1] -> current_user.id = {current_user.id}, current_user.tipo = {current_user.tipo}")
    
    if current_user.tipo != 'participante':
        print(f"DEBUG [2] -> Redirecionando: usuário não é participante (tipo: {current_user.tipo})")
        return redirect(url_for('routes.dashboard'))

    print(f"DEBUG [3] -> Usuário é participante, cliente_id = {current_user.cliente_id}")

    # Se o participante está associado a um cliente, buscamos a config desse cliente
    config_cliente = None
    
    # CORREÇÃO 1: Buscar TODOS os eventos, não só do cliente do participante
    print(f"DEBUG [4] -> Buscando eventos do cliente_id = {current_user.cliente_id} e eventos globais")
    # Incluir eventos do cliente E eventos sem cliente (eventos globais)
    eventos = []
    if current_user.cliente_id:
        eventos_cliente = Evento.query.filter_by(cliente_id=current_user.cliente_id).all()
        print(f"DEBUG [5] -> Encontrados {len(eventos_cliente)} eventos do cliente")
        eventos.extend(eventos_cliente)
    
    # Buscar eventos sem cliente associado (eventos globais)
    eventos_globais = Evento.query.filter_by(cliente_id=None).all()
    print(f"DEBUG [6] -> Encontrados {len(eventos_globais)} eventos globais")
    eventos.extend(eventos_globais)
    
    print(f"DEBUG [7] -> Total de eventos (cliente + globais): {len(eventos)}")
    
    # Se não houver eventos específicos, usamos o primeiro como fallback (comportamento anterior)
    evento = eventos[0] if eventos else None
    print(f"DEBUG [8] -> Evento padrão selecionado: {evento.id if evento else None}")
    
    # Verifica se há formulários disponíveis para preenchimento associados ao cliente do participante
    formularios_disponiveis = False
    if current_user.cliente_id:
        print(f"DEBUG [9] -> Verificando formulários disponíveis para cliente_id = {current_user.cliente_id}")
        form_count = Formulario.query.filter_by(cliente_id=current_user.cliente_id).count()
        formularios_disponiveis = form_count > 0
        print(f"DEBUG [10] -> Formulários disponíveis: {formularios_disponiveis} (total: {form_count})")
    
    if current_user.cliente_id:
        print(f"DEBUG [11] -> Buscando configuração do cliente_id = {current_user.cliente_id}")
        from models import ConfiguracaoCliente
        config_cliente = ConfiguracaoCliente.query.filter_by(cliente_id=current_user.cliente_id).first()
        
        # Se não existir ainda, pode criar com valores padrão
        if not config_cliente:
            print(f"DEBUG [12] -> Configuração não encontrada, criando padrão para cliente_id = {current_user.cliente_id}")
            config_cliente = ConfiguracaoCliente(
                cliente_id=current_user.cliente_id,
                permitir_checkin_global=False,
                habilitar_feedback=False,
                habilitar_certificado_individual=False
            )
            db.session.add(config_cliente)
            db.session.commit()
            print(f"DEBUG [13] -> Configuração padrão criada e salva com sucesso")
        else:
            print(f"DEBUG [14] -> Configuração encontrada para cliente_id = {current_user.cliente_id}")
    
    # Agora definimos as variáveis que o template utiliza
    permitir_checkin = config_cliente.permitir_checkin_global if config_cliente else False
    habilitar_feedback = config_cliente.habilitar_feedback if config_cliente else False
    habilitar_certificado = config_cliente.habilitar_certificado_individual if config_cliente else False
    
    print(f"DEBUG [15] -> Configurações: checkin={permitir_checkin}, feedback={habilitar_feedback}, certificado={habilitar_certificado}")

    # CORREÇÃO: Limpar inscrições inválidas (oficina_id = None)
    inscricoes_validas = [i for i in current_user.inscricoes if i.oficina_id is not None]
    print(f"DEBUG [16] -> Filtrando inscrições válidas: {len(current_user.inscricoes)} -> {len(inscricoes_validas)}")
    
    # Obter os eventos em que o participante está inscrito
    print(f"DEBUG [17] -> Buscando eventos inscritos para participante_id = {current_user.id}")
    eventos_inscritos = []
    for inscricao in inscricoes_validas:
        print(f"DEBUG [18] -> Verificando inscrição: {inscricao.id}, oficina_id: {inscricao.oficina_id}")
        if inscricao.oficina and inscricao.oficina.evento_id:
            eventos_inscritos.append(inscricao.oficina.evento_id)
            print(f"DEBUG [19] -> Adicionado evento_id: {inscricao.oficina.evento_id}")
    
    # Remover duplicatas
    eventos_inscritos_original = eventos_inscritos.copy()
    eventos_inscritos = list(set(eventos_inscritos))
    print(f"DEBUG [20] -> Eventos inscritos: {len(eventos_inscritos_original)} -> {len(eventos_inscritos)} (após remover duplicatas)")
    
    # Obter IDs de todos os eventos carregados
    eventos_disponiveis_ids = [e.id for e in eventos] if eventos else []
    print(f"DEBUG [21] -> IDs de eventos disponíveis: {eventos_disponiveis_ids}")
    
    # CORREÇÃO 2: Verificar se há eventos inscritos que não estão na lista de eventos disponíveis
    eventos_inscritos_ausentes = [e_id for e_id in eventos_inscritos if e_id not in eventos_disponiveis_ids]
    
    if eventos_inscritos_ausentes:
        print(f"DEBUG [22] -> Eventos inscritos ausentes: {eventos_inscritos_ausentes}, buscando detalhes")
        # Buscar os eventos ausentes
        eventos_ausentes = Evento.query.filter(Evento.id.in_(eventos_inscritos_ausentes)).all()
        # Adicionar aos eventos disponíveis
        eventos.extend(eventos_ausentes)
        # Atualizar lista de IDs
        eventos_disponiveis_ids = [e.id for e in eventos]
        print(f"DEBUG [23] -> Total de eventos após adicionar ausentes: {len(eventos)}")
    
    # Combinar todos os IDs de eventos (disponíveis + inscritos) e remover duplicatas
    todos_eventos_ids = list(set(eventos_disponiveis_ids + eventos_inscritos))
    print(f"DEBUG [24] -> IDs de todos os eventos (disponíveis + inscritos): {todos_eventos_ids}")
    
    # Buscar detalhes de todos os eventos e verificar se já ocorreram
    eventos_info = {}
    data_atual = datetime.now().date()
    print(f"DEBUG [25] -> Data atual = {data_atual}")
    
    print(f"DEBUG [26] -> Processando informações de {len(todos_eventos_ids)} eventos")
    for evento_id in todos_eventos_ids:
        print(f"DEBUG [27] -> Buscando evento_id = {evento_id}")
        evento_obj = Evento.query.get(evento_id)
        if evento_obj:
            print(f"DEBUG [28] -> Evento {evento_obj.id} - {evento_obj.nome}")
            print(f"DEBUG [29] -> data_inicio = {evento_obj.data_inicio}, data_fim = {evento_obj.data_fim}")
            
            # CORREÇÃO 3: Um evento só é considerado encerrado se:
            # 1. Tiver uma data_fim definida E
            # 2. A data_fim for anterior à data atual
            ja_ocorreu = False
            
            # Se data_fim for nula, consideramos o evento como não encerrado (aberto)
            if evento_obj.data_fim:
                print(f"DEBUG [30] -> Processando data_fim para evento_id = {evento_obj.id}")
                # Converter para data se for datetime
                if hasattr(evento_obj.data_fim, 'date'):
                    data_fim = evento_obj.data_fim.date()
                    print(f"DEBUG [31] -> Convertendo datetime para date: {data_fim}")
                else:
                    # Se já for um objeto date
                    data_fim = evento_obj.data_fim
                    print(f"DEBUG [32] -> Já é objeto date: {data_fim}")
                
                # Só consideramos encerrado se a data_fim for anterior à data atual
                ja_ocorreu = data_fim < data_atual
                print(f"DEBUG [33] -> data_fim ({data_fim}) < data_atual ({data_atual})? {ja_ocorreu}")
            else:
                print(f"DEBUG [34] -> Evento sem data_fim, considerado aberto")
            
            # Verificar se é um evento futuro (data_inicio posterior à data atual)
            eh_futuro = False
            if evento_obj.data_inicio:
                print(f"DEBUG [35] -> Processando data_inicio para evento_id = {evento_obj.id}")
                if hasattr(evento_obj.data_inicio, 'date'):
                    data_inicio = evento_obj.data_inicio.date()
                    print(f"DEBUG [36] -> Convertendo datetime para date: {data_inicio}")
                else:
                    data_inicio = evento_obj.data_inicio
                    print(f"DEBUG [37] -> Já é objeto date: {data_inicio}")
                
                eh_futuro = data_inicio > data_atual
                print(f"DEBUG [38] -> É evento futuro (data_inicio ({data_inicio}) > data_atual ({data_atual}))? {eh_futuro}")
            else:
                print(f"DEBUG [39] -> Evento sem data_inicio definida, considerado atual")
            
            # Registrar status e informações do evento
            status = 'Futuro' if eh_futuro else ('Encerrado' if ja_ocorreu else 'Atual')
            eventos_info[evento_id] = {
                'id': evento_obj.id,
                'nome': evento_obj.nome,
                'data_inicio': evento_obj.data_inicio,
                'data_fim': evento_obj.data_fim,
                'ja_ocorreu': ja_ocorreu,
                'eh_futuro': eh_futuro,
                'status': status
            }
            print(f"DEBUG [40] -> Evento {evento_obj.id} classificado como '{status}'")
    
    # CORREÇÃO 4: Buscar oficinas para cada evento individualmente
    print(f"DEBUG [41] -> Buscando oficinas para todos os eventos e oficinas sem eventos")
    oficinas = []
    
    # Primeiro, buscar oficinas dos eventos identificados
    if todos_eventos_ids:
        print(f"DEBUG [42] -> Buscando oficinas para eventos IDs: {todos_eventos_ids}")
        oficinas_eventos = Oficina.query.filter(Oficina.evento_id.in_(todos_eventos_ids)).all()
        oficinas.extend(oficinas_eventos)
        print(f"DEBUG [43] -> Encontradas {len(oficinas_eventos)} oficinas para os eventos")
    
    # CORREÇÃO 5: Adicionar oficinas sem evento associado (evento_id=None)
    print(f"DEBUG [44] -> Buscando oficinas sem evento associado")
    if current_user.cliente_id:
        # Buscar oficinas do cliente sem evento associado
        oficinas_sem_evento_cliente = Oficina.query.filter_by(
            cliente_id=current_user.cliente_id, 
            evento_id=None
        ).all()
        oficinas.extend(oficinas_sem_evento_cliente)
        print(f"DEBUG [45] -> Encontradas {len(oficinas_sem_evento_cliente)} oficinas do cliente sem evento")
    
    # Buscar oficinas globais sem evento associado
    oficinas_sem_evento_globais = Oficina.query.filter_by(
        cliente_id=None, 
        evento_id=None
    ).all()
    oficinas.extend(oficinas_sem_evento_globais)
    print(f"DEBUG [46] -> Encontradas {len(oficinas_sem_evento_globais)} oficinas globais sem evento")
    
    # CORREÇÃO 6: Se não encontramos oficinas mas temos eventos, buscar todas as oficinas do sistema
    if len(oficinas) == 0 and len(eventos) > 0:
        print(f"DEBUG [47] -> Não foram encontradas oficinas. Buscando oficinas do cliente")
        if current_user.cliente_id:
            oficinas_cliente = Oficina.query.filter_by(cliente_id=current_user.cliente_id).all()
            oficinas.extend(oficinas_cliente)
            print(f"DEBUG [48] -> Encontradas {len(oficinas_cliente)} oficinas pelo cliente_id")
    
    print(f"DEBUG [49] -> Total de oficinas encontradas: {len(oficinas)}")

    # CORREÇÃO 7: Filtrar inscrições válidas (com oficina_id não nulo)
    print(f"DEBUG [50] -> Montando lista de inscrições do participante_id = {current_user.id}")
    inscricoes_ids = [i.oficina_id for i in inscricoes_validas if i.oficina_id is not None]
    print(f"DEBUG [51] -> Participante inscrito em {len(inscricoes_ids)} oficinas válidas: {inscricoes_ids}")
    
    # Lógica para professores verem horários disponíveis
    print(f"DEBUG [52] -> Buscando horários disponíveis para visitação")
    horarios_disponiveis = HorarioVisitacao.query.filter(
        HorarioVisitacao.vagas_disponiveis > 0,
        HorarioVisitacao.data >= datetime.now().date()
    ).all()
    print(f"DEBUG [53] -> Encontrados {len(horarios_disponiveis)} horários disponíveis")

    # Carregar regras de inscrição para o tipo de inscrição do usuário por evento
    print(f"DEBUG [54] -> Carregando regras de inscrição para tipo_inscricao_id = {current_user.tipo_inscricao_id}")
    regras_inscricao = {}
    
    if current_user.tipo_inscricao_id:
        print(f"DEBUG [55] -> Verificando regras para cada evento")
        # Para cada evento disponível, verificar as regras para o tipo de inscrição do usuário
        for evento_id in todos_eventos_ids:
            print(f"DEBUG [56] -> Buscando regra para evento_id = {evento_id}, tipo_inscricao_id = {current_user.tipo_inscricao_id}")
            regra = RegraInscricaoEvento.query.filter_by(
                evento_id=evento_id,
                tipo_inscricao_id=current_user.tipo_inscricao_id
            ).first()
            
            if regra:
                oficinas_permitidas = regra.get_oficinas_permitidas_list()
                regras_inscricao[evento_id] = {
                    'limite_oficinas': regra.limite_oficinas,
                    'oficinas_permitidas': set(oficinas_permitidas)
                }
                print(f"DEBUG [57] -> Regra encontrada: limite={regra.limite_oficinas}, oficinas_permitidas={len(oficinas_permitidas)}")
            else:
                print(f"DEBUG [58] -> Nenhuma regra encontrada para este evento/tipo de inscrição")
    
    # Contar inscrições do usuário por evento
    print(f"DEBUG [59] -> Contando inscrições do usuário por evento")
    inscricoes_por_evento = {}
    for inscricao in inscricoes_validas:
        if inscricao.oficina and inscricao.oficina.evento_id:
            evento_id = inscricao.oficina.evento_id
            if evento_id not in inscricoes_por_evento:
                inscricoes_por_evento[evento_id] = 0
            inscricoes_por_evento[evento_id] += 1
            print(f"DEBUG [60] -> Evento {evento_id}: {inscricoes_por_evento[evento_id]} inscrições")
    
    # Monte a estrutura que o template "dashboard_participante.html" precisa
    print(f"DEBUG [61] -> Formatando {len(oficinas)} oficinas para exibição")
    oficinas_formatadas = []
    
    # CORREÇÃO 8: Garantir que todas as oficinas estejam devidamente processadas
    for idx, oficina in enumerate(oficinas):
        print(f"DEBUG [62] -> Processando oficina {idx+1}/{len(oficinas)}: id={oficina.id}, título={oficina.titulo}")
        
        print(f"DEBUG [63] -> Buscando dias para oficina_id = {oficina.id}")
        dias = OficinaDia.query.filter_by(oficina_id=oficina.id).all()
        print(f"DEBUG [64] -> Encontrados {len(dias)} dias para esta oficina")
        
        # Verificar se a oficina está disponível para o tipo de inscrição do usuário
        disponivel_para_inscricao = True
        motivo_indisponibilidade = None
        
        # CORREÇÃO 9: Tratamento para oficinas sem evento
        evento_id = oficina.evento_id
        evento_nome = "Atividades Gerais"  # Nome padrão para oficinas sem evento
        evento_encerrado = False
        evento_futuro = False
        evento_info = None
        
        # Se a oficina tiver evento associado, pegar informações do evento
        if evento_id:
            print(f"DEBUG [65] -> Oficina tem evento_id = {evento_id}")
            if oficina.evento:
                evento_nome = oficina.evento.nome
                print(f"DEBUG [66] -> Nome do evento: {evento_nome}")
            
            # Verificar disponibilidade para tipo de inscrição
            if current_user.tipo_inscricao_id:
                print(f"DEBUG [67] -> Verificando disponibilidade para tipo_inscricao_id = {current_user.tipo_inscricao_id}, evento_id = {evento_id}")
                # Obter regras para este evento específico
                regra_evento = regras_inscricao.get(evento_id, {})
                
                # Se há oficinas permitidas definidas e esta oficina não está na lista
                oficinas_permitidas = regra_evento.get('oficinas_permitidas', set())
                if oficinas_permitidas:
                    print(f"DEBUG [68] -> Verificando se oficina {oficina.id} está entre as {len(oficinas_permitidas)} oficinas permitidas")
                    if oficina.id not in oficinas_permitidas:
                        disponivel_para_inscricao = False
                        motivo_indisponibilidade = "Seu tipo de inscrição não permite acesso a esta oficina"
                        print(f"DEBUG [69] -> Oficina indisponível: não está na lista de permitidas")
                
                # Se há limite de oficinas e o usuário já atingiu o limite
                limite_oficinas = regra_evento.get('limite_oficinas', 0)
                if limite_oficinas > 0 and evento_id in inscricoes_por_evento:
                    print(f"DEBUG [70] -> Verificando limite de inscrições: atual={inscricoes_por_evento[evento_id]}, limite={limite_oficinas}")
                    if inscricoes_por_evento[evento_id] >= limite_oficinas and oficina.id not in inscricoes_ids:
                        disponivel_para_inscricao = False
                        motivo_indisponibilidade = f"Você já atingiu o limite de {limite_oficinas} oficinas para seu tipo de inscrição"
                        print(f"DEBUG [71] -> Oficina indisponível: limite de inscrições atingido")
            
            # Obter informações do evento desta oficina
            print(f"DEBUG [72] -> Buscando informações do evento para oficina_id = {oficina.id}, evento_id = {evento_id}")
            if evento_id in eventos_info:
                evento_info = eventos_info[evento_id]
                evento_encerrado = evento_info['ja_ocorreu']
                evento_futuro = evento_info['eh_futuro']
                print(f"DEBUG [73] -> Evento encontrado: encerrado={evento_encerrado}, futuro={evento_futuro}, status={evento_info['status']}")
            else:
                # CORREÇÃO 10: Se o evento não estiver em eventos_info, criar uma entrada padrão
                print(f"DEBUG [74] -> Evento não encontrado nas informações processadas, adicionando informação padrão")
                if oficina.evento:
                    # Calcular status baseado nas datas do evento
                    ja_ocorreu = False
                    eh_futuro = False
                    if oficina.evento.data_fim:
                        data_fim = oficina.evento.data_fim
                        if hasattr(data_fim, 'date'):
                            data_fim = data_fim.date()
                        ja_ocorreu = data_fim < data_atual
                    
                    if oficina.evento.data_inicio:
                        data_inicio = oficina.evento.data_inicio
                        if hasattr(data_inicio, 'date'):
                            data_inicio = data_inicio.date()
                        eh_futuro = data_inicio > data_atual
                    
                    status = 'Futuro' if eh_futuro else ('Encerrado' if ja_ocorreu else 'Atual')
                    
                    eventos_info[evento_id] = {
                        'id': evento_id,
                        'nome': evento_nome,
                        'data_inicio': oficina.evento.data_inicio,
                        'data_fim': oficina.evento.data_fim,
                        'ja_ocorreu': ja_ocorreu,
                        'eh_futuro': eh_futuro,
                        'status': status
                    }
                    
                    evento_encerrado = ja_ocorreu
                    evento_futuro = eh_futuro
                    evento_info = eventos_info[evento_id]
                    print(f"DEBUG [75] -> Adicionando evento {evento_id} como {status}")
        else:
            # CORREÇÃO 11: Para oficinas sem evento, considerar como eventos atuais
            print(f"DEBUG [76] -> Oficina sem evento_id, tratando como 'Atividades Gerais'")
            evento_id = 'sem_evento'  # Um marcador para identificar
            evento_encerrado = False
            evento_futuro = False
        
        print(f"DEBUG [77] -> Adicionando oficina formatada: id={oficina.id}, disponível={disponivel_para_inscricao}")
        oficinas_formatadas.append({
            'id': oficina.id,
            'titulo': oficina.titulo,
            'descricao': oficina.descricao,
            'ministrantes': [m.nome for m in oficina.ministrantes_associados] if hasattr(oficina, 'ministrantes_associados') and oficina.ministrantes_associados else [],
            'vagas': oficina.vagas,
            'carga_horaria': oficina.carga_horaria,
            'dias': dias,
            'evento_id': evento_id,
            'evento_nome': evento_nome,
            'tipo_inscricao': oficina.tipo_inscricao,
            'disponivel_para_inscricao': disponivel_para_inscricao,
            'motivo_indisponibilidade': motivo_indisponibilidade,
            'evento_encerrado': evento_encerrado,
            'evento_futuro': evento_futuro,
            'evento_data_inicio': evento_info['data_inicio'] if evento_info else None,
            'evento_data_fim': evento_info['data_fim'] if evento_info else None,
            'evento_status': evento_info['status'] if evento_info else 'Atual'
        })

    print(f"DEBUG [78] -> Estatísticas finais de oficinas:")
    print(f"DEBUG [79] -> Total de oficinas formatadas: {len(oficinas_formatadas)}")
    print(f"DEBUG [80] -> Oficinas de eventos abertos: {len([o for o in oficinas_formatadas if not o['evento_encerrado']])}")
    print(f"DEBUG [81] -> Oficinas de eventos futuros: {len([o for o in oficinas_formatadas if o['evento_futuro']])}")
    print(f"DEBUG [82] -> Oficinas de eventos encerrados: {len([o for o in oficinas_formatadas if o['evento_encerrado']])}")
    print(f"DEBUG [83] -> Oficinas sem evento: {len([o for o in oficinas_formatadas if o['evento_id'] == 'sem_evento'])}")
    print(f"DEBUG [84] -> Oficinas disponíveis para inscrição: {len([o for o in oficinas_formatadas if o['disponivel_para_inscricao']])}")

    # CORREÇÃO 12: Garantir que mostraremos todas as oficinas, mesmo se todos os eventos estiverem encerrados
    print(f"DEBUG [85] -> Renderizando template dashboard_participante.html")
    
     # NOVA CORREÇÃO: agrupar eventos por status usando eventos_info
    eventos_futuros = [e for e in eventos if eventos_info.get(e.id, {}).get('eh_futuro')]
    eventos_atuais = [e for e in eventos if not eventos_info.get(e.id, {}).get('eh_futuro')
                       and not eventos_info.get(e.id, {}).get('ja_ocorreu')]
    eventos_encerrados = [e for e in eventos if eventos_info.get(e.id, {}).get('ja_ocorreu')]
    print(f"DEBUG -> Eventos futuros: {len(eventos_futuros)}, atuais: {len(eventos_atuais)}, encerrados: {len(eventos_encerrados)}")
    
    return render_template(
        'dashboard_participante.html',
        config_cliente=config_cliente,
        usuario=current_user,
        evento=evento,
        eventos=eventos,
        eventos_futuros=eventos_futuros,
        eventos_atuais=eventos_atuais,
        eventos_encerrados=eventos_encerrados,
        oficinas=oficinas_formatadas,
        eventos_inscritos=eventos_inscritos,
        todos_eventos_ids=todos_eventos_ids,
        eventos_info=eventos_info,
        data_atual=data_atual,
        permitir_checkin_global=permitir_checkin,
        habilitar_feedback=habilitar_feedback,
        habilitar_certificado_individual=habilitar_certificado,
        formularios_disponiveis=formularios_disponiveis,
        horarios_disponiveis=horarios_disponiveis,
        # Variáveis para diagnóstico
        debug_total_oficinas=len(oficinas_formatadas),
        debug_oficinas_encerradas=len([o for o in oficinas_formatadas if o['evento_encerrado']]),
        debug_oficinas_futuras=len([o for o in oficinas_formatadas if o['evento_futuro']]),
        debug_oficinas_atuais=len([o for o in oficinas_formatadas if not o['evento_encerrado'] and not o['evento_futuro']]),
        debug_oficinas_sem_evento=len([o for o in oficinas_formatadas if o['evento_id'] == 'sem_evento']),
        # CORREÇÃO: Forçar exibição de todos os tipos de eventos
        forcar_exibir_encerrados=True
    )


# ===========================
#    GESTÃO DE OFICINAS - ADMIN
# ===========================
@routes.route('/criar_oficina', methods=['GET', 'POST'])
@login_required
def criar_oficina():
    if current_user.tipo not in ['admin', 'cliente']:
        flash('Acesso negado!', 'danger')
        return redirect(url_for('routes.dashboard'))

    estados = obter_estados()
    ministrantes_disponiveis = (
        Ministrante.query.filter_by(cliente_id=current_user.id).all()
        if current_user.tipo == 'cliente'
        else Ministrante.query.all()
    )
    clientes_disponiveis = Cliente.query.all() if current_user.tipo == 'admin' else []
    eventos_disponiveis = (
        Evento.query.filter_by(cliente_id=current_user.id).all()
        if current_user.tipo == 'cliente'
        else Evento.query.all()
    )

    if request.method == 'POST':
        print("Dados recebidos do formulário:", request.form)  # Log para depuração

        # Captura os campos do formulário
        titulo = request.form.get('titulo')
        descricao = request.form.get('descricao')
        ministrante_id = request.form.get('ministrante_id') or None
        vagas = request.form.get('vagas')
        carga_horaria = request.form.get('carga_horaria')
        estado = request.form.get('estado')
        cidade = request.form.get('cidade')
        opcoes_checkin = request.form.get('opcoes_checkin')
        palavra_correta = request.form.get('palavra_correta')
        evento_id = request.form.get('evento_id')

        # Validação básica dos campos obrigatórios
        if not all([titulo, descricao, vagas, carga_horaria, estado, cidade, evento_id]):
            flash("Erro: Todos os campos obrigatórios devem ser preenchidos!", "danger")
            return render_template(
                'criar_oficina.html',
                estados=estados,
                ministrantes=ministrantes_disponiveis,
                clientes=clientes_disponiveis,
                eventos=eventos_disponiveis,
                datas=request.form.getlist('data[]'),
                horarios_inicio=request.form.getlist('horario_inicio[]'),
                horarios_fim=request.form.getlist('horario_fim[]')
            )

        # Definir o cliente da oficina
        cliente_id = (
            request.form.get('cliente_id') if current_user.tipo == 'admin' else current_user.id
        )

        # Verifica se o cliente possui habilitação de pagamento
        inscricao_gratuita = (
            True if request.form.get('inscricao_gratuita') == 'on' else False
            if current_user.habilita_pagamento else True
        )

        try:
            # Cria a nova oficina
            # Determina o tipo de inscrição com base nos dados do formulário
            tipo_inscricao = request.form.get('tipo_inscricao', 'com_inscricao_com_limite')
            
            # Obtém os valores dos novos campos tipo_oficina e tipo_oficina_outro
            tipo_oficina = request.form.get('tipo_oficina', 'Oficina')
            tipo_oficina_outro = None
            if tipo_oficina == 'outros':
                tipo_oficina_outro = request.form.get('tipo_oficina_outro')
                
            nova_oficina = Oficina(
                titulo=titulo,
                descricao=descricao,
                ministrante_id=ministrante_id,
                vagas=int(vagas),  # Este valor será ajustado no __init__ conforme o tipo_inscricao
                carga_horaria=carga_horaria,
                estado=estado,
                cidade=cidade,
                cliente_id=cliente_id,
                evento_id=evento_id,
                opcoes_checkin=opcoes_checkin,
                palavra_correta=palavra_correta,
                tipo_inscricao=tipo_inscricao,
                tipo_oficina=tipo_oficina,
                tipo_oficina_outro=tipo_oficina_outro
            )
            nova_oficina.inscricao_gratuita = inscricao_gratuita
            db.session.add(nova_oficina)
            db.session.flush()  # Garante que o ID da oficina esteja disponível
            
       

            # Adiciona os tipos de inscrição (se não for gratuita)
            if not inscricao_gratuita:
                nomes_tipos = request.form.getlist('nome_tipo[]')
                precos = request.form.getlist('preco_tipo[]')
                if not nomes_tipos or not precos:
                    raise ValueError("Tipos de inscrição e preços são obrigatórios para oficinas pagas.")
                for nome, preco in zip(nomes_tipos, precos):
                    novo_tipo = InscricaoTipo(
                        oficina_id=nova_oficina.id,
                        nome=nome,
                        preco=float(preco)
                    )
                    db.session.add(novo_tipo)

            # Adiciona os dias e horários
            datas = request.form.getlist('data[]')
            horarios_inicio = request.form.getlist('horario_inicio[]')
            horarios_fim = request.form.getlist('horario_fim[]')
            if not datas or len(datas) != len(horarios_inicio) or len(datas) != len(horarios_fim):
                raise ValueError("Datas e horários inconsistentes.")
            for i in range(len(datas)):
                novo_dia = OficinaDia(
                    oficina_id=nova_oficina.id,
                    data=datetime.strptime(datas[i], '%Y-%m-%d').date(),
                    horario_inicio=horarios_inicio[i],
                    horario_fim=horarios_fim[i]
                )
                db.session.add(novo_dia)
            
                # 3) Captura lista de IDs de ministrantes extras
            ids_extras = request.form.getlist('ministrantes_ids[]')  # array
            for mid in ids_extras:
                m = Ministrante.query.get(int(mid))
                if m:
                    nova_oficina.ministrantes_associados.append(m)

            db.session.commit()
            flash('Atividade criada com sucesso!', 'success')
            return redirect(
                url_for('routes.dashboard_cliente' if current_user.tipo == 'cliente' else 'routes.dashboard')
            )

        except Exception as e:
            db.session.rollback()
            print(f"Erro ao criar oficina: {str(e)}")  # Log do erro
            flash(f"Erro ao criar oficina: {str(e)}", "danger")
            return render_template(
                'criar_oficina.html',
                estados=estados,
                ministrantes=ministrantes_disponiveis,
                clientes=clientes_disponiveis,
                eventos=eventos_disponiveis,
                datas=request.form.getlist('data[]'),
                horarios_inicio=request.form.getlist('horario_inicio[]'),
                horarios_fim=request.form.getlist('horario_fim[]')
            )

    return render_template(
        'criar_oficina.html',
        estados=estados,
        ministrantes=ministrantes_disponiveis,
        clientes=clientes_disponiveis,
        eventos=eventos_disponiveis
    )

@routes.route('/get_cidades/<estado_sigla>')
def get_cidades(estado_sigla):
    cidades = obter_cidades(estado_sigla)
    print(f"📌 Estado recebido: {estado_sigla}, Cidades encontradas: {cidades}")
    return jsonify(cidades)


@routes.route('/editar_oficina/<int:oficina_id>', methods=['GET', 'POST'])
@login_required
def editar_oficina(oficina_id):
    oficina = Oficina.query.get_or_404(oficina_id)

    if current_user.tipo == 'cliente' and oficina.cliente_id != current_user.id:
        flash('Você não tem permissão para editar esta atividade.', 'danger')
        return redirect(url_for('routes.dashboard_cliente'))

    estados = obter_estados()
    if current_user.tipo == 'cliente':
        ministrantes = Ministrante.query.filter_by(cliente_id=current_user.id).all()
        eventos_disponiveis = Evento.query.filter_by(cliente_id=current_user.id).all()
    else:
        ministrantes = Ministrante.query.all()
        eventos_disponiveis = Evento.query.all()

    clientes_disponiveis = Cliente.query.all() if current_user.tipo == 'admin' else []

    if request.method == 'POST':
        oficina.titulo = request.form.get('titulo')
        oficina.descricao = request.form.get('descricao')
        ministrante_id = request.form.get('ministrante_id') or None
        oficina.ministrante_id = ministrante_id
        oficina.carga_horaria = request.form.get('carga_horaria')
        oficina.estado = request.form.get('estado')
        oficina.cidade = request.form.get('cidade')
        oficina.opcoes_checkin = request.form.get('opcoes_checkin')
        oficina.palavra_correta = request.form.get('palavra_correta')
        oficina.evento_id = request.form.get('evento_id')  # Atualiza o evento_id
        
        # Atualiza os campos tipo_oficina e tipo_oficina_outro
        tipo_oficina = request.form.get('tipo_oficina', 'Oficina')
        oficina.tipo_oficina = tipo_oficina
        if tipo_oficina == 'outros':
            oficina.tipo_oficina_outro = request.form.get('tipo_oficina_outro')
        else:
            oficina.tipo_oficina_outro = None
        
        # Atualiza o tipo de inscrição e ajusta as vagas conforme necessário
        tipo_inscricao = request.form.get('tipo_inscricao', 'com_inscricao_com_limite')
        oficina.tipo_inscricao = tipo_inscricao
        
        # Define o valor de vagas com base no tipo de inscrição
        if tipo_inscricao == 'sem_inscricao':
            oficina.vagas = 0  # Não é necessário controlar vagas
        elif tipo_inscricao == 'com_inscricao_sem_limite':
            oficina.vagas = 9999  # Um valor alto para representar "sem limite"
        else:  # com_inscricao_com_limite
            oficina.vagas = int(request.form.get('vagas'))

        # Permitir que apenas admins alterem o cliente
        if current_user.tipo == 'admin':
            oficina.cliente_id = request.form.get('cliente_id') or None
            
        # Atualiza o campo inscricao_gratuita
        if current_user.habilita_pagamento:
            oficina.inscricao_gratuita = True if request.form.get('inscricao_gratuita') == 'on' else False
        else:
            oficina.inscricao_gratuita = True

        try:
            # Atualizar os dias e horários
            datas = request.form.getlist('data[]')
            horarios_inicio = request.form.getlist('horario_inicio[]')
            horarios_fim = request.form.getlist('horario_fim[]')
            oficina.ministrantes_associados = []  # Limpa os ministrantes associados

            # Capturar IDs dos ministrantes selecionados
            ministrantes_ids = request.form.getlist('ministrantes_ids[]')
            for mid in ministrantes_ids:
                m = Ministrante.query.get(int(mid))
                if m:
                    oficina.ministrantes_associados.append(m)

            if not datas or len(datas) != len(horarios_inicio) or len(datas) != len(horarios_fim):
                raise ValueError("Datas e horários inconsistentes.")

            # Apagar os registros antigos para evitar duplicação
            OficinaDia.query.filter_by(oficina_id=oficina.id).delete()

            for i in range(len(datas)):
                novo_dia = OficinaDia(
                    oficina_id=oficina.id,
                    data=datetime.strptime(datas[i], '%Y-%m-%d').date(),
                    horario_inicio=horarios_inicio[i],
                    horario_fim=horarios_fim[i]
                )
                db.session.add(novo_dia)
                
            # Atualiza os tipos de inscrição (se não for gratuita)
            if not oficina.inscricao_gratuita:
                # Remove os tipos de inscrição antigos
                InscricaoTipo.query.filter_by(oficina_id=oficina.id).delete()
                
                # Adiciona os novos tipos de inscrição
                nomes_tipos = request.form.getlist('nome_tipo[]')
                precos = request.form.getlist('preco_tipo[]')
                if not nomes_tipos or not precos:
                    raise ValueError("Tipos de inscrição e preços são obrigatórios para oficinas pagas.")
                for nome, preco in zip(nomes_tipos, precos):
                    novo_tipo = InscricaoTipo(
                        oficina_id=oficina.id,
                        nome=nome,
                        preco=float(preco)
                    )
                    db.session.add(novo_tipo)

            db.session.commit()
            flash('Oficina editada com sucesso!', 'success')
            return redirect(url_for('routes.dashboard_cliente' if current_user.tipo == 'cliente' else 'routes.dashboard'))

        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao editar oficina: {str(e)}', 'danger')
            return render_template(
                'editar_oficina.html',
                oficina=oficina,
                estados=estados,
                ministrantes=ministrantes,
                clientes=clientes_disponiveis,
                eventos=eventos_disponiveis
            )

    return render_template(
        'editar_oficina.html',
        oficina=oficina,
        estados=estados,
        ministrantes=ministrantes,
        clientes=clientes_disponiveis,
        eventos=eventos_disponiveis
    )

@routes.route('/excluir_oficina/<int:oficina_id>', methods=['POST'])
@login_required
def excluir_oficina(oficina_id):
    oficina = Oficina.query.get_or_404(oficina_id)

    # 🚨 Cliente só pode excluir oficinas que ele criou
    if current_user.tipo == 'cliente' and oficina.cliente_id != current_user.id:
        flash('Você não tem permissão para excluir esta oficina.', 'danger')
        return redirect(url_for('routes.dashboard_cliente'))

    try:
        print(f"📌 [DEBUG] Excluindo oficina ID: {oficina_id}")

        # 1️⃣ **Excluir check-ins relacionados à oficina**
        db.session.query(Checkin).filter_by(oficina_id=oficina.id).delete()
        print("✅ [DEBUG] Check-ins removidos.")

        # 2️⃣ **Excluir inscrições associadas à oficina**
        db.session.query(Inscricao).filter_by(oficina_id=oficina.id).delete()
        print("✅ [DEBUG] Inscrições removidas.")

        # 3️⃣ **Excluir registros de datas da oficina (OficinaDia)**
        db.session.query(OficinaDia).filter_by(oficina_id=oficina.id).delete()
        print("✅ [DEBUG] Dias da oficina removidos.")

        # 4️⃣ **Excluir materiais da oficina**
        db.session.query(MaterialOficina).filter_by(oficina_id=oficina.id).delete()
        print("✅ [DEBUG] Materiais da oficina removidos.")

        # 5️⃣ **Excluir relatórios associados à oficina**
        db.session.query(RelatorioOficina).filter_by(oficina_id=oficina.id).delete()
        print("✅ [DEBUG] Relatórios da oficina removidos.")

        # 6️⃣ **Excluir associações com ministrantes na tabela de associação**
        from sqlalchemy import text
        db.session.execute(
            text('DELETE FROM oficina_ministrantes_association WHERE oficina_id = :oficina_id'),
            {'oficina_id': oficina.id}
        )
        print("✅ [DEBUG] Associações com ministrantes removidas.")

        # 7️⃣ **Excluir a própria oficina**
        db.session.delete(oficina)
        db.session.commit()
        print("✅ [DEBUG] Oficina removida com sucesso!")
        flash('Oficina excluída com sucesso!', 'success')

    except Exception as e:
        db.session.rollback()
        print(f"❌ [ERRO] Erro ao excluir oficina {oficina_id}: {str(e)}")
        flash(f'Erro ao excluir oficina: {str(e)}', 'danger')

    return redirect(url_for('routes.dashboard_cliente' if current_user.tipo == 'cliente' else 'routes.dashboard'))


# ===========================
# INSCRIÇÃO EM OFICINAS - PARTICIPANTE
# ===========================
@routes.route('/inscrever/<int:oficina_id>', methods=['POST'])
@login_required
def inscrever(oficina_id):
    if current_user.tipo != 'participante':
        return jsonify({
            'success': False,
            'message': 'Apenas participantes podem se inscrever.'
        })

    oficina = Oficina.query.get(oficina_id)
    if not oficina:
        return jsonify({
            'success': False,
            'message': 'Oficina não encontrada!'
        })

    # Verifica disponibilidade de vagas com base no tipo de inscrição
    if oficina.tipo_inscricao == 'sem_inscricao':
        # Não é necessário verificar vagas para oficinas sem inscrição
        pass
    elif oficina.tipo_inscricao == 'com_inscricao_sem_limite':
        # Não há limite de vagas
        pass
    elif oficina.vagas <= 0:
        return jsonify({
            'success': False,
            'message': 'Esta oficina está lotada!'
        })

    # Evita duplicidade
    if Inscricao.query.filter_by(usuario_id=current_user.id, oficina_id=oficina.id).first():
        return jsonify({
            'success': False,
            'message': 'Você já está inscrito nesta oficina!'
        })
    
    # Verificar regras de inscrição baseadas no tipo de inscrição do participante
    if oficina.evento_id and current_user.tipo_inscricao_id:
        # Buscar regras para o tipo de inscrição do participante
        regra = RegraInscricaoEvento.query.filter_by(
            evento_id=oficina.evento_id,
            tipo_inscricao_id=current_user.tipo_inscricao_id
        ).first()
        
        if regra:
            # Verificar se esta oficina está na lista de oficinas permitidas
            oficinas_permitidas = regra.get_oficinas_permitidas_list()
            if oficinas_permitidas and oficina.id not in oficinas_permitidas:
                return jsonify({
                    'success': False,
                    'message': 'Seu tipo de inscrição não permite acesso a esta oficina.'
                })
            
            # Verificar se o participante já atingiu o limite de oficinas
            if regra.limite_oficinas > 0:
                # Contar quantas oficinas o participante já está inscrito neste evento
                inscricoes_evento = Inscricao.query.join(Oficina).filter(
                    Inscricao.usuario_id == current_user.id,
                    Oficina.evento_id == oficina.evento_id
                ).count()
                
                if inscricoes_evento >= regra.limite_oficinas:
                    return jsonify({
                        'success': False,
                        'message': f'Você já atingiu o limite de {regra.limite_oficinas} oficinas para seu tipo de inscrição.'
                    })

    # Decrementa vagas se for uma oficina com inscrição limitada
    if oficina.tipo_inscricao == 'com_inscricao_com_limite':
        oficina.vagas -= 1
    
    # No formulário de inscrição, capture o id do tipo de inscrição escolhido:
    tipo_inscricao_id = request.form.get('tipo_inscricao_id')  # Pode ser None se for gratuita
    
    # Criar a inscrição
    inscricao = Inscricao(
        usuario_id=current_user.id, 
        oficina_id=oficina.id, 
        cliente_id=current_user.cliente_id,
        evento_id=oficina.evento_id  # Importante: associar ao evento da oficina
    )
    
    if tipo_inscricao_id:
        inscricao.tipo_inscricao_id = tipo_inscricao_id
        # Aqui você pode chamar a função que integra com o Mercado Pago
        # Exemplo: url_pagamento = iniciar_pagamento(inscricao)
    
    try:
        db.session.add(inscricao)
        
        # IMPORTANTE: Se o usuário não estiver associado a nenhum evento ainda,
        # associar este usuário ao evento da oficina para manter compatibilidade com o sistema
        if not current_user.evento_id and oficina.evento_id:
            current_user.evento_id = oficina.evento_id
            
        db.session.commit()

        # Gera o comprovante
        try:
            pdf_path = gerar_comprovante_pdf(current_user, oficina, inscricao)

            assunto = f"Confirmação de Inscrição - {oficina.titulo}"
            corpo_texto = f"Olá {current_user.nome},\n\nVocê se inscreveu na oficina '{oficina.titulo}'.\nSegue o comprovante de inscrição em anexo."

            enviar_email(
                destinatario=current_user.email,
                nome_participante=current_user.nome,
                nome_oficina=oficina.titulo,
                assunto=assunto,
                corpo_texto=corpo_texto,
                anexo_path=pdf_path
            )
        except Exception as e:
            logger.error(f"❌ ERRO ao enviar e-mail: {e}", exc_info=True)
            # Continuamos mesmo se houver erro no e-mail, pois a inscrição já foi concluída
            
        return jsonify({
            'success': True,
            'message': 'Inscrição realizada com sucesso!',
            'pdf_url': url_for('routes.baixar_comprovante', oficina_id=oficina.id)
        })
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"❌ ERRO ao realizar inscrição: {e}", exc_info=True)
        return jsonify({
            'success': False,
            'message': f'Erro ao realizar inscrição: {str(e)}'
        })

@routes.route('/remover_inscricao/<int:oficina_id>', methods=['POST'])
@login_required
def remover_inscricao(oficina_id):
    inscricao = Inscricao.query.filter_by(usuario_id=current_user.id, oficina_id=oficina_id).first()
    if not inscricao:
        flash('Você não está inscrito nesta oficina!', 'warning')
        return redirect(url_for('routes.dashboard_participante'))

    oficina = Oficina.query.get(oficina_id)
    if oficina:
        oficina.vagas += 1

    db.session.delete(inscricao)
    db.session.commit()
    flash('Inscrição removida com sucesso!', 'success')
    return redirect(url_for('routes.dashboard_participante'))


# ===========================
#   COMPROVANTE DE INSCRIÇÃO (PDF)
# ===========================
@routes.route('/leitor_checkin', methods=['GET'])
@login_required
def leitor_checkin():
    print("➡️ Entrou em /leitor_checkin")

    token = request.args.get('token')
    print(f"➡️ token recebido: {token}")

    if not token:
        mensagem = "Token não fornecido ou inválido."
        print(f"➡️ Erro: {mensagem}")
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'status': 'erro', 'mensagem': mensagem}), 400
        flash(mensagem, "danger")
        return redirect(url_for('routes.dashboard'))

    inscricao = Inscricao.query.filter_by(qr_code_token=token).first()
    print(f"➡️ Inscricao encontrada: {inscricao} (ID: {inscricao.id if inscricao else 'None'})")

    if not inscricao:
        mensagem = "Inscrição não encontrada para este token."
        print(f"➡️ Erro: {mensagem}")
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'status': 'erro', 'mensagem': mensagem}), 404
        flash(mensagem, "danger")
        return redirect(url_for('routes.dashboard'))

    # Verifica se é check-in de evento ou oficina
    if inscricao.evento_id:
        print(f"➡️ Esta inscrição pertence a um EVENTO (ID={inscricao.evento_id})")
        checkin_existente = Checkin.query.filter_by(
            usuario_id=inscricao.usuario_id,
            evento_id=inscricao.evento_id
        ).first()
        if checkin_existente:
            mensagem = "Check-in de evento já foi realizado!"
            print(f"➡️ Aviso: {mensagem}")
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'status': 'repetido', 'mensagem': mensagem}), 200
            flash(mensagem, "warning")
            return redirect(url_for('routes.dashboard'))

        novo_checkin = Checkin(
            usuario_id=inscricao.usuario_id,
            evento_id=inscricao.evento_id,
            palavra_chave="QR-EVENTO"
        )
        dados_checkin = {
            'participante': inscricao.usuario.nome,
            'evento': inscricao.evento.nome,
            'data_hora': novo_checkin.data_hora.strftime('%d/%m/%Y %H:%M:%S')
        }
        print(f"➡️ Check-in de EVENTO criado. Usuario={inscricao.usuario_id}, Evento={inscricao.evento_id}")

    elif inscricao.oficina_id:
        print(f"➡️ Esta inscrição pertence a uma OFICINA (ID={inscricao.oficina_id})")
        checkin_existente = Checkin.query.filter_by(
            usuario_id=inscricao.usuario_id,
            oficina_id=inscricao.oficina_id
        ).first()
        if checkin_existente:
            mensagem = "Check-in da oficina já foi realizado!"
            print(f"➡️ Aviso: {mensagem}")
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'status': 'repetido', 'mensagem': mensagem}), 200
            flash(mensagem, "warning")
            return redirect(url_for('routes.dashboard'))

        novo_checkin = Checkin(
            usuario_id=inscricao.usuario_id,
            oficina_id=inscricao.oficina_id,
            palavra_chave="QR-OFICINA"
        )
        dados_checkin = {
            'participante': inscricao.usuario.nome,
            'oficina': inscricao.oficina.titulo,
            'data_hora': novo_checkin.data_hora.strftime('%d/%m/%Y %H:%M:%S')
        }
        print(f"➡️ Check-in de OFICINA criado. Usuario={inscricao.usuario_id}, Oficina={inscricao.oficina_id}")

    else:
        mensagem = "Inscrição sem evento ou oficina vinculada."
        print(f"➡️ Erro: {mensagem}")
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'status': 'erro', 'mensagem': mensagem}), 400
        flash(mensagem, "danger")
        return redirect(url_for('routes.dashboard'))

    db.session.add(novo_checkin)
    db.session.commit()
    print(f"➡️ Check-in salvo no banco: ID={novo_checkin.id}")

    # Emitir via Socket.IO (se você usa Socket.IO)
    socketio.emit('novo_checkin', dados_checkin, broadcast=True)
    print(f"➡️ Socket.IO emit -> {dados_checkin}")

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        print("➡️ Retornando JSON para AJAX")
        return jsonify({'status': 'ok', **dados_checkin})

    flash("Check-in realizado com sucesso!", "success")
    print("➡️ Check-in concluído e flash exibido, redirecionando para dashboard.")
    return redirect(url_for('routes.dashboard'))

@routes.route('/baixar_comprovante/<int:oficina_id>')
@login_required
def baixar_comprovante(oficina_id):
    oficina = Oficina.query.get(oficina_id)
    if not oficina:
        flash('Oficina não encontrada!', 'danger')
        return redirect(url_for('routes.dashboard_participante'))

    # Busca a inscrição do usuário logado nessa oficina
    inscricao = Inscricao.query.filter_by(usuario_id=current_user.id, oficina_id=oficina.id).first()
    if not inscricao:
        flash('Você não está inscrito nesta oficina.', 'danger')
        return redirect(url_for('routes.dashboard_participante'))

    # Agora chamamos a função com o parâmetro adicional "inscricao"
    pdf_path = gerar_comprovante_pdf(current_user, oficina, inscricao)
    return send_file(pdf_path, as_attachment=True)


# ===========================
# GERAÇÃO DE PDFs (Inscritos, Lista de Frequência, Certificados, Check-ins, Oficina)
# ===========================

def gerar_lista_frequencia_pdf(oficina, pdf_path):
    """
    Generates a modern and professional attendance list PDF for a workshop.
    
    Args:
        oficina: The workshop object containing all relevant information
        pdf_path: The file path where the PDF will be saved
    """
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.units import mm, inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    import os
    from datetime import datetime

    # Create custom styles
    styles = getSampleStyleSheet()
    
    # Custom title style
    title_style = ParagraphStyle(
        name='CustomTitle',
        parent=styles['Title'],
        fontSize=16,
        textColor=colors.HexColor("#023E8A"),
        spaceAfter=10,
        alignment=TA_CENTER
    )
    
    # Custom heading styles
    heading_style = ParagraphStyle(
        name='CustomHeading',
        parent=styles['Heading2'],
        fontSize=12,
        textColor=colors.HexColor("#023E8A"),
        spaceBefore=12,
        spaceAfter=6
    )
    
    # Custom normal text style
    normal_style = ParagraphStyle(
        name='CustomNormal',
        parent=styles['Normal'],
        fontSize=10,
        leading=14,
        spaceBefore=6,
        spaceAfter=6
    )
    
    # Info style for workshop details
    info_style = ParagraphStyle(
        name='InfoStyle',
        parent=styles['Normal'],
        fontSize=10,
        leading=14,
        leftIndent=5 * mm,
        textColor=colors.HexColor("#444444")
    )
    
    # Setup document with proper margins
    doc = SimpleDocTemplate(
        pdf_path,
        pagesize=letter,
        leftMargin=20 * mm,
        rightMargin=20 * mm,
        topMargin=20 * mm,
        bottomMargin=20 * mm
    )
    
    elements = []
    
    # Add header with logo (if available)
    logo_path = os.path.join("static", "logos", "company_logo.png")
    if os.path.exists(logo_path):
        elements.append(Image(logo_path, width=50 * mm, height=15 * mm, hAlign='CENTER'))
        elements.append(Spacer(1, 5 * mm))
    
    # Add title and current date
    current_date = datetime.now().strftime("%d/%m/%Y")
    elements.append(Paragraph(f"LISTA DE FREQUÊNCIA", title_style))
    elements.append(Paragraph(f"<i>Gerado em {current_date}</i>", ParagraphStyle(
        name='date_style', parent=normal_style, alignment=TA_CENTER, fontSize=8, textColor=colors.gray
    )))
    elements.append(Spacer(1, 10 * mm))
    
    # Workshop information in a visually appealing box
    workshop_info = [
        [Paragraph("<b>INFORMAÇÕES DA OFICINA</b>", ParagraphStyle(
            name='workshop_header_style', parent=heading_style, textColor=colors.white, alignment=TA_CENTER
        ))]
    ]
    
    workshop_info_table = Table(workshop_info, colWidths=[doc.width])
    workshop_info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#023E8A")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('ROUNDEDCORNERS', [5, 5, 5, 5]),
    ]))
    elements.append(workshop_info_table)
    elements.append(Spacer(1, 2 * mm))
    
    # Workshop details
    elements.append(Paragraph(f"<b>Título:</b> {oficina.titulo}", info_style))
    
    ministrante_nome = oficina.ministrante_obj.nome if oficina.ministrante_obj else 'N/A'
    elements.append(Paragraph(f"<b>Ministrante:</b> {ministrante_nome}", info_style))
    
    elements.append(Paragraph(f"<b>Local:</b> {oficina.cidade}, {oficina.estado}", info_style))
    
    elements.append(Paragraph("<b>Carga Horária:</b> {0} horas".format(oficina.carga_horaria), info_style))
    
    # Dates and times
    if oficina.dias:
        elements.append(Paragraph("<b>Datas e Horários:</b>", info_style))
        
        dates_data = []
        for dia in oficina.dias:
            data_formatada = dia.data.strftime('%d/%m/%Y')
            horario = f"{dia.horario_inicio} às {dia.horario_fim}"
            dates_data.append([
                Paragraph(data_formatada, normal_style),
                Paragraph(horario, normal_style)
            ])
        
        if dates_data:
            dates_table = Table(dates_data, colWidths=[doc.width/2 - 10*mm, doc.width/2 - 10*mm])
            dates_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor("#F8F9FA")),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('LEFTPADDING', (0, 0), (-1, -1), 6),
                ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ]))
            elements.append(dates_table)
    else:
        elements.append(Paragraph("<b>Datas:</b> Nenhuma data registrada", info_style))
    
    elements.append(Spacer(1, 15 * mm))
    
    # Attendance list header
    elements.append(Paragraph("LISTA DE PRESENÇA", heading_style))
    elements.append(Spacer(1, 5 * mm))
    
    # Attendance table with signature column
    table_data = [
        [
            Paragraph("<b>Nº</b>", normal_style),
            Paragraph("<b>Nome Completo</b>", normal_style),
            Paragraph("<b>Assinatura</b>", normal_style)
        ]
    ]
    
    # Add rows for each participant
    for i, inscricao in enumerate(oficina.inscritos, 1):
        table_data.append([
            Paragraph(str(i), normal_style),
            Paragraph(inscricao.usuario.nome, normal_style),
            ""  # Signature space
        ])
    
    # Add empty rows if needed (to ensure at least 15 rows)
    current_rows = len(table_data) - 1  # Exclude header
    if current_rows < 15:
        for i in range(current_rows + 1, 16):
            table_data.append([
                Paragraph(str(i), normal_style),
                "",  # Empty name
                ""   # Signature space
            ])
    
    # Create the table with appropriate width distribution
    table = Table(table_data, colWidths=[15*mm, 85*mm, 70*mm])
    
    # Apply styles to the table
    table.setStyle(TableStyle([
        # Header row styling
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#023E8A")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        
        # Data rows styling
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # Center the numbers
        ('ALIGN', (1, 1), (1, -1), 'LEFT'),    # Left-align the names
        
        # Grid styling
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
        ('BOX', (0, 0), (-1, -1), 1, colors.HexColor("#023E8A")),
        
        # Alternating row colors for better readability
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8F9FA")]),
        
        # Cell padding
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        
        # Line height for signature spaces
        ('LINEBELOW', (2, 1), (2, -1), 0.5, colors.HexColor("#AAAAAA")),
    ]))
    
    elements.append(table)
    
    # Footer with signature fields
    elements.append(Spacer(1, 30 * mm))
    
    # Create signature lines
    signature_data = [
        [
            Paragraph("_______________________________", ParagraphStyle(name="signature_line", parent=normal_style, alignment=TA_CENTER)),
            "",
            Paragraph("_______________________________", ParagraphStyle(name="signature_line", parent=normal_style, alignment=TA_CENTER))
        ],
        [
            Paragraph("Ministrante", ParagraphStyle(name="signature_label", parent=normal_style, alignment=TA_CENTER)),
            "",
            Paragraph("Coordenador", ParagraphStyle(name="signature_label", parent=normal_style, alignment=TA_CENTER))
        ]
    ]
    
    signature_table = Table(signature_data, colWidths=[doc.width/3, doc.width/3, doc.width/3])
    signature_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    elements.append(signature_table)
    
    # Add page numbers
    def add_page_number(canvas, doc):
        page_num = canvas.getPageNumber()
        text = f"Página {page_num}"
        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(colors.grey)
        canvas.drawRightString(
            doc.pagesize[0] - doc.rightMargin, 
            doc.bottomMargin/2, 
            text
        )
    
    # Build the PDF with page numbers
    doc.build(elements, onFirstPage=add_page_number, onLaterPages=add_page_number)    
    
@routes.route('/gerar_pdf_inscritos/<int:oficina_id>', methods=['GET'])
@login_required
def gerar_pdf_inscritos_pdf(oficina_id):
    """
    Gera um PDF com a lista de inscritos para uma oficina específica,
    com layout moderno e organizado.
    """
    # Importações necessárias
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm, cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
    from reportlab.platypus import PageBreak, Flowable
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    import os
    from flask import send_file
    from datetime import datetime
    
    # Busca a oficina no banco de dados
    oficina = Oficina.query.get_or_404(oficina_id)
    
    # Preparar o diretório para salvar o PDF
    pdf_filename = f"inscritos_oficina_{oficina.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    diretorio = os.path.join("static", "comprovantes")
    os.makedirs(diretorio, exist_ok=True)
    pdf_path = os.path.join(diretorio, pdf_filename)

    # Configurar estilos personalizados
    styles = getSampleStyleSheet()
    
    # Estilo de título modernizado
    title_style = ParagraphStyle(
        name='CustomTitle',
        parent=styles['Title'],
        fontSize=18,
        alignment=TA_CENTER,
        spaceAfter=6 * mm,
        fontName='Helvetica-Bold',
        textColor=colors.HexColor('#023E8A')
    )
    
    # Estilo para subtítulos
    subtitle_style = ParagraphStyle(
        name='CustomSubtitle',
        parent=styles['Heading2'],
        fontSize=14,
        alignment=TA_LEFT,
        spaceAfter=3 * mm,
        fontName='Helvetica-Bold',
        textColor=colors.HexColor('#0077B6')
    )
    
    # Estilo para texto normal
    normal_style = ParagraphStyle(
        name='CustomNormal',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=2 * mm,
        fontName='Helvetica'
    )
    
    # Estilo para texto em tabelas (para permitir quebra de linha)
    table_text_style = ParagraphStyle(
        name='TableText',
        parent=styles['Normal'],
        fontSize=9,
        fontName='Helvetica',
        leading=12,
        wordWrap='CJK'
    )
    
    # Estilo para rodapé
    footer_style = ParagraphStyle(
        name='Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.darkgrey,
        alignment=TA_CENTER
    )

    # Crie uma classe personalizada para linha horizontal
    class HorizontalLine(Flowable):
        def __init__(self, width, thickness=1):
            Flowable.__init__(self)
            self.width = width
            self.thickness = thickness
        
        def draw(self):
            self.canv.setStrokeColor(colors.HexColor('#0077B6'))
            self.canv.setLineWidth(self.thickness)
            self.canv.line(0, 0, self.width, 0)
    
    # Criar um documento PDF
    doc = SimpleDocTemplate(
        pdf_path, 
        pagesize=A4,
        leftMargin=2.5*cm, 
        rightMargin=2.5*cm, 
        topMargin=2*cm, 
        bottomMargin=2*cm
    )
    
    # Lista para armazenar elementos do PDF
    elements = []
    
    # Verificar se há um logo personalizado para o cliente
    # Se a oficina estiver associada a um cliente e o cliente tiver um logo
    logo_path = None
    if oficina.cliente_id:
        cliente = Cliente.query.get(oficina.cliente_id)
        if cliente and hasattr(cliente, 'logo_certificado') and cliente.logo_certificado:
            logo_path = cliente.logo_certificado
    
    # Se encontrou logo personalizado, adiciona ao PDF
    if logo_path and os.path.exists(logo_path):
        logo = Image(logo_path)
        logo.drawHeight = 2 * cm
        logo.drawWidth = 5 * cm
        elements.append(logo)
        elements.append(Spacer(1, 5 * mm))
    
    # Título principal
    elements.append(Paragraph(f"Lista de Inscritos", title_style))
    elements.append(Paragraph(f"{oficina.titulo}", subtitle_style))
    elements.append(HorizontalLine(doc.width))
    elements.append(Spacer(1, 5 * mm))
    
    # Informações da oficina em formato mais elegante
    elements.append(Paragraph("<b>Detalhes da Oficina</b>", subtitle_style))
    
    ministrante_nome = oficina.ministrante_obj.nome if oficina.ministrante_obj else 'Não atribuído'
    elements.append(Paragraph(f"<b>Ministrante:</b> {ministrante_nome}", normal_style))
    elements.append(Paragraph(f"<b>Local:</b> {oficina.cidade}, {oficina.estado}", normal_style))
    elements.append(Paragraph(f"<b>Carga Horária:</b> {oficina.carga_horaria} horas", normal_style))
    
    # Criar uma tabela para as datas e horários se houver dados
    if oficina.dias and len(oficina.dias) > 0:
        elements.append(Paragraph("<b>Datas e Horários:</b>", normal_style))
        
        date_data = [["Data", "Início", "Término"]]
        for dia in oficina.dias:
            data_formatada = dia.data.strftime('%d/%m/%Y')
            # Convertendo os valores para Paragraph para permitir quebra de linha
            date_data.append([
                Paragraph(data_formatada, table_text_style),
                Paragraph(dia.horario_inicio, table_text_style),
                Paragraph(dia.horario_fim, table_text_style)
            ])
        
        date_table = Table(date_data, colWidths=[doc.width * 0.4, doc.width * 0.3, doc.width * 0.3])
        date_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#EBF2FA')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#023E8A')),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            # Define que o texto pode quebrar dentro das células
            ('WORDWRAP', (0, 0), (-1, -1), True),
        ]))
        elements.append(date_table)
    else:
        elements.append(Paragraph("<b>Datas:</b> Nenhuma data registrada", normal_style))
    
    elements.append(Spacer(1, 8 * mm))
    elements.append(HorizontalLine(doc.width))
    elements.append(Spacer(1, 8 * mm))
    
    # Adicionar contador de inscritos
    total_inscritos = len(oficina.inscritos) if oficina.inscritos else 0
    elements.append(Paragraph(f"<b>Total de Inscritos:</b> {total_inscritos}", subtitle_style))
    elements.append(Spacer(1, 5 * mm))
    
    # Tabela de inscritos com estilo moderno
    if oficina.inscritos and len(oficina.inscritos) > 0:
        table_data = [["#", "Nome", "CPF", "E-mail"]]
        
        for idx, inscricao in enumerate(oficina.inscritos, 1):
            # Verifica se é um objeto mapeado ou um objeto de modelo regular
            if hasattr(inscricao, 'usuario'):
                nome = inscricao.usuario.nome
                cpf = inscricao.usuario.cpf
                email = inscricao.usuario.email
            else:
                nome = inscricao.get('nome', 'N/A')
                cpf = inscricao.get('cpf', 'N/A')
                email = inscricao.get('email', 'N/A')
                
            # Formatação de CPF se necessário (adicionar pontos e traço)
            if cpf and len(cpf) == 11 and cpf.isdigit():
                cpf = f"{cpf[:3]}.{cpf[3:6]}.{cpf[6:9]}-{cpf[9:]}"
                
            # Usando Paragraph para permitir quebra de linha em cada coluna
            table_data.append([
                Paragraph(str(idx), table_text_style),
                Paragraph(nome, table_text_style),
                Paragraph(cpf, table_text_style),
                Paragraph(email, table_text_style)
            ])
        
        # Definir larguras das colunas para melhor distribuição
        col_widths = [doc.width * 0.05, doc.width * 0.35, doc.width * 0.25, doc.width * 0.35]
        
        # Criar tabela com estilo moderno
        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            # Cabeçalho com cor de fundo
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#023E8A')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),  # Centraliza a coluna de números
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),    # Alinha nomes à esquerda
            ('ALIGN', (2, 0), (2, -1), 'CENTER'),  # Centraliza CPFs
            ('ALIGN', (3, 0), (3, -1), 'LEFT'),    # Alinha e-mails à esquerda
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            # Linhas alternadas para melhor legibilidade
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
            # Bordas mais sutis
            ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
            # Configuração para permitir quebra de linha
            ('WORDWRAP', (0, 0), (-1, -1), True),
        ]))
        elements.append(table)
    else:
        elements.append(Paragraph("Não há inscritos nesta oficina.", normal_style))
    
    # Adiciona espaço para assinatura
    elements.append(Spacer(1, 2 * cm))
    elements.append(HorizontalLine(doc.width * 0.4))
    elements.append(Paragraph("Assinatura do Coordenador", footer_style))
    
    # Adiciona rodapé com data de geração
    elements.append(Spacer(1, 2 * cm))
    current_date = datetime.now().strftime("%d/%m/%Y %H:%M")
    elements.append(HorizontalLine(doc.width))
    elements.append(Spacer(1, 3 * mm))
    elements.append(Paragraph(f"Documento gerado em {current_date} | AppFiber", footer_style))
    
    # Construir o PDF
    doc.build(elements)
    
    # Retorna o arquivo para download
    return send_file(pdf_path, as_attachment=True, download_name=pdf_filename)

    
@routes.route('/gerar_lista_frequencia/<int:oficina_id>')
@login_required
def gerar_lista_frequencia(oficina_id, pdf_path=None):
    """
    Generates a modern and professional attendance list PDF for a workshop.
    
    Args:
        oficina_id: The ID of the workshop
        pdf_path: The file path where the PDF will be saved (optional)
    """
    # Importações necessárias
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    import os
    from datetime import datetime
    from flask import send_file
    from models import Oficina, Inscricao

    # Obter a oficina real pelo ID
    oficina = Oficina.query.get(oficina_id)
    if not oficina:
        raise ValueError("Oficina não encontrada!")

    # Se pdf_path não for fornecido, gere um caminho padrão
    if pdf_path is None:
        import tempfile
        pdf_path = os.path.join(tempfile.gettempdir(), f"lista_frequencia_{oficina_id}.pdf")

    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        name='CustomTitle',
        parent=styles['Title'],
        fontSize=16,
        textColor=colors.HexColor("#023E8A"),
        spaceAfter=10,
        alignment=TA_CENTER
    )

    doc = SimpleDocTemplate(pdf_path, pagesize=letter)
    elements = []

    # Cabeçalho
    elements.append(Paragraph("LISTA DE FREQUÊNCIA", title_style))
    elements.append(Paragraph(f"<i>{oficina.titulo}</i>", styles['Normal']))
    elements.append(Spacer(1, 12))

    # Dados da oficina
    ministrante_nome = oficina.ministrante_obj.nome if oficina.ministrante_obj else 'N/A'
    elements.append(Paragraph(f"<b>Ministrante:</b> {ministrante_nome}", styles['Normal']))
    elements.append(Paragraph(f"<b>Local:</b> {oficina.cidade}, {oficina.estado}", styles['Normal']))
    elements.append(Paragraph(f"<b>Carga Horária:</b> {oficina.carga_horaria} horas", styles['Normal']))
    elements.append(Spacer(1, 12))

    # Tabela de frequência
    table_data = [["Nº", "Nome Completo", "Assinatura"]]

    # Buscando participantes reais inscritos
    inscricoes = Inscricao.query.filter_by(oficina_id=oficina_id).all()

    for i, inscricao in enumerate(inscricoes, 1):
        nome_participante = inscricao.usuario.nome if inscricao.usuario else 'N/A'
        # Observe que usamos Paragraph para permitir a quebra de linha no nome
        table_data.append([
            Paragraph(str(i), styles['Normal']),
            Paragraph(nome_participante, styles['Normal']),
            ""
        ])

    table = Table(table_data, colWidths=[30*mm, 100*mm, 60*mm])

    # Aplica estilos, incluindo WORDWRAP
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#023E8A")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8F9FA")]),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 5),
        ('RIGHTPADDING', (0, 0), (-1, -1), 5),
        ('WORDWRAP', (0, 0), (-1, -1), True),
    ]))

    elements.append(table)
    elements.append(Spacer(1, 12))

    # Rodapé para assinaturas
    elements.append(Spacer(1, 24))
    elements.append(Paragraph("________________________", styles['Normal']))
    elements.append(Paragraph("Ministrante", styles['Normal']))

    elements.append(Spacer(1, 48))
    elements.append(Paragraph("________________________", styles['Normal']))
    elements.append(Paragraph("Coordenador", styles['Normal']))

    # Gera PDF
    doc.build(elements)

    return send_file(pdf_path)


@routes.route('/gerar_certificados/<int:oficina_id>', methods=['GET'])
@login_required
def gerar_certificados(oficina_id):
    if current_user.tipo not in ['admin', 'cliente']:
        flash("Apenas administradores podem gerar certificados.", "danger")
        

    oficina = Oficina.query.get(oficina_id)
    if not oficina:
        flash("Oficina não encontrada!", "danger")
        return redirect(url_for('routes.dashboard'))

    inscritos = oficina.inscritos
    if not inscritos:
        flash("Não há inscritos nesta oficina para gerar certificados!", "warning")
        return redirect(url_for('routes.dashboard'))

    pdf_path = f"static/certificados/certificados_oficina_{oficina.id}.pdf"
    os.makedirs(os.path.dirname(pdf_path), exist_ok=True)

    # Agora chama a função ajustada
    gerar_certificados_pdf(oficina, inscritos, pdf_path)

    flash("Certificados gerados com sucesso!", "success")
    return send_file(pdf_path, as_attachment=True)


@routes.route('/checkin/<int:oficina_id>', methods=['GET', 'POST'])
@login_required
def checkin(oficina_id):
    oficina = Oficina.query.get_or_404(oficina_id)
    cliente_id_oficina = oficina.cliente_id

    config_cliente = ConfiguracaoCliente.query.filter_by(cliente_id=cliente_id_oficina).first()
    if not config_cliente or not config_cliente.permitir_checkin_global:
        flash("Check-in indisponível para esta oficina!", "danger")
        return redirect(url_for('routes.dashboard_participante'))

    if request.method == 'POST':
        palavra_escolhida = request.form.get('palavra_escolhida')
        if not palavra_escolhida:
            flash("Selecione uma opção de check-in.", "danger")
            return redirect(url_for('routes.checkin', oficina_id=oficina_id))

        inscricao = Inscricao.query.filter_by(usuario_id=current_user.id, oficina_id=oficina.id).first()
        if not inscricao:
            flash("Você não está inscrito nesta oficina!", "danger")
            return redirect(url_for('routes.checkin', oficina_id=oficina_id))

        if inscricao.checkin_attempts >= 2:
            flash("Você excedeu o número de tentativas de check-in.", "danger")
            return redirect(url_for('routes.dashboard_participante'))

        # ❗Verificar se já existe check-in anterior
        checkin_existente = Checkin.query.filter_by(usuario_id=current_user.id, oficina_id=oficina.id).first()
        if checkin_existente:
            flash("Você já realizou o check-in!", "warning")
            return redirect(url_for('routes.dashboard_participante'))

        if palavra_escolhida.strip() != oficina.palavra_correta.strip():
            inscricao.checkin_attempts += 1
            db.session.commit()
            flash("Palavra-chave incorreta!", "danger")
            return redirect(url_for('routes.checkin', oficina_id=oficina_id))

        checkin = Checkin(
            usuario_id=current_user.id,
            oficina_id=oficina.id,
            palavra_chave=palavra_escolhida,
            cliente_id=oficina.cliente_id,
            evento_id=oficina.evento_id
        )
        db.session.add(checkin)
        db.session.commit()
        flash("Check-in realizado com sucesso!", "success")
        return redirect(url_for('routes.dashboard_participante'))

    opcoes = oficina.opcoes_checkin.split(',') if oficina.opcoes_checkin else []
    return render_template('checkin.html', oficina=oficina, opcoes=opcoes)




@routes.route('/oficina/<int:oficina_id>/checkins', methods=['GET'])
@login_required
def lista_checkins(oficina_id):
    if current_user.tipo not in ['admin', 'cliente']:
        flash("Acesso Autorizado!", "danger")

    oficina = Oficina.query.get_or_404(oficina_id)
    checkins = Checkin.query.filter_by(oficina_id=oficina_id).all()

    usuarios_checkin = [{
        'nome': checkin.usuario.nome,
        'cpf': checkin.usuario.cpf,
        'email': checkin.usuario.email,
        'data_hora': checkin.data_hora
    } for checkin in checkins]

    # Coleta todos os ministrantes associados à oficina
    ministrantes = [m.nome for m in oficina.ministrantes_associados] if oficina.ministrantes_associados else []

    return render_template(
        'lista_checkins.html',
        oficina=oficina,
        usuarios_checkin=usuarios_checkin,
        ministrantes=ministrantes
    )


@routes.route('/gerar_pdf_checkins/<int:oficina_id>', methods=['GET'])
@login_required
def gerar_pdf_checkins(oficina_id):
    """
    Gera um PDF moderno e bem estruturado com a lista de check-ins para uma oficina específica.
    
    Args:
        oficina_id (int): ID da oficina para gerar o relatório de check-ins
        
    Returns:
        Flask response: Arquivo PDF para download
    """
    import pytz
    from datetime import datetime
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, letter, inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Image, LongTable
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from flask import send_file
    import os
    
    # Busca dados necessários do banco de dados
    oficina = Oficina.query.get_or_404(oficina_id)
    checkins = Checkin.query.filter_by(oficina_id=oficina_id).order_by(Checkin.data_hora.desc()).all()
    dias = OficinaDia.query.filter_by(oficina_id=oficina_id).order_by(OficinaDia.data).all()
    
    # Cria pasta se não existir
    output_dir = os.path.join("static", "relatorios")
    os.makedirs(output_dir, exist_ok=True)
    
    # Define o caminho do arquivo PDF
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    pdf_filename = f"checkins_oficina_{oficina.id}_{timestamp}.pdf"
    pdf_path = os.path.join(output_dir, pdf_filename)
    
    # Configuração do timezone para Brasil
    brasil_tz = pytz.timezone("America/Sao_Paulo")
    
    def formatar_data_hora(dt):
        """Converte datetime para timezone Brasil e formata adequadamente"""
        if dt.tzinfo is None:
            dt = pytz.utc.localize(dt)
        dt_local = dt.astimezone(brasil_tz)
        return dt_local.strftime("%d/%m/%Y %H:%M:%S")
    
    # Configuração de estilos
    styles = getSampleStyleSheet()
    
    # Estilo para título principal
    titulo_style = ParagraphStyle(
        name="TituloPrincipal",
        parent=styles["Heading1"],
        fontSize=20,
        alignment=TA_CENTER,
        spaceBefore=12,
        spaceAfter=24,
        textColor=colors.HexColor("#023E8A")
    )
    
    # Estilo para subtítulos
    subtitulo_style = ParagraphStyle(
        name="Subtitulo",
        parent=styles["Heading2"],
        fontSize=16,
        spaceBefore=12,
        spaceAfter=6,
        textColor=colors.HexColor("#0077B6")
    )
    
    # Estilo para informações
    info_style = ParagraphStyle(
        name="Info",
        parent=styles["Normal"],
        fontSize=12,
        spaceAfter=3,
        leading=14
    )
    
    # Estilo para rodapé
    footer_style = ParagraphStyle(
        name="Footer",
        parent=styles["Normal"],
        fontSize=9,
        textColor=colors.gray,
        alignment=TA_CENTER
    )
    
    # Configuração do documento
    doc = SimpleDocTemplate(
        pdf_path,
        pagesize=landscape(letter),
        leftMargin=0.5*inch,
        rightMargin=0.5*inch,
        topMargin=0.5*inch,
        bottomMargin=0.5*inch,
        title=f"Check-ins - {oficina.titulo}",
        author="AppFiber"
    )
    
    # Lista para elementos do PDF
    elementos = []
    
    # ==== CABEÇALHO DO DOCUMENTO ====
    elementos.append(Paragraph(f"Relatório de Check-ins", titulo_style))
    
    # ==== INFORMAÇÕES DA OFICINA ====
    elementos.append(Paragraph("Informações da Oficina", subtitulo_style))
    
    # Tabela de informações da oficina
    ministrante_nome = oficina.ministrante_obj.nome if oficina.ministrante_obj else 'Não atribuído'
    cliente_nome = oficina.cliente.nome if oficina.cliente else 'Administração'
    
    info_data = [
        ["Título:", oficina.titulo],
        ["Ministrante:", ministrante_nome],
        ["Local:", f"{oficina.cidade}, {oficina.estado}"],
        ["Cliente:", cliente_nome],
        ["Vagas totais:", str(oficina.vagas)],
        ["Check-ins registrados:", str(len(checkins))]
    ]
    
    # Calcular a taxa de participação
    total_inscritos = len(oficina.inscritos) if oficina.inscritos else 0
    if total_inscritos > 0:
        taxa_checkin = (len(checkins) / total_inscritos) * 100
        info_data.append(["Taxa de participação:", f"{taxa_checkin:.1f}%"])
    
    # Tabela de informações
    info_table = Table(info_data, colWidths=[2*inch, 5*inch])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor("#f8f9fa")),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor("#0077B6")),
        ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 7),
        ('TOPPADDING', (0, 0), (-1, -1), 7),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
    ]))
    
    elementos.append(info_table)
    elementos.append(Spacer(1, 0.2*inch))
    
    # ==== DATAS E HORÁRIOS ====
    if dias:
        elementos.append(Paragraph("Cronograma da Oficina", subtitulo_style))
        
        # Tabela com as datas e horários
        cronograma_data = [["Data", "Horário de Início", "Horário de Término"]]
        
        for dia in dias:
            cronograma_data.append([
                dia.data.strftime('%d/%m/%Y'),
                dia.horario_inicio,
                dia.horario_fim
            ])
        
        cronograma_table = Table(cronograma_data, colWidths=[2.5*inch, 2.5*inch, 2.5*inch])
        cronograma_table.setStyle(TableStyle([
            # Cabeçalho
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#023E8A")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            
            # Corpo da tabela
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor("#f8f9fa")),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('GRID', (0, 0), (-1, -1), 1, colors.lightgrey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 7),
            ('TOPPADDING', (0, 0), (-1, -1), 7),
        ]))
        
        elementos.append(cronograma_table)
        elementos.append(Spacer(1, 0.3*inch))
    
    # ==== LISTA DE CHECK-INS ====
    if checkins:
        elementos.append(Paragraph("Lista de Check-ins Realizados", subtitulo_style))
        
        # Cabeçalho da tabela de check-ins
        checkins_data = [["Nome", "CPF", "E-mail", "Data e Hora", "Método"]]
        
        # Preparar dados para tabela de check-ins
        for checkin in checkins:
            # Determinar o método de check-in
            metodo = "QR Code" if checkin.palavra_chave == "QR-AUTO" else "Manual"
            
            checkins_data.append([
                checkin.usuario.nome,
                checkin.usuario.cpf,
                checkin.usuario.email,
                formatar_data_hora(checkin.data_hora),
                metodo
            ])
        
        # Criar tabela de check-ins com suporte a múltiplas páginas
        checkins_table = LongTable(
            checkins_data, 
            colWidths=[2.5*inch, 1.5*inch, 2.5*inch, 1.7*inch, 1*inch],
            repeatRows=1
        )
        
        checkins_table.setStyle(TableStyle([
            # Cabeçalho
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#023E8A")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            
            # Corpo da tabela
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),   # Nome alinhado à esquerda
            ('ALIGN', (1, 1), (1, -1), 'CENTER'), # CPF centralizado
            ('ALIGN', (2, 1), (2, -1), 'LEFT'),   # Email alinhado à esquerda
            ('ALIGN', (3, 1), (3, -1), 'CENTER'), # Data centralizada
            ('ALIGN', (4, 1), (4, -1), 'CENTER'), # Método centralizado
            
            # Alternar cores das linhas (usando ROWBACKGROUNDS para aplicar automaticamente)
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8f9fa")]),
            
            # Bordas e formatação
            ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            
            # Destaque para QR Code
            ('TEXTCOLOR', (4, 1), (4, -1), colors.HexColor("#0077B6")),
            ('FONTNAME', (4, 1), (4, -1), 'Helvetica-Bold'),
        ]))
        
        # Linhas alternadas em cor diferente
        for i in range(1, len(checkins_data)):
            if i % 2 == 0:
                checkins_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, i), (-1, i), colors.HexColor("#f8f9fa"))
                ]))
        
        elementos.append(checkins_table)
    else:
        elementos.append(Paragraph("Nenhum check-in registrado para esta oficina.", info_style))
    
    # ==== RODAPÉ E INFORMAÇÕES FINAIS ====
    elementos.append(Spacer(1, 0.5*inch))
    
    # Adiciona data e hora de geração do relatório
    data_geracao = datetime.now(brasil_tz).strftime("%d/%m/%Y %H:%M:%S")
    elementos.append(Paragraph(f"Relatório gerado em {data_geracao}", footer_style))
    elementos.append(Paragraph("AppFiber - Sistema de Gestão Educacional", footer_style))
    
    # Constrói o documento
    doc.build(elementos)
    
    return send_file(pdf_path, as_attachment=True, download_name=f"checkins_{oficina.titulo.replace(' ', '_')}.pdf")

@routes.route('/gerar_pdf/<int:oficina_id>')
def gerar_pdf(oficina_id):
    oficina = Oficina.query.get(oficina_id)
    if not oficina:
        flash("Oficina não encontrada!", "danger")
        return redirect(url_for('routes.dashboard'))
    pdf_path = os.path.join("static", "pdfs")
    os.makedirs(pdf_path, exist_ok=True)
    pdf_file = os.path.join(pdf_path, f"oficina_{oficina_id}.pdf")
    c = canvas.Canvas(pdf_file, pagesize=landscape(A4))
    width, height = landscape(A4)
    logo_path = os.path.join("static", "logom.png")
    if os.path.exists(logo_path):
        logo = ImageReader(logo_path)
        c.drawImage(logo, width / 2 - 100, height - 100, width=200, height=80, preserveAspectRatio=True, mask='auto')
    c.setLineWidth(2)
    c.line(50, height - 120, width - 50, height - 120)
    c.setFont("Helvetica-Bold", 36)
    c.setFillColorRGB(0, 0, 0.7)
    c.drawCentredString(width / 2, height - 180, oficina.titulo.upper())
    c.setFont("Helvetica-Bold", 22)
    c.setFillColorRGB(0, 0, 0)
    ministrante_nome = oficina.ministrante_obj.nome if oficina.ministrante_obj else 'N/A'
    c.drawCentredString(width / 2, height - 230, f"Ministrante: {ministrante_nome}")
    c.setLineWidth(1)
    c.line(50, height - 250, width - 50, height - 250)
    c.setFont("Helvetica-Bold", 20)
    c.setFillColorRGB(0.1, 0.1, 0.1)
    c.drawCentredString(width / 2, height - 280, "Datas e Horários")
    c.setFont("Helvetica", 16)
    c.setFillColorRGB(0, 0, 0)
    y_pos = height - 300
    for dia in oficina.dias:
        c.drawCentredString(width / 2, y_pos, f"{dia.data.strftime('%d/%m/%Y')} - {dia.horario_inicio} às {dia.horario_fim}")
        y_pos -= 30
    jornada_path = os.path.join("static", "jornada2025.png")
    if os.path.exists(jornada_path):
        jornada = ImageReader(jornada_path)
        x_centered = (width - 600) / 2
        c.drawImage(jornada, x_centered, 20, width=600, height=240, preserveAspectRatio=True, mask='auto')
    c.save()
    return send_file(pdf_file, as_attachment=True, download_name=f"oficina_{oficina_id}.pdf")


@routes.route('/gerar_certificado_individual_admin', methods=['POST'])
@login_required
def gerar_certificado_individual_admin():
    if current_user.tipo not in ['admin', 'cliente']:
        flash("Acesso Autorizado!", "danger")
        

    oficina_id = request.form.get('oficina_id')
    usuario_id = request.form.get('usuario_id')
    
    if not oficina_id or not usuario_id:
        flash("Oficina ou participante não informado.", "danger")
        return redirect(url_for('routes.dashboard'))

    # Busca a oficina
    oficina = Oficina.query.get(oficina_id)
    if not oficina:
        flash("Oficina não encontrada!", "danger")
        return redirect(url_for('routes.dashboard'))

    # Verifica se o participante está inscrito na oficina
    inscricao = Inscricao.query.filter_by(oficina_id=oficina_id, usuario_id=usuario_id).first()
    if not inscricao:
        flash("O participante não está inscrito nesta oficina!", "danger")
        return redirect(url_for('routes.dashboard'))

    # Define o caminho do PDF e gera o certificado
    pdf_path = f"static/certificados/certificado_{usuario_id}_{oficina_id}_admin.pdf"
    os.makedirs(os.path.dirname(pdf_path), exist_ok=True)

    # Gera o certificado utilizando a função existente; observe que passamos uma lista contendo só essa inscrição
    gerar_certificados_pdf(oficina, [inscricao], pdf_path)

    flash("Certificado individual gerado com sucesso!", "success")
    return send_file(pdf_path, as_attachment=True)


# ===========================
#   SORTEIOS
# ===========================

@routes.route('/criar_sorteio', methods=['GET', 'POST'])
@login_required
def criar_sorteio():
    """
    Rota para criar um novo sorteio.
    GET: Exibe o formulário de criação
    POST: Processa o formulário e cria o sorteio
    """
    # Verificar se o usuário é um cliente
    if current_user.tipo != 'cliente':
        flash('Acesso negado. Apenas clientes podem gerenciar sorteios.', 'danger')
        return redirect(url_for('routes.dashboard'))
    
    # Listar eventos do cliente para o formulário
    eventos = Evento.query.filter_by(cliente_id=current_user.id).all()
    
    if request.method == 'POST':
        titulo = request.form.get('titulo')
        premio = request.form.get('premio')
        descricao = request.form.get('descricao')
        tipo_sorteio = request.form.get('tipo_sorteio')
        evento_id = request.form.get('evento_id')
        oficina_id = request.form.get('oficina_id') if tipo_sorteio == 'oficina' else None
        
        # Validações básicas
        if not titulo or not premio or not evento_id:
            flash('Por favor, preencha todos os campos obrigatórios.', 'danger')
            return render_template('criar_sorteio.html', eventos=eventos)
        
        # Criar novo sorteio
        sorteio = Sorteio(
            titulo=titulo,
            premio=premio,
            descricao=descricao,
            cliente_id=current_user.id,
            evento_id=evento_id,
            oficina_id=oficina_id,
            status='pendente'
        )
        
        try:
            db.session.add(sorteio)
            db.session.commit()
            flash('Sorteio criado com sucesso!', 'success')
            return redirect(url_for('routes.gerenciar_sorteios'))
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao criar sorteio: {str(e)}', 'danger')
    
    return render_template('criar_sorteio.html', eventos=eventos)

@routes.route('/gerenciar_sorteios')
@login_required
def gerenciar_sorteios():
    """
    Rota para listar e gerenciar os sorteios existentes.
    """
    # Verificar se o usuário é um cliente
    if current_user.tipo != 'cliente':
        flash('Acesso negado. Apenas clientes podem gerenciar sorteios.', 'danger')
        return redirect(url_for('routes.dashboard'))
    
    # Filtros da URL
    evento_id = request.args.get('evento_id', type=int)
    status = request.args.get('status')
    
    # Consulta base
    query = Sorteio.query.filter_by(cliente_id=current_user.id)
    
    # Aplicar filtros
    if evento_id:
        query = query.filter_by(evento_id=evento_id)
    if status:
        query = query.filter_by(status=status)
    
    # Ordenar por data, mais recentes primeiro
    sorteios = query.order_by(Sorteio.data_sorteio.desc()).all()
    
    # Listar eventos do cliente para o filtro
    eventos = Evento.query.filter_by(cliente_id=current_user.id).all()
    
    return render_template('gerenciar_sorteios.html', 
                           sorteios=sorteios, 
                           eventos=eventos)

@routes.route('/api/oficinas_evento/<int:evento_id>')
@login_required
def get_oficinas_evento(evento_id):
    """
    API para obter as oficinas de um evento específico.
    """
    try:
        oficinas = Oficina.query.filter_by(evento_id=evento_id).all()
        return jsonify({
            'success': True,
            'oficinas': [{'id': o.id, 'titulo': o.titulo} for o in oficinas]
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@routes.route('/api/participantes_contagem')
@login_required
def get_participantes_contagem():
    """
    API para contar o número de participantes elegíveis para um sorteio.
    """
    tipo = request.args.get('tipo')
    id_param = request.args.get('id', type=int)
    
    
    if not tipo or not id_param:
        return jsonify({'success': False, 'message': 'Parâmetros inválidos'})
    
    try:
        if tipo == 'evento':
            # Contar participantes de um evento - filtrar diretamente pelo evento_id na tabela Usuario
            total = Usuario.query.filter_by(evento_id=id_param).count()
        elif tipo == 'oficina':
            # Contar participantes de uma oficina
            total = Usuario.query.join(
                Inscricao, Usuario.id == Inscricao.usuario_id
            ).filter(
                Inscricao.oficina_id == id_param
            ).count()
        else:
            return jsonify({'success': False, 'message': 'Tipo inválido'})
        
        return jsonify({'success': True, 'total': total})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})
    """
    API para contar o número de participantes elegíveis para um sorteio.
    """
    tipo = request.args.get('tipo')
    id_param = request.args.get('id', type=int)
    
    if not tipo or not id_param:
        return jsonify({'success': False, 'message': 'Parâmetros inválidos'})
    
    try:
        if tipo == 'evento':
            # Contar participantes de um evento
            total = Inscricao.query.filter_by(evento_id=id_param).count()
        elif tipo == 'oficina':
            # Contar participantes de uma oficina
            total = Inscricao.query.filter_by(oficina_id=id_param).count()
        else:
            return jsonify({'success': False, 'message': 'Tipo inválido'})
        
        return jsonify({'success': True, 'total': total})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@routes.route('/api/sorteio/<int:sorteio_id>')
@login_required
def get_sorteio(sorteio_id):
    """
    API para obter detalhes de um sorteio específico.
    """
    try:
        sorteio = Sorteio.query.get_or_404(sorteio_id)
        
        # Verificar se o usuário atual é o dono do sorteio
        if sorteio.cliente_id != current_user.id:
            return jsonify({'success': False, 'message': 'Acesso negado'})
        
        # Verificar se o sorteio foi realizado
        if sorteio.status != 'realizado' or not sorteio.ganhador:
            return jsonify({'success': False, 'message': 'Este sorteio ainda não foi realizado'})
        
        # Formatar dados do sorteio
        return jsonify({
            'success': True,
            'sorteio': {
                'id': sorteio.id,
                'titulo': sorteio.titulo,
                'premio': sorteio.premio,
                'descricao': sorteio.descricao,
                'data_sorteio': sorteio.data_sorteio.isoformat(),
                'status': sorteio.status,
                'ganhador': {
                    'id': sorteio.ganhador.id,
                    'nome': sorteio.ganhador.nome,
                    'email': sorteio.ganhador.email
                }
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@routes.route('/realizar_sorteio/<int:sorteio_id>', methods=['POST'])
@login_required
def realizar_sorteio(sorteio_id):
    """
    Rota para realizar um sorteio, selecionando um ganhador aleatório.
    """
    try:
        sorteio = Sorteio.query.get_or_404(sorteio_id)
        
        # Verificar se o usuário atual é o dono do sorteio
        if sorteio.cliente_id != current_user.id:
            return jsonify({'success': False, 'message': 'Acesso negado'})
        
        # Verificar se o sorteio já foi realizado ou cancelado
        if sorteio.status != 'pendente':
            return jsonify({'success': False, 'message': f'O sorteio não pode ser realizado porque está com status: {sorteio.status}'})
        
        # Buscar participantes elegíveis baseado no tipo de sorteio (evento ou oficina)
        if sorteio.oficina_id:
            # Sorteio para uma oficina específica - buscar usuários através da tabela de inscrições
            participantes_ids = db.session.query(Inscricao.usuario_id).filter_by(oficina_id=sorteio.oficina_id).all()
            if not participantes_ids:
                return jsonify({'success': False, 'message': 'Não há participantes elegíveis para este sorteio'})
            
            # Extrair os IDs dos usuários da lista de tuplas
            usuario_ids = [id[0] for id in participantes_ids]
            
            # Buscar os objetos de usuário
            participantes = Usuario.query.filter(Usuario.id.in_(usuario_ids)).all()
        else:
            # Sorteio para todo o evento - buscar usuários diretamente
            participantes = Usuario.query.filter_by(evento_id=sorteio.evento_id).all()
        
        if not participantes:
            return jsonify({'success': False, 'message': 'Não há participantes elegíveis para este sorteio'})
        
        # Selecionar um participante aleatoriamente
        ganhador = random.choice(participantes)
        
        # Atualizar o sorteio com o ganhador
        sorteio.ganhador_id = ganhador.id
        sorteio.status = 'realizado'
        sorteio.data_sorteio = datetime.utcnow()
        
        db.session.commit()
        
        # Formatar dados do sorteio realizado
        return jsonify({
            'success': True,
            'message': 'Sorteio realizado com sucesso!',
            'sorteio': {
                'id': sorteio.id,
                'titulo': sorteio.titulo,
                'premio': sorteio.premio,
                'descricao': sorteio.descricao,
                'data_sorteio': sorteio.data_sorteio.isoformat(),
                'status': sorteio.status,
                'ganhador': {
                    'id': ganhador.id,
                    'nome': ganhador.nome,
                    'email': ganhador.email
                }
            }
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})
    """
    Rota para realizar um sorteio, selecionando um ganhador aleatório.
    """
    try:
        sorteio = Sorteio.query.get_or_404(sorteio_id)
        
        # Verificar se o usuário atual é o dono do sorteio
        if sorteio.cliente_id != current_user.id:
            return jsonify({'success': False, 'message': 'Acesso negado'})
        
        # Verificar se o sorteio já foi realizado ou cancelado
        if sorteio.status != 'pendente':
            return jsonify({'success': False, 'message': f'O sorteio não pode ser realizado porque está com status: {sorteio.status}'})
        
        # Buscar participantes elegíveis baseado no tipo de sorteio (evento ou oficina)
        participantes = []
        if sorteio.oficina_id:
            # Sorteio para uma oficina específica - buscar usuários diretamente
            participantes = Usuario.query.filter_by(oficina_id=sorteio.oficina_id).all()
        else:
            # Sorteio para todo o evento - buscar usuários diretamente
            participantes = Usuario.query.filter_by(evento_id=sorteio.evento_id).all()
        
        if not participantes:
            return jsonify({'success': False, 'message': 'Não há participantes elegíveis para este sorteio'})
        
        # Selecionar um participante aleatoriamente
        ganhador = random.choice(participantes)
        
        # Atualizar o sorteio com o ganhador
        sorteio.ganhador_id = ganhador.id
        sorteio.status = 'realizado'
        sorteio.data_sorteio = datetime.utcnow()
        
        db.session.commit()
        
        # Formatar dados do sorteio realizado
        return jsonify({
            'success': True,
            'message': 'Sorteio realizado com sucesso!',
            'sorteio': {
                'id': sorteio.id,
                'titulo': sorteio.titulo,
                'premio': sorteio.premio,
                'descricao': sorteio.descricao,
                'data_sorteio': sorteio.data_sorteio.isoformat(),
                'status': sorteio.status,
                'ganhador': {
                    'id': ganhador.id,
                    'nome': ganhador.nome,
                    'email': ganhador.email
                }
            }
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})
    """
    Rota para realizar um sorteio, selecionando um ganhador aleatório.
    """
    try:
        sorteio = Sorteio.query.get_or_404(sorteio_id)
        
        # Verificar se o usuário atual é o dono do sorteio
        if sorteio.cliente_id != current_user.id:
            return jsonify({'success': False, 'message': 'Acesso negado'})
        
        # Verificar se o sorteio já foi realizado ou cancelado
        if sorteio.status != 'pendente':
            return jsonify({'success': False, 'message': f'O sorteio não pode ser realizado porque está com status: {sorteio.status}'})
        
        # Buscar participantes elegíveis baseado no tipo de sorteio (evento ou oficina)
        if sorteio.oficina_id:
            # Sorteio para uma oficina específica
            participantes = Inscricao.query.filter_by(oficina_id=sorteio.oficina_id).all()
        else:
            # Sorteio para todo o evento
            participantes = Inscricao.query.filter_by(evento_id=sorteio.evento_id).all()
        
        if not participantes:
            return jsonify({'success': False, 'message': 'Não há participantes elegíveis para este sorteio'})
        
        # Selecionar um participante aleatoriamente
        inscricao_sorteada = random.choice(participantes)
        
        # Atualizar o sorteio com o ganhador
        sorteio.ganhador_id = inscricao_sorteada.usuario_id
        sorteio.status = 'realizado'
        sorteio.data_sorteio = datetime.utcnow()
        
        db.session.commit()
        
        # Buscar dados completos do ganhador
        ganhador = Usuario.query.get(inscricao_sorteada.usuario_id)
        
        # Formatar dados do sorteio realizado
        return jsonify({
            'success': True,
            'message': 'Sorteio realizado com sucesso!',
            'sorteio': {
                'id': sorteio.id,
                'titulo': sorteio.titulo,
                'premio': sorteio.premio,
                'descricao': sorteio.descricao,
                'data_sorteio': sorteio.data_sorteio.isoformat(),
                'status': sorteio.status,
                'ganhador': {
                    'id': ganhador.id,
                    'nome': ganhador.nome,
                    'email': ganhador.email
                }
            }
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})

@routes.route('/cancelar_sorteio/<int:sorteio_id>', methods=['POST'])
@login_required
def cancelar_sorteio(sorteio_id):
    """
    Rota para cancelar um sorteio.
    """
    try:
        sorteio = Sorteio.query.get_or_404(sorteio_id)
        
        # Verificar se o usuário atual é o dono do sorteio
        if sorteio.cliente_id != current_user.id:
            return jsonify({'success': False, 'message': 'Acesso negado'})
        
        # Verificar se o sorteio já foi realizado ou cancelado
        if sorteio.status != 'pendente':
            return jsonify({'success': False, 'message': f'O sorteio não pode ser cancelado porque está com status: {sorteio.status}'})
        
        # Cancelar o sorteio
        sorteio.status = 'cancelado'
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Sorteio cancelado com sucesso!'
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})

# ===========================
#   RESET DE SENHA VIA CPF
# ===========================
@routes.route('/esqueci_senha_cpf', methods=['GET', 'POST'])
def esqueci_senha_cpf():
    if request.method == 'POST':
        cpf = request.form.get('cpf')
        usuario = Usuario.query.filter_by(cpf=cpf).first()
        if usuario:
            session['reset_user_id'] = usuario.id
            return redirect(url_for('routes.reset_senha_cpf'))
        else:
            flash('CPF não encontrado!', 'danger')
            return redirect(url_for('routes.esqueci_senha_cpf'))
    return render_template('esqueci_senha_cpf.html')

@routes.route('/reset_senha_cpf', methods=['GET', 'POST'])
def reset_senha_cpf():
    user_id = session.get('reset_user_id')
    if not user_id:
        flash('Nenhum usuário selecionado para redefinição!', 'danger')
        return redirect(url_for('routes.esqueci_senha_cpf'))
    usuario = Usuario.query.get(user_id)
    if not usuario:
        flash('Usuário não encontrado no banco de dados!', 'danger')
        return redirect(url_for('routes.esqueci_senha_cpf'))
    if request.method == 'POST':
        nova_senha = request.form.get('nova_senha')
        confirmar_senha = request.form.get('confirmar_senha')
        if not nova_senha or nova_senha != confirmar_senha:
            flash('As senhas não coincidem ou são inválidas.', 'danger')
            return redirect(url_for('routes.reset_senha_cpf'))
        usuario.senha = generate_password_hash(nova_senha)
        db.session.commit()
        session.pop('reset_user_id', None)
        flash('Senha redefinida com sucesso! Faça login novamente.', 'success')
        return redirect(url_for('routes.login'))
    return render_template('reset_senha_cpf.html', usuario=usuario)

@routes.route("/excluir_todas_oficinas", methods=["POST"])
@login_required
def excluir_todas_oficinas():
    if current_user.tipo not in ['admin', 'cliente']:
        flash('Acesso Autorizado!', 'danger')
        

    try:
        if current_user.tipo == 'admin':
            oficinas = Oficina.query.all()
        else:  # Cliente só pode excluir suas próprias oficinas
            oficinas = Oficina.query.filter_by(cliente_id=current_user.id).all()

        if not oficinas:
            flash("Não há oficinas para excluir.", "warning")
            return redirect(url_for("routes.dashboard_cliente" if current_user.tipo == 'cliente' else "routes.dashboard"))

        for oficina in oficinas:
            db.session.query(Checkin).filter_by(oficina_id=oficina.id).delete()
            db.session.query(Inscricao).filter_by(oficina_id=oficina.id).delete()
            db.session.query(OficinaDia).filter_by(oficina_id=oficina.id).delete()
            db.session.query(MaterialOficina).filter_by(oficina_id=oficina.id).delete()
            db.session.query(RelatorioOficina).filter_by(oficina_id=oficina.id).delete()
            db.session.delete(oficina)

        db.session.commit()
        flash("Oficinas excluídas com sucesso!", "success")

    except Exception as e:
        db.session.rollback()
        flash(f"Erro ao excluir oficinas: {str(e)}", "danger")

    return redirect(url_for("routes.dashboard_cliente" if current_user.tipo == 'cliente' else "routes.dashboard"))


@routes.route("/importar_usuarios", methods=["POST"])
def importar_usuarios():
    if "arquivo" not in request.files:
        flash("Nenhum arquivo enviado!", "danger")
        return redirect(url_for("routes.dashboard"))
    arquivo = request.files["arquivo"]
    if arquivo.filename == "":
        flash("Nenhum arquivo selecionado.", "danger")
        return redirect(url_for("routes.dashboard"))
    if arquivo and arquivo_permitido(arquivo.filename):
        filename = secure_filename(arquivo.filename)
        filepath = os.path.join(app.config["UPLOAD_FOLDER"], filename)
        arquivo.save(filepath)
        try:
            print("📌 [DEBUG] Lendo o arquivo Excel...")
            df = pd.read_excel(filepath, dtype={"cpf": str})
            print(f"📌 [DEBUG] Colunas encontradas: {df.columns.tolist()}")
            colunas_obrigatorias = ["nome", "cpf", "email", "senha", "formacao", "tipo"]
            if not all(col in df.columns for col in colunas_obrigatorias):
                flash("Erro: O arquivo deve conter as colunas: " + ", ".join(colunas_obrigatorias), "danger")
                return redirect(url_for("routes.dashboard"))
            total_importados = 0
            for _, row in df.iterrows():
                cpf_str = str(row["cpf"]).strip()
                usuario_existente = Usuario.query.filter_by(email=row["email"]).first()
                if usuario_existente:
                    print(f"⚠️ [DEBUG] Usuário com e-mail {row['email']} já existe. Pulando...")
                    continue
                usuario_existente = Usuario.query.filter_by(cpf=cpf_str).first()
                if usuario_existente:
                    print(f"⚠️ [DEBUG] Usuário com CPF {cpf_str} já existe. Pulando...")
                    continue
                senha_hash = generate_password_hash(str(row["senha"]))
                novo_usuario = Usuario(
                    nome=row["nome"],
                    cpf=cpf_str,
                    email=row["email"],
                    senha=senha_hash,
                    formacao=row["formacao"],
                    tipo=row["tipo"]
                )
                db.session.add(novo_usuario)
                total_importados += 1
                print(f"✅ [DEBUG] Usuário '{row['nome']}' cadastrado com sucesso!")
            db.session.commit()
            flash(f"{total_importados} usuários importados com sucesso!", "success")
        except Exception as e:
            db.session.rollback()
            print(f"❌ [ERRO] Erro ao importar usuários: {str(e)}")
            flash(f"Erro ao processar o arquivo: {str(e)}", "danger")
        os.remove(filepath)
    else:
        flash("Formato de arquivo inválido. Envie um arquivo Excel (.xlsx)", "danger")
    return redirect(url_for("routes.dashboard"))

@routes.route("/toggle_checkin_global_cliente", methods=["POST"])
@login_required
def toggle_checkin_global_cliente():
    # Permite apenas clientes acessarem esta rota
    #if current_user.tipo != "cliente":
        #flash("Acesso Autorizado!", "danger")
        
        
    
    # Para clientes, já utiliza o próprio ID
    cliente_id = current_user.id

    from models import ConfiguracaoCliente
    config_cliente = ConfiguracaoCliente.query.filter_by(cliente_id=cliente_id).first()
    if not config_cliente:
        # Cria uma nova configuração para esse cliente, se não existir
        config_cliente = ConfiguracaoCliente(
            cliente_id=cliente_id,
            permitir_checkin_global=False,
            habilitar_feedback=False,
            habilitar_certificado_individual=False
        )
        db.session.add(config_cliente)
        db.session.commit()

    # Inverte o valor de permitir_checkin_global e persiste
    config_cliente.permitir_checkin_global = not config_cliente.permitir_checkin_global
    db.session.commit()

    return jsonify({
        "success": True,
        "value": config_cliente.permitir_checkin_global,  # True ou False
        "message": "Check-in Global atualizado com sucesso!"
    })


@routes.route("/toggle_feedback_cliente", methods=["POST"])
@login_required
def toggle_feedback_cliente():
    # Permite apenas clientes
    #if current_user.tipo != "cliente":
        #flash("Acesso Autorizado!", "danger")
        
    
    cliente_id = current_user.id
    config_cliente = ConfiguracaoCliente.query.filter_by(cliente_id=cliente_id).first()
    if not config_cliente:
        config_cliente = ConfiguracaoCliente(
            cliente_id=cliente_id,
            permitir_checkin_global=False,
            habilitar_feedback=False,
            habilitar_certificado_individual=False
        )
        db.session.add(config_cliente)
        db.session.commit()

    config_cliente.habilitar_feedback = not config_cliente.habilitar_feedback
    db.session.commit()

    return jsonify({
        "success": True,
        "value": config_cliente.habilitar_feedback,
        "message": "Feedback atualizado com sucesso!"
    })


@routes.route("/toggle_certificado_cliente", methods=["POST"])
@login_required
def toggle_certificado_cliente():
    # Permite apenas clientes
    #if current_user.tipo != "cliente":
        #flash("Acesso Autorizado!", "danger")
        
    
    cliente_id = current_user.id
    config_cliente = ConfiguracaoCliente.query.filter_by(cliente_id=cliente_id).first()
    if not config_cliente:
        config_cliente = ConfiguracaoCliente(
            cliente_id=cliente_id,
            permitir_checkin_global=False,
            habilitar_feedback=False,
            habilitar_certificado_individual=False
        )
        db.session.add(config_cliente)
        db.session.commit()

    config_cliente.habilitar_certificado_individual = not config_cliente.habilitar_certificado_individual
    db.session.commit()

    return jsonify({
        "success": True,
        "value": config_cliente.habilitar_certificado_individual,
        "message": "Certificado Individual atualizado com sucesso!"
    })


@routes.route("/toggle_certificado_individual", methods=["POST"])
@login_required
def toggle_certificado_individual():
    # Permite apenas clientes (já que esta rota altera uma configuração global de certificado)
    #if current_user.tipo != "cliente":
        #flash("Acesso Autorizado!", "danger")
        
    
    config = Configuracao.query.first()
    if not config:
        config = Configuracao(
            permitir_checkin_global=False,
            habilitar_feedback=False,
            habilitar_certificado_individual=False
        )
        db.session.add(config)
    config.habilitar_certificado_individual = not config.habilitar_certificado_individual
    db.session.commit()

    status = "ativado" if config.habilitar_certificado_individual else "desativado"
    flash(f"Certificado individual {status} com sucesso!", "success")
    return redirect(url_for("routes.dashboard_cliente"))



# ===========================
#         FEEDBACK
# ===========================
@routes.route('/feedback/<int:oficina_id>', methods=['GET', 'POST'])
@login_required
def feedback(oficina_id):
    oficina = Oficina.query.get_or_404(oficina_id)
    if current_user.tipo != 'participante':
        flash('Apenas participantes podem enviar feedback.', 'danger')
        return redirect(url_for('routes.dashboard'))
    if request.method == 'POST':
        try:
            rating = int(request.form.get('rating', 0))
        except ValueError:
            rating = 0
        comentario = request.form.get('comentario', '').strip()
        if rating < 1 or rating > 5:
            flash('A avaliação deve ser entre 1 e 5 estrelas.', 'danger')
            return redirect(url_for('routes.feedback', oficina_id=oficina_id))
        novo_feedback = Feedback(
            usuario_id=current_user.id,
            oficina_id=oficina.id,
            rating=rating,
            comentario=comentario
        )
        db.session.add(novo_feedback)
        db.session.commit()
        flash('Feedback enviado com sucesso!', 'success')
        return redirect(url_for('routes.dashboard_participante'))
    return render_template('feedback.html', oficina=oficina)

@routes.route('/feedback_oficina/<int:oficina_id>')
@login_required
def feedback_oficina(oficina_id):
    oficina = Oficina.query.get_or_404(oficina_id)  # Primeiro
    if current_user.tipo not in ['admin', 'cliente'] or (current_user.tipo == 'cliente' and oficina.cliente_id != current_user.id and oficina.cliente_id is not None):
        flash('Você não tem permissão para visualizar o feedback desta oficina.', 'danger')
        

    # Obtendo clientes para filtro (somente admin pode visualizar)
    clientes = Cliente.query.all() if current_user.tipo == 'admin' else []

    # Cálculo das estatísticas gerais (sem os filtros da query abaixo)
    total_feedbacks_all = Feedback.query.filter_by(oficina_id=oficina_id).all()
    total_count = len(total_feedbacks_all)
    total_avg = (sum(fb.rating for fb in total_feedbacks_all) / total_count) if total_count > 0 else 0

    feedbacks_usuarios = Feedback.query.filter(
        Feedback.oficina_id == oficina_id,
        Feedback.usuario_id.isnot(None)
    ).all()
    count_usuarios = len(feedbacks_usuarios)
    avg_usuarios = (sum(fb.rating for fb in feedbacks_usuarios) / count_usuarios) if count_usuarios > 0 else 0

    feedbacks_ministrantes = Feedback.query.filter(
        Feedback.oficina_id == oficina_id,
        Feedback.ministrante_id.isnot(None)
    ).all()
    count_ministrantes = len(feedbacks_ministrantes)
    avg_ministrantes = (sum(fb.rating for fb in feedbacks_ministrantes) / count_ministrantes) if count_ministrantes > 0 else 0

    # Filtros
    tipo = request.args.get('tipo')
    estrelas = request.args.get('estrelas')
    cliente_filter = request.args.get('cliente_id')

    query = Feedback.query.join(Oficina).filter(Feedback.oficina_id == oficina_id)

    # Filtra pelo tipo de feedback (usuário ou ministrante)
    if tipo == 'usuario':
        query = query.filter(Feedback.usuario_id.isnot(None))
    elif tipo == 'ministrante':
        query = query.filter(Feedback.ministrante_id.isnot(None))

    # Filtra pelo número de estrelas
    if estrelas and estrelas.isdigit():
        query = query.filter(Feedback.rating == int(estrelas))

    # Filtra pelo cliente selecionado (somente admins)
    if current_user.tipo == 'admin' and cliente_filter and cliente_filter.isdigit():
        query = query.filter(Oficina.cliente_id == int(cliente_filter))

    feedbacks = query.order_by(Feedback.created_at.desc()).all()

    
    return render_template('feedback_oficina.html', oficina=oficina, feedbacks=feedbacks,
                           total_count=total_count, total_avg=total_avg,
                           count_ministrantes=count_ministrantes, avg_ministrantes=avg_ministrantes,
                           count_usuarios=count_usuarios, avg_usuarios=avg_usuarios,  is_admin=current_user.tipo == 'admin', clientes=clientes, cliente_filter=cliente_filter)



def gerar_pdf_feedback(oficina, feedbacks, pdf_path):
    """
    Gera um PDF elegante com os feedbacks de uma oficina.
    
    Args:
        oficina: Objeto da oficina com informações como título
        feedbacks: Lista de objetos de feedback contendo avaliações e comentários
        pdf_path: Caminho onde o PDF será salvo
    """
    from reportlab.platypus import Table, TableStyle, Paragraph, Spacer, SimpleDocTemplate, PageBreak, Image
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch, mm
    from datetime import datetime
    import pytz
    import os
    
    # Função para converter um datetime para o fuso de Brasília
    def convert_to_brasilia(dt):
        brasilia_tz = pytz.timezone("America/Sao_Paulo")
        # Se o datetime não for "aware", assume-se que está em UTC
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=pytz.utc)
        return dt.astimezone(brasilia_tz)
    
    # Criar estilos personalizados
    styles = getSampleStyleSheet()
    
    # Título com estilo moderno
    title_style = ParagraphStyle(
        'TitleStyle',
        parent=styles['Title'],
        fontSize=24,
        fontName='Helvetica-Bold',
        alignment=1,  # Centralizado
        spaceAfter=20,
        textColor=colors.HexColor('#1A365D')  # Azul escuro elegante
    )
    
    # Estilo para o subtítulo
    subtitle_style = ParagraphStyle(
        'SubtitleStyle',
        parent=styles['Heading2'],
        fontSize=16,
        fontName='Helvetica-Bold',
        alignment=1,
        spaceAfter=15,
        textColor=colors.HexColor('#2A4365')  # Azul médio
    )
    
    # Estilo para o texto normal
    normal_style = ParagraphStyle(
        'NormalStyle',
        parent=styles['Normal'],
        fontSize=11,
        leading=14,
        fontName='Helvetica'
    )
    
    # Estilo para o cabeçalho da tabela
    header_style = ParagraphStyle(
        'HeaderStyle',
        parent=styles['Heading4'],
        fontSize=12,
        fontName='Helvetica-Bold',
        alignment=1,
        textColor=colors.white,
        leading=14
    )
    
    # Estilo para o rodapé
    footer_style = ParagraphStyle(
        'FooterStyle',
        parent=styles['Normal'],
        fontSize=9,
        fontName='Helvetica-Oblique',
        textColor=colors.HexColor('#4A5568'),  # Cinza escuro
        alignment=1
    )
    
    # Estilo para comentários
    comment_style = ParagraphStyle(
        'CommentStyle',
        parent=styles['Normal'],
        fontSize=10,
        leading=14,
        fontName='Helvetica',
        firstLineIndent=0,
        spaceBefore=3,
        spaceAfter=3
    )
    
    # Cria o documento em modo paisagem com margens aprimoradas
    doc = SimpleDocTemplate(
        pdf_path, 
        pagesize=landscape(letter), 
        leftMargin=0.75*inch, 
        rightMargin=0.75*inch,
        topMargin=0.5*inch,
        bottomMargin=0.5*inch
    )
    
    available_width = doc.width  # largura disponível após as margens
    
    elements = []
    
    # Adicionar logotipo ou imagem header (opcional)
    logo_path = os.path.join("static", "logo.png")
    if os.path.exists(logo_path):
        # Adiciona um espaço antes do logo
        elements.append(Spacer(1, 0.2 * inch))
        
        # Centraliza o logo
        logo = Image(logo_path, width=1.5*inch, height=0.75*inch)
        elements.append(logo)
        
        # Adiciona um espaço após o logo
        elements.append(Spacer(1, 0.3 * inch))
    
    # Título principal
    elements.append(Paragraph(f"Relatório de Feedback", title_style))
    
    # Subtítulo com informações da oficina
    elements.append(Paragraph(f"Oficina: {oficina.titulo}", subtitle_style))
    
    # Adicionar informações da data de geração
    now = convert_to_brasilia(datetime.utcnow())
    elements.append(Paragraph(f"Gerado em: {now.strftime('%d/%m/%Y às %H:%M')}", normal_style))
    
    # Informações gerais (pode-se adicionar ministrate, datas, etc.)
    ministrante_nome = oficina.ministrante_obj.nome if hasattr(oficina, 'ministrante_obj') and oficina.ministrante_obj else 'N/A'
    elements.append(Paragraph(f"Ministrante: {ministrante_nome}", normal_style))
    
    # Verificar se oficina tem atributo 'cidade' e 'estado'
    if hasattr(oficina, 'cidade') and hasattr(oficina, 'estado'):
        elements.append(Paragraph(f"Local: {oficina.cidade}, {oficina.estado}", normal_style))
    
    # Calcular estatísticas de avaliação
    if feedbacks:
        total_ratings = len(feedbacks)
        avg_rating = sum(fb.rating for fb in feedbacks) / total_ratings if total_ratings > 0 else 0
        elements.append(Paragraph(f"Avaliação média: {avg_rating:.1f}/5.0 ({total_ratings} avaliações)", normal_style))
    
    # Adicionar espaço antes da tabela
    elements.append(Spacer(1, 0.4 * inch))
    
    # Linha decorativa antes da tabela
    elements.append(Table([['']], colWidths=[doc.width], 
                          style=TableStyle([('LINEABOVE', (0, 0), (-1, 0), 1, colors.HexColor('#3182CE'))])))
    elements.append(Spacer(1, 0.3 * inch))
    
    # Título da seção de feedbacks
    elements.append(Paragraph("Detalhes dos Feedbacks", subtitle_style))
    elements.append(Spacer(1, 0.2 * inch))
    
    # Cabeçalho da tabela com Paragraph para melhor formatação
    header = [
        Paragraph("Usuário", header_style),
        Paragraph("Avaliação", header_style),
        Paragraph("Comentário", header_style),
        Paragraph("Data", header_style)
    ]
    table_data = [header]
    
    # Prepara os dados da tabela convertendo os horários para o fuso local
    for fb in feedbacks:
        # Criar string de estrelas
        filled_star = "★"  # Estrela preenchida
        empty_star = "☆"   # Estrela vazia
        rating_str = filled_star * fb.rating + empty_star * (5 - fb.rating)
        
        # Formatar data local
        dt_local = convert_to_brasilia(fb.created_at)
        data_str = dt_local.strftime('%d/%m/%Y %H:%M')
        
        # Determinar o nome do autor
        nome_autor = fb.usuario.nome if hasattr(fb, 'usuario') and fb.usuario is not None else (
                     fb.ministrante.nome if hasattr(fb, 'ministrante') and fb.ministrante is not None else "Desconhecido")
        
        # Garante que o comentário não seja None
        comentario_text = fb.comentario or "Sem comentários adicionais."
        
        # Utiliza Paragraph para permitir quebra de linha em comentários longos
        comentario_paragraph = Paragraph(comentario_text, comment_style)
        
        row = [
            Paragraph(nome_autor, normal_style),
            Paragraph(rating_str, normal_style),
            comentario_paragraph,
            Paragraph(data_str, normal_style)
        ]
        table_data.append(row)
    
    # Cria o documento em modo paisagem com margens aprimoradas
    doc = SimpleDocTemplate(
        pdf_path, 
        pagesize=landscape(letter), 
        leftMargin=0.75*inch, 
        rightMargin=0.75*inch,
        topMargin=0.5*inch,
        bottomMargin=0.5*inch
    )
    
    available_width = doc.width  # largura disponível após as margens
    
    # Define as larguras das colunas em porcentagem da largura disponível
    col_widths = [
        available_width * 0.20,  # Usuário
        available_width * 0.15,  # Avaliação
        available_width * 0.45,  # Comentário
        available_width * 0.20   # Data
    ]
    
    # Cria e estiliza a tabela
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    
    # Cores suaves e modernas
    header_bg_color = colors.HexColor('#2C5282')  # Azul escuro
    alt_row_color = colors.HexColor('#EBF8FF')    # Azul bem claro
    grid_color = colors.HexColor('#CBD5E0')       # Cinza claro
    
    table.setStyle(TableStyle([
        # Cabeçalho
        ('BACKGROUND', (0, 0), (-1, 0), header_bg_color),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        
        # Linhas alternadas para facilitar a leitura
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, alt_row_color]),
        
        # Grade fina e elegante
        ('GRID', (0, 0), (-1, -1), 0.5, grid_color),
        
        # Alinhamento
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),       # Cabeçalho centralizado
        ('ALIGN', (1, 1), (1, -1), 'CENTER'),       # Coluna de avaliação centralizada
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),         # Coluna de usuários à esquerda
        ('ALIGN', (3, 1), (3, -1), 'CENTER'),       # Coluna de datas centralizada
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),     # Alinhamento vertical no meio
        
        # Espaçamento interno
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 1), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 10),
    ]))
    
    elements.append(table)
    
    # Adiciona espaço antes do rodapé
    elements.append(Spacer(1, 0.4 * inch))
    
    # Linha decorativa antes do rodapé
    elements.append(Table([['']], colWidths=[doc.width], 
                          style=TableStyle([('LINEABOVE', (0, 0), (-1, 0), 0.5, colors.HexColor('#CBD5E0'))])))
    
    # Adiciona espaço após a linha
    elements.append(Spacer(1, 0.2 * inch))
    
    # Rodapé com horário local e informações adicionais
    footer_text = "Este relatório é um documento confidencial e de uso interno. "
    footer_text += f"Gerado via AppFiber em {now.strftime('%d/%m/%Y às %H:%M')}."
    elements.append(Paragraph(footer_text, footer_style))
    
    # Construir o PDF
    doc.build(elements)


@routes.route('/gerar_pdf_feedback/<int:oficina_id>')
@login_required
def gerar_pdf_feedback_route(oficina_id):
    if current_user.tipo != 'admin' and current_user.tipo != 'cliente':
        flash('Acesso Autorizado!', 'danger')
        
    oficina = Oficina.query.get_or_404(oficina_id)
    
    # Replicar a lógica de filtragem usada na rota feedback_oficina
    query = Feedback.query.filter(Feedback.oficina_id == oficina_id)
    tipo = request.args.get('tipo')
    if tipo == 'usuario':
        query = query.filter(Feedback.usuario_id.isnot(None))
    elif tipo == 'ministrante':
        query = query.filter(Feedback.ministrante_id.isnot(None))
    estrelas = request.args.get('estrelas')
    if estrelas and estrelas.isdigit():
        query = query.filter(Feedback.rating == int(estrelas))
    
    feedbacks = query.order_by(Feedback.created_at.desc()).all()
    
    pdf_folder = os.path.join("static", "feedback_pdfs")
    os.makedirs(pdf_folder, exist_ok=True)
    pdf_filename = f"feedback_{oficina.id}.pdf"
    pdf_path = os.path.join(pdf_folder, pdf_filename)
    gerar_pdf_feedback(oficina, feedbacks, pdf_path)
    return send_file(pdf_path, as_attachment=True)


@routes.route("/toggle_feedback", methods=["POST"])
@login_required
def toggle_feedback():
    if current_user.tipo != "admin":
        flash("Acesso negado!", "danger")
        return redirect(url_for("routes.dashboard_participante"))
    config = Configuracao.query.first()
    if not config:
        config = Configuracao(permitir_checkin_global=False, habilitar_feedback=False)
        db.session.add(config)
    config.habilitar_feedback = not config.habilitar_feedback
    db.session.commit()
    flash(f"Feedback global {'ativado' if config.habilitar_feedback else 'desativado'} com sucesso!", "success")
    return redirect(url_for("routes.dashboard"))

@routes.route('/feedback_ministrante/<int:oficina_id>', methods=['GET', 'POST'])
@login_required
def feedback_ministrante(oficina_id):
    # Verifica se o usuário é um ministrante
    if current_user.tipo != 'ministrante':
        flash('Apenas ministrantes podem enviar feedback por aqui.', 'danger')
        return redirect(url_for('routes.dashboard_ministrante'))
    
    oficina = Oficina.query.get_or_404(oficina_id)
    
    if request.method == 'POST':
        try:
            rating = int(request.form.get('rating', 0))
        except ValueError:
            rating = 0
        comentario = request.form.get('comentario', '').strip()
        if rating < 1 or rating > 5:
            flash('A avaliação deve ser entre 1 e 5 estrelas.', 'danger')
            return redirect(url_for('routes.feedback_ministrante', oficina_id=oficina_id))
        
        novo_feedback = Feedback(
            ministrante_id=current_user.id,  # Salva o id do ministrante
            oficina_id=oficina.id,
            rating=rating,
            comentario=comentario
        )
        db.session.add(novo_feedback)
        db.session.commit()
        flash('Feedback enviado com sucesso!', 'success')
        return redirect(url_for('routes.dashboard_ministrante'))
    
    # Reaproveita o template existente (feedback.html) ou crie um específico se desejar
    return render_template('feedback.html', oficina=oficina)


@routes.route('/gerar_certificado/<int:oficina_id>', methods=['GET'])
@login_required
def gerar_certificado_individual(oficina_id):
    """
    Gera um certificado individual para o usuário logado em uma oficina específica.
    """
    oficina = Oficina.query.get(oficina_id)
    if not oficina:
        flash("Oficina não encontrada!", "danger")
        return redirect(url_for('routes.dashboard_participante'))

    # Verifica se o usuário está inscrito na oficina
    inscricao = Inscricao.query.filter_by(usuario_id=current_user.id, oficina_id=oficina.id).first()
    if not inscricao:
        flash("Você não está inscrito nesta oficina!", "danger")
        return redirect(url_for('routes.dashboard_participante'))

    # Define o caminho do certificado
    pdf_path = f"static/certificados/certificado_{current_user.id}_{oficina.id}.pdf"
    os.makedirs(os.path.dirname(pdf_path), exist_ok=True)

    # Gera o certificado (mesmo layout do admin, mas apenas para o usuário logado)
    gerar_certificados_pdf(oficina, [inscricao], pdf_path)

    # Retorna o arquivo PDF gerado
    return send_file(pdf_path, as_attachment=True)


# ===========================
#   CADASTRO DE MINISTRANTE
# ===========================
import logging
from flask import Flask, render_template, redirect, url_for, flash, request, jsonify
from werkzeug.security import generate_password_hash
from extensions import db
from models import Ministrante, EventoInscricaoTipo, Sorteio, Inscricao, Usuario, Oficina
from flask_login import login_required
import random

# Configure o logger (isso pode ser configurado globalmente no seu app)
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

from flask import current_app, flash, redirect, request, url_for, render_template
from flask_login import login_required, current_user
import os, uuid

@routes.route('/cadastro_ministrante', methods=['GET', 'POST'])
@login_required
def cadastro_ministrante():
    if current_user.tipo not in ['admin', 'cliente']:
        flash('Apenas administradores e clientes podem cadastrar ministrantes!', 'danger')
        return redirect(url_for('routes.dashboard'))

    clientes = Cliente.query.all() if current_user.tipo == 'admin' else []

    if request.method == 'POST':
        nome = request.form.get('nome')
        formacao = request.form.get('formacao')
        categorias_formacao = request.form.getlist('categorias_formacao')
        categorias_str = ','.join(categorias_formacao)  # Transforma lista em string
        foto = request.files.get('foto')
        areas = request.form.get('areas')
        cpf = request.form.get('cpf')
        pix = request.form.get('pix')
        cidade = request.form.get('cidade')
        estado = request.form.get('estado')
        email = request.form.get('email')
        senha = generate_password_hash(request.form.get('senha'))

        # Se for admin, pega o cliente_id do form
        # Se for cliente, assume o id do current_user
        cliente_id = request.form.get('cliente_id') if current_user.tipo == 'admin' else current_user.id

        # Gera caminho único para foto
        foto_path = None
        if foto and foto.filename:
            original_filename = secure_filename(foto.filename)   # ex.: foto.jpg
            ext = original_filename.rsplit('.', 1)[1].lower()    # pega a extensão ex.: jpg
            unique_name = f"{uuid.uuid4()}.{ext}"                # ex.: 123e4567-e89b-12d3-a456-426614174000.jpg
            caminho_foto = os.path.join(
                current_app.root_path, 
                'static', 
                'uploads', 
                'ministrantes', 
                unique_name
            )
            os.makedirs(os.path.dirname(caminho_foto), exist_ok=True)
            foto.save(caminho_foto) 
            foto_path = f'uploads/ministrantes/{unique_name}'    # caminho relativo à pasta static

        # Agora criamos o objeto Ministrante, passando foto_path
        novo_ministrante = Ministrante(
            nome=nome,
            formacao=formacao,
            categorias_formacao=categorias_str,
            foto=foto_path,  # Passamos o caminho aqui (ou None se não houve upload)
            areas_atuacao=areas,
            cpf=cpf,
            pix=pix,
            cidade=cidade,
            estado=estado,
            email=email,
            senha=senha,
            cliente_id=cliente_id
        )

        try:
            db.session.add(novo_ministrante)
            db.session.commit()
            flash('Cadastro realizado com sucesso!', 'success')
            # Redireciona para o dashboard adequado (admin / cliente)
            return redirect(url_for('routes.dashboard_cliente' if current_user.tipo == 'cliente' else 'routes.dashboard'))
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao cadastrar ministrante: {str(e)}', 'danger')

    return render_template('cadastro_ministrante.html', clientes=clientes)



@routes.route('/dashboard_ministrante')
@login_required
def dashboard_ministrante():
    # Log para depuração: exibir o tipo do current_user e seus atributos
    import logging
    logger = logging.getLogger(__name__)
    print(f"current_user: {current_user}, type: {type(current_user)}")
    # Se estiver usando UserMixin, current_user pode não ter o atributo 'tipo'
    # Então, usamos isinstance para verificar se é Ministrante.
    if not isinstance(current_user, Ministrante):
        print("current_user não é uma instância de Ministrante")
        flash('Acesso negado!', 'danger')
        return redirect(url_for('routes.home'))

    # Busca o ministrante logado com base no email (ou use current_user diretamente)
    ministrante_logado = Ministrante.query.filter_by(email=current_user.email).first()
    if not ministrante_logado:
        print("Ministrante não encontrado no banco de dados")
        flash('Ministrante não encontrado!', 'danger')
        return redirect(url_for('routes.home'))

    # Buscar as oficinas deste ministrante
    oficinas_do_ministrante = Oficina.query.filter_by(ministrante_id=ministrante_logado.id).all()
    # Carrega a configuração e define habilitar_feedback
    config = Configuracao.query.first()
    habilitar_feedback = config.habilitar_feedback if config else False
    print(f"Foram encontradas {len(oficinas_do_ministrante)} oficinas para o ministrante {ministrante_logado.email}")

    return render_template(
        'dashboard_ministrante.html',
        ministrante=ministrante_logado,
        oficinas=oficinas_do_ministrante,
        habilitar_feedback=habilitar_feedback
    )

@routes.route('/enviar_relatorio/<int:oficina_id>', methods=['GET', 'POST'])
@login_required
def enviar_relatorio(oficina_id):
    if current_user.tipo != 'ministrante':
        flash('Acesso negado!', 'danger')
        return redirect(url_for('routes.home'))

    oficina = Oficina.query.get_or_404(oficina_id)
    ministrante_logado = Ministrante.query.filter_by(email=current_user.email).first()

    if oficina.ministrante_id != ministrante_logado.id:
        flash('Você não é responsável por esta oficina!', 'danger')
        return redirect(url_for('routes.dashboard_ministrante'))

    if request.method == 'POST':
        metodologia = request.form.get('metodologia')
        resultados = request.form.get('resultados')

        # Upload de fotos/vídeos se desejado
        arquivo_midia = request.files.get('arquivo_midia')
        midia_path = None
        if arquivo_midia:
            filename = secure_filename(arquivo_midia.filename)
            pasta_uploads = os.path.join('uploads', 'relatorios')
            os.makedirs(pasta_uploads, exist_ok=True)
            caminho_arquivo = os.path.join(pasta_uploads, filename)
            arquivo_midia.save(caminho_arquivo)
            midia_path = caminho_arquivo

        novo_relatorio = RelatorioOficina(
            oficina_id=oficina.id,
            ministrante_id=ministrante_logado.id,
            metodologia=metodologia,
            resultados=resultados,
            fotos_videos_path=midia_path
        )
        db.session.add(novo_relatorio)
        db.session.commit()

        flash("Relatório enviado com sucesso!", "success")
        return redirect(url_for('routes.dashboard_ministrante'))

    return render_template('enviar_relatorio.html', oficina=oficina)

@routes.route('/upload_material/<int:oficina_id>', methods=['GET', 'POST'])
@login_required
def upload_material(oficina_id):
    # Verifica se o usuário é um ministrante
    from models import Ministrante  # Certifique-se de importar se necessário
    if not hasattr(current_user, 'tipo') or current_user.tipo != 'ministrante':
        flash('Acesso negado!', 'danger')
        return redirect(url_for('routes.home'))
    
    # Buscar a oficina e verificar se o ministrante logado é responsável por ela
    oficina = Oficina.query.get_or_404(oficina_id)
    ministrante_logado = Ministrante.query.filter_by(email=current_user.email).first()
    if not ministrante_logado or oficina.ministrante_id != ministrante_logado.id:
        flash('Você não é responsável por esta oficina!', 'danger')
        return redirect(url_for('routes.dashboard_ministrante'))
    
    if request.method == 'POST':
        arquivo = request.files.get('arquivo')
        if arquivo:
            filename = secure_filename(arquivo.filename)
            pasta_uploads = os.path.join('uploads', 'materiais')
            os.makedirs(pasta_uploads, exist_ok=True)
            caminho_arquivo = os.path.join(pasta_uploads, filename)
            arquivo.save(caminho_arquivo)
            
            novo_material = MaterialOficina(
                oficina_id=oficina.id,
                nome_arquivo=filename,
                caminho_arquivo=caminho_arquivo
            )
            db.session.add(novo_material)
            db.session.commit()
            
            flash('Material anexado com sucesso!', 'success')
            return redirect(url_for('routes.dashboard_ministrante'))
        else:
            flash('Nenhum arquivo foi enviado.', 'danger')
    
    return render_template('upload_material.html', oficina=oficina)

@routes.route('/editar_ministrante/<int:ministrante_id>', methods=['GET', 'POST'])
@login_required
def editar_ministrante(ministrante_id):
    ministrante = Ministrante.query.get_or_404(ministrante_id)

    if current_user.tipo == 'cliente' and ministrante.cliente_id != current_user.id:
        flash('Acesso negado!', 'danger')
        return redirect(url_for('routes.dashboard_cliente'))

    clientes = Cliente.query.all() if current_user.tipo == 'admin' else None
    ids_extras = request.form.getlist('ministrantes_ids[]')
    
    for mid in ids_extras:
        m = Ministrante.query.get(int(mid))
        if m:
            oficina.ministrantes_associados.append(m)

    if request.method == 'POST':
        ministrante.nome = request.form.get('nome')
        ministrante.formacao = request.form.get('formacao')
        categorias_formacao = request.form.getlist('categorias_formacao')
        ministrante.categorias_formacao = ','.join(categorias_formacao)

        ministrante.areas_atuacao = request.form.get('areas')
        ministrante.cpf = request.form.get('cpf')
        ministrante.pix = request.form.get('pix')
        ministrante.cidade = request.form.get('cidade')
        ministrante.estado = request.form.get('estado')
        ministrante.email = request.form.get('email')

        if current_user.tipo == 'admin':
            novo_cliente_id = request.form.get('cliente_id')
            ministrante.cliente_id = novo_cliente_id if novo_cliente_id else None

        nova_senha = request.form.get('senha')
        if nova_senha:
            ministrante.senha = generate_password_hash(nova_senha)

        foto = request.files.get('foto')
        if foto and foto.filename:
            filename = secure_filename(foto.filename)
            caminho_foto = os.path.join(current_app.root_path, 'static/uploads/ministrantes', filename)
            os.makedirs(os.path.dirname(caminho_foto), exist_ok=True)
            foto.save(caminho_foto)
            ministrante.foto = f'uploads/ministrantes/{filename}'

        db.session.commit()
        flash('Ministrante atualizado com sucesso!', 'success')
        return redirect(url_for('routes.gerenciar_ministrantes'))

    return render_template('editar_ministrante.html', ministrante=ministrante, clientes=clientes)

@routes.route('/excluir_ministrante/<int:ministrante_id>', methods=['POST'])
@login_required
def excluir_ministrante(ministrante_id):
    ministrante = Ministrante.query.get_or_404(ministrante_id)

    if current_user.tipo == 'cliente' and ministrante.cliente_id != current_user.id:
        flash('Acesso negado!', 'danger')
        return redirect(url_for('routes.dashboard_cliente'))

    db.session.delete(ministrante)
    db.session.commit()
    flash('Ministrante excluído com sucesso!', 'success')
    return redirect(url_for('routes.gerenciar_ministrantes'))


@routes.route('/gerenciar_ministrantes', methods=['GET'])
@login_required
def gerenciar_ministrantes():
    if current_user.tipo not in ['admin', 'cliente']:
        flash('Acesso negado!', 'danger')
        return redirect(url_for('routes.dashboard'))

    ministrantes = Ministrante.query.filter_by(cliente_id=current_user.id).all() if current_user.tipo == 'cliente' else Ministrante.query.all()

    return render_template('gerenciar_ministrantes.html', ministrantes=ministrantes)



@routes.route('/gerenciar_inscricoes', methods=['GET', 'POST'])
@login_required
def gerenciar_inscricoes():
    if current_user.tipo not in ['admin', 'cliente']:
        flash('Acesso Autorizado!', 'danger')
        
    # Se o usuário for cliente, filtra apenas as oficinas e inscrições associadas a ele
    if current_user.tipo == 'cliente':
        oficinas = Oficina.query.filter_by(cliente_id=current_user.id).all()
        inscritos = Inscricao.query.join(Oficina).filter(Oficina.cliente_id == current_user.id).all()
    else:
        # Se for admin, mostra todos os registros
        oficinas = Oficina.query.all()
        inscritos = Inscricao.query.all()
    return render_template('gerenciar_inscricoes.html', oficinas=oficinas, inscritos=inscritos)



@routes.route('/admin_scan')
@login_required
def admin_scan():
    if current_user.tipo not in ('admin', 'cliente'):
        flash("Acesso Autorizado!", "danger")
        
    return render_template("scan_qr.html")

@routes.route('/relatorios/<path:filename>')
@login_required
def get_relatorio_file(filename):
    # Ajuste o caminho para a pasta de relatórios
    pasta_uploads = os.path.join('uploads', 'relatorios')
    return send_from_directory(pasta_uploads, filename)

@routes.route('/gerar_pdf_checkins_qr', methods=['GET'])
@login_required
def gerar_pdf_checkins_qr():
    import os
    import pytz
    from datetime import datetime
    from collections import defaultdict
    from flask import flash, redirect, url_for, send_file
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, LongTable, TableStyle
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    from models import Checkin, Usuario, Oficina
    from extensions import db
    from flask_login import current_user
    from sqlalchemy import or_

    print(f"\n📥 DEBUG: Cliente logado: {current_user.id} ({current_user.nome})")

    checkins_qr = (
        Checkin.query
        .outerjoin(Checkin.oficina)
        .outerjoin(Checkin.usuario)
        .filter(
            Checkin.palavra_chave.in_(['QR-AUTO', 'QR-EVENTO', 'QR-OFICINA', 'QR‑OFICINA']),
            or_(
                Usuario.cliente_id == current_user.id,
                Oficina.cliente_id == current_user.id,
                Checkin.cliente_id == current_user.id
            )
        )
        .order_by(Checkin.data_hora.desc())
        .all()
    )

    print(f"📊 DEBUG: Total de check-ins encontrados: {len(checkins_qr)}")
    for i, ck in enumerate(checkins_qr, start=1):
        print(f" - Checkin {i}: Usuario={ck.usuario.nome if ck.usuario else 'N/A'} | "
              f"Email={ck.usuario.email if ck.usuario else 'N/A'} | "
              f"Palavra={ck.palavra_chave} | "
              f"Oficina={'N/A' if not ck.oficina else ck.oficina.titulo} | "
              f"ClienteID={ck.cliente_id}")

    if not checkins_qr:
        flash("Não há check-ins via QR Code para gerar o relatório.", "warning")
        return redirect(url_for('routes.dashboard'))

    # (continua com sua lógica atual do PDF sem alterações)
    # 2. Configuração do documento
    current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
    pdf_filename = f"checkins_qr_{current_time}.pdf"
    pdf_dir = os.path.join("static", "relatorios")
    os.makedirs(pdf_dir, exist_ok=True)
    pdf_path = os.path.join(pdf_dir, pdf_filename)

    # 3. Definição de estilos personalizados
    styles = getSampleStyleSheet()
    
    # Estilo para o título principal
    title_style = ParagraphStyle(
        name='CustomTitle',
        parent=styles['Title'],
        fontSize=16,
        textColor=colors.HexColor("#023E8A"),
        spaceAfter=12,
        alignment=TA_CENTER
    )
    
    # Estilo para subtítulos (oficinas)
    subtitle_style = ParagraphStyle(
        name='CustomSubtitle',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor("#1A75CF"),
        spaceBefore=15,
        spaceAfter=10,
        borderWidth=1,
        borderColor=colors.HexColor("#1A75CF"),
        borderPadding=5,
        borderRadius=3
    )
    
    # Estilo para o rodapé
    footer_style = ParagraphStyle(
        name='Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.gray,
        alignment=TA_RIGHT
    )

    doc = SimpleDocTemplate(
        pdf_path,
        pagesize=landscape(A4),
        rightMargin=15*mm,
        leftMargin=15*mm,
        topMargin=20*mm,
        bottomMargin=15*mm
    )
    elements = []

    # 5. Cabeçalho do relatório
    elements.append(Paragraph("Relatório de Check-ins via QR Code", title_style))
    elements.append(Paragraph(f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}", footer_style))
    elements.append(Spacer(1, 10*mm))

    # 6. Configuração do fuso horário (Brasil)
    brasilia_tz = pytz.timezone("America/Sao_Paulo")
    
    def convert_to_brasilia(dt):
        """Converte datetime para horário de Brasília."""
        if dt is None:
            return None
        # Se o datetime não for "aware", assume-se que está em UTC
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=pytz.utc)
        return dt.astimezone(brasilia_tz)

    # 7. Agrupamento dos check-ins por oficina
    oficinas_grupos = defaultdict(list)
    for checkin in checkins_qr:
        oficina_titulo = checkin.oficina.titulo if checkin.oficina else "Oficina não especificada"
        oficinas_grupos[oficina_titulo].append(checkin)

    # 8. Definição do estilo de tabela
    table_style = TableStyle([
        # Cabeçalho
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#8ecde6")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        
        # Corpo da tabela
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        
        # Linhas zebradas
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
        
        # Bordas
        ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ('BOX', (0, 0), (-1, -1), 1, colors.grey),
        
        # Repetir cabeçalho em novas páginas
        ('REPEATROWS', (0, 0), (0, 0)),
        
        # Habilitar quebra de linha
        ('WORDWRAP', (0, 0), (-1, -1), True),
    ])

    # 9. Gerar tabelas para cada oficina
    total_checkins = 0
    
    for oficina_titulo, checkins in oficinas_grupos.items():
        total_oficina = len(checkins)
        total_checkins += total_oficina
        
        # Adicionar subtítulo da oficina
        elements.append(Paragraph(f"Oficina: {oficina_titulo} ({total_oficina} check-ins)", subtitle_style))
        
        # Preparar dados da tabela
        # Usamos Paragraph para cada célula, o que permite o WORDWRAP aplicado acima.
        table_data = [[
            Paragraph("Nome do Participante", styles["Normal"]),
            Paragraph("E-mail", styles["Normal"]),
            Paragraph("Data/Hora do Check-in", styles["Normal"])
        ]]
        
        for ck in checkins:
            usuario = ck.usuario
            nome = usuario.nome if usuario else "N/A"
            email = usuario.email if usuario else "N/A"
            
            # Converter para horário de Brasília
            dt_local = convert_to_brasilia(ck.data_hora)
            data_str = dt_local.strftime('%d/%m/%Y %H:%M') if dt_local else "N/A"
            
            table_data.append([
                Paragraph(nome, styles["Normal"]),
                Paragraph(email, styles["Normal"]),
                Paragraph(data_str, styles["Normal"])
            ])
        
        # Definir larguras das colunas
        col_widths = [
            doc.width * 0.35,
            doc.width * 0.35,
            doc.width * 0.3
        ]
        
        table = LongTable(table_data, colWidths=col_widths)
        table.setStyle(table_style)
        elements.append(table)
        elements.append(Spacer(1, 8*mm))

    # 10. Resumo final
    elements.append(Spacer(1, 10*mm))
    elements.append(Paragraph(f"Total de check-ins: {total_checkins}", styles["Heading3"]))
    
    # 11. Rodapé com informações do sistema
    footer_text = f"Documento gerado pelo sistema AppFiber em {datetime.now().strftime('%d/%m/%Y às %H:%M:%S')}"
    elements.append(Spacer(1, 20*mm))
    elements.append(Paragraph(footer_text, footer_style))

    # 12. Gerar o PDF
    doc.build(elements)
    
    # 13. Retornar o arquivo para download
    return send_file(pdf_path, as_attachment=True, download_name=pdf_filename)

@routes.route('/gerenciar_participantes', methods=['GET'])
@login_required
def gerenciar_participantes():
    # Verifique se é admin
    if current_user.tipo != 'admin':
        flash('Acesso negado!', 'danger')
        return redirect(url_for('routes.dashboard'))

    # Busca todos os usuários cujo tipo é 'participante'
    participantes = Usuario.query.filter_by(tipo='participante').all()

    # Renderiza um template parcial (ou completo). Você pode renderizar
    # a página inteira ou só retornar JSON. Aqui vamos supor que renderiza a modal.
    return render_template('gerenciar_participantes.html', participantes=participantes)

@routes.route('/excluir_participante/<int:participante_id>', methods=['POST'])
@login_required
def excluir_participante(participante_id):
    if current_user.tipo != 'admin':
        flash('Acesso negado!', 'danger')
        return redirect(url_for('routes.dashboard'))
    
    participante = Usuario.query.get_or_404(participante_id)
    if participante.tipo != 'participante':
        flash('Esse usuário não é um participante.', 'danger')
        return redirect(url_for('routes.dashboard'))
    
    db.session.delete(participante)
    db.session.commit()
    flash('Participante excluído com sucesso!', 'success')
    return redirect(url_for('routes.dashboard'))

@routes.route('/editar_participante_admin/<int:participante_id>', methods=['POST'])
@login_required
def editar_participante_admin(participante_id):
    if current_user.tipo != 'admin':
        flash('Acesso negado!', 'danger')
        return redirect(url_for('routes.dashboard'))
    
    participante = Usuario.query.get_or_404(participante_id)
    if participante.tipo != 'participante':
        flash('Esse usuário não é um participante.', 'danger')
        return redirect(url_for('routes.dashboard'))

    # Captura os dados do form
    nome = request.form.get('nome')
    cpf = request.form.get('cpf')
    email = request.form.get('email')
    formacao = request.form.get('formacao')
    nova_senha = request.form.get('senha')

    # Atualiza
    participante.nome = nome
    participante.cpf = cpf
    participante.email = email
    participante.formacao = formacao
    if nova_senha:
        participante.senha = generate_password_hash(nova_senha)

    db.session.commit()
    flash('Participante atualizado com sucesso!', 'success')
    return redirect(url_for('routes.dashboard'))

@routes.route('/gerar_relatorio_mensagem', methods=['GET'])
@login_required
def gerar_relatorio_mensagem():
    from sqlalchemy import func
    
    # Se quiser só as oficinas do cliente, verifique se current_user é admin ou cliente:
    is_admin = (current_user.tipo == 'admin')
    if is_admin:
        total_oficinas = Oficina.query.count()
        # Buscar todas as oficinas para calcular o total de vagas considerando o tipo_inscricao
        oficinas = Oficina.query.options(db.joinedload(Oficina.inscritos)).all()
        total_inscricoes = Inscricao.query.count()
        eventos = Evento.query.all()
    else:
        total_oficinas = Oficina.query.filter_by(cliente_id=current_user.id).count()
        # Buscar oficinas do cliente para calcular o total de vagas considerando o tipo_inscricao
        oficinas = Oficina.query.filter_by(cliente_id=current_user.id).options(db.joinedload(Oficina.inscritos)).all()
        total_inscricoes = Inscricao.query.join(Oficina).filter(Oficina.cliente_id == current_user.id).count()
        eventos = Evento.query.filter_by(cliente_id=current_user.id).all()
    
    # Novo cálculo do total_vagas conforme solicitado:
    # 1. Soma as vagas das oficinas com tipo_inscricao 'com_inscricao_com_limite'
    # 2. Soma o número de inscritos nas oficinas com tipo_inscricao 'com_inscricao_sem_limite'
    total_vagas = 0
    for of in oficinas:
        if of.tipo_inscricao == 'com_inscricao_com_limite':
            total_vagas += of.vagas
        elif of.tipo_inscricao == 'com_inscricao_sem_limite':
            total_vagas += len(of.inscritos)
    
    # Cálculo de adesão
    percentual_adesao = (total_inscricoes / total_vagas) * 100 if total_vagas > 0 else 0

    # Monta a mensagem com emojis e loop
    total_eventos = len(eventos)
    mensagem = (
        "📊 *Relatório do Sistema*\n\n"
        f"✅ *Total de Eventos:* {total_eventos}\n"
        f"✅ *Total de Oficinas:* {total_oficinas}\n"
        f"✅ *Vagas Ofertadas:* {total_vagas}\n"
        f"✅ *Vagas Preenchidas:* {total_inscricoes}\n"
        f"✅ *% de Adesão:* {percentual_adesao:.2f}%\n\n"
        "----------------------------------------\n"
    )
    
    # Agrupar oficinas por evento
    for evento in eventos:
        # Buscar oficinas deste evento
        if is_admin:
            oficinas_evento = Oficina.query.filter_by(evento_id=evento.id).all()
        else:
            oficinas_evento = Oficina.query.filter_by(evento_id=evento.id, cliente_id=current_user.id).all()
        
        # Se não houver oficinas neste evento, pular
        if not oficinas_evento:
            continue
            
        # Adicionar cabeçalho do evento
        mensagem += f"\n🎪 *EVENTO: {evento.nome}*\n"
        mensagem += f"📌 *Total de Oficinas no Evento:* {len(oficinas_evento)}\n"
        
        # Adicionar dados de cada oficina do evento
        for oficina in oficinas_evento:
            # Conta inscritos
            num_inscritos = Inscricao.query.filter_by(oficina_id=oficina.id).count()
            
            # Calcula ocupação considerando o tipo de inscrição
            if oficina.tipo_inscricao == 'sem_inscricao':
                ocupacao = 0  # Não é relevante calcular ocupação
                vagas_texto = "N/A (sem inscrição)"
            elif oficina.tipo_inscricao == 'com_inscricao_sem_limite':
                ocupacao = 100  # Sempre 100% pois aceita qualquer número de inscritos
                vagas_texto = "Ilimitadas"
            else:  # com_inscricao_com_limite
                ocupacao = (num_inscritos / oficina.vagas)*100 if oficina.vagas else 0
                vagas_texto = str(oficina.vagas)
            
            # Determina o texto amigável para o tipo de inscrição
            tipo_inscricao_texto = "Sem inscrição"
            if oficina.tipo_inscricao == "com_inscricao_sem_limite":
                tipo_inscricao_texto = "Inscrição sem limite de vagas"
            elif oficina.tipo_inscricao == "com_inscricao_com_limite":
                tipo_inscricao_texto = "Inscrição com vagas limitadas"
        
        mensagem += (
            f"\n🎓 *Oficina:* {oficina.titulo}\n"
                f"🔹 *Tipo de Inscrição:* {tipo_inscricao_texto}\n"
                f"🔹 *Vagas:* {vagas_texto}\n"
            f"🔹 *Inscritos:* {num_inscritos}\n"
            f"🔹 *Ocupação:* {ocupacao:.2f}%\n"
        )
        
        mensagem += "----------------------------------------\n"

    return mensagem


@routes.route('/cancelar_inscricoes_lote', methods=['POST'])
@login_required
def cancelar_inscricoes_lote():
    # Verifica se é admin
    if current_user.tipo != 'admin':
        flash("Acesso negado!", "danger")
        return redirect(url_for('routes.dashboard'))

    # Pega os IDs marcados
    inscricao_ids = request.form.getlist('inscricao_ids')
    if not inscricao_ids:
        flash("Nenhuma inscrição selecionada!", "warning")
        return redirect(url_for('routes.dashboard'))

    # Converte para int
    inscricao_ids = list(map(int, inscricao_ids))

    try:
        # Busca todas as inscrições com esses IDs
        inscricoes = Inscricao.query.filter(Inscricao.id.in_(inscricao_ids)).all()
        # Cancela removendo do banco
        for insc in inscricoes:
            db.session.delete(insc)

        db.session.commit()
        flash(f"Foram canceladas {len(inscricoes)} inscrições!", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Erro ao cancelar inscrições: {e}", "danger")

    return redirect(url_for('routes.dashboard'))

from models import Inscricao, Oficina, Usuario   # ➊ certifique-se do import

@routes.route('/inscricoes_lote', methods=['POST'])
@login_required
def inscricoes_lote():
    # ─── 0. Permissão ──────────────────────────────────────────────
    if current_user.tipo not in {'admin', 'cliente'}:
        flash("Acesso negado!", "danger")
        return redirect(url_for('routes.dashboard'))

    # ─── 1. Dados do formulário ────────────────────────────────────
    inscricao_ids      = list(map(int, request.form.getlist('inscricao_ids')))
    oficina_destino_id = request.form.get('oficina_destino', type=int)
    acao               = request.form.get('acao', 'mover')  # mover | copiar

    if not inscricao_ids or not oficina_destino_id:
        flash("Selecione inscrições e oficina de destino.", "warning")
        return redirect(url_for('routes.dashboard'))

    oficina_destino = Oficina.query.get_or_404(oficina_destino_id)
    inscricoes      = Inscricao.query.filter(Inscricao.id.in_(inscricao_ids)).all()

    # ─── 2. Restrições de cliente ─────────────────────────────────
    if current_user.tipo == 'cliente' and any(i.cliente_id != current_user.id for i in inscricoes):
        flash("Há inscrições de outro cliente selecionadas!", "danger")
        return redirect(url_for('routes.dashboard'))

    # ─── 3. Verificar vagas de uma vez só ─────────────────────────
    if oficina_destino.vagas < len(inscricoes):
        flash(f"Não há vagas suficientes! "
              f"(Disponíveis: {oficina_destino.vagas}, Necessárias: {len(inscricoes)})",
              "danger")
        return redirect(url_for('routes.dashboard'))

    # ─── 4. Processar lote ────────────────────────────────────────
    try:
        usuarios_afetados = set()          # para atualizar evento_id depois

        for insc in inscricoes:
            usuarios_afetados.add(insc.usuario_id)

            if acao == 'mover':
                # devolve vaga na origem (se houver)
                if insc.oficina:
                    insc.oficina.vagas += 1

                # ocupa vaga na destino
                oficina_destino.vagas -= 1

                # atualiza inscrição existente
                insc.oficina_id = oficina_destino.id
                insc.evento_id  = oficina_destino.evento_id

            else:  # acao == 'copiar'
                # ocupa vaga na destino
                oficina_destino.vagas -= 1

                # clona inscrição
                nova = Inscricao(
                    usuario_id       = insc.usuario_id,
                    cliente_id       = insc.cliente_id,
                    oficina_id       = oficina_destino.id,
                    evento_id        = oficina_destino.evento_id,
                    status_pagamento = insc.status_pagamento
                )
                db.session.add(nova)

        # ─── 5. Sincronizar usuario.evento_id ─────────────────────
        for uid in usuarios_afetados:
            usuario = Usuario.query.get(uid)
            # se o sistema admite apenas 1 evento corrente por usuário:
            if usuario and usuario.evento_id != oficina_destino.evento_id:
                usuario.evento_id = oficina_destino.evento_id

        db.session.commit()
        verbo = "movida(s)" if acao == 'mover' else "copiada(s)"
        flash(f"{len(inscricoes)} inscrição(ões) {verbo} com sucesso!", "success")

    except Exception as e:
        db.session.rollback()
        flash(f"Erro ao processar inscrições: {e}", "danger")

    return redirect(url_for('routes.gerenciar_inscricoes'))



@routes.route("/mover_inscricoes_lote", methods=["POST"])
@login_required
def mover_inscricoes_lote():
    # ░░░ 1. Permissão --------------------------------------------------------
    if current_user.tipo not in {"admin", "cliente"}:
        flash("Acesso negado!", "danger")
        return redirect(url_for("routes.dashboard"))

    # ░░░ 2. Coleta de parâmetros --------------------------------------------
    inscricao_ids      = list(map(int, request.form.getlist("inscricao_ids")))
    oficina_destino_id = request.form.get("oficina_destino", type=int)

    if not inscricao_ids:
        flash("Nenhuma inscrição selecionada!", "warning")
        return redirect(url_for("routes.dashboard"))

    if not oficina_destino_id:
        flash("Nenhuma oficina de destino selecionada!", "warning")
        return redirect(url_for("routes.dashboard"))

    try:
        # ░░░ 3. Objetos de referência ---------------------------------------
        primeira_insc   = Inscricao.query.get_or_404(inscricao_ids[0])
        oficina_origem  = primeira_insc.oficina
        evento_id       = oficina_origem.evento_id
        oficina_destino = Oficina.query.get_or_404(oficina_destino_id)

        # 3.1 – Garantir que destino pertence ao mesmo evento
        if oficina_destino.evento_id != evento_id:
            flash("A oficina de destino deve pertencer ao mesmo evento!", "danger")
            return redirect(url_for("routes.dashboard"))

        # 3.2 – Buscar inscrições a mover
        inscricoes = (
            Inscricao.query
            .filter(Inscricao.id.in_(inscricao_ids))
            .all()
        )

        # 3.3 – Checar se todas são do mesmo evento
        if any(insc.oficina.evento_id != evento_id for insc in inscricoes):
            flash("Todas as inscrições devem pertencer ao mesmo evento!", "danger")
            return redirect(url_for("routes.dashboard"))

        # 3.4 – Verificar vagas
        if oficina_destino.vagas < len(inscricoes):
            flash(
                f"Não há vagas suficientes na oficina de destino! "
                f"(Disponível: {oficina_destino.vagas}, Necessário: {len(inscricoes)})",
                "danger",
            )
            return redirect(url_for("routes.dashboard"))

        # ░░░ 4. Garante inscrição-evento para quem não tem -------------------
        usuario_ids = {insc.usuario_id for insc in inscricoes}

        existentes = (
            Inscricao.query
            .filter(
                Inscricao.evento_id == evento_id,
                Inscricao.oficina_id.is_(None),           # inscrição “geral” do evento
                Inscricao.usuario_id.in_(usuario_ids)
            )
            .with_entities(Inscricao.usuario_id)
            .all()
        )
        ja_inscritos_evento = {row.usuario_id for row in existentes}
        faltantes = usuario_ids - ja_inscritos_evento

        novas_inscricoes_evento = [
            Inscricao(
                usuario_id=u_id,
                cliente_id=primeira_insc.cliente_id,
                evento_id=evento_id,
                status_pagamento="paid"   # ajuste conforme sua regra de cobrança
            )
            for u_id in faltantes
        ]
        db.session.add_all(novas_inscricoes_evento)

        # ░░░ 5. Move as inscrições de oficina -------------------------------
        for insc in inscricoes:
            # devolve vaga na oficina de origem
            insc.oficina.vagas += 1
            # retira vaga da oficina destino
            oficina_destino.vagas -= 1
            # efetiva a troca
            insc.oficina_id = oficina_destino_id
            # (mantém vínculo com o mesmo evento pela coerência)
            insc.evento_id  = evento_id

        db.session.commit()

        flash(
            f"{len(inscricoes)} inscrição(ões) movida(s) para "
            f"“{oficina_destino.titulo}”. "
            f"{len(novas_inscricoes_evento)} inscrição(ões) de evento criadas.",
            "success"
        )

    except Exception as e:
        db.session.rollback()
        flash(f"Erro ao mover inscrições: {e}", "danger")

    return redirect(url_for("routes.gerenciar_inscricoes"))


@routes.route('/api/oficinas_mesmo_evento/<int:oficina_id>')
@login_required
def oficinas_mesmo_evento(oficina_id):
    oficina = Oficina.query.get_or_404(oficina_id)
    evento = oficina.evento
    oficinas = Oficina.query.filter(
        Oficina.evento_id == oficina.evento_id,
        Oficina.id != oficina_id
    ).all()
    return jsonify({
        'evento_nome': evento.nome if evento else 'Sem evento',
        'oficinas': [
            {
                'id': o.id,
                'titulo': o.titulo,
                'vagas': o.vagas,
                'evento_nome': evento.nome if evento else 'Sem evento'
            } for o in oficinas
        ]
    })
    
# routes.py  (ou onde ficam suas rotas API)
@routes.route('/api/eventos_com_oficinas')
@login_required
def eventos_com_oficinas():
    # ① Se for cliente, mostra só os eventos dele.  Admin vê todos.
    q = Evento.query
    if current_user.tipo == 'cliente':
        q = q.filter_by(cliente_id=current_user.id)

    eventos = q.order_by(Evento.nome).all()
    payload = []
    for ev in eventos:
        payload.append({
            'id': ev.id,
            'nome': ev.nome,
            'oficinas': [
                {'id': o.id, 'titulo': o.titulo, 'vagas': o.vagas}
                for o in ev.oficinas
            ]
        })
    return jsonify(payload)


@routes.route('/cancelar_inscricao/<int:inscricao_id>', methods=['GET','POST'])
@login_required
def cancelar_inscricao(inscricao_id):
    # Allow both admin and client access
    if current_user.tipo not in ['admin', 'cliente']:
        flash("Acesso negado!", "danger")
        return redirect(url_for('routes.dashboard'))

    # Get inscription
    insc = Inscricao.query.get_or_404(inscricao_id)
    
    # For clients, verify they own the workshop/event
    if current_user.tipo == 'cliente':
        oficina = Oficina.query.get(insc.oficina_id)
        if oficina.cliente_id != current_user.id:
            flash("Você não tem permissão para cancelar esta inscrição!", "danger")
            return redirect(url_for('routes.dashboard_cliente'))

    try:
        db.session.delete(insc)
        db.session.commit()
        flash("Inscrição cancelada com sucesso!", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Erro ao cancelar inscrição: {e}", "danger")

    # Redirect to appropriate dashboard based on user type
    if current_user.tipo == 'admin':
        return redirect(url_for('routes.dashboard'))
    else:
        return redirect(url_for('routes.dashboard_cliente'))
    


@routes.route('/dashboard_cliente')
@login_required
def dashboard_cliente():
    if current_user.tipo != 'cliente':
        return redirect(url_for('routes.dashboard'))

    print(f"📌 [DEBUG] Cliente autenticado: {current_user.email} (ID: {current_user.id})")
    print("Usuário logado:", current_user.email)
    print("ID:", current_user.id)
    print("Tipo:", current_user.tipo if hasattr(current_user, 'tipo') else "N/A")


    eventos = Evento.query.filter_by(cliente_id=current_user.id).all()
    print("✅ Eventos:", eventos)
  
    

    # Mostra apenas as oficinas criadas por este cliente OU pelo admin (cliente_id nulo)
    oficinas = Oficina.query.filter_by(cliente_id=current_user.id).all()

    # Cálculo das estatísticas
    total_oficinas = len(oficinas)
    
    # Novo cálculo do total_vagas conforme solicitado:
    # 1. Soma as vagas das oficinas com tipo_inscricao 'com_inscricao_com_limite'
    # 2. Soma o número de inscritos nas oficinas com tipo_inscricao 'com_inscricao_sem_limite'
    total_vagas = 0
    for of in oficinas:
        if of.tipo_inscricao == 'com_inscricao_com_limite':
            total_vagas += of.vagas
        elif of.tipo_inscricao == 'com_inscricao_sem_limite':
            total_vagas += len(of.inscritos)
    
    total_inscricoes = Inscricao.query.join(Oficina).filter(
        (Oficina.cliente_id == current_user.id) | (Oficina.cliente_id.is_(None))
    ).count()
    percentual_adesao = (total_inscricoes / total_vagas) * 100 if total_vagas > 0 else 0

    # Busca apenas check-ins via QR dos usuários do cliente
    from sqlalchemy import or_

    checkins_via_qr = (
        Checkin.query
        .outerjoin(Checkin.oficina)
        .outerjoin(Checkin.usuario)
        .filter(
            Checkin.palavra_chave.in_(['QR-AUTO', 'QR-EVENTO']),
            or_(
                Usuario.cliente_id == current_user.id,
                Oficina.cliente_id == current_user.id,
                Checkin.cliente_id == current_user.id  # segurança extra
            )
        )
        .order_by(Checkin.data_hora.desc())
        .all()
    )


    # Se for para filtrar pela coluna Inscricao.cliente_id:
    inscritos = Inscricao.query.filter(
        (Inscricao.cliente_id == current_user.id) | (Inscricao.cliente_id.is_(None))
    ).all()
    
     # Buscar eventos ativos
    eventos_ativos = Evento.query.filter_by(cliente_id=current_user.id).all()
    total_eventos = len(eventos_ativos)
    
    # Dados para cards
    agendamentos_totais = db.session.query(func.count(AgendamentoVisita.id)).join(
        HorarioVisitacao, AgendamentoVisita.horario_id == HorarioVisitacao.id
    ).join(
        Evento, HorarioVisitacao.evento_id == Evento.id
    ).filter(
        Evento.cliente_id == current_user.id
    ).scalar() or 0
    
    agendamentos_confirmados = db.session.query(func.count(AgendamentoVisita.id)).join(
        HorarioVisitacao, AgendamentoVisita.horario_id == HorarioVisitacao.id
    ).join(
        Evento, HorarioVisitacao.evento_id == Evento.id
    ).filter(
        Evento.cliente_id == current_user.id,
        AgendamentoVisita.status == 'confirmado'
    ).scalar() or 0
    
    agendamentos_realizados = db.session.query(func.count(AgendamentoVisita.id)).join(
        HorarioVisitacao, AgendamentoVisita.horario_id == HorarioVisitacao.id
    ).join(
        Evento, HorarioVisitacao.evento_id == Evento.id
    ).filter(
        Evento.cliente_id == current_user.id,
        AgendamentoVisita.status == 'realizado'
    ).scalar() or 0
    
    agendamentos_cancelados = db.session.query(func.count(AgendamentoVisita.id)).join(
        HorarioVisitacao, AgendamentoVisita.horario_id == HorarioVisitacao.id
    ).join(
        Evento, HorarioVisitacao.evento_id == Evento.id
    ).filter(
        Evento.cliente_id == current_user.id,
        AgendamentoVisita.status == 'cancelado'
    ).scalar() or 0
    
    # Total de visitantes
    total_visitantes = db.session.query(func.sum(AgendamentoVisita.quantidade_alunos)).join(
        HorarioVisitacao, AgendamentoVisita.horario_id == HorarioVisitacao.id
    ).join(
        Evento, HorarioVisitacao.evento_id == Evento.id
    ).filter(
        Evento.cliente_id == current_user.id,
        AgendamentoVisita.status.in_(['confirmado', 'realizado'])
    ).scalar() or 0
    
    # Agendamentos para hoje
    hoje = datetime.utcnow().date()
    agendamentos_hoje = AgendamentoVisita.query.join(
        HorarioVisitacao, AgendamentoVisita.horario_id == HorarioVisitacao.id
    ).join(
        Evento, HorarioVisitacao.evento_id == Evento.id
    ).filter(
        Evento.cliente_id == current_user.id,
        HorarioVisitacao.data == hoje,
        AgendamentoVisita.status == 'confirmado'
    ).order_by(
        HorarioVisitacao.horario_inicio
    ).all()
    
    # Próximos agendamentos (próximos 7 dias, excluindo hoje)
    data_limite = hoje + timedelta(days=7)
    proximos_agendamentos = AgendamentoVisita.query.join(
        HorarioVisitacao, AgendamentoVisita.horario_id == HorarioVisitacao.id
    ).join(
        Evento, HorarioVisitacao.evento_id == Evento.id
    ).filter(
        Evento.cliente_id == current_user.id,
        HorarioVisitacao.data > hoje,
        HorarioVisitacao.data <= data_limite,
        AgendamentoVisita.status == 'confirmado'
    ).order_by(
        HorarioVisitacao.data,
        HorarioVisitacao.horario_inicio
    ).limit(5).all()
    
    # Calcular ocupação média (vagas preenchidas / capacidade total) 
    ocupacao_query = db.session.query(
        func.sum(HorarioVisitacao.capacidade_total - HorarioVisitacao.vagas_disponiveis).label('ocupadas'),
        func.sum(HorarioVisitacao.capacidade_total).label('total')
    ).join(
        Evento, HorarioVisitacao.evento_id == Evento.id
    ).filter(
        Evento.cliente_id == current_user.id,
        HorarioVisitacao.data >= hoje
    ).first()
    
    ocupacao_media = 0
    if ocupacao_query and ocupacao_query.total and ocupacao_query.total > 0:
        ocupacao_media = (ocupacao_query.ocupadas / ocupacao_query.total) * 100

    
    # Buscar config específica do cliente
    config_cliente = ConfiguracaoCliente.query.filter_by(cliente_id=current_user.id).first()
    # Se não existir, cria:
    if not config_cliente:
        config_cliente = ConfiguracaoCliente(
            cliente_id=current_user.id,
            permitir_checkin_global=False,
            habilitar_feedback=False,
            habilitar_certificado_individual=False
        )
        db.session.add(config_cliente)
        db.session.commit()

    return render_template(
        'dashboard_cliente.html',
        usuario=current_user,
        oficinas=oficinas,
        total_oficinas=total_oficinas,
        total_vagas=total_vagas,
        total_inscricoes=total_inscricoes,
        percentual_adesao=percentual_adesao,
        checkins_via_qr=checkins_via_qr,
        inscritos=inscritos,
        config_cliente=config_cliente,
        eventos_ativos=eventos_ativos,
        agendamentos_totais=agendamentos_totais,
        agendamentos_confirmados=agendamentos_confirmados,
        agendamentos_realizados=agendamentos_realizados,
        agendamentos_cancelados=agendamentos_cancelados,
        total_visitantes=total_visitantes,
        agendamentos_hoje=agendamentos_hoje,
        proximos_agendamentos=proximos_agendamentos,
        ocupacao_media=ocupacao_media,
        total_eventos=total_eventos,
        eventos=eventos
    )
    
def obter_configuracao_do_cliente(cliente_id):
    config = ConfiguracaoCliente.query.filter_by(cliente_id=cliente_id).first()
    if not config:
        config = ConfiguracaoCliente(
            cliente_id=cliente_id,
            permitir_checkin_global=False,
            habilitar_feedback=False,
            habilitar_certificado_individual=False,
        )
        db.session.add(config)
        db.session.commit()
    return config


@routes.route('/oficinas_disponiveis')
@login_required
def oficinas_disponiveis():
    oficinas = Oficina.query.filter_by(cliente_id=current_user.cliente_id).all()
    return render_template('oficinas.html', oficinas=oficinas)

@routes.route('/gerar_link', methods=['GET', 'POST'])
@login_required
def gerar_link():
    if current_user.tipo not in ['cliente', 'admin']:
        return "Forbidden", 403

    cliente_id = current_user.id

    if request.method == 'POST':
        # Obtém os dados JSON da requisição
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': 'Nenhum dado enviado'}), 400

        evento_id = data.get('evento_id')
        slug_customizado = data.get('slug_customizado')

        if not evento_id:
            return jsonify({'success': False, 'message': 'Evento não especificado'}), 400

        # Verifica se o evento pertence ao cliente
        evento = Evento.query.filter_by(id=evento_id, cliente_id=cliente_id).first()
        if not evento and current_user.tipo != 'admin':
            return jsonify({'success': False, 'message': 'Evento inválido ou não autorizado'}), 403

        # Gera um token único
        novo_token = str(uuid.uuid4())

        # Valida e limpa o slug personalizado
        if slug_customizado:
            slug_customizado = slug_customizado.strip().lower().replace(' ', '-')
            if LinkCadastro.query.filter_by(slug_customizado=slug_customizado).first():
                return jsonify({'success': False, 'message': 'Slug já está em uso'}), 400
        else:
            slug_customizado = None

        # Cria o link de cadastro no banco
        novo_link = LinkCadastro(
            cliente_id=cliente_id,
            evento_id=evento_id,
            token=novo_token,
            slug_customizado=slug_customizado
        )
        db.session.add(novo_link)
        db.session.commit()

        # Define a URL base dependendo do ambiente
        if request.host.startswith("127.0.0.1") or "localhost" in request.host:
            base_url = "http://127.0.0.1:5000"  # URL local
        else:
            base_url = "https://appfiber.com.br"  # URL de produção

        # Gera o link final
        if slug_customizado:
            link_gerado = f"{base_url}/inscricao/{slug_customizado}"
        else:
            link_gerado = f"{base_url}{url_for('routes.cadastro_participante', token=novo_token)}"

        return jsonify({'success': True, 'link': link_gerado})

    # Para GET, verificamos se é uma solicitação para listar links de um evento específico
    evento_id = request.args.get('evento_id')
    if evento_id:
        # Verificar se o evento pertence ao cliente atual
        evento = Evento.query.filter_by(id=evento_id, cliente_id=cliente_id).first()
        if not evento and current_user.tipo != 'admin':
            return jsonify({'success': False, 'links': [], 'message': 'Evento não autorizado'}), 403
        
        # Buscar todos os links para este evento
        links = LinkCadastro.query.filter_by(evento_id=evento_id, cliente_id=cliente_id).all()
        
        # Montar a lista de links com URLs completas
        links_list = []
        for link in links:
            if request.host.startswith("127.0.0.1") or "localhost" in request.host:
                base_url = "http://127.0.0.1:5000"
            else:
                base_url = "https://appfiber.com.br"
                
            if link.slug_customizado:
                url = f"{base_url}/inscricao/{link.slug_customizado}"
            else:
                url = f"{base_url}{url_for('routes.cadastro_participante', token=link.token)}"
                
            links_list.append({
                'id': link.id,
                'token': link.token,
                'slug': link.slug_customizado,
                'url': url,
                'criado_em': link.criado_em.isoformat()
            })
            
        return jsonify({'success': True, 'links': links_list})
        
    # Para GET sem evento_id, apenas retorna os eventos disponíveis
    eventos = Evento.query.filter_by(cliente_id=cliente_id).all()
    return jsonify({'eventos': [{'id': e.id, 'nome': e.nome} for e in eventos]})

@routes.route('/excluir_link', methods=['POST'])
@login_required
def excluir_link():
    if current_user.tipo not in ['cliente', 'admin']:
        return jsonify({'success': False, 'message': 'Não autorizado'}), 403
        
    data = request.get_json()
    if not data or 'link_id' not in data:
        return jsonify({'success': False, 'message': 'ID do link não fornecido'}), 400
        
    link_id = data['link_id']
    link = LinkCadastro.query.get(link_id)
    
    if not link:
        return jsonify({'success': False, 'message': 'Link não encontrado'}), 404
        
    if link.cliente_id != current_user.id and current_user.tipo != 'admin':
        return jsonify({'success': False, 'message': 'Não autorizado a excluir este link'}), 403
        
    db.session.delete(link)
    db.session.commit()
    
    return jsonify({'success': True, 'message': 'Link excluído com sucesso'})

@routes.route('/evento/<int:evento_id>/inscricao')
def inscricao_evento(evento_id):
    """
    Redireciona para a página de inscrição do evento usando o LinkCadastro existente
    ou para a inscrição direta caso o evento permita
    """
    evento = Evento.query.get_or_404(evento_id)
    
    # 1. Primeiro verifica se o evento está ativo
    if evento.status != 'ativo':
        flash('Este evento não está disponível para inscrições no momento.', 'warning')
        return redirect(url_for('routes.visualizar_evento', evento_id=evento_id))
    
    # 2. Busca um link de cadastro para este evento
    link = LinkCadastro.query.filter_by(evento_id=evento_id).first()
    
    if link:
        # Se existe um link personalizado, usa ele
        if link.slug_customizado:
            return redirect(url_for('routes.inscricao', slug=link.slug_customizado))
        # Senão, usa o token
        return redirect(url_for('routes.cadastro_participante', token=link.token))
    
    # 3. Se não existe link, verifica se o evento permite inscrição direta
    if evento.publico and not evento.requer_aprovacao:
        # Aqui está a correção - não usamos mais link.token pois link é None
        # Criamos um token temporário ou usamos um método alternativo
        return redirect(url_for('routes.cadastro_participante', evento_id=evento_id))
    
    # 4. Fallback final
    flash('Este evento não possui inscrições abertas no momento.', 'warning')
    return redirect(url_for('routes.visualizar_evento', evento_id=evento_id))
    
@routes.route('/inscricao/<slug_customizado>', methods=['GET'])
def inscricao_personalizada(slug_customizado):
    # Busca o LinkCadastro pelo slug personalizado
    link = LinkCadastro.query.filter_by(slug_customizado=slug_customizado).first()
    if not link or not link.evento_id:
        return "Link inválido ou sem evento associado", 404

    # Redireciona para a rota cadastro_participante com o token
    return redirect(url_for('routes.cadastro_participante', token=link.token))


@app.route('/inscricao/<token>', methods=['GET', 'POST'])
def inscricao(token):
    cliente = Cliente.query.filter_by(token=token).first()
    
    if not cliente:
        return "Link inválido", 404

    if request.method == 'POST':
        novo_usuario = Usuario(
            email=request.form['email'],
            senha_hash=generate_password_hash(request.form['senha']),
            cliente_id=cliente.id,
            tipo="usuario"
        )
        db.session.add(novo_usuario)
        db.session.commit()
        return redirect(url_for('login'))

    return render_template('inscricao.html', cliente=cliente)

@app.route('/superadmin_dashboard')
@login_required
def superadmin_dashboard():
    if not current_user.is_superuser():
        return abort(403)

    clientes = Cliente.query.all()
    return render_template('dashboard_superadmin.html', clientes=clientes)


@routes.route('/toggle_cliente/<int:cliente_id>')
@login_required
def toggle_cliente(cliente_id):
    if current_user.tipo != 'admin':
        flash('Acesso negado!', 'danger')
        return redirect(url_for('routes.dashboard'))
    
    cliente = Cliente.query.get_or_404(cliente_id)
    print(f"Antes: {cliente.ativo}")
    cliente.ativo = not cliente.ativo  
    print(f"Depois: {cliente.ativo}")
    

    db.session.commit()
    flash(f"Cliente {'ativado' if cliente.ativo else 'desativado'} com sucesso", "success")
    return redirect(url_for('routes.dashboard'))


@routes.route('/cadastrar_cliente', methods=['GET', 'POST'])
@login_required
def cadastrar_cliente():
    if session.get('user_type') != 'admin':  # Apenas admin pode cadastrar clientes
        abort(403)

    if request.method == 'POST':
        nome = request.form['nome']
        email = request.form['email']
        senha = request.form['senha']

        # Verifica se o e-mail já está cadastrado
        cliente_existente = Cliente.query.filter_by(email=email).first()
        if cliente_existente:
            flash("Já existe um cliente com esse e-mail!", "danger")
            return redirect(url_for('routes.cadastrar_cliente'))

        # Cria o cliente
        habilita_pagamento = True if request.form.get('habilita_pagamento') == 'on' else False
        novo_cliente = Cliente(
            nome=request.form['nome'],
            email=request.form['email'],
            senha=request.form['senha'],
            habilita_pagamento=habilita_pagamento
        )


        db.session.add(novo_cliente)
        db.session.commit()

        flash("Cliente cadastrado com sucesso!", "success")
        return redirect(url_for('routes.dashboard'))

    return render_template('cadastrar_cliente.html')


@routes.route('/oficinas', methods=['GET'])
@login_required
def listar_oficinas():
    if session.get('user_type') == 'participante':
        oficinas = Oficina.query.filter_by(cliente_id=current_user.cliente_id).all()  # ✅ Mostra apenas oficinas do Cliente que registrou o usuário
    else:
        oficinas = Oficina.query.all()

    return render_template('oficinas.html', oficinas=oficinas)

@routes.route('/editar_cliente/<int:cliente_id>', methods=['GET', 'POST'])
@login_required
def editar_cliente(cliente_id):
    if current_user.tipo != 'admin':
        flash('Acesso negado!', 'danger')
        return redirect(url_for('routes.dashboard'))

    cliente = Cliente.query.get_or_404(cliente_id)
    if request.method == 'POST':
        cliente.nome = request.form.get('nome')
        cliente.email = request.form.get('email')
        nova_senha = request.form.get('senha')
        if nova_senha:  # Só atualiza a senha se fornecida
            cliente.senha = generate_password_hash(nova_senha)
        
        # Debug: exibe o valor recebido do checkbox
        debug_checkbox = request.form.get('habilita_pagamento')
        print("DEBUG: Valor recebido do checkbox 'habilita_pagamento':", debug_checkbox)
        # Se você tiver um logger configurado, pode usar:
        # logger.debug("Valor recebido do checkbox 'habilita_pagamento': %s", debug_checkbox)
        
        cliente.habilita_pagamento = True if debug_checkbox == 'on' else False
        
        # Debug: exibe o valor que está sendo salvo
        print("DEBUG: Valor salvo em cliente.habilita_pagamento:", cliente.habilita_pagamento)
        # logger.debug("Valor salvo em cliente.habilita_pagamento: %s", cliente.habilita_pagamento)

        try:
            db.session.commit()
            flash("Cliente atualizado com sucesso!", "success")
        except Exception as e:
            db.session.rollback()
            print("DEBUG: Erro ao atualizar cliente:", e)
            # logger.error("Erro ao atualizar cliente: %s", e, exc_info=True)
            flash(f"Erro ao atualizar cliente: {str(e)}", "danger")
        return redirect(url_for('routes.dashboard'))
    
    return render_template('editar_cliente.html', cliente=cliente)


# imports extras no topo do arquivo -----------------------------------------
from sqlalchemy.orm import joinedload

# ----------------------------------------------------------------------------
@routes.route('/formularios', methods=['GET'])
@login_required
def listar_formularios():
    """
    •  Superusuário   → vê todos os formulários do sistema.
    •  Demais perfis  → vê apenas formulários ligados ao seu cliente.
    Cada formulário vem com TODAS as respostas + usuário de cada resposta
    para evitar consultas N+1.
    """
    # Base query com eager-load das relações necessárias
    q = (Formulario.query
            .options(
                joinedload(Formulario.respostas)
                    .joinedload(RespostaFormulario.usuario)   # carrega o autor da resposta
            )
            .order_by(Formulario.nome)
        )

    # 1. Se o usuário for superadmin, não filtra nada
    if getattr(current_user, "tipo", None) == "superadmin":
        formularios = q.all()

    # 2. Caso contrário, mostra apenas o que pertence ao cliente dele
    else:
        cliente_id = getattr(current_user, "cliente_id", None)

        # fallback: se o modelo antigo gravou cliente_id = current_user.id
        # usa o id do usuário como último recurso
        if not cliente_id:
            cliente_id = current_user.id

        formularios = q.filter(Formulario.cliente_id == cliente_id).all()

    # Envia tudo para o novo template que lista formulários + respostas
    return render_template('formularios.html',
                           formularios=formularios)


@routes.route('/formularios/novo', methods=['GET', 'POST'])
@login_required
def criar_formulario():
    if request.method == 'POST':
        nome = request.form.get('nome')
        descricao = request.form.get('descricao')
        
        novo_formulario = Formulario(
            nome=nome,
            descricao=descricao,
            cliente_id=current_user.id  # Relaciona com o cliente logado
        )
        db.session.add(novo_formulario)
        db.session.commit()
        flash('Formulário criado com sucesso!', 'success')
        return redirect(url_for('routes.listar_formularios'))
    
    return render_template('criar_formulario.html')

@routes.route('/formularios/<int:formulario_id>/editar', methods=['GET', 'POST'])
@login_required
def editar_formulario(formulario_id):
    formulario = Formulario.query.get_or_404(formulario_id)

    if request.method == 'POST':
        formulario.nome = request.form.get('nome')
        formulario.descricao = request.form.get('descricao')
        db.session.commit()
        flash('Formulário atualizado!', 'success')
        return redirect(url_for('routes.listar_formularios'))

    return render_template('editar_formulario.html', formulario=formulario)

@routes.route('/formularios/<int:formulario_id>/deletar', methods=['POST'])
@login_required
def deletar_formulario(formulario_id):
    formulario = Formulario.query.get_or_404(formulario_id)
    db.session.delete(formulario)
    db.session.commit()
    flash('Formulário deletado com sucesso!', 'success')
    return redirect(url_for('routes.listar_formularios'))

@routes.route('/formularios/<int:formulario_id>/campos', methods=['GET', 'POST'])
@login_required
def gerenciar_campos(formulario_id):
    formulario = Formulario.query.get_or_404(formulario_id)

    if request.method == 'POST':
        nome = request.form.get('nome')
        tipo = request.form.get('tipo')
        opcoes = request.form.get('opcoes', '').strip()
        obrigatorio = request.form.get('obrigatorio') == 'on'
        tamanho_max = request.form.get('tamanho_max') or None
        regex_validacao = request.form.get('regex_validacao') or None

        novo_campo = CampoFormulario(
            formulario_id=formulario.id,
            nome=nome,
            tipo=tipo,
            opcoes=opcoes if tipo in ['dropdown', 'checkbox', 'radio'] else None,
            obrigatorio=obrigatorio,
            tamanho_max=int(tamanho_max) if tamanho_max else None,
            regex_validacao=regex_validacao
        )

        db.session.add(novo_campo)
        db.session.commit()
        flash('Campo adicionado com sucesso!', 'success')

        return redirect(url_for('routes.gerenciar_campos', formulario_id=formulario.id))

    return render_template('gerenciar_campos.html', formulario=formulario)

@routes.route('/campos/<int:campo_id>/editar', methods=['GET', 'POST'])
@login_required
def editar_campo(campo_id):
    campo = CampoFormulario.query.get_or_404(campo_id)

    if request.method == 'POST':
        campo.nome = request.form.get('nome')
        campo.tipo = request.form.get('tipo')
        campo.opcoes = request.form.get('opcoes', '').strip() if campo.tipo in ['dropdown', 'checkbox', 'radio'] else None
        campo.obrigatorio = request.form.get('obrigatorio') == 'on'
        campo.tamanho_max = request.form.get('tamanho_max') or None
        campo.regex_validacao = request.form.get('regex_validacao') or None

        db.session.commit()
        flash('Campo atualizado com sucesso!', 'success')

        return redirect(url_for('routes.gerenciar_campos', formulario_id=campo.formulario_id))

    return render_template('editar_campo.html', campo=campo)

@routes.route('/campos/<int:campo_id>/deletar', methods=['POST'])
@login_required
def deletar_campo(campo_id):
    campo = CampoFormulario.query.get_or_404(campo_id)
    formulario_id = campo.formulario_id
    db.session.delete(campo)
    db.session.commit()
    flash('Campo removido com sucesso!', 'success')

    return redirect(url_for('routes.gerenciar_campos', formulario_id=formulario_id))

@routes.route('/formularios/<int:formulario_id>/preencher', methods=['GET', 'POST'])
@login_required
def preencher_formulario(formulario_id):
    formulario = Formulario.query.get_or_404(formulario_id)

    if request.method == 'POST':
        resposta_formulario = RespostaFormulario(
            formulario_id=formulario.id,
            usuario_id=current_user.id
        )
        db.session.add(resposta_formulario)
        db.session.commit()

        for campo in formulario.campos:
            valor = request.form.get(str(campo.id))
            if campo.tipo == 'file' and 'file_' + str(campo.id) in request.files:
                arquivo = request.files['file_' + str(campo.id)]
                if arquivo.filename:
                    filename = secure_filename(arquivo.filename)
                    caminho_arquivo = os.path.join('uploads', 'respostas', filename)
                    os.makedirs(os.path.dirname(caminho_arquivo), exist_ok=True)
                    arquivo.save(caminho_arquivo)
                    valor = caminho_arquivo  # Salva o caminho do arquivo

            resposta_campo = RespostaCampo(
                resposta_formulario_id=resposta_formulario.id,
                campo_id=campo.id,
                valor=valor
            )
            db.session.add(resposta_campo)

        db.session.commit()
        flash("Formulário enviado com sucesso!", "success")
        return redirect(url_for('routes.dashboard_participante'))

    return render_template('preencher_formulario.html', formulario=formulario)


@routes.route('/formularios_participante', methods=['GET'])
@login_required
def listar_formularios_participante():
    if current_user.tipo != 'participante':
        flash("Acesso negado!", "danger")
        return redirect(url_for('routes.dashboard'))

    # Busca apenas formulários disponíveis para o participante
    # Filtra formulários criados pelo mesmo cliente ao qual o participante está associado
    cliente_id = current_user.cliente_id
    
    if not cliente_id:
        flash("Você não está associado a nenhum cliente.", "warning")
        return redirect(url_for('routes.dashboard_participante'))
        
    # Busca formulários criados pelo cliente do participante
    formularios = Formulario.query.filter_by(cliente_id=cliente_id).all()
    
    # Não há relação direta entre formulários e ministrantes no modelo atual,
    # então estamos filtrando apenas pelo cliente_id do participante

    if not formularios:
        flash("Nenhum formulário disponível no momento.", "warning")
        return redirect(url_for('routes.dashboard_participante'))

    return render_template('formularios_participante.html', formularios=formularios)

@routes.route('/respostas/<int:resposta_id>', methods=['GET'])
@login_required
def visualizar_resposta(resposta_id):
    resposta = RespostaFormulario.query.get_or_404(resposta_id)

    # Se quiser, confira se o current_user é o dono da resposta
    if resposta.usuario_id != current_user.id:
        flash("Você não tem permissão para ver esta resposta.", "danger")
        return redirect(url_for('routes.dashboard_participante'))

    return render_template('visualizar_resposta.html', resposta=resposta)


from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

@routes.route('/formularios/<int:formulario_id>/exportar_pdf')
@login_required
def gerar_pdf_respostas(formulario_id):
    """
    Gera um PDF formatado e organizado das respostas de um formulário específico.
    
    Args:
        formulario_id: ID do formulário para buscar as respostas
        
    Returns:
        Um arquivo PDF para download
    """
    # Importações necessárias
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, 
        Spacer, Image, PageBreak
    )
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    import pytz
    import os
    from flask import send_file, current_app
    import time
    from datetime import datetime
    
    # Busca o formulário e as respostas
    formulario = Formulario.query.get_or_404(formulario_id)
    respostas = RespostaFormulario.query.filter_by(formulario_id=formulario.id).all()
    
    # Verifica se há respostas
    if not respostas:
        return None, "Não existem respostas para este formulário"

    # Define nome e caminho do arquivo PDF
    timestamp = int(time.time())
    pdf_filename = f"respostas_{formulario.id}_{timestamp}.pdf"
    pdf_folder = os.path.join(current_app.static_folder, "reports")
    os.makedirs(pdf_folder, exist_ok=True)
    pdf_path = os.path.join(pdf_folder, pdf_filename)
    
    # Configura o documento com margens adequadas
    doc = SimpleDocTemplate(
        pdf_path,
        pagesize=letter,
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36,
        title=f"Respostas - {formulario.nome}"
    )
    
    # Configuração de estilos customizados
    styles = getSampleStyleSheet()
    
    # Estilo para o título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles["Title"],
        fontSize=18,
        spaceAfter=20,
        textColor=colors.HexColor("#023E8A"),
        alignment=TA_CENTER
    )
    
    # Estilo para cabeçalhos
    header_style = ParagraphStyle(
        'HeaderStyle',
        parent=styles["Heading2"],
        fontSize=14,
        textColor=colors.white,
        alignment=TA_LEFT
    )
    
    # Estilo para o conteúdo
    content_style = ParagraphStyle(
        'ContentStyle',
        parent=styles["Normal"],
        fontSize=10,
        leading=14,
        spaceAfter=6
    )
    
    # Estilo para o rodapé
    footer_style = ParagraphStyle(
        'FooterStyle',
        parent=styles["Normal"],
        fontSize=8,
        textColor=colors.gray,
        alignment=TA_CENTER
    )

    # Lista para armazenar os elementos do PDF
    elements = []
    
    # Tenta adicionar um logo se existir
    logo_path = os.path.join(current_app.static_folder, "img", "logo.png")
    if os.path.exists(logo_path):
        # Configura o logo centralizado
        logo = Image(logo_path)
        logo.drawHeight = 0.8 * inch
        logo.drawWidth = 2 * inch
        elements.append(logo)
        elements.append(Spacer(1, 0.25 * inch))
    
    # Título do PDF
    title = Paragraph(f"Respostas do Formulário: {formulario.nome}", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.2 * inch))
    
    # Adiciona informações sobre o formulário
    if formulario.descricao:
        desc = Paragraph(f"<i>{formulario.descricao}</i>", content_style)
        elements.append(desc)
        elements.append(Spacer(1, 0.2 * inch))
    
    # Data de geração do relatório
    report_date = Paragraph(
        f"Relatório gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M:%S')}",
        content_style
    )
    elements.append(report_date)
    elements.append(Spacer(1, 0.3 * inch))
    
    # Função para converter datetime para o horário de Brasília
    def convert_to_brasilia(dt):
        brasilia_tz = pytz.timezone("America/Sao_Paulo")
        # Se o datetime não for "aware", assume-se que está em UTC
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=pytz.utc)
        return dt.astimezone(brasilia_tz)
    
    # Cria tabela
    data = []
    header = [
        Paragraph("<b>Participante</b>", header_style),
        Paragraph("<b>Data de Envio</b>", header_style),
        Paragraph("<b>Respostas</b>", header_style)
    ]
    data.append(header)
    
    # Preenche as linhas da tabela com cada resposta
    for resposta in respostas:
        # Informações do usuário
        usuario = resposta.usuario.nome if resposta.usuario else "N/A"
        
        # Conversão de data para horário local
        dt_local = convert_to_brasilia(resposta.data_submissao)
        data_envio = dt_local.strftime('%d/%m/%Y %H:%M')
        
        # Formatação do status, se disponível
        status_text = ""
        if hasattr(resposta, 'status_avaliacao') and resposta.status_avaliacao:
            status_color = {
                'Aprovada': '#28a745',
                'Aprovada com ressalvas': '#ffc107',
                'Negada': '#dc3545',
                'Não Avaliada': '#6c757d'
            }.get(resposta.status_avaliacao, '#6c757d')
            
            status_text = f"<br/><b>Status:</b> <font color='{status_color}'>{resposta.status_avaliacao}</font>"
        
        # Formatação das respostas com melhor estruturação
        resposta_text = f"<b>Respostas de {usuario}</b>{status_text}<br/><br/>"
        
        for campo in resposta.respostas_campos:
            valor = campo.valor if campo.valor else "N/A"
            
            # Se for caminho de arquivo, mostra apenas o nome do arquivo
            if campo.campo.tipo == 'file' and valor and '/' in valor:
                arquivo = valor.split('/')[-1]
                valor = f"<i>Arquivo: {arquivo}</i>"
                
            resposta_text += f"<b>{campo.campo.nome}:</b><br/>{valor}<br/><br/>"
        
        # Criação dos parágrafos para a tabela
        usuario_cell = Paragraph(f"<b>{usuario}</b>", content_style)
        data_cell = Paragraph(data_envio, content_style)
        resposta_cell = Paragraph(resposta_text, content_style)
        
        # Adiciona a linha à tabela
        data.append([usuario_cell, data_cell, resposta_cell])
    
    # Define a largura das colunas (distribuição percentual)
    available_width = doc.width
    col_widths = [
        available_width * 0.25,  # Nome (25%)
        available_width * 0.15,  # Data (15%)
        available_width * 0.60   # Respostas (60%)
    ]
    
    # Criação da tabela com os dados e larguras definidas
    table = Table(data, colWidths=col_widths, repeatRows=1)
    
    # Estilo da tabela
    table_style = TableStyle([
        # Cabeçalho
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#023E8A")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        
        # Bordas externas da tabela
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
        
        # Linhas horizontais mais finas
        ('LINEBELOW', (0, 0), (-1, -1), 0.5, colors.grey),
        
        # Alinhamento do texto
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        
        # Configurações de padding
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        
        # Zebra striping para facilitar a leitura
        ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
    ])
    
    table.setStyle(table_style)
    elements.append(table)
    
    # Adicionando rodapé
    elements.append(Spacer(1, 0.5 * inch))
    footer = Paragraph(
        f"© {datetime.now().year} - Documento gerado pelo sistema AppFiber - Página 1",
        footer_style
    )
    elements.append(footer)
    
    # Constrói o PDF
    doc.build(
        elements,
        onFirstPage=lambda canvas, doc: add_page_number(canvas, doc, 1),
        onLaterPages=lambda canvas, doc: add_page_number(canvas, doc)
    )
    
    # Retorna o arquivo para download
    return send_file(pdf_path, as_attachment=True)

def add_page_number(canvas, doc, page_num=None):
    """
    Adiciona o número de página ao rodapé.
    
    Args:
        canvas: O canvas do ReportLab
        doc: O documento
        page_num: Número específico de página (opcional)
    """
    page = page_num if page_num else canvas._pageNumber
    text = f"© {datetime.now().year} - Documento gerado pelo sistema AppFiber - Página {page}"
    
    # Define estilo e posição
    canvas.saveState()
    canvas.setFont("Helvetica", 8)
    canvas.setFillColor(colors.grey)
    
    # Posiciona na parte inferior central
    text_width = canvas.stringWidth(text, "Helvetica", 8)
    x = (doc.pagesize[0] - text_width) / 2
    canvas.drawString(x, 20, text)
    canvas.restoreState()
    
    
@routes.route('/formularios/<int:formulario_id>/exportar_csv')
@login_required
def exportar_csv(formulario_id):
    import csv
    import io
    import pytz
    from flask import Response, stream_with_context

    formulario = Formulario.query.get_or_404(formulario_id)
    respostas = RespostaFormulario.query.filter_by(formulario_id=formulario.id).all()

    csv_filename = f"respostas_{formulario.id}.csv"
    
    # Função para converter datetime para o fuso de Brasília
    def convert_to_brasilia(dt):
        brasilia_tz = pytz.timezone("America/Sao_Paulo")
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=pytz.utc)
        return dt.astimezone(brasilia_tz)

    # Função geradora que cria o CSV linha a linha
    def generate():
        output = io.StringIO()
        writer = csv.writer(output, delimiter=',')
        
        # Cabeçalho do CSV: Usuário, Data de Envio e os nomes dos campos do formulário
        header = ["Usuário", "Data de Envio"] + [campo.nome for campo in formulario.campos]
        writer.writerow(header)
        yield output.getvalue()
        output.seek(0)
        output.truncate(0)
        
        # Preenche as linhas com as respostas
        for resposta in respostas:
            usuario_nome = resposta.usuario.nome if resposta.usuario else "N/A"
            data_envio = convert_to_brasilia(resposta.data_submissao).strftime('%d/%m/%Y %H:%M')
            row = [usuario_nome, data_envio]
            for campo in formulario.campos:
                valor = next((resp.valor for resp in resposta.respostas_campos if resp.campo_id == campo.id), "")
                row.append(valor)
            writer.writerow(row)
            yield output.getvalue()
            output.seek(0)
            output.truncate(0)

    return Response(
        stream_with_context(generate()),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename={csv_filename}"}
    )


@routes.route('/respostas/<path:filename>')
@login_required
def get_resposta_file(filename):
    print(">> get_resposta_file foi chamado com:", filename)
    uploads_folder = os.path.join('uploads', 'respostas')
    return send_from_directory(uploads_folder, filename)

from sqlalchemy import text  # Adicione esta importação no topo do arquivo!

from sqlalchemy import text

@routes.route('/formularios/<int:formulario_id>/excluir', methods=['POST'])
@login_required
def excluir_formulario(formulario_id):
    formulario = Formulario.query.get_or_404(formulario_id)

    try:
        # 1️⃣ Exclui FeedbackCampo associados às respostas do formulário (SQL textual corrigido)
        db.session.execute(text('''
            DELETE FROM feedback_campo
            WHERE resposta_campo_id IN (
                SELECT id FROM respostas_campo
                WHERE resposta_formulario_id IN (
                    SELECT id FROM respostas_formulario
                    WHERE formulario_id = :fid
                )
            );
        '''), {'fid': formulario_id})

        # 2️⃣ Exclui RespostaCampo
        RespostaCampo.query.filter(
            RespostaCampo.resposta_formulario_id.in_(
                db.session.query(RespostaFormulario.id).filter_by(formulario_id=formulario_id)
            )
        ).delete(synchronize_session=False)

        # 3️⃣ Exclui RespostaFormulario
        RespostaFormulario.query.filter_by(formulario_id=formulario_id).delete()

        # 4️⃣ Exclui o Formulário
        formulario = Formulario.query.get_or_404(formulario_id)
        db.session.delete(formulario)

        db.session.commit()

        flash("Formulário e todos os dados relacionados excluídos com sucesso!", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Erro ao excluir formulário: {str(e)}", "danger")

    return redirect(url_for('routes.listar_formularios'))

@routes.route('/upload_personalizacao_certificado', methods=['GET', 'POST'])
@login_required
def upload_personalizacao_certificado():
    
    cliente = Cliente.query.get(current_user.id)
    templates = CertificadoTemplate.query.filter_by(cliente_id=current_user.id).all()

    if request.method == 'POST':
        logo_file = request.files.get('logo_certificado')
        fundo_file = request.files.get('fundo_certificado')
        ass_file = request.files.get('assinatura_certificado')

        # Exemplo de pasta
        pasta_uploads = os.path.join('uploads', 'personalizacao')
        os.makedirs(pasta_uploads, exist_ok=True)

        # Se o cliente enviar algo, salvamos e atualizamos o path
        if logo_file and logo_file.filename:
            filename_logo = secure_filename(logo_file.filename)
            caminho_logo = os.path.join(pasta_uploads, filename_logo)
            logo_file.save(caminho_logo)
            current_user.logo_certificado = caminho_logo  # Salva no banco

        if fundo_file and fundo_file.filename:
            filename_fundo = secure_filename(fundo_file.filename)
            caminho_fundo = os.path.join(pasta_uploads, filename_fundo)
            fundo_file.save(caminho_fundo)
            current_user.fundo_certificado = caminho_fundo

        if ass_file and ass_file.filename:
            filename_ass = secure_filename(ass_file.filename)
            caminho_ass = os.path.join(pasta_uploads, filename_ass)
            ass_file.save(caminho_ass)
            current_user.assinatura_certificado = caminho_ass

        db.session.commit()
        flash("Personalização salva com sucesso!", "success")
        return redirect(url_for('routes.dashboard_cliente'))

    return render_template('upload_personalizacao_cert.html', templates=templates, cliente=cliente)



@routes.route("/api/configuracao_cliente_atual", methods=["GET"])
@login_required
def configuracao_cliente_atual():
    """Retorna o estado atual das configurações do cliente logado em JSON."""
    cliente_id = current_user.id
    config_cliente = ConfiguracaoCliente.query.filter_by(cliente_id=cliente_id).first()
    if not config_cliente:
        config_cliente = ConfiguracaoCliente(
            cliente_id=cliente_id,
            permitir_checkin_global=False,
            habilitar_feedback=False,
            habilitar_certificado_individual=False
        )
        db.session.add(config_cliente)
        db.session.commit()

    return jsonify({
        "success": True,
        "permitir_checkin_global": config_cliente.permitir_checkin_global,
        "habilitar_feedback": config_cliente.habilitar_feedback,
        "habilitar_certificado_individual": config_cliente.habilitar_certificado_individual
    })
    
@routes.route('/gerar_etiquetas/<int:cliente_id>', methods=['GET'])
@login_required
def gerar_etiquetas(cliente_id):
    """Gera um PDF de etiquetas para o cliente"""
    if current_user.tipo != 'cliente' or current_user.id != cliente_id:
        flash("Acesso negado!", "danger")
        return redirect(url_for('routes.dashboard_cliente'))

    pdf_path = gerar_etiquetas_pdf(cliente_id)
    if not pdf_path:
        flash("Nenhum usuário encontrado para gerar etiquetas!", "warning")
        return redirect(url_for('routes.dashboard_cliente'))

    return send_file(pdf_path, as_attachment=True)

@routes.route('/respostas', methods=['GET'])
@login_required
def listar_respostas():
    # Verifica se o usuário é cliente ou ministrante
    if current_user.tipo not in ['cliente', 'ministrante']:
        flash('Acesso negado!', 'danger')
        return redirect(url_for('routes.dashboard'))

    # --- Se for cliente ---
    if current_user.tipo == 'cliente':
        respostas = (
            RespostaFormulario.query
            .join(Usuario)
            .join(Formulario)
            .filter(
                Usuario.cliente_id == current_user.id,
                Formulario.cliente_id == current_user.id
            )
            .order_by(RespostaFormulario.data_submissao.desc())
            .all()
        )

    # --- Se for ministrante ---
    elif current_user.tipo == 'ministrante':
        # Você pode adaptar essa parte para buscar participantes das oficinas que ele ministra
        respostas = (
            RespostaFormulario.query
            .join(Usuario)
            .join(Inscricao, Inscricao.usuario_id == Usuario.id)
            .join(Oficina, Inscricao.oficina_id == Oficina.id)
            .filter(Oficina.ministrante_id == current_user.id)
            .order_by(RespostaFormulario.data_submissao.desc())
            .all()
        )

    # Se não houver respostas
    if not respostas:
        flash('Não há respostas disponíveis no momento.', 'info')
        return redirect(url_for('routes.dashboard'))

    formulario = respostas[0].formulario

    return render_template(
        'listar_respostas.html',
        formulario=formulario,
        respostas=respostas
    )


@routes.route('/formularios/<int:formulario_id>/respostas_ministrante', methods=['GET'])
@login_required
def listar_respostas_ministrante(formulario_id):
    # 1) Verifica se o current_user é ministrante
    if not isinstance(current_user, Ministrante):
        flash('Apenas ministrantes têm acesso a esta tela.', 'danger')
        return redirect(url_for('routes.dashboard_ministrante'))

    formulario = Formulario.query.get_or_404(formulario_id)
    # 2) Carrega as respostas
    respostas = RespostaFormulario.query.filter_by(formulario_id=formulario.id).all()

    return render_template(
        'listar_respostas_ministrante.html',
        formulario=formulario,
        respostas=respostas
    )

@routes.route(
    '/respostas/<int:resposta_id>/feedback', 
    methods=['GET', 'POST'], 
    endpoint='dar_feedback_resposta_formulario'
)
@login_required
def dar_feedback_resposta(resposta_id):
    # só Ministrantes ou Clientes
    if not (isinstance(current_user, Ministrante) or current_user.tipo == 'cliente'):
        flash('Apenas clientes e ministrantes podem dar feedback.', 'danger')
        return redirect(url_for('routes.dashboard'))

    resposta = RespostaFormulario.query.get_or_404(resposta_id)
    lista_campos = resposta.formulario.campos
    resposta_campos = resposta.respostas_campos

    if request.method == 'POST':
        for rc in resposta_campos:
            nome_textarea = f"feedback_{rc.id}"
            texto = request.form.get(nome_textarea, "").strip()
            if not texto:
                continue

            fb = FeedbackCampo(
                resposta_campo_id=rc.id,
                ministrante_id=current_user.id if isinstance(current_user, Ministrante) else None,
                cliente_id    =current_user.id if current_user.tipo == 'cliente' else None,
                texto_feedback=texto
            )
            db.session.add(fb)

        db.session.commit()
        flash("Feedback registrado com sucesso!", "success")
        return redirect(url_for('routes.dar_feedback_resposta_formulario', resposta_id=resposta_id))

    return render_template(
        'dar_feedback_resposta.html',
        resposta=resposta,
        lista_campos=lista_campos,
        resposta_campos=resposta_campos
    )


@routes.route('/definir_status_inline', methods=['POST'])
@login_required
def definir_status_inline():
    # 1) Pega valores do form
    resposta_id = request.form.get('resposta_id')
    novo_status = request.form.get('status_avaliacao')

    # 2) Valida
    if not resposta_id or not novo_status:
        flash("Dados incompletos!", "danger")
        return redirect(request.referrer or url_for('routes.dashboard'))

    # 3) Busca a resposta no banco
    resposta = RespostaFormulario.query.get(resposta_id)
    if not resposta:
        flash("Resposta não encontrada!", "warning")
        return redirect(request.referrer or url_for('routes.dashboard'))

    # 4) Atualiza
    resposta.status_avaliacao = novo_status
    db.session.commit()

    flash("Status atualizado com sucesso!", "success")

    # Redireciona para a mesma página (listar_respostas) ou usa request.referrer
    # Se estiver em /formularios/<id>/respostas_ministrante, podemos redirecionar
    return redirect(request.referrer or url_for('routes.listar_respostas',
                                                formulario_id=resposta.formulario_id))



@routes.route('/inscrever_participantes_lote', methods=['POST'])
@login_required
def inscrever_participantes_lote():
    print("📌 [DEBUG] Iniciando processo de inscrição em lote...")

    oficina_id = request.form.get('oficina_id')
    usuario_ids = request.form.getlist('usuario_ids')

    print(f"📌 [DEBUG] Oficina selecionada: {oficina_id}")
    print(f"📌 [DEBUG] Usuários selecionados: {usuario_ids}")

    if not oficina_id or not usuario_ids:
        flash('Oficina ou participantes não selecionados corretamente.', 'warning')
        print("❌ [DEBUG] Erro: Oficina ou participantes não foram selecionados corretamente.")
        return redirect(url_for('routes.dashboard'))

    oficina = Oficina.query.get(oficina_id)
    if not oficina:
        flash('Oficina não encontrada!', 'danger')
        print("❌ [DEBUG] Erro: Oficina não encontrada no banco de dados.")
        return redirect(url_for('routes.dashboard'))

    inscritos_sucesso = 0
    erros = 0

    try:
        for usuario_id in usuario_ids:
            print(f"🔄 [DEBUG] Tentando inscrever usuário {usuario_id} na oficina {oficina.titulo}...")

            ja_inscrito = Inscricao.query.filter_by(usuario_id=usuario_id, oficina_id=oficina_id).first()

            if ja_inscrito:
                print(f"⚠️ [DEBUG] Usuário {usuario_id} já está inscrito na oficina. Pulando...")
                continue  # Evita duplicação

            # Verifica se há vagas disponíveis
            if oficina.vagas <= 0:
                print(f"❌ [DEBUG] Sem vagas para a oficina {oficina.titulo}. Usuário {usuario_id} não pode ser inscrito.")
                erros += 1
                continue

            # 🔥 SOLUÇÃO: Passando cliente_id corretamente para a Inscricao
            nova_inscricao = Inscricao(
                usuario_id=usuario_id,
                oficina_id=oficina_id,
                cliente_id=oficina.cliente_id  # Obtém o cliente_id da própria oficina
            )

            db.session.add(nova_inscricao)
            oficina.vagas -= 1  # Reduz a quantidade de vagas disponíveis

            inscritos_sucesso += 1
            print(f"✅ [DEBUG] Usuário {usuario_id} inscrito com sucesso!")

        db.session.commit()
        flash(f'{inscritos_sucesso} participantes inscritos com sucesso! {erros} não foram inscritos por falta de vagas.', 'success')
        print(f"🎯 [DEBUG] {inscritos_sucesso} inscrições concluídas. {erros} falharam.")

    except Exception as e:
        db.session.rollback()
        flash(f"Erro ao inscrever participantes em lote: {str(e)}", "danger")
        print(f"❌ [DEBUG] Erro ao inscrever participantes: {e}")

    return redirect(url_for('routes.dashboard'))

@routes.route('/configurar_evento', methods=['GET', 'POST'])
@login_required
def configurar_evento():
    if current_user.tipo != 'cliente':
        flash('Acesso negado!', 'danger')
        return redirect(url_for('routes.dashboard_cliente'))

    # Lista todos os eventos do cliente
    eventos = Evento.query.filter_by(cliente_id=current_user.id).all()
    
    # Evento selecionado (por padrão, None até que o usuário escolha)
    evento_id = request.args.get('evento_id') or (request.form.get('evento_id') if request.method == 'POST' else None)
    evento = None
    if evento_id:
        # Carregamento eager de todos os relacionamentos necessários
        evento = Evento.query.options(
            db.joinedload(Evento.tipos_inscricao),
            db.joinedload(Evento.lotes).joinedload(LoteInscricao.tipos_inscricao)
        ).filter_by(id=evento_id, cliente_id=current_user.id).first()

    if request.method == 'POST':
        nome = request.form.get('nome')
        descricao = request.form.get('descricao')
        programacao = request.form.get('programacao')
        localizacao = request.form.get('localizacao')
        link_mapa = request.form.get('link_mapa')
        inscricao_gratuita = request.form.get('inscricao_gratuita') == 'on'  # Checkbox retorna 'on' se marcado
        habilitar_lotes = request.form.get('habilitar_lotes') == 'on'  # Novo campo
        nomes_tipos = request.form.getlist('nome_tipo[]')  # Lista de nomes dos tipos
        precos_tipos = request.form.getlist('preco_tipo[]')  # Lista de preços dos tipos
        
        banner = request.files.get('banner')
        banner_url = evento.banner_url if evento else None
        
        if banner:
            filename = secure_filename(banner.filename)
            caminho_banner = os.path.join('static/banners', filename)
            os.makedirs(os.path.dirname(caminho_banner), exist_ok=True)
            banner.save(caminho_banner)
            banner_url = url_for('static', filename=f'banners/{filename}', _external=True)

        try:
            if evento:  # Atualizar evento existente
                evento.nome = nome
                evento.descricao = descricao
                evento.programacao = programacao
                evento.localizacao = localizacao
                evento.link_mapa = link_mapa
                evento.inscricao_gratuita = inscricao_gratuita
                evento.habilitar_lotes = habilitar_lotes  # Novo campo
                if banner_url:
                    evento.banner_url = banner_url

                # Remover regras de inscrição associadas para evitar violação de chave estrangeira
                RegraInscricaoEvento.query.filter_by(evento_id=evento.id).delete()
                
                # Verifica se existem usuários vinculados aos tipos de inscrição deste evento
                tipos_com_usuarios = db.session.query(EventoInscricaoTipo.id).join(
                    Usuario, Usuario.tipo_inscricao_id == EventoInscricaoTipo.id
                ).filter(
                    EventoInscricaoTipo.evento_id == evento.id
                ).all()
                
                # Lista de IDs de tipos que têm usuários vinculados
                ids_tipos_com_usuarios = [tipo[0] for tipo in tipos_com_usuarios]
                
                # Verifica quais tipos têm referências na tabela lote_tipo_inscricao
                tipos_com_lotes = db.session.query(LoteTipoInscricao.tipo_inscricao_id).join(
                    EventoInscricaoTipo, LoteTipoInscricao.tipo_inscricao_id == EventoInscricaoTipo.id
                ).filter(
                    EventoInscricaoTipo.evento_id == evento.id
                ).distinct().all()
                
                # Lista de IDs de tipos que têm lotes vinculados
                ids_tipos_com_lotes = [tipo[0] for tipo in tipos_com_lotes]
                
                # Primeiro, remover as referências de lote_tipo_inscricao para os tipos que serão removidos
                ids_tipos_para_preservar = list(set(ids_tipos_com_usuarios))
                
                # Lista de IDs que foram enviados pelo formulário
                ids_tipos_enviados = []
                for i, nome_tipo in enumerate(nomes_tipos):
                    if nome_tipo and i < len(request.form.getlist('id_tipo[]')):
                        tipo_id = request.form.getlist('id_tipo[]')[i]
                        if tipo_id and tipo_id.isdigit():
                            ids_tipos_enviados.append(int(tipo_id))
                
                # Adicionar os tipos enviados pelo formulário à lista de preservação
                ids_tipos_para_preservar.extend([tid for tid in ids_tipos_enviados if tid not in ids_tipos_para_preservar])
                
                # Remover referências em lote_tipo_inscricao para tipos que serão excluídos
                for tipo_id in ids_tipos_com_lotes:
                    if tipo_id not in ids_tipos_para_preservar:
                        LoteTipoInscricao.query.filter_by(tipo_inscricao_id=tipo_id).delete()
                
                # Agora podemos excluir os tipos de inscrição com segurança
                if ids_tipos_para_preservar:
                    # Exclui apenas os tipos que NÃO estão na lista de preservação
                    EventoInscricaoTipo.query.filter(
                        EventoInscricaoTipo.evento_id == evento.id,
                        ~EventoInscricaoTipo.id.in_(ids_tipos_para_preservar)
                    ).delete(synchronize_session=False)
                else:
                    # Se não houver tipos para preservar, primeiro remova todas as referências
                    LoteTipoInscricao.query.filter(
                        LoteTipoInscricao.tipo_inscricao_id.in_(
                            db.session.query(EventoInscricaoTipo.id).filter_by(evento_id=evento.id)
                        )
                    ).delete(synchronize_session=False)
                    # Depois exclua todos os tipos
                    EventoInscricaoTipo.query.filter_by(evento_id=evento.id).delete()
                
                # Adicionar novos tipos ou atualizar existentes
                tipos_inscricao = []
                for nome_tipo, preco_tipo in zip(nomes_tipos, precos_tipos):
                    if nome_tipo:  # Só adicionar se o nome for preenchido
                        # Se for inscrição gratuita, definir preço como 0.00
                        preco_efetivo = 0.0 if inscricao_gratuita else float(preco_tipo)
                        
                        # Verificar se já existe um tipo com este nome
                        tipo_existente = None
                        for tipo_id in ids_tipos_para_preservar:
                            tipo = EventoInscricaoTipo.query.get(tipo_id)
                            if tipo and tipo.nome == nome_tipo:
                                tipo_existente = tipo
                                break
                        
                        if tipo_existente:
                            # Atualiza o preço do tipo existente
                            tipo_existente.preco = preco_efetivo
                            tipos_inscricao.append(tipo_existente)
                        else:
                            # Cria um novo tipo
                            tipo = EventoInscricaoTipo(
                                evento_id=evento.id,
                                nome=nome_tipo,
                                preco=preco_efetivo
                            )
                            db.session.add(tipo)
                            db.session.flush()  # Para obter o ID do tipo
                            tipos_inscricao.append(tipo)
                
                # Processar os lotes somente se habilitar_lotes for True
                if habilitar_lotes:
                    lote_ids = request.form.getlist('lote_id[]')
                    lote_nomes = request.form.getlist('lote_nome[]')
                    lote_ordens = request.form.getlist('lote_ordem[]')
                    lote_ativo = [val == '1' for val in request.form.getlist('lote_ativo[]')]
                    lote_usar_data = [val == 'on' for val in request.form.getlist('lote_usar_data[]')]
                    lote_data_inicio = request.form.getlist('lote_data_inicio[]')
                    lote_data_fim = request.form.getlist('lote_data_fim[]')
                    lote_usar_qtd = [val == 'on' for val in request.form.getlist('lote_usar_qtd[]')]
                    lote_qtd_maxima = request.form.getlist('lote_qtd_maxima[]')
                    
                    # Verificar quais lotes possuem inscrições
                    lotes_com_inscricoes = db.session.query(LoteInscricao.id).join(
                        Inscricao, Inscricao.lote_id == LoteInscricao.id
                    ).filter(
                        LoteInscricao.evento_id == evento.id
                    ).all()
                    
                    # Lista de IDs de lotes com inscrições vinculadas
                    ids_lotes_com_inscricoes = [lote_id[0] for lote_id in lotes_com_inscricoes]
                    
                    # IDs de lotes que devem ser preservados (informados pelo cliente via form)
                    preservar_ids_lote = request.form.get('preservar_ids_lote', '').split(',')
                    preservar_ids_lote = [int(id) for id in preservar_ids_lote if id and id.isdigit()]
                    
                    # Lotes a serem removidos
                    lotes_para_remover = LoteInscricao.query.filter(
                        LoteInscricao.evento_id == evento.id,
                        ~LoteInscricao.id.in_(preservar_ids_lote),
                        ~LoteInscricao.id.in_(ids_lotes_com_inscricoes)
                    ).all()
                    
                    # Remover os registros de lote_tipo_inscricao antes de remover os lotes
                    for lote in lotes_para_remover:
                        LoteTipoInscricao.query.filter_by(lote_id=lote.id).delete()
                    
                    # Agora é seguro remover os lotes
                    for lote in lotes_para_remover:
                        db.session.delete(lote)
                    
                    # Processar cada lote
                    for i, nome in enumerate(lote_nomes):
                        if nome:
                            # Determinar configurações do lote
                            data_inicio_lote = None
                            data_fim_lote = None
                            qtd_maxima = None
                            ativo = True if i < len(lote_ativo) and lote_ativo[i] else False
                            
                            if i < len(lote_usar_data) and lote_usar_data[i]:
                                if i < len(lote_data_inicio) and lote_data_inicio[i]:
                                    data_inicio_lote = datetime.strptime(lote_data_inicio[i], '%Y-%m-%d')
                                if i < len(lote_data_fim) and lote_data_fim[i]:
                                    data_fim_lote = datetime.strptime(lote_data_fim[i], '%Y-%m-%d')
                            
                            if i < len(lote_usar_qtd) and lote_usar_qtd[i]:
                                if i < len(lote_qtd_maxima) and lote_qtd_maxima[i]:
                                    qtd_maxima = int(lote_qtd_maxima[i])
                            
                            # Verificar se é um lote existente ou novo
                            lote_id = lote_ids[i] if i < len(lote_ids) and lote_ids[i] else None
                            
                            if lote_id:
                                # Atualizar lote existente
                                lote = LoteInscricao.query.get(int(lote_id))
                                if lote:
                                    lote.nome = nome
                                    lote.data_inicio = data_inicio_lote
                                    lote.data_fim = data_fim_lote
                                    lote.qtd_maxima = qtd_maxima
                                    lote.ordem = int(lote_ordens[i]) if i < len(lote_ordens) else i+1
                                    lote.ativo = ativo
                            else:
                                # Criar novo lote
                                lote = LoteInscricao(
                                    evento_id=evento.id,
                                    nome=nome,
                                    data_inicio=data_inicio_lote,
                                    data_fim=data_fim_lote,
                                    qtd_maxima=qtd_maxima,
                                    ordem=int(lote_ordens[i]) if i < len(lote_ordens) else i+1,
                                    ativo=ativo
                                )
                                db.session.add(lote)
                                db.session.flush()  # Para obter o ID
                            
                            # Processar preços dos tipos de inscrição para este lote
                            for tipo in tipos_inscricao:
                                # O formato do nome do campo: lote_tipo_preco_[lote_id]_[tipo_id]
                                preco_key = f'lote_tipo_preco_{lote.id}_{tipo.id}'
                                preco_valor = request.form.get(preco_key)
                                
                                if preco_valor is not None:
                                    # Se for gratuito, todos os preços são 0
                                    preco_final = 0.0 if inscricao_gratuita else float(preco_valor)
                                    
                                    # Verificar se já existe um registro de preço para este lote e tipo
                                    lote_tipo = LoteTipoInscricao.query.filter_by(
                                        lote_id=lote.id, 
                                        tipo_inscricao_id=tipo.id
                                    ).first()
                                    
                                    if lote_tipo:
                                        # Atualizar preço existente
                                        lote_tipo.preco = preco_final
                                    else:
                                        # Criar novo registro de preço
                                        novo_lote_tipo = LoteTipoInscricao(
                                            lote_id=lote.id,
                                            tipo_inscricao_id=tipo.id,
                                            preco=preco_final
                                        )
                                        db.session.add(novo_lote_tipo)
                
            else:  # Criar novo evento
                evento = Evento(
                    cliente_id=current_user.id,
                    nome=nome,
                    descricao=descricao,
                    programacao=programacao,
                    localizacao=localizacao,
                    link_mapa=link_mapa,
                    banner_url=banner_url,
                    inscricao_gratuita=inscricao_gratuita,
                    habilitar_lotes=habilitar_lotes  # Novo campo
                )
                db.session.add(evento)
                db.session.flush()  # Gera o ID do evento antes de adicionar os tipos

                # Adicionar tipos de inscrição
                tipos_inscricao = []
                for nome_tipo, preco_tipo in zip(nomes_tipos, precos_tipos):
                    if nome_tipo:  # Só adicionar se o nome for preenchido
                        # Se for inscrição gratuita, definir preço como 0.00
                        preco_efetivo = 0.0 if inscricao_gratuita else float(preco_tipo)
                        
                        tipo = EventoInscricaoTipo(
                            evento_id=evento.id,
                            nome=nome_tipo,
                            preco=preco_efetivo
                        )
                        db.session.add(tipo)
                        db.session.flush()  # Para obter o ID
                        tipos_inscricao.append(tipo)
                
                # Adicionar lotes de inscrição somente se habilitar_lotes for True
                if habilitar_lotes:
                    lote_nomes = request.form.getlist('lote_nome[]')
                    lote_ordens = request.form.getlist('lote_ordem[]')
                    lote_ativo = [val == '1' for val in request.form.getlist('lote_ativo[]')]
                    lote_usar_data = [val == 'on' for val in request.form.getlist('lote_usar_data[]')]
                    lote_data_inicio = request.form.getlist('lote_data_inicio[]')
                    lote_data_fim = request.form.getlist('lote_data_fim[]')
                    lote_usar_qtd = [val == 'on' for val in request.form.getlist('lote_usar_qtd[]')]
                    lote_qtd_maxima = request.form.getlist('lote_qtd_maxima[]')
                    
                    # Criar cada lote
                    for i, nome in enumerate(lote_nomes):
                        if nome:
                            # Determinar configurações do lote
                            data_inicio_lote = None
                            data_fim_lote = None
                            qtd_maxima = None
                            ativo = True if i < len(lote_ativo) and lote_ativo[i] else False
                            
                            if i < len(lote_usar_data) and lote_usar_data[i]:
                                if i < len(lote_data_inicio) and lote_data_inicio[i]:
                                    data_inicio_lote = datetime.strptime(lote_data_inicio[i], '%Y-%m-%d')
                                if i < len(lote_data_fim) and lote_data_fim[i]:
                                    data_fim_lote = datetime.strptime(lote_data_fim[i], '%Y-%m-%d')
                            
                            if i < len(lote_usar_qtd) and lote_usar_qtd[i]:
                                if i < len(lote_qtd_maxima) and lote_qtd_maxima[i]:
                                    qtd_maxima = int(lote_qtd_maxima[i])
                            
                            # Criar o lote
                            lote = LoteInscricao(
                                evento_id=evento.id,
                                nome=nome,
                                data_inicio=data_inicio_lote,
                                data_fim=data_fim_lote,
                                qtd_maxima=qtd_maxima,
                                ordem=int(lote_ordens[i]) if i < len(lote_ordens) else i+1,
                                ativo=ativo
                            )
                            db.session.add(lote)
                            db.session.flush()  # Para obter o ID do lote
                            
                            # Processar preços por tipo de inscrição
                            for j, tipo in enumerate(tipos_inscricao):
                                # Para novos lotes e tipos, o formato é diferente
                                preco_key = f'lote_tipo_preco_new_{i}_{tipo.id}'
                                # Verificar também o formato alternativo para compatibilidade
                                preco_valor = request.form.get(preco_key) or request.form.get(f'lote_tipo_preco_new_{i}_new_{j}')
                                
                                if preco_valor is not None:
                                    # Se for gratuito, todos os preços são 0
                                    preco_final = 0.0 if inscricao_gratuita else float(preco_valor)
                                    
                                    novo_lote_tipo = LoteTipoInscricao(
                                        lote_id=lote.id,
                                        tipo_inscricao_id=tipo.id,
                                        preco=preco_final
                                    )
                                    db.session.add(novo_lote_tipo)

            db.session.commit()
            
            # Recarregar o evento com todos os relacionamentos para exibição
            if evento:
                evento = Evento.query.options(
                    db.joinedload(Evento.tipos_inscricao),
                    db.joinedload(Evento.lotes).joinedload(LoteInscricao.tipos_inscricao)
                ).filter_by(id=evento.id, cliente_id=current_user.id).first()
                
            flash('Evento salvo com sucesso!', 'success')
            return redirect(url_for('routes.dashboard_cliente'))
        
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao salvar evento: {str(e)}', 'danger')
            # Adicionar log para debugging
            print(f"Erro ao salvar evento: {str(e)}")
            import traceback
            traceback.print_exc()

    return render_template('configurar_evento.html', eventos=eventos, evento=evento)


from collections import defaultdict
from datetime import datetime

@routes.route('/exibir_evento/<int:evento_id>')
@login_required
def exibir_evento(evento_id):
    # 1) Carrega o evento
    evento = Evento.query.get_or_404(evento_id)

    # 2) Carrega as oficinas do cliente vinculado ao evento
    #    (Aqui assumimos que evento.cliente_id é o mesmo que Oficina.cliente_id)
    oficinas = Oficina.query.filter_by(cliente_id=evento.cliente_id).all()

    # 3) Monta uma estrutura para agrupar por data
    #    grouped_oficinas[ "DD/MM/AAAA" ] = [ { 'titulo': ..., 'ministrante': ..., 'inicio': ..., 'fim': ... }, ... ]
    grouped_oficinas = defaultdict(list)

   # No trecho onde você monta grouped_oficinas
    for oficina in oficinas:
        for dia in oficina.dias:
            data_str = dia.data.strftime('%d/%m/%Y')
            temp_group[data_str].append({
                'titulo': oficina.titulo,
                'descricao': oficina.descricao,
                'ministrante': oficina.ministrante_obj,  # Objeto ministrante completo em vez de só o nome
                'horario_inicio': dia.horario_inicio,
                'horario_fim': dia.horario_fim
            })

    # Ordena as datas no dicionário pela data real (opcional)
    # Precisamos converter a string "DD/MM/AAAA" para datetime para ordenar:
    sorted_keys = sorted(
        grouped_oficinas.keys(), 
        key=lambda d: datetime.strptime(d, '%d/%m/%Y')
    )

    # 4) Renderiza o template passando o evento e a programação agrupada
    return render_template(
        'exibir_evento.html',
        evento=evento,
        sorted_keys=sorted_keys,
        grouped_oficinas=grouped_oficinas
    )

@routes.route('/criar_evento', methods=['GET', 'POST'])
@login_required
def criar_evento():
    if current_user.tipo != 'cliente':
        flash('Acesso negado!', 'danger')
        return redirect(url_for('routes.dashboard_cliente'))
    
    # Para evitar o erro 'evento is undefined' no template
    evento = None

    if request.method == 'POST':
        nome = request.form.get('nome')
        descricao = request.form.get('descricao')
        programacao = request.form.get('programacao')
        localizacao = request.form.get('localizacao')
        link_mapa = request.form.get('link_mapa')

        banner = request.files.get('banner')
        banner_url = None
        
        if banner:
            filename = secure_filename(banner.filename)
            caminho_banner = os.path.join('static/banners', filename)
            os.makedirs(os.path.dirname(caminho_banner), exist_ok=True)
            banner.save(caminho_banner)
            banner_url = url_for('static', filename=f'banners/{filename}', _external=False)
        
        # Processar campos de data
        data_inicio_str = request.form.get('data_inicio')
        data_fim_str = request.form.get('data_fim')
        hora_inicio_str = request.form.get('hora_inicio')
        hora_fim_str = request.form.get('hora_fim')
        
        data_inicio = datetime.strptime(data_inicio_str, '%Y-%m-%d') if data_inicio_str else None
        data_fim = datetime.strptime(data_fim_str, '%Y-%m-%d') if data_fim_str else None
        
        from datetime import time
        hora_inicio = time.fromisoformat(hora_inicio_str) if hora_inicio_str else None
        hora_fim = time.fromisoformat(hora_fim_str) if hora_fim_str else None
        
        # Verificar se é gratuito
        inscricao_gratuita = (request.form.get('inscricao_gratuita') == 'on')
        
        # Verificar se habilita lotes
        habilitar_lotes = (request.form.get('habilitar_lotes') == 'on')
    
        # Cria o objeto Evento
        novo_evento = Evento(
            cliente_id=current_user.id,
            nome=nome,
            descricao=descricao,
            programacao=programacao,
            localizacao=localizacao,
            link_mapa=link_mapa,
            banner_url=banner_url,
            data_inicio=data_inicio,
            data_fim=data_fim,
            hora_inicio=hora_inicio,
            hora_fim=hora_fim,
            inscricao_gratuita=inscricao_gratuita,  # Salvar a flag no evento
            habilitar_lotes=habilitar_lotes  # Nova flag para habilitar lotes
        )

        try:
            db.session.add(novo_evento)
            db.session.flush()  # precisamos do ID para criar tipos de inscrição

            # Se o cliente tiver pagamento habilitado, tratar tipos de inscrição
            if current_user.habilita_pagamento:
                nomes_tipos = request.form.getlist('nome_tipo[]')
                precos = request.form.getlist('preco_tipo[]')
                
                # Verificar se os tipos de inscrição foram fornecidos
                if not nomes_tipos:
                    raise ValueError("É necessário definir pelo menos um tipo de inscrição.")
                
                # Para eventos gratuitos, definir todos os preços como 0.00
                if inscricao_gratuita:
                    precos = ['0.00'] * len(nomes_tipos)
                
                # Criar tipos de inscrição para o evento
                tipos_inscricao = []
                for i, nome in enumerate(nomes_tipos):
                    if nome.strip():  # Só criar se o nome não estiver vazio
                        preco = 0.0 if inscricao_gratuita else float(precos[i])
                        novo_tipo = EventoInscricaoTipo(
                            evento_id=novo_evento.id,
                            nome=nome,
                            preco=preco
                        )
                        db.session.add(novo_tipo)
                        db.session.flush()  # Para obter o ID do tipo
                        tipos_inscricao.append(novo_tipo)
                
                # Processar os lotes de inscrição somente se habilitar_lotes for True
                if habilitar_lotes:
                    lote_nomes = request.form.getlist('lote_nome[]')
                    lote_ordens = request.form.getlist('lote_ordem[]')
                    lote_usar_data = [item == 'on' for item in request.form.getlist('lote_usar_data[]')]
                    lote_data_inicio = request.form.getlist('lote_data_inicio[]')
                    lote_data_fim = request.form.getlist('lote_data_fim[]')
                    lote_usar_qtd = [item == 'on' for item in request.form.getlist('lote_usar_qtd[]')]
                    lote_qtd_maxima = request.form.getlist('lote_qtd_maxima[]')
                    
                    # Criar cada lote
                    for i, nome in enumerate(lote_nomes):
                        if nome.strip():
                            # Determinar se usa data ou quantidade
                            data_inicio_lote = None
                            data_fim_lote = None
                            qtd_maxima = None
                            
                            if i < len(lote_usar_data) and lote_usar_data[i]:
                                if i < len(lote_data_inicio) and lote_data_inicio[i]:
                                    data_inicio_lote = datetime.strptime(lote_data_inicio[i], '%Y-%m-%d')
                                if i < len(lote_data_fim) and lote_data_fim[i]:
                                    data_fim_lote = datetime.strptime(lote_data_fim[i], '%Y-%m-%d')
                            
                            if i < len(lote_usar_qtd) and lote_usar_qtd[i]:
                                if i < len(lote_qtd_maxima) and lote_qtd_maxima[i]:
                                    qtd_maxima = int(lote_qtd_maxima[i])
                            
                            # Criar o lote
                            novo_lote = LoteInscricao(
                                evento_id=novo_evento.id,
                                nome=nome,
                                data_inicio=data_inicio_lote,
                                data_fim=data_fim_lote,
                                qtd_maxima=qtd_maxima,
                                ordem=int(lote_ordens[i]) if i < len(lote_ordens) and lote_ordens[i] else i+1,
                                ativo=True
                            )
                            db.session.add(novo_lote)
                            db.session.flush()  # Para obter o ID do lote
                            
                            # Processar preços por tipo de inscrição para este lote
                            for j, tipo in enumerate(tipos_inscricao):
                                # O formato do name é lote_tipo_preco_0_1 onde 0 é o índice do lote e 1 é o índice do tipo
                                preco_key = f'lote_tipo_preco_{i}_{j}'
                                preco_lote = request.form.get(preco_key)
                                
                                if preco_lote:
                                    # Se o evento for gratuito, todos os preços são 0
                                    preco_valor = 0.0 if inscricao_gratuita else float(preco_lote)
                                    
                                    novo_preco = LoteTipoInscricao(
                                        lote_id=novo_lote.id,
                                        tipo_inscricao_id=tipo.id,
                                        preco=preco_valor
                                    )
                                    db.session.add(novo_preco)
            
            db.session.commit()
            flash('Evento criado com sucesso!', 'success')
            flash('Agora você pode configurar as regras de inscrição para este evento.', 'info')
            return redirect(url_for('routes.dashboard_cliente'))
        
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao criar evento: {str(e)}', 'danger')

    # Retorna ao template, passando o 'evento' mesmo que seja None
    return render_template('criar_evento.html', evento=evento)


@routes.route('/configurar_regras_inscricao', methods=['GET', 'POST'])
@login_required
def configurar_regras_inscricao():
    if current_user.tipo != 'cliente':
        flash('Acesso negado!', 'danger')
        return redirect(url_for('routes.dashboard_cliente'))
    
    # Lista todos os eventos do cliente
    eventos = Evento.query.filter_by(cliente_id=current_user.id).all()
    
    # Evento selecionado (por padrão, None até que o usuário escolha)
    evento_id = request.args.get('evento_id') or (request.form.get('evento_id') if request.method == 'POST' else None)
    evento = None
    oficinas = []
    regras = {}
    
    if evento_id:
        evento = Evento.query.filter_by(id=evento_id, cliente_id=current_user.id).first()
        if evento:
            # Carrega oficinas do evento
            oficinas = Oficina.query.filter_by(evento_id=evento.id).all()
            
            # Carrega regras existentes
            regras_db = RegraInscricaoEvento.query.filter_by(evento_id=evento.id).all()
            for regra in regras_db:
                regras[regra.tipo_inscricao_id] = {
                    'limite_oficinas': regra.limite_oficinas,
                    'oficinas_permitidas_list': regra.get_oficinas_permitidas_list()
                }
    
    if request.method == 'POST' and evento:
        try:
            # Primeiro, remove todas as regras existentes para este evento
            RegraInscricaoEvento.query.filter_by(evento_id=evento.id).delete()
            
            # Processa cada tipo de inscrição
            for tipo in evento.tipos_inscricao_evento:
                limite_oficinas = int(request.form.get(f'limite_oficinas_{tipo.id}', 0))
                oficinas_permitidas = request.form.getlist(f'oficinas_{tipo.id}[]')
                
                # Cria nova regra
                nova_regra = RegraInscricaoEvento(
                    evento_id=evento.id,
                    tipo_inscricao_id=tipo.id,
                    limite_oficinas=limite_oficinas
                )
                
                # Define as oficinas permitidas
                nova_regra.set_oficinas_permitidas_list(oficinas_permitidas)
                
                db.session.add(nova_regra)
            
            db.session.commit()
            flash('Regras de inscrição configuradas com sucesso!', 'success')
            return redirect(url_for('routes.dashboard_cliente'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao configurar regras: {str(e)}', 'danger')
    
    return render_template(
        'configurar_regras_inscricao.html', 
        eventos=eventos, 
        evento=evento, 
        oficinas=oficinas,
        regras=regras
    )

@routes.route('/evento/<identifier>')
def pagina_evento(identifier):
    evento = Evento.query.filter_by(token=identifier).first_or_404()

    oficinas = Oficina.query.filter_by(evento_id=evento.id).order_by(Oficina.data, Oficina.horario_inicio).all()

    # Agrupando oficinas por data
    from collections import defaultdict
    grouped_oficinas = defaultdict(list)
    ministrantes_set = set()

    for oficina in oficinas:
        data_str = oficina.data.strftime('%d/%m/%Y')
        grouped_oficinas[data_str].append(oficina)
        if oficina.ministrante:
            ministrantes_set.add(oficina.ministrante)

    sorted_keys = sorted(grouped_oficinas.keys(), key=lambda date: datetime.strptime(date, '%d/%m/%Y'))

    # Garante que estamos enviando uma lista e não um conjunto
    ministrantes = list(ministrantes_set)

    return render_template(
        'pagina_evento.html',
        evento=evento,
        grouped_oficinas=grouped_oficinas,
        sorted_keys=sorted_keys,
        ministrantes=ministrantes  # Passa os ministrantes para o template
    )

@routes.route('/formulario_templates', methods=['GET'])
@login_required
def listar_templates():
    if current_user.tipo not in ['admin', 'cliente']:
        flash('Acesso negado!', 'danger')
        return redirect(url_for('routes.dashboard'))
    
    # Filter by cliente if not admin
    if current_user.tipo == 'cliente':
        templates = FormularioTemplate.query.filter(
            (FormularioTemplate.cliente_id == current_user.id) | 
            (FormularioTemplate.is_default == True)
        ).all()
    else:  # Admin sees all templates
        templates = FormularioTemplate.query.all()
        
    return render_template('templates_formulario.html', templates=templates)

@routes.route('/formulario_templates/novo', methods=['GET', 'POST'])
@login_required
def criar_template():
    if current_user.tipo not in ['admin', 'cliente']:
        flash('Acesso negado!', 'danger')
        return redirect(url_for('routes.dashboard'))
    
    if request.method == 'POST':
        nome = request.form.get('nome')
        descricao = request.form.get('descricao')
        categoria = request.form.get('categoria')
        is_default = request.form.get('is_default') == 'on'
        
        # Only admin can create default templates
        if current_user.tipo != 'admin' and is_default:
            is_default = False
        
        novo_template = FormularioTemplate(
            nome=nome,
            descricao=descricao,
            categoria=categoria,
            is_default=is_default,
            cliente_id=None if is_default else current_user.id
        )
        
        db.session.add(novo_template)
        db.session.commit()
        
        flash('Template criado com sucesso!', 'success')
        return redirect(url_for('routes.gerenciar_campos_template', template_id=novo_template.id))
    
    return render_template('criar_template.html')

@routes.route('/formulario_templates/<int:template_id>/campos', methods=['GET', 'POST'])
@login_required
def gerenciar_campos_template(template_id):
    template = FormularioTemplate.query.get_or_404(template_id)
    
    # Check permissions
    if current_user.tipo != 'admin' and template.cliente_id != current_user.id and not template.is_default:
        flash('Acesso negado!', 'danger')
        return redirect(url_for('routes.listar_templates'))
    
    if request.method == 'POST':
        nome = request.form.get('nome')
        tipo = request.form.get('tipo')
        opcoes = request.form.get('opcoes', '').strip()
        obrigatorio = request.form.get('obrigatorio') == 'on'
        ordem = request.form.get('ordem', 0)
        
        novo_campo = CampoFormularioTemplate(
            template_id=template.id,
            nome=nome,
            tipo=tipo,
            opcoes=opcoes if tipo in ['dropdown', 'checkbox', 'radio'] else None,
            obrigatorio=obrigatorio,
            ordem=ordem
        )
        
        db.session.add(novo_campo)
        db.session.commit()
        
        flash('Campo adicionado com sucesso!', 'success')
        return redirect(url_for('routes.gerenciar_campos_template', template_id=template.id))
    
    return render_template('gerenciar_campos_template.html', template=template)

@routes.route('/formulario_templates/<int:template_id>/usar', methods=['GET', 'POST'])
@login_required
def usar_template(template_id):
    template = FormularioTemplate.query.get_or_404(template_id)
    
    if request.method == 'POST':
        nome = request.form.get('nome')
        descricao = request.form.get('descricao')
        
        # Create new form from template
        novo_formulario = Formulario(
            nome=nome,
            descricao=descricao,
            cliente_id=current_user.id
        )
        db.session.add(novo_formulario)
        db.session.flush()  # Get ID before committing
        
        # Copy fields from template
        for campo_template in sorted(template.campos, key=lambda x: x.ordem):
            novo_campo = CampoFormulario(
                formulario_id=novo_formulario.id,
                nome=campo_template.nome,
                tipo=campo_template.tipo,
                opcoes=campo_template.opcoes,
                obrigatorio=campo_template.obrigatorio
            )
            db.session.add(novo_campo)
        
        db.session.commit()
        flash('Formulário criado com sucesso a partir do template!', 'success')
        return redirect(url_for('routes.listar_formularios'))
    
    return render_template('usar_template.html', template=template)

# Rotas para gerenciamento de agendamentos (para professores/participantes)
@routes.route('/professor/eventos_disponiveis')
@login_required
def eventos_disponiveis_professor():
    # Apenas participantes (professores) podem acessar
    if current_user.tipo != 'professor':
        flash('Acesso negado! Esta área é exclusiva para professores.', 'danger')
        return redirect(url_for('routes.dashboard'))
    
    # Buscar eventos disponíveis para agendamento
    eventos = Evento.query.filter(
        Evento.data_inicio <= datetime.utcnow(),
        Evento.data_fim >= datetime.utcnow(),
        Evento.status == 'ativo'
    ).all()
    
    return render_template(
        'professor/eventos_disponiveis.html',
        eventos=eventos
    )


@routes.route('/professor/evento/<int:evento_id>')
@login_required
def detalhes_evento_professor(evento_id):
    # Apenas participantes (professores) podem acessar
    if current_user.tipo != 'professor':
        flash('Acesso negado! Esta área é exclusiva para professores.', 'danger')
        return redirect(url_for('routes.dashboard'))
    
    evento = Evento.query.get_or_404(evento_id)
    
    # Verificar se o professor está bloqueado
    bloqueio = ProfessorBloqueado.query.filter_by(
        professor_id=current_user.id,
        evento_id=evento_id
    ).filter(ProfessorBloqueado.data_final >= datetime.utcnow()).first()
    
    # Buscar salas do evento
    salas = SalaVisitacao.query.filter_by(evento_id=evento_id).all()
    
    return render_template(
        'professor/detalhes_evento.html',
        evento=evento,
        bloqueio=bloqueio,
        salas=salas
    )


@routes.route('/professor/horarios_disponiveis/<int:evento_id>')
@login_required
def horarios_disponiveis_professor(evento_id):
    # Apenas participantes (professores) podem acessar
    if current_user.tipo != 'professor':
        flash('Acesso negado! Esta área é exclusiva para professores.', 'danger')
        return redirect(url_for('routes.dashboard'))
    
    # Verificar se o professor está bloqueado
    bloqueio = ProfessorBloqueado.query.filter_by(
        professor_id=current_user.id,
        evento_id=evento_id
    ).filter(ProfessorBloqueado.data_final >= datetime.utcnow()).first()
    
    if bloqueio:
        flash(f'Você está temporariamente bloqueado até {bloqueio.data_final.strftime("%d/%m/%Y")}. Motivo: {bloqueio.motivo}', 'danger')
        return redirect(url_for('routes.eventos_disponiveis_professor'))
    
    evento = Evento.query.get_or_404(evento_id)
    
    # Filtrar por data
    data_filtro = request.args.get('data')
    
    # Base da consulta - horários com vagas disponíveis
    query = HorarioVisitacao.query.filter_by(
        evento_id=evento_id
    ).filter(HorarioVisitacao.vagas_disponiveis > 0)
    
    # Filtrar apenas datas futuras (a partir de amanhã)
    amanha = datetime.now().date() + timedelta(days=1)
    query = query.filter(HorarioVisitacao.data >= amanha)
    
    # Aplicar filtro por data específica
    if data_filtro:
        data_filtrada = datetime.strptime(data_filtro, '%Y-%m-%d').date()
        query = query.filter(HorarioVisitacao.data == data_filtrada)
    
    # Ordenar por data e horário
    horarios = query.order_by(
        HorarioVisitacao.data,
        HorarioVisitacao.horario_inicio
    ).all()
    
    # Agrupar horários por data para facilitar a visualização
    horarios_por_data = {}
    for horario in horarios:
        data_str = horario.data.strftime('%Y-%m-%d')
        if data_str not in horarios_por_data:
            horarios_por_data[data_str] = []
        horarios_por_data[data_str].append(horario)
    
    return render_template(
        'professor/horarios_disponiveis.html',
        evento=evento,
        horarios_por_data=horarios_por_data,
        data_filtro=data_filtro
    )


@routes.route('/professor/criar_agendamento/<int:horario_id>', methods=['GET', 'POST'])
@login_required
def criar_agendamento_professor(horario_id):
    # Apenas participantes (professores) podem acessar
    if current_user.tipo != 'professor':
        flash('Acesso negado! Esta área é exclusiva para professores.', 'danger')
        return redirect(url_for('routes.dashboard'))
    
    horario = HorarioVisitacao.query.get_or_404(horario_id)
    evento = horario.evento
    
    # Verificar se o professor está bloqueado
    bloqueio = ProfessorBloqueado.query.filter_by(
        professor_id=current_user.id,
        evento_id=evento.id
    ).filter(ProfessorBloqueado.data_final >= datetime.utcnow()).first()
    
    if bloqueio:
        flash(f'Você está temporariamente bloqueado até {bloqueio.data_final.strftime("%d/%m/%Y")}. Motivo: {bloqueio.motivo}', 'danger')
        return redirect(url_for('routes.eventos_disponiveis_professor'))
    
    # Verificar se ainda há vagas
    if horario.vagas_disponiveis <= 0:
        flash('Não há mais vagas disponíveis para este horário!', 'warning')
        return redirect(url_for('routes.horarios_disponiveis_professor', evento_id=evento.id))
    
    # Buscar salas para seleção
    salas = SalaVisitacao.query.filter_by(evento_id=evento.id).all()
    
    if request.method == 'POST':
        # Validar campos obrigatórios
        escola_nome = request.form.get('escola_nome')
        escola_codigo_inep = request.form.get('escola_codigo_inep')
        turma = request.form.get('turma')
        nivel_ensino = request.form.get('nivel_ensino')
        quantidade_alunos = request.form.get('quantidade_alunos', type=int)
        salas_selecionadas = request.form.getlist('salas_selecionadas')
        
        if not escola_nome or not turma or not nivel_ensino or not quantidade_alunos:
            flash('Preencha todos os campos obrigatórios!', 'danger')
        elif quantidade_alunos <= 0:
            flash('A quantidade de alunos deve ser maior que zero!', 'danger')
        elif quantidade_alunos > horario.vagas_disponiveis:
            flash(f'Não há vagas suficientes! Disponíveis: {horario.vagas_disponiveis}', 'danger')
        else:
            # Criar o agendamento
            agendamento = AgendamentoVisita(
                horario_id=horario.id,
                professor_id=current_user.id,
                escola_nome=escola_nome,
                escola_codigo_inep=escola_codigo_inep,
                turma=turma,
                nivel_ensino=nivel_ensino,
                quantidade_alunos=quantidade_alunos,
                salas_selecionadas=','.join(salas_selecionadas) if salas_selecionadas else None
            )
            
            # Atualizar vagas disponíveis
            horario.vagas_disponiveis -= quantidade_alunos
            
            db.session.add(agendamento)
            
            try:
                db.session.commit()
                flash('Agendamento realizado com sucesso!', 'success')
                
                # Redirecionar para a página de adicionar alunos
                return redirect(url_for('routes.adicionar_alunos_agendamento', agendamento_id=agendamento.id))
            except Exception as e:
                db.session.rollback()
                flash(f'Erro ao realizar agendamento: {str(e)}', 'danger')
    
    return render_template(
        'professor/criar_agendamento.html',
        horario=horario,
        evento=evento,
        salas=salas
    )


@routes.route('/professor/adicionar_alunos/<int:agendamento_id>', methods=['GET', 'POST'])
@login_required
def adicionar_alunos_agendamento(agendamento_id):
    # Apenas participantes (professores) podem acessar
    if current_user.tipo != 'professor':
        flash('Acesso negado! Esta área é exclusiva para professores.', 'danger')
        return redirect(url_for('routes.dashboard'))
    
    agendamento = AgendamentoVisita.query.get_or_404(agendamento_id)
    
    # Verificar se o agendamento pertence ao professor
    if agendamento.professor_id != current_user.id:
        flash('Acesso negado! Este agendamento não pertence a você.', 'danger')
        return redirect(url_for('routes.meus_agendamentos'))
    
    # Lista de alunos já adicionados
    alunos = AlunoVisitante.query.filter_by(agendamento_id=agendamento.id).all()
    
    if request.method == 'POST':
        nome = request.form.get('nome')
        cpf = request.form.get('cpf')
        
        if nome:
            # Validar CPF se fornecido
            if cpf and len(cpf.replace('.', '').replace('-', '')) != 11:
                flash('CPF inválido. Digite apenas os números ou deixe em branco.', 'danger')
            else:
                aluno = AlunoVisitante(
                    agendamento_id=agendamento.id,
                    nome=nome,
                    cpf=cpf
                )
                db.session.add(aluno)
                
                try:
                    db.session.commit()
                    flash('Aluno adicionado com sucesso!', 'success')
                    # Recarregar a página para mostrar o aluno adicionado
                    return redirect(url_for('routes.adicionar_alunos_agendamento', agendamento_id=agendamento.id))
                except Exception as e:
                    db.session.rollback()
                    flash(f'Erro ao adicionar aluno: {str(e)}', 'danger')
        else:
            flash('Nome do aluno é obrigatório!', 'danger')
    
    return render_template(
        'professor/adicionar_alunos.html',
        agendamento=agendamento,
        alunos=alunos,
        total_adicionados=len(alunos),
        quantidade_esperada=agendamento.quantidade_alunos
    )


@routes.route('/professor/remover_aluno/<int:aluno_id>', methods=['POST'])
@login_required
def remover_aluno_agendamento(aluno_id):
    # Apenas participantes (professores) podem acessar
    if current_user.tipo != 'professor':
        flash('Acesso negado! Esta área é exclusiva para professores.', 'danger')
        return redirect(url_for('routes.dashboard'))
    
    aluno = AlunoVisitante.query.get_or_404(aluno_id)
    agendamento = aluno.agendamento
    
    # Verificar se o agendamento pertence ao professor
    if agendamento.professor_id != current_user.id:
        flash('Acesso negado! Este aluno não pertence a um agendamento seu.', 'danger')
        return redirect(url_for('routes.meus_agendamentos'))
    
    try:
        db.session.delete(aluno)
        db.session.commit()
        flash('Aluno removido com sucesso!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao remover aluno: {str(e)}', 'danger')
    
    return redirect(url_for('routes.adicionar_alunos_agendamento', agendamento_id=agendamento.id))


@routes.route('/professor/importar_alunos/<int:agendamento_id>', methods=['GET', 'POST'])
@login_required
def importar_alunos_agendamento(agendamento_id):
    # Apenas participantes (professores) podem acessar
    if current_user.tipo != 'professor':
        flash('Acesso negado! Esta área é exclusiva para professores.', 'danger')
        return redirect(url_for('routes.dashboard'))
    
    agendamento = AgendamentoVisita.query.get_or_404(agendamento_id)
    
    # Verificar se o agendamento pertence ao professor
    if agendamento.professor_id != current_user.id:
        flash('Acesso negado! Este agendamento não pertence a você.', 'danger')
        return redirect(url_for('routes.meus_agendamentos'))
    
    if request.method == 'POST':
        # Verificar se foi enviado um arquivo
        if 'arquivo_csv' not in request.files:
            flash('Nenhum arquivo selecionado!', 'danger')
            return redirect(request.url)
        
        arquivo = request.files['arquivo_csv']
        
        # Verificar se o arquivo tem nome
        if arquivo.filename == '':
            flash('Nenhum arquivo selecionado!', 'danger')
            return redirect(request.url)
        
        # Verificar se o arquivo é CSV
        if arquivo and arquivo.filename.endswith('.csv'):
            try:
                # Ler o conteúdo do arquivo
                conteudo = arquivo.read().decode('utf-8')
                linhas = conteudo.splitlines()
                
                # Contar alunos adicionados
                alunos_adicionados = 0
                
                # Processar cada linha do CSV
                for linha in linhas:
                    if ',' in linha:
                        # Formato esperado: Nome,CPF (opcional)
                        partes = linha.split(',')
                        nome = partes[0].strip()
                        cpf = partes[1].strip() if len(partes) > 1 else None
                        
                        if nome:
                            aluno = AlunoVisitante(
                                agendamento_id=agendamento.id,
                                nome=nome,
                                cpf=cpf
                            )
                            db.session.add(aluno)
                            alunos_adicionados += 1
                
                if alunos_adicionados > 0:
                    db.session.commit()
                    flash(f'{alunos_adicionados} alunos importados com sucesso!', 'success')
                else:
                    flash('Nenhum aluno encontrado no arquivo!', 'warning')
                
                return redirect(url_for('routes.adicionar_alunos_agendamento', agendamento_id=agendamento.id))
            except Exception as e:
                db.session.rollback()
                flash(f'Erro ao processar arquivo: {str(e)}', 'danger')
        else:
            flash('Arquivo deve estar no formato CSV!', 'danger')
    
    return render_template(
        'professor/importar_alunos.html',
        agendamento=agendamento
    )

from datetime import date, datetime, timedelta
@routes.route('/detalhes_agendamento/<int:agendamento_id>')
@login_required
def detalhes_agendamento(agendamento_id):
    if current_user.tipo != 'professor':
        flash("Acesso negado!", "danger")
        return redirect(url_for('routes.dashboard_professor'))

    # Buscar o agendamento no banco de dados
    agendamento = AgendamentoVisita.query.get_or_404(agendamento_id)

    # Buscar informações do horário e evento associados ao agendamento
    horario = HorarioVisitacao.query.get(agendamento.horario_id)
    evento = Evento.query.get(horario.evento_id) if horario else None

    # Buscar lista de alunos vinculados ao agendamento
    alunos = AlunoVisitante.query.filter_by(agendamento_id=agendamento.id).all()

    return render_template(
        'professor/detalhes_agendamento.html',
        agendamento=agendamento,
        horario=horario,
        evento=evento,
        alunos=alunos
    )

@routes.route('/professor/meus_agendamentos')
@login_required
def meus_agendamentos():
    # Apenas participantes (professores) podem acessar
    if current_user.tipo != 'professor':
        flash('Acesso negado! Esta área é exclusiva para professores.', 'danger')
        return redirect(url_for('routes.login'))
    
    # Filtros
    status = request.args.get('status')
    
    # Base da consulta
    query = AgendamentoVisita.query.filter_by(professor_id=current_user.id)
    
    # Aplicar filtros
    if status:
        query = query.filter(AgendamentoVisita.status == status)
    
    # Ordenar por data/horário
    agendamentos = query.join(
        HorarioVisitacao, AgendamentoVisita.horario_id == HorarioVisitacao.id
    ).order_by(
        HorarioVisitacao.data,
        HorarioVisitacao.horario_inicio
    ).all()
    
    return render_template(
        'professor/meus_agendamentos.html',
        agendamentos=agendamentos,
        status_filtro=status,
        today=date.today,
        hoje=date.today()
    )


@routes.route('/professor/cancelar_agendamento/<int:agendamento_id>', methods=['GET', 'POST'])
@login_required
def cancelar_agendamento_professor(agendamento_id):
    # Apenas participantes (professores) podem acessar
    if current_user.tipo != 'professor':
        flash('Acesso negado! Esta área é exclusiva para professores.', 'danger')
        return redirect(url_for('routes.dashboard'))
    
    agendamento = AgendamentoVisita.query.get_or_404(agendamento_id)
    
    # Verificar se o agendamento pertence ao professor
    if agendamento.professor_id != current_user.id:
        flash('Acesso negado! Este agendamento não pertence a você.', 'danger')
        return redirect(url_for('routes.meus_agendamentos'))
    
    # Verificar se o agendamento já foi cancelado
    if agendamento.status == 'cancelado':
        flash('Este agendamento já foi cancelado!', 'warning')
        return redirect(url_for('routes.meus_agendamentos'))
    
    # Verificar se o agendamento já foi realizado
    if agendamento.status == 'realizado':
        flash('Este agendamento já foi realizado e não pode ser cancelado!', 'warning')
        return redirect(url_for('routes.meus_agendamentos'))
    
    # Verificar prazo de cancelamento
    horario = agendamento.horario
    config = ConfiguracaoAgendamento.query.filter_by(evento_id=horario.evento_id).first()
    
    if config:
        # Calcular prazo limite para cancelamento
        data_hora_visita = datetime.combine(horario.data, horario.horario_inicio)
        prazo_limite = data_hora_visita - timedelta(hours=config.prazo_cancelamento)
        
        # Verificar se está dentro do prazo
        if datetime.utcnow() > prazo_limite:
            # Cancelamento fora do prazo - bloquear professor
            data_final_bloqueio = datetime.utcnow() + timedelta(days=config.tempo_bloqueio)
            
            # Criar registro de bloqueio
            bloqueio = ProfessorBloqueado(
                professor_id=current_user.id,
                evento_id=horario.evento_id,
                data_final=data_final_bloqueio,
                motivo=f"Cancelamento fora do prazo ({config.prazo_cancelamento}h antes) para o agendamento #{agendamento.id}"
            )
            db.session.add(bloqueio)
            
            flash(f'Atenção! Cancelamento fora do prazo. Você ficará bloqueado por {config.tempo_bloqueio} dias para novos agendamentos neste evento.', 'warning')
    
    if request.method == 'POST':
        # Restaurar vagas
        horario.vagas_disponiveis += agendamento.quantidade_alunos
        
        # Cancelar agendamento
        agendamento.status = 'cancelado'
        agendamento.data_cancelamento = datetime.utcnow()
        
        try:
            db.session.commit()
            flash('Agendamento cancelado com sucesso!', 'success')
            return redirect(url_for('routes.meus_agendamentos'))
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao cancelar agendamento: {str(e)}', 'danger')
    
    prazo_limite = data_hora_visita - timedelta(hours=config.prazo_cancelamento)
    
    return render_template(
        'professor/cancelar_agendamento.html',
        agendamento=agendamento,
        horario=horario,
        prazo_limite=prazo_limite
    )


@routes.route('/professor/imprimir_agendamento/<int:agendamento_id>')
@login_required
def imprimir_agendamento_professor(agendamento_id):
    # Apenas participantes (professores) podem acessar
    if current_user.tipo != 'professor':
        flash('Acesso negado! Esta área é exclusiva para professores.', 'danger')
        return redirect(url_for('routes.dashboard'))
    
    agendamento = AgendamentoVisita.query.get_or_404(agendamento_id)
    
    # Verificar se o agendamento pertence ao professor
    if agendamento.professor_id != current_user.id:
        flash('Acesso negado! Este agendamento não pertence a você.', 'danger')
        return redirect(url_for('routes.meus_agendamentos'))
    
    horario = agendamento.horario
    evento = horario.evento
    
    # Buscar salas selecionadas para visitação
    salas_ids = agendamento.salas_selecionadas.split(',') if agendamento.salas_selecionadas else []
    salas = SalaVisitacao.query.filter(SalaVisitacao.id.in_(salas_ids)).all() if salas_ids else []
    
    # Buscar alunos participantes
    alunos = AlunoVisitante.query.filter_by(agendamento_id=agendamento.id).all()
    
    # Gerar PDF para impressão
    pdf_filename = f"agendamento_{agendamento_id}.pdf"
    pdf_path = os.path.join("static", "agendamentos", pdf_filename)
    os.makedirs(os.path.dirname(pdf_path), exist_ok=True)
    
    # Chamar função para gerar PDF
    gerar_pdf_comprovante_agendamento(agendamento, horario, evento, salas, alunos, pdf_path)
    
    return send_file(pdf_path, as_attachment=True)


@routes.route('/professor/qrcode_agendamento/<int:agendamento_id>')
@login_required
def qrcode_agendamento_professor(agendamento_id):
    # Apenas participantes (professores) podem acessar
    if current_user.tipo != 'professor':
        flash('Acesso negado! Esta área é exclusiva para professores.', 'danger')
        return redirect(url_for('routes.dashboard'))
    
    agendamento = AgendamentoVisita.query.get_or_404(agendamento_id)
    
    # Verificar se o agendamento pertence ao professor
    if agendamento.professor_id != current_user.id:
        flash('Acesso negado! Este agendamento não pertence a você.', 'danger')
        return redirect(url_for('routes.meus_agendamentos'))
    
    # Página que exibe o QR Code para check-in
    return render_template(
        'professor/qrcode_agendamento.html',
        agendamento=agendamento,
        token=agendamento.qr_code_token
    )
    
    # Funções utilitárias para geração de PDFs

from datetime import datetime
from fpdf import FPDF # type: ignore
import qrcode # type: ignore
from PIL import Image # type: ignore
import io

import io
import qrcode
from fpdf import FPDF
from datetime import datetime
from models import AgendamentoVisita  # Ajuste conforme a sua importação


def gerar_pdf_comprovante_agendamento(agendamento, horario, evento, salas, alunos, caminho_pdf):
    """
    Gera um PDF com o comprovante de agendamento para o professor.
    
    Args:
        agendamento: Objeto AgendamentoVisita
        horario: Objeto HorarioVisitacao
        evento: Objeto Evento
        salas: Lista de objetos SalaVisitacao
        alunos: Lista de objetos AlunoVisitante
        caminho_pdf: Caminho onde o PDF será salvo
    """
    
    # 1) Obter todos os agendamentos do professor para compor o relatório
    agendamentos = AgendamentoVisita.query.filter_by(professor_id=agendamento.professor_id).all()
    
    # 2) Cria objeto PDF
    pdf = FPDF()
    pdf.add_page()

    # 3) -------------------------------
    #    SEÇÃO: Relatório de Agendamentos
    # 3.1) Título
    pdf.set_font('Arial', 'B', 16)
    pdf.cell(190, 10, f'Relatório de Agendamentos - {evento.nome}', 0, 1, 'C')
    
    # 3.2) Cabeçalho do evento
    pdf.set_font('Arial', 'B', 12)
    pdf.cell(190, 10, f'Evento: {evento.nome}', 0, 1)
    
    pdf.set_font('Arial', '', 12)
    if evento.data_inicio and evento.data_fim:
        pdf.cell(
            190, 10,
            f'Período: {evento.data_inicio.strftime("%d/%m/%Y")} a {evento.data_fim.strftime("%d/%m/%Y")}',
            0, 1
        )
    else:
        pdf.cell(190, 10, 'Período: não informado', 0, 1)

    pdf.cell(190, 10, f'Local: {evento.localizacao or "Não informado"}', 0, 1)

    # 3.3) Total de agendamentos
    pdf.set_font('Arial', 'B', 12)
    pdf.cell(190, 10, f'Total de agendamentos: {len(agendamentos)}', 0, 1)
    
    # 3.4) Resumo por status
    status_count = {'confirmado': 0, 'cancelado': 0, 'realizado': 0}
    alunos_esperados = 0
    alunos_presentes = 0
    
    for ag in agendamentos:
        status = ag.status
        status_count[status] = status_count.get(status, 0) + 1
        
        # Contar alunos
        alunos_esperados += ag.quantidade_alunos
        if ag.status == 'realizado':
            presentes = sum(1 for aluno in ag.alunos if aluno.presente)
            alunos_presentes += presentes
    
    pdf.cell(190, 10, f'Confirmados: {status_count["confirmado"]}', 0, 1)
    pdf.cell(190, 10, f'Cancelados: {status_count["cancelado"]}', 0, 1)
    pdf.cell(190, 10, f'Realizados: {status_count["realizado"]}', 0, 1)
    pdf.cell(190, 10, f'Total de alunos esperados: {alunos_esperados}', 0, 1)
    
    if alunos_presentes > 0:
        presenca = (alunos_presentes / alunos_esperados) * 100 if alunos_esperados > 0 else 0
        pdf.cell(
            190, 10,
            f'Total de alunos presentes: {alunos_presentes} ({presenca:.1f}%)',
            0, 1
        )
    
    # 3.5) Listagem de agendamentos
    pdf.ln(10)
    pdf.set_font('Arial', 'B', 14)
    pdf.cell(190, 10, 'Listagem de Agendamentos', 0, 1, 'C')
    
    # Cabeçalho da tabela
    pdf.set_font('Arial', 'B', 10)
    pdf.cell(20, 10, 'ID', 1, 0, 'C')
    pdf.cell(30, 10, 'Data', 1, 0, 'C')
    pdf.cell(30, 10, 'Horário', 1, 0, 'C')
    pdf.cell(50, 10, 'Escola', 1, 0, 'C')
    pdf.cell(30, 10, 'Alunos', 1, 0, 'C')
    pdf.cell(30, 10, 'Status', 1, 1, 'C')
    
    pdf.set_font('Arial', '', 9)
    for ag in agendamentos:
        h = ag.horario  # HorarioVisitacao
        escola_nome = ag.escola_nome
        if len(escola_nome) > 20:
            escola_nome = escola_nome[:17] + '...'
        
        pdf.cell(20, 10, str(ag.id), 1, 0, 'C')
        if h and h.data:
            pdf.cell(30, 10, h.data.strftime('%d/%m/%Y'), 1, 0, 'C')
        else:
            pdf.cell(30, 10, '--/--/----', 1, 0, 'C')
        
        if h and h.horario_inicio and h.horario_fim:
            pdf.cell(
                30, 10,
                f"{h.horario_inicio.strftime('%H:%M')} - {h.horario_fim.strftime('%H:%M')}",
                1, 0, 'C'
            )
        else:
            pdf.cell(30, 10, '---', 1, 0, 'C')
        
        pdf.cell(50, 10, escola_nome, 1, 0, 'L')
        pdf.cell(30, 10, str(ag.quantidade_alunos), 1, 0, 'C')
        pdf.cell(30, 10, ag.status.capitalize(), 1, 1, 'C')
    
    # Rodapé do relatório
    pdf.ln(10)
    pdf.set_font('Arial', 'I', 10)
    pdf.cell(
        190, 10,
        f'Relatório gerado em {datetime.now().strftime("%d/%m/%Y %H:%M")}',
        0, 1, 'C'
    )

    # 4) --------------------------------
    #    SEÇÃO: Comprovante de Agendamento (para este agendamento específico)
    # Precisamos de nova página para não escrever em cima do relatório
    pdf.add_page()

    pdf.set_font('Arial', 'B', 16)
    pdf.cell(190, 10, 'Comprovante de Agendamento', 0, 1, 'C')

    # Informações do evento
    pdf.set_font('Arial', 'B', 12)
    pdf.cell(190, 10, f'Evento: {evento.nome}', 0, 1)
    
    # Informações do agendamento
    pdf.set_font('Arial', '', 12)
    pdf.cell(190, 10, f'Cód. do Agendamento: #{agendamento.id}', 0, 1)

    if horario and horario.data:
        pdf.cell(190, 10, f'Data: {horario.data.strftime("%d/%m/%Y")}', 0, 1)
    else:
        pdf.cell(190, 10, 'Data: não informada', 0, 1)

    if horario and horario.horario_inicio and horario.horario_fim:
        pdf.cell(
            190, 10,
            f'Horário: {horario.horario_inicio.strftime("%H:%M")} às {horario.horario_fim.strftime("%H:%M")}',
            0, 1
        )
    else:
        pdf.cell(190, 10, 'Horário: não informado', 0, 1)
    
    if agendamento.professor:
        pdf.cell(190, 10, f'Professor: {agendamento.professor.nome}', 0, 1)
    else:
        pdf.cell(190, 10, 'Professor: --', 0, 1)

    pdf.cell(190, 10, f'Escola: {agendamento.escola_nome}', 0, 1)
    pdf.cell(190, 10, f'Turma: {agendamento.turma} - {agendamento.nivel_ensino}', 0, 1)
    pdf.cell(190, 10, f'Total de Alunos: {agendamento.quantidade_alunos}', 0, 1)
    
    if agendamento.data_agendamento:
        pdf.cell(
            190, 10,
            f'Agendado em: {agendamento.data_agendamento.strftime("%d/%m/%Y %H:%M")}',
            0, 1
        )
    else:
        pdf.cell(190, 10, 'Agendado em: --/--/---- --:--', 0, 1)
    
    # Status do agendamento
    pdf.set_font('Arial', 'B', 12)
    pdf.cell(190, 10, f'Status: {agendamento.status.upper()}', 0, 1)
    
    # Salas selecionadas
    if salas:
        pdf.ln(5)
        pdf.set_font('Arial', 'B', 12)
        pdf.cell(190, 10, 'Salas selecionadas:', 0, 1)
        pdf.set_font('Arial', '', 12)
        for sala in salas:
            pdf.cell(190, 10, f'- {sala.nome}', 0, 1)
    
    # Informações de check-in
    pdf.ln(5)
    pdf.set_font('Arial', 'B', 12)
    pdf.cell(190, 10, 'Informações para Check-in:', 0, 1)
    pdf.set_font('Arial', '', 12)
    if agendamento.checkin_realizado and agendamento.data_checkin:
        pdf.cell(
            190, 10,
            f'Check-in realizado em: {agendamento.data_checkin.strftime("%d/%m/%Y %H:%M")}',
            0, 1
        )
    else:
        pdf.cell(
            190, 10,
            'Apresente este comprovante no dia da visita para realizar o check-in.',
            0, 1
        )
    
    # 5) QR Code para check-in
    qr = qrcode.QRCode(version=1, box_size=10, border=5)
    qr.add_data(agendamento.qr_code_token)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    
    # Salvar imagem QR em buffer
    buffer = io.BytesIO()
    img.save(buffer, format='PNG')
    buffer.seek(0)
    
    # Adicionar QR Code
    pdf.ln(10)
    pdf.cell(190, 10, 'QR Code para Check-in:', 0, 1, 'C')
    # FPDF não aceita BytesIO diretamente como caminho
    # Precisamos salvar em arquivo temporário OU usar a abordagem "tempfile"
    # Abaixo, exemplificamos salvando em um arquivo temporário
    import tempfile

    with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as temp_img:
        temp_img.write(buffer.getvalue())
        temp_img_path = temp_img.name

    pdf.image(temp_img_path, x=75, y=pdf.get_y(), w=60)
    
    # Rodapé do comprovante
    pdf.ln(65)
    pdf.set_font('Arial', 'I', 10)
    pdf.cell(
        190, 10,
        'Este documento é seu comprovante oficial de agendamento.',
        0, 1, 'C'
    )
    pdf.cell(
        190, 10,
        f'Emitido em {datetime.now().strftime("%d/%m/%Y %H:%M")}',
        0, 1, 'C'
    )
    
    # 6) Finalmente, salvar o PDF (apenas uma vez!)
    pdf.output(caminho_pdf)


def gerar_pdf_relatorio_agendamentos(evento, agendamentos, caminho_pdf):
    """
    Gera um PDF com o relatório de agendamentos para o cliente/organizador.
    
    Args:
        evento: Objeto Evento
        agendamentos: Lista de objetos AgendamentoVisita
        caminho_pdf: Caminho onde o PDF será salvo
    """
    pdf = FPDF()
    pdf.add_page()
    
    # Configurações iniciais
    pdf.set_font('Arial', 'B', 16)
    
# Funções para manipulação de QR Code e checkin
import qrcode
from PIL import Image
import io
import os
from flask import send_file

def gerar_qrcode_url(token, tamanho=200):
    """
    Gera uma imagem QR Code para um token de agendamento
    
    Args:
        token: Token único do agendamento
        tamanho: Tamanho da imagem em pixels
        
    Returns:
        BytesIO: Buffer contendo a imagem do QR Code em formato PNG
    """
    # Preparar a URL para o QR Code (pode ser um endpoint de check-in)
    url = f"/checkin?token={token}"
    
    # Gerar QR Code
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(url)
    qr.make(fit=True)
    
    # Criar imagem
    img = qr.make_image(fill_color="black", back_color="white")
    
    # Salvar em um buffer em memória
    buffer = io.BytesIO()
    img.save(buffer, format="PNG")
    buffer.seek(0)
    
    return buffer


@routes.route('/api/qrcode/<token>')
def gerar_qrcode_token(token):
    """
    Endpoint para gerar e retornar uma imagem QR Code para um token
    """
    buffer = gerar_qrcode_url(token)
    return send_file(buffer, mimetype='image/png')


@routes.route('/checkin')
def checkin_token():
    """
    Endpoint para processamento de check-in via QR Code
    """
    token = request.args.get('token')
    
    if not token:
        flash('Token inválido ou não fornecido', 'danger')
        return redirect(url_for('routes.dashboard'))
    
    # Se não estiver logado, redirecionar para login
    if not current_user.is_authenticated:
        # Salvar token na sessão para usar após o login
        session['checkin_token'] = token
        flash('Faça login para processar o check-in', 'info')
        return redirect(url_for('auth.login', next=request.url))
    
    # Verificar se é um cliente (organizador)
    if current_user.tipo != 'cliente':
        flash('Apenas organizadores podem realizar check-in', 'danger')
        return redirect(url_for('routes.dashboard'))
    
    # Buscar o agendamento pelo token
    agendamento = AgendamentoVisita.query.filter_by(qr_code_token=token).first()
    
    if not agendamento:
        flash('Agendamento não encontrado', 'danger')
        return redirect(url_for('routes.dashboard_cliente'))
    
    # Verificar se o agendamento pertence a um evento do cliente
    evento_id = agendamento.horario.evento_id
    evento = Evento.query.get(evento_id)
    
    if evento.cliente_id != current_user.id:
        flash('Este agendamento não pertence a um evento seu', 'danger')
        return redirect(url_for('routes.dashboard_cliente'))
    
    # Verificar se já foi realizado check-in
    if agendamento.checkin_realizado:
        flash('Check-in já realizado para este agendamento', 'warning')
    else:
        # Realizar check-in
        return redirect(url_for(
            'routes.checkin_agendamento', 
            agendamento_id=agendamento.id,
            token=token
        ))
    
    # Redirecionar para detalhes do agendamento
    return redirect(url_for('routes.detalhes_agendamento', agendamento_id=agendamento.id))


@routes.route('/presenca_aluno/<int:aluno_id>', methods=['POST'])
@login_required
def marcar_presenca_aluno(aluno_id):
    """
    Marca presença individual de um aluno
    """
    if current_user.tipo != 'cliente':
        flash('Acesso negado!', 'danger')
        return redirect(url_for('routes.dashboard'))
    
    aluno = AlunoVisitante.query.get_or_404(aluno_id)
    agendamento = aluno.agendamento
    
    # Verificar se o agendamento pertence a um evento do cliente
    evento_id = agendamento.horario.evento_id
    evento = Evento.query.get(evento_id)
    
    if evento.cliente_id != current_user.id:
        flash('Este agendamento não pertence a um evento seu', 'danger')
        return redirect(url_for('routes.dashboard_cliente'))
    
    # Alternar estado de presença
    aluno.presente = not aluno.presente
    
    try:
        db.session.commit()
        if aluno.presente:
            flash(f'Presença de {aluno.nome} registrada com sucesso!', 'success')
        else:
            flash(f'Presença de {aluno.nome} removida com sucesso!', 'info')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao marcar presença: {str(e)}', 'danger')
    
    return redirect(url_for('routes.detalhes_agendamento', agendamento_id=agendamento.id))

# Módulo de notificações para agendamentos
from datetime import datetime, timedelta
from flask import render_template
from extensions import db, mail
from flask_mail import Message
import threading

class NotificacaoAgendamento:
    """
    Classe para gerenciar notificações de agendamentos
    """
    
    @staticmethod
    def enviar_email_confirmacao(agendamento):
        """
        Envia email de confirmação para o professor após um agendamento
        
        Args:
            agendamento: Objeto AgendamentoVisita
        """
        professor = agendamento.professor
        horario = agendamento.horario
        evento = horario.evento
        
        assunto = f"Confirmação de Agendamento - {evento.nome}"
        
        # Preparar o corpo do email
        corpo_html = render_template(
            'emails/confirmacao_agendamento.html',
            professor=professor,
            agendamento=agendamento,
            horario=horario,
            evento=evento
        )
        
        # Criar mensagem
        msg = Message(
            subject=assunto,
            recipients=[professor.email],
            html=corpo_html
        )
        
        # Enviar em um thread separado para não bloquear a resposta ao usuário
        thread = threading.Thread(target=NotificacaoAgendamento._enviar_email_async, args=[msg])
        thread.start()
    
    @staticmethod
    def enviar_email_cancelamento(agendamento):
        """
        Envia email de confirmação de cancelamento para o professor
        
        Args:
            agendamento: Objeto AgendamentoVisita
        """
        professor = agendamento.professor
        horario = agendamento.horario
        evento = horario.evento
        
        assunto = f"Confirmação de Cancelamento - {evento.nome}"
        
        # Preparar o corpo do email
        corpo_html = render_template(
            'emails/cancelamento_agendamento.html',
            professor=professor,
            agendamento=agendamento,
            horario=horario,
            evento=evento
        )
        
        # Criar mensagem
        msg = Message(
            subject=assunto,
            recipients=[professor.email],
            html=corpo_html
        )
        
        # Enviar em um thread separado
        thread = threading.Thread(target=NotificacaoAgendamento._enviar_email_async, args=[msg])
        thread.start()
    
    @staticmethod
    def enviar_lembrete_visita(agendamento):
        """
        Envia lembrete de visita para o professor um dia antes
        
        Args:
            agendamento: Objeto AgendamentoVisita
        """
        professor = agendamento.professor
        horario = agendamento.horario
        evento = horario.evento
        
        assunto = f"Lembrete de Visita Amanhã - {evento.nome}"
        
        # Preparar o corpo do email
        corpo_html = render_template(
            'emails/lembrete_visita.html',
            professor=professor,
            agendamento=agendamento,
            horario=horario,
            evento=evento
        )
        
        # Criar mensagem
        msg = Message(
            subject=assunto,
            recipients=[professor.email],
            html=corpo_html
        )
        
        # Enviar em um thread separado
        thread = threading.Thread(target=NotificacaoAgendamento._enviar_email_async, args=[msg])
        thread.start()
    
    @staticmethod
    def notificar_cliente_novo_agendamento(agendamento):
        """
        Notifica o cliente/organizador sobre um novo agendamento
        
        Args:
            agendamento: Objeto AgendamentoVisita
        """
        horario = agendamento.horario
        evento = horario.evento
        cliente = evento.cliente
        
        assunto = f"Novo Agendamento - {evento.nome}"
        
        # Preparar o corpo do email
        corpo_html = render_template(
            'emails/notificacao_novo_agendamento.html',
            cliente=cliente,
            agendamento=agendamento,
            horario=horario,
            evento=evento
        )
        
        # Criar mensagem
        msg = Message(
            subject=assunto,
            recipients=[cliente.email],
            html=corpo_html
        )
        
        # Enviar em um thread separado
        thread = threading.Thread(target=NotificacaoAgendamento._enviar_email_async, args=[msg])
        thread.start()
    
    @staticmethod
    def _enviar_email_async(msg):
        """
        Função interna para enviar email de forma assíncrona
        """
        with mail.connect() as conn:
            conn.send(msg)
    
    @staticmethod
    def processar_lembretes_diarios():
        """
        Tarefa agendada para enviar lembretes diários de visitas
        Deve ser executada uma vez por dia através de um agendador como Celery ou cron
        """
        # Data de amanhã
        amanha = datetime.utcnow().date() + timedelta(days=1)
        
        # Buscar todos os agendamentos confirmados para amanhã
        query = AgendamentoVisita.query.join(
            HorarioVisitacao, AgendamentoVisita.horario_id == HorarioVisitacao.id
        ).filter(
            HorarioVisitacao.data == amanha,
            AgendamentoVisita.status == 'confirmado'
        )
        
        agendamentos = query.all()
        
        # Enviar lembretes
        for agendamento in agendamentos:
            NotificacaoAgendamento.enviar_lembrete_visita(agendamento)


# Integrar notificações com as rotas
def integrar_notificacoes():
    """
    Integra as notificações com as rotas existentes
    Esta função deve ser chamada na inicialização da aplicação
    """
    # Sobrescrever a função de criação de agendamento para incluir notificação
    original_criar_agendamento = routes.criar_agendamento_professor
    
    @routes.route('/professor/criar_agendamento/<int:horario_id>', methods=['GET', 'POST'])
    @login_required
    def criar_agendamento_com_notificacao(horario_id):
        response = original_criar_agendamento(horario_id)
        
        # Verificar se o agendamento foi criado com sucesso
        # Isso é um hack - na prática seria melhor refatorar a função original
        # para retornar o agendamento criado ou um status
        if request.method == 'POST' and 'success' in session.get('_flashes', []):
            # Buscar o último agendamento criado pelo professor
            agendamento = AgendamentoVisita.query.filter_by(
                professor_id=current_user.id
            ).order_by(AgendamentoVisita.id.desc()).first()
            
            if agendamento:
                # Enviar notificações
                NotificacaoAgendamento.enviar_email_confirmacao(agendamento)
                NotificacaoAgendamento.notificar_cliente_novo_agendamento(agendamento)
        
        return response
    
    # Substituir a rota original pela nova versão com notificação
    routes.criar_agendamento_professor = criar_agendamento_com_notificacao
    
    # Fazer o mesmo para a função de cancelamento
    original_cancelar_agendamento = routes.cancelar_agendamento_professor
    
    @routes.route('/professor/cancelar_agendamento/<int:agendamento_id>', methods=['GET', 'POST'])
    @login_required
    def cancelar_agendamento_com_notificacao(agendamento_id):
        # Buscar o agendamento antes que seja cancelado
        agendamento = AgendamentoVisita.query.get_or_404(agendamento_id)
        
        response = original_cancelar_agendamento(agendamento_id)
        
        # Verificar se o cancelamento foi bem-sucedido
        if request.method == 'POST' and 'success' in session.get('_flashes', []):
            # Enviar notificação de cancelamento
            NotificacaoAgendamento.enviar_email_cancelamento(agendamento)
        
        return response
    
    # Substituir a rota original pela nova versão com notificação
    routes.cancelar_agendamento_professor = cancelar_agendamento_com_notificacao
    
@routes.route('/eventos_agendamento')
@login_required
def eventos_agendamento():
    # Verificar se é um cliente
    if current_user.tipo != 'cliente':
        flash('Acesso negado!', 'danger')
        return redirect(url_for('routes.dashboard'))
    
    hoje = datetime.utcnow().date()
    
    # Eventos ativos (em andamento)
    eventos_ativos = Evento.query.filter_by(
        cliente_id=current_user.id
    ).filter(
        and_(
            Evento.data_inicio <= hoje,
            Evento.data_fim >= hoje,
            Evento.status == 'ativo'
        )
    ).all()
    
    # Eventos futuros
    eventos_futuros = Evento.query.filter_by(
        cliente_id=current_user.id
    ).filter(
        and_(
            Evento.data_inicio > hoje,
            Evento.status == 'ativo'
        )
    ).all()
    
    # Eventos passados
    eventos_passados = Evento.query.filter_by(
        cliente_id=current_user.id
    ).filter(
        Evento.data_fim < hoje
    ).order_by(
        Evento.data_fim.desc()
    ).limit(10).all()
    
    # Contar agendamentos para cada evento
    for evento in eventos_ativos + eventos_futuros + eventos_passados:
        evento.agendamentos_count = db.session.query(func.count(AgendamentoVisita.id)).join(
            HorarioVisitacao, AgendamentoVisita.horario_id == HorarioVisitacao.id
        ).filter(
            HorarioVisitacao.evento_id == evento.id
        ).scalar() or 0
    
    return render_template(
        'eventos_agendamento.html',
        eventos_ativos=eventos_ativos,
        eventos_futuros=eventos_futuros,
        eventos_passados=eventos_passados
    )


@routes.route('/relatorio_geral_agendamentos')
@login_required
def relatorio_geral_agendamentos():
    # Verificar se é um cliente
    if current_user.tipo != 'cliente':
        flash('Acesso negado!', 'danger')
        return redirect(url_for('routes.dashboard'))
    
    # Filtros de data
    data_inicio = request.args.get('data_inicio')
    data_fim = request.args.get('data_fim')
    
    if data_inicio:
        data_inicio = datetime.strptime(data_inicio, '%Y-%m-%d').date()
    else:
        # Padrão: último mês
        data_inicio = datetime.utcnow().date() - timedelta(days=30)
    
    if data_fim:
        data_fim = datetime.strptime(data_fim, '%Y-%m-%d').date()
    else:
        data_fim = datetime.utcnow().date()
    
    # Buscar todos os eventos do cliente no período
    eventos = Evento.query.filter_by(
        cliente_id=current_user.id
    ).all()
    
    # Estatísticas gerais
    estatisticas = {}
    
    # Para cada evento, coletar estatísticas
    for evento in eventos:
        # Contar agendamentos por status
        confirmados = db.session.query(func.count(AgendamentoVisita.id)).join(
            HorarioVisitacao, AgendamentoVisita.horario_id == HorarioVisitacao.id
        ).filter(
            HorarioVisitacao.evento_id == evento.id,
            AgendamentoVisita.status == 'confirmado',
            HorarioVisitacao.data >= data_inicio,
            HorarioVisitacao.data <= data_fim
        ).scalar() or 0
        
        realizados = db.session.query(func.count(AgendamentoVisita.id)).join(
            HorarioVisitacao, AgendamentoVisita.horario_id == HorarioVisitacao.id
        ).filter(
            HorarioVisitacao.evento_id == evento.id,
            AgendamentoVisita.status == 'realizado',
            HorarioVisitacao.data >= data_inicio,
            HorarioVisitacao.data <= data_fim
        ).scalar() or 0
        
        cancelados = db.session.query(func.count(AgendamentoVisita.id)).join(
            HorarioVisitacao, AgendamentoVisita.horario_id == HorarioVisitacao.id
        ).filter(
            HorarioVisitacao.evento_id == evento.id,
            AgendamentoVisita.status == 'cancelado',
            HorarioVisitacao.data >= data_inicio,
            HorarioVisitacao.data <= data_fim
        ).scalar() or 0
        
        # Total de visitantes
        visitantes = db.session.query(func.sum(AgendamentoVisita.quantidade_alunos)).join(
            HorarioVisitacao, AgendamentoVisita.horario_id == HorarioVisitacao.id
        ).filter(
            HorarioVisitacao.evento_id == evento.id,
            AgendamentoVisita.status.in_(['confirmado', 'realizado']),
            HorarioVisitacao.data >= data_inicio,
            HorarioVisitacao.data <= data_fim
        ).scalar() or 0
        
        # Guardar estatísticas
        estatisticas[evento.id] = {
            'nome': evento.nome,
            'confirmados': confirmados,
            'realizados': realizados,
            'cancelados': cancelados,
            'total': confirmados + realizados + cancelados,
            'visitantes': visitantes
        }
    
    # Gerar PDF com estatísticas
    if request.args.get('gerar_pdf'):
        pdf_filename = f"relatorio_geral_{data_inicio.strftime('%Y%m%d')}_{data_fim.strftime('%Y%m%d')}.pdf"
        pdf_path = os.path.join("static", "relatorios", pdf_filename)
        os.makedirs(os.path.dirname(pdf_path), exist_ok=True)
        
        # Chamar função para gerar PDF
        gerar_pdf_relatorio_geral(eventos, estatisticas, data_inicio, data_fim, pdf_path)
        
        return send_file(pdf_path, as_attachment=True)
    
    return render_template(
        'relatorio_geral_agendamentos.html',
        eventos=eventos,
        estatisticas=estatisticas,
        filtros={
            'data_inicio': data_inicio,
            'data_fim': data_fim
        }
    )


@routes.route('/editar_horario_agendamento', methods=['POST'])
@login_required
def editar_horario_agendamento():
    # Verificar se é um cliente
    if current_user.tipo != 'cliente':
        flash('Acesso negado!', 'danger')
        return redirect(url_for('routes.dashboard'))
    
    horario_id = request.form.get('horario_id', type=int)
    horario = HorarioVisitacao.query.get_or_404(horario_id)
    evento = horario.evento
    
    # Verificar se o evento pertence ao cliente
    if evento.cliente_id != current_user.id:
        flash('Este evento não pertence a você!', 'danger')
        return redirect(url_for('routes.dashboard_cliente'))
    
    # Coletar dados do formulário
    horario_inicio = request.form.get('horario_inicio')
    horario_fim = request.form.get('horario_fim')
    capacidade_total = request.form.get('capacidade_total', type=int)
    vagas_disponiveis = request.form.get('vagas_disponiveis', type=int)
    
    if horario_inicio and horario_fim and capacidade_total is not None and vagas_disponiveis is not None:
        # Converter string para time
        horario.horario_inicio = datetime.strptime(horario_inicio, '%H:%M').time()
        horario.horario_fim = datetime.strptime(horario_fim, '%H:%M').time()
        
        # Verificar se a capacidade é menor que o número de agendamentos existentes
        agendamentos_count = db.session.query(func.count(AgendamentoVisita.id)).filter_by(
            horario_id=horario.id,
            status='confirmado'
        ).scalar() or 0
        
        agendamentos_alunos = db.session.query(func.sum(AgendamentoVisita.quantidade_alunos)).filter_by(
            horario_id=horario.id,
            status='confirmado'
        ).scalar() or 0
        
        if capacidade_total < agendamentos_alunos:
            flash(f'Não é possível reduzir a capacidade para {capacidade_total}. Já existem {agendamentos_alunos} alunos agendados.', 'danger')
            return redirect(url_for('routes.listar_horarios_agendamento', evento_id=evento.id))
        
        if vagas_disponiveis > capacidade_total:
            flash('As vagas disponíveis não podem ser maiores que a capacidade total!', 'danger')
            return redirect(url_for('routes.listar_horarios_agendamento', evento_id=evento.id))
        
        # Atualizar horário
        horario.capacidade_total = capacidade_total
        horario.vagas_disponiveis = vagas_disponiveis
        
        try:
            db.session.commit()
            flash('Horário atualizado com sucesso!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao atualizar horário: {str(e)}', 'danger')
    else:
        flash('Preencha todos os campos!', 'danger')
    
    return redirect(url_for('routes.listar_horarios_agendamento', evento_id=evento.id))


@routes.route('/excluir_horario_agendamento', methods=['POST'])
@login_required
def excluir_horario_agendamento():
    # Verificar se é um cliente
    if current_user.tipo != 'cliente':
        flash('Acesso negado!', 'danger')
        return redirect(url_for('routes.dashboard'))
    
    horario_id = request.form.get('horario_id', type=int)
    horario = HorarioVisitacao.query.get_or_404(horario_id)
    evento_id = horario.evento_id
    evento = horario.evento
    
    # Verificar se o evento pertence ao cliente
    if evento.cliente_id != current_user.id:
        flash('Este evento não pertence a você!', 'danger')
        return redirect(url_for('routes.dashboard_cliente'))
    
    # Verificar se existem agendamentos para este horário
    agendamentos = AgendamentoVisita.query.filter_by(
        horario_id=horario.id,
        status='confirmado'
    ).all()
    
    if agendamentos:
        # Cancelar todos os agendamentos
        for agendamento in agendamentos:
            agendamento.status = 'cancelado'
            agendamento.data_cancelamento = datetime.utcnow()
            
            # Enviar notificação de cancelamento
            # (Aqui pode-se adicionar código para enviar emails de cancelamento)
    
    try:
        # Excluir o horário
        db.session.delete(horario)
        db.session.commit()
        
        if agendamentos:
            flash(f'Horário excluído e {len(agendamentos)} agendamentos cancelados!', 'success')
        else:
            flash('Horário excluído com sucesso!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao excluir horário: {str(e)}', 'danger')
    
    return redirect(url_for('routes.listar_horarios_agendamento', evento_id=evento_id))

from fpdf import FPDF
from datetime import datetime

def gerar_pdf_relatorio_geral(eventos, estatisticas, data_inicio, data_fim, caminho_pdf):
    """
    Gera um relatório geral em PDF com estatísticas de agendamentos para todos os eventos.
    
    Args:
        eventos: Lista de objetos Evento
        estatisticas: Dicionário com estatísticas por evento
        data_inicio: Data inicial do período do relatório
        data_fim: Data final do período do relatório
        caminho_pdf: Caminho onde o PDF será salvo
    """
    pdf = FPDF()
    pdf.add_page()
    
    # Configurar fonte
    pdf.set_font('Arial', 'B', 16)
    
    # Título
    pdf.cell(190, 10, 'Relatório Geral de Agendamentos', 0, 1, 'C')
    pdf.set_font('Arial', '', 12)
    pdf.cell(190, 10, f'Período: {data_inicio.strftime("%d/%m/%Y")} a {data_fim.strftime("%d/%m/%Y")}', 0, 1, 'C')
    
    # Data e hora de geração
    pdf.set_font('Arial', 'I', 10)
    pdf.cell(190, 10, f'Gerado em: {datetime.now().strftime("%d/%m/%Y %H:%M")}', 0, 1, 'R')
    
    # Cálculo de totais
    total_confirmados = 0
    total_realizados = 0
    total_cancelados = 0
    total_visitantes = 0
    
    for stats in estatisticas.values():
        total_confirmados += stats['confirmados']
        total_realizados += stats['realizados']
        total_cancelados += stats['cancelados']
        total_visitantes += stats['visitantes']
    
    total_agendamentos = total_confirmados + total_realizados + total_cancelados
    
    # Resumo geral
    pdf.ln(5)
    pdf.set_font('Arial', 'B', 14)
    pdf.cell(190, 10, 'Resumo Geral', 0, 1)
    
    pdf.set_font('Arial', '', 12)
    pdf.cell(95, 10, f'Total de Agendamentos: {total_agendamentos}', 0, 0)
    pdf.cell(95, 10, f'Total de Visitantes: {total_visitantes}', 0, 1)
    
    pdf.cell(95, 10, f'Agendamentos Confirmados: {total_confirmados}', 0, 0)
    pdf.cell(95, 10, f'Agendamentos Realizados: {total_realizados}', 0, 1)
    
    pdf.cell(95, 10, f'Agendamentos Cancelados: {total_cancelados}', 0, 1)
    
    # Calcular taxas
    if total_agendamentos > 0:
        taxa_cancelamento = (total_cancelados / total_agendamentos) * 100
        pdf.cell(190, 10, f'Taxa de Cancelamento: {taxa_cancelamento:.1f}%', 0, 1)
    
    if total_confirmados + total_realizados > 0:
        taxa_conclusao = (total_realizados / (total_confirmados + total_realizados)) * 100
        pdf.cell(190, 10, f'Taxa de Conclusão: {taxa_conclusao:.1f}%', 0, 1)
    
    # Detalhes por evento
    pdf.ln(10)
    pdf.set_font('Arial', 'B', 14)
    pdf.cell(190, 10, 'Detalhes por Evento', 0, 1)
    
    # Cabeçalho da tabela
    pdf.set_font('Arial', 'B', 10)
    pdf.cell(60, 10, 'Evento', 1, 0, 'C')
    pdf.cell(25, 10, 'Confirmados', 1, 0, 'C')
    pdf.cell(25, 10, 'Realizados', 1, 0, 'C')
    pdf.cell(25, 10, 'Cancelados', 1, 0, 'C')
    pdf.cell(25, 10, 'Visitantes', 1, 0, 'C')
    pdf.cell(30, 10, 'Taxa Conclusão', 1, 1, 'C')
    
    # Dados da tabela
    pdf.set_font('Arial', '', 10)
    for evento_id, stats in estatisticas.items():
        evento_nome = stats['nome']
        
        # Limitar tamanho do nome para caber na tabela
        if len(evento_nome) > 30:
            evento_nome = evento_nome[:27] + '...'
        
        pdf.cell(60, 10, evento_nome, 1, 0)
        pdf.cell(25, 10, str(stats['confirmados']), 1, 0, 'C')
        pdf.cell(25, 10, str(stats['realizados']), 1, 0, 'C')
        pdf.cell(25, 10, str(stats['cancelados']), 1, 0, 'C')
        pdf.cell(25, 10, str(stats['visitantes']), 1, 0, 'C')
        
        # Calcular taxa de conclusão
        if stats['confirmados'] + stats['realizados'] > 0:
            taxa = (stats['realizados'] / (stats['confirmados'] + stats['realizados'])) * 100
            pdf.cell(30, 10, f'{taxa:.1f}%', 1, 1, 'C')
        else:
            pdf.cell(30, 10, 'N/A', 1, 1, 'C')
    
    # Análise e recomendações
    pdf.ln(10)
    pdf.set_font('Arial', 'B', 14)
    pdf.cell(190, 10, 'Análise e Recomendações', 0, 1)
    
    pdf.set_font('Arial', '', 12)
    if total_agendamentos > 0:
        if taxa_cancelamento > 30:
            pdf.multi_cell(190, 10, '- Alta taxa de cancelamento. Considere revisar suas políticas de cancelamento.')
            pdf.multi_cell(190, 10, '- Envie lembretes com mais frequência para professores com agendamentos confirmados.')
        else:
            pdf.multi_cell(190, 10, '- Taxa de cancelamento está em níveis aceitáveis.')
        
        if total_realizados < total_confirmados:
            pdf.multi_cell(190, 10, '- Implemente um sistema de lembretes mais eficiente para aumentar o comparecimento.')
            
        if total_visitantes < 100:
            pdf.multi_cell(190, 10, '- Divulgue mais seus eventos entre escolas e professores para aumentar a quantidade de visitantes.')
    else:
        pdf.multi_cell(190, 10, '- Ainda não há dados suficientes para recomendações personalizadas.')
    
    pdf.multi_cell(190, 10, '- Continue monitorando os agendamentos e ajustando a capacidade disponível conforme necessário.')
    
    # Rodapé
    pdf.ln(10)
    pdf.set_font('Arial', 'I', 10)
    pdf.cell(190, 10, 'Este relatório é gerado automaticamente pelo sistema de agendamentos.', 0, 1, 'C')
    
    # Salvar o PDF
    pdf.output(caminho_pdf)


def gerar_pdf_relatorio_agendamentos(evento, agendamentos, caminho_pdf):
    """
    Gera um relatório em PDF com a lista de agendamentos para um evento específico.
    
    Args:
        evento: Objeto Evento
        agendamentos: Lista de objetos AgendamentoVisita
        caminho_pdf: Caminho onde o PDF será salvo
    """
    pdf = FPDF()
    pdf.add_page()
    
    # Configurar fonte
    pdf.set_font('Arial', 'B', 16)
    
    # Título
    pdf.cell(190, 10, f'Relatório de Agendamentos - {evento.nome}', 0, 1, 'C')
    
    # Informações do evento
    pdf.set_font('Arial', '', 12)
    pdf.cell(190, 10, f'Local: {evento.local}', 0, 1)
    pdf.cell(190, 10, f'Período: {evento.data_inicio.strftime("%d/%m/%Y")} a {evento.data_fim.strftime("%d/%m/%Y")}', 0, 1)
    
    # Data e hora de geração
    pdf.set_font('Arial', 'I', 10)
    pdf.cell(190, 10, f'Gerado em: {datetime.now().strftime("%d/%m/%Y %H:%M")}', 0, 1, 'R')
    
    # Estatísticas
    pdf.ln(5)
    pdf.set_font('Arial', 'B', 14)
    pdf.cell(190, 10, 'Estatísticas', 0, 1)
    
    # Contadores
    total_agendamentos = len(agendamentos)
    confirmados = sum(1 for a in agendamentos if a.status == 'confirmado')
    realizados = sum(1 for a in agendamentos if a.status == 'realizado')
    cancelados = sum(1 for a in agendamentos if a.status == 'cancelado')
    
    total_alunos = sum(a.quantidade_alunos for a in agendamentos if a.status in ['confirmado', 'realizado'])
    presentes = 0
    for a in agendamentos:
        if a.status == 'realizado':
            presentes += sum(1 for aluno in a.alunos if aluno.presente)
    
    pdf.set_font('Arial', '', 12)
    pdf.cell(95, 10, f'Total de Agendamentos: {total_agendamentos}', 0, 0)
    pdf.cell(95, 10, f'Total de Alunos: {total_alunos}', 0, 1)
    
    pdf.cell(95, 10, f'Confirmados: {confirmados}', 0, 0)
    pdf.cell(95, 10, f'Realizados: {realizados}', 0, 1)
    
    pdf.cell(95, 10, f'Cancelados: {cancelados}', 0, 0)
    if realizados > 0:
        pdf.cell(95, 10, f'Alunos Presentes: {presentes}', 0, 1)
    
    # Lista de agendamentos
    pdf.ln(10)
    pdf.set_font('Arial', 'B', 14)
    pdf.cell(190, 10, 'Lista de Agendamentos', 0, 1)
    
    # Cabeçalho da tabela
    pdf.set_font('Arial', 'B', 9)
    pdf.cell(15, 10, 'ID', 1, 0, 'C')
    pdf.cell(25, 10, 'Data', 1, 0, 'C')
    pdf.cell(20, 10, 'Horário', 1, 0, 'C')
    pdf.cell(50, 10, 'Escola', 1, 0, 'C')
    pdf.cell(35, 10, 'Professor', 1, 0, 'C')
    pdf.cell(15, 10, 'Alunos', 1, 0, 'C')
    pdf.cell(30, 10, 'Status', 1, 1, 'C')
    
    # Dados da tabela
    pdf.set_font('Arial', '', 8)
    for agendamento in agendamentos:
        horario = agendamento.horario
        
        # Limitar tamanho dos nomes para caber na tabela
        escola_nome = agendamento.escola_nome
        if len(escola_nome) > 25:
            escola_nome = escola_nome[:22] + '...'
        
        professor_nome = agendamento.professor.nome
        if len(professor_nome) > 18:
            professor_nome = professor_nome[:15] + '...'
        
        pdf.cell(15, 8, str(agendamento.id), 1, 0, 'C')
        pdf.cell(25, 8, horario.data.strftime('%d/%m/%Y'), 1, 0, 'C')
        pdf.cell(20, 8, horario.horario_inicio.strftime('%H:%M'), 1, 0, 'C')
        pdf.cell(50, 8, escola_nome, 1, 0, 'L')
        pdf.cell(35, 8, professor_nome, 1, 0, 'L')
        pdf.cell(15, 8, str(agendamento.quantidade_alunos), 1, 0, 'C')
        
        status_txt = agendamento.status.capitalize()
        if agendamento.checkin_realizado:
            status_txt += " ✓"
        
        pdf.cell(30, 8, status_txt, 1, 1, 'C')
    
    # Rodapé
    pdf.ln(10)
    pdf.set_font('Arial', 'I', 10)
    pdf.cell(190, 10, 'Este relatório é gerado automaticamente pelo sistema de agendamentos.', 0, 1, 'C')
    
    # Salvar o PDF
    pdf.output(caminho_pdf)
    
# Rotas para a aba de agendamentos no dashboard
from datetime import datetime, timedelta
from sqlalchemy import and_, func, or_
from flask import render_template, redirect, url_for, flash, request, send_file, session, jsonify

# Importe os modelos necessários
from models import (
    Evento, 
    ConfiguracaoAgendamento, 
    SalaVisitacao, 
    HorarioVisitacao, 
    AgendamentoVisita, 
    AlunoVisitante, 
    ProfessorBloqueado
)

@routes.route('/dashboard_aba_agendamentos')
@login_required
def dashboard_aba_agendamentos():
    """
    Rota para carregar os dados da aba de agendamentos no dashboard do cliente.
    Esta rota é projetada para ser chamada via AJAX para popular a aba de agendamentos.
    """
    # Verificar se é um cliente
    if current_user.tipo != 'cliente':
        return jsonify(error='Acesso negado'), 403
    
    # Buscar eventos ativos
    eventos_ativos = Evento.query.filter_by(
        cliente_id=current_user.id
    ).filter(
        and_(
            Evento.data_inicio <= datetime.utcnow(),
            Evento.data_fim >= datetime.utcnow(),
            Evento.status == 'ativo'
        )
    ).all()
    
    # Dados para cards
    agendamentos_totais = db.session.query(func.count(AgendamentoVisita.id)).join(
        HorarioVisitacao, AgendamentoVisita.horario_id == HorarioVisitacao.id
    ).join(
        Evento, HorarioVisitacao.evento_id == Evento.id
    ).filter(
        Evento.cliente_id == current_user.id
    ).scalar() or 0
    
    agendamentos_confirmados = db.session.query(func.count(AgendamentoVisita.id)).join(
        HorarioVisitacao, AgendamentoVisita.horario_id == HorarioVisitacao.id
    ).join(
        Evento, HorarioVisitacao.evento_id == Evento.id
    ).filter(
        Evento.cliente_id == current_user.id,
        AgendamentoVisita.status == 'confirmado'
    ).scalar() or 0
    
    agendamentos_realizados = db.session.query(func.count(AgendamentoVisita.id)).join(
        HorarioVisitacao, AgendamentoVisita.horario_id == HorarioVisitacao.id
    ).join(
        Evento, HorarioVisitacao.evento_id == Evento.id
    ).filter(
        Evento.cliente_id == current_user.id,
        AgendamentoVisita.status == 'realizado'
    ).scalar() or 0
    
    agendamentos_cancelados = db.session.query(func.count(AgendamentoVisita.id)).join(
        HorarioVisitacao, AgendamentoVisita.horario_id == HorarioVisitacao.id
    ).join(
        Evento, HorarioVisitacao.evento_id == Evento.id
    ).filter(
        Evento.cliente_id == current_user.id,
        AgendamentoVisita.status == 'cancelado'
    ).scalar() or 0
    
    # Total de visitantes
    total_visitantes = db.session.query(func.sum(AgendamentoVisita.quantidade_alunos)).join(
        HorarioVisitacao, AgendamentoVisita.horario_id == HorarioVisitacao.id
    ).join(
        Evento, HorarioVisitacao.evento_id == Evento.id
    ).filter(
        Evento.cliente_id == current_user.id,
        AgendamentoVisita.status.in_(['confirmado', 'realizado'])
    ).scalar() or 0
    
    # Agendamentos para hoje
    hoje = datetime.utcnow().date()
    agendamentos_hoje = AgendamentoVisita.query.join(
        HorarioVisitacao, AgendamentoVisita.horario_id == HorarioVisitacao.id
    ).join(
        Evento, HorarioVisitacao.evento_id == Evento.id
    ).filter(
        Evento.cliente_id == current_user.id,
        HorarioVisitacao.data == hoje,
        AgendamentoVisita.status == 'confirmado'
    ).order_by(
        HorarioVisitacao.horario_inicio
    ).all()
    
    # Próximos agendamentos (próximos 7 dias, excluindo hoje)
    data_limite = hoje + timedelta(days=7)
    proximos_agendamentos = AgendamentoVisita.query.join(
        HorarioVisitacao, AgendamentoVisita.horario_id == HorarioVisitacao.id
    ).join(
        Evento, HorarioVisitacao.evento_id == Evento.id
    ).filter(
        Evento.cliente_id == current_user.id,
        HorarioVisitacao.data > hoje,
        HorarioVisitacao.data <= data_limite,
        AgendamentoVisita.status == 'confirmado'
    ).order_by(
        HorarioVisitacao.data,
        HorarioVisitacao.horario_inicio
    ).limit(5).all()
    
    # Calcular ocupação média (vagas preenchidas / capacidade total) 
    ocupacao_query = db.session.query(
        func.sum(HorarioVisitacao.capacidade_total - HorarioVisitacao.vagas_disponiveis).label('ocupadas'),
        func.sum(HorarioVisitacao.capacidade_total).label('total')
    ).join(
        Evento, HorarioVisitacao.evento_id == Evento.id
    ).filter(
        Evento.cliente_id == current_user.id,
        HorarioVisitacao.data >= hoje
    ).first()
    
    ocupacao_media = 0
    if ocupacao_query and ocupacao_query.total and ocupacao_query.total > 0:
        ocupacao_media = (ocupacao_query.ocupadas / ocupacao_query.total) * 100
    
    # Se for uma requisição AJAX, retornar JSON com os dados
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify({
            'eventos_ativos_count': len(eventos_ativos),
            'agendamentos_totais': agendamentos_totais,
            'agendamentos_confirmados': agendamentos_confirmados,
            'agendamentos_realizados': agendamentos_realizados,
            'agendamentos_cancelados': agendamentos_cancelados,
            'total_visitantes': total_visitantes,
            'ocupacao_media': round(ocupacao_media, 1) if ocupacao_media else 0,
            # Não é possível enviar objetos complexos via JSON, então apenas enviamos
            # um sinal de que há ou não agendamentos
            'tem_agendamentos_hoje': len(agendamentos_hoje) > 0,
            'tem_proximos_agendamentos': len(proximos_agendamentos) > 0
        })
    
    # Renderizar o template HTML da aba ou redirecionar para o dashboard
    # Dependendo de como sua aplicação lida com as abas
    return render_template(
        'partials/dashboard_agendamentos_aba.html',
        eventos_ativos=eventos_ativos,
        agendamentos_totais=agendamentos_totais,
        agendamentos_confirmados=agendamentos_confirmados,
        agendamentos_realizados=agendamentos_realizados,
        agendamentos_cancelados=agendamentos_cancelados,
        total_visitantes=total_visitantes,
        agendamentos_hoje=agendamentos_hoje,
        proximos_agendamentos=proximos_agendamentos,
        ocupacao_media=ocupacao_media
    )

@routes.route('/dashboard_aba_agendamentos_hoje')
@login_required
def dashboard_aba_agendamentos_hoje():
    """
    Rota para obter apenas os agendamentos de hoje para atualização dinâmica.
    """
    if current_user.tipo != 'cliente':
        return jsonify(error='Acesso negado'), 403
    
    hoje = datetime.utcnow().date()
    agendamentos_hoje = AgendamentoVisita.query.join(
        HorarioVisitacao, AgendamentoVisita.horario_id == HorarioVisitacao.id
    ).join(
        Evento, HorarioVisitacao.evento_id == Evento.id
    ).filter(
        Evento.cliente_id == current_user.id,
        HorarioVisitacao.data == hoje,
        AgendamentoVisita.status == 'confirmado'
    ).order_by(
        HorarioVisitacao.horario_inicio
    ).all()
    
    return render_template(
        'partials/agendamentos_hoje_lista.html',
        agendamentos_hoje=agendamentos_hoje
    )

@routes.route('/dashboard_aba_proximos_agendamentos')
@login_required
def dashboard_aba_proximos_agendamentos():
    """
    Rota para obter apenas os próximos agendamentos para atualização dinâmica.
    """
    if current_user.tipo != 'cliente':
        return jsonify(error='Acesso negado'), 403
    
    hoje = datetime.utcnow().date()
    data_limite = hoje + timedelta(days=7)
    proximos_agendamentos = AgendamentoVisita.query.join(
        HorarioVisitacao, AgendamentoVisita.horario_id == HorarioVisitacao.id
    ).join(
        Evento, HorarioVisitacao.evento_id == Evento.id
    ).filter(
        Evento.cliente_id == current_user.id,
        HorarioVisitacao.data > hoje,
        HorarioVisitacao.data <= data_limite,
        AgendamentoVisita.status == 'confirmado'
    ).order_by(
        HorarioVisitacao.data,
        HorarioVisitacao.horario_inicio
    ).limit(5).all()
    
    return render_template(
        'partials/proximos_agendamentos_lista.html',
        proximos_agendamentos=proximos_agendamentos
    )

# Função auxiliar para definir os valores na sessão
def set_dashboard_agendamentos_data():
    """
    Função auxiliar para calcular e armazenar em sessão os dados para a aba de agendamentos.
    Chamada antes de renderizar o dashboard principal para garantir que os dados estejam disponíveis.
    """
    if current_user.tipo != 'cliente':
        return
    
    # Buscar eventos ativos
    eventos_ativos = Evento.query.filter_by(
        cliente_id=current_user.id
    ).filter(
        and_(
            Evento.data_inicio <= datetime.utcnow(),
            Evento.data_fim >= datetime.utcnow(),
            Evento.status == 'ativo'
        )
    ).all()
    
    # Dados para cards
    agendamentos_totais = db.session.query(func.count(AgendamentoVisita.id)).join(
        HorarioVisitacao, AgendamentoVisita.horario_id == HorarioVisitacao.id
    ).join(
        Evento, HorarioVisitacao.evento_id == Evento.id
    ).filter(
        Evento.cliente_id == current_user.id
    ).scalar() or 0
    
    agendamentos_confirmados = db.session.query(func.count(AgendamentoVisita.id)).join(
        HorarioVisitacao, AgendamentoVisita.horario_id == HorarioVisitacao.id
    ).join(
        Evento, HorarioVisitacao.evento_id == Evento.id
    ).filter(
        Evento.cliente_id == current_user.id,
        AgendamentoVisita.status == 'confirmado'
    ).scalar() or 0
    
    agendamentos_realizados = db.session.query(func.count(AgendamentoVisita.id)).join(
        HorarioVisitacao, AgendamentoVisita.horario_id == HorarioVisitacao.id
    ).join(
        Evento, HorarioVisitacao.evento_id == Evento.id
    ).filter(
        Evento.cliente_id == current_user.id,
        AgendamentoVisita.status == 'realizado'
    ).scalar() or 0
    
    agendamentos_cancelados = db.session.query(func.count(AgendamentoVisita.id)).join(
        HorarioVisitacao, AgendamentoVisita.horario_id == HorarioVisitacao.id
    ).join(
        Evento, HorarioVisitacao.evento_id == Evento.id
    ).filter(
        Evento.cliente_id == current_user.id,
        AgendamentoVisita.status == 'cancelado'
    ).scalar() or 0
    
    # Total de visitantes
    total_visitantes = db.session.query(func.sum(AgendamentoVisita.quantidade_alunos)).join(
        HorarioVisitacao, AgendamentoVisita.horario_id == HorarioVisitacao.id
    ).join(
        Evento, HorarioVisitacao.evento_id == Evento.id
    ).filter(
        Evento.cliente_id == current_user.id,
        AgendamentoVisita.status.in_(['confirmado', 'realizado'])
    ).scalar() or 0
    
    # Agendamentos para hoje
    hoje = datetime.utcnow().date()
    agendamentos_hoje = AgendamentoVisita.query.join(
        HorarioVisitacao, AgendamentoVisita.horario_id == HorarioVisitacao.id
    ).join(
        Evento, HorarioVisitacao.evento_id == Evento.id
    ).filter(
        Evento.cliente_id == current_user.id,
        HorarioVisitacao.data == hoje,
        AgendamentoVisita.status == 'confirmado'
    ).order_by(
        HorarioVisitacao.horario_inicio
    ).all()
    
    # Próximos agendamentos (próximos 7 dias, excluindo hoje)
    data_limite = hoje + timedelta(days=7)
    proximos_agendamentos = AgendamentoVisita.query.join(
        HorarioVisitacao, AgendamentoVisita.horario_id == HorarioVisitacao.id
    ).join(
        Evento, HorarioVisitacao.evento_id == Evento.id
    ).filter(
        Evento.cliente_id == current_user.id,
        HorarioVisitacao.data > hoje,
        HorarioVisitacao.data <= data_limite,
        AgendamentoVisita.status == 'confirmado'
    ).order_by(
        HorarioVisitacao.data,
        HorarioVisitacao.horario_inicio
    ).limit(5).all()
    
    # Calcular ocupação média (vagas preenchidas / capacidade total) 
    ocupacao_query = db.session.query(
        func.sum(HorarioVisitacao.capacidade_total - HorarioVisitacao.vagas_disponiveis).label('ocupadas'),
        func.sum(HorarioVisitacao.capacidade_total).label('total')
    ).join(
        Evento, HorarioVisitacao.evento_id == Evento.id
    ).filter(
        Evento.cliente_id == current_user.id,
        HorarioVisitacao.data >= hoje
    ).first()
    
    ocupacao_media = 0
    if ocupacao_query and ocupacao_query.total and ocupacao_query.total > 0:
        ocupacao_media = (ocupacao_query.ocupadas / ocupacao_query.total) * 100
    
    # Armazenar valores na sessão para uso no template principal
    session['dashboard_agendamentos'] = {
        'eventos_ativos': len(eventos_ativos),
        'agendamentos_totais': agendamentos_totais,
        'agendamentos_confirmados': agendamentos_confirmados,
        'agendamentos_realizados': agendamentos_realizados,
        'agendamentos_cancelados': agendamentos_cancelados,
        'total_visitantes': total_visitantes,
        'ocupacao_media': round(ocupacao_media, 1) if ocupacao_media else 0,
        'tem_agendamentos_hoje': len(agendamentos_hoje) > 0,
        'tem_proximos_agendamentos': len(proximos_agendamentos) > 0,
        'timestamp': datetime.utcnow().timestamp()
    }
    
    # Passar os objetos para o contexto global
    return {
        'eventos_ativos': eventos_ativos,
        'agendamentos_totais': agendamentos_totais,
        'agendamentos_confirmados': agendamentos_confirmados,
        'agendamentos_realizados': agendamentos_realizados,
        'agendamentos_cancelados': agendamentos_cancelados,
        'total_visitantes': total_visitantes,
        'agendamentos_hoje': agendamentos_hoje,
        'proximos_agendamentos': proximos_agendamentos,
        'ocupacao_media': ocupacao_media
    }

@routes.route('/confirmar_checkin/<int:agendamento_id>', methods=['GET', 'POST'])
@login_required
def confirmar_checkin(agendamento_id):
    """
    Página de confirmação de check-in com detalhes do agendamento e lista de alunos.
    """
    # Verificar se é um cliente
    if current_user.tipo != 'cliente':
        flash('Acesso negado! Esta área é exclusiva para organizadores.', 'danger')
        return redirect(url_for('routes.dashboard'))
    
    agendamento = AgendamentoVisita.query.get_or_404(agendamento_id)
    
    # Verificar se o agendamento pertence a um evento do cliente
    evento = agendamento.horario.evento
    if evento.cliente_id != current_user.id:
        flash('Este agendamento não pertence a um evento seu!', 'danger')
        return redirect(url_for('routes.checkin_qr_agendamento'))
    
    # Verificar se já foi feito check-in
    if agendamento.checkin_realizado:
        flash('Check-in já foi realizado para este agendamento!', 'warning')
        return redirect(url_for('routes.detalhes_agendamento', agendamento_id=agendamento.id))
    
    if request.method == 'POST':
        # Realizar check-in
        agendamento.checkin_realizado = True
        agendamento.data_checkin = datetime.utcnow()
        agendamento.status = 'realizado'
        
        # Verificar se há alunos marcados como presentes
        alunos_presentes = request.form.getlist('alunos_presentes')
        
        # Marcar alunos como presentes
        for aluno in agendamento.alunos:
            aluno.presente = str(aluno.id) in alunos_presentes
        
        try:
            db.session.commit()
            flash('Check-in realizado com sucesso!', 'success')
            return redirect(url_for('routes.detalhes_agendamento', agendamento_id=agendamento.id))
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao realizar check-in: {str(e)}', 'danger')
    
    return render_template(
        'confirmar_checkin.html',
        agendamento=agendamento,
        horario=agendamento.horario,
        evento=evento
    )


@routes.route('/processar_qrcode', methods=['POST'])
@login_required
def processar_qrcode():
    """
    Endpoint AJAX para processar dados do QR Code escaneado.
    """
    # Verificar se é um cliente
    if current_user.tipo != 'cliente':
        return jsonify({'success': False, 'message': 'Acesso negado!'})
    
    # Obter dados do request
    data = request.json
    token = data.get('token')
    
    if not token:
        return jsonify({'success': False, 'message': 'Token não fornecido!'})
    
    # Buscar agendamento pelo token
    agendamento = AgendamentoVisita.query.filter_by(qr_code_token=token).first()
    
    if not agendamento:
        return jsonify({'success': False, 'message': 'Agendamento não encontrado!'})
    
    # Verificar se o agendamento pertence a um evento do cliente
    evento = agendamento.horario.evento
    if evento.cliente_id != current_user.id:
        return jsonify({'success': False, 'message': 'Este agendamento não pertence a um evento seu!'})
    
    # Verificar se já foi feito check-in
    if agendamento.checkin_realizado:
        return jsonify({
            'success': False, 
            'message': 'Check-in já realizado!',
            'redirect': url_for('routes.detalhes_agendamento', agendamento_id=agendamento.id)
        })
    
    # Retornar sucesso e URL para confirmação
    return jsonify({
        'success': True,
        'message': 'Agendamento encontrado!',
        'redirect': url_for('routes.confirmar_checkin', agendamento_id=agendamento.id)
    })

@routes.route('/configurar_agendamentos/<int:evento_id>', methods=['GET', 'POST'])
@login_required
def configurar_agendamentos(evento_id):
    """
    Rota para configurar as regras de agendamento para um evento específico.
    Permite definir horários disponíveis, prazos, capacidade, etc.
    """
    # Apenas clientes podem configurar agendamentos
    if current_user.tipo != 'cliente':
        flash('Acesso negado! Esta área é exclusiva para organizadores.', 'danger')
        return redirect(url_for('routes.dashboard'))
    
    evento = Evento.query.get_or_404(evento_id)
    
    # Verificar se o evento pertence ao cliente
    if evento.cliente_id != current_user.id:
        flash('Este evento não pertence a você!', 'danger')
        return redirect(url_for('routes.dashboard_cliente'))
    
    # Verificar se já existe configuração
    config = ConfiguracaoAgendamento.query.filter_by(
        cliente_id=current_user.id,
        evento_id=evento_id
    ).first()
    
    if request.method == 'POST':
        if config:
            # Atualizar configuração existente
            config.prazo_cancelamento = request.form.get('prazo_cancelamento', type=int)
            config.tempo_bloqueio = request.form.get('tempo_bloqueio', type=int)
            config.capacidade_padrao = request.form.get('capacidade_padrao', type=int)
            config.intervalo_minutos = request.form.get('intervalo_minutos', type=int)
            
            # Converter string para time
            hora_inicio = request.form.get('horario_inicio')
            hora_fim = request.form.get('horario_fim')
            config.horario_inicio = datetime.strptime(hora_inicio, '%H:%M').time()
            config.horario_fim = datetime.strptime(hora_fim, '%H:%M').time()
            
            # Dias da semana selecionados
            dias_semana = request.form.getlist('dias_semana')
            config.dias_semana = ','.join(dias_semana)
        else:
            # Criar nova configuração
            hora_inicio = request.form.get('horario_inicio')
            hora_fim = request.form.get('horario_fim')
            
            config = ConfiguracaoAgendamento(
                cliente_id=current_user.id,
                evento_id=evento_id,
                prazo_cancelamento=request.form.get('prazo_cancelamento', type=int),
                tempo_bloqueio=request.form.get('tempo_bloqueio', type=int),
                capacidade_padrao=request.form.get('capacidade_padrao', type=int),
                intervalo_minutos=request.form.get('intervalo_minutos', type=int),
                horario_inicio=datetime.strptime(hora_inicio, '%H:%M').time(),
                horario_fim=datetime.strptime(hora_fim, '%H:%M').time(),
                dias_semana=','.join(request.form.getlist('dias_semana'))
            )
            db.session.add(config)
        
        try:
            db.session.commit()
            flash('Configurações de agendamento salvas com sucesso!', 'success')
            return redirect(url_for('routes.gerar_horarios_agendamento', evento_id=evento_id))
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao salvar configurações: {str(e)}', 'danger')
    
    return render_template(
        'configurar_agendamentos.html',
        evento=evento,
        config=config
    )


@routes.route('/gerar_horarios_agendamento/<int:evento_id>', methods=['GET', 'POST'])
@login_required
def gerar_horarios_agendamento(evento_id):
    """
    Página para gerar horários de agendamento com base nas configurações.
    """
    if current_user.tipo != 'cliente':
        flash('Acesso negado! Esta área é exclusiva para organizadores.', 'danger')
        return redirect(url_for('routes.dashboard'))
    
    evento = Evento.query.get_or_404(evento_id)
    
    # Verificar se o evento pertence ao cliente
    if evento.cliente_id != current_user.id:
        flash('Este evento não pertence a você!', 'danger')
        return redirect(url_for('routes.dashboard_cliente'))
    
    config = ConfiguracaoAgendamento.query.filter_by(evento_id=evento_id).first_or_404()
    
    if request.method == 'POST':
        # Obter datas do form
        data_inicial = datetime.strptime(request.form.get('data_inicial'), '%Y-%m-%d').date()
        data_final = datetime.strptime(request.form.get('data_final'), '%Y-%m-%d').date()
        
        # Converter dias da semana para ints
        dias_permitidos = [int(dia) for dia in config.dias_semana.split(',')]
        
        # Gerar horários
        data_atual = data_inicial
        horarios_criados = 0
        
        while data_atual <= data_final:
            # Verificar se o dia da semana é permitido (0=Segunda, 6=Domingo na função weekday())
            # Ajuste: convert 0-6 (seg-dom) do input para 0-6 (seg-dom) do Python (que usa 0=seg, 6=dom)
            if data_atual.weekday() in dias_permitidos:
                # Horário atual começa no início configurado
                horario_atual = datetime.combine(data_atual, config.horario_inicio)
                hora_final = datetime.combine(data_atual, config.horario_fim)
                
                # Continuar gerando slots até atingir o horário final
                while horario_atual < hora_final:
                    # Calcular horário de término do slot
                    horario_fim = horario_atual + timedelta(minutes=config.intervalo_minutos)
                    
                    # Não ultrapassar o horário final do dia
                    if horario_fim > hora_final:
                        horario_fim = hora_final
                    
                    # Verificar se já existe esse horário
                    horario_existente = HorarioVisitacao.query.filter_by(
                        evento_id=evento_id,
                        data=data_atual,
                        horario_inicio=horario_atual.time(),
                        horario_fim=horario_fim.time()
                    ).first()
                    
                    if not horario_existente:
                        # Criar novo horário
                        novo_horario = HorarioVisitacao(
                            evento_id=evento_id,
                            data=data_atual,
                            horario_inicio=horario_atual.time(),
                            horario_fim=horario_fim.time(),
                            capacidade_total=config.capacidade_padrao,
                            vagas_disponiveis=config.capacidade_padrao
                        )
                        db.session.add(novo_horario)
                        horarios_criados += 1
                    
                    # Avançar para o próximo slot
                    horario_atual = horario_fim
            
            # Avançar para o próximo dia
            data_atual += timedelta(days=1)
        
        # Salvar alterações no banco
        try:
            db.session.commit()
            flash(f'{horarios_criados} horários de visitação foram criados com sucesso!', 'success')
            return redirect(url_for('routes.listar_horarios_agendamento', evento_id=evento_id))
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao gerar horários: {str(e)}', 'danger')
    
    return render_template(
        'gerar_horarios_agendamento.html',
        evento=evento,
        config=config
    )


@routes.route('/listar_horarios_agendamento/<int:evento_id>')
@login_required
def listar_horarios_agendamento(evento_id):
    """
    Página para listar e gerenciar os horários de agendamento disponíveis.
    """
    if current_user.tipo != 'cliente':
        flash('Acesso negado! Esta área é exclusiva para organizadores.', 'danger')
        return redirect(url_for('routes.dashboard'))
    
    evento = Evento.query.get_or_404(evento_id)
    
    # Verificar se o evento pertence ao cliente
    if evento.cliente_id != current_user.id:
        flash('Este evento não pertence a você!', 'danger')
        return redirect(url_for('routes.dashboard_cliente'))
    
    # Agrupar horários por data
    horarios = HorarioVisitacao.query.filter_by(evento_id=evento_id).order_by(
        HorarioVisitacao.data,
        HorarioVisitacao.horario_inicio
    ).all()
    
    # Agrupar horários por data para facilitar a visualização
    horarios_por_data = {}
    for horario in horarios:
        data_str = horario.data.strftime('%Y-%m-%d')
        if data_str not in horarios_por_data:
            horarios_por_data[data_str] = []
        horarios_por_data[data_str].append(horario)
    
    return render_template(
        'listar_horarios_agendamento.html',
        evento=evento,
        horarios_por_data=horarios_por_data
    )


@routes.route('/salas_visitacao/<int:evento_id>', methods=['GET', 'POST'])
@login_required
def salas_visitacao(evento_id):
    """
    Página para gerenciar as salas disponíveis para visitação.
    """
    if current_user.tipo != 'cliente':
        flash('Acesso negado! Esta área é exclusiva para organizadores.', 'danger')
        return redirect(url_for('routes.dashboard'))
    
    evento = Evento.query.get_or_404(evento_id)
    
    # Verificar se o evento pertence ao cliente
    if evento.cliente_id != current_user.id:
        flash('Este evento não pertence a você!', 'danger')
        return redirect(url_for('routes.dashboard_cliente'))
    
    if request.method == 'POST':
        nome = request.form.get('nome')
        descricao = request.form.get('descricao')
        capacidade = request.form.get('capacidade', type=int)
        
        if nome and capacidade:
            nova_sala = SalaVisitacao(
                nome=nome,
                descricao=descricao,
                capacidade=capacidade,
                evento_id=evento_id
            )
            db.session.add(nova_sala)
            
            try:
                db.session.commit()
                flash('Sala de visitação cadastrada com sucesso!', 'success')
            except Exception as e:
                db.session.rollback()
                flash(f'Erro ao cadastrar sala: {str(e)}', 'danger')
    
    # Listar salas existentes
    salas = SalaVisitacao.query.filter_by(evento_id=evento_id).all()
    
    return render_template(
        'salas_visitacao.html',
        evento=evento,
        salas=salas
    )

@routes.route('/gerar_relatorio_agendamentos/<int:evento_id>')
@login_required
def gerar_relatorio_agendamentos(evento_id):
    """
    Gera um relatório em PDF dos agendamentos para um evento.
    """
    if current_user.tipo != 'cliente':
        flash('Acesso negado!', 'danger')
        return redirect(url_for('routes.dashboard'))
    
    evento = Evento.query.get_or_404(evento_id)
    
    # Verificar se o evento pertence ao cliente
    if evento.cliente_id != current_user.id:
        flash('Este evento não pertence a você!', 'danger')
        return redirect(url_for('routes.dashboard_cliente'))
    
    # Filtros (mesmos da listagem)
    data_inicio = request.args.get('data_inicio')
    data_fim = request.args.get('data_fim')
    status = request.args.get('status')
    escola = request.args.get('escola')
    
    # Base da consulta
    query = AgendamentoVisita.query.join(
        HorarioVisitacao, AgendamentoVisita.horario_id == HorarioVisitacao.id
    ).filter(HorarioVisitacao.evento_id == evento_id)
    
    # Aplicar filtros
    if data_inicio:
        data_inicio = datetime.strptime(data_inicio, '%Y-%m-%d').date()
        query = query.filter(HorarioVisitacao.data >= data_inicio)
    
    if data_fim:
        data_fim = datetime.strptime(data_fim, '%Y-%m-%d').date()
        query = query.filter(HorarioVisitacao.data <= data_fim)
    
    if status:
        query = query.filter(AgendamentoVisita.status == status)
    
    if escola:
        query = query.filter(AgendamentoVisita.escola_nome.ilike(f'%{escola}%'))
    
    # Ordenar por data/horário
    agendamentos = query.order_by(
        HorarioVisitacao.data,
        HorarioVisitacao.horario_inicio
    ).all()
    
    # Gerar PDF
    pdf_filename = f"relatorio_agendamentos_{evento_id}.pdf"
    pdf_path = os.path.join("static", "relatorios", pdf_filename)
    os.makedirs(os.path.dirname(pdf_path), exist_ok=True)
    
    # Chamar função para gerar PDF
    gerar_pdf_relatorio_agendamentos(evento, agendamentos, pdf_path)
    
    return send_file(pdf_path, as_attachment=True)

@routes.route('/dashboard-agendamentos')
@login_required
def dashboard_agendamentos():
    # Inicializar variáveis vazias/padrão para o template
    eventos_ativos = []
    agendamentos_totais = 0
    total_visitantes = 0
    ocupacao_media = 0
    agendamentos_confirmados = 0
    agendamentos_realizados = 0
    agendamentos_cancelados = 0
    agendamentos_hoje = []
    proximos_agendamentos = []
    todos_agendamentos = []
    periodos_agendamento = []
    config_agendamento = None
    
    # Buscar eventos do cliente atual
    try:
        eventos_ativos = Evento.query.filter_by(cliente_id=current_user.id).all()
    except Exception as e:
        flash(f"Erro ao buscar eventos: {str(e)}", "danger")
    
    # Verificar se temos dados suficientes para mostrar a página básica
    return render_template('dashboard_agendamentos.html', 
                          eventos_ativos=eventos_ativos,
                          agendamentos_totais=agendamentos_totais,
                          total_visitantes=total_visitantes,
                          ocupacao_media=ocupacao_media,
                          agendamentos_confirmados=agendamentos_confirmados,
                          agendamentos_realizados=agendamentos_realizados,
                          agendamentos_cancelados=agendamentos_cancelados,
                          agendamentos_hoje=agendamentos_hoje,
                          proximos_agendamentos=proximos_agendamentos,
                          todos_agendamentos=todos_agendamentos,
                          periodos_agendamento=periodos_agendamento,
                          config_agendamento=config_agendamento)

@routes.route('/criar-agendamento', methods=['GET', 'POST'])
@login_required
def criar_agendamento():
    """
    Rota para criação de um novo agendamento.
    """
    # Inicialização de variáveis
    form_erro = None
    eventos = []
    
    # Buscar eventos disponíveis do cliente atual
    try:
        eventos = Evento.query.filter_by(cliente_id=current_user.id).all()
    except Exception as e:
        flash(f"Erro ao buscar eventos: {str(e)}", "danger")
    
    # Processar o formulário quando enviado via POST
    if request.method == 'POST':
        try:
            # Obter dados do formulário
            evento_id = request.form.get('evento_id')
            data = request.form.get('data')
            horario_id = request.form.get('horario_id')
            escola_nome = request.form.get('escola_nome')
            nome_responsavel = request.form.get('nome_responsavel')
            email_responsavel = request.form.get('email_responsavel')
            telefone_escola = request.form.get('telefone_escola')
            turma = request.form.get('turma')
            quantidade_alunos = request.form.get('quantidade_alunos')
            faixa_etaria = request.form.get('faixa_etaria')
            observacoes = request.form.get('observacoes')
            
            # Validar dados obrigatórios
            if not evento_id or not data or not horario_id or not escola_nome or not quantidade_alunos:
                form_erro = "Preencha todos os campos obrigatórios."
                flash(form_erro, "danger")
            else:
                # Aqui você adicionaria o código para criar um novo agendamento
                # baseado nos modelos que você tem disponíveis
                
                # Como não sabemos a estrutura exata do seu modelo de Agendamento,
                # vamos apenas exibir uma mensagem de sucesso
                flash("Agendamento criado com sucesso! Implementação completa pendente.", "success")
                
                # Redirecionar para o dashboard de agendamentos
                return redirect(url_for('routes.dashboard_agendamentos'))
                
        except Exception as e:
            form_erro = f"Erro ao processar o formulário: {str(e)}"
            flash(form_erro, "danger")
    
    # Renderizar o template com o formulário
    return render_template('criar_agendamento.html', 
                          eventos=eventos,
                          form_erro=form_erro)

@routes.route('/api/horarios-disponiveis')
@login_required
def horarios_disponiveis():
    """
    API para obter horários disponíveis para agendamento, baseado em evento e data.
    """
    evento_id = request.args.get('evento_id')
    data = request.args.get('data')
    
    # Verificar se os parâmetros foram fornecidos
    if not evento_id or not data:
        return jsonify({
            'success': False,
            'message': 'Parâmetros evento_id e data são obrigatórios'
        }), 400
    
    try:
        # Como não sabemos a estrutura exata do seu modelo de Horario,
        # vamos retornar dados simulados para teste
        # Em uma implementação real, você buscaria horários do banco de dados
        
        # Simular alguns horários como exemplo
        horarios_exemplo = [
            {
                'id': 1,
                'horario_inicio': '08:00',
                'horario_fim': '10:00',
                'vagas_disponiveis': 30
            },
            {
                'id': 2,
                'horario_inicio': '10:30',
                'horario_fim': '12:30',
                'vagas_disponiveis': 25
            },
            {
                'id': 3,
                'horario_inicio': '14:00',
                'horario_fim': '16:00',
                'vagas_disponiveis': 20
            }
        ]
        
        return jsonify({
            'success': True,
            'horarios': horarios_exemplo
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Erro ao buscar horários: {str(e)}'
        }), 500

@routes.route('/configurar-horarios-agendamento', methods=['GET', 'POST'])
@login_required
def configurar_horarios_agendamento():
    """
    Rota para configuração de horários disponíveis para agendamentos.
    """
    # Inicialização de variáveis
    form_erro = None
    eventos = []
    horarios_existentes = []
    evento_selecionado = None
    evento_id = request.args.get('evento_id', None)
    
    # Buscar eventos disponíveis do cliente atual
    try:
        eventos = Evento.query.filter_by(cliente_id=current_user.id).all()
        
        # Se um evento foi especificado na URL, buscamos seus detalhes
        if evento_id:
            evento_selecionado = Evento.query.filter_by(id=evento_id, cliente_id=current_user.id).first()
            
            # Buscar horários existentes para o evento selecionado
            if evento_selecionado:
                horarios_existentes = HorarioVisitacao.query.filter_by(evento_id=evento_id).all()
                
                # Transformar os horários do banco em dicionários para facilitar o uso no template
                horarios_existentes = [
                    {
                        'id': h.id,
                        'data': h.data.strftime('%Y-%m-%d'),
                        'horario_inicio': h.horario_inicio.strftime('%H:%M'),
                        'horario_fim': h.horario_fim.strftime('%H:%M'),
                        'capacidade': h.capacidade_total,
                        'agendamentos': h.capacidade_total - h.vagas_disponiveis
                    } for h in horarios_existentes
                ]
    except Exception as e:
        flash(f"Erro ao buscar eventos: {str(e)}", "danger")
    
    # Processar o formulário quando enviado via POST
    if request.method == 'POST':
        try:
            # Determinar o tipo de ação
            acao = request.form.get('acao')
            
            if acao == 'adicionar':
                # Obter dados do formulário para adicionar novo horário
                evento_id = request.form.get('evento_id')
                data = request.form.get('data')
                horario_inicio = request.form.get('horario_inicio')
                horario_fim = request.form.get('horario_fim')
                capacidade = request.form.get('capacidade')
                
                # Validar dados obrigatórios
                if not evento_id or not data or not horario_inicio or not horario_fim or not capacidade:
                    form_erro = "Preencha todos os campos obrigatórios."
                    flash(form_erro, "danger")
                else:
                    # Converte string para data e hora
                    data_obj = datetime.strptime(data, '%Y-%m-%d').date()
                    horario_inicio_obj = datetime.strptime(horario_inicio, '%H:%M').time()
                    horario_fim_obj = datetime.strptime(horario_fim, '%H:%M').time()
                    capacidade_int = int(capacidade)
                    
                    # Criar novo horário de visitação
                    novo_horario = HorarioVisitacao(
                        evento_id=evento_id,
                        data=data_obj,
                        horario_inicio=horario_inicio_obj,
                        horario_fim=horario_fim_obj,
                        capacidade_total=capacidade_int,
                        vagas_disponiveis=capacidade_int
                    )
                    
                    db.session.add(novo_horario)
                    db.session.commit()
                    
                    flash(f"Horário adicionado com sucesso para o dia {data} das {horario_inicio} às {horario_fim}!", "success")
                    
                    # Redirecionar para a mesma página com o evento selecionado
                    return redirect(url_for('routes.configurar_horarios_agendamento', evento_id=evento_id))
            
            elif acao == 'excluir':
                # Obter ID do horário a ser excluído
                horario_id = request.form.get('horario_id')
                evento_id = request.form.get('evento_id')
                
                if not horario_id:
                    flash("ID do horário não fornecido.", "danger")
                else:
                    # Verificar se existem agendamentos para este horário
                    horario = HorarioVisitacao.query.get(horario_id)
                    
                    if horario:
                        # Verificar se há agendamentos para este horário
                        agendamentos = AgendamentoVisita.query.filter_by(horario_id=horario_id).first()
                        
                        if agendamentos:
                            flash("Não é possível excluir um horário que possui agendamentos.", "danger")
                        else:
                            db.session.delete(horario)
                            db.session.commit()
                            flash("Horário excluído com sucesso!", "success")
                    else:
                        flash("Horário não encontrado.", "danger")
                    
                    # Redirecionar para a mesma página com o evento selecionado
                    return redirect(url_for('routes.configurar_horarios_agendamento', evento_id=evento_id))
            
            elif acao == 'adicionar_periodo':
                # Obter dados do formulário para adicionar vários horários em um período
                evento_id = request.form.get('evento_id')
                data_inicio = request.form.get('data_inicio')
                data_fim = request.form.get('data_fim')
                dias_semana = request.form.getlist('dias_semana')
                horario_inicio = request.form.get('horario_inicio')
                horario_fim = request.form.get('horario_fim')
                capacidade = request.form.get('capacidade')
                
                # Validar dados obrigatórios
                if not evento_id or not data_inicio or not data_fim or not dias_semana or not horario_inicio or not horario_fim or not capacidade:
                    form_erro = "Preencha todos os campos obrigatórios."
                    flash(form_erro, "danger")
                else:
                    # Converter strings para objetos de data
                    data_inicio_obj = datetime.strptime(data_inicio, '%Y-%m-%d').date()
                    data_fim_obj = datetime.strptime(data_fim, '%Y-%m-%d').date()
                    horario_inicio_obj = datetime.strptime(horario_inicio, '%H:%M').time()
                    horario_fim_obj = datetime.strptime(horario_fim, '%H:%M').time()
                    capacidade_int = int(capacidade)
                    
                    # Converter dias da semana para inteiros (0=Segunda, 6=Domingo)
                    dias = [int(dia) for dia in dias_semana]
                    
                    # Criar horários para todas as datas no período que correspondem aos dias da semana selecionados
                    delta = data_fim_obj - data_inicio_obj
                    horarios_criados = 0
                    
                    for i in range(delta.days + 1):
                        data_atual = data_inicio_obj + timedelta(days=i)
                        # weekday() retorna 0 para segunda e 6 para domingo
                        if data_atual.weekday() in dias:
                            # Verificar se já existe um horário para esta data e período
                            horario_existente = HorarioVisitacao.query.filter_by(
                                evento_id=evento_id,
                                data=data_atual,
                                horario_inicio=horario_inicio_obj,
                                horario_fim=horario_fim_obj
                            ).first()
                            
                            if not horario_existente:
                                novo_horario = HorarioVisitacao(
                                    evento_id=evento_id,
                                    data=data_atual,
                                    horario_inicio=horario_inicio_obj,
                                    horario_fim=horario_fim_obj,
                                    capacidade_total=capacidade_int,
                                    vagas_disponiveis=capacidade_int
                                )
                                
                                db.session.add(novo_horario)
                                horarios_criados += 1
                    
                    if horarios_criados > 0:
                        db.session.commit()
                        flash(f"{horarios_criados} horários configurados com sucesso no período de {data_inicio} a {data_fim}!", "success")
                    else:
                        flash("Nenhum horário novo foi criado. Verifique se já existem horários para as datas selecionadas.", "warning")
                    
                    # Redirecionar para a mesma página com o evento selecionado
                    return redirect(url_for('routes.configurar_horarios_agendamento', evento_id=evento_id))
                
        except Exception as e:
            form_erro = f"Erro ao processar o formulário: {str(e)}"
            flash(form_erro, "danger")
            db.session.rollback()
    
    # Adicione esta linha para verificar se a função editar_horario existe
    has_editar_horario = hasattr(routes, 'editar_horario')
    
    # Obter configuração de agendamento se existir
    configuracao = None
    if evento_selecionado:
        configuracao = ConfiguracaoAgendamento.query.filter_by(
            cliente_id=current_user.id,
            evento_id=evento_selecionado.id
        ).first()
    
    # Renderizar o template com o formulário
    return render_template('configurar_horarios_agendamento.html', 
                          eventos=eventos,
                          evento_selecionado=evento_selecionado,
                          horarios_existentes=horarios_existentes,
                          form_erro=form_erro,
                          has_editar_horario=has_editar_horario,
                          configuracao=configuracao)

@routes.route('/criar-evento-agendamento', methods=['GET', 'POST'])
@login_required
def criar_evento_agendamento():
    """
    Rota para criação de um novo evento para agendamentos.
    """
    # Verificação de permissão
    if current_user.tipo != 'cliente':
        flash('Acesso negado!', 'danger')
        return redirect(url_for('routes.dashboard_cliente'))
    
    # Inicialização de variáveis
    form_erro = None
    
    # Processar o formulário quando enviado via POST
    if request.method == 'POST':
        try:
            # Obter dados do formulário
            nome = request.form.get('nome')
            descricao = request.form.get('descricao')
            programacao = request.form.get('programacao')
            localizacao = request.form.get('local')
            link_mapa = request.form.get('link_mapa')
            data_inicio = request.form.get('data_inicio')
            data_fim = request.form.get('data_fim')
            hora_inicio = request.form.get('hora_inicio')
            hora_fim = request.form.get('hora_fim')
            capacidade_padrao = request.form.get('capacidade_padrao')
            requer_aprovacao = 'requer_aprovacao' in request.form
            publico = 'publico' in request.form
            
            # Validar dados obrigatórios
            if not nome or not data_inicio or not data_fim or not capacidade_padrao:
                form_erro = "Preencha todos os campos obrigatórios."
                flash(form_erro, "danger")
            elif data_fim < data_inicio:
                form_erro = "A data de fim deve ser posterior à data de início."
                flash(form_erro, "danger")
            else:
                # Processar upload de banner, se houver
                banner = request.files.get('banner')
                banner_url = None
                
                if banner and banner.filename:
                    filename = secure_filename(banner.filename)
                    caminho_banner = os.path.join('static/banners', filename)
                    os.makedirs(os.path.dirname(caminho_banner), exist_ok=True)
                    banner.save(caminho_banner)
                    banner_url = url_for('static', filename=f'banners/{filename}', _external=True)
                
                # Converter strings para objetos de data/hora
                data_inicio_obj = datetime.strptime(data_inicio, '%Y-%m-%d') if data_inicio else None
                data_fim_obj = datetime.strptime(data_fim, '%Y-%m-%d') if data_fim else None
                
                # Processar hora (se necessário)
                from datetime import time
                hora_inicio_obj = time.fromisoformat(hora_inicio) if hora_inicio else None
                hora_fim_obj = time.fromisoformat(hora_fim) if hora_fim else None
                
                # Criar o objeto Evento
                novo_evento = Evento(
                    cliente_id=current_user.id,
                    nome=nome,
                    descricao=descricao,
                    programacao=programacao,
                    localizacao=localizacao,
                    link_mapa=link_mapa,
                    banner_url=banner_url,
                    data_inicio=data_inicio_obj,
                    data_fim=data_fim_obj,
                    hora_inicio=hora_inicio_obj,
                    hora_fim=hora_fim_obj,
                    capacidade_padrao=int(capacidade_padrao),
                    requer_aprovacao=requer_aprovacao,
                    publico=publico
                )
                
                try:
                    db.session.add(novo_evento)
                    db.session.flush()  # Obter o ID do evento antes de criar tipos de inscrição
                    
                    # Processar tipos de inscrição se o cliente tiver pagamento habilitado
                    if current_user.habilita_pagamento:
                        inscricao_gratuita = (request.form.get('inscricao_gratuita') == 'on')
                        novo_evento.inscricao_gratuita = inscricao_gratuita
                        
                        # Adicionar tipos de inscrição se não for gratuito
                        if not inscricao_gratuita:
                            nomes_tipos = request.form.getlist('nome_tipo[]')
                            precos = request.form.getlist('preco_tipo[]')
                            
                            if not nomes_tipos or not precos:
                                raise ValueError("Tipos de inscrição e preços são obrigatórios quando o evento não é gratuito.")
                            
                            for nome, preco in zip(nomes_tipos, precos):
                                if nome and preco:
                                    novo_tipo = EventoInscricaoTipo(
                                        evento_id=novo_evento.id,
                                        nome=nome,
                                        preco=float(preco)
                                    )
                                    db.session.add(novo_tipo)
                    
                    db.session.commit()
                    flash(f"Evento '{nome}' criado com sucesso! Você pode agora configurar os horários.", "success")
                    return redirect(url_for('routes.configurar_horarios_agendamento', evento_id=novo_evento.id))
                
                except Exception as e:
                    db.session.rollback()
                    form_erro = f"Erro ao salvar evento: {str(e)}"
                    flash(form_erro, "danger")
        
        except Exception as e:
            form_erro = f"Erro ao processar o formulário: {str(e)}"
            flash(form_erro, "danger")
    
    # Renderizar o template com o formulário
    return render_template('criar_evento_agendamento.html', form_erro=form_erro)
    
@routes.route('/importar-agendamentos', methods=['GET', 'POST'])
@login_required
def importar_agendamentos():
    """
    Rota para importação de agendamentos a partir de um arquivo CSV ou Excel.
    """
    # Inicialização de variáveis
    form_erro = None
    eventos = []
    importacao_resultado = None
    
    # Buscar eventos disponíveis do cliente atual
    try:
        eventos = Evento.query.filter_by(cliente_id=current_user.id).all()
    except Exception as e:
        flash(f"Erro ao buscar eventos: {str(e)}", "danger")
    
    # Processar o formulário quando enviado via POST
    if request.method == 'POST':
        try:
            # Verificar se um arquivo foi enviado
            if 'arquivo' not in request.files:
                form_erro = "Nenhum arquivo enviado."
                flash(form_erro, "danger")
            else:
                arquivo = request.files['arquivo']
                
                # Verificar se o arquivo tem nome
                if arquivo.filename == '':
                    form_erro = "Nenhum arquivo selecionado."
                    flash(form_erro, "danger")
                else:
                    # Verificar a extensão do arquivo
                    if not arquivo.filename.endswith(('.csv', '.xlsx', '.xls')):
                        form_erro = "Formato de arquivo não suportado. Use CSV ou Excel (.xlsx, .xls)."
                        flash(form_erro, "danger")
                    else:
                        # Obter o evento selecionado
                        evento_id = request.form.get('evento_id')
                        if not evento_id:
                            form_erro = "Selecione um evento para importar os agendamentos."
                            flash(form_erro, "danger")
                        else:
                            # Processar o arquivo (CSV ou Excel)
                            # Aqui teríamos a lógica para ler o arquivo e importar os agendamentos
                            # Como não temos acesso ao modelo real de Agendamento, usaremos dados simulados
                            
                            # Simular resultados da importação
                            importacao_resultado = {
                                'total_registros': 15,
                                'importados': 12,
                                'ignorados': 3,
                                'detalhes': [
                                    {'linha': 2, 'status': 'sucesso', 'mensagem': 'Importado com sucesso'},
                                    {'linha': 5, 'status': 'sucesso', 'mensagem': 'Importado com sucesso'},
                                    {'linha': 8, 'status': 'erro', 'mensagem': 'Data inválida'},
                                    {'linha': 10, 'status': 'erro', 'mensagem': 'Horário não disponível'},
                                    {'linha': 12, 'status': 'erro', 'mensagem': 'Capacidade excedida'}
                                ]
                            }
                            
                            flash(f"Importação concluída! {importacao_resultado['importados']} agendamentos importados, {importacao_resultado['ignorados']} ignorados.", "success")
                
        except Exception as e:
            form_erro = f"Erro ao processar a importação: {str(e)}"
            flash(form_erro, "danger")
    
    # Renderizar o template com o formulário
    return render_template('importar_agendamentos.html', 
                          eventos=eventos,
                          form_erro=form_erro,
                          importacao_resultado=importacao_resultado)
    
@routes.route('/download-modelo-importacao')
@login_required
def download_modelo_importacao():
    """
    Rota para baixar um modelo de planilha para importação de agendamentos.
    """
    try:
        # Aqui você criaria um arquivo Excel ou CSV com as colunas necessárias
        # Como exemplo, vamos apenas retornar uma resposta simulando o download
        
        # Em uma implementação real, você usaria bibliotecas como xlsxwriter ou pandas
        # para criar o arquivo e depois enviá-lo como resposta
        
        # Exemplo simplificado (apenas para demonstração):
        from io import BytesIO
        import csv
        
        # Criar um buffer de memória para o CSV
        output = BytesIO()
        writer = csv.writer(output)
        
        # Escrever o cabeçalho
        writer.writerow(['Data', 'Horário', 'Escola/Instituição', 'Nome do Responsável', 
                         'E-mail', 'Telefone', 'Turma', 'Quantidade de Alunos'])
        
        # Escrever algumas linhas de exemplo
        writer.writerow(['20/03/2025', '09:00', 'Escola Exemplo', 'João Silva', 
                         'joao.silva@email.com', '(11) 98765-4321', '5º Ano A', '25'])
        writer.writerow(['21/03/2025', '14:30', 'Colégio Modelo', 'Maria Oliveira', 
                         'maria.oliveira@email.com', '(11) 91234-5678', '8º Ano B', '30'])
        
        # Preparar a resposta
        output.seek(0)
        
        return send_file(
            output,
            mimetype='text/csv',
            as_attachment=True,
            download_name='modelo_importacao_agendamentos.csv'
        )
        
    except Exception as e:
        flash(f"Erro ao gerar o modelo: {str(e)}", "danger")
        return redirect(url_for('routes.importar_agendamentos'))


@routes.route('/exportar-log-importacao')
@login_required
def exportar_log_importacao():
    """
    Rota para exportar o log detalhado da última importação.
    """
    try:
        # Aqui você buscaria os logs de importação do banco de dados
        # Como exemplo, vamos apenas retornar um arquivo CSV com dados simulados
        
        from io import BytesIO
        import csv
        
        # Criar um buffer de memória para o CSV
        output = BytesIO()
        writer = csv.writer(output)
        
        # Escrever o cabeçalho
        writer.writerow(['Linha', 'Status', 'Mensagem', 'Dados Originais'])
        
        # Escrever algumas linhas de exemplo
        writer.writerow(['1', 'Cabeçalho', 'Ignorado', 'Data,Horário,Escola,...'])
        writer.writerow(['2', 'Sucesso', 'Importado com sucesso', '20/03/2025,09:00,Escola Exemplo,...'])
        writer.writerow(['3', 'Sucesso', 'Importado com sucesso', '20/03/2025,14:00,Escola Modelo,...'])
        writer.writerow(['4', 'Erro', 'Data inválida', '32/03/2025,10:00,Escola Inválida,...'])
        writer.writerow(['5', 'Erro', 'Horário não disponível', '21/03/2025,18:00,Escola Teste,...'])
        writer.writerow(['6', 'Erro', 'Capacidade excedida', '22/03/2025,09:00,Escola Grande,...'])
        
        # Preparar a resposta
        output.seek(0)
        
        return send_file(
            output,
            mimetype='text/csv',
            as_attachment=True,
            download_name='log_importacao_agendamentos.csv'
        )
        
    except Exception as e:
        flash(f"Erro ao gerar o log: {str(e)}", "danger")
        return redirect(url_for('routes.importar_agendamentos'))
    
@routes.route('/api/toggle-agendamento-publico', methods=['POST'])
@login_required
def toggle_agendamento_publico():
    """
    Alternar o status de agendamento público (se visitantes podem agendar pelo site).
    Esta rota é chamada via AJAX a partir da página de configurações.
    """
    try:
        # Buscar configuração atual de agendamento do cliente
        config_agendamento = None
        
        # Verificar se existe um modelo ConfigAgendamento
        # Esta verificação ajuda a evitar erros se o modelo não existir
        if 'ConfigAgendamento' in globals():
            config_agendamento = ConfigAgendamento.query.filter_by(cliente_id=current_user.id).first()
        
            # Se não existir, criar uma nova configuração
            if not config_agendamento:
                config_agendamento = ConfigAgendamento(
                    cliente_id=current_user.id,
                    agendamento_publico=False
                )
                db.session.add(config_agendamento)
            
            # Alternar o status
            config_agendamento.agendamento_publico = not config_agendamento.agendamento_publico
            
            # Salvar alterações no banco de dados
            db.session.commit()
            
            # Retornar o novo status
            return jsonify({
                'success': True,
                'value': config_agendamento.agendamento_publico
            })
        else:
            # Se o modelo não existir, simule a operação para fins de demonstração
            # Em um ambiente de produção, você implementaria isso com seu modelo real
            return jsonify({
                'success': True,
                'value': True,  # Valor simulado
                'message': 'Operação simulada: modelo ConfigAgendamento não encontrado'
            })
            
    except Exception as e:
        # Log de erro para depuração
        print(f"Erro ao alternar status de agendamento público: {str(e)}")
        
        # Retornar erro para a aplicação
        return jsonify({
            'success': False,
            'message': f"Erro ao alternar status: {str(e)}"
        }), 500

@routes.route('/api/toggle-aprovacao-manual', methods=['POST'])
@login_required
def toggle_aprovacao_manual():
    """
    Alternar o status de aprovação manual de agendamentos.
    Quando ativado, os agendamentos novos ficam com status pendente até aprovação.
    Esta rota é chamada via AJAX a partir da página de configurações.
    """
    try:
        # Buscar configuração atual de agendamento do cliente
        config_agendamento = None
        
        # Verificar se existe um modelo ConfigAgendamento
        # Esta verificação ajuda a evitar erros se o modelo não existir
        if 'ConfigAgendamento' in globals():
            config_agendamento = ConfigAgendamento.query.filter_by(cliente_id=current_user.id).first()
        
            # Se não existir, criar uma nova configuração
            if not config_agendamento:
                config_agendamento = ConfigAgendamento(
                    cliente_id=current_user.id,
                    aprovacao_manual=False
                )
                db.session.add(config_agendamento)
            
            # Alternar o status
            config_agendamento.aprovacao_manual = not config_agendamento.aprovacao_manual
            
            # Salvar alterações no banco de dados
            db.session.commit()
            
            # Retornar o novo status
            return jsonify({
                'success': True,
                'value': config_agendamento.aprovacao_manual
            })
        else:
            # Se o modelo não existir, simule a operação para fins de demonstração
            # Em um ambiente de produção, você implementaria isso com seu modelo real
            return jsonify({
                'success': True,
                'value': True,  # Valor simulado
                'message': 'Operação simulada: modelo ConfigAgendamento não encontrado'
            })
            
    except Exception as e:
        # Log de erro para depuração
        print(f"Erro ao alternar status de aprovação manual: {str(e)}")
        
        # Retornar erro para a aplicação
        return jsonify({
            'success': False,
            'message': f"Erro ao alternar status: {str(e)}"
        }), 500

@routes.route('/api/toggle-limite-capacidade', methods=['POST'])
@login_required
def toggle_limite_capacidade():
    """
    Alternar a aplicação do limite de capacidade para agendamentos.
    Quando ativado, o sistema verifica se há vagas disponíveis antes de permitir o agendamento.
    Esta rota é chamada via AJAX a partir da página de configurações.
    """
    try:
        # Buscar configuração atual de agendamento do cliente
        config_agendamento = None
        
        # Verificar se existe um modelo ConfigAgendamento
        # Esta verificação ajuda a evitar erros se o modelo não existir
        if 'ConfigAgendamento' in globals():
            config_agendamento = ConfigAgendamento.query.filter_by(cliente_id=current_user.id).first()
        
            # Se não existir, criar uma nova configuração
            if not config_agendamento:
                config_agendamento = ConfigAgendamento(
                    cliente_id=current_user.id,
                    aplicar_limite_capacidade=True  # O padrão é aplicar o limite
                )
                db.session.add(config_agendamento)
            
            # Alternar o status
            config_agendamento.aplicar_limite_capacidade = not config_agendamento.aplicar_limite_capacidade
            
            # Salvar alterações no banco de dados
            db.session.commit()
            
            # Retornar o novo status
            return jsonify({
                'success': True,
                'value': config_agendamento.aplicar_limite_capacidade
            })
        else:
            # Se o modelo não existir, simule a operação para fins de demonstração
            # Em um ambiente de produção, você implementaria isso com seu modelo real
            return jsonify({
                'success': True,
                'value': True,  # Valor simulado
                'message': 'Operação simulada: modelo ConfigAgendamento não encontrado'
            })
            
    except Exception as e:
        # Log de erro para depuração
        print(f"Erro ao alternar status de limite de capacidade: {str(e)}")
        
        # Retornar erro para a aplicação
        return jsonify({
            'success': False,
            'message': f"Erro ao alternar status: {str(e)}"
        }), 500

@routes.route('/salvar-config-agendamento', methods=['POST'])
@login_required
def salvar_config_agendamento():
    """
    Salvar as configurações gerais do sistema de agendamentos.
    Esta rota processa o formulário enviado pela página de configurações.
    """
    try:
        # Obter dados do formulário
        capacidade_maxima = request.form.get('capacidade_maxima', type=int)
        dias_antecedencia = request.form.get('dias_antecedencia', type=int)
        
        # Validar dados
        if not capacidade_maxima or capacidade_maxima < 1:
            flash("A capacidade máxima deve ser um número positivo.", "danger")
            return redirect(url_for('routes.dashboard_agendamentos', _anchor='configuracoes'))
            
        if not dias_antecedencia or dias_antecedencia < 1:
            flash("Os dias de antecedência devem ser um número positivo.", "danger")
            return redirect(url_for('routes.dashboard_agendamentos', _anchor='configuracoes'))
            
        # Buscar configuração atual de agendamento do cliente
        config_agendamento = None
        
        # Verificar se existe um modelo ConfigAgendamento
        if 'ConfigAgendamento' in globals():
            config_agendamento = ConfigAgendamento.query.filter_by(cliente_id=current_user.id).first()
        
            # Se não existir, criar uma nova configuração
            if not config_agendamento:
                config_agendamento = ConfigAgendamento(
                    cliente_id=current_user.id,
                    capacidade_maxima=30,  # Valor padrão
                    dias_antecedencia=30,  # Valor padrão
                    agendamento_publico=True,
                    aprovacao_manual=False,
                    aplicar_limite_capacidade=True
                )
                db.session.add(config_agendamento)
            
            # Atualizar configurações
            config_agendamento.capacidade_maxima = capacidade_maxima
            config_agendamento.dias_antecedencia = dias_antecedencia
            
            # Salvar alterações no banco de dados
            db.session.commit()
            
            flash("Configurações de agendamento salvas com sucesso!", "success")
        else:
            # Se o modelo não existir, apenas exibir mensagem de sucesso simulado
            flash("Configurações salvas com sucesso! (Modo de demonstração)", "success")
        
        # Obter valores opcionais adicionais
        # Pode-se adicionar mais campos conforme necessário
        enviar_lembretes = 'enviar_lembretes' in request.form
        periodo_lembrete = request.form.get('periodo_lembrete', type=int)
        template_email = request.form.get('template_email')
        
        # Se você tiver campos adicionais, o código para salvá-los seria inserido aqui
            
        # Redirecionar de volta para a página de configurações
        return redirect(url_for('routes.dashboard_agendamentos', _anchor='configuracoes'))
            
    except Exception as e:
        # Log de erro para depuração
        print(f"Erro ao salvar configurações de agendamento: {str(e)}")
        
        # Notificar o usuário
        flash(f"Erro ao salvar configurações: {str(e)}", "danger")
        
        # Redirecionar de volta para a página de configurações
        return redirect(url_for('routes.dashboard_agendamentos', _anchor='configuracoes'))

@routes.route('/criar-periodo-agendamento', methods=['GET', 'POST'])
@login_required
def criar_periodo_agendamento():
    """
    Rota para criação de período de agendamento.
    Um período define um intervalo de datas em que o agendamento está disponível.
    """
    # Inicialização de variáveis
    form_erro = None
    eventos = []
    
    # Buscar eventos disponíveis do cliente atual
    try:
        eventos = Evento.query.filter_by(cliente_id=current_user.id).all()
    except Exception as e:
        flash(f"Erro ao buscar eventos: {str(e)}", "danger")
    
    # Processar o formulário quando enviado via POST
    if request.method == 'POST':
        try:
            # Obter dados do formulário
            evento_id = request.form.get('evento_id')
            data_inicio = request.form.get('data_inicio')
            data_fim = request.form.get('data_fim')
            hora_inicio = request.form.get('hora_inicio')
            hora_fim = request.form.get('hora_fim')
            intervalo_min = request.form.get('intervalo_min', type=int)
            capacidade = request.form.get('capacidade', type=int)
            dias_semana = request.form.getlist('dias_semana')  # Lista de dias selecionados (0-6)
            
            # Validar dados obrigatórios
            if not evento_id or not data_inicio or not data_fim or not hora_inicio or not hora_fim or not capacidade:
                form_erro = "Preencha todos os campos obrigatórios."
                flash(form_erro, "danger")
            elif not dias_semana:
                form_erro = "Selecione pelo menos um dia da semana."
                flash(form_erro, "danger")
            else:
                # Verificar se data de fim é posterior à data de início
                if data_fim < data_inicio:
                    form_erro = "A data de fim deve ser posterior à data de início."
                    flash(form_erro, "danger")
                else:
                    # Aqui você adicionaria o código para criar um novo período de agendamento
                    # e os horários relacionados baseado nos dias da semana selecionados
                    
                    # Como não sabemos a estrutura exata do seu modelo,
                    # vamos apenas exibir uma mensagem de sucesso simulada
                    
                    # Converter lista de strings para dias da semana
                    dias_nomes = {
                        '0': 'Domingo',
                        '1': 'Segunda',
                        '2': 'Terça',
                        '3': 'Quarta',
                        '4': 'Quinta',
                        '5': 'Sexta',
                        '6': 'Sábado'
                    }
                    dias_selecionados = [dias_nomes.get(dia, '') for dia in dias_semana if dia in dias_nomes]
                    dias_texto = ", ".join(dias_selecionados)
                    
                    flash(f"Período de agendamento criado com sucesso! Horários configurados para {dias_texto} das {hora_inicio} às {hora_fim}.", "success")
                    
                    # Redirecionar para a página de configuração de horários
                    return redirect(url_for('routes.configurar_horarios_agendamento', evento_id=evento_id))
                
        except Exception as e:
            form_erro = f"Erro ao processar o formulário: {str(e)}"
            flash(form_erro, "danger")
    
    # Renderizar o template com o formulário
    return render_template('criar_periodo_agendamento.html', 
                          eventos=eventos,
                          form_erro=form_erro)

@routes.route('/excluir-todos-agendamentos', methods=['POST'])
@login_required
def excluir_todos_agendamentos():
    """
    Rota para excluir todos os agendamentos do cliente atual.
    Esta é uma operação perigosa e irreversível, por isso requer uma confirmação
    e é acessível apenas via método POST.
    """
    try:
        # Verificar se o cliente tem permissão para excluir agendamentos
        if not current_user.is_admin and not current_user.is_cliente:
            flash("Você não tem permissão para realizar esta operação.", "danger")
            return redirect(url_for('routes.dashboard_agendamentos'))
        
        # Buscar todos os eventos do cliente
        eventos = Evento.query.filter_by(cliente_id=current_user.id).all()
        
        # Contador para registrar quantos agendamentos foram excluídos
        total_excluidos = 0
        
        # Verificar se existe um modelo Agendamento
        # Esta verificação ajuda a evitar erros se o modelo não existir
        if 'Agendamento' in globals():
            # Para cada evento, buscar todos os horários e seus agendamentos
            for evento in eventos:
                # Se você tiver um relacionamento direto entre Evento e Horario
                horarios = Horario.query.filter_by(evento_id=evento.id).all()
                
                for horario in horarios:
                    # Buscar agendamentos deste horário
                    agendamentos = Agendamento.query.filter_by(horario_id=horario.id).all()
                    
                    # Excluir cada agendamento
                    for agendamento in agendamentos:
                        db.session.delete(agendamento)
                        total_excluidos += 1
            
            # Commit das alterações ao banco de dados
            db.session.commit()
            
            # Notificar o usuário do sucesso da operação
            flash(f"Todos os agendamentos foram excluídos com sucesso. Total de {total_excluidos} agendamentos removidos.", "success")
        else:
            # Se o modelo não existir, simule a operação para fins de demonstração
            flash("Operação simulada: Todos os agendamentos foram excluídos com sucesso.", "success")
        
        # Redirecionar para o dashboard de agendamentos
        return redirect(url_for('routes.dashboard_agendamentos'))
            
    except Exception as e:
        # Em caso de erro, fazer rollback das alterações
        if 'db' in globals() and hasattr(db, 'session'):
            db.session.rollback()
        
        # Log do erro para depuração
        print(f"Erro ao excluir agendamentos: {str(e)}")
        
        # Notificar o usuário do erro
        flash(f"Erro ao excluir agendamentos: {str(e)}", "danger")
        
        # Redirecionar para o dashboard
        return redirect(url_for('routes.dashboard_agendamentos'))
    
@routes.route('/resetar-configuracoes-agendamento', methods=['POST'])
@login_required
def resetar_configuracoes_agendamento():
    """
    Rota para resetar as configurações de agendamento para valores padrão.
    Esta operação restaura as configurações originais, mas não afeta os agendamentos existentes.
    """
    try:
        # Verificar se o cliente tem permissão para resetar configurações
        if not current_user.is_admin and not current_user.is_cliente:
            flash("Você não tem permissão para realizar esta operação.", "danger")
            return redirect(url_for('routes.dashboard_agendamentos'))
        
        # Buscar configuração atual de agendamento do cliente
        config_agendamento = None
        
        # Verificar se existe um modelo ConfigAgendamento
        if 'ConfigAgendamento' in globals():
            config_agendamento = ConfigAgendamento.query.filter_by(cliente_id=current_user.id).first()
        
            # Se existir, resetar para os valores padrão
            if config_agendamento:
                config_agendamento.capacidade_maxima = 30
                config_agendamento.dias_antecedencia = 30
                config_agendamento.agendamento_publico = True
                config_agendamento.aprovacao_manual = False
                config_agendamento.aplicar_limite_capacidade = True
                
                # Adicionar outras configurações padrão conforme necessário
                
                # Salvar alterações no banco de dados
                db.session.commit()
                
                flash("Configurações de agendamento resetadas para os valores padrão com sucesso!", "success")
            else:
                # Se não existir, criar uma nova configuração com valores padrão
                config_agendamento = ConfigAgendamento(
                    cliente_id=current_user.id,
                    capacidade_maxima=30,
                    dias_antecedencia=30,
                    agendamento_publico=True,
                    aprovacao_manual=False,
                    aplicar_limite_capacidade=True
                )
                
                db.session.add(config_agendamento)
                db.session.commit()
                
                flash("Configurações de agendamento criadas com valores padrão!", "success")
        else:
            # Se o modelo não existir, simule a operação para fins de demonstração
            flash("Operação simulada: Configurações resetadas para valores padrão.", "success")
        
        # Redirecionar para o dashboard de agendamentos, aba configurações
        return redirect(url_for('routes.dashboard_agendamentos', _anchor='configuracoes'))
            
    except Exception as e:
        # Em caso de erro, fazer rollback das alterações
        if 'db' in globals() and hasattr(db, 'session'):
            db.session.rollback()
        
        # Log do erro para depuração
        print(f"Erro ao resetar configurações: {str(e)}")
        
        # Notificar o usuário do erro
        flash(f"Erro ao resetar configurações: {str(e)}", "danger")
        
        # Redirecionar para o dashboard
        return redirect(url_for('routes.dashboard_agendamentos', _anchor='configuracoes'))
    
@routes.route('/exportar-agendamentos')
@login_required
def exportar_agendamentos():
    """
    Rota para exportar agendamentos do cliente atual em formato CSV ou Excel.
    Recebe parâmetros opcionais por query string para filtrar os dados.
    """
    try:
        # Obter parâmetros de filtro
        formato = request.args.get('formato', 'csv')  # csv ou excel
        evento_id = request.args.get('evento_id')
        data_inicio = request.args.get('data_inicio')
        data_fim = request.args.get('data_fim')
        status = request.args.get('status')  # todos, confirmados, realizados, cancelados
        
        # Validar o formato solicitado
        if formato not in ['csv', 'excel']:
            flash("Formato de exportação inválido. Use 'csv' ou 'excel'.", "danger")
            return redirect(url_for('routes.dashboard_agendamentos'))
        
        # Buscar dados para exportação
        # Em uma implementação real, você buscaria os agendamentos do banco de dados
        # baseado nos filtros fornecidos
        
        # Como não sabemos a estrutura exata do seu modelo,
        # vamos criar dados simulados para demonstração
        
        # Dados simulados para exportação
        agendamentos_dados = [
            {
                'id': 1,
                'data': '2025-03-20',
                'horario': '09:00 - 11:00',
                'evento': 'Feira de Ciências 2025',
                'escola': 'Escola Modelo',
                'responsavel': 'João Silva',
                'email': 'joao.silva@email.com',
                'telefone': '(11) 98765-4321',
                'turma': '5º Ano A',
                'alunos': 25,
                'status': 'confirmado',
                'data_criacao': '2025-02-15 14:30:22'
            },
            {
                'id': 2,
                'data': '2025-03-21',
                'horario': '14:00 - 16:00',
                'evento': 'Feira de Ciências 2025',
                'escola': 'Colégio Exemplo',
                'responsavel': 'Maria Oliveira',
                'email': 'maria.oliveira@email.com',
                'telefone': '(11) 91234-5678',
                'turma': '8º Ano B',
                'alunos': 30,
                'status': 'confirmado',
                'data_criacao': '2025-02-16 10:15:45'
            },
            {
                'id': 3,
                'data': '2025-03-22',
                'horario': '09:00 - 11:00',
                'evento': 'Feira de Ciências 2025',
                'escola': 'Instituto Educacional',
                'responsavel': 'Carlos Santos',
                'email': 'carlos.santos@email.com',
                'telefone': '(11) 95555-1234',
                'turma': '2º Ano EM',
                'alunos': 35,
                'status': 'cancelado',
                'data_criacao': '2025-02-17 09:22:10'
            }
        ]
        
        # Exportar para CSV
        if formato == 'csv':
            from io import StringIO
            import csv
            
            # Criar buffer de memória para o CSV
            output = StringIO()
            writer = csv.writer(output)
            
            # Escrever cabeçalho
            writer.writerow(['ID', 'Data', 'Horário', 'Evento', 'Escola', 'Responsável', 'Email', 
                            'Telefone', 'Turma', 'Alunos', 'Status', 'Data de Criação'])
            
            # Escrever linhas de dados
            for agendamento in agendamentos_dados:
                writer.writerow([
                    agendamento['id'],
                    agendamento['data'],
                    agendamento['horario'],
                    agendamento['evento'],
                    agendamento['escola'],
                    agendamento['responsavel'],
                    agendamento['email'],
                    agendamento['telefone'],
                    agendamento['turma'],
                    agendamento['alunos'],
                    agendamento['status'],
                    agendamento['data_criacao']
                ])
            
            # Preparar a resposta
            output.seek(0)
            
            return Response(
                output.getvalue(),
                mimetype='text/csv',
                headers={
                    'Content-Disposition': 'attachment; filename=agendamentos.csv',
                    'Content-Type': 'text/csv; charset=utf-8'
                }
            )
        
        # Exportar para Excel
        elif formato == 'excel':
            try:
                # Tentar importar a biblioteca xlsxwriter
                import xlsxwriter
                from io import BytesIO
                
                # Criar buffer de memória para o Excel
                output = BytesIO()
                
                # Criar workbook e adicionar uma planilha
                workbook = xlsxwriter.Workbook(output)
                worksheet = workbook.add_worksheet('Agendamentos')
                
                # Formatar cabeçalho
                header_format = workbook.add_format({
                    'bold': True,
                    'bg_color': '#4B5563',
                    'color': 'white',
                    'border': 1
                })
                
                # Formatar células normais
                cell_format = workbook.add_format({
                    'border': 1
                })
                
                # Formatar células de status
                status_formats = {
                    'confirmado': workbook.add_format({
                        'border': 1,
                        'bg_color': '#DBEAFE'  # Azul claro
                    }),
                    'realizado': workbook.add_format({
                        'border': 1,
                        'bg_color': '#DCFCE7'  # Verde claro
                    }),
                    'cancelado': workbook.add_format({
                        'border': 1,
                        'bg_color': '#FEE2E2'  # Vermelho claro
                    })
                }
                
                # Definir cabeçalho
                headers = ['ID', 'Data', 'Horário', 'Evento', 'Escola', 'Responsável', 'Email', 
                          'Telefone', 'Turma', 'Alunos', 'Status', 'Data de Criação']
                
                # Escrever cabeçalho
                for col, header in enumerate(headers):
                    worksheet.write(0, col, header, header_format)
                
                # Escrever dados
                for row, agendamento in enumerate(agendamentos_dados, start=1):
                    status = agendamento['status']
                    status_format = status_formats.get(status, cell_format)
                    
                    worksheet.write(row, 0, agendamento['id'], cell_format)
                    worksheet.write(row, 1, agendamento['data'], cell_format)
                    worksheet.write(row, 2, agendamento['horario'], cell_format)
                    worksheet.write(row, 3, agendamento['evento'], cell_format)
                    worksheet.write(row, 4, agendamento['escola'], cell_format)
                    worksheet.write(row, 5, agendamento['responsavel'], cell_format)
                    worksheet.write(row, 6, agendamento['email'], cell_format)
                    worksheet.write(row, 7, agendamento['telefone'], cell_format)
                    worksheet.write(row, 8, agendamento['turma'], cell_format)
                    worksheet.write(row, 9, agendamento['alunos'], cell_format)
                    worksheet.write(row, 10, agendamento['status'], status_format)
                    worksheet.write(row, 11, agendamento['data_criacao'], cell_format)
                
                # Ajustar largura das colunas automaticamente
                for col, header in enumerate(headers):
                    col_width = max(len(header), 12)  # Mínimo de 12 caracteres
                    worksheet.set_column(col, col, col_width)
                
                # Fechar o workbook
                workbook.close()
                
                # Preparar a resposta
                output.seek(0)
                
                return send_file(
                    output,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True,
                    download_name='agendamentos.xlsx'
                )
                
            except ImportError:
                # Caso a biblioteca xlsxwriter não esteja disponível, fallback para CSV
                flash("Biblioteca para exportação Excel não disponível. Exportando como CSV.", "warning")
                
                # Chamada recursiva usando formato CSV
                return exportar_agendamentos(formato='csv')
        
    except Exception as e:
        # Log de erro para depuração
        print(f"Erro ao exportar agendamentos: {str(e)}")
        
        # Notificar o usuário do erro
        flash(f"Erro ao exportar agendamentos: {str(e)}", "danger")
        
        # Redirecionar para o dashboard
        return redirect(url_for('routes.dashboard_agendamentos'))

@routes.route('/sala_visitacao/<int:sala_id>/excluir', methods=['POST'])
@login_required
def excluir_sala_visitacao(sala_id):
    """
    Excluir uma sala de visitação existente.
    
    Args:
        sala_id (int): ID da sala de visitação a ser excluída
        
    Returns:
        Redirecionamento para a página de listagem de salas
    """
    # Verificar permissões do usuário (apenas administradores)
    if current_user.perfil.lower() != 'administrador':
        flash('Você não tem permissão para excluir salas de visitação.', 'danger')
        return redirect(url_for('routes.index'))
    
    # Buscar a sala pelo ID
    sala = SalaVisitacao.query.get_or_404(sala_id)
    
    # Verificar se existem agendamentos relacionados
    agendamentos = Agendamento.query.filter_by(sala_id=sala_id).count()
    if agendamentos > 0:
        flash(f'Não é possível excluir esta sala pois existem {agendamentos} agendamentos associados a ela.', 'warning')
        return redirect(url_for('routes.salas_visitacao', evento_id=sala.evento_id))
    
    # Guardar o evento_id para usar no redirecionamento
    evento_id = sala.evento_id
    
    # Excluir a sala
    try:
        db.session.delete(sala)
        db.session.commit()
        flash(f'Sala "{sala.nome}" excluída com sucesso!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao excluir sala: {str(e)}', 'danger')
    
    # Redirecionar para a lista de salas do evento
    return redirect(url_for('routes.salas_visitacao', evento_id=evento_id))


@routes.route('/agendamentos/exportar/pdf', methods=['GET'])
@login_required
def exportar_agendamentos_pdf():
    """Exporta a lista de agendamentos em PDF"""
    # Implementar lógica para gerar PDF
    # Pode usar bibliotecas como ReportLab, WeasyPrint, etc.
    
    # Exemplo simples com ReportLab
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from io import BytesIO
    
    # Criar o buffer de memória para o PDF
    buffer = BytesIO()
    
    # Configurar o documento
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    
    # Título
    styles = getSampleStyleSheet()
    elements.append(Paragraph("Relatório de Agendamentos", styles['Title']))
    elements.append(Spacer(1, 12))
    
    # Filtrar agendamentos (usar mesma lógica de filtragem da view)
    # Adaptar conforme necessário
    if current_user.tipo == 'admin':
        agendamentos = AgendamentoVisita.query.all()
    else:
        agendamentos = AgendamentoVisita.query.filter_by(professor_id=current_user.id).all()
    
    # Dados da tabela
    data = [['ID', 'Escola', 'Professor', 'Data', 'Horário', 'Turma', 'Status']]
    
    for agendamento in agendamentos:
        data.append([
            str(agendamento.id),
            agendamento.escola_nome,
            agendamento.professor.nome,
            agendamento.horario.data.strftime('%d/%m/%Y'),
            f"{agendamento.horario.hora_inicio} - {agendamento.horario.hora_fim}",
            agendamento.turma,
            agendamento.status.capitalize()
        ])
    
    # Criar tabela
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    
    # Construir PDF
    doc.build(elements)
    
    # Preparar o response
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name="agendamentos.pdf",
        mimetype="application/pdf"
    )

@routes.route('/agendamentos/exportar/csv', methods=['GET'])
@login_required
def exportar_agendamentos_csv():
    """Exporta a lista de agendamentos em CSV"""
    import csv
    from io import StringIO
    
    # Filtrar agendamentos (usar mesma lógica de filtragem da view)
    if current_user.tipo == 'admin':
        agendamentos = AgendamentoVisita.query.all()
    else:
        agendamentos = AgendamentoVisita.query.filter_by(professor_id=current_user.id).all()
    
    # Criar o buffer de memória para o CSV
    output = StringIO()
    writer = csv.writer(output)
    
    # Escrever cabeçalho
    writer.writerow(['ID', 'Escola', 'Professor', 'Data', 'Horário', 'Turma', 'Nível de Ensino', 
                    'Quantidade de Alunos', 'Status', 'Data do Agendamento'])
    
    # Escrever dados
    for agendamento in agendamentos:
        writer.writerow([
            agendamento.id,
            agendamento.escola_nome,
            agendamento.professor.nome,
            agendamento.horario.data.strftime('%d/%m/%Y'),
            f"{agendamento.horario.hora_inicio} - {agendamento.horario.hora_fim}",
            agendamento.turma,
            agendamento.nivel_ensino,
            agendamento.quantidade_alunos,
            agendamento.status,
            agendamento.data_agendamento.strftime('%d/%m/%Y %H:%M')
        ])
    
    # Preparar o response
    output.seek(0)
    return Response(
        output,
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment;filename=agendamentos.csv"}
    )

@routes.route('/visualizar/<int:agendamento_id>', methods=['GET'])
@login_required
def visualizar_agendamento(agendamento_id):
    """
    Rota para visualizar os detalhes de um agendamento específico.
    
    :param agendamento_id: ID do agendamento a ser visualizado
    :return: Template renderizado com os detalhes do agendamento ou resposta JSON
    """
    # Buscar o agendamento pelo ID
    agendamento = AgendamentoVisita.query.get_or_404(agendamento_id)
    
    # Verificar permissões (somente o professor que criou ou um administrador pode ver)
    if current_user.id != agendamento.professor_id and not current_user.is_admin:
        abort(403, "Você não tem permissão para visualizar este agendamento")
    
    # Buscar informações adicionais
    # Se as salas estiverem armazenadas como IDs separados por vírgula
    salas_ids = []
    if agendamento.salas_selecionadas:
        salas_ids = [int(sala_id.strip()) for sala_id in agendamento.salas_selecionadas.split(',')]
        
    # Determinar o formato de resposta (HTML ou JSON)
    if request.headers.get('Accept') == 'application/json':
        # Resposta JSON para API
        return jsonify({
            'id': agendamento.id,
            'horario': {
                'id': agendamento.horario.id,
                'data': agendamento.horario.data.strftime('%d/%m/%Y'),
                'hora_inicio': agendamento.horario.hora_inicio.strftime('%H:%M'),
                'hora_fim': agendamento.horario.hora_fim.strftime('%H:%M')
            },
            'professor': {
                'id': agendamento.professor.id,
                'nome': agendamento.professor.nome,
                'email': agendamento.professor.email
            },
            'escola': {
                'nome': agendamento.escola_nome,
                'codigo_inep': agendamento.escola_codigo_inep
            },
            'turma': agendamento.turma,
            'nivel_ensino': agendamento.nivel_ensino,
            'quantidade_alunos': agendamento.quantidade_alunos,
            'status': agendamento.status,
            'checkin_realizado': agendamento.checkin_realizado,
            'data_agendamento': agendamento.data_agendamento.strftime('%d/%m/%Y %H:%M') if agendamento.data_agendamento else None,
            'data_cancelamento': agendamento.data_cancelamento.strftime('%d/%m/%Y %H:%M') if agendamento.data_cancelamento else None,
            'data_checkin': agendamento.data_checkin.strftime('%d/%m/%Y %H:%M') if agendamento.data_checkin else None,
            'qr_code_token': agendamento.qr_code_token,
            'salas_selecionadas': salas_ids
        })
    
    # Resposta HTML para interface web
    return render_template(
        'agendamento/visualizar.html',
        agendamento=agendamento,
        salas_ids=salas_ids
    )
    
@routes.route('/editar_agendamento/<int:agendamento_id>', methods=['GET', 'POST'])
@login_required
def editar_agendamento(agendamento_id):
    # Busca o agendamento no banco de dados
    agendamento = AgendamentoVisita.query.get_or_404(agendamento_id)
    
    # Verifica permissões (apenas o próprio professor, administradores ou clientes podem editar)
    if current_user.tipo not in ['admin', 'cliente'] and current_user.id != agendamento.professor_id:
        flash('Você não tem permissão para editar este agendamento.', 'danger')
        return redirect(url_for('routes.listar_agendamentos'))
    
    # Busca horários disponíveis para edição
    horarios_disponiveis = HorarioVisitacao.query.filter_by(disponivel=True).all()
    # Adiciona o horário atual do agendamento, caso ele não esteja mais disponível
    if agendamento.horario not in horarios_disponiveis:
        horarios_disponiveis.append(agendamento.horario)
    
    # Carrega as possíveis salas para visitação
    from models import Sala  # Importa o modelo Sala (assumindo que existe)
    salas = Sala.query.all()
    
    # Pega as salas já selecionadas
    salas_selecionadas = []
    if agendamento.salas_selecionadas:
        salas_selecionadas = [int(sala_id) for sala_id in agendamento.salas_selecionadas.split(',')]
    
    if request.method == 'POST':
        # Captura os dados do formulário
        horario_id = request.form.get('horario_id')
        escola_nome = request.form.get('escola_nome')
        escola_codigo_inep = request.form.get('escola_codigo_inep')
        turma = request.form.get('turma')
        nivel_ensino = request.form.get('nivel_ensino')
        quantidade_alunos = request.form.get('quantidade_alunos')
        salas_ids = request.form.getlist('salas')
        
        # Validação básica
        if not all([horario_id, escola_nome, turma, nivel_ensino, quantidade_alunos]):
            flash('Por favor, preencha todos os campos obrigatórios.', 'danger')
            return render_template(
                'editar_agendamento.html',
                agendamento=agendamento,
                horarios=horarios_disponiveis,
                salas=salas,
                salas_selecionadas=salas_selecionadas
            )
        
        try:
            # Atualiza os dados do agendamento
            agendamento.horario_id = horario_id
            agendamento.escola_nome = escola_nome
            agendamento.escola_codigo_inep = escola_codigo_inep
            agendamento.turma = turma
            agendamento.nivel_ensino = nivel_ensino
            agendamento.quantidade_alunos = int(quantidade_alunos)
            
            # Atualiza as salas selecionadas
            agendamento.salas_selecionadas = ','.join(salas_ids) if salas_ids else None
            
            db.session.commit()
            flash('Agendamento atualizado com sucesso!', 'success')
            return redirect(url_for('routes.listar_agendamentos'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao atualizar agendamento: {str(e)}', 'danger')
    
    # Renderiza o template com os dados do agendamento
    return render_template(
        'editar_agendamento.html',
        agendamento=agendamento,
        horarios=horarios_disponiveis,
        salas=salas,
        salas_selecionadas=salas_selecionadas
    )
    
@routes.route('/atualizar_status/<int:agendamento_id>', methods=['PUT'])
@login_required
def atualizar_status_agendamento(agendamento_id):
    """
    Atualiza o status de um agendamento de visita.
    
    Parâmetros:
    - agendamento_id: ID do agendamento a ser atualizado
    
    Corpo da requisição:
    {
        "status": "confirmado|cancelado|realizado",
        "checkin_realizado": true|false  (opcional)
    }
    
    Retorna:
    - 200: Agendamento atualizado com sucesso
    - 400: Dados inválidos
    - 403: Usuário não tem permissão
    - 404: Agendamento não encontrado
    """
    # Buscar o agendamento pelo ID
    agendamento = AgendamentoVisita.query.get(agendamento_id)
    
    # Verificar se o agendamento existe
    if not agendamento:
        return jsonify({"erro": "Agendamento não encontrado"}), 404
    
    # Verificar permissões: apenas o professor que criou ou um administrador pode alterar
    if current_user.id != agendamento.professor_id and not current_user.is_admin:
        return jsonify({"erro": "Você não tem permissão para alterar este agendamento"}), 403
    
    # Obter os dados do request
    dados = request.get_json()
    
    if not dados:
        return jsonify({"erro": "Nenhum dado fornecido"}), 400
    
    # Validar o status
    novo_status = dados.get('status')
    if novo_status and novo_status not in ['confirmado', 'cancelado', 'realizado']:
        return jsonify({"erro": "Status inválido. Use 'confirmado', 'cancelado' ou 'realizado'"}), 400
    
    # Atualizar o status
    if novo_status:
        agendamento.status = novo_status
        
        # Se for cancelado, registrar a data de cancelamento
        if novo_status == 'cancelado' and not agendamento.data_cancelamento:
            agendamento.data_cancelamento = datetime.utcnow()
    
    # Verificar se houve alteração no check-in
    if 'checkin_realizado' in dados:
        checkin = dados.get('checkin_realizado')
        
        # Se check-in está sendo realizado agora
        if checkin and not agendamento.checkin_realizado:
            agendamento.checkin_realizado = True
            agendamento.data_checkin = datetime.utcnow()
            # Se houve check-in e o status não foi alterado, atualizar para 'realizado'
            if not novo_status:
                agendamento.status = 'realizado'
        # Se check-in está sendo desfeito
        elif not checkin and agendamento.checkin_realizado:
            agendamento.checkin_realizado = False
            agendamento.data_checkin = None
    
    try:
        # Salvar as alterações no banco de dados
        db.session.commit()
        
        # Formatar resposta
        resposta = {
            "mensagem": "Agendamento atualizado com sucesso",
            "agendamento": {
                "id": agendamento.id,
                "status": agendamento.status,
                "checkin_realizado": agendamento.checkin_realizado,
                "data_checkin": agendamento.data_checkin.isoformat() if agendamento.data_checkin else None,
                "data_cancelamento": agendamento.data_cancelamento.isoformat() if agendamento.data_cancelamento else None
            }
        }
        
        return jsonify(resposta), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({"erro": f"Erro ao atualizar agendamento: {str(e)}"}), 500

# Rota para realizar check-in via QR Code
@routes.route('/checkin/<string:qr_code_token>', methods=['POST'])
@login_required
def checkin_agendamento(qr_code_token):
    """
    Realiza o check-in de um agendamento através do token QR Code.
    
    Parâmetros:
    - qr_code_token: Token único do QR Code do agendamento
    
    Retorna:
    - 200: Check-in realizado com sucesso
    - 403: Usuário não tem permissão
    - 404: Agendamento não encontrado
    - 409: Check-in já realizado
    """
    # Buscar o agendamento pelo token do QR Code
    agendamento = AgendamentoVisita.query.filter_by(qr_code_token=qr_code_token).first()
    
    # Verificar se o agendamento existe
    if not agendamento:
        return jsonify({"erro": "Agendamento não encontrado"}), 404
    
    # Verificar se o check-in já foi realizado
    if agendamento.checkin_realizado:
        return jsonify({"erro": "Check-in já foi realizado para este agendamento"}), 409
    
    # Verificar permissões: apenas o professor que criou ou um administrador pode realizar check-in
    if current_user.id != agendamento.professor_id and not current_user.is_admin:
        return jsonify({"erro": "Você não tem permissão para realizar check-in neste agendamento"}), 403
    
    # Realizar o check-in
    agendamento.checkin_realizado = True
    agendamento.data_checkin = datetime.utcnow()
    agendamento.status = 'realizado'
    
    try:
        # Salvar as alterações no banco de dados
        db.session.commit()
        
        # Formatar resposta
        resposta = {
            "mensagem": "Check-in realizado com sucesso",
            "agendamento": {
                "id": agendamento.id,
                "status": agendamento.status,
                "checkin_realizado": agendamento.checkin_realizado,
                "data_checkin": agendamento.data_checkin.isoformat()
            }
        }
        
        return jsonify(resposta), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({"erro": f"Erro ao realizar check-in: {str(e)}"}), 500

@routes.route('/agendamentos', methods=['GET'])
@login_required
def listar_agendamentos():
    """
    Lista os agendamentos de visitas com opções de filtro.
    Administradores veem todos os agendamentos, professores veem apenas os próprios.
    """
    # Definir os parâmetros de filtro
    page = request.args.get('page', 1, type=int)
    data_inicio = request.args.get('data_inicio')
    data_fim = request.args.get('data_fim')
    status = request.args.get('status')
    participante_id = request.args.get('participante_id')
    oficina_id = request.args.get('oficina_id')
    cliente_id = request.args.get('cliente_id')

    # Base da query
    query = AgendamentoVisita.query

    # Filtrar por tipo de usuário
    if current_user.tipo == 'participante' or current_user.tipo == 'professor':
        # Professores/participantes só veem seus próprios agendamentos
        query = query.filter(AgendamentoVisita.professor_id == current_user.id)
    elif current_user.tipo == 'cliente':
        # Clientes veem agendamentos relacionados a eles
        # Aqui precisaria de uma lógica para filtrar por cliente, se aplicável
        if current_user.id:
            # Filtrar agendamentos relacionados ao cliente
            # Esta lógica depende da sua estrutura de dados
            pass
    
    # Filtros dos parâmetros da URL
    if data_inicio:
        data_inicio_dt = datetime.strptime(data_inicio, '%Y-%m-%d')
        query = query.filter(AgendamentoVisita.horario.has(data_agendamento >= data_inicio_dt))
    
    if data_fim:
        data_fim_dt = datetime.strptime(data_fim, '%Y-%m-%d')
        query = query.filter(AgendamentoVisita.horario.has(data_agendamento <= data_fim_dt))
    
    if status:
        query = query.filter(AgendamentoVisita.status == status)
    
    if participante_id:
        query = query.filter(AgendamentoVisita.professor_id == participante_id)
    
    if oficina_id:
        # Se você relacionar agendamentos com oficinas
        # query = query.filter(AgendamentoVisita.oficina_id == oficina_id)
        pass
    
    if cliente_id and current_user.tipo == 'admin':
        # Se você relacionar agendamentos com clientes
        # query = query.filter(AgendamentoVisita.cliente_id == cliente_id)
        pass
    
    # Ordenação
    query = query.order_by(AgendamentoVisita.data_agendamento.desc())
    
    # Paginação
    pagination = query.paginate(page=page, per_page=10, error_out=False)
    agendamentos = pagination.items
    
    # Dados para os filtros de formulário
    oficinas = Oficina.query.all()
    participantes = Usuario.query.filter_by(tipo='participante').all()
    clientes = []
    if current_user.tipo == 'admin':
        clientes = Cliente.query.all()
    
    return render_template(
        'listar_agendamentos.html',
        agendamentos=agendamentos,
        pagination=pagination,
        oficinas=oficinas,
        participantes=participantes,
        clientes=clientes
    )
    
@routes.route('/processar_qrcode_agendamento', methods=['POST'])
@login_required
def processar_qrcode_agendamento():
    """
    Processa o QR Code lido e retorna informações sobre o agendamento.
    """
    if not request.is_json:
        return jsonify({
            'success': False,
            'message': 'Formato de requisição inválido. Envie um JSON.'
        }), 400
    
    data = request.get_json()
    token = data.get('token')
    
    if not token:
        return jsonify({
            'success': False,
            'message': 'Token não fornecido.'
        }), 400
    
    try:
        # Busca o agendamento pelo token do QR Code
        agendamento = AgendamentoVisita.query.filter_by(qr_code_token=token).first()
        
        if not agendamento:
            return jsonify({
                'success': False,
                'message': 'Agendamento não encontrado. Verifique o QR Code e tente novamente.'
            }), 404
        
        # Verifica se o agendamento foi cancelado
        if agendamento.status == 'cancelado':
            return jsonify({
                'success': False,
                'message': 'Este agendamento foi cancelado.'
            }), 400
        
        # Verifica se o check-in já foi realizado
        if agendamento.checkin_realizado:
            return jsonify({
                'success': False,
                'message': 'Check-in já realizado para este agendamento.',
                'redirect': url_for('routes.confirmar_checkin_agendamento', token=token)
            }), 200
        
        # Redireciona para a página de confirmação de check-in
        return jsonify({
            'success': True,
            'message': 'Agendamento encontrado!',
            'redirect': url_for('routes.confirmar_checkin_agendamento', token=token)
        }), 200
    
    except Exception as e:
        print(f"Erro ao processar QR code: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Erro ao processar o QR Code: {str(e)}'
        }), 500

      
@routes.route('/confirmar_checkin_agendamento/<token>', methods=['GET', 'POST'])
@login_required
def confirmar_checkin_agendamento(token):
    """
    Exibe página de confirmação e processa o check-in de um agendamento via QR code.
    """
    # Busca o agendamento pelo token
    agendamento = AgendamentoVisita.query.filter_by(qr_code_token=token).first()
    
    if not agendamento:
        flash('Agendamento não encontrado. Verifique o QR Code e tente novamente.', 'danger')
        return redirect(url_for('routes.checkin_qr_agendamento'))
    
    # Busca informações relacionadas
    evento = Evento.query.get(agendamento.horario.evento_id)
    horario = agendamento.horario
    
    # Se for POST, realiza o check-in
    if request.method == 'POST':
        try:
            # Atualiza o status do agendamento
            if not agendamento.checkin_realizado:
                agendamento.checkin_realizado = True
                agendamento.data_checkin = datetime.utcnow()
                agendamento.status = 'realizado'
                
                # Processa os alunos presentes
                alunos_presentes = request.form.getlist('alunos_presentes')
                for aluno in agendamento.alunos:
                    aluno.presente = str(aluno.id) in alunos_presentes
                
                db.session.commit()
                
                flash('Check-in realizado com sucesso!', 'success')
            else:
                flash('Este agendamento já teve check-in realizado anteriormente.', 'warning')
            
            return redirect(url_for('routes.dashboard_cliente'))
        
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao processar check-in: {str(e)}', 'danger')
    
    # Renderiza a página de confirmação
    return render_template('confirmar_checkin.html', 
                          agendamento=agendamento,
                          evento=evento,
                          horario=horario)


@routes.route('/checkin_qr_agendamento', methods=['GET'])
@login_required
def checkin_qr_agendamento():
    """
    Página para escanear QR code para check-in de agendamentos.
    """
    token = request.args.get('token')
    
    # Se um token foi fornecido via parâmetro de URL, redireciona para a confirmação
    if token:
        agendamento = AgendamentoVisita.query.filter_by(qr_code_token=token).first()
        if agendamento:
            return redirect(url_for('routes.confirmar_checkin_agendamento', token=token))
        else:
            flash('Agendamento não encontrado. Verifique o token e tente novamente.', 'danger')
    
    # Renderiza a página do scanner QR Code
    return render_template('checkin_qr_agendamento.html')

# Adicione isto ao seu arquivo routes.py
@routes.route('/professor/eventos_disponiveis')
def professor_eventos_disponiveis():
    
    # Buscar eventos disponíveis para o professor
    eventos = Evento.query.filter_by(cliente_id=current_user.cliente_id).all()
    
    # Renderizar o template com os eventos
    return render_template('professor/eventos_disponiveis.html', eventos=eventos)

@routes.route('/cadastro_professor', methods=['GET', 'POST'])
def cadastro_professor():
    if request.method == 'POST':
        # Coletar dados do formulário
        nome = request.form.get('nome')
        email = request.form.get('email')
        cpf = request.form.get('cpf')
        senha = request.form.get('senha')
        formacao = request.form.get('formacao')

        # Verificar se email ou CPF já existem
        usuario_existente = Usuario.query.filter(
            (Usuario.email == email) | (Usuario.cpf == cpf)
        ).first()

        if usuario_existente:
            flash('Email ou CPF já cadastrado!', 'danger')
            return render_template('cadastro_professor.html')

        # Criar novo usuário professor
        novo_professor = Usuario(
            nome=nome,
            email=email,
            cpf=cpf,
            senha=generate_password_hash(senha),
            formacao=formacao,
            tipo='professor'
        )

        try:
            db.session.add(novo_professor)
            db.session.commit()
            flash('Cadastro realizado com sucesso!', 'success')
            return redirect(url_for('routes.login'))
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao cadastrar: {str(e)}', 'danger')

    return render_template('cadastro_professor.html')

@routes.route('/agendar_visita/<int:horario_id>', methods=['GET', 'POST'])
@login_required
def agendar_visita(horario_id):
    if not current_user.is_professor():
        flash('Apenas professores podem fazer agendamentos.', 'danger')
        return redirect(url_for('routes.dashboard_participante'))

    horario = HorarioVisitacao.query.get_or_404(horario_id)

    if request.method == 'POST':
        # Coletar detalhes do agendamento
        escola_nome = request.form.get('escola_nome')
        turma = request.form.get('turma')
        nivel_ensino = request.form.get('nivel_ensino')
        quantidade_alunos = int(request.form.get('quantidade_alunos'))

        # Validar vagas disponíveis
        if quantidade_alunos > horario.vagas_disponiveis:
            flash('Quantidade de alunos excede vagas disponíveis.', 'danger')
            return redirect(url_for('routes.agendar_visita', horario_id=horario_id))

        # Criar agendamento
        novo_agendamento = AgendamentoVisita(
            horario_id=horario.id,
            professor_id=current_user.id,
            escola_nome=escola_nome,
            turma=turma,
            nivel_ensino=nivel_ensino,
            quantidade_alunos=quantidade_alunos
        )

        # Reduzir vagas disponíveis
        horario.vagas_disponiveis -= quantidade_alunos

        try:
            db.session.add(novo_agendamento)
            db.session.commit()
            flash('Agendamento realizado com sucesso!', 'success')
            return redirect(url_for('routes.dashboard_participante'))
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao agendar: {str(e)}', 'danger')

    return render_template('agendar_visita.html', horario=horario)

@routes.route('/adicionar_alunos', methods=['GET', 'POST'])
@login_required
def adicionar_alunos():
    """
    Rota para adicionar alunos (participantes) em lote ou individualmente.
    Suporta upload de arquivo CSV/Excel e também entrada manual de dados.
    """
    if current_user.tipo not in ['admin', 'cliente']:
        flash('Você não tem permissão para adicionar alunos.', 'danger')
        return redirect(url_for('routes.dashboard'))

    if request.method == 'POST':
        # Verifica se há upload de arquivo
        arquivo = request.files.get('arquivo')
        
        # Verifica se há entrada manual
        nome = request.form.get('nome')
        cpf = request.form.get('cpf')
        email = request.form.get('email')
        formacao = request.form.get('formacao')
        estados = request.form.getlist('estados[]')
        cidades = request.form.getlist('cidades[]')

        # Processamento de upload de arquivo
        if arquivo and arquivo.filename:
            try:
                # Usa Pandas para ler o arquivo de upload
                df = pd.read_excel(arquivo, dtype={'cpf': str})
                
                # Verificar colunas obrigatórias
                colunas_obrigatorias = ['nome', 'cpf', 'email', 'formacao']
                if not all(col in df.columns for col in colunas_obrigatorias):
                    flash(f"Erro: O arquivo deve conter as colunas: {', '.join(colunas_obrigatorias)}", "danger")
                    return redirect(url_for('routes.adicionar_alunos'))

                # Processamento em lote
                alunos_adicionados = 0
                for _, row in df.iterrows():
                    cpf_str = str(row['cpf']).strip()
                    
                    # Verifica se o usuário já existe
                    usuario_existente = Usuario.query.filter(
                        (Usuario.cpf == cpf_str) | (Usuario.email == row['email'])
                    ).first()

                    if usuario_existente:
                        print(f"⚠️ Usuário {row['nome']} já existe. Pulando...")
                        continue

                    novo_usuario = Usuario(
                        nome=row['nome'],
                        cpf=cpf_str,
                        email=row['email'],
                        senha=generate_password_hash(str(row['cpf'])),  # Senha inicial como CPF
                        formacao=row.get('formacao', 'Não informada'),
                        tipo='participante',
                        cliente_id=current_user.id  # Vincula ao cliente logado
                    )
                    
                    # Tratamento de estados e cidades do arquivo, se existirem
                    if 'estados' in df.columns and 'cidades' in df.columns:
                        novo_usuario.estados = str(row.get('estados', ''))
                        novo_usuario.cidades = str(row.get('cidades', ''))

                    db.session.add(novo_usuario)
                    alunos_adicionados += 1

                db.session.commit()
                flash(f"✅ {alunos_adicionados} alunos importados com sucesso!", "success")
                return redirect(url_for('routes.dashboard'))

            except Exception as e:
                db.session.rollback()
                flash(f"Erro ao processar arquivo: {str(e)}", "danger")
                print(f"❌ Erro na importação: {e}")
                return redirect(url_for('routes.adicionar_alunos'))

        # Processamento de entrada manual
        elif nome and cpf and email and formacao:
            try:
                # Verifica se o usuário já existe
                usuario_existente = Usuario.query.filter(
                    (Usuario.cpf == cpf) | (Usuario.email == email)
                ).first()

                if usuario_existente:
                    flash(f"Usuário com CPF {cpf} ou email {email} já existe.", "warning")
                    return redirect(url_for('routes.adicionar_alunos'))

                novo_usuario = Usuario(
                    nome=nome,
                    cpf=cpf,
                    email=email,
                    senha=generate_password_hash(cpf),  # Senha inicial como CPF
                    formacao=formacao,
                    tipo='participante',
                    cliente_id=current_user.id,  # Vincula ao cliente logado
                    estados=','.join(estados) if estados else None,
                    cidades=','.join(cidades) if cidades else None
                )

                db.session.add(novo_usuario)
                db.session.commit()
                flash("Aluno adicionado com sucesso!", "success")
                return redirect(url_for('routes.dashboard'))

            except Exception as e:
                db.session.rollback()
                flash(f"Erro ao adicionar aluno: {str(e)}", "danger")
                print(f"❌ Erro na adição manual: {e}")
                return redirect(url_for('routes.adicionar_alunos'))

        else:
            flash("Dados insuficientes para adicionar aluno.", "warning")
            return redirect(url_for('routes.adicionar_alunos'))

    # GET: Renderiza o formulário
    estados = obter_estados()
    return render_template('adicionar_alunos.html', estados=estados)

@routes.route('/importar_alunos', methods=['GET', 'POST'])
@login_required
def importar_alunos():
    """
    Rota específica para importação em lote de alunos (participantes).
    Suporta upload de arquivos Excel e CSV.
    """
    if current_user.tipo not in ['admin', 'cliente']:
        flash('Você não tem permissão para importar alunos.', 'danger')
        return redirect(url_for('routes.dashboard'))

    if request.method == 'POST':
        # Verifica se há upload de arquivo
        arquivo = request.files.get('arquivo')
        
        if not arquivo or not arquivo.filename:
            flash('Nenhum arquivo selecionado.', 'warning')
            return redirect(url_for('routes.importar_alunos'))

        try:
            # Determina o tipo de arquivo e usa a biblioteca correta
            if arquivo.filename.endswith(('.xlsx', '.xls')):
                df = pd.read_excel(arquivo, dtype={'cpf': str})
            elif arquivo.filename.endswith('.csv'):
                df = pd.read_csv(arquivo, dtype={'cpf': str})
            else:
                flash('Formato de arquivo não suportado. Use .xlsx, .xls ou .csv', 'danger')
                return redirect(url_for('routes.importar_alunos'))
            
            # Verificar colunas obrigatórias
            colunas_obrigatorias = ['nome', 'cpf', 'email', 'formacao']
            colunas_faltantes = [col for col in colunas_obrigatorias if col not in df.columns]
            
            if colunas_faltantes:
                flash(f"Erro: Colunas faltantes no arquivo: {', '.join(colunas_faltantes)}", "danger")
                return redirect(url_for('routes.importar_alunos'))

            # Processamento em lote
            alunos_adicionados = 0
            alunos_duplicados = 0
            alunos_invalidos = 0

            # Criar log de importação
            log_importacao = []

            for index, row in df.iterrows():
                try:
                    # Limpeza e validação de dados
                    nome = str(row['nome']).strip()
                    cpf = str(row['cpf']).strip()
                    email = str(row['email']).strip().lower()
                    formacao = str(row.get('formacao', 'Não informada')).strip()

                    # Validações básicas
                    if not nome or not cpf or not email:
                        log_importacao.append(f"Linha {index + 2}: Dados incompletos")
                        alunos_invalidos += 1
                        continue

                    # Verifica se o usuário já existe
                    usuario_existente = Usuario.query.filter(
                        (Usuario.cpf == cpf) | (Usuario.email == email)
                    ).first()

                    if usuario_existente:
                        log_importacao.append(f"Linha {index + 2}: Usuário já existe (CPF ou email duplicado)")
                        alunos_duplicados += 1
                        continue

                    # Tratamento de estados e cidades (se existirem no arquivo)
                    estados = row.get('estados', '') if 'estados' in df.columns else ''
                    cidades = row.get('cidades', '') if 'cidades' in df.columns else ''

                    # Cria novo usuário
                    novo_usuario = Usuario(
                        nome=nome,
                        cpf=cpf,
                        email=email,
                        senha=generate_password_hash(cpf),  # Senha inicial como CPF
                        formacao=formacao,
                        tipo='participante',
                        cliente_id=current_user.id,  # Vincula ao cliente logado
                        estados=str(estados),
                        cidades=str(cidades)
                    )
                    
                    db.session.add(novo_usuario)
                    alunos_adicionados += 1
                    log_importacao.append(f"Linha {index + 2}: Usuário {nome} importado com sucesso")

                except Exception as e:
                    log_importacao.append(f"Linha {index + 2}: Erro - {str(e)}")
                    alunos_invalidos += 1

            # Commit final
            db.session.commit()

            # Criar arquivo de log
            log_filename = f"log_importacao_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
            log_path = os.path.join('static', 'logs', log_filename)
            os.makedirs(os.path.dirname(log_path), exist_ok=True)
            
            with open(log_path, 'w', encoding='utf-8') as log_file:
                log_file.write("\n".join(log_importacao))

            # Mensagem resumo
            flash(f"""
                Importação concluída:
                ✅ Alunos adicionados: {alunos_adicionados}
                ⚠️ Alunos duplicados: {alunos_duplicados}
                ❌ Alunos inválidos: {alunos_invalidos}
                📄 Log de importação salvo.
            """, "info")

            # Redireciona com o log
            return render_template('resultado_importacao.html', 
                                   adicionados=alunos_adicionados, 
                                   duplicados=alunos_duplicados, 
                                   invalidos=alunos_invalidos,
                                   log_filename=log_filename)

        except Exception as e:
            db.session.rollback()
            flash(f"Erro ao processar arquivo: {str(e)}", "danger")
            print(f"❌ Erro na importação: {e}")
            return redirect(url_for('routes.importar_alunos'))

    # GET: Renderiza o formulário de importação
    return render_template('importar_alunos.html')

@routes.route('/cancelar_agendamento/<int:agendamento_id>', methods=['POST'])
@login_required
def cancelar_agendamento(agendamento_id):
    # Busca o agendamento no banco de dados
    agendamento = AgendamentoVisita.query.get_or_404(agendamento_id)

    # Verifica se o usuário é o professor que fez o agendamento ou admin
    if current_user.tipo != 'admin' and current_user.id != agendamento.professor_id:
        flash("Você não tem permissão para cancelar este agendamento!", "danger")
        return redirect(url_for('routes.dashboard'))

    # Atualiza o status do agendamento para cancelado e a data do cancelamento
    agendamento.status = 'cancelado'
    agendamento.data_cancelamento = datetime.utcnow()

    # Atualiza a capacidade do horário, devolvendo as vagas
    horario = agendamento.horario
    horario.vagas_disponiveis += agendamento.quantidade_alunos

    try:
        db.session.commit()
        flash("Agendamento cancelado com sucesso!", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Erro ao cancelar agendamento: {str(e)}", "danger")

    return redirect(url_for('routes.dashboard_professor' if current_user.tipo == 'participante' else 'routes.dashboard'))

@routes.route('/eventos_disponiveis', methods=['GET'])
@login_required
def eventos_disponiveis():
    if current_user.tipo != 'participante':
        flash('Acesso negado! Esta área é exclusiva para participantes.', 'danger')
        return redirect(url_for('routes.dashboard'))

    eventos = Evento.query.filter(Evento.data_inicio >= datetime.utcnow()).order_by(Evento.data_inicio).all()

    return render_template('eventos_disponiveis.html', eventos=eventos)

@routes.route('/listar_eventos_disponiveis', methods=['GET'])
@login_required
def listar_eventos_disponiveis():
    if current_user.tipo != 'professor':
        flash('Acesso negado!', 'danger')
        return redirect(url_for('routes.dashboard'))

    eventos = Evento.query.all()
    return render_template('eventos_disponiveis.html', eventos=eventos)



@routes.route('/detalhes_evento/<int:evento_id>', methods=['GET'])
@login_required
def detalhes_evento(evento_id):
    evento = Evento.query.get_or_404(evento_id)

    # Carrega as oficinas associadas ao evento
    oficinas = Oficina.query.filter_by(evento_id=evento_id).order_by(Oficina.titulo).all()

    return render_template('detalhes_evento.html', evento=evento, oficinas=oficinas)


@routes.route('/qrcode_agendamento/<int:agendamento_id>', methods=['GET'])
@login_required
def qrcode_agendamento(agendamento_id):
    agendamento = AgendamentoVisita.query.get_or_404(agendamento_id)

    # Dados que estarão no QR Code
    qr_data = f"Agendamento ID: {agendamento.id}, Evento: {agendamento.horario.evento.nome}, Data: {agendamento.horario.data.strftime('%d/%m/%Y')}"

    # Gerando QR Code
    qr = qrcode.QRCode(version=1, box_size=10, border=5)
    qr.add_data(qr_data)
    qr.make(fit=True)
    img = qr.make_image(fill='black', back_color='white')

    # Envia o QR Code como imagem
    buf = io.BytesIO()
    img.save(buf, 'PNG')
    buf.seek(0)

    return send_file(buf, mimetype='image/png', as_attachment=False, download_name=f'qrcode_agendamento_{agendamento.id}.png')

@routes.route('/api/horarios_disponiveis', methods=['GET'])
@login_required
def horarios_disponiveis_api():
    if current_user.tipo != 'professor':
        return jsonify({"error": "Acesso não permitido"}), 403

    horarios = HorarioVisitacao.query.filter(HorarioVisitacao.vagas_disponiveis > 0).all()
    eventos = []

    for horario in horarios:
        eventos.append({
            "id": horario.id,
            "title": f"Disponível ({horario.vagas_disponiveis} vagas)",
            "start": f"{horario.data}T{horario.horario_inicio}",
            "end": f"{horario.data}T{horario.horario_fim}",
            "url": url_for('routes.agendar_visita', horario_id=horario.id)
        })

    return jsonify(eventos)
@routes.route("/importar_oficinas", methods=["POST"])
@login_required
def importar_oficinas():
    """
    Exemplo de rota para importar oficinas de um arquivo Excel (.xlsx).
    Inclui o cadastro da própria oficina e também das datas (OficinaDia).
    """
    # 1. Verificar se foi enviado um arquivo
    if "arquivo" not in request.files:
        flash("Nenhum arquivo enviado!", "danger")
        return redirect(url_for("routes.dashboard_cliente"))
    
    arquivo = request.files["arquivo"]
    if arquivo.filename == "":
        flash("Nenhum arquivo selecionado.", "danger")
        return redirect(url_for("routes.dashboard_cliente"))

    # Verifica se a extensão é permitida (.xlsx)
    ALLOWED_EXTENSIONS = {"xlsx"}
    def arquivo_permitido(filename):
        return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

    if not arquivo_permitido(arquivo.filename):
        flash("Formato de arquivo inválido. Envie um arquivo Excel (.xlsx)", "danger")
        return redirect(url_for("routes.dashboard_cliente"))

    # 2. Salvar o arquivo em local temporário
    from werkzeug.utils import secure_filename
    import os
    filename = secure_filename(arquivo.filename)
    filepath = os.path.join(current_app.config["UPLOAD_FOLDER"], filename)
    arquivo.save(filepath)

    import pandas as pd
    from datetime import datetime
    from models import Oficina, OficinaDia
    from sqlalchemy.exc import IntegrityError

    try:
        # 3. Ler o arquivo Excel
        df = pd.read_excel(filepath)
        df.columns = df.columns.str.strip()  # tirar espaços extras do nome das colunas

        # Exemplo de colunas que esperamos:
        #   titulo, descricao, ministrante_id, vagas, carga_horaria,
        #   estado, cidade, datas, horarios_inicio, horarios_fim
        #
        # Onde:
        #   - "datas" pode ser uma string com várias datas separadas por vírgula, ex: "01/05/2025,02/05/2025"
        #   - "horarios_inicio" idem, ex: "08:00,09:00"
        #   - "horarios_fim" idem, ex: "12:00,13:00"
        #   - O número de datas deve bater com o número de horários_inicio e horários_fim
        
        colunas_obrigatorias = [
            "titulo", "descricao", "ministrante_id",
            "vagas", "carga_horaria", "estado", "cidade",
            "datas", "horarios_inicio", "horarios_fim"
        ]
        for col in colunas_obrigatorias:
            if col not in df.columns:
                flash(f"Erro: Coluna '{col}' não encontrada no arquivo Excel!", "danger")
                os.remove(filepath)
                return redirect(url_for("routes.dashboard_cliente"))

        # 4. Percorrer cada linha do DataFrame e criar as oficinas
        total_oficinas_criadas = 0
        for index, row in df.iterrows():
            try:
                # Converter alguns campos para o tipo adequado
                 # Tratar ministrante_id
                raw_m_id = row["ministrante_id"]
                if pd.isna(raw_m_id) or raw_m_id == '':
                    # Se estiver vazio, defina None ou crie lógica de fallback
                    ministrante_id = None
                else:
                    ministrante_id = int(raw_m_id)  # aqui converte para int, se tiver valor
                
                 # Tratar vagas
                raw_vagas = row["vagas"]
                if pd.isna(raw_vagas) or raw_vagas == '':
                    vagas = 0
                else:
                    vagas = int(raw_vagas)
                    
                carga_horaria = str(row["carga_horaria"])
                estado = str(row["estado"]).upper().strip()  # ex: "SP"
                cidade = str(row["cidade"]).strip()

                # Criar a oficina principal
                nova_oficina = Oficina(
                    titulo=row["titulo"],
                    descricao=row["descricao"],
                    ministrante_id=ministrante_id,
                    vagas=vagas,
                    carga_horaria=carga_horaria,
                    estado=estado,
                    cidade=cidade
                )

                # Se quiser vincular a um cliente específico:
                # nova_oficina.cliente_id = current_user.id

                db.session.add(nova_oficina)
                db.session.flush()  # para garantir que nova_oficina.id exista

                # Lendo as datas e horários
                # Supondo que cada coluna seja uma string com valores separados por vírgula
                datas_str = str(row["datas"]).strip()                # ex.: "01/05/2025,02/05/2025"
                horarios_inicio_str = str(row["horarios_inicio"]).strip()  # ex.: "08:00,09:00"
                horarios_fim_str = str(row["horarios_fim"]).strip()        # ex.: "12:00,13:00"

                datas_list = datas_str.split(",")
                hi_list = horarios_inicio_str.split(",")
                hf_list = horarios_fim_str.split(",")

                # Checa se todos os arrays têm mesmo tamanho
                if not (len(datas_list) == len(hi_list) == len(hf_list)):
                    raise ValueError(f"As colunas 'datas', 'horarios_inicio' e 'horarios_fim' devem ter a mesma quantidade de itens na linha {index+1}.")

                # 5. Para cada data, criar um registro OficinaDia
                for i in range(len(datas_list)):
                    data_str = datas_list[i].strip()
                    hi_str = hi_list[i].strip()
                    hf_str = hf_list[i].strip()

                    # Converter data_str para datetime.date (formato padrão dd/mm/yyyy)
                    try:
                        data_formatada = datetime.strptime(data_str, "%d/%m/%Y").date()
                    except ValueError:
                        raise ValueError(f"Data inválida na linha {index+1}: '{data_str}'. Formato esperado: DD/MM/YYYY.")

                    novo_dia = OficinaDia(
                        oficina_id=nova_oficina.id,
                        data=data_formatada,
                        horario_inicio=hi_str,
                        horario_fim=hf_str
                    )
                    db.session.add(novo_dia)

                db.session.commit()
                total_oficinas_criadas += 1
            
            except IntegrityError as e:
                db.session.rollback()
                current_app.logger.error(f"[Linha {index+1}] Erro de integridade: {e}")
            except Exception as e:
                db.session.rollback()
                current_app.logger.error(f"[Linha {index+1}] Erro ao criar oficina: {e}")

        flash(f"Foram importadas {total_oficinas_criadas} oficinas com sucesso, incluindo as datas!", "success")
    
    except Exception as e:
        db.session.rollback()
        flash(f"Erro ao processar o arquivo: {str(e)}", "danger")

    # Remover o arquivo temporário e redirecionar de volta
    os.remove(filepath)
    return redirect(url_for("routes.dashboard_cliente"))


@routes.route('/gerar_modelo/<string:tipo>', methods=['GET'])
@login_required
def gerar_modelo(tipo):
    """
    Gera um arquivo Excel (XLSX) em memória com colunas obrigatórias
    para importação de Usuários ou Oficinas. Retorna o arquivo para download.
    
    Use:
      /gerar_modelo/usuarios  -> para Modelo de Usuários
      /gerar_modelo/oficinas  -> para Modelo de Oficinas
    """
    # 1. Cria o Workbook em memória
    wb = Workbook()
    ws = wb.active

    if tipo.lower() == 'usuarios':
        ws.title = "ModeloUsuarios"

        # Exemplo de colunas do model Usuario:
        #   nome, cpf, email, senha, formacao, tipo
        colunas = [
            "nome", "cpf", "email", "senha", "formacao", "tipo"
        ]
        ws.append(colunas)

        # Exemplo de linha de demonstração
        ws.append([
            "Fulano de Tal",     # nome
            "123.456.789-00",    # cpf
            "fulano@email.com",  # email
            "senha123",          # senha
            "Graduado em X",     # formacao
            "participante"       # tipo: pode ser admin, cliente, participante, etc.
        ])

        # Nome do arquivo para download
        nome_arquivo = "modelo_usuarios.xlsx"

    elif tipo.lower() == 'oficinas':
        ws.title = "ModeloOficinas"

        # Exemplo de colunas do model Oficina (e OficinaDia):
        #   titulo, descricao, ministrante_id, vagas, carga_horaria,
        #   estado, cidade, datas, horarios_inicio, horarios_fim
        colunas = [
            "titulo", "descricao", "ministrante_id",
            "vagas", "carga_horaria", "estado", "cidade",
            "datas", "horarios_inicio", "horarios_fim"
        ]
        ws.append(colunas)

        # Exemplo de linha de demonstração
        ws.append([
            "Oficina Exemplo",              # titulo
            "Descricao da oficina",         # descricao
            1,                              # ministrante_id
            30,                             # vagas
            "4h",                           # carga_horaria
            "SP",                           # estado
            "São Paulo",                    # cidade
            "01/09/2025,02/09/2025",        # datas (separado por vírgula)
            "08:00,08:00",                  # horarios_inicio (mesma quantidade de itens de datas)
            "12:00,12:00"                   # horarios_fim
        ])

        nome_arquivo = "modelo_oficinas.xlsx"

    else:
        # Se não for "usuarios" nem "oficinas", retorna 400 (Bad Request)
        abort(400, "Tipo inválido. Use 'usuarios' ou 'oficinas'.")

    # 2. Salva o Workbook em um buffer de memória
    output = BytesIO()
    wb.save(output)
    output.seek(0)  # Volta para o início do buffer

    # 3. Retorna o arquivo
    return send_file(
        output,
        as_attachment=True,
        download_name=nome_arquivo,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
@routes.route('/adicionar_patrocinadores_categorizados', methods=['POST'])
@login_required
def adicionar_patrocinadores_categorizados():
    if current_user.tipo not in ['admin', 'cliente']:
        flash("Acesso negado!", "danger")
        return redirect(url_for('routes.dashboard'))

    evento_id = request.form.get('evento_id')
    if not evento_id:
        flash("Evento não selecionado!", "danger")
        return redirect(url_for('routes.dashboard_cliente'))

    qtd_realizacao = int(request.form.get('qtd_realizacao', 0))
    qtd_patrocinio = int(request.form.get('qtd_patrocinio', 0))
    qtd_organizacao = int(request.form.get('qtd_organizacao', 0))
    qtd_apoio = int(request.form.get('qtd_apoio', 0))

    def salvar_uploads(categoria_label, qtd):
        imported_count = 0
        for i in range(qtd):
            key = f'{categoria_label.lower()}_{i}'
            if key in request.files:
                file = request.files[key]
                if file and file.filename.strip():
                    filename = secure_filename(file.filename)
                    upload_folder = os.path.join('static', 'uploads', 'patrocinadores')
                    os.makedirs(upload_folder, exist_ok=True)
                    file.save(os.path.join(upload_folder, filename))

                    logo_path = os.path.join('uploads', 'patrocinadores', filename)

                    # Ajuste aqui para categoria com inicial maiúscula
                    novo_pat = Patrocinador(
                        evento_id=evento_id,
                        logo_path=logo_path,
                        categoria=categoria_label.capitalize()  # Realizacao, Patrocinio, Organizacao, Apoio
                    )
                    db.session.add(novo_pat)
                    imported_count += 1
        return imported_count

    total_importados = 0
    total_importados += salvar_uploads('realizacao', qtd_realizacao)
    total_importados += salvar_uploads('patrocinio', qtd_patrocinio)
    total_importados += salvar_uploads('organizacao', qtd_organizacao)
    total_importados += salvar_uploads('apoio', qtd_apoio)

    db.session.commit()

    flash(f"Patrocinadores adicionados com sucesso! Total: {total_importados}", "success")
    return redirect(url_for('routes.dashboard_cliente'))

@routes.route('/templates_certificado', methods=['GET', 'POST'])
@login_required
def templates_certificado():
    if request.method == 'POST':
        titulo = request.form['titulo']
        conteudo = request.form['conteudo']
        novo_template = CertificadoTemplate(cliente_id=current_user.id, titulo=titulo, conteudo=conteudo)
        db.session.add(novo_template)
        db.session.commit()
        flash('Template cadastrado com sucesso!', 'success')

    templates = CertificadoTemplate.query.filter_by(cliente_id=current_user.id).all()
    return render_template('templates_certificado.html', templates=templates)

@routes.route('/set_template_ativo/<int:template_id>', methods=['POST'])
@login_required
def set_template_ativo(template_id):
    CertificadoTemplate.query.filter_by(cliente_id=current_user.id).update({'ativo': False})
    template = CertificadoTemplate.query.get(template_id)
    template.ativo = True
    db.session.commit()
    flash('Template definido como ativo com sucesso!', 'success')
    return redirect(url_for('routes.templates_certificado'))

@routes.route('/gerar_certificado_evento', methods=['POST'])
@login_required
def gerar_certificado_evento():
    texto_personalizado = request.form.get('texto_personalizado', '')
    oficinas_ids = request.form.getlist('oficinas_selecionadas')

    oficinas = Oficina.query.filter(Oficina.id.in_(oficinas_ids)).all()
    total_horas = sum(int(of.carga_horaria) for of in oficinas)

    # Capturar template ativo
    template = CertificadoTemplate.query.filter_by(cliente_id=current_user.id, ativo=True).first()
    if not template:
        flash('Nenhum template ativo encontrado!', 'danger')
        return redirect(url_for('routes.dashboard_cliente'))

    pdf_path = gerar_certificado_personalizado(current_user, oficinas, total_horas, texto_personalizado, template.conteudo)
    return send_file(pdf_path, as_attachment=True)

@routes.route('/salvar_personalizacao_certificado', methods=['POST'])
@login_required
def salvar_personalizacao_certificado():
    cliente = Cliente.query.get(current_user.id)

    for campo in ['logo_certificado', 'assinatura_certificado', 'fundo_certificado']:
        arquivo = request.files.get(campo)
        if arquivo:
            filename = secure_filename(arquivo.filename)
            path = os.path.join('static/uploads/certificados', filename)
            arquivo.save(path)
            setattr(cliente, campo, path)

    cliente.texto_personalizado = request.form.get('texto_personalizado')
    db.session.commit()

    flash('Personalizações salvas com sucesso!', 'success')
    return redirect(url_for('routes.upload_personalizacao_certificado'))

@routes.route('/ativar_template_certificado/<int:template_id>', methods=['POST'])
@login_required
def ativar_template_certificado(template_id):
    CertificadoTemplate.query.filter_by(cliente_id=current_user.id).update({'ativo': False})
    template = CertificadoTemplate.query.get_or_404(template_id)
    template.ativo = True
    db.session.commit()

    flash('Template ativado com sucesso!', 'success')
    return redirect(url_for('routes.upload_personalizacao_certificado'))


@routes.route('/editar_template_certificado/<int:template_id>', methods=['POST'])
@login_required
def editar_template_certificado(template_id):
    template = CertificadoTemplate.query.get_or_404(template_id)

    if template.cliente_id != current_user.id:
        flash('Você não tem permissão para editar este template.', 'danger')
        return redirect(url_for('routes.upload_personalizacao_certificado'))

    novo_titulo = request.form.get('titulo')
    novo_conteudo = request.form.get('conteudo')

    if not novo_titulo or not novo_conteudo:
        flash('Todos os campos são obrigatórios.', 'warning')
        return redirect(url_for('routes.upload_personalizacao_certificado'))

    template.titulo = novo_titulo
    template.conteudo = novo_conteudo

    db.session.commit()
    flash('Template atualizado com sucesso!', 'success')
    return redirect(url_for('routes.upload_personalizacao_certificado'))

@routes.route('/desativar_template_certificado/<int:template_id>', methods=['POST'])
@login_required
def desativar_template_certificado(template_id):
    template = CertificadoTemplate.query.get_or_404(template_id)

    if template.cliente_id != current_user.id:
        flash('Você não tem permissão para alterar esse template.', 'danger')
        return redirect(url_for('routes.upload_personalizacao_certificado'))

    template.ativo = False
    db.session.commit()
    flash('Template desativado com sucesso!', 'info')
    return redirect(url_for('routes.upload_personalizacao_certificado'))

@routes.route('/remover_patrocinador/<int:patrocinador_id>', methods=['POST'])
@login_required
def remover_patrocinador(patrocinador_id):
    if current_user.tipo not in ['admin', 'cliente']:
        flash("Acesso negado!", "danger")
        return redirect(url_for('routes.dashboard'))

    patrocinador = Patrocinador.query.get_or_404(patrocinador_id)

    # Se for cliente, verifica se realmente é dele
    if current_user.tipo == 'cliente':
        # Busca o evento do patrocinador e verifica se pertence ao cliente
        if not patrocinador.evento or patrocinador.evento.cliente_id != current_user.id:
            flash("Você não tem permissão para remover esse patrocinador.", "danger")
            return redirect(url_for('routes.listar_patrocinadores'))

    try:
        db.session.delete(patrocinador)
        db.session.commit()
        flash("Patrocinador removido com sucesso!", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Erro ao remover patrocinador: {e}", "danger")

    return redirect(url_for('routes.listar_patrocinadores'))


@routes.route('/patrocinadores', methods=['GET'])
@login_required
def listar_patrocinadores():
    # Verifica se é admin ou cliente
    if current_user.tipo not in ['admin', 'cliente']:
        flash("Acesso negado!", "danger")
        return redirect(url_for('routes.dashboard'))
    
    # Se for admin, traz todos; se for cliente, traz só do cliente
    if current_user.tipo == 'admin':
        patrocinadores = Patrocinador.query.all()
    else:
        # Busca os eventos do cliente
        eventos_cliente = Evento.query.filter_by(cliente_id=current_user.id).all()
        evento_ids = [ev.id for ev in eventos_cliente]
        # Traz patrocinadores apenas dos eventos do cliente
        patrocinadores = Patrocinador.query.filter(Patrocinador.evento_id.in_(evento_ids)).all()

    return render_template(
        'listar_patrocinadores.html', 
        patrocinadores=patrocinadores
    )

@routes.route('/gerenciar_patrocinadores')
@login_required
def gerenciar_patrocinadores():
    """Lista todos os patrocinadores, de todas as categorias."""
    if current_user.tipo not in ['admin','cliente']:
        flash("Acesso negado!", "danger")
        return redirect(url_for('routes.dashboard'))

    # Se for admin, traz todos. Se for cliente, filtra pelos eventos do cliente
    if current_user.tipo == 'admin':
        patrocinadores = Patrocinador.query.all()
    else:
        # Buscar eventos do cliente e extrair seus IDs
        eventos_cliente = Evento.query.filter_by(cliente_id=current_user.id).all()
        eventos_ids = [ev.id for ev in eventos_cliente]
        patrocinadores = Patrocinador.query.filter(Patrocinador.evento_id.in_(eventos_ids)).all()

    return render_template('gerenciar_patrocinadores.html', patrocinadores=patrocinadores)

@routes.route('/remover_foto_patrocinador/<int:patrocinador_id>', methods=['POST'])
@login_required
def remover_foto_patrocinador(patrocinador_id):
    """Remove a foto de patrocinador (categoria: Realização, Organização, Apoio, Patrocínio)."""
    if current_user.tipo not in ['admin','cliente']:
        flash("Acesso negado!", "danger")
        return redirect(url_for('routes.dashboard'))

    pat = Patrocinador.query.get_or_404(patrocinador_id)

    # Se for cliente, verifica se esse patrocinador é dele
    if current_user.tipo == 'cliente':
        if not pat.evento or pat.evento.cliente_id != current_user.id:
            flash("Você não tem permissão para remover este registro.", "danger")
            return redirect(url_for('routes.gerenciar_patrocinadores'))

    try:
        db.session.delete(pat)
        db.session.commit()
        flash("Logo removida com sucesso!", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Erro ao remover: {e}", "danger")

    return redirect(url_for('routes.gerenciar_patrocinadores'))

@routes.route('/adicionar_campo_personalizado', methods=['POST'])
@login_required
def adicionar_campo_personalizado():
    nome_campo = request.form.get('nome_campo')
    tipo_campo = request.form.get('tipo_campo')
    obrigatorio = bool(request.form.get('obrigatorio'))

    novo_campo = CampoPersonalizadoCadastro(
        cliente_id=current_user.id,
        nome=nome_campo,
        tipo=tipo_campo,
        obrigatorio=obrigatorio
    )
    db.session.add(novo_campo)
    db.session.commit()

    flash('Campo personalizado adicionado com sucesso!', 'success')
    return redirect(url_for('routes.dashboard_cliente'))

@routes.route('/remover_campo_personalizado/<int:campo_id>', methods=['POST'])
@login_required
def remover_campo_personalizado(campo_id):
    campo = CampoPersonalizadoCadastro.query.get_or_404(campo_id)

    if campo.cliente_id != current_user.id:
        flash('Você não tem permissão para remover este campo.', 'danger')
        return redirect(url_for('routes.dashboard_cliente'))

    db.session.delete(campo)
    db.session.commit()

    flash('Campo personalizado removido com sucesso!', 'success')
    return redirect(url_for('routes.dashboard_cliente'))

@routes.route('/gerar_folder_evento/<int:evento_id>')
def gerar_folder_evento(evento_id):

    """
    Gera um PDF em formato de folder com três dobras (trifold) contendo
    a programação do evento para download. Usa canvas para desenho direto,
    com quebra de texto automática e layout compacto.
    
    Args:
        evento_id: ID do evento para obter a programação
        
    Returns:
        Arquivo PDF para download
    """
    # Importações necessárias
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm, cm, inch
    from reportlab.platypus import Paragraph
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT, TA_JUSTIFY
    from reportlab.pdfgen import canvas
    import os
    from datetime import datetime
    from flask import current_app
    from flask import send_file
    import textwrap
    
    # Função para quebrar texto em várias linhas
    def wrap_text(text, width, font_name, font_size):
        """Quebra o texto em múltiplas linhas baseado na largura disponível."""
        words = text.split()
        lines = []
        current_line = []
        
        for word in words:
            # Testar se adicionando esta palavra ainda cabe na linha
            test_line = current_line + [word]
            test_text = ' '.join(test_line)
            text_width = c.stringWidth(test_text, font_name, font_size)
            
            if text_width <= width:
                current_line.append(word)
            else:
                # A linha atual está cheia, adicioná-la à lista e começar uma nova
                if current_line:  # Verificar se não está vazia
                    lines.append(' '.join(current_line))
                    current_line = [word]
                else:
                    # Se a linha estiver vazia, significa que uma única palavra é maior que a largura
                    # Adicionar a palavra mesmo assim (será cortada na exibição)
                    lines.append(word)
                    current_line = []
        
        # Adicionar a última linha
        if current_line:
            lines.append(' '.join(current_line))
            
        return lines
    
    # Busca o evento no banco de dados
    evento = Evento.query.get_or_404(evento_id)
    
    # Preparar dados para o PDF
    oficinas = Oficina.query.filter_by(evento_id=evento_id).all()
    
    # Agrupar oficinas por data
    from collections import defaultdict
    grouped_oficinas = defaultdict(list)
    
    for oficina in oficinas:
        for dia in oficina.dias:
            data_str = dia.data.strftime('%d/%m/%Y')
            
            # Obter nome do ministrante principal
            ministrante_nome = oficina.ministrante_obj.nome if oficina.ministrante_obj else 'N/A'
            
            # Obter foto do ministrante principal
            ministrante_foto = None
            if oficina.ministrante_obj and hasattr(oficina.ministrante_obj, 'foto') and oficina.ministrante_obj.foto:
                ministrante_foto = oficina.ministrante_obj.foto
            
            # Obter lista de ministrantes associados
            ministrantes_associados = []
            if hasattr(oficina, 'ministrantes_associados'):
                for m in oficina.ministrantes_associados:
                    ministrantes_associados.append({
                        'nome': m.nome,
                        'foto': m.foto if hasattr(m, 'foto') else None
                    })
            
            grouped_oficinas[data_str].append({
                'titulo': oficina.titulo,
                'descricao': oficina.descricao,
                'ministrante': ministrante_nome,
                'ministrante_foto': ministrante_foto,
                'ministrantes_associados': ministrantes_associados,
                'horario_inicio': dia.horario_inicio,
                'horario_fim': dia.horario_fim,
                'tipo_oficina': oficina.tipo_oficina if hasattr(oficina, 'tipo_oficina') and oficina.tipo_oficina != 'outros' else 
                               (oficina.tipo_oficina_outro if hasattr(oficina, 'tipo_oficina_outro') else 'Oficina')
            })
    
    # Ordenar datas
    sorted_keys = sorted(grouped_oficinas.keys(), key=lambda d: datetime.strptime(d, '%d/%m/%Y'))
    
    # Preparar o diretório para salvar o PDF
    pdf_filename = f"folder_trifold_{evento_id}{datetime.now().strftime('%Y%m%d%H%M%S')}.pdf"
    folder_dir = os.path.join("static", "folders")
    os.makedirs(folder_dir, exist_ok=True)
    pdf_path = os.path.join(folder_dir, pdf_filename)
    
    # Definição de cores
    primary_color = colors.HexColor('#1a237e')      # Azul escuro
    secondary_color = colors.HexColor('#283593')    # Azul médio
    accent_color = colors.HexColor('#5c6bc0')       # Azul claro
    light_bg = colors.HexColor('#f5f5f5')           # Cinza claro
    text_dark = colors.HexColor('#212121')          # Quase preto
    text_medium = colors.HexColor('#424242')        # Cinza escuro
    highlight_color = colors.HexColor('#03a9f4')    # Azul claro para destaque
    
    # Configuração de estilos personalizados para os parágrafos
    styles = getSampleStyleSheet()
    
    # Estilo para descrições
    description_style = ParagraphStyle(
        name='Description',
        parent=styles['Normal'],
        fontSize=9,  # Reduzido
        alignment=TA_JUSTIFY,
        spaceAfter=2 * mm,
        leading=11,
        textColor=text_medium
    )
    
    # Estilo para títulos de atividades
    activity_title_style = ParagraphStyle(
        name='ActivityTitle',
        parent=styles['Heading4'],
        fontSize=12,  # Reduzido
        alignment=TA_LEFT,
        spaceAfter=1 * mm,
        fontName='Helvetica-Bold',
        textColor=secondary_color,
        leading=15
    )
    
    # Criar o PDF diretamente com canvas
    pagesize = landscape(A4)
    c = canvas.Canvas(pdf_path, pagesize=pagesize)
    
    # Dimensões da página
    page_width, page_height = pagesize
    
    # Largura útil (removendo margens)
    margin = 10*mm
    gutter = 5*mm
    usable_width = page_width - 2*margin
    
    # Largura de cada painel
    panel_width = (usable_width - 2*gutter) / 3
    
    # Posições X dos painéis
    panel1_x = margin  # Painel da direita (frente)
    panel2_x = margin + panel_width + gutter  # Painel do meio
    panel3_x = margin + 2*panel_width + 2*gutter  # Painel da esquerda (verso)
    
    # Desenhar linhas para marcar as dobras
    c.setStrokeColor(colors.gray)
    c.setDash([3, 3])
    c.setLineWidth(0.5)
    
    # Primeira dobra
    fold1_x = panel1_x + panel_width + gutter/2
    c.line(fold1_x, 5*mm, fold1_x, page_height - 5*mm)
    
    # Segunda dobra
    fold2_x = panel2_x + panel_width + gutter/2
    c.line(fold2_x, 5*mm, fold2_x, page_height - 5*mm)
    
    # Desenhar moldura de cada painel para visualização
    c.setStrokeColor(colors.black)
    c.setDash([1, 2])
    c.setLineWidth(0.5)
    
    # Painel 1 (direita)
    c.rect(panel1_x, margin, panel_width, page_height - 2*margin, stroke=1, fill=0)
    
    # Painel 2 (meio)
    c.rect(panel2_x, margin, panel_width, page_height - 2*margin, stroke=1, fill=0)
    
    # Painel 3 (esquerda)
    c.rect(panel3_x, margin, panel_width, page_height - 2*margin, stroke=1, fill=0)
    
    # === PAINEL 1 (CAPA) ===
    # Posição Y inicial para este painel
    y_pos = page_height - 2*margin
    
    # Título do evento em destaque
    c.setFont("Helvetica-Bold", 22)
    c.setFillColor(primary_color)
    
    # Banner do evento se disponível
    if evento.banner_url:
        try:
            # Tentar construir caminho correto da imagem
            img_path = None
            if evento.banner_url.startswith('http'):
                img_path = evento.banner_url
            elif evento.banner_url.startswith('/static/'):
                img_path = os.path.join(current_app.root_path, 'static', evento.banner_url.replace('/static/', ''))
            else:
                img_path = os.path.join(current_app.root_path, 'static', evento.banner_url)
                
            if img_path and (img_path.startswith('http') or os.path.exists(img_path)):
                img_width = panel_width * 0.9
                img_height = img_width / 2  # Proporção da imagem
                
                # Centralizar a imagem
                img_x = panel1_x + (panel_width - img_width) / 2
                c.drawImage(img_path, img_x, y_pos - img_height, width=img_width, height=img_height)
                
                # Ajustar posição Y
                y_pos -= (img_height + 15*mm)
        except Exception as e:
            print(f"Erro ao carregar imagem: {str(e)}")
            y_pos -= 10*mm  # Ajustar menos se a imagem falhar
    
    # Título do evento
    title_text = evento.nome
    title_width = c.stringWidth(title_text, "Helvetica-Bold", 22)
    
    # Se o título for muito largo, dividir em várias linhas
    if title_width > panel_width * 0.9:
        title_lines = wrap_text(title_text, panel_width * 0.9, "Helvetica-Bold", 22)
        line_height = 25  # Altura de linha em pontos
        
        # Desenhar cada linha do título centralizada
        for line in title_lines:
            line_width = c.stringWidth(line, "Helvetica-Bold", 22)
            c.drawString(panel1_x + (panel_width - line_width) / 2, y_pos - 10*mm, line)
            y_pos -= line_height
        
        y_pos -= 5*mm  # Espaço adicional depois do título
    else:
        # Centralizar texto
        c.drawString(panel1_x + (panel_width - title_width) / 2, y_pos - 10*mm, title_text)
        y_pos -= 20*mm
    
    # Informações do evento
    info_text = []
    
    # Datas
    if hasattr(evento, 'data_inicio') and evento.data_inicio and hasattr(evento, 'data_fim') and evento.data_fim:
        data_evento = f"De {evento.data_inicio.strftime('%d/%m/%Y')} a {evento.data_fim.strftime('%d/%m/%Y')}"
        info_text.append(data_evento)
    
    # Local
    if evento.localizacao:
        info_text.append(f"Local: {evento.localizacao}")
    
    # Desenhar informações
    c.setFont("Helvetica", 10)
    c.setFillColor(text_dark)
    
    # Desenhar cada linha de informação
    for line in info_text:
        line_width = c.stringWidth(line, "Helvetica", 10)
        
        # Se a linha for muito grande, quebrar
        if line_width > panel_width * 0.9:
            wrapped_lines = wrap_text(line, panel_width * 0.9, "Helvetica", 10)
            for wrapped in wrapped_lines:
                wrapped_width = c.stringWidth(wrapped, "Helvetica", 10)
                c.drawString(panel1_x + (panel_width - wrapped_width) / 2, y_pos - 5*mm, wrapped)
                y_pos -= 12*mm
        else:
            c.drawString(panel1_x + (panel_width - line_width) / 2, y_pos - 5*mm, line)
            y_pos -= 12*mm
    
    # Deixar espaço
    y_pos -= 5*mm
    
    # Adicionar descrição do evento se disponível
    if hasattr(evento, 'descricao') and evento.descricao:
        c.setFont("Helvetica-Bold", 16)
        c.setFillColor(secondary_color)
        about_text = "Sobre o Evento"
        about_width = c.stringWidth(about_text, "Helvetica-Bold", 16)
        c.drawString(panel1_x + (panel_width - about_width) / 2, y_pos - 8*mm, about_text)
        y_pos -= 15*mm
        
        # Descrição em parágrafo limitada a 500 caracteres
        desc_text = evento.descricao[:500] + ("..." if len(evento.descricao) > 500 else "")
        
        # Usar Paragraph para descrição com quebra automática
        desc_p = Paragraph(desc_text, description_style)
        w, h = desc_p.wrap(panel_width * 0.9, 100*mm)
        desc_p.drawOn(c, panel1_x + (panel_width - w) / 2, y_pos - h)
        y_pos -= (h + 10*mm)
    
    # Contato
    if hasattr(evento, 'email_contato') and evento.email_contato:
        c.setFont("Helvetica", 8)
        c.setFillColor(colors.gray)
        contact_text = f"Contato: {evento.email_contato}"
        contact_width = c.stringWidth(contact_text, "Helvetica", 8)
        c.drawString(panel1_x + (panel_width - contact_width) / 2, margin + 10*mm, contact_text)
    
    # === DIVISÃO DA PROGRAMAÇÃO ENTRE PAINÉIS 2 E 3 ===
    # Selecionamos metade das datas para o painel 2 e metade para o painel 3
    half_point = len(sorted_keys) // 2
    panel2_dates = sorted_keys[:half_point]
    panel3_dates = sorted_keys[half_point:]
    
    # Se houver apenas uma data, colocamos no painel 2
    if not panel2_dates and panel3_dates:
        panel2_dates = [panel3_dates[0]]
        panel3_dates = panel3_dates[1:]
    
    # === PAINEL 2 (PROGRAMAÇÃO - PARTE 1) ===
    y_pos = page_height - 2*margin
    
    # Título da programação
    c.setFont("Helvetica-Bold", 16)
    c.setFillColor(secondary_color)
    prog_text = "PROGRAMAÇÃO"
    prog_width = c.stringWidth(prog_text, "Helvetica-Bold", 16)
    c.drawString(panel2_x + (panel_width - prog_width) / 2, y_pos - 8*mm, prog_text)
    y_pos -= 20*mm
    
    # Função para desenhar uma data com suas atividades
    def draw_date_activities(date_key, start_x, start_y, width):
        y = start_y
        
        # Nome do dia da semana
        weekday_names = {
            0: "Segunda-feira", 
            1: "Terça-feira", 
            2: "Quarta-feira", 
            3: "Quinta-feira", 
            4: "Sexta-feira", 
            5: "Sábado", 
            6: "Domingo"
        }
        day_date = datetime.strptime(date_key, '%d/%m/%Y')
        weekday = weekday_names[day_date.weekday()]
        
        # Fundo para o título do dia
        c.setFillColor(primary_color)
        c.rect(start_x + 5*mm, y - 8*mm, width - 10*mm, 12*mm, fill=1, stroke=0)
        
        # Texto do dia
        c.setFont("Helvetica-Bold", 14)
        c.setFillColor(colors.white)
        day_text = f"{weekday} - {date_key}"
        c.drawString(start_x + 10*mm, y - 5*mm, day_text)
        y -= 20*mm
        
        # Ordenar atividades pelo horário
        sorted_atividades = sorted(grouped_oficinas[date_key], key=lambda x: x['horario_inicio'])
        
        # Desenhar cada atividade
        for atividade in sorted_atividades:
            # Criar texto com título e horário na mesma linha
            c.setFont("Helvetica-Bold", 12)
            c.setFillColor(secondary_color)
            
            # Horário com fonte menor
            horario = f"{atividade['horario_inicio']} - {atividade['horario_fim']}"
            c.setFont("Helvetica", 8)
            c.setFillColor(text_medium)
            horario_width = c.stringWidth(horario, "Helvetica", 8)
            
            # Título da atividade
            c.setFont("Helvetica-Bold", 12)
            c.setFillColor(secondary_color)
            
            # Título com quebra automática (considerando espaço para o horário)
            titulo_width = width - 20*mm  # Largura disponível para o título
            wrapped_title = wrap_text(atividade['titulo'], titulo_width, "Helvetica-Bold", 12)
            
            # Desenhar a primeira linha do título com o horário
            c.drawString(start_x + 10*mm, y - 5*mm, wrapped_title[0])
            
            # Desenhar o horário ao lado da primeira linha (se couber) ou na mesma altura
            c.setFont("Helvetica", 8)
            c.setFillColor(text_medium)
            
            # Calcular posição X para o horário (alinhado à direita do painel)
            horario_x = start_x + width - 10*mm - horario_width
            c.drawString(horario_x, y - 5*mm, horario)
            
            # Desenhar resto do título se houver mais linhas
            if len(wrapped_title) > 1:
                c.setFont("Helvetica-Bold", 12)
                c.setFillColor(secondary_color)
                line_height = 14
                for i in range(1, len(wrapped_title)):
                    y -= line_height
                    c.drawString(start_x + 10*mm, y - 5*mm, wrapped_title[i])
            
            y -= 15*mm
            
            # Tipo de oficina
            c.setFont("Helvetica-Oblique", 9)
            c.setFillColor(highlight_color)
            c.drawString(start_x + 10*mm, y - 5*mm, atividade['tipo_oficina'])
            
            y -= 10*mm
            
            # Ministrante
            if atividade['ministrante'] != 'N/A':
                c.setFont("Helvetica-Oblique", 9)
                c.setFillColor(accent_color)
                ministrante_text = f"Ministrante: {atividade['ministrante']}"
                
                # Verificar se precisa quebrar o texto do ministrante
                ministrante_width = c.stringWidth(ministrante_text, "Helvetica-Oblique", 9)
                if ministrante_width > width - 20*mm:
                    ministrante_lines = wrap_text(ministrante_text, width - 20*mm, "Helvetica-Oblique", 9)
                    line_height = 11
                    for line in ministrante_lines:
                        c.drawString(start_x + 10*mm, y - 5*mm, line)
                        y -= line_height
                else:
                    c.drawString(start_x + 10*mm, y - 5*mm, ministrante_text)
                    y -= 10*mm
            
            # Descrição resumida
            if atividade['descricao']:
                # Limitar a descrição para economizar espaço
                desc_short = atividade['descricao'][:150] + ("..." if len(atividade['descricao']) > 150 else "")
                
                # Usar Paragraph para descrição com quebra automática
                desc_p = Paragraph(desc_short, description_style)
                desc_w, desc_h = desc_p.wrap(width - 20*mm, 50*mm)
                desc_p.drawOn(c, start_x + 10*mm, y - desc_h)
                y -= (desc_h + 5*mm)
            
            # Separador entre atividades
            c.setStrokeColor(light_bg)
            c.setLineWidth(0.5)
            c.line(start_x + 10*mm, y - 5*mm, start_x + width - 10*mm, y - 5*mm)
            y -= 10*mm
        
        return y
    
    # Desenhar programação para o painel 2
    for date_key in panel2_dates:
        y_pos = draw_date_activities(date_key, panel2_x, y_pos, panel_width)
    
    # === PAINEL 3 (PROGRAMAÇÃO - PARTE 2 OU INFORMAÇÕES ADICIONAIS) ===
    y_pos = page_height - 2*margin
    
    # Se houver mais programação, continua no painel 3
    if panel3_dates:
        # Título da programação (continuação)
        c.setFont("Helvetica-Bold", 16)
        c.setFillColor(secondary_color)
        prog2_text = "PROGRAMAÇÃO (cont.)"
        prog2_width = c.stringWidth(prog2_text, "Helvetica-Bold", 16)
        c.drawString(panel3_x + (panel_width - prog2_width) / 2, y_pos - 8*mm, prog2_text)
        y_pos -= 20*mm
        
        # Desenhar programação para o painel 3
        for date_key in panel3_dates:
            y_pos = draw_date_activities(date_key, panel3_x, y_pos, panel_width)
    else:
        # Se não houver mais programação, adicionar informações adicionais
        c.setFont("Helvetica-Bold", 16)
        c.setFillColor(secondary_color)
        info_add_text = "INFORMAÇÕES ADICIONAIS"
        info_add_width = c.stringWidth(info_add_text, "Helvetica-Bold", 16)
        c.drawString(panel3_x + (panel_width - info_add_width) / 2, y_pos - 8*mm, info_add_text)
        y_pos -= 20*mm
        
        # Informações adicionais
        if hasattr(evento, 'informacoes_adicionais') and evento.informacoes_adicionais:
            # Usar Paragraph para texto com quebra automática
            info_p = Paragraph(evento.informacoes_adicionais, description_style)
            w, h = info_p.wrap(panel_width * 0.9, 100*mm)
            info_p.drawOn(c, panel3_x + (panel_width - w) / 2, y_pos - h)
        else:
            # Mensagem genérica
            c.setFont("Helvetica", 9)
            c.setFillColor(text_medium)
            default_text = "Para mais informações, entre em contato conosco ou visite nosso site."
            default_p = Paragraph(default_text, description_style)
            w, h = default_p.wrap(panel_width * 0.9, 50*mm)
            default_p.drawOn(c, panel3_x + (panel_width - w) / 2, y_pos - h)
    
    # Salvar o PDF
    c.save()
    
    # Retornar o arquivo para download
    return send_file(pdf_path, as_attachment=True, download_name=f"Folder_{evento.nome.replace(' ', '_')}.pdf")

@routes.route('/excluir_cliente/<int:cliente_id>', methods=['POST'])
@login_required
def excluir_cliente(cliente_id):
    if current_user.tipo != 'admin':
        flash("Apenas administradores podem excluir clientes.", "danger")
        return redirect(url_for('routes.dashboard'))

    cliente = Cliente.query.get_or_404(cliente_id)

    try:
        from sqlalchemy import or_, text
        from models import (
            Usuario, Oficina, OficinaDia, Checkin, Inscricao, MaterialOficina, RelatorioOficina,
            Ministrante, Evento, Patrocinador, CampoPersonalizadoCadastro, LinkCadastro,
            ConfiguracaoCliente, CertificadoTemplate, Feedback, RespostaCampo, RespostaFormulario,
            ConfiguracaoAgendamento, HorarioVisitacao, SalaVisitacao
        )

        # ===============================
        # 1️⃣ PARTICIPANTES
        # ===============================
        participantes = Usuario.query.filter_by(cliente_id=cliente.id).all()

        with db.session.no_autoflush:
            for usuario in participantes:
                Checkin.query.filter_by(usuario_id=usuario.id).delete()
                Inscricao.query.filter_by(usuario_id=usuario.id).delete()
                Feedback.query.filter_by(usuario_id=usuario.id).delete()
                RespostaCampo.query.filter_by(resposta_formulario_id=usuario.id).delete()
                RespostaFormulario.query.filter_by(usuario_id=usuario.id).delete()

        Usuario.query.filter_by(cliente_id=cliente.id).delete()

        # ===============================
        # 2️⃣ OFICINAS
        # ===============================
        oficinas = Oficina.query.filter_by(cliente_id=cliente.id).all()

        for oficina in oficinas:
            Checkin.query.filter_by(oficina_id=oficina.id).delete()
            Inscricao.query.filter_by(oficina_id=oficina.id).delete()
            OficinaDia.query.filter_by(oficina_id=oficina.id).delete()
            MaterialOficina.query.filter_by(oficina_id=oficina.id).delete()
            RelatorioOficina.query.filter_by(oficina_id=oficina.id).delete()

            db.session.execute(
                text('DELETE FROM oficina_ministrantes_association WHERE oficina_id = :oficina_id'),
                {'oficina_id': oficina.id}
            )

            db.session.delete(oficina)

        # ===============================
        # 3️⃣ MINISTRANTES
        # ===============================
        Ministrante.query.filter_by(cliente_id=cliente.id).delete()

        # ===============================
        # 4️⃣ EVENTOS E DEPENDÊNCIAS
        # ===============================
        eventos = Evento.query.filter(
            or_(Evento.cliente_id == cliente.id, Evento.cliente_id == None)
        ).all()

        for evento in eventos:
            with db.session.no_autoflush:
                HorarioVisitacao.query.filter_by(evento_id=evento.id).delete()
                ConfiguracaoAgendamento.query.filter_by(evento_id=evento.id).delete()
                Patrocinador.query.filter_by(evento_id=evento.id).delete()
                SalaVisitacao.query.filter_by(evento_id=evento.id).delete()  # ✅ NOVO

                db.session.delete(evento)

        # ===============================
        # 5️⃣ CONFIGURAÇÕES E VINCULAÇÕES
        # ===============================
        CertificadoTemplate.query.filter_by(cliente_id=cliente.id).delete()
        CampoPersonalizadoCadastro.query.filter_by(cliente_id=cliente.id).delete()
        LinkCadastro.query.filter_by(cliente_id=cliente.id).delete()
        ConfiguracaoCliente.query.filter_by(cliente_id=cliente.id).delete()

        # ===============================
        # 6️⃣ EXCLUI O CLIENTE
        # ===============================
        db.session.delete(cliente)
        db.session.commit()

        flash('Cliente e todos os dados vinculados foram excluídos com sucesso!', 'success')

    except Exception as e:
        db.session.rollback()
        flash(f"Erro ao excluir cliente: {str(e)}", 'danger')

    return redirect(url_for('routes.dashboard'))

@routes.route('/admin/evento/<int:evento_id>/inscritos')
@login_required
def listar_inscritos_evento(evento_id):
    if current_user.tipo != 'cliente':
        flash("Acesso restrito.", "danger")
        return redirect(url_for('routes.dashboard'))

    evento = Evento.query.get_or_404(evento_id)
    inscricoes = Inscricao.query.filter_by(evento_id=evento.id).all()

    return render_template('listar_inscritos_evento.html', evento=evento, inscricoes=inscricoes)


@routes.route('/admin/inscricao/<int:inscricao_id>/editar', methods=['GET', 'POST'])
@login_required
def editar_inscricao_evento(inscricao_id):
    inscricao = Inscricao.query.get_or_404(inscricao_id)
    tipos = InscricaoTipo.query.filter_by(oficina_id=inscricao.oficina_id).all()

    if request.method == 'POST':
        tipo_id = request.form.get('tipo_inscricao_id')
        inscricao.tipo_inscricao_id = tipo_id
        db.session.commit()
        flash("Inscrição atualizada com sucesso!", "success")
        return redirect(url_for('listar_inscritos_evento', evento_id=inscricao.evento_id))

    return render_template('editar_inscricao_evento.html', inscricao=inscricao, tipos=tipos)


@routes.route('/admin/inscricao/<int:inscricao_id>/excluir', methods=['POST'])
@login_required
def excluir_inscricao_evento(inscricao_id):
    inscricao = Inscricao.query.get_or_404(inscricao_id)
    evento_id = inscricao.evento_id
    db.session.delete(inscricao)
    db.session.commit()
    flash("Inscrição excluída com sucesso.", "warning")
    return redirect(url_for('listar_inscritos_evento', evento_id=evento_id))

@routes.route('/leitor_checkin_json', methods=['POST'])
@login_required
def leitor_checkin_json():
    """
    Recebe o token do QR‑Code, grava o check‑in
    (evento ou oficina) e avisa via Socket.IO.
    Sempre grava o cliente_id para que apareça
    na lista filtrada por cliente.
    """
    from datetime import datetime      # ← agora importado
    import sys

    data = request.get_json(silent=True) or {}
    token = (data.get('token') or '').strip()

    if not token:
        return jsonify(status='error',
                       message='Token não fornecido!'), 400

    # procura a inscrição pelo token
    inscricao = Inscricao.query.filter_by(qr_code_token=token).first()
    if not inscricao:
        return jsonify(status='error',
                       message='Inscrição não encontrada.'), 404

    cliente_id   = inscricao.cliente_id        # já existe no modelo :contentReference[oaicite:0]{index=0}
    sala_cliente = f"cliente_{cliente_id}"     # sala usada no Socket.IO

    # ---------- EVENTO ----------
    if inscricao.evento_id:
        # evita duplicidade
        if Checkin.query.filter_by(usuario_id=inscricao.usuario_id,
                                   evento_id=inscricao.evento_id).first():
            return jsonify(status='warning',
                           message='Check‑in do evento já foi realizado!')

        novo = Checkin(usuario_id = inscricao.usuario_id,
                       evento_id  = inscricao.evento_id,
                       cliente_id = cliente_id,          # ★ grava aqui
                       palavra_chave = "QR‑EVENTO")
    # ---------- OFICINA ----------
    elif inscricao.oficina_id:
        if Checkin.query.filter_by(usuario_id=inscricao.usuario_id,
                                   oficina_id=inscricao.oficina_id).first():
            return jsonify(status='warning',
                           message='Check‑in da oficina já foi realizado!')

        novo = Checkin(usuario_id = inscricao.usuario_id,
                       oficina_id = inscricao.oficina_id,
                       cliente_id = cliente_id,          # ★ grava aqui
                       palavra_chave = "QR‑OFICINA")
    else:
        return jsonify(status='error',
                       message='Inscrição sem evento ou oficina.'), 400

    db.session.add(novo)
    db.session.commit()

    # prepara carga para front‑end
    payload = {
        'participante': inscricao.usuario.nome,
        'data_hora'   : novo.data_hora.strftime('%d/%m/%Y %H:%M:%S')
    }
    if novo.evento_id:
        payload['evento'] = inscricao.evento.nome
    else:
        payload['oficina'] = inscricao.oficina.titulo

    # emite apenas para quem estiver na sala do cliente
    socketio.emit('novo_checkin', payload,
                  namespace='/checkins', room=sala_cliente)

    return jsonify(status='success',
                   message='Check‑in realizado com sucesso!',
                   **payload)


from utils import formatar_brasilia  # coloque no topo do routes.py

@routes.route('/lista_checkins_json')
@login_required
def lista_checkins_json():
    """
    Retorna os últimos check‑ins do cliente logado em formato JSON.
    Identifica se o check‑in é de evento ou de oficina.
    """
    if current_user.is_cliente():
        base_query = Checkin.query.filter_by(cliente_id=current_user.id)
    else:
        base_query = Checkin.query

    checkins = (base_query
                .order_by(Checkin.data_hora.desc())
                .limit(50)
                .all())

    resultado = []
    for c in checkins:
        data_formatada = formatar_brasilia(c.data_hora)

        if c.evento_id:
            resultado.append({
                'participante': c.usuario.nome,
                'evento'      : c.evento.nome if c.evento else "Evento Desconhecido",
                'data_hora'   : data_formatada,
                'tipo_checkin': 'evento'
            })
        elif c.oficina_id:
            resultado.append({
                'participante': c.usuario.nome,
                'oficina'     : c.oficina.titulo if c.oficina else "Oficina Desconhecida",
                'data_hora'   : data_formatada,
                'tipo_checkin': 'oficina'
            })
        else:
            resultado.append({
                'participante': c.usuario.nome,
                'atividade'   : "N/A",
                'data_hora'   : data_formatada,
                'tipo_checkin': 'nenhum'
            })

    return jsonify(status='success', checkins=resultado)


@routes.route('/remover_fotos_selecionadas', methods=['POST'])
def remover_fotos_selecionadas():
    # Recupera a lista de IDs selecionados no checkbox
    ids_selecionados = request.form.getlist('fotos_selecionadas')

    if not ids_selecionados:
        flash('Nenhuma imagem foi selecionada para remoção.', 'warning')
        return redirect(url_for('routes.gerenciar_patrocinadores'))
    
    # Converte cada id para inteiro
    ids_selecionados = [int(x) for x in ids_selecionados]

    # Busca cada patrocinador correspondente e remove do DB
    for pat_id in ids_selecionados:
        pat = Patrocinador.query.get(pat_id)
        if pat:
            # Se quiser, remova também o arquivo de imagem físico, se armazenar localmente
            # Exemplo: os.remove(os.path.join(current_app.static_folder, pat.logo_path))

            db.session.delete(pat)
    
    db.session.commit()
    flash('Fotos removidas com sucesso!', 'success')
    return redirect(url_for('routes.gerenciar_patrocinadores'))
    # Se não houver fotos selecionadas, redireciona para a página de gerenciamento
    
@routes.route('/gerar_evento_qrcode_pdf/<int:evento_id>')
@login_required
def gerar_evento_qrcode_pdf(evento_id):
    """
    Gera um PDF contendo o QR Code do evento para credenciamento.
    O PDF terá informações do evento e do participante, além do código QR.
    """

    import os
    import uuid
    from datetime import datetime
    from flask import send_file, flash, redirect, url_for
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    import qrcode

    # 1) Verifica se há configuração do cliente e se está habilitado o QR Code de evento
    config_cliente = ConfiguracaoCliente.query.filter_by(cliente_id=current_user.cliente_id).first()
    if not config_cliente or not config_cliente.habilitar_qrcode_evento_credenciamento:
        flash("A geração de QR Code para credenciamento de evento está desabilitada para este cliente.", "danger")
        return redirect(url_for('routes.dashboard_participante'))

    # 2) Localiza o evento
    evento = Evento.query.get_or_404(evento_id)

    # 3) Verifica se o participante está inscrito nesse evento, senão cria a inscrição automaticamente
    inscricao = Inscricao.query.filter_by(usuario_id=current_user.id, evento_id=evento_id).first()
    if not inscricao:
        inscricao = Inscricao(
            usuario_id=current_user.id,
            cliente_id=current_user.cliente_id,
            evento_id=evento_id
        )
        db.session.add(inscricao)
        db.session.commit()

    # 4) Caso não tenha token gerado, gera agora
    if not inscricao.qr_code_token:
        novo_token = str(uuid.uuid4())
        inscricao.qr_code_token = novo_token
        db.session.commit()

    token = inscricao.qr_code_token

    # 5) Gera a imagem do QR Code
    output_dir = os.path.join("static", "tmp")
    os.makedirs(output_dir, exist_ok=True)
    qr_filename = f"qr_evento_{evento_id}_user_{current_user.id}.png"
    qr_path = os.path.join(output_dir, qr_filename)

    # Criando QR code com estilo
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_H,
        box_size=10,
        border=4,
    )
    qr.add_data(token)
    qr.make(fit=True)

    # Criando QR code colorido
    qr_img = qr.make_image(fill_color="#0066CC", back_color="white")
    qr_img.save(qr_path)

    # 6) Cria um PDF com ReportLab
    pdf_filename = f"evento_{evento_id}_qrcode_{current_user.id}.pdf"
    pdf_output_dir = os.path.join("static", "comprovantes")
    os.makedirs(pdf_output_dir, exist_ok=True)
    pdf_path = os.path.join(pdf_output_dir, pdf_filename)

    # Registro de fontes personalizadas (se disponíveis)
    # Descomente as linhas abaixo se tiver arquivos de fontes
    # pdfmetrics.registerFont(TTFont('Montserrat', 'static/fonts/Montserrat-Regular.ttf'))
    # pdfmetrics.registerFont(TTFont('MontserratBold', 'static/fonts/Montserrat-Bold.ttf'))

    # Cores e estilos
    cor_primaria = colors.HexColor("#0066CC")  # Azul
    cor_secundaria = colors.HexColor("#333333")  # Cinza escuro
    cor_destaque = colors.HexColor("#FF9900")  # Laranja

    # Configurações do PDF
    c = canvas.Canvas(pdf_path, pagesize=A4)
    width, height = A4
    c.setTitle("Comprovante de Inscrição - " + evento.nome)

    # Fundo do cabeçalho
    c.setFillColor(cor_primaria)
    c.rect(0, height-4*cm, width, 4*cm, fill=True, stroke=False)

    # Título no cabeçalho
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 22)
    c.drawCentredString(width/2, height-2.5*cm, "COMPROVANTE DE INSCRIÇÃO")

    # Logo da organização (se disponível)
    # c.drawImage("static/img/logo.png", 1*cm, height-3*cm, width=2*cm, height=2*cm)

    # Data e hora de geração do comprovante
    import datetime
    data_geracao = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
    c.setFont("Helvetica", 9)
    c.drawRightString(width-1*cm, height-3.5*cm, f"Gerado em: {data_geracao}")

    # Caixa principal de conteúdo
    c.setFillColor(colors.white)
    c.roundRect(1*cm, 5*cm, width-2*cm, height-10*cm, 10, fill=True, stroke=False)

    # Sombra sutil para a caixa
    c.setFillColor(colors.HexColor("#EEEEEE"))
    c.roundRect(1.1*cm, 4.9*cm, width-2*cm, height-10*cm, 10, fill=True, stroke=False)
    c.setFillColor(colors.white)
    c.roundRect(1*cm, 5*cm, width-2*cm, height-10*cm, 10, fill=True, stroke=False)

    # Informações do evento
    c.setFillColor(cor_secundaria)
    c.setFont("Helvetica-Bold", 18)
    c.drawString(2*cm, height-5.5*cm, evento.nome)

    # Linha decorativa
    c.setStrokeColor(cor_destaque)
    c.setLineWidth(3)
    c.line(2*cm, height-5.8*cm, width-2*cm, height-5.8*cm)

    # Informações detalhadas
    c.setFont("Helvetica-Bold", 12)
    c.setFillColor(cor_primaria)
    c.drawString(2*cm, height-7*cm, "DETALHES DO EVENTO:")

    c.setFillColor(cor_secundaria)
    c.setFont("Helvetica", 12)
    data_evento = evento.data_inicio.strftime("%d/%m/%Y") if evento.data_inicio else "Data indefinida"
    hora_evento = evento.hora_inicio.strftime("%H:%M") if hasattr(evento, 'hora_inicio') and evento.hora_inicio else "Horário não especificado"
    localizacao = evento.localizacao or "Local não especificado"

    y_position = height-8*cm
    c.setFont("Helvetica-Bold", 10)
    c.drawString(2*cm, y_position, "Data:")
    c.setFont("Helvetica", 10)
    c.drawString(4*cm, y_position, data_evento)

    y_position -= 0.7*cm
    c.setFont("Helvetica-Bold", 10)
    c.drawString(2*cm, y_position, "Horário:")
    c.setFont("Helvetica", 10)
    c.drawString(4*cm, y_position, hora_evento)

    y_position -= 0.7*cm
    c.setFont("Helvetica-Bold", 10)
    c.drawString(2*cm, y_position, "Local:")
    c.setFont("Helvetica", 10)
    # Gerencia texto de localização longo para não sobrepor o QR code
    # Calcula a largura máxima disponível para o texto (evitando a área do QR code)
    texto_max_largura = width - 15*cm  # Deixa margem para o QR code à direita

    # Importa ferramentas para texto multilinha
    from reportlab.platypus import Paragraph
    from reportlab.lib.styles import ParagraphStyle

    # Define o estilo do parágrafo
    style = ParagraphStyle(
        name='Normal',
        fontName='Helvetica',
        fontSize=10,
        leading=12  # Espaçamento entre linhas
    )

    # Cria o parágrafo com o texto da localização
    p = Paragraph(localizacao, style)

    # Organiza o parágrafo dentro do espaço disponível
    text_width, text_height = p.wrapOn(c, texto_max_largura, 4*cm)

    # Desenha o parágrafo
    p.drawOn(c, 4*cm, y_position - text_height + 10)

    # Ajusta a posição vertical baseada na altura do texto
    y_position -= (text_height + 0.3*cm)

    # Informações do participante
    y_position -= 0.7*cm
    c.setFont("Helvetica-Bold", 12)
    c.setFillColor(cor_primaria)
    c.drawString(2*cm, y_position, "DADOS DO PARTICIPANTE:")
    y_position -= 0.7*cm

    c.setFillColor(cor_secundaria)
    c.setFont("Helvetica-Bold", 10)
    c.drawString(2*cm, y_position, "Nome:")
    c.setFont("Helvetica", 10)
    c.drawString(4*cm, y_position, current_user.nome)

    # Adicionar e-mail se disponível
    if hasattr(current_user, 'email'):
        y_position -= 0.7*cm
        c.setFont("Helvetica-Bold", 10)
        c.drawString(2*cm, y_position, "E-mail:")
        c.setFont("Helvetica", 10)
        c.drawString(4*cm, y_position, current_user.email)

    # Código de inscrição
    y_position -= 0.7*cm
    c.setFont("Helvetica-Bold", 10)
    c.drawString(2*cm, y_position, "Código:")
    c.setFont("Helvetica", 10)
    codigo_inscricao = token[:8].upper()  # Primeiros 8 caracteres do token
    c.drawString(4*cm, y_position, codigo_inscricao)

    # QR Code com título
    c.setFont("Helvetica-Bold", 12)
    c.setFillColor(cor_primaria)
    c.drawString(width-7*cm, height-7*cm, "QR CODE DE ACESSO")

    # Borda decorativa ao redor do QR Code
    c.setStrokeColor(cor_destaque)
    c.setLineWidth(2)
    c.roundRect(width-7*cm, height-13*cm, 5*cm, 5*cm, 5, fill=False, stroke=True)

    # Inserir o QR Code
    c.drawImage(qr_path, width-6.5*cm, height-12.5*cm, width=4*cm, height=4*cm)

    # Instruções
    c.setFont("Helvetica-Oblique", 9)
    c.setFillColor(cor_secundaria)
    c.drawCentredString(width-4.5*cm, height-13.5*cm, "Apresente este QR Code na entrada do evento")

    # Rodapé
    c.setFillColor(cor_primaria)
    c.rect(0, 0, width, 2*cm, fill=True, stroke=False)

    c.setFillColor(colors.white)
    c.setFont("Helvetica", 9)
    c.drawCentredString(width/2, 1*cm, "Este é um comprovante oficial de inscrição. Em caso de dúvidas, entre em contato conosco.")

    # Número da inscrição
    numero_inscricao = f"#{current_user.id:06d}"
    c.setFont("Helvetica-Bold", 10)
    c.drawString(width-3*cm, 1*cm, numero_inscricao)

    # Finaliza o PDF
    c.showPage()
    c.save()

    # 7) Retorna o PDF para download
    return send_file(pdf_path, as_attachment=True, download_name=pdf_filename)



@routes.route("/toggle_qrcode_evento_credenciamento", methods=["POST"])
@login_required
def toggle_qrcode_evento_credenciamento():
    # Garante que apenas o cliente (dono) possa mudar
    if current_user.tipo != 'cliente':
        return jsonify({"success": False, "message": "Acesso negado!"}), 403

    # Busca (ou cria) a configuração desse cliente
    config_cliente = ConfiguracaoCliente.query.filter_by(cliente_id=current_user.id).first()
    if not config_cliente:
        config_cliente = ConfiguracaoCliente(
            cliente_id=current_user.id,
            permitir_checkin_global=False,
            habilitar_feedback=False,
            habilitar_certificado_individual=False,
            habilitar_qrcode_evento_credenciamento=False
        )
        db.session.add(config_cliente)
        db.session.commit()

    # Inverte o valor atual
    config_cliente.habilitar_qrcode_evento_credenciamento = not config_cliente.habilitar_qrcode_evento_credenciamento
    db.session.commit()

    return jsonify({
        "success": True,
        "value": config_cliente.habilitar_qrcode_evento_credenciamento,
        "message": "Habilitação de QRCode de Evento atualizada!"
    })
    
from flask import current_app  # já pode deixar no topo

@routes.route("/exportar_checkins_evento_pdf/<int:evento_id>")
@login_required
def exportar_checkins_evento_pdf(evento_id):
    import os
    from io import BytesIO
    from flask import send_file, flash, redirect, url_for
    from datetime import datetime
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm, mm
    from reportlab.lib import colors
    from reportlab.platypus import Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus.flowables import Image
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    from reportlab.platypus.doctemplate import SimpleDocTemplate
    from reportlab.platypus.frames import Frame
    from reportlab.platypus.tableofcontents import TableOfContents

    print(f"[DEBUG] Usuário acessou exportação de check-ins do evento ID {evento_id}")

    if current_user.is_cliente():
            evento = Evento.query.filter_by(id=evento_id, cliente_id=current_user.id).first()
            if not evento:
                flash("Evento não encontrado ou não pertence ao seu acesso.", "danger")
                print("[DEBUG] Evento não pertence ao cliente logado.")
                return redirect(url_for("routes.dashboard_cliente"))
    else:
        evento = Evento.query.get_or_404(evento_id)

    print(f"[DEBUG] Evento encontrado: {evento.nome} (ID: {evento.id})")

    checkins = (
        Checkin.query
        .filter_by(evento_id=evento.id)
        .filter(Checkin.palavra_chave == 'QR-EVENTO')
        .all()
    )
    print(f"[DEBUG] Total de check-ins encontrados: {len(checkins)}")

    if not checkins:
        flash("Nenhum check-in encontrado para este evento.", "warning")
        return redirect(url_for("routes.dashboard_cliente"))

    # DEBUG de alguns dados dos check-ins
    for c in checkins[:5]:
        print(f"[DEBUG] Check-in: Nome={c.usuario.nome}, Data={c.data_hora}, Palavra-chave={c.palavra_chave}")

    # ...continua o restante da geração do PDF normalmente

    # Cria PDF em memória
    buffer = BytesIO()
    
    # Define cores do tema do relatório
    cor_primaria = colors.HexColor("#1E88E5")  # Azul moderno
    cor_secundaria = colors.HexColor("#E0E0E0")  # Cinza claro
    cor_destaque = colors.HexColor("#FF5722")   # Laranja para destaques
    cor_texto = colors.HexColor("#333333")      # Cinza escuro para texto
    
    # Configuração do documento
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        title=f"Check-ins do Evento: {evento.nome}",
        author="Sistema de Eventos",
        leftMargin=2*cm,
        rightMargin=2*cm,
        topMargin=2*cm,
        bottomMargin=2*cm
    )
    
    # Estilos
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        name='Titulo',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=16,
        textColor=cor_primaria,
        spaceAfter=12,
        spaceBefore=12,
    ))
    
    styles.add(ParagraphStyle(
        name='Subtitulo',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=12,
        textColor=cor_primaria,
        spaceAfter=6,
    ))
    
    # Modificar o estilo Normal existente em vez de adicionar um novo
    styles['Normal'].fontName = 'Helvetica'
    styles['Normal'].fontSize = 10
    styles['Normal'].textColor = cor_texto
    styles['Normal'].spaceAfter = 6
    
    styles.add(ParagraphStyle(
        name='InfoEvento',
        parent=styles['Normal'],
        fontName='Helvetica-Oblique',
        fontSize=10,
        textColor=cor_texto,
        spaceAfter=10,
    ))
    
    styles.add(ParagraphStyle(
        name='Rodape',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8,
        textColor=cor_texto,
        alignment=TA_CENTER,
    ))
    
    # Lista para elementos do documento
    elementos = []
    
    # Função para cabeçalho e rodapé
    def adicionar_cabecalho_rodape(canvas, doc):
        canvas.saveState()
        largura, altura = A4
        
        # Cabeçalho
        canvas.setFillColor(cor_primaria)
        canvas.rect(0, altura - 1.5*cm, largura, 1.5*cm, fill=1)
        
        canvas.setFillColor(colors.white)
        canvas.setFont('Helvetica-Bold', 14)
        canvas.drawString(2*cm, altura - 1*cm, "RELATÓRIO DE CHECK-INS")
        
        # Rodapé
        canvas.setFillColor(cor_secundaria)
        canvas.rect(0, 0, largura, 1*cm, fill=1)
        
        canvas.setFillColor(cor_texto)
        canvas.setFont('Helvetica', 8)
        canvas.drawString(2*cm, 0.5*cm, f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
        canvas.drawString(largura/2, 0.5*cm, f"Página {doc.page}")
        
        # Logo ou ícone (exemplo)
        # Se tiver um logo, substituir a linha abaixo
        canvas.setFillColor(cor_destaque)
        canvas.circle(largura - 3*cm, altura - 0.75*cm, 0.5*cm, fill=1)
        
        canvas.restoreState()
    
    # Título do relatório
    elementos.append(Paragraph(f"Relatório de Check-ins", styles['Titulo']))
    
    # Informações do evento
    elementos.append(Paragraph(f"<b>Evento:</b> {evento.nome}", styles['Subtitulo']))
    
    # Adicionar mais informações do evento se disponíveis
    if hasattr(evento, 'data'):
        elementos.append(Paragraph(f"<b>Data do evento:</b> {evento.data.strftime('%d/%m/%Y')}", styles['InfoEvento']))
    if hasattr(evento, 'local'):
        elementos.append(Paragraph(f"<b>Local:</b> {evento.local}", styles['InfoEvento']))
    
    elementos.append(Paragraph(f"<b>Total de check-ins:</b> {len(checkins)}", styles['InfoEvento']))
    elementos.append(Spacer(1, 0.5*cm))
    
    # Resumo estatístico se houver dados suficientes
    if len(checkins) > 1:
        # Dados para tabela de resumo
        resumo_data = []
        
        # Exemplo: distribuição por hora (ajuste conforme necessário)
        hora_counts = {}
        for checkin in checkins:
            hora = checkin.data_hora.hour
            hora_counts[hora] = hora_counts.get(hora, 0) + 1
        
        # Mostrar distribuição por horas se relevante
        if len(hora_counts) > 1:
            elementos.append(Paragraph("Distribuição de check-ins por hora:", styles['Subtitulo']))
            
            # Criar dados para tabela de distribuição
            hora_data = [["Hora", "Quantidade", "Percentual"]]
            for hora in sorted(hora_counts.keys()):
                qtd = hora_counts[hora]
                percentual = (qtd / len(checkins)) * 100
                hora_str = f"{hora}:00 - {hora+1}:00"
                hora_data.append([hora_str, str(qtd), f"{percentual:.1f}%"])
            
            # Criar e estilizar tabela de distribuição
            tabela_horas = Table(hora_data, colWidths=[5*cm, 3*cm, 3*cm])
            tabela_horas.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), cor_primaria),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.5, cor_texto),
                ('BOX', (0, 0), (-1, -1), 1, cor_primaria),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, cor_secundaria.clone(alpha=0.3)]),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
            ]))
            
            elementos.append(tabela_horas)
            elementos.append(Spacer(1, 0.5*cm))
    
    # Tabela principal de check-ins
    elementos.append(Paragraph("Lista de Check-ins Realizados:", styles['Subtitulo']))
    elementos.append(Spacer(1, 0.2*cm))
    
    # Dados para tabela principal
    dados_tabela = [["Nome", "CPF", "Data/Hora", "Palavra-chave"]]
    
    # Preencher dados na tabela
    for checkin in checkins:
        usuario = checkin.usuario
        dados_tabela.append([
            usuario.nome[:40],
            usuario.cpf or "-",
            checkin.data_hora.strftime('%d/%m/%Y %H:%M'),
            checkin.palavra_chave or "-"
        ])
    
    # Criar e estilizar tabela
    tabela = Table(dados_tabela, colWidths=[5*cm, 3*cm, 4*cm, 4*cm])
    tabela.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), cor_primaria),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, cor_texto),
        ('BOX', (0, 0), (-1, -1), 1, cor_primaria),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, cor_secundaria.clone(alpha=0.2)]),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
    ]))
    
    elementos.append(tabela)
    
    # Adicionar informações finais ou observações
    elementos.append(Spacer(1, 1*cm))
    elementos.append(Paragraph("* Este relatório é gerado automaticamente pelo sistema de eventos.", styles['Rodape']))
    
    # Construir documento com cabeçalho e rodapé personalizados
    doc.build(elementos, onFirstPage=adicionar_cabecalho_rodape, onLaterPages=adicionar_cabecalho_rodape)
    
    buffer.seek(0)

    filename = f"checkins_evento_{evento.id}.pdf"
    return send_file(buffer, as_attachment=True, download_name=filename, mimetype='application/pdf')


@routes.route('/submeter_trabalho', methods=['GET', 'POST'])
@login_required
def submeter_trabalho():
    if current_user.tipo != 'participante':
        flash('Apenas participantes podem submeter trabalhos.', 'danger')
        return redirect(url_for('routes.dashboard_participante'))

    if request.method == 'POST':
        titulo = request.form.get('titulo')
        resumo = request.form.get('resumo')
        area_tematica = request.form.get('area_tematica')
        arquivo = request.files.get('arquivo_pdf')

        if not all([titulo, resumo, area_tematica, arquivo]):
            flash('Todos os campos são obrigatórios!', 'danger')
            return redirect(url_for('routes.submeter_trabalho'))

        filename = secure_filename(arquivo.filename)
        caminho_pdf = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        arquivo.save(caminho_pdf)

        trabalho = TrabalhoCientifico(
            titulo=titulo,
            resumo=resumo,
            area_tematica=area_tematica,
            arquivo_pdf=caminho_pdf,
            usuario_id=current_user.id,
            evento_id=current_user.evento_id  # opcionalmente usa o evento do usuário
        )
        db.session.add(trabalho)
        db.session.commit()

        flash("Trabalho submetido com sucesso!", "success")
        return redirect(url_for('routes.dashboard_participante'))

    return render_template('submeter_trabalho.html')


@routes.route('/avaliar_trabalhos')
@login_required
def avaliar_trabalhos():
    if current_user.tipo != 'cliente' and not current_user.is_superuser():
        flash('Apenas administradores ou avaliadores têm acesso.', 'danger')
        return redirect(url_for('routes.dashboard'))

    trabalhos = TrabalhoCientifico.query.filter(TrabalhoCientifico.status != 'aceito').all()
    return render_template('avaliar_trabalhos.html', trabalhos=trabalhos)

@routes.route('/avaliar_trabalho/<int:trabalho_id>', methods=['GET', 'POST'])
@login_required
def avaliar_trabalho(trabalho_id):
    trabalho = TrabalhoCientifico.query.get_or_404(trabalho_id)

    if request.method == 'POST':
        estrelas = request.form.get('estrelas')
        nota = request.form.get('nota')
        conceito = request.form.get('conceito')
        comentario = request.form.get('comentario')

        avaliacao = AvaliacaoTrabalho(
            trabalho_id=trabalho.id,
            avaliador_id=current_user.id,
            estrelas=int(estrelas) if estrelas else None,
            nota=float(nota) if nota else None,
            conceito=conceito,
            comentario=comentario
        )
        db.session.add(avaliacao)

        trabalho.status = 'avaliado'
        db.session.commit()

        flash("Avaliação registrada!", "success")
        return redirect(url_for('routes.avaliar_trabalhos'))

    return render_template('avaliar_trabalho.html', trabalho=trabalho)


@routes.route('/meus_trabalhos')
@login_required
def meus_trabalhos():
    if current_user.tipo != 'participante':
        return redirect(url_for('routes.dashboard_participante'))

    trabalhos = TrabalhoCientifico.query.filter_by(usuario_id=current_user.id).all()
    return render_template('meus_trabalhos.html', trabalhos=trabalhos)

@routes.route('/submeter_trabalho', methods=['GET', 'POST'])
@login_required
def nova_submissao():
    if current_user.tipo != 'participante':
        flash('Apenas participantes podem submeter trabalhos.', 'danger')
        return redirect(url_for('routes.dashboard'))

    if request.method == 'POST':
        titulo = request.form.get('titulo')
        resumo = request.form.get('resumo')
        area_tematica = request.form.get('area_tematica')
        arquivo = request.files.get('arquivo_pdf')

        if not all([titulo, resumo, area_tematica, arquivo]):
            flash('Todos os campos são obrigatórios!', 'warning')
            return redirect(url_for('routes.nova_submissao'))

        # Garante diretório e salva o arquivo PDF
        filename = secure_filename(arquivo.filename)
        caminho_pdf = os.path.join('static/uploads/trabalhos', filename)
        os.makedirs(os.path.dirname(caminho_pdf), exist_ok=True)
        arquivo.save(caminho_pdf)

        # Registra trabalho no banco
        trabalho = TrabalhoCientifico(
            titulo=titulo,
            resumo=resumo,
            area_tematica=area_tematica,
            arquivo_pdf=caminho_pdf,
            usuario_id=current_user.id,
            evento_id=current_user.evento_id if hasattr(current_user, 'evento_id') else None
        )

        db.session.add(trabalho)
        db.session.commit()

        flash('Trabalho submetido com sucesso!', 'success')
        return redirect(url_for('routes.meus_trabalhos'))
    
@routes.route('/toggle_submissao_trabalhos')
@login_required
def toggle_submissao_trabalhos_cliente():
    if current_user.tipo != 'cliente':
        flash("Apenas clientes podem alterar essa configuração.", "warning")
        return redirect(url_for('routes.dashboard'))

    config = current_user.configuracao

    if not config:
        flash("Cliente sem configuração associada.", "danger")
        return redirect(url_for('routes.dashboard'))

    config.habilitar_submissao_trabalhos = not config.habilitar_submissao_trabalhos
    db.session.commit()
    flash("Configuração de submissão de trabalhos atualizada!", "success")
    return redirect(url_for('routes.dashboard'))

  
# routes.py  (já existe)
@routes.route("/pagamento_sucesso")
def pagamento_sucesso():
    payment_id        = request.args.get("payment_id")
    external_ref      = request.args.get("external_reference")  # id da inscrição
    inscricao         = Inscricao.query.get_or_404(external_ref)

    # 1) marca como aprovado, se ainda não estiver
    if inscricao.status_pagamento != "approved":
        inscricao.status_pagamento = "approved"
        db.session.commit()

    # 2) faz login do usuário automaticamente
    login_user(inscricao.usuario)
    session['user_type'] = 'participante'

    flash("Pagamento aprovado! Bem‑vindo(a) 😉", "success")
    return redirect(url_for("routes.dashboard_participante"))



@routes.route("/pagamento_pendente")
def pagamento_pendente():
    flash("Pagamento pendente. Você poderá concluir mais tarde.", "warning")
    return redirect(url_for("routes.login"))


@routes.route("/pagamento_falha")
def pagamento_falha():
    flash("Pagamento não foi concluído. Tente novamente.", "danger")
    return redirect(url_for("routes.login"))


@routes.route("/webhook_mp", methods=["POST"])
def webhook_mp():
    data = request.get_json(silent=True) or {}
    if data.get("type") == "payment":
        sdk = mercadopago.SDK(os.getenv("MERCADOPAGO_ACCESS_TOKEN"))
        pay = sdk.payment().get(data["data"]["id"])["response"]

        if pay["status"] == "approved":
            ref = pay["external_reference"]
            insc = Inscricao.query.get(int(ref))
            if insc and insc.status_pagamento != "approved":
                insc.status_pagamento = "approved"
                db.session.commit()
    return "", 200  


from flask import request, send_file, redirect, url_for, flash
from io import BytesIO
from datetime import datetime
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from flask_login import login_required, current_user
from models import Checkin, Evento
import unicodedata

@routes.route('/exportar_checkins_pdf_opcoes', methods=['GET'])
@login_required
def exportar_checkins_pdf_opcoes():
    from flask import request
    import unicodedata
    from io import BytesIO
    from datetime import datetime
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from models import Checkin, Evento, Oficina

    def normalizar(texto):
        return unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode("utf-8").strip()

    tipo = normalizar(request.args.get('tipo', '').upper())
    evento_id = request.args.get('evento_id')

    if not current_user.is_cliente():
        flash("Acesso negado.", "danger")
        return redirect(url_for('routes.dashboard'))

    # Query base
    base_query = Checkin.query.filter(Checkin.cliente_id == current_user.id)

    if evento_id and evento_id != 'todos':
        base_query = base_query.filter(Checkin.evento_id == int(evento_id))

    if tipo and tipo != 'TODOS':
        base_query = base_query.filter(Checkin.palavra_chave == tipo)

    checkins = base_query.order_by(Checkin.data_hora.desc()).all()

    if not checkins:
        flash("Nenhum check-in encontrado para os filtros aplicados.", "info")
        return redirect(url_for('routes.dashboard'))

    # Gerar PDF
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    
    # Cores modernas
    azul_principal = colors.HexColor('#2196F3')
    azul_claro = colors.HexColor('#BBDEFB')
    cinza_escuro = colors.HexColor('#424242')
    cinza_claro = colors.HexColor('#EEEEEE')
    
    def adicionar_cabecalho_rodape(pagina, num_pagina, total_paginas):
        # Cabeçalho
        pagina.setFillColor(azul_principal)
        pagina.rect(0, height - 2*cm, width, 2*cm, fill=1)
        
        pagina.setFillColor(colors.white)
        pagina.setFont("Helvetica-Bold", 16)
        titulo_relatorio = "Relatório de Check-ins"
        if tipo != 'TODOS':
            titulo_relatorio += f" - {tipo}"
        pagina.drawCentredString(width/2, height - 1.2*cm, titulo_relatorio)
        
        pagina.setFont("Helvetica", 10)
        data_relatorio = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        pagina.drawString(width - 5*cm, height - 1.7*cm, data_relatorio)
        
        # Rodapé
        pagina.setFillColor(cinza_escuro)
        pagina.rect(0, 0, width, 1*cm, fill=1)
        pagina.setFillColor(colors.white)
        pagina.setFont("Helvetica", 8)
        pagina.drawString(1*cm, 0.4*cm, f"Página {num_pagina} de {total_paginas}")
        pagina.drawRightString(width - 1*cm, 0.4*cm, "Sistema de Check-ins")
    
    # Função para criar nova página
    def nova_pagina(num_pagina, total_paginas):
        p.showPage()
        adicionar_cabecalho_rodape(p, num_pagina, total_paginas)
        return height - 3*cm  # Retornar posição Y após o cabeçalho
    
    # Calcular número total de páginas (estimativa)
    # Vamos estimar 15 itens por página
    total_paginas = 1 + (len(checkins) // 15)
    
    # Agrupando check-ins por evento e oficina
    agrupados = {}
    for c in checkins:
        chave = f"EVENTO: {c.evento.nome}" if c.evento else f"OFICINA: {c.oficina.titulo if c.oficina else 'Sem título'}"
        if chave not in agrupados:
            agrupados[chave] = []
        agrupados[chave].append(c)
    
    # Iniciar primeira página
    pagina_atual = 1
    adicionar_cabecalho_rodape(p, pagina_atual, total_paginas)
    y = height - 3*cm  # Posição inicial após o cabeçalho
    
    for secao, lista in agrupados.items():
        # Verificar se tem espaço na página
        if y < 8*cm:
            y = nova_pagina(pagina_atual, total_paginas)
            pagina_atual += 1
        
        # Desenhar cabeçalho da seção
        p.setFont("Helvetica-Bold", 12)
        p.setFillColor(azul_principal)
        p.drawString(1*cm, y, secao)
        y -= 0.8*cm
        
        # Desenhar cabeçalho da tabela
        p.setFillColor(azul_principal)
        p.rect(1*cm, y - 0.7*cm, width - 2*cm, 0.7*cm, fill=1)
        
        p.setFillColor(colors.white)
        p.setFont("Helvetica-Bold", 10)
        p.drawString(1.2*cm, y - 0.4*cm, "Participante")
        p.drawString(7*cm, y - 0.4*cm, "Palavra-chave")
        p.drawString(13*cm, y - 0.4*cm, "Data/Hora")
        
        y -= 0.7*cm
        
        # Listar check-ins desta seção
        linha_alternada = True
        for c in lista:
            # Verificar se tem espaço na página
            if y < 2*cm:
                y = nova_pagina(pagina_atual, total_paginas)
                pagina_atual += 1
                
                # Redesenhar cabeçalho da tabela
                p.setFillColor(azul_principal)
                p.rect(1*cm, y - 0.7*cm, width - 2*cm, 0.7*cm, fill=1)
                
                p.setFillColor(colors.white)
                p.setFont("Helvetica-Bold", 10)
                p.drawString(1.2*cm, y - 0.4*cm, "Participante")
                p.drawString(7*cm, y - 0.4*cm, "Palavra-chave")
                p.drawString(13*cm, y - 0.4*cm, "Data/Hora")
                
                y -= 0.7*cm
            
            # Alternar cores de fundo para linhas
            if linha_alternada:
                p.setFillColor(cinza_claro)
                p.rect(1*cm, y - 0.5*cm, width - 2*cm, 0.5*cm, fill=1)
            linha_alternada = not linha_alternada
            
            # Desenhar dados
            p.setFillColor(cinza_escuro)
            p.setFont("Helvetica", 9)
            p.drawString(1.2*cm, y - 0.3*cm, c.usuario.nome[:28])
            p.drawString(7*cm, y - 0.3*cm, c.palavra_chave[:20])
            p.drawString(13*cm, y - 0.3*cm, c.data_hora.strftime('%d/%m/%Y %H:%M'))
            
            # Descer para próxima linha
            y -= 0.5*cm
        
        # Espaço entre seções
        y -= 1*cm
    
    # Finalizar PDF
    p.save()
    buffer.seek(0)
    
    nome_arquivo = f"checkins_{evento_id or 'todos'}_{tipo or 'todos'}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return send_file(buffer, as_attachment=True, download_name=nome_arquivo, mimetype='application/pdf')

# routes_admin.py  (ou onde guarda rotas de admin)
from decimal import Decimal
from flask_login import login_required, current_user
from models import Configuracao
from extensions import db

@routes.route("/atualizar_taxa", methods=["POST"])
@login_required
def atualizar_taxa():
    if current_user.tipo != "cliente":
        abort(403)

    valor = request.form.get("taxa_percentual", "0").replace(",", ".")
    try:
        perc = round(Decimal(valor), 2)          # 0–100
        assert 0 <= perc <= 100
    except:
        flash("Percentual inválido", "danger")
        return redirect(request.referrer or url_for("routes.dashboard_admin"))

    cfg = Configuracao.query.first()
    if not cfg:
        cfg = Configuracao()
        db.session.add(cfg)

    cfg.taxa_percentual_inscricao = perc
    db.session.commit()
    flash("Percentual salvo!", "success")
    return redirect(request.referrer or url_for("routes.dashboard_admin"))
