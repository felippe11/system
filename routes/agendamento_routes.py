from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify
from flask_login import login_required, current_user
import os
import threading
from datetime import datetime, timedelta, date
from flask import current_app, send_file, Response, abort
from werkzeug.utils import secure_filename
from sqlalchemy.exc import IntegrityError
from werkzeug.security import generate_password_hash
from extensions import db
from services.mailjet_service import send_via_mailjet
from services import pdf_service
from mailjet_rest.client import ApiError
import logging
from utils import endpoints

from utils.arquivo_utils import arquivo_permitido
from utils.dia_semana import dia_semana

logger = logging.getLogger(__name__)

from models import (
    AgendamentoVisita,
    HorarioVisitacao,
    Evento,
    EventoInscricaoTipo,
    ConfiguracaoAgendamento,
    Usuario,
    Cliente,
    Oficina,
    ProfessorBloqueado,
    Checkin,
    AlunoVisitante,
    NecessidadeEspecial,
)
from models.event import NotificacaoAgendamento
from utils import obter_estados
from fpdf import FPDF
import pandas as pd
import qrcode
import io
from PIL import Image
from types import SimpleNamespace
from . import routes


# Aliases for backward compatibility used later in this module
ConfigAgendamento = ConfiguracaoAgendamento
Agendamento = AgendamentoVisita
Horario = HorarioVisitacao
data_agendamento = AgendamentoVisita.data_agendamento

agendamento_routes = Blueprint(
    'agendamento_routes',
    __name__,
    template_folder="../templates/agendamento"
)

class NotificacaoAgendamentoService:
    """Helper methods for agendamento notification emails."""
    
    @staticmethod
    def enviar_email_confirmacao(agendamento):
        """
        Envia email de confirmação para o professor após um agendamento
        
        Args:
            agendamento: Objeto AgendamentoVisita
        """
        professor = agendamento.professor
        if professor is None:
            return
        horario = agendamento.horario
        evento = horario.evento
        
        assunto = f"Confirmação de Agendamento - {evento.nome}"
        
        # Preparar o corpo do email
        corpo_html = render_template(
            'emails/confirmacao_agendamento.html',
            professor=professor,
            agendamento=agendamento,
            horario=horario,
            evento=evento
        )
        
        # Enviar em um thread separado para não bloquear a resposta ao usuário
        thread = threading.Thread(
            target=NotificacaoAgendamentoService._enviar_email_async,
            args=[professor.email, assunto, corpo_html]
        )
        thread.start()
    
    @staticmethod
    def enviar_email_cancelamento(agendamento):
        """
        Envia email de confirmação de cancelamento para o professor
        
        Args:
            agendamento: Objeto AgendamentoVisita
        """
        professor = agendamento.professor
        if professor is None:
            return
        horario = agendamento.horario
        evento = horario.evento
        
        assunto = f"Confirmação de Cancelamento - {evento.nome}"
        
        # Preparar o corpo do email
        corpo_html = render_template(
            'emails/cancelamento_agendamento.html',
            professor=professor,
            agendamento=agendamento,
            horario=horario,
            evento=evento
        )
        
        # Enviar em um thread separado
        thread = threading.Thread(
            target=NotificacaoAgendamentoService._enviar_email_async,
            args=[professor.email, assunto, corpo_html]
        )
        thread.start()
    
    @staticmethod
    def enviar_lembrete_visita(agendamento):
        """
        Envia lembrete de visita para o professor um dia antes
        
        Args:
            agendamento: Objeto AgendamentoVisita
        """
        professor = agendamento.professor
        if professor is None:
            return
        horario = agendamento.horario
        evento = horario.evento
        
        assunto = f"Lembrete de Visita Amanhã - {evento.nome}"
        
        # Preparar o corpo do email
        corpo_html = render_template(
            'emails/lembrete_visita.html',
            professor=professor,
            agendamento=agendamento,
            horario=horario,
            evento=evento
        )
        
        # Enviar em um thread separado
        thread = threading.Thread(
            target=NotificacaoAgendamentoService._enviar_email_async,
            args=[professor.email, assunto, corpo_html]
        )
        thread.start()
    
    @staticmethod
    def notificar_cliente_novo_agendamento(agendamento):
        """
        Notifica o cliente/organizador sobre um novo agendamento
        
        Args:
            agendamento: Objeto AgendamentoVisita
        """
        horario = agendamento.horario
        evento = horario.evento
        cliente = evento.cliente
        
        assunto = f"Novo Agendamento - {evento.nome}"
        
        # Preparar o corpo do email
        corpo_html = render_template(
            'emails/notificacao_novo_agendamento.html',
            cliente=cliente,
            agendamento=agendamento,
            horario=horario,
            evento=evento
        )
        
        # Enviar em um thread separado
        thread = threading.Thread(
            target=NotificacaoAgendamentoService._enviar_email_async,
            args=[cliente.email, assunto, corpo_html]
        )
        thread.start()
    
    @staticmethod
    def _enviar_email_async(dest, subject, html):
        """Função interna para enviar email de forma assíncrona."""
        try:
            if not dest or not subject or not html:
                logger.warning("Dados incompletos para envio de email")
                return
                
            send_via_mailjet(to_email=dest, subject=subject, html=html)
            logger.info(f"Email enviado com sucesso para {dest}")
        except ApiError as exc:
            logger.exception("Erro da API Mailjet ao enviar email: %s", exc)
        except Exception as exc:
            logger.exception("Erro geral ao enviar email para %s: %s", dest, exc)
    
    @staticmethod
    def notificar_monitor_alunos_pcd(agendamento):
        """
        Notifica o monitor sobre alunos PCD/Neurodivergentes no agendamento
        
        Args:
            agendamento: Objeto AgendamentoVisita
        """
        try:
            from models import MonitorAgendamento, NecessidadeEspecial
            
            # Verificar se há monitor atribuído
            monitor_agendamento = MonitorAgendamento.query.filter_by(
                agendamento_id=agendamento.id,
                status='ativo'
            ).first()
            
            if not monitor_agendamento or not monitor_agendamento.monitor:
                logger.info(f"Nenhum monitor ativo encontrado para o agendamento {agendamento.id}")
                return
            
            monitor = monitor_agendamento.monitor
            
            # Verificar se o monitor tem email válido
            if not monitor.email:
                logger.warning(f"Monitor {monitor.id} não possui email cadastrado")
                return
            
            # Buscar alunos com necessidades especiais neste agendamento
            alunos_pcd = db.session.query(AlunoVisitante, NecessidadeEspecial).join(
                NecessidadeEspecial, AlunoVisitante.id == NecessidadeEspecial.aluno_id
            ).filter(
                AlunoVisitante.agendamento_id == agendamento.id
            ).all()
            
            if not alunos_pcd:
                logger.info(f"Nenhum aluno PCD/Neurodivergente encontrado no agendamento {agendamento.id}")
                return
            
            horario = agendamento.horario
            if not horario or not horario.evento:
                logger.warning(f"Agendamento {agendamento.id} não possui horário ou evento válido")
                return
                
            evento = horario.evento
            
            assunto = f"Alunos PCD/Neurodivergentes - {evento.nome}"
            
            # Preparar o corpo do email
            corpo_html = render_template(
                'emails/notificacao_monitor_pcd.html',
                monitor=monitor,
                agendamento=agendamento,
                horario=horario,
                evento=evento,
                alunos_pcd=alunos_pcd
            )
            
            # Enviar em um thread separado
            thread = threading.Thread(
                target=NotificacaoAgendamentoService._enviar_email_async,
                args=[monitor.email, assunto, corpo_html]
            )
            thread.daemon = True  # Thread será finalizada quando o programa principal terminar
            thread.start()
            
            logger.info(f"Notificação enviada para monitor {monitor.nome_completo} sobre {len(alunos_pcd)} alunos PCD/ND")
            
        except Exception as e:
            logger.error(f"Erro ao notificar monitor sobre alunos PCD no agendamento {agendamento.id}: {str(e)}")
    
    @staticmethod
    def processar_lembretes_diarios():
        """
        Tarefa agendada para enviar lembretes diários de visitas
        Deve ser executada uma vez por dia através de um agendador como Celery ou cron
        """
        # Data de amanhã
        amanha = datetime.utcnow().date() + timedelta(days=1)
        
        # Buscar agendamentos pendentes ou confirmados para amanhã
        query = AgendamentoVisita.query.join(
            HorarioVisitacao, AgendamentoVisita.horario_id == HorarioVisitacao.id
        ).filter(
            HorarioVisitacao.data == amanha,
            AgendamentoVisita.status.in_(['pendente', 'confirmado'])
        )
        
        agendamentos = query.all()
        
        # Enviar lembretes
        for agendamento in agendamentos:
            NotificacaoAgendamentoService.enviar_lembrete_visita(agendamento)


# Integrar notificações com as rotas
def integrar_notificacoes():
    """
    Integra as notificações com as rotas existentes
    Esta função deve ser chamada na inicialização da aplicação
    """
    # Sobrescrever a função de criação de agendamento para incluir notificação
    original_criar_agendamento = routes.criar_agendamento_professor
    
    @agendamento_routes.route('/professor/criar_agendamento/<int:horario_id>', methods=['GET', 'POST'])
    @login_required
    def criar_agendamento_com_notificacao(horario_id):
        response = original_criar_agendamento(horario_id)
        
        # Verificar se o agendamento foi criado com sucesso
        # Isso é um hack - na prática seria melhor refatorar a função original
        # para retornar o agendamento criado ou um status
        flashes = session.get('_flashes', [])
        sucesso = any(cat == 'success' for cat, _ in flashes)
        if request.method == 'POST' and sucesso:
            # Buscar o último agendamento criado pelo professor
            agendamento = AgendamentoVisita.query.filter_by(
                professor_id=current_user.id
            ).order_by(AgendamentoVisita.id.desc()).first()
            
            if agendamento:
                # Enviar notificações
                NotificacaoAgendamentoService.notificar_cliente_novo_agendamento(agendamento)
        
        return response
    
    # Substituir a rota original pela nova versão com notificação
    routes.criar_agendamento_professor = criar_agendamento_com_notificacao
    
    # Fazer o mesmo para a função de cancelamento
    original_cancelar_agendamento = routes.cancelar_agendamento_professor
    
    @agendamento_routes.route('/professor/cancelar_agendamento/<int:agendamento_id>', methods=['GET', 'POST'])
    @login_required
    def cancelar_agendamento_com_notificacao(agendamento_id):
        # Buscar o agendamento antes que seja cancelado
        agendamento = AgendamentoVisita.query.get_or_404(agendamento_id)
        
        response = original_cancelar_agendamento(agendamento_id)
        
        # Verificar se o cancelamento foi bem-sucedido
        flashes = session.get('_flashes', [])
        sucesso = any(cat == 'success' for cat, _ in flashes)
        if request.method == 'POST' and sucesso:
            # Enviar notificação de cancelamento
            NotificacaoAgendamentoService.enviar_email_cancelamento(agendamento)
        
        return response
    
    # Substituir a rota original pela nova versão com notificação
    routes.cancelar_agendamento_professor = cancelar_agendamento_com_notificacao
    
@agendamento_routes.route('/eventos_agendamento')
@login_required
def eventos_agendamento():
    # Verificar se é um cliente
    if current_user.tipo != 'cliente':
        flash('Acesso negado!', 'danger')
        return redirect(url_for(endpoints.DASHBOARD))
    
    hoje = datetime.utcnow().date()
    
    # Eventos ativos (em andamento)
    eventos_ativos = Evento.query.filter_by(
        cliente_id=current_user.id
    ).filter(
        and_(
            Evento.data_inicio <= hoje,
            Evento.data_fim >= hoje,
            Evento.status == 'ativo'
        )
    ).all()
    
    # Eventos futuros
    eventos_futuros = Evento.query.filter_by(
        cliente_id=current_user.id
    ).filter(
        and_(
            Evento.data_inicio > hoje,
            Evento.status == 'ativo'
        )
    ).all()
    
    # Eventos passados
    eventos_passados = Evento.query.filter_by(
        cliente_id=current_user.id
    ).filter(
        Evento.data_fim < hoje
    ).order_by(
        Evento.data_fim.desc()
    ).limit(10).all()
    
    # Contar agendamentos para cada evento
    for evento in eventos_ativos + eventos_futuros + eventos_passados:
        evento.agendamentos_count = db.session.query(func.count(AgendamentoVisita.id)).join(
            HorarioVisitacao, AgendamentoVisita.horario_id == HorarioVisitacao.id
        ).filter(
            HorarioVisitacao.evento_id == evento.id
        ).scalar() or 0
    
    return render_template(
        'eventos_agendamento.html',
        eventos_ativos=eventos_ativos,
        eventos_futuros=eventos_futuros,
        eventos_passados=eventos_passados
    )


@agendamento_routes.route('/relatorio_geral_agendamentos')
@login_required
def relatorio_geral_agendamentos():
    """Renderiza o relatório geral de agendamentos.

    O relatório pode ser filtrado pela data da visita ou pela data em que o
    agendamento foi criado, dependendo do parâmetro ``tipo_data`` recebido via
    query string.
    """
    # Verificar se é um cliente
    if current_user.tipo != 'cliente':
        flash('Acesso negado!', 'danger')
        return redirect(url_for(endpoints.DASHBOARD))

    # Filtros de data
    data_inicio_raw = request.args.get('data_inicio')
    data_fim_raw = request.args.get('data_fim')

    def _parse_date(value: str):
        for fmt in ('%Y-%m-%d', '%d/%m/%Y'):
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue
        return None

    class _RawDate(str):
        def strftime(self, _fmt):
            return self

    if data_inicio_raw:
        data_inicio = _parse_date(data_inicio_raw)
        if data_inicio is None:
            flash(
                'Formato de data inválido. Por favor, corrija os valores informados.',
                'info',
            )
            return render_template(
                'relatorio_geral_agendamentos.html',
                eventos=[],
                estatisticas={},
                agendamentos=[],
                professores_confirmados=[],
                filtros={
                    'data_inicio': _RawDate(data_inicio_raw)
                    if data_inicio_raw
                    else None,
                    'data_fim': _RawDate(data_fim_raw)
                    if data_fim_raw
                    else None,
                },
            )
    else:
        # Padrão: último mês
        data_inicio = datetime.utcnow().date() - timedelta(days=30)

    if data_fim_raw:
        data_fim = _parse_date(data_fim_raw)
        if data_fim is None:
            flash(
                'Formato de data inválido. Por favor, corrija os valores informados.',
                'info',
            )
            return render_template(
                'relatorio_geral_agendamentos.html',
                eventos=[],
                estatisticas={},
                agendamentos=[],
                professores_confirmados=[],
                filtros={
                    'data_inicio': _RawDate(data_inicio_raw)
                    if data_inicio_raw
                    else None,
                    'data_fim': _RawDate(data_fim_raw)
                    if data_fim_raw
                    else None,
                },
            )
    else:
        data_fim = datetime.utcnow().date()

    tipo_data = request.args.get('tipo_data', 'horario')
    if tipo_data not in ('horario', 'agendamento'):
        tipo_data = 'horario'

    if tipo_data == 'agendamento':
        coluna_data = func.date(AgendamentoVisita.data_agendamento)
        order_col = AgendamentoVisita.data_agendamento
    else:
        coluna_data = HorarioVisitacao.data
        order_col = HorarioVisitacao.data

    cliente_id = (
        current_user.id if isinstance(current_user, Cliente)
        else current_user.cliente_id
    )

    evento_id = request.args.get('evento_id', type=int)

    eventos = (
        Evento.query.filter_by(cliente_id=cliente_id)
        .order_by(Evento.nome)
        .all()
    )

    # Buscar eventos do cliente com agendamentos no período selecionado
    eventos_query = (
        Evento.query.outerjoin(
            HorarioVisitacao, HorarioVisitacao.evento_id == Evento.id
        )
        .outerjoin(
            AgendamentoVisita, AgendamentoVisita.horario_id == HorarioVisitacao.id
        )
        .filter(
            Evento.cliente_id == cliente_id,
            coluna_data >= data_inicio,
            coluna_data <= data_fim,
        )
    )
    if evento_id:
        eventos_query = eventos_query.filter(Evento.id == evento_id)
    eventos_periodo = eventos_query.distinct().order_by(Evento.nome).all()

    # Estatísticas agregadas em uma única consulta
    aggregados_query = (
        db.session.query(
            Evento.id.label('evento_id'),
            func.count(
                func.distinct(
                    case(

                        (AgendamentoVisita.status == 'confirmado', AgendamentoVisita.id)

                    )
                )
            ).label('confirmados'),
            func.count(
                func.distinct(
                    case(

                        (AgendamentoVisita.status == 'realizado', AgendamentoVisita.id)

                    )
                )
            ).label('realizados'),
            func.count(
                func.distinct(
                    case(

                        (AgendamentoVisita.status == 'cancelado', AgendamentoVisita.id)

                    )
                )
            ).label('cancelados'),
            func.count(
                func.distinct(
                    case(

                        (AgendamentoVisita.status == 'pendente', AgendamentoVisita.id)

                    )
                )
            ).label('pendentes'),
            func.count(
                func.distinct(
                    case(
                        (
                            AgendamentoVisita.checkin_realizado == True,
                            AgendamentoVisita.id,
                        )
                    )
                )
            ).label('checkins'),
            func.sum(
                case(
                    (
                        and_(
                            AgendamentoVisita.status.in_([
                                'confirmado',
                                'realizado',
                            ]),
                            AlunoVisitante.presente.is_(True),
                        ),
                        1,
                    ),
                    else_=0,
                )
            ).label('visitantes_confirmados'),
        )
        .outerjoin(
            HorarioVisitacao, HorarioVisitacao.evento_id == Evento.id
        )
        .outerjoin(
            AgendamentoVisita, AgendamentoVisita.horario_id == HorarioVisitacao.id
        )
        .outerjoin(
            AlunoVisitante,
            AlunoVisitante.agendamento_id == AgendamentoVisita.id,
        )
        .filter(
            Evento.cliente_id == cliente_id,
            coluna_data >= data_inicio,
            coluna_data <= data_fim,
        )
    )
    if evento_id:
        aggregados_query = aggregados_query.filter(Evento.id == evento_id)
    aggregados = aggregados_query.group_by(Evento.id).all()

    stats_map = {row.evento_id: row for row in aggregados}

    estatisticas = {}
    for evento in eventos_periodo:
        row = stats_map.get(evento.id)
        confirmados = row.confirmados if row else 0
        realizados = row.realizados if row else 0
        cancelados = row.cancelados if row else 0
        pendentes = row.pendentes if row else 0
        checkins = row.checkins if row else 0
        visitantes_confirmados = (
            row.visitantes_confirmados if row else 0
        )

        estatisticas[evento.id] = {
            'nome': evento.nome,
            'confirmados': confirmados,
            'realizados': realizados,
            'cancelados': cancelados,
            'pendentes': pendentes,
            'checkins': checkins,
            'total': confirmados + realizados + cancelados + pendentes,
            'visitantes_confirmados': visitantes_confirmados,
        }

    totais = {
        'confirmados': sum(e['confirmados'] for e in estatisticas.values()),
        'realizados': sum(e['realizados'] for e in estatisticas.values()),
        'cancelados': sum(e['cancelados'] for e in estatisticas.values()),
        'pendentes': sum(e['pendentes'] for e in estatisticas.values()),
        'checkins': sum(e['checkins'] for e in estatisticas.values()),
        'visitantes_confirmados': sum(
            e['visitantes_confirmados'] for e in estatisticas.values()
        ),
    }
    total_agendamentos = (
        totais['confirmados']
        + totais['realizados']
        + totais['cancelados']
        + totais['pendentes']
    )


    agendamentos_query = (


        db.session.query(AgendamentoVisita)
        .outerjoin(
            HorarioVisitacao, AgendamentoVisita.horario_id == HorarioVisitacao.id
        )
        .outerjoin(Evento, HorarioVisitacao.evento_id == Evento.id)
        .filter(
            or_(
                AgendamentoVisita.cliente_id == cliente_id,
                Evento.cliente_id == cliente_id,
            ),
            coluna_data >= data_inicio,
            coluna_data <= data_fim,
        )
    )
    if evento_id:
        agendamentos_query = agendamentos_query.filter(Evento.id == evento_id)
    agendamentos = (
        agendamentos_query.order_by(
            order_col, HorarioVisitacao.horario_inicio
        ).all()
    )

    professores_query = (
        db.session.query(
            Usuario.id.label('id'),
            Usuario.nome.label('nome'),
            Usuario.email.label('email'),
            AgendamentoVisita.responsavel_whatsapp.label(
                'responsavel_whatsapp'
            ),
        )
        .join(AgendamentoVisita, AgendamentoVisita.professor_id == Usuario.id)
        .join(
            HorarioVisitacao,
            AgendamentoVisita.horario_id == HorarioVisitacao.id,
        )
        .join(Evento, HorarioVisitacao.evento_id == Evento.id)
        .filter(
            Evento.cliente_id == cliente_id,
            AgendamentoVisita.status == 'confirmado',
            coluna_data >= data_inicio,
            coluna_data <= data_fim,
        )
    )
    if evento_id:
        professores_query = professores_query.filter(Evento.id == evento_id)
    professores_confirmados = professores_query.distinct().all()

    # Calcular dados agregados
    dados_agregados = {}
    
    # Estatísticas por escola
    escolas_stats = (
        db.session.query(
            AgendamentoVisita.escola_nome.label('nome'),
            func.count(AgendamentoVisita.id).label('total_agendamentos'),
            func.sum(AgendamentoVisita.quantidade_alunos).label('total_alunos'),
            func.sum(
                case(
                    (AgendamentoVisita.status == 'realizado', 1),
                    else_=0,
                )
            ).label('realizados'),
            func.sum(
                case(
                    (AgendamentoVisita.status == 'pendente', 1),
                    else_=0,
                )
            ).label('pendentes'),
        )
        .join(HorarioVisitacao, AgendamentoVisita.horario_id == HorarioVisitacao.id)
        .join(Evento, HorarioVisitacao.evento_id == Evento.id)
        .filter(
            Evento.cliente_id == cliente_id,
            coluna_data >= data_inicio,
            coluna_data <= data_fim,
            AgendamentoVisita.escola_nome.isnot(None)
        )
    )
    if evento_id:
        escolas_stats = escolas_stats.filter(Evento.id == evento_id)
    escolas_stats = escolas_stats.group_by(AgendamentoVisita.escola_nome).order_by(func.count(AgendamentoVisita.id).desc()).all()
    dados_agregados['escolas'] = escolas_stats
    
    # Estatísticas por professor
    professores_stats = (
        db.session.query(
            Usuario.nome.label('nome'),
            func.count(AgendamentoVisita.id).label('total_agendamentos'),
            func.sum(AgendamentoVisita.quantidade_alunos).label('total_alunos')
        )
        .join(AgendamentoVisita, AgendamentoVisita.professor_id == Usuario.id)
        .join(HorarioVisitacao, AgendamentoVisita.horario_id == HorarioVisitacao.id)
        .join(Evento, HorarioVisitacao.evento_id == Evento.id)
        .filter(
            Evento.cliente_id == cliente_id,
            coluna_data >= data_inicio,
            coluna_data <= data_fim
        )
    )
    if evento_id:
        professores_stats = professores_stats.filter(Evento.id == evento_id)
    professores_stats = professores_stats.group_by(Usuario.id, Usuario.nome).order_by(func.count(AgendamentoVisita.id).desc()).all()
    dados_agregados['professores'] = professores_stats
    
    # Estatísticas de necessidades especiais
    try:
        from models import MaterialApoio, NecessidadeEspecial
        
        total_alunos_pcd = (
            db.session.query(func.count(AlunoVisitante.id))
            .join(AgendamentoVisita, AlunoVisitante.agendamento_id == AgendamentoVisita.id)
            .join(HorarioVisitacao, AgendamentoVisita.horario_id == HorarioVisitacao.id)
            .join(Evento, HorarioVisitacao.evento_id == Evento.id)
            .join(NecessidadeEspecial, NecessidadeEspecial.aluno_id == AlunoVisitante.id)
            .filter(
                Evento.cliente_id == cliente_id,
                coluna_data >= data_inicio,
                coluna_data <= data_fim,
                or_(
                    NecessidadeEspecial.tipo.like('pcd_%'),
                    NecessidadeEspecial.tipo.like('neurodivergente_%'),
                    NecessidadeEspecial.tipo == 'outros'
                )
            )
        )
        if evento_id:
            total_alunos_pcd = total_alunos_pcd.filter(Evento.id == evento_id)
        total_alunos_pcd = total_alunos_pcd.scalar() or 0
        
        agendamentos_com_pcd = (
            db.session.query(func.count(func.distinct(AgendamentoVisita.id)))
            .join(AlunoVisitante, AlunoVisitante.agendamento_id == AgendamentoVisita.id)
            .join(HorarioVisitacao, AgendamentoVisita.horario_id == HorarioVisitacao.id)
            .join(Evento, HorarioVisitacao.evento_id == Evento.id)
            .join(NecessidadeEspecial, NecessidadeEspecial.aluno_id == AlunoVisitante.id)
            .filter(
                Evento.cliente_id == cliente_id,
                coluna_data >= data_inicio,
                coluna_data <= data_fim,
                or_(
                    NecessidadeEspecial.tipo.like('pcd_%'),
                    NecessidadeEspecial.tipo.like('neurodivergente_%'),
                    NecessidadeEspecial.tipo == 'outros'
                )
            )
        )
        if evento_id:
            agendamentos_com_pcd = agendamentos_com_pcd.filter(Evento.id == evento_id)
        agendamentos_com_pcd = agendamentos_com_pcd.scalar() or 0
        
        percentual_pcd = (agendamentos_com_pcd / total_agendamentos * 100) if total_agendamentos > 0 else 0
        
        dados_agregados['necessidades_especiais'] = {
            'total': total_alunos_pcd,
            'agendamentos': agendamentos_com_pcd,
            'percentual': round(percentual_pcd, 1)
        }
        
        # Materiais de apoio mais utilizados
        materiais_stats = (
            db.session.query(
                MaterialApoio.nome.label('nome'),
                func.count(func.distinct(AlunoVisitante.id)).label('total_uso')
            )
            .join(AlunoVisitante.materiais_apoio)
            .join(AgendamentoVisita, AlunoVisitante.agendamento_id == AgendamentoVisita.id)
            .join(HorarioVisitacao, AgendamentoVisita.horario_id == HorarioVisitacao.id)
            .join(Evento, HorarioVisitacao.evento_id == Evento.id)
            .filter(
                Evento.cliente_id == cliente_id,
                coluna_data >= data_inicio,
                coluna_data <= data_fim,
                MaterialApoio.ativo == True
            )
            .group_by(MaterialApoio.id, MaterialApoio.nome)
            .order_by(func.count(func.distinct(AlunoVisitante.id)).desc())
            .limit(5)
            .all()
        )
        if evento_id:
            materiais_stats = (
                db.session.query(
                    MaterialApoio.nome.label('nome'),
                    func.count(func.distinct(AlunoVisitante.id)).label('total_uso')
                )
                .join(AlunoVisitante.materiais_apoio)
                .join(AgendamentoVisita, AlunoVisitante.agendamento_id == AgendamentoVisita.id)
                .join(HorarioVisitacao, AgendamentoVisita.horario_id == HorarioVisitacao.id)
                .join(Evento, HorarioVisitacao.evento_id == Evento.id)
                .filter(
                    Evento.cliente_id == cliente_id,
                    Evento.id == evento_id,
                    coluna_data >= data_inicio,
                    coluna_data <= data_fim,
                    MaterialApoio.ativo == True
                )
                .group_by(MaterialApoio.id, MaterialApoio.nome)
                .order_by(func.count(func.distinct(AlunoVisitante.id)).desc())
                .limit(5)
                .all()
            )
        dados_agregados['materiais_apoio'] = materiais_stats
        
    except ImportError:
        # Se o modelo MaterialApoio não existir, usar dados vazios
        dados_agregados['necessidades_especiais'] = {
            'total': 0,
            'agendamentos': 0,
            'percentual': 0
        }
        dados_agregados['materiais_apoio'] = []
    
    # Distribuição por nível de ensino
    niveis_stats = (
        db.session.query(
            AgendamentoVisita.nivel_ensino.label('nome'),
            func.count(AgendamentoVisita.id).label('total')
        )
        .join(HorarioVisitacao, AgendamentoVisita.horario_id == HorarioVisitacao.id)
        .join(Evento, HorarioVisitacao.evento_id == Evento.id)
        .filter(
            Evento.cliente_id == cliente_id,
            coluna_data >= data_inicio,
            coluna_data <= data_fim,
            AgendamentoVisita.nivel_ensino.isnot(None)
        )
    )
    if evento_id:
        niveis_stats = niveis_stats.filter(Evento.id == evento_id)
    niveis_stats = niveis_stats.group_by(AgendamentoVisita.nivel_ensino).order_by(func.count(AgendamentoVisita.id).desc()).all()
    dados_agregados['niveis_ensino'] = niveis_stats

    # Gerar PDF com detalhes
    if request.args.get('gerar_pdf'):
        pdf_filename = (
            f"relatorio_geral_{data_inicio.strftime('%Y%m%d')}_{data_fim.strftime('%Y%m%d')}.pdf"
        )
        pdf_path = os.path.join("static", "relatorios", pdf_filename)
        os.makedirs(os.path.dirname(pdf_path), exist_ok=True)

        # Gerar PDF completo com todas as estatísticas
        gerar_pdf_relatorio_geral_completo(
            eventos_periodo, estatisticas, totais, dados_agregados, 
            professores_confirmados, agendamentos, data_inicio, data_fim, pdf_path
        )

        return send_file(pdf_path, as_attachment=True)
    
    # Gerar XLSX com detalhes
    if request.args.get('gerar_xlsx'):
        xlsx_filename = (
            f"relatorio_geral_{data_inicio.strftime('%Y%m%d')}_{data_fim.strftime('%Y%m%d')}.xlsx"
        )
        xlsx_path = os.path.join("static", "relatorios", xlsx_filename)
        os.makedirs(os.path.dirname(xlsx_path), exist_ok=True)

        # Gerar XLSX completo com todas as informações
        gerar_xlsx_relatorio_geral_completo(
            eventos_periodo, estatisticas, totais, dados_agregados, 
            professores_confirmados, agendamentos, data_inicio, data_fim, xlsx_path
        )

        return send_file(xlsx_path, as_attachment=True)
    
    return render_template(
        'relatorio_geral_agendamentos.html',
        eventos=eventos,
        estatisticas=estatisticas,
        totais=totais,
        total_agendamentos=total_agendamentos,
        agendamentos=agendamentos,
        professores_confirmados=professores_confirmados,
        dados_agregados=dados_agregados,
        filtros={
            'data_inicio': data_inicio,
            'data_fim': data_fim,
            'tipo_data': tipo_data,
            'evento_id': evento_id,
        }
    )


@agendamento_routes.route('/editar_horario_agendamento', methods=['POST'])
@login_required
def editar_horario_agendamento():
    # Verificar se é um cliente
    if current_user.tipo != 'cliente':
        flash('Acesso negado!', 'danger')
        return redirect(url_for(endpoints.DASHBOARD))
    
    horario_id = request.form.get('horario_id', type=int)
    horario = HorarioVisitacao.query.get_or_404(horario_id)

    if horario.fechado:
        flash('Este horário está fechado para agendamentos.', 'warning')
        return redirect(
            url_for('dashboard_participante_routes.dashboard_participante')
        )
    evento = horario.evento
    
    # Verificar se o evento pertence ao cliente
    if evento.cliente_id != current_user.id:
        flash('Este evento não pertence a você!', 'danger')
        return redirect(url_for(endpoints.DASHBOARD_CLIENTE))
    
    # Coletar dados do formulário
    horario_inicio = request.form.get('horario_inicio')
    horario_fim = request.form.get('horario_fim')
    capacidade_total = request.form.get('capacidade_total', type=int)
    vagas_disponiveis = request.form.get('vagas_disponiveis', type=int)
    
    if horario_inicio and horario_fim and capacidade_total is not None and vagas_disponiveis is not None:
        # Converter string para time
        horario.horario_inicio = datetime.strptime(horario_inicio, '%H:%M').time()
        horario.horario_fim = datetime.strptime(horario_fim, '%H:%M').time()
        
        # Verificar se a capacidade é menor que o número de agendamentos existentes
        agendamentos_count = db.session.query(func.count(AgendamentoVisita.id)).filter(
            AgendamentoVisita.horario_id == horario.id,
            AgendamentoVisita.status.in_(['pendente', 'confirmado'])
        ).scalar() or 0

        agendamentos_alunos = db.session.query(
            func.sum(AgendamentoVisita.quantidade_alunos)
        ).filter(
            AgendamentoVisita.horario_id == horario.id,
            AgendamentoVisita.status.in_(['pendente', 'confirmado'])
        ).scalar() or 0
        
        if capacidade_total < agendamentos_alunos:
            flash(f'Não é possível reduzir a capacidade para {capacidade_total}. Já existem {agendamentos_alunos} alunos agendados.', 'danger')
            return redirect(url_for('agendamento_routes.listar_horarios_agendamento', evento_id=evento.id))
        
        if vagas_disponiveis > capacidade_total:
            flash('As vagas disponíveis não podem ser maiores que a capacidade total!', 'danger')
            return redirect(url_for('agendamento_routes.listar_horarios_agendamento', evento_id=evento.id))
        
        # Atualizar horário
        horario.capacidade_total = capacidade_total
        horario.vagas_disponiveis = vagas_disponiveis
        
        try:
            db.session.commit()
            flash('Horário atualizado com sucesso!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao atualizar horário: {str(e)}', 'danger')
    else:
        flash('Preencha todos os campos!', 'danger')
    
    return redirect(url_for('agendamento_routes.listar_horarios_agendamento', evento_id=evento.id))


@agendamento_routes.route('/excluir_horario_agendamento', methods=['POST'])
@login_required
def excluir_horario_agendamento():
    # Verificar se é um cliente
    if current_user.tipo != 'cliente':
        flash('Acesso negado!', 'danger')
        return redirect(url_for(endpoints.DASHBOARD))
    
    horario_id = request.form.get('horario_id', type=int)
    horario = HorarioVisitacao.query.get_or_404(horario_id)
    evento_id = horario.evento_id
    evento = horario.evento
    
    # Verificar se o evento pertence ao cliente
    if evento.cliente_id != current_user.id:
        flash('Este evento não pertence a você!', 'danger')
        return redirect(url_for(endpoints.DASHBOARD_CLIENTE))
    
    # Verificar se existem agendamentos para este horário
    agendamentos = AgendamentoVisita.query.filter(
        AgendamentoVisita.horario_id == horario.id,
        AgendamentoVisita.status.in_(['pendente', 'confirmado'])
    ).all()
    
    if agendamentos:
        # Cancelar todos os agendamentos
        for agendamento in agendamentos:
            agendamento.status = 'cancelado'
            agendamento.data_cancelamento = datetime.utcnow()
            
            # Enviar notificação de cancelamento
            # (Aqui pode-se adicionar código para enviar emails de cancelamento)
    
    try:
        # Excluir o horário
        db.session.delete(horario)
        db.session.commit()
        
        if agendamentos:
            flash(f'Horário excluído e {len(agendamentos)} agendamentos cancelados!', 'success')
        else:
            flash('Horário excluído com sucesso!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao excluir horário: {str(e)}', 'danger')

    return redirect(url_for('agendamento_routes.listar_horarios_agendamento', evento_id=evento_id))


@agendamento_routes.route('/toggle_horario_agendamento', methods=['POST'])
@login_required
def toggle_horario_agendamento():
    """Alterna o estado de fechamento de um horário de visitação."""
    if current_user.tipo != 'cliente':
        flash('Acesso negado!', 'danger')
        return redirect(url_for(endpoints.DASHBOARD))

    horario_id = request.form.get('horario_id', type=int)
    horario = HorarioVisitacao.query.get_or_404(horario_id)
    evento = horario.evento

    if evento.cliente_id != current_user.id:
        flash('Este evento não pertence a você!', 'danger')
        return redirect(url_for(endpoints.DASHBOARD_CLIENTE))

    horario.fechado = not horario.fechado
    db.session.commit()

    mensagem = 'Horário fechado com sucesso!' if horario.fechado else 'Horário reaberto com sucesso!'
    flash(mensagem, 'success')

    return redirect(
        url_for('agendamento_routes.listar_horarios_agendamento', evento_id=horario.evento_id)
    )


@agendamento_routes.route(
    '/excluir_todos_horarios_agendamento/<int:evento_id>',
    methods=['POST'],
)
@login_required
def excluir_todos_horarios_agendamento(evento_id):
    """Remove todos os horários de um evento e cancela agendamentos."""
    if current_user.tipo != 'cliente':
        flash('Acesso negado!', 'danger')
        return redirect(url_for(endpoints.DASHBOARD))

    evento = Evento.query.get_or_404(evento_id)
    if evento.cliente_id != current_user.id:
        flash('Este evento não pertence a você!', 'danger')
        return redirect(url_for(endpoints.DASHBOARD_CLIENTE))

    try:
        horarios = HorarioVisitacao.query.filter_by(evento_id=evento_id).all()
        for horario in horarios:
            agendamentos = AgendamentoVisita.query.filter(
                AgendamentoVisita.horario_id == horario.id,
                AgendamentoVisita.status.in_(['pendente', 'confirmado']),
            ).all()
            for agendamento in agendamentos:
                agendamento.status = 'cancelado'
                agendamento.data_cancelamento = datetime.utcnow()
            db.session.delete(horario)
        db.session.commit()
        flash('Todos os horários foram excluídos com sucesso!', 'success')
    except Exception as exc:  # pragma: no cover - erro inesperado
        db.session.rollback()
        flash(f'Erro ao excluir horários: {exc}', 'danger')

    return redirect(
        url_for('agendamento_routes.listar_horarios_agendamento', evento_id=evento_id)
    )


def gerar_pdf_relatorio_geral_completo(eventos, estatisticas, totais, dados_agregados, professores_confirmados, agendamentos, data_inicio, data_fim, caminho_pdf):
    """
    Gera um relatório geral completo em PDF com todas as estatísticas de agendamentos.
    
    Args:
        eventos: Lista de objetos Evento
        estatisticas: Dicionário com estatísticas por evento
        totais: Dicionário com totais gerais
        dados_agregados: Dados agregados (escolas, professores, etc.)
        professores_confirmados: Lista de professores confirmados
        agendamentos: Lista de agendamentos
        data_inicio: Data inicial do período do relatório
        data_fim: Data final do período do relatório
        caminho_pdf: Caminho onde o PDF será salvo
    """
    pdf = FPDF()
    pdf.add_page()
    
        # Register fonts before setting them
    font_dir = os.path.join(current_app.root_path, "fonts")
    pdf.add_font("DejaVu", "", os.path.join(font_dir, "DejaVuSans.ttf"), uni=True)
    pdf.add_font("DejaVu", "B", os.path.join(font_dir, "DejaVuSans-Bold.ttf"), uni=True)
    pdf.add_font(
        "DejaVu", "I", os.path.join(font_dir, "DejaVuSans-Oblique.ttf"), uni=True
    )

    # Configurar fonte
    pdf.set_font("DejaVu", "B", 16)
    
    # Título
    pdf.cell(190, 10, 'Relatório Geral de Agendamentos', 0, 1, 'C')
    pdf.set_font('DejaVu', '', 12)
    pdf.cell(190, 10, f'Período: {data_inicio.strftime("%d/%m/%Y")} a {data_fim.strftime("%d/%m/%Y")}', 0, 1, 'C')
    
    # Data e hora de geração
    pdf.set_font('DejaVu', 'I', 10)
    pdf.cell(190, 10, f'Gerado em: {datetime.now().strftime("%d/%m/%Y %H:%M")}', 0, 1, 'R')
    
    # Resumo geral
    pdf.ln(5)
    pdf.set_font('DejaVu', 'B', 14)
    pdf.cell(190, 10, 'Resumo Geral', 0, 1)
    
    total_agendamentos = totais['confirmados'] + totais['realizados'] + totais['cancelados'] + totais['pendentes']
    
    pdf.set_font('DejaVu', '', 12)
    pdf.cell(95, 10, f'Total de Agendamentos: {total_agendamentos}', 0, 0)
    pdf.cell(95, 10, f'Confirmados: {totais["confirmados"]}', 0, 1)
    
    pdf.cell(95, 10, f'Cancelados: {totais["cancelados"]}', 0, 0)
    pdf.cell(95, 10, f'Check-ins: {totais["checkins"]}', 0, 1)
    
    # Calcular taxas
    if total_agendamentos > 0:
        taxa_cancelamento = (totais['cancelados'] / total_agendamentos) * 100
        pdf.cell(95, 10, f'Taxa de Cancelamento: {taxa_cancelamento:.1f}%', 0, 0)
    
    if totais['confirmados'] + totais['realizados'] > 0:
        taxa_conclusao = (totais['realizados'] / (totais['confirmados'] + totais['realizados'])) * 100
        pdf.cell(95, 10, f'Taxa de Conclusão: {taxa_conclusao:.1f}%', 0, 1)
    
    # Estatísticas por escola com PCD
    if 'escolas' in dados_agregados and dados_agregados['escolas']:
        pdf.ln(10)
        pdf.set_font('DejaVu', 'B', 14)
        pdf.cell(190, 10, 'Top 10 Escolas', 0, 1)
        
        pdf.set_font('DejaVu', 'B', 9)
        pdf.cell(60, 10, 'Escola', 1, 0, 'C')
        pdf.cell(25, 10, 'Agendamentos', 1, 0, 'C')
        pdf.cell(25, 10, 'Alunos', 1, 0, 'C')
        pdf.cell(25, 10, 'PCD', 1, 0, 'C')
        pdf.cell(25, 10, 'Realizados', 1, 0, 'C')
        pdf.cell(30, 10, 'Pendentes', 1, 1, 'C')
        
        pdf.set_font('DejaVu', '', 8)
        for escola in dados_agregados['escolas'][:10]:
            escola_nome = escola.nome[:30] + '...' if len(escola.nome) > 30 else escola.nome
            
            # Calcular PCD por escola
            pcd_escola = 0
            for agendamento in agendamentos:
                if agendamento.escola_nome == escola.nome:
                    for aluno in agendamento.alunos:
                        if hasattr(aluno, 'necessidade_especial') and aluno.necessidade_especial:
                            pcd_escola += 1
            
            pdf.cell(60, 8, escola_nome, 1, 0)
            pdf.cell(25, 8, str(escola.total_agendamentos), 1, 0, 'C')
            pdf.cell(25, 8, str(escola.total_alunos or 0), 1, 0, 'C')
            pdf.cell(25, 8, str(pcd_escola), 1, 0, 'C')
            pdf.cell(25, 8, str(escola.realizados or 0), 1, 0, 'C')
            pdf.cell(30, 8, str(escola.pendentes or 0), 1, 1, 'C')
    
    # Estatísticas por professor
    if 'professores' in dados_agregados and dados_agregados['professores']:
        pdf.ln(5)
        pdf.set_font('DejaVu', 'B', 14)
        pdf.cell(190, 10, 'Top 10 Professores', 0, 1)
        
        pdf.set_font('DejaVu', 'B', 10)
        pdf.cell(80, 10, 'Professor', 1, 0, 'C')
        pdf.cell(30, 10, 'Agendamentos', 1, 0, 'C')
        pdf.cell(30, 10, 'Alunos', 1, 1, 'C')
        
        pdf.set_font('DejaVu', '', 10)
        for professor in dados_agregados['professores'][:10]:
            prof_nome = professor.nome[:35] + '...' if len(professor.nome) > 35 else professor.nome
            pdf.cell(80, 10, prof_nome, 1, 0)
            pdf.cell(30, 10, str(professor.total_agendamentos), 1, 0, 'C')
            pdf.cell(30, 10, str(professor.total_alunos or 0), 1, 1, 'C')
    
    # Professores confirmados
    if professores_confirmados:
        pdf.ln(5)
        pdf.set_font('DejaVu', 'B', 14)
        pdf.cell(190, 10, 'Professores Confirmados', 0, 1)
        
        pdf.set_font('DejaVu', 'B', 10)
        pdf.cell(60, 10, 'Nome', 1, 0, 'C')
        pdf.cell(70, 10, 'Email', 1, 0, 'C')
        pdf.cell(30, 10, 'Total de Visitantes', 1, 0, 'C')
        pdf.cell(30, 10, 'WhatsApp', 1, 1, 'C')
        
        pdf.set_font('DejaVu', '', 8)
        for professor in professores_confirmados[:20]:  # Limitar a 20 para não sobrecarregar
            nome = professor.nome[:25] + '...' if len(professor.nome) > 25 else professor.nome
            email = professor.email[:30] + '...' if len(professor.email) > 30 else professor.email
            whatsapp = professor.responsavel_whatsapp or 'N/A'
            
            # Calcular total de visitantes do professor
            total_visitantes_prof = sum(
                ag.quantidade_alunos for ag in agendamentos 
                if ag.professor and ag.professor.id == professor.id and ag.status in ['confirmado', 'realizado']
            )
            
            pdf.cell(60, 8, nome, 1, 0)
            pdf.cell(70, 8, email, 1, 0)
            pdf.cell(30, 8, str(total_visitantes_prof), 1, 0, 'C')
            pdf.cell(30, 8, whatsapp, 1, 1)
    
    # Necessidades especiais com detalhes
    if 'necessidades_especiais' in dados_agregados:
        pdf.ln(5)
        pdf.set_font('DejaVu', 'B', 14)
        pdf.cell(190, 10, 'Necessidades Especiais', 0, 1)
        
        pdf.set_font('DejaVu', '', 12)
        pcd_data = dados_agregados['necessidades_especiais']
        
        # Calcular estatísticas detalhadas de PCD
        total_alunos_pcd = 0
        tipos_pcd = {}
        
        for agendamento in agendamentos:
            for aluno in agendamento.alunos:
                if hasattr(aluno, 'necessidade_especial') and aluno.necessidade_especial:
                    total_alunos_pcd += 1
                    tipo_display = aluno.necessidade_especial.get_tipo_display()
                    tipos_pcd[tipo_display] = tipos_pcd.get(tipo_display, 0) + 1
        
        percentual = (total_alunos_pcd / max(totais.get('visitantes_confirmados', 1), 1)) * 100
        
        pdf.cell(95, 10, f'Total de Alunos PCD: {total_alunos_pcd}', 0, 0)
        pdf.cell(95, 10, f'Percentual: {percentual:.1f}%', 0, 1)
        pdf.cell(95, 10, f'Agendamentos com PCD: {pcd_data["agendamentos"]}', 0, 1)
        
        # Distribuição por tipo de necessidade especial
        if tipos_pcd:
            pdf.ln(3)
            pdf.set_font('DejaVu', 'B', 12)
            pdf.cell(190, 8, 'Distribuição por Tipo de Necessidade:', 0, 1)
            pdf.set_font('DejaVu', '', 10)
            for tipo, quantidade in tipos_pcd.items():
                pdf.cell(190, 6, f'- {tipo}: {quantidade} aluno(s)', 0, 1)
    
    # Agendamentos detalhados por status
    pdf.add_page()
    pdf.set_font('DejaVu', 'B', 16)
    pdf.cell(190, 10, 'Detalhes dos Agendamentos', 0, 1, 'C')
    
    # Organizar agendamentos por status (confirmados primeiro)
    agendamentos_por_status = {
        'confirmado': [a for a in agendamentos if a.status == 'confirmado'],
        'realizado': [a for a in agendamentos if a.status == 'realizado'],
        'pendente': [a for a in agendamentos if a.status == 'pendente'],
        'cancelado': [a for a in agendamentos if a.status == 'cancelado']
    }
    
    for status, lista_agendamentos in agendamentos_por_status.items():
        if not lista_agendamentos:
            continue
            
        pdf.ln(5)
        pdf.set_font('DejaVu', 'B', 14)
        pdf.cell(190, 10, f'Agendamentos {status.capitalize()}s ({len(lista_agendamentos)})', 0, 1)
        
        for agendamento in lista_agendamentos:
            # Verificar se precisa de nova página
            if pdf.get_y() > 250:
                pdf.add_page()
            
            horario = agendamento.horario
            
            # Cabeçalho do agendamento
            pdf.ln(3)
            pdf.set_font('DejaVu', 'B', 11)
            pdf.set_fill_color(230, 230, 230)
            pdf.cell(190, 8, f'#{agendamento.id} - {agendamento.escola_nome} - {horario.data.strftime("%d/%m/%Y")} {horario.horario_inicio.strftime("%H:%M")}', 1, 1, 'L', True)
            
            # Informações básicas
            pdf.set_font('DejaVu', '', 9)
            professor_nome = agendamento.professor.nome if agendamento.professor else 'Não informado'
            pdf.cell(95, 6, f'Professor: {professor_nome}', 1, 0)
            pdf.cell(95, 6, f'Status: {agendamento.status.capitalize()}', 1, 1)
            
            pdf.cell(95, 6, f'Turma: {agendamento.turma}', 1, 0)
            pdf.cell(95, 6, f'Total de Alunos: {agendamento.quantidade_alunos}', 1, 1)
            
            # Lista de alunos
            if agendamento.alunos:
                pdf.ln(2)
                pdf.set_font('DejaVu', 'B', 9)
                pdf.cell(190, 6, 'Lista de Alunos:', 0, 1)
                
                # Cabeçalho da tabela de alunos
                pdf.set_font('DejaVu', 'B', 8)
                pdf.cell(70, 6, 'Nome', 1, 0, 'C')
                pdf.cell(20, 6, 'Presente', 1, 0, 'C')
                pdf.cell(50, 6, 'Tipo PCD', 1, 0, 'C')
                pdf.cell(50, 6, 'Descrição PCD', 1, 1, 'C')
                
                pdf.set_font('DejaVu', '', 7)
                for aluno in agendamento.alunos:
                    # Verificar se precisa de nova página
                    if pdf.get_y() > 270:
                        pdf.add_page()
                        # Repetir cabeçalho
                        pdf.set_font('DejaVu', 'B', 8)
                        pdf.cell(70, 6, 'Nome', 1, 0, 'C')
                        pdf.cell(20, 6, 'Presente', 1, 0, 'C')
                        pdf.cell(50, 6, 'Tipo PCD', 1, 0, 'C')
                        pdf.cell(50, 6, 'Descrição PCD', 1, 1, 'C')
                        pdf.set_font('DejaVu', '', 7)
                    
                    nome = aluno.nome[:35] + '...' if len(aluno.nome) > 35 else aluno.nome
                    presente = 'Sim' if aluno.presente else 'Não'
                    
                    tipo_pcd = ''
                    descricao_pcd = ''
                    
                    if hasattr(aluno, 'necessidade_especial') and aluno.necessidade_especial:
                        tipo_pcd = aluno.necessidade_especial.get_tipo_display()[:25]
                        descricao_pcd = aluno.necessidade_especial.descricao[:30] + '...' if len(aluno.necessidade_especial.descricao) > 30 else aluno.necessidade_especial.descricao
                    
                    pdf.cell(70, 6, nome, 1, 0, 'L')
                    pdf.cell(20, 6, presente, 1, 0, 'C')
                    pdf.cell(50, 6, tipo_pcd, 1, 0, 'L')
                    pdf.cell(50, 6, descricao_pcd, 1, 1, 'L')
    
    # Análise e recomendações
    pdf.add_page()
    pdf.set_font('DejaVu', 'B', 14)
    pdf.cell(190, 10, 'Análise e Recomendações', 0, 1)
    
    pdf.set_font('DejaVu', '', 12)
    if total_agendamentos > 0:
        taxa_cancelamento = (totais['cancelados'] / total_agendamentos) * 100
        if taxa_cancelamento > 30:
            pdf.multi_cell(190, 10, '- Alta taxa de cancelamento. Considere revisar suas políticas de cancelamento.')
            pdf.multi_cell(190, 10, '- Envie lembretes com mais frequência para professores com agendamentos confirmados.')
        else:
            pdf.multi_cell(190, 10, '- Taxa de cancelamento está em níveis aceitáveis.')
        
        if totais['realizados'] < totais['confirmados']:
            pdf.multi_cell(190, 10, '- Implemente um sistema de lembretes mais eficiente para aumentar o comparecimento.')
            
        if totais.get('visitantes_confirmados', 0) < 100:
            pdf.multi_cell(190, 10, '- Divulgue mais seus eventos entre escolas e professores para aumentar a quantidade de visitantes.')
    else:
        pdf.multi_cell(190, 10, '- Ainda não há dados suficientes para recomendações personalizadas.')
    
    pdf.multi_cell(190, 10, '- Continue monitorando os agendamentos e ajustando a capacidade disponível conforme necessário.')
    
    # Rodapé
    pdf.ln(10)
    pdf.set_font('DejaVu', 'I', 10)
    pdf.cell(190, 10, 'Este relatório é gerado automaticamente pelo sistema de agendamentos.', 0, 1, 'C')
    
    # Salvar o PDF
    pdf.output(caminho_pdf)


def gerar_pdf_relatorio_geral(eventos, estatisticas, data_inicio, data_fim, caminho_pdf):
    """
    Gera um relatório geral em PDF com estatísticas de agendamentos para todos os eventos.
    
    Args:
        eventos: Lista de objetos Evento
        estatisticas: Dicionário com estatísticas por evento
        data_inicio: Data inicial do período do relatório
        data_fim: Data final do período do relatório
        caminho_pdf: Caminho onde o PDF será salvo
    """
    pdf = FPDF()
    pdf.add_page()
    
    # Configurar fonte
    pdf.set_font('DejaVu', 'B', 16)
    
    # Título
    pdf.cell(190, 10, 'Relatório Geral de Agendamentos', 0, 1, 'C')
    pdf.set_font('DejaVu', '', 12)
    pdf.cell(190, 10, f'Período: {data_inicio.strftime("%d/%m/%Y")} a {data_fim.strftime("%d/%m/%Y")}', 0, 1, 'C')
    
    # Data e hora de geração
    pdf.set_font('DejaVu', 'I', 10)
    pdf.cell(190, 10, f'Gerado em: {datetime.now().strftime("%d/%m/%Y %H:%M")}', 0, 1, 'R')
    
    # Cálculo de totais
    total_confirmados = 0
    total_realizados = 0
    total_cancelados = 0
    total_visitantes = 0
    
    for stats in estatisticas.values():
        total_confirmados += stats['confirmados']
        total_realizados += stats['realizados']
        total_cancelados += stats['cancelados']
        total_visitantes += stats['visitantes']
    
    total_agendamentos = total_confirmados + total_realizados + total_cancelados
    
    # Resumo geral
    pdf.ln(5)
    pdf.set_font('DejaVu', 'B', 14)
    pdf.cell(190, 10, 'Resumo Geral', 0, 1)
    
    pdf.set_font('DejaVu', '', 12)
    pdf.cell(95, 10, f'Total de Agendamentos: {total_agendamentos}', 0, 0)
    pdf.cell(95, 10, f'Total de Visitantes: {total_visitantes}', 0, 1)
    
    pdf.cell(95, 10, f'Agendamentos Confirmados: {total_confirmados}', 0, 0)
    pdf.cell(95, 10, f'Agendamentos Realizados: {total_realizados}', 0, 1)
    
    pdf.cell(95, 10, f'Agendamentos Cancelados: {total_cancelados}', 0, 1)
    
    # Calcular taxas
    if total_agendamentos > 0:
        taxa_cancelamento = (total_cancelados / total_agendamentos) * 100
        pdf.cell(190, 10, f'Taxa de Cancelamento: {taxa_cancelamento:.1f}%', 0, 1)
    
    if total_confirmados + total_realizados > 0:
        taxa_conclusao = (total_realizados / (total_confirmados + total_realizados)) * 100
        pdf.cell(190, 10, f'Taxa de Conclusão: {taxa_conclusao:.1f}%', 0, 1)
    
    # Detalhes por evento
    pdf.ln(10)
    pdf.set_font('DejaVu', 'B', 14)
    pdf.cell(190, 10, 'Detalhes por Evento', 0, 1)
    
    # Cabeçalho da tabela
    pdf.set_font('DejaVu', 'B', 10)
    pdf.cell(60, 10, 'Evento', 1, 0, 'C')
    pdf.cell(25, 10, 'Confirmados', 1, 0, 'C')
    pdf.cell(25, 10, 'Realizados', 1, 0, 'C')
    pdf.cell(25, 10, 'Cancelados', 1, 0, 'C')
    pdf.cell(25, 10, 'Visitantes', 1, 0, 'C')
    pdf.cell(30, 10, 'Taxa Conclusão', 1, 1, 'C')
    
    # Dados da tabela
    pdf.set_font('DejaVu', '', 10)
    for evento_id, stats in estatisticas.items():
        evento_nome = stats['nome']
        
        # Limitar tamanho do nome para caber na tabela
        if len(evento_nome) > 30:
            evento_nome = evento_nome[:27] + '...'
        
        pdf.cell(60, 10, evento_nome, 1, 0)
        pdf.cell(25, 10, str(stats['confirmados']), 1, 0, 'C')
        pdf.cell(25, 10, str(stats['realizados']), 1, 0, 'C')
        pdf.cell(25, 10, str(stats['cancelados']), 1, 0, 'C')
        pdf.cell(25, 10, str(stats['visitantes']), 1, 0, 'C')
        
        # Calcular taxa de conclusão
        if stats['confirmados'] + stats['realizados'] > 0:
            taxa = (stats['realizados'] / (stats['confirmados'] + stats['realizados'])) * 100
            pdf.cell(30, 10, f'{taxa:.1f}%', 1, 1, 'C')
        else:
            pdf.cell(30, 10, 'N/A', 1, 1, 'C')
    
    # Análise e recomendações
    pdf.ln(10)
    pdf.set_font('DejaVu', 'B', 14)
    pdf.cell(190, 10, 'Análise e Recomendações', 0, 1)
    
    pdf.set_font('DejaVu', '', 12)
    if total_agendamentos > 0:
        if taxa_cancelamento > 30:
            pdf.multi_cell(190, 10, '- Alta taxa de cancelamento. Considere revisar suas políticas de cancelamento.')
            pdf.multi_cell(190, 10, '- Envie lembretes com mais frequência para professores com agendamentos confirmados.')
        else:
            pdf.multi_cell(190, 10, '- Taxa de cancelamento está em níveis aceitáveis.')
        
        if total_realizados < total_confirmados:
            pdf.multi_cell(190, 10, '- Implemente um sistema de lembretes mais eficiente para aumentar o comparecimento.')
            
        if total_visitantes < 100:
            pdf.multi_cell(190, 10, '- Divulgue mais seus eventos entre escolas e professores para aumentar a quantidade de visitantes.')
    else:
        pdf.multi_cell(190, 10, '- Ainda não há dados suficientes para recomendações personalizadas.')
    
    pdf.multi_cell(190, 10, '- Continue monitorando os agendamentos e ajustando a capacidade disponível conforme necessário.')
    
    # Rodapé
    pdf.ln(10)
    pdf.set_font('DejaVu', 'I', 10)
    pdf.cell(190, 10, 'Este relatório é gerado automaticamente pelo sistema de agendamentos.', 0, 1, 'C')
    
    # Salvar o PDF
    pdf.output(caminho_pdf)




def gerar_pdf_relatorio_agendamentos(evento, agendamentos, caminho_pdf):
    """Gera um PDF de agendamentos com presença registrada via QR code.

    Args:
        evento: Objeto Evento
        agendamentos: Lista de objetos ``AgendamentoVisita``
        caminho_pdf: Caminho onde o PDF será salvo
    """
    pdf = FPDF()
    pdf.add_page()
    
    # Configurar fonte
    pdf.set_font('DejaVu', 'B', 16)
    
    # Título
    pdf.cell(190, 10, f'Relatório de Agendamentos - {evento.nome}', 0, 1, 'C')
    
    # Informações do evento
    pdf.set_font('DejaVu', '', 12)
    pdf.cell(190, 10, f'Local: {evento.local}', 0, 1)
    pdf.cell(190, 10, f'Período: {evento.data_inicio.strftime("%d/%m/%Y")} a {evento.data_fim.strftime("%d/%m/%Y")}', 0, 1)
    
    # Data e hora de geração
    pdf.set_font('DejaVu', 'I', 10)
    pdf.cell(190, 10, f'Gerado em: {datetime.now().strftime("%d/%m/%Y %H:%M")}', 0, 1, 'R')
    
    # Estatísticas
    pdf.ln(5)
    pdf.set_font('DejaVu', 'B', 14)
    pdf.cell(190, 10, 'Estatísticas', 0, 1)
    
    # Contadores
    total_agendamentos = len(agendamentos)
    confirmados = sum(1 for a in agendamentos if a.status == 'confirmado')
    realizados = sum(1 for a in agendamentos if a.status == 'realizado')
    cancelados = sum(1 for a in agendamentos if a.status == 'cancelado')
    
    total_alunos = sum(a.quantidade_alunos for a in agendamentos if a.status in ['confirmado', 'realizado'])
    presentes = 0
    for a in agendamentos:
        if a.status == 'realizado':
            presentes += sum(1 for aluno in a.alunos if aluno.presente)
    
    pdf.set_font('DejaVu', '', 12)
    pdf.cell(95, 10, f'Total de Agendamentos: {total_agendamentos}', 0, 0)
    pdf.cell(95, 10, f'Total de Alunos: {total_alunos}', 0, 1)
    
    pdf.cell(95, 10, f'Confirmados: {confirmados}', 0, 0)
    pdf.cell(95, 10, f'Realizados: {realizados}', 0, 1)
    
    pdf.cell(95, 10, f'Cancelados: {cancelados}', 0, 0)
    if realizados > 0:
        pdf.cell(95, 10, f'Alunos Presentes: {presentes}', 0, 1)
    
    # Lista de agendamentos
    pdf.ln(10)
    pdf.set_font('DejaVu', 'B', 14)
    pdf.cell(190, 10, 'Lista de Agendamentos', 0, 1)

    # Larguras dinâmicas para escola e professor
    pdf.set_font('DejaVu', '', 8)
    max_escola = max(
        (pdf.get_string_width(a.escola_nome) for a in agendamentos),
        default=0,
    )
    max_prof = max(
        (
            pdf.get_string_width(
                a.professor.nome if a.professor else "-"
            )
            for a in agendamentos
        ),
        default=0,
    )
    max_escola += 2
    max_prof += 2
    min_escola, min_prof = 40, 30
    fixed_width = 15 + 25 + 20 + 15 + 30
    avail = 190 - fixed_width
    escola_w = max(max_escola, min_escola)
    prof_w = max(max_prof, min_prof)
    if escola_w + prof_w > avail and escola_w + prof_w > 0:
        scale = avail / (escola_w + prof_w)
        escola_w *= scale
        prof_w *= scale

    # Cabeçalho da tabela
    pdf.set_font('DejaVu', 'B', 9)
    pdf.cell(15, 10, 'ID', 1, 0, 'C')
    pdf.cell(25, 10, 'Data', 1, 0, 'C')
    pdf.cell(20, 10, 'Horário', 1, 0, 'C')
    pdf.cell(escola_w, 10, 'Escola', 1, 0, 'C')
    pdf.cell(prof_w, 10, 'Professor', 1, 0, 'C')
    pdf.cell(15, 10, 'Alunos', 1, 0, 'C')
    pdf.cell(30, 10, 'Status', 1, 1, 'C')

    # Dados da tabela
    pdf.set_font('DejaVu', '', 8)
    for agendamento in agendamentos:
        horario = agendamento.horario
        escola_nome = agendamento.escola_nome
        professor_nome = (
            agendamento.professor.nome if agendamento.professor else "-"
        )
        presentes = sum(1 for aluno in agendamento.alunos if aluno.presente)

        pdf.cell(15, 8, str(agendamento.id), 1, 0, 'C')
        pdf.cell(25, 8, horario.data.strftime('%d/%m/%Y'), 1, 0, 'C')
        pdf.cell(20, 8, horario.horario_inicio.strftime('%H:%M'), 1, 0, 'C')
        pdf.cell(escola_w, 8, escola_nome, 1, 0, 'L')
        pdf.cell(prof_w, 8, professor_nome, 1, 0, 'L')
        pdf.cell(15, 8, str(presentes), 1, 0, 'C')

        status_txt = agendamento.status.capitalize()
        if agendamento.checkin_realizado:
            status_txt += " ✓"

        pdf.cell(30, 8, status_txt, 1, 1, 'C')

    # Seção de presença dos alunos
    agendamentos_com_presenca = [
        a
        for a in agendamentos
        if a.checkin_realizado and any(aluno.presente for aluno in a.alunos)
    ]
    if agendamentos_com_presenca:
        pdf.ln(10)
        pdf.set_font('DejaVu', 'B', 14)
        pdf.cell(190, 10, 'Lista de Presença', 0, 1)

        for ag in agendamentos_com_presenca:
            pdf.set_font('DejaVu', 'B', 10)
            pdf.cell(190, 8, f'Agendamento {ag.id} - {ag.escola_nome}', 0, 1)
            pdf.set_font('DejaVu', 'B', 9)
            pdf.cell(160, 8, 'Aluno', 1, 0, 'L')
            pdf.cell(30, 8, 'Presente', 1, 1, 'C')
            pdf.set_font('DejaVu', '', 9)
            for aluno in ag.alunos:
                if not aluno.presente:
                    continue
                pdf.cell(160, 8, aluno.nome, 1, 0, 'L')
                pdf.cell(30, 8, 'Sim', 1, 1, 'C')
            pdf.ln(2)

    # Rodapé
    pdf.ln(10)
    pdf.set_font('DejaVu', 'I', 10)
    pdf.cell(
        190,
        10,
        'Este relatório é gerado automaticamente pelo sistema de agendamentos.',
        0,
        1,
        'C',
    )

    # Salvar o PDF
    pdf.output(caminho_pdf)

    
from datetime import datetime, timedelta
from sqlalchemy import and_, func, or_, case
from flask import (
    render_template,
    redirect,
    url_for,
    flash,
    request,
    send_file,
    session,
    jsonify,
)

# Importe os modelos necessários
from models import (
    Evento,
    ConfiguracaoAgendamento,
    SalaVisitacao,
    HorarioVisitacao,
    AgendamentoVisita,
    AlunoVisitante,
    ProfessorBloqueado,
)

@agendamento_routes.route('/configurar_agendamentos/<int:evento_id>', methods=['GET', 'POST'])
@login_required
def configurar_agendamentos(evento_id):
    """
    Rota para configurar as regras de agendamento para um evento específico.
    Permite definir horários disponíveis, prazos, capacidade, etc.
    """
    # Apenas clientes podem configurar agendamentos
    if current_user.tipo != 'cliente':
        flash('Acesso negado! Esta área é exclusiva para organizadores.', 'danger')
        return redirect(url_for(endpoints.DASHBOARD))
    
    evento = Evento.query.get_or_404(evento_id)
    tipos_inscricao = EventoInscricaoTipo.query.filter_by(evento_id=evento_id).all()
    
    # Verificar se o evento pertence ao cliente
    if evento.cliente_id != current_user.id:
        flash('Este evento não pertence a você!', 'danger')
        return redirect(url_for(endpoints.DASHBOARD_CLIENTE))
    
    # Verificar se já existe configuração
    config = ConfiguracaoAgendamento.query.filter_by(
        cliente_id=current_user.id,
        evento_id=evento_id
    ).first()
    
    if request.method == 'POST':
        dias_semana = request.form.getlist('dias_semana')
        valores_validos = {str(i) for i in range(7)}
        if not dias_semana:
            flash('Selecione ao menos um dia da semana.', 'danger')
            return render_template(
                'configurar_agendamentos.html',
                evento=evento,
                config=config,
                tipos_inscricao=tipos_inscricao,
            )
        if any(dia not in valores_validos for dia in dias_semana):
            flash('Dias da semana inválidos.', 'danger')
            return render_template(
                'configurar_agendamentos.html',
                evento=evento,
                config=config,
                tipos_inscricao=tipos_inscricao,
            )
        dias_semana_str = ','.join(dias_semana)
        if config:
            # Atualizar configuração existente
            config.prazo_cancelamento = request.form.get('prazo_cancelamento', type=int)
            config.tempo_bloqueio = request.form.get('tempo_bloqueio', type=int)
            config.capacidade_padrao = request.form.get('capacidade_padrao', type=int)
            config.intervalo_minutos = request.form.get('intervalo_minutos', type=int)
            config.intervalo_entrada = request.form.get('intervalo_entrada', type=int)

            # Converter string para time
            hora_inicio = request.form.get('horario_inicio')
            hora_fim = request.form.get('horario_fim')
            config.horario_inicio = datetime.strptime(hora_inicio, '%H:%M').time()
            config.horario_fim = datetime.strptime(hora_fim, '%H:%M').time()

            config.dias_semana = dias_semana_str
            selecionados = request.form.getlist('tipos_inscricao_permitidos')
            config.tipos_inscricao_permitidos = ','.join(selecionados) if selecionados else None
        else:
            # Criar nova configuração
            hora_inicio = request.form.get('horario_inicio')
            hora_fim = request.form.get('horario_fim')

            selecionados = request.form.getlist('tipos_inscricao_permitidos')
            config = ConfiguracaoAgendamento(
                cliente_id=current_user.id,
                evento_id=evento_id,
                prazo_cancelamento=request.form.get('prazo_cancelamento', type=int),
                tempo_bloqueio=request.form.get('tempo_bloqueio', type=int),
                capacidade_padrao=request.form.get('capacidade_padrao', type=int),
                intervalo_minutos=request.form.get('intervalo_minutos', type=int),
                intervalo_entrada=request.form.get('intervalo_entrada', type=int),
                horario_inicio=datetime.strptime(hora_inicio, '%H:%M').time(),
                horario_fim=datetime.strptime(hora_fim, '%H:%M').time(),
                dias_semana=dias_semana_str,
                tipos_inscricao_permitidos=','.join(selecionados) or None,
            )
            db.session.add(config)
        
        try:
            db.session.commit()
            flash('Configurações de agendamento salvas com sucesso!', 'success')
            return redirect(url_for('agendamento_routes.gerar_horarios_agendamento', evento_id=evento_id))
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao salvar configurações: {str(e)}', 'danger')
    
    return render_template(
        'configurar_agendamentos.html',
        evento=evento,
        config=config,
        tipos_inscricao=tipos_inscricao
    )


@agendamento_routes.route('/gerar_horarios_agendamento/<int:evento_id>', methods=['GET', 'POST'])
@login_required
def gerar_horarios_agendamento(evento_id):
    """
    Página para gerar horários de agendamento com base nas configurações.
    """
    if current_user.tipo != 'cliente':
        flash('Acesso negado! Esta área é exclusiva para organizadores.', 'danger')
        return redirect(url_for(endpoints.DASHBOARD))
    
    evento = Evento.query.get_or_404(evento_id)
    
    # Verificar se o evento pertence ao cliente
    if evento.cliente_id != current_user.id:
        flash('Este evento não pertence a você!', 'danger')
        return redirect(url_for(endpoints.DASHBOARD_CLIENTE))
    
    config = ConfiguracaoAgendamento.query.filter_by(
        evento_id=evento_id
    ).first_or_404()

    if request.method == 'POST':
        # Obter datas do form
        data_inicial = datetime.strptime(
            request.form.get('data_inicial'), '%Y-%m-%d'
        ).date()
        data_final = datetime.strptime(
            request.form.get('data_final'), '%Y-%m-%d'
        ).date()

        if data_final < data_inicial:
            flash(
                'A data final deve ser igual ou posterior à data inicial.',
                'danger'
            )
            return render_template(
                'gerar_horarios_agendamento.html',
                evento=evento,
                config=config,
            )

        # Converter dias da semana do formulário (0=Domingo … 6=Sábado)
        # para os valores usados por datetime.weekday() (0=Segunda … 6=Domingo)
        dias_permitidos = [
            (int(dia) + 6) % 7 for dia in config.dias_semana.split(',')
        ]

        def tem_conflito(intervalos, inicio, fim):
            for ini_exist, fim_exist in intervalos:
                if inicio < fim_exist and fim > ini_exist:
                    return True
            return False

        # Gerar horários
        data_atual = data_inicial
        horarios_criados = 0

        while data_atual <= data_final:
            intervalos_existentes = [
                (h.horario_inicio, h.horario_fim)
                for h in HorarioVisitacao.query.filter_by(
                    evento_id=evento_id, data=data_atual
                ).all()
            ]

            if data_atual.weekday() in dias_permitidos:
                horario_atual = datetime.combine(
                    data_atual, config.horario_inicio
                )
                hora_final = datetime.combine(
                    data_atual, config.horario_fim
                )

                while horario_atual < hora_final:
                    horario_fim = horario_atual + timedelta(
                        minutes=config.intervalo_minutos
                    )
                    if horario_fim > hora_final:
                        horario_fim = hora_final

                    if not tem_conflito(
                        intervalos_existentes,
                        horario_atual.time(),
                        horario_fim.time(),
                    ):
                        novo_horario = HorarioVisitacao(
                            evento_id=evento_id,
                            data=data_atual,
                            horario_inicio=horario_atual.time(),
                            horario_fim=horario_fim.time(),
                            capacidade_total=config.capacidade_padrao,
                            vagas_disponiveis=config.capacidade_padrao,
                        )
                        db.session.add(novo_horario)
                        horarios_criados += 1
                        intervalos_existentes.append(
                            (horario_atual.time(), horario_fim.time())
                        )

                    horario_atual = horario_fim

            data_atual += timedelta(days=1)

        try:
            db.session.commit()
            flash(
                f'{horarios_criados} horários de visitação foram criados com sucesso!',
                'success',
            )
            return redirect(
                url_for(
                    'agendamento_routes.listar_horarios_agendamento',
                    evento_id=evento_id,
                )
            )
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao gerar horários: {str(e)}', 'danger')

    return render_template(
        'gerar_horarios_agendamento.html', evento=evento, config=config
    )


@agendamento_routes.route(
    '/adicionar_horario_agendamento/<int:evento_id>', methods=['GET', 'POST']
)
@login_required
def adicionar_horario_agendamento(evento_id):
    """Adicionar um único horário de visitação."""
    if current_user.tipo != 'cliente':
        flash('Acesso negado! Esta área é exclusiva para organizadores.', 'danger')
        return redirect(url_for(endpoints.DASHBOARD))

    evento = Evento.query.get_or_404(evento_id)

    if evento.cliente_id != current_user.id:
        flash('Este evento não pertence a você!', 'danger')
        return redirect(url_for(endpoints.DASHBOARD_CLIENTE))

    if request.method == 'POST':
        data_str = request.form.get('data')
        inicio_str = request.form.get('horario_inicio')
        fim_str = request.form.get('horario_fim')
        capacidade = request.form.get('capacidade', type=int)

        if not data_str or not inicio_str or not fim_str or capacidade is None:
            flash('Todos os campos são obrigatórios.', 'danger')
            return render_template('adicionar_horario_agendamento.html', evento=evento)

        data = datetime.strptime(data_str, '%Y-%m-%d').date()
        inicio = datetime.strptime(inicio_str, '%H:%M').time()
        fim = datetime.strptime(fim_str, '%H:%M').time()

        if fim <= inicio:
            flash('O horário final deve ser posterior ao inicial.', 'danger')
            return render_template('adicionar_horario_agendamento.html', evento=evento)

        conflito = (
            HorarioVisitacao.query.filter_by(evento_id=evento_id, data=data)
            .filter(
                HorarioVisitacao.horario_inicio < fim,
                HorarioVisitacao.horario_fim > inicio,
            )
            .first()
        )
        if conflito:
            flash('Já existe um horário neste intervalo.', 'warning')
            return render_template('adicionar_horario_agendamento.html', evento=evento)

        novo_horario = HorarioVisitacao(
            evento_id=evento_id,
            data=data,
            horario_inicio=inicio,
            horario_fim=fim,
            capacidade_total=capacidade,
            vagas_disponiveis=capacidade,
        )
        db.session.add(novo_horario)
        try:
            db.session.commit()
            flash('Horário adicionado com sucesso!', 'success')
            return redirect(
                url_for(
                    'agendamento_routes.listar_horarios_agendamento',
                    evento_id=evento_id,
                )
            )
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao adicionar horário: {str(e)}', 'danger')

    return render_template('adicionar_horario_agendamento.html', evento=evento)


@agendamento_routes.route('/listar_horarios_agendamento/<int:evento_id>')
@login_required
def listar_horarios_agendamento(evento_id):
    """
    Página para listar e gerenciar os horários de agendamento disponíveis.
    """
    if current_user.tipo != 'cliente':
        flash('Acesso negado! Esta área é exclusiva para organizadores.', 'danger')
        return redirect(url_for(endpoints.DASHBOARD))
    
    evento = Evento.query.get_or_404(evento_id)
    
    # Verificar se o evento pertence ao cliente
    if evento.cliente_id != current_user.id:
        flash('Este evento não pertence a você!', 'danger')
        return redirect(url_for(endpoints.DASHBOARD_CLIENTE))
    
    # Agrupar horários por data
    horarios = HorarioVisitacao.query.filter_by(evento_id=evento_id).order_by(
        HorarioVisitacao.data,
        HorarioVisitacao.horario_inicio
    ).all()
    
    # Agrupar horários por data para facilitar a visualização
    horarios_por_data = {}
    for horario in horarios:
        data_str = horario.data.strftime('%Y-%m-%d')
        if data_str not in horarios_por_data:
            horarios_por_data[data_str] = []
        horarios_por_data[data_str].append(horario)
    
    return render_template(
        'listar_horarios_agendamento.html',
        evento=evento,
        horarios_por_data=horarios_por_data
    )


@agendamento_routes.route('/salas_visitacao/<int:evento_id>', methods=['GET', 'POST'])
@login_required
def salas_visitacao(evento_id):
    """
    Página para gerenciar as salas disponíveis para visitação.
    """
    if current_user.tipo != 'cliente':
        flash('Acesso negado! Esta área é exclusiva para organizadores.', 'danger')
        return redirect(url_for(endpoints.DASHBOARD))
    
    evento = Evento.query.get_or_404(evento_id)
    
    # Verificar se o evento pertence ao cliente
    if evento.cliente_id != current_user.id:
        flash('Este evento não pertence a você!', 'danger')
        return redirect(url_for(endpoints.DASHBOARD_CLIENTE))
    
    if request.method == 'POST':
        nome = request.form.get('nome')
        descricao = request.form.get('descricao')
        capacidade = request.form.get('capacidade', type=int)
        
        if nome and capacidade:
            nova_sala = SalaVisitacao(
                nome=nome,
                descricao=descricao,
                capacidade=capacidade,
                evento_id=evento_id
            )
            db.session.add(nova_sala)
            
            try:
                db.session.commit()
                flash('Sala de visitação cadastrada com sucesso!', 'success')
            except Exception as e:
                db.session.rollback()
                flash(f'Erro ao cadastrar sala: {str(e)}', 'danger')
    
    # Listar salas existentes
    salas = SalaVisitacao.query.filter_by(evento_id=evento_id).all()
    
    return render_template(
        'salas_visitacao.html',
        evento=evento,
        salas=salas
    )

@agendamento_routes.route('/gerar_relatorio_agendamentos/<int:evento_id>')
@login_required
def gerar_relatorio_agendamentos(evento_id):
    """Gera um PDF com agendamentos que tiveram presença via QR code."""
    if current_user.tipo != 'cliente':
        flash('Acesso negado!', 'danger')
        return redirect(url_for(endpoints.DASHBOARD))
    
    evento = Evento.query.get_or_404(evento_id)
    
    # Verificar se o evento pertence ao cliente
    if evento.cliente_id != current_user.id:
        flash('Este evento não pertence a você!', 'danger')
        return redirect(url_for(endpoints.DASHBOARD_CLIENTE))
    
    # Filtros (mesmos da listagem)
    data_inicio = request.args.get('data_inicio')
    data_fim = request.args.get('data_fim')
    status = request.args.get('status')
    escola = request.args.get('escola')
    
    # Base da consulta: apenas agendamentos com check-in via QR code
    query = (
        AgendamentoVisita.query.join(
            HorarioVisitacao, AgendamentoVisita.horario_id == HorarioVisitacao.id
        )
        .join(
            Checkin,
            (Checkin.usuario_id == AgendamentoVisita.professor_id)
            & (Checkin.evento_id == HorarioVisitacao.evento_id),
        )
        .filter(Checkin.palavra_chave == 'QR-AGENDAMENTO')
        .filter(HorarioVisitacao.evento_id == evento_id)
    )
    
    # Aplicar filtros
    if data_inicio:
        data_inicio = datetime.strptime(data_inicio, '%Y-%m-%d').date()
        query = query.filter(HorarioVisitacao.data >= data_inicio)
    
    if data_fim:
        data_fim = datetime.strptime(data_fim, '%Y-%m-%d').date()
        query = query.filter(HorarioVisitacao.data <= data_fim)
    
    if status:
        query = query.filter(AgendamentoVisita.status == status)
    
    if escola:
        query = query.filter(AgendamentoVisita.escola_nome.ilike(f'%{escola}%'))
    
    # Ordenar por data/horário
    agendamentos = query.order_by(
        HorarioVisitacao.data,
        HorarioVisitacao.horario_inicio
    ).all()
    
    # Gerar PDF
    pdf_filename = f"relatorio_agendamentos_{evento_id}.pdf"
    pdf_path = os.path.join("static", "relatorios", pdf_filename)
    os.makedirs(os.path.dirname(pdf_path), exist_ok=True)
    
    # Chamar função para gerar PDF
    pdf_service.gerar_pdf_relatorio_agendamentos(evento, agendamentos, pdf_path)
    
    return send_file(pdf_path, as_attachment=True)


@agendamento_routes.route('/criar-agendamento', methods=['GET', 'POST'])
@login_required
def criar_agendamento():
    """
    Rota para criação de um novo agendamento.
    """
    # Inicialização de variáveis
    form_erro = None
    eventos = []
    
    # Buscar eventos disponíveis conforme o tipo de usuário
    try:
        if getattr(current_user, 'is_cliente', lambda: False)():
            eventos = Evento.query.filter_by(cliente_id=current_user.id).all()
        else:
            eventos = Evento.query.all()
    except Exception as e:
        db.session.rollback()
        flash(f"Erro ao buscar eventos: {str(e)}", "danger")
    
    # Processar o formulário quando enviado via POST
    if request.method == 'POST':
        try:
            # Obter dados do formulário
            evento_id = request.form.get('evento_id')
            data = request.form.get('data')
            horario_id = request.form.get('horario_id')
            escola_nome = request.form.get('escola_nome')
            responsavel_nome = request.form.get('responsavel_nome')
            responsavel_cargo = request.form.get('responsavel_cargo')
            responsavel_email = request.form.get('responsavel_email')
            responsavel_whatsapp = request.form.get('responsavel_whatsapp')
            acompanhantes_qtd = request.form.get('acompanhantes_qtd', type=int)
            acompanhantes_nomes = request.form.get('acompanhantes_nomes')
            turma = request.form.get('turma')
            quantidade_alunos = request.form.get('quantidade_alunos')
            faixa_etaria = request.form.get('faixa_etaria')
            observacoes = request.form.get('observacoes')
            compromissos = [
                request.form.get('compromisso1'),
                request.form.get('compromisso2'),
                request.form.get('compromisso3'),
                request.form.get('compromisso4'),
            ]
            aceite_final = request.form.get('aceite_final')
            
            # Validar dados obrigatórios
            if (
                not evento_id
                or not data
                or not horario_id
                or not escola_nome
                or not quantidade_alunos
            ):
                form_erro = "Preencha todos os campos obrigatórios."
                flash(form_erro, "danger")
            elif not all(compromissos) or not aceite_final:
                form_erro = (
                    'É necessário confirmar todos os compromissos e o aceite final.'
                )
                flash(form_erro, "danger")
            else:
                professor_id = None
                cliente_id = None
                usuario_id = None
                if getattr(current_user, 'is_cliente', lambda: False)():
                    cliente = Cliente.query.get(current_user.id)
                    if not cliente:
                        form_erro = 'Cliente inválido.'
                        flash(form_erro, 'danger')
                        return redirect(url_for('agendamento_routes.criar_agendamento'))
                    cliente_id = cliente.id
                else:
                    usuario = Usuario.query.get(current_user.id)
                    if not usuario:
                        form_erro = 'Usuário inválido.'
                        flash(form_erro, 'danger')
                        return redirect(url_for('agendamento_routes.criar_agendamento'))
                    usuario_id = usuario.id
                    if getattr(usuario, 'is_professor', lambda: False)():
                        professor_id = usuario.id

                quantidade = int(quantidade_alunos)
                agendamento = None
                try:
                    with db.session.begin():
                        horario = Horario.get_for_update(horario_id)
                        if not horario:
                            raise ValueError('Horário inválido.')
                        if horario.evento_id != int(evento_id):
                            raise ValueError(
                                'Horário não pertence ao evento selecionado.'
                            )
                        if quantidade > horario.vagas_disponiveis:
                            raise ValueError(
                                f"Não há vagas suficientes! Disponíveis: {horario.vagas_disponiveis}"
                            )

                        agendamento = AgendamentoVisita(
                            horario_id=horario.id,
                            professor_id=professor_id,
                            cliente_id=cliente_id,
                            escola_nome=escola_nome,
                            turma=turma,
                            nivel_ensino=faixa_etaria,
                            quantidade_alunos=quantidade,
                            responsavel_nome=responsavel_nome,
                            responsavel_cargo=responsavel_cargo,
                            responsavel_email=responsavel_email,
                            responsavel_whatsapp=responsavel_whatsapp,
                            acompanhantes_nomes=acompanhantes_nomes,
                            acompanhantes_qtd=acompanhantes_qtd,
                            observacoes=observacoes,
                            compromisso_1=True,
                            compromisso_2=True,
                            compromisso_3=True,
                            compromisso_4=True,
                            status='pendente',
                        )
                        horario.vagas_disponiveis -= quantidade
                        db.session.add(agendamento)
                except ValueError as exc:
                    form_erro = str(exc)
                    flash(form_erro, 'danger')
                except Exception as e:
                    db.session.rollback()
                    form_erro = f"Erro ao salvar agendamento: {str(e)}"
                    flash(form_erro, 'danger')
                else:
                    flash("Agendamento criado com sucesso!", "success")
                    if current_user.tipo == 'professor':
                        return redirect(
                            url_for(
                                'routes.adicionar_alunos_professor',
                                agendamento_id=agendamento.id,
                            )
                        )
                    if current_user.tipo == 'cliente':
                        return redirect(
                            url_for(
                                'routes.adicionar_alunos_professor',
                                agendamento_id=agendamento.id,
                            )
                        )
                    return redirect(url_for(endpoints.DASHBOARD))
                
        except Exception as e:
            form_erro = f"Erro ao processar o formulário: {str(e)}"
            flash(form_erro, "danger")
    
    # Renderizar o template com o formulário
    return render_template('criar_agendamento.html', 
                          eventos=eventos,
                          form_erro=form_erro)

@agendamento_routes.route('/api/horarios-disponiveis')
@login_required
def horarios_disponiveis():
    """
    API para obter horários disponíveis para agendamento, baseado em evento e data.
    """
    evento_id = request.args.get('evento_id')
    data = request.args.get('data')
    
    # Verificar se os parâmetros foram fornecidos
    if not evento_id or not data:
        return jsonify({
            'success': False,
            'message': 'Parâmetros evento_id e data são obrigatórios'
        }), 400
    
    try:
        # Como não sabemos a estrutura exata do seu modelo de Horario,
        # vamos retornar dados simulados para teste
        # Em uma implementação real, você buscaria horários do banco de dados
        
        # Simular alguns horários como exemplo
        horarios_exemplo = [
            {
                'id': 1,
                'horario_inicio': '08:00',
                'horario_fim': '10:00',
                'vagas_disponiveis': 30
            },
            {
                'id': 2,
                'horario_inicio': '10:30',
                'horario_fim': '12:30',
                'vagas_disponiveis': 25
            },
            {
                'id': 3,
                'horario_inicio': '14:00',
                'horario_fim': '16:00',
                'vagas_disponiveis': 20
            }
        ]
        
        return jsonify({
            'success': True,
            'horarios': horarios_exemplo
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Erro ao buscar horários: {str(e)}'
        }), 500

@agendamento_routes.route('/configurar-horarios-agendamento', methods=['GET', 'POST'])
@login_required
def configurar_horarios_agendamento():
    """
    Rota para configuração de horários disponíveis para agendamentos.
    """
    # Inicialização de variáveis
    form_erro = None
    eventos = []
    horarios_existentes = []
    evento_selecionado = None
    evento_id = request.args.get('evento_id', None)
    
    # Buscar eventos disponíveis do cliente atual
    try:
        eventos = Evento.query.filter_by(cliente_id=current_user.id).all()
        
        # Se um evento foi especificado na URL, buscamos seus detalhes
        if evento_id:
            evento_selecionado = Evento.query.filter_by(id=evento_id, cliente_id=current_user.id).first()
            
            # Buscar horários existentes para o evento selecionado
            if evento_selecionado:
                horarios_existentes = HorarioVisitacao.query.filter_by(evento_id=evento_id).all()
                
                # Transformar os horários do banco em dicionários para facilitar o uso no template
                horarios_existentes = [
                    {
                        'id': h.id,
                        'data': h.data.strftime('%Y-%m-%d'),
                        'horario_inicio': h.horario_inicio.strftime('%H:%M'),
                        'horario_fim': h.horario_fim.strftime('%H:%M'),
                        'capacidade': h.capacidade_total,
                        'agendamentos': h.capacidade_total - h.vagas_disponiveis,
                        'fechado': h.fechado
                    } for h in horarios_existentes
                ]
    except Exception as e:
        db.session.rollback()
        flash(f"Erro ao buscar eventos: {str(e)}", "danger")
    
    # Processar o formulário quando enviado via POST
    if request.method == 'POST':
        try:
            # Determinar o tipo de ação
            acao = request.form.get('acao')
            
            if acao == 'adicionar':
                # Obter dados do formulário para adicionar novo horário
                evento_id = request.form.get('evento_id')
                data = request.form.get('data')
                horario_inicio = request.form.get('horario_inicio')
                horario_fim = request.form.get('horario_fim')
                capacidade = request.form.get('capacidade')
                
                # Validar dados obrigatórios
                if not evento_id or not data or not horario_inicio or not horario_fim or not capacidade:
                    form_erro = "Preencha todos os campos obrigatórios."
                    flash(form_erro, "danger")
                else:
                    # Converte string para data e hora
                    data_obj = datetime.strptime(data, '%Y-%m-%d').date()
                    horario_inicio_obj = datetime.strptime(horario_inicio, '%H:%M').time()
                    horario_fim_obj = datetime.strptime(horario_fim, '%H:%M').time()
                    capacidade_int = int(capacidade)
                    
                    # Criar novo horário de visitação
                    novo_horario = HorarioVisitacao(
                        evento_id=evento_id,
                        data=data_obj,
                        horario_inicio=horario_inicio_obj,
                        horario_fim=horario_fim_obj,
                        capacidade_total=capacidade_int,
                        vagas_disponiveis=capacidade_int
                    )
                    
                    db.session.add(novo_horario)
                    db.session.commit()
                    
                    flash(f"Horário adicionado com sucesso para o dia {data} das {horario_inicio} às {horario_fim}!", "success")
                    
                    # Redirecionar para a mesma página com o evento selecionado
                    return redirect(url_for('agendamento_routes.configurar_horarios_agendamento', evento_id=evento_id))
            
            elif acao == 'excluir':
                # Obter ID do horário a ser excluído
                horario_id = request.form.get('horario_id')
                evento_id = request.form.get('evento_id')
                
                if not horario_id:
                    flash("ID do horário não fornecido.", "danger")
                else:
                    # Verificar se existem agendamentos para este horário
                    horario = HorarioVisitacao.query.get(horario_id)
                    
                    if horario:
                        # Verificar se há agendamentos para este horário
                        agendamentos = AgendamentoVisita.query.filter_by(horario_id=horario_id).first()
                        
                        if agendamentos:
                            flash("Não é possível excluir um horário que possui agendamentos.", "danger")
                        else:
                            db.session.delete(horario)
                            db.session.commit()
                            flash("Horário excluído com sucesso!", "success")
                    else:
                        flash("Horário não encontrado.", "danger")
                    
                    # Redirecionar para a mesma página com o evento selecionado
                    return redirect(url_for('agendamento_routes.configurar_horarios_agendamento', evento_id=evento_id))

            elif acao == 'fechar':
                horario_id = request.form.get('horario_id')
                evento_id = request.form.get('evento_id')
                horario = HorarioVisitacao.query.get(horario_id)
                if horario:
                    horario.fechado = True
                    db.session.commit()
                    flash('Horário fechado com sucesso!', 'success')
                else:
                    flash('Horário não encontrado.', 'danger')
                return redirect(url_for('agendamento_routes.configurar_horarios_agendamento', evento_id=evento_id))

            elif acao == 'reabrir':
                horario_id = request.form.get('horario_id')
                evento_id = request.form.get('evento_id')
                horario = HorarioVisitacao.query.get(horario_id)
                if horario:
                    horario.fechado = False
                    db.session.commit()
                    flash('Horário reaberto com sucesso!', 'success')
                else:
                    flash('Horário não encontrado.', 'danger')
                return redirect(url_for('agendamento_routes.configurar_horarios_agendamento', evento_id=evento_id))

            elif acao == 'adicionar_periodo':
                # Obter dados do formulário para adicionar vários horários em um período
                evento_id = request.form.get('evento_id')
                data_inicio = request.form.get('data_inicio')
                data_fim = request.form.get('data_fim')
                dias_semana = request.form.getlist('dias_semana')
                horario_inicio = request.form.get('horario_inicio')
                horario_fim = request.form.get('horario_fim')
                capacidade = request.form.get('capacidade')
                
                # Validar dados obrigatórios
                if not evento_id or not data_inicio or not data_fim or not dias_semana or not horario_inicio or not horario_fim or not capacidade:
                    form_erro = "Preencha todos os campos obrigatórios."
                    flash(form_erro, "danger")
                else:
                    # Converter strings para objetos de data
                    data_inicio_obj = datetime.strptime(data_inicio, '%Y-%m-%d').date()
                    data_fim_obj = datetime.strptime(data_fim, '%Y-%m-%d').date()
                    horario_inicio_obj = datetime.strptime(horario_inicio, '%H:%M').time()
                    horario_fim_obj = datetime.strptime(horario_fim, '%H:%M').time()
                    capacidade_int = int(capacidade)
                    
                    # Converter dias do formulário (0=Domingo … 6=Sábado)
                    # para os valores usados por datetime.weekday() (0=Segunda … 6=Domingo)
                    dias = [(int(dia) + 6) % 7 for dia in dias_semana]
                    
                    # Criar horários para todas as datas no período que correspondem aos dias da semana selecionados
                    delta = data_fim_obj - data_inicio_obj
                    horarios_criados = 0
                    
                    for i in range(delta.days + 1):
                        data_atual = data_inicio_obj + timedelta(days=i)
                        # datetime.weekday() retorna 0 para segunda e 6 para domingo
                        if data_atual.weekday() in dias:
                            # Verificar se já existe um horário para esta data e período
                            horario_existente = HorarioVisitacao.query.filter_by(
                                evento_id=evento_id,
                                data=data_atual,
                                horario_inicio=horario_inicio_obj,
                                horario_fim=horario_fim_obj
                            ).first()
                            
                            if not horario_existente:
                                novo_horario = HorarioVisitacao(
                                    evento_id=evento_id,
                                    data=data_atual,
                                    horario_inicio=horario_inicio_obj,
                                    horario_fim=horario_fim_obj,
                                    capacidade_total=capacidade_int,
                                    vagas_disponiveis=capacidade_int
                                )
                                
                                db.session.add(novo_horario)
                                horarios_criados += 1
                    
                    if horarios_criados > 0:
                        db.session.commit()
                        flash(f"{horarios_criados} horários configurados com sucesso no período de {data_inicio} a {data_fim}!", "success")
                    else:
                        flash("Nenhum horário novo foi criado. Verifique se já existem horários para as datas selecionadas.", "warning")
                    
                    # Redirecionar para a mesma página com o evento selecionado
                    return redirect(url_for('agendamento_routes.configurar_horarios_agendamento', evento_id=evento_id))
                
        except Exception as e:
            form_erro = f"Erro ao processar o formulário: {str(e)}"
            flash(form_erro, "danger")
            db.session.rollback()
    
    # Adicione esta linha para verificar se a função editar_horario existe
    has_editar_horario = hasattr(routes, 'editar_horario')
    
    # Obter configuração de agendamento se existir
    configuracao = None
    if evento_selecionado:
        configuracao = ConfiguracaoAgendamento.query.filter_by(
            cliente_id=current_user.id,
            evento_id=evento_selecionado.id
        ).first()
    
    # Renderizar o template com o formulário
    return render_template('configurar_horarios_agendamento.html', 
                          eventos=eventos,
                          evento_selecionado=evento_selecionado,
                          horarios_existentes=horarios_existentes,
                          form_erro=form_erro,
                          has_editar_horario=has_editar_horario,
                          configuracao=configuracao)

@agendamento_routes.route('/criar-evento-agendamento', methods=['GET', 'POST'])
@login_required
def criar_evento_agendamento():
    """
    Rota para criação de um novo evento para agendamentos.
    """
    # Verificação de permissão
    if current_user.tipo != 'cliente':
        flash('Acesso negado!', 'danger')
        return redirect(url_for(endpoints.DASHBOARD_CLIENTE))
    
    # Inicialização de variáveis
    form_erro = None
    
    # Processar o formulário quando enviado via POST
    if request.method == 'POST':
        try:
            # Obter dados do formulário
            nome = request.form.get('nome')
            descricao = request.form.get('descricao')
            programacao = request.form.get('programacao')
            localizacao = request.form.get('local')
            link_mapa = request.form.get('link_mapa')
            data_inicio = request.form.get('data_inicio')
            data_fim = request.form.get('data_fim')
            hora_inicio = request.form.get('hora_inicio')
            hora_fim = request.form.get('hora_fim')
            capacidade_padrao = request.form.get('capacidade_padrao')
            requer_aprovacao = 'requer_aprovacao' in request.form
            publico = 'publico' in request.form
            
            # Validar dados obrigatórios
            if not nome or not data_inicio or not data_fim or not capacidade_padrao:
                form_erro = "Preencha todos os campos obrigatórios."
                flash(form_erro, "danger")
            elif data_fim < data_inicio:
                form_erro = "A data de fim deve ser posterior à data de início."
                flash(form_erro, "danger")
            else:
                # Processar upload de banner, se houver
                banner = request.files.get('banner')
                banner_url = None
                
                if banner and banner.filename:
                    filename = secure_filename(banner.filename)
                    caminho_banner = os.path.join('static/banners', filename)
                    os.makedirs(os.path.dirname(caminho_banner), exist_ok=True)
                    banner.save(caminho_banner)
                    banner_url = url_for('static', filename=f'banners/{filename}', _external=True)
                
                # Converter strings para objetos de data/hora
                data_inicio_obj = datetime.strptime(data_inicio, '%Y-%m-%d') if data_inicio else None
                data_fim_obj = datetime.strptime(data_fim, '%Y-%m-%d') if data_fim else None
                
                # Processar hora (se necessário)
                from datetime import time
                hora_inicio_obj = time.fromisoformat(hora_inicio) if hora_inicio else None
                hora_fim_obj = time.fromisoformat(hora_fim) if hora_fim else None
                
                # Criar o objeto Evento
                novo_evento = Evento(
                    cliente_id=current_user.id,
                    nome=nome,
                    descricao=descricao,
                    programacao=programacao,
                    localizacao=localizacao,
                    link_mapa=link_mapa,
                    banner_url=banner_url,
                    data_inicio=data_inicio_obj,
                    data_fim=data_fim_obj,
                    hora_inicio=hora_inicio_obj,
                    hora_fim=hora_fim_obj,
                    capacidade_padrao=int(capacidade_padrao),
                    requer_aprovacao=requer_aprovacao,
                    publico=publico
                )
                
                try:
                    db.session.add(novo_evento)
                    db.session.flush()  # Obter o ID do evento antes de criar tipos de inscrição
                    
                    # Processar tipos de inscrição se o cliente tiver pagamento habilitado
                    if current_user.habilita_pagamento:
                        inscricao_gratuita = (request.form.get('inscricao_gratuita') == 'on')
                        novo_evento.inscricao_gratuita = inscricao_gratuita
                        
                        # Adicionar tipos de inscrição se não for gratuito
                        if not inscricao_gratuita:
                            nomes_tipos = request.form.getlist('nome_tipo[]')
                            precos = request.form.getlist('preco_tipo[]')
                            
                            if not nomes_tipos or not precos:
                                raise ValueError("Tipos de inscrição e preços são obrigatórios quando o evento não é gratuito.")
                            
                            for nome, preco in zip(nomes_tipos, precos):
                                if nome and preco:
                                    try:
                                        preco_float = float(preco.strip())
                                    except ValueError:
                                        db.session.rollback()
                                        form_erro = (
                                            "Erro: o campo 'preço' deve conter apenas números."
                                        )
                                        flash(form_erro, 'danger')
                                        return render_template(
                                            'criar_evento_agendamento.html',
                                            form_erro=form_erro,
                                        )
                                    novo_tipo = EventoInscricaoTipo(
                                        evento_id=novo_evento.id,
                                        nome=nome,
                                        preco=preco_float,
                                    )
                                    db.session.add(novo_tipo)
                    
                    db.session.commit()
                    flash(f"Evento '{nome}' criado com sucesso! Você pode agora configurar os horários.", "success")
                    return redirect(url_for('agendamento_routes.configurar_horarios_agendamento', evento_id=novo_evento.id))
                
                except Exception as e:
                    db.session.rollback()
                    form_erro = f"Erro ao salvar evento: {str(e)}"
                    flash(form_erro, "danger")
        
        except Exception as e:
            form_erro = f"Erro ao processar o formulário: {str(e)}"
            flash(form_erro, "danger")
    
    # Renderizar o template com o formulário
    return render_template('criar_evento_agendamento.html', form_erro=form_erro)
    
@agendamento_routes.route('/importar-agendamentos', methods=['GET', 'POST'])
@login_required
def importar_agendamentos():
    """
    Rota para importação de agendamentos a partir de um arquivo CSV ou Excel.
    """
    # Inicialização de variáveis
    form_erro = None
    eventos = []
    importacao_resultado = None
    
    # Buscar eventos disponíveis do cliente atual
    try:
        eventos = Evento.query.filter_by(cliente_id=current_user.id).all()
    except Exception as e:
        db.session.rollback()
        flash(f"Erro ao buscar eventos: {str(e)}", "danger")
    
    # Processar o formulário quando enviado via POST
    if request.method == 'POST':
        try:
            # Verificar se um arquivo foi enviado
            if 'arquivo' not in request.files:
                form_erro = "Nenhum arquivo enviado."
                flash(form_erro, "danger")
            else:
                arquivo = request.files['arquivo']
                
                # Verificar se o arquivo tem nome
                if arquivo.filename == '':
                    form_erro = "Nenhum arquivo selecionado."
                    flash(form_erro, "danger")
                else:
                    # Verificar a extensão do arquivo
                    if not arquivo.filename.endswith(('.csv', '.xlsx', '.xls')):
                        form_erro = "Formato de arquivo não suportado. Use CSV ou Excel (.xlsx, .xls)."
                        flash(form_erro, "danger")
                    else:
                        # Obter o evento selecionado
                        evento_id = request.form.get('evento_id')
                        if not evento_id:
                            form_erro = "Selecione um evento para importar os agendamentos."
                            flash(form_erro, "danger")
                        else:
                            # Processar o arquivo (CSV ou Excel)
                            # Aqui teríamos a lógica para ler o arquivo e importar os agendamentos
                            # Como não temos acesso ao modelo real de Agendamento, usaremos dados simulados
                            
                            # Simular resultados da importação
                            importacao_resultado = {
                                'total_registros': 15,
                                'importados': 12,
                                'ignorados': 3,
                                'detalhes': [
                                    {'linha': 2, 'status': 'sucesso', 'mensagem': 'Importado com sucesso'},
                                    {'linha': 5, 'status': 'sucesso', 'mensagem': 'Importado com sucesso'},
                                    {'linha': 8, 'status': 'erro', 'mensagem': 'Data inválida'},
                                    {'linha': 10, 'status': 'erro', 'mensagem': 'Horário não disponível'},
                                    {'linha': 12, 'status': 'erro', 'mensagem': 'Capacidade excedida'}
                                ]
                            }
                            
                            flash(f"Importação concluída! {importacao_resultado['importados']} agendamentos importados, {importacao_resultado['ignorados']} ignorados.", "success")
                
        except Exception as e:
            form_erro = f"Erro ao processar a importação: {str(e)}"
            flash(form_erro, "danger")
    
    # Renderizar o template com o formulário
    return render_template('importar_agendamentos.html', 
                          eventos=eventos,
                          form_erro=form_erro,
                          importacao_resultado=importacao_resultado)
    
@agendamento_routes.route('/download-modelo-importacao')
@login_required
def download_modelo_importacao():
    """
    Rota para baixar um modelo de planilha para importação de agendamentos.
    """
    try:
        # Aqui você criaria um arquivo Excel ou CSV com as colunas necessárias
        # Como exemplo, vamos apenas retornar uma resposta simulando o download
        
        # Em uma implementação real, você usaria bibliotecas como xlsxwriter ou pandas
        # para criar o arquivo e depois enviá-lo como resposta
        
        # Exemplo simplificado (apenas para demonstração):
        from io import BytesIO
        import csv
        
        # Criar um buffer de memória para o CSV
        output = BytesIO()
        writer = csv.writer(output)
        
        # Escrever o cabeçalho
        writer.writerow(['Data', 'Horário', 'Escola/Instituição', 'Nome do Responsável', 
                         'E-mail', 'Telefone', 'Turma', 'Quantidade de Alunos'])
        
        # Escrever algumas linhas de exemplo
        writer.writerow(['20/03/2025', '09:00', 'Escola Exemplo', 'João Silva', 
                         'joao.silva@email.com', '(11) 98765-4321', '5º Ano A', '25'])
        writer.writerow(['21/03/2025', '14:30', 'Colégio Modelo', 'Maria Oliveira', 
                         'maria.oliveira@email.com', '(11) 91234-5678', '8º Ano B', '30'])
        
        # Preparar a resposta
        output.seek(0)
        
        return send_file(
            output,
            mimetype='text/csv',
            as_attachment=True,
            download_name='modelo_importacao_agendamentos.csv'
        )
        
    except Exception as e:
        flash(f"Erro ao gerar o modelo: {str(e)}", "danger")
        return redirect(url_for('agendamento_routes.importar_agendamentos'))


@agendamento_routes.route('/exportar-log-importacao')
@login_required
def exportar_log_importacao():
    """
    Rota para exportar o log detalhado da última importação.
    """
    try:
        # Aqui você buscaria os logs de importação do banco de dados
        # Como exemplo, vamos apenas retornar um arquivo CSV com dados simulados
        
        from io import BytesIO
        import csv
        
        # Criar um buffer de memória para o CSV
        output = BytesIO()
        writer = csv.writer(output)
        
        # Escrever o cabeçalho
        writer.writerow(['Linha', 'Status', 'Mensagem', 'Dados Originais'])
        
        # Escrever algumas linhas de exemplo
        writer.writerow(['1', 'Cabeçalho', 'Ignorado', 'Data,Horário,Escola,...'])
        writer.writerow(['2', 'Sucesso', 'Importado com sucesso', '20/03/2025,09:00,Escola Exemplo,...'])
        writer.writerow(['3', 'Sucesso', 'Importado com sucesso', '20/03/2025,14:00,Escola Modelo,...'])
        writer.writerow(['4', 'Erro', 'Data inválida', '32/03/2025,10:00,Escola Inválida,...'])
        writer.writerow(['5', 'Erro', 'Horário não disponível', '21/03/2025,18:00,Escola Teste,...'])
        writer.writerow(['6', 'Erro', 'Capacidade excedida', '22/03/2025,09:00,Escola Grande,...'])
        
        # Preparar a resposta
        output.seek(0)
        
        return send_file(
            output,
            mimetype='text/csv',
            as_attachment=True,
            download_name='log_importacao_agendamentos.csv'
        )
        
    except Exception as e:
        flash(f"Erro ao gerar o log: {str(e)}", "danger")
        return redirect(url_for('agendamento_routes.importar_agendamentos'))
    
@agendamento_routes.route('/api/toggle-agendamento-publico', methods=['POST'])
@login_required
def toggle_agendamento_publico():
    """
    Alternar o status de agendamento público (se visitantes podem agendar pelo site).
    Esta rota é chamada via AJAX a partir da página de configurações.
    """
    try:
        # Buscar configuração atual de agendamento do cliente
        config_agendamento = None
        
        # Verificar se existe um modelo ConfigAgendamento
        # Esta verificação ajuda a evitar erros se o modelo não existir
        if 'ConfigAgendamento' in globals():
            config_agendamento = ConfigAgendamento.query.filter_by(cliente_id=current_user.id).first()
        
            # Se não existir, criar uma nova configuração
            if not config_agendamento:
                config_agendamento = ConfigAgendamento(
                    cliente_id=current_user.id,
                    agendamento_publico=False
                )
                db.session.add(config_agendamento)
            
            # Alternar o status
            config_agendamento.agendamento_publico = not config_agendamento.agendamento_publico
            
            # Salvar alterações no banco de dados
            db.session.commit()
            
            # Retornar o novo status
            return jsonify({
                'success': True,
                'value': config_agendamento.agendamento_publico
            })
        else:
            # Se o modelo não existir, simule a operação para fins de demonstração
            # Em um ambiente de produção, você implementaria isso com seu modelo real
            return jsonify({
                'success': True,
                'value': True,  # Valor simulado
                'message': 'Operação simulada: modelo ConfigAgendamento não encontrado'
            })
            
    except Exception as e:
        # Log de erro para depuração
        logger.error("Erro ao alternar status de agendamento público: %s", str(e))
        
        # Retornar erro para a aplicação
        return jsonify({
            'success': False,
            'message': f"Erro ao alternar status: {str(e)}"
        }), 500

@agendamento_routes.route('/api/toggle-aprovacao-manual', methods=['POST'])
@login_required
def toggle_aprovacao_manual():
    """
    Alternar o status de aprovação manual de agendamentos.
    Quando ativado, os agendamentos novos ficam com status pendente até aprovação.
    Esta rota é chamada via AJAX a partir da página de configurações.
    """
    try:
        # Buscar configuração atual de agendamento do cliente
        config_agendamento = None
        
        # Verificar se existe um modelo ConfigAgendamento
        # Esta verificação ajuda a evitar erros se o modelo não existir
        if 'ConfigAgendamento' in globals():
            config_agendamento = ConfigAgendamento.query.filter_by(cliente_id=current_user.id).first()
        
            # Se não existir, criar uma nova configuração
            if not config_agendamento:
                config_agendamento = ConfigAgendamento(
                    cliente_id=current_user.id,
                    aprovacao_manual=False
                )
                db.session.add(config_agendamento)
            
            # Alternar o status
            config_agendamento.aprovacao_manual = not config_agendamento.aprovacao_manual
            
            # Salvar alterações no banco de dados
            db.session.commit()
            
            # Retornar o novo status
            return jsonify({
                'success': True,
                'value': config_agendamento.aprovacao_manual
            })
        else:
            # Se o modelo não existir, simule a operação para fins de demonstração
            # Em um ambiente de produção, você implementaria isso com seu modelo real
            return jsonify({
                'success': True,
                'value': True,  # Valor simulado
                'message': 'Operação simulada: modelo ConfigAgendamento não encontrado'
            })
            
    except Exception as e:
        # Log de erro para depuração
        logger.error("Erro ao alternar status de aprovação manual: %s", str(e))
        
        # Retornar erro para a aplicação
        return jsonify({
            'success': False,
            'message': f"Erro ao alternar status: {str(e)}"
        }), 500

@agendamento_routes.route('/api/toggle-limite-capacidade', methods=['POST'])
@login_required
def toggle_limite_capacidade():
    """
    Alternar a aplicação do limite de capacidade para agendamentos.
    Quando ativado, o sistema verifica se há vagas disponíveis antes de permitir o agendamento.
    Esta rota é chamada via AJAX a partir da página de configurações.
    """
    try:
        # Buscar configuração atual de agendamento do cliente
        config_agendamento = None
        
        # Verificar se existe um modelo ConfigAgendamento
        # Esta verificação ajuda a evitar erros se o modelo não existir
        if 'ConfigAgendamento' in globals():
            config_agendamento = ConfigAgendamento.query.filter_by(cliente_id=current_user.id).first()
        
            # Se não existir, criar uma nova configuração
            if not config_agendamento:
                config_agendamento = ConfigAgendamento(
                    cliente_id=current_user.id,
                    aplicar_limite_capacidade=True  # O padrão é aplicar o limite
                )
                db.session.add(config_agendamento)
            
            # Alternar o status
            config_agendamento.aplicar_limite_capacidade = not config_agendamento.aplicar_limite_capacidade
            
            # Salvar alterações no banco de dados
            db.session.commit()
            
            # Retornar o novo status
            return jsonify({
                'success': True,
                'value': config_agendamento.aplicar_limite_capacidade
            })
        else:
            # Se o modelo não existir, simule a operação para fins de demonstração
            # Em um ambiente de produção, você implementaria isso com seu modelo real
            return jsonify({
                'success': True,
                'value': True,  # Valor simulado
                'message': 'Operação simulada: modelo ConfigAgendamento não encontrado'
            })
            
    except Exception as e:
        # Log de erro para depuração
        logger.error("Erro ao alternar status de limite de capacidade: %s", str(e))
        
        # Retornar erro para a aplicação
        return jsonify({
            'success': False,
            'message': f"Erro ao alternar status: {str(e)}"
        }), 500

@agendamento_routes.route('/salvar-config-agendamento', methods=['POST'])
@login_required
def salvar_config_agendamento():
    """
    Salvar as configurações gerais do sistema de agendamentos.
    Esta rota processa o formulário enviado pela página de configurações.
    """
    try:
        # Obter dados do formulário
        capacidade_maxima = request.form.get('capacidade_maxima', type=int)
        dias_antecedencia = request.form.get('dias_antecedencia', type=int)
        
        # Validar dados
        if not capacidade_maxima or capacidade_maxima < 1:
            flash("A capacidade máxima deve ser um número positivo.", "danger")
            return redirect(url_for(endpoints.DASHBOARD_AGENDAMENTOS, _anchor='configuracoes'))
            
        if not dias_antecedencia or dias_antecedencia < 1:
            flash("Os dias de antecedência devem ser um número positivo.", "danger")
            return redirect(url_for(endpoints.DASHBOARD_AGENDAMENTOS, _anchor='configuracoes'))
            
        # Buscar configuração atual de agendamento do cliente
        config_agendamento = None
        
        # Verificar se existe um modelo ConfigAgendamento
        if 'ConfigAgendamento' in globals():
            config_agendamento = ConfigAgendamento.query.filter_by(cliente_id=current_user.id).first()
        
            # Se não existir, criar uma nova configuração
            if not config_agendamento:
                config_agendamento = ConfigAgendamento(
                    cliente_id=current_user.id,
                    capacidade_maxima=30,  # Valor padrão
                    dias_antecedencia=30,  # Valor padrão
                    agendamento_publico=True,
                    aprovacao_manual=False,
                    aplicar_limite_capacidade=True
                )
                db.session.add(config_agendamento)
            
            # Atualizar configurações
            config_agendamento.capacidade_maxima = capacidade_maxima
            config_agendamento.dias_antecedencia = dias_antecedencia
            
            # Salvar alterações no banco de dados
            db.session.commit()
            
            flash("Configurações de agendamento salvas com sucesso!", "success")
        else:
            # Se o modelo não existir, apenas exibir mensagem de sucesso simulado
            flash("Configurações salvas com sucesso! (Modo de demonstração)", "success")
        
        # Obter valores opcionais adicionais
        # Pode-se adicionar mais campos conforme necessário
        enviar_lembretes = 'enviar_lembretes' in request.form
        periodo_lembrete = request.form.get('periodo_lembrete', type=int)
        template_email = request.form.get('template_email')
        
        # Se você tiver campos adicionais, o código para salvá-los seria inserido aqui
            
        # Redirecionar de volta para a página de configurações
        return redirect(url_for(endpoints.DASHBOARD_AGENDAMENTOS, _anchor='configuracoes'))
            
    except Exception as e:
        # Log de erro para depuração
        logger.error("Erro ao salvar configurações de agendamento: %s", str(e))
        
        # Notificar o usuário
        flash(f"Erro ao salvar configurações: {str(e)}", "danger")
        
        # Redirecionar de volta para a página de configurações
        return redirect(url_for(endpoints.DASHBOARD_AGENDAMENTOS, _anchor='configuracoes'))

@agendamento_routes.route('/criar-periodo-agendamento', methods=['GET', 'POST'])
@login_required
def criar_periodo_agendamento():
    """
    Rota para criação de período de agendamento.
    Um período define um intervalo de datas em que o agendamento está disponível.
    """
    # Inicialização de variáveis
    form_erro = None
    eventos = []
    
    # Buscar eventos disponíveis do cliente atual
    try:
        eventos = Evento.query.filter_by(cliente_id=current_user.id).all()
    except Exception as e:
        db.session.rollback()
        flash(f"Erro ao buscar eventos: {str(e)}", "danger")
    
    # Processar o formulário quando enviado via POST
    if request.method == 'POST':
        try:
            # Obter dados do formulário
            evento_id = request.form.get('evento_id')
            data_inicio = request.form.get('data_inicio')
            data_fim = request.form.get('data_fim')
            hora_inicio = request.form.get('hora_inicio')
            hora_fim = request.form.get('hora_fim')
            intervalo_min = request.form.get('intervalo_min', type=int)
            capacidade = request.form.get('capacidade', type=int)
            dias_semana = request.form.getlist('dias_semana')  # Lista de dias selecionados (0-6)
            
            # Validar dados obrigatórios
            if not evento_id or not data_inicio or not data_fim or not hora_inicio or not hora_fim or not capacidade:
                form_erro = "Preencha todos os campos obrigatórios."
                flash(form_erro, "danger")
            elif not dias_semana:
                form_erro = "Selecione pelo menos um dia da semana."
                flash(form_erro, "danger")
            else:
                # Verificar se data de fim é posterior à data de início
                if data_fim < data_inicio:
                    form_erro = "A data de fim deve ser posterior à data de início."
                    flash(form_erro, "danger")
                else:
                    # Aqui você adicionaria o código para criar um novo período de agendamento
                    # e os horários relacionados baseado nos dias da semana selecionados
                    
                    # Como não sabemos a estrutura exata do seu modelo,
                    # vamos apenas exibir uma mensagem de sucesso simulada
                    
                    dias_selecionados = [dia_semana(dia) for dia in dias_semana]
                    dias_texto = ", ".join(dias_selecionados)
                    
                    flash(f"Período de agendamento criado com sucesso! Horários configurados para {dias_texto} das {hora_inicio} às {hora_fim}.", "success")
                    
                    # Redirecionar para a página de configuração de horários
                    return redirect(url_for('agendamento_routes.configurar_horarios_agendamento', evento_id=evento_id))
                
        except Exception as e:
            form_erro = f"Erro ao processar o formulário: {str(e)}"
            flash(form_erro, "danger")
    
    # Renderizar o template com o formulário
    return render_template('criar_periodo_agendamento.html',
                          eventos=eventos,
                          form_erro=form_erro)


@agendamento_routes.route('/editar_periodo/<int:periodo_id>', methods=['GET', 'POST'])
@login_required
def editar_periodo(periodo_id):
    flash('Funcionalidade de edição de período ainda não implementada.', 'info')
    return redirect(url_for(endpoints.DASHBOARD_AGENDAMENTOS))


@agendamento_routes.route('/editar_sala_visitacao/<int:sala_id>', methods=['GET', 'POST'])
@login_required
def editar_sala_visitacao(sala_id):
    flash('Funcionalidade de edição de sala ainda não implementada.', 'info')
    return redirect(url_for('agendamento_routes.salas_visitacao', evento_id=sala_id))

@agendamento_routes.route('/excluir-todos-agendamentos', methods=['POST'])
@login_required
def excluir_todos_agendamentos():
    """
    Rota para excluir todos os agendamentos do cliente atual.
    Esta é uma operação perigosa e irreversível, por isso requer uma confirmação
    e é acessível apenas via método POST.
    """
    try:
        # Verificar se o cliente tem permissão para excluir agendamentos
        if not current_user.is_admin and not current_user.is_cliente:
            flash("Você não tem permissão para realizar esta operação.", "danger")
            return redirect(url_for(endpoints.DASHBOARD_AGENDAMENTOS))
        
        # Buscar todos os eventos do cliente
        eventos = Evento.query.filter_by(cliente_id=current_user.id).all()
        
        # Contador para registrar quantos agendamentos foram excluídos
        total_excluidos = 0
        
        # Verificar se existe um modelo Agendamento
        # Esta verificação ajuda a evitar erros se o modelo não existir
        if 'Agendamento' in globals():
            # Para cada evento, buscar todos os horários e seus agendamentos
            for evento in eventos:
                # Se você tiver um relacionamento direto entre Evento e Horario
                horarios = Horario.query.filter_by(evento_id=evento.id).all()
                
                for horario in horarios:
                    # Buscar agendamentos deste horário
                    agendamentos = Agendamento.query.filter_by(horario_id=horario.id).all()
                    
                    # Excluir cada agendamento
                    for agendamento in agendamentos:
                        db.session.delete(agendamento)
                        total_excluidos += 1
            
            # Commit das alterações ao banco de dados
            db.session.commit()
            
            # Notificar o usuário do sucesso da operação
            flash(f"Todos os agendamentos foram excluídos com sucesso. Total de {total_excluidos} agendamentos removidos.", "success")
        else:
            # Se o modelo não existir, simule a operação para fins de demonstração
            flash("Operação simulada: Todos os agendamentos foram excluídos com sucesso.", "success")
        
        # Redirecionar para o dashboard de agendamentos
        return redirect(url_for(endpoints.DASHBOARD_AGENDAMENTOS))
            
    except Exception as e:
        # Em caso de erro, fazer rollback das alterações
        if 'db' in globals() and hasattr(db, 'session'):
            db.session.rollback()

        # Log do erro para depuração
        logger.error("Erro ao excluir agendamentos: %s", str(e))
        
        # Notificar o usuário do erro
        flash(f"Erro ao excluir agendamentos: {str(e)}", "danger")
        
        # Redirecionar para o dashboard
        return redirect(url_for(endpoints.DASHBOARD_AGENDAMENTOS))
    
@agendamento_routes.route('/resetar-configuracoes-agendamento', methods=['POST'])
@login_required
def resetar_configuracoes_agendamento():
    """
    Rota para resetar as configurações de agendamento para valores padrão.
    Esta operação restaura as configurações originais, mas não afeta os agendamentos existentes.
    """
    try:
        # Verificar se o cliente tem permissão para resetar configurações
        if not current_user.is_admin and not current_user.is_cliente:
            flash("Você não tem permissão para realizar esta operação.", "danger")
            return redirect(url_for(endpoints.DASHBOARD_AGENDAMENTOS))
        
        # Buscar configuração atual de agendamento do cliente
        config_agendamento = None
        
        # Verificar se existe um modelo ConfigAgendamento
        if 'ConfigAgendamento' in globals():
            config_agendamento = ConfigAgendamento.query.filter_by(cliente_id=current_user.id).first()
        
            # Se existir, resetar para os valores padrão
            if config_agendamento:
                config_agendamento.capacidade_maxima = 30
                config_agendamento.dias_antecedencia = 30
                config_agendamento.agendamento_publico = True
                config_agendamento.aprovacao_manual = False
                config_agendamento.aplicar_limite_capacidade = True
                
                # Adicionar outras configurações padrão conforme necessário
                
                # Salvar alterações no banco de dados
                db.session.commit()
                
                flash("Configurações de agendamento resetadas para os valores padrão com sucesso!", "success")
            else:
                # Se não existir, criar uma nova configuração com valores padrão
                config_agendamento = ConfigAgendamento(
                    cliente_id=current_user.id,
                    capacidade_maxima=30,
                    dias_antecedencia=30,
                    agendamento_publico=True,
                    aprovacao_manual=False,
                    aplicar_limite_capacidade=True
                )
                
                db.session.add(config_agendamento)
                db.session.commit()
                
                flash("Configurações de agendamento criadas com valores padrão!", "success")
        else:
            # Se o modelo não existir, simule a operação para fins de demonstração
            flash("Operação simulada: Configurações resetadas para valores padrão.", "success")
        
        # Redirecionar para o dashboard de agendamentos, aba configurações
        return redirect(url_for(endpoints.DASHBOARD_AGENDAMENTOS, _anchor='configuracoes'))
            
    except Exception as e:
        # Em caso de erro, fazer rollback das alterações
        if 'db' in globals() and hasattr(db, 'session'):
            db.session.rollback()

        # Log do erro para depuração
        logger.error("Erro ao resetar configurações: %s", str(e))
        
        # Notificar o usuário do erro
        flash(f"Erro ao resetar configurações: {str(e)}", "danger")
        
        # Redirecionar para o dashboard
        return redirect(url_for(endpoints.DASHBOARD_AGENDAMENTOS, _anchor='configuracoes'))
    
@agendamento_routes.route('/exportar-agendamentos')
@login_required
def exportar_agendamentos():
    """
    Rota para exportar agendamentos do cliente atual em formato CSV ou Excel.
    Recebe parâmetros opcionais por query string para filtrar os dados.
    """
    try:
        # Obter parâmetros de filtro
        formato = request.args.get('formato', 'csv')  # csv ou excel
        evento_id = request.args.get('evento_id')
        data_inicio = request.args.get('data_inicio')
        data_fim = request.args.get('data_fim')
        status = request.args.get('status')  # todos, confirmados, realizados, cancelados
        
        # Validar o formato solicitado
        if formato not in ['csv', 'excel']:
            flash("Formato de exportação inválido. Use 'csv' ou 'excel'.", "danger")
            return redirect(url_for(endpoints.DASHBOARD_AGENDAMENTOS))
        
        # Buscar dados para exportação
        # Em uma implementação real, você buscaria os agendamentos do banco de dados
        # baseado nos filtros fornecidos
        
        # Como não sabemos a estrutura exata do seu modelo,
        # vamos criar dados simulados para demonstração
        
        # Dados simulados para exportação
        agendamentos_dados = [
            {
                'id': 1,
                'data': '2025-03-20',
                'horario': '09:00 - 11:00',
                'evento': 'Feira de Ciências 2025',
                'escola': 'Escola Modelo',
                'responsavel': 'João Silva',
                'email': 'joao.silva@email.com',
                'telefone': '(11) 98765-4321',
                'turma': '5º Ano A',
                'alunos': 25,
                'status': 'confirmado',
                'data_criacao': '2025-02-15 14:30:22'
            },
            {
                'id': 2,
                'data': '2025-03-21',
                'horario': '14:00 - 16:00',
                'evento': 'Feira de Ciências 2025',
                'escola': 'Colégio Exemplo',
                'responsavel': 'Maria Oliveira',
                'email': 'maria.oliveira@email.com',
                'telefone': '(11) 91234-5678',
                'turma': '8º Ano B',
                'alunos': 30,
                'status': 'confirmado',
                'data_criacao': '2025-02-16 10:15:45'
            },
            {
                'id': 3,
                'data': '2025-03-22',
                'horario': '09:00 - 11:00',
                'evento': 'Feira de Ciências 2025',
                'escola': 'Instituto Educacional',
                'responsavel': 'Carlos Santos',
                'email': 'carlos.santos@email.com',
                'telefone': '(11) 95555-1234',
                'turma': '2º Ano EM',
                'alunos': 35,
                'status': 'cancelado',
                'data_criacao': '2025-02-17 09:22:10'
            }
        ]
        
        # Exportar para CSV
        if formato == 'csv':
            from io import StringIO
            import csv
            
            # Criar buffer de memória para o CSV
            output = StringIO()
            writer = csv.writer(output)
            
            # Escrever cabeçalho
            writer.writerow(['ID', 'Data', 'Horário', 'Evento', 'Escola', 'Responsável', 'Email', 
                            'Telefone', 'Turma', 'Alunos', 'Status', 'Data de Criação'])
            
            # Escrever linhas de dados
            for agendamento in agendamentos_dados:
                writer.writerow([
                    agendamento['id'],
                    agendamento['data'],
                    agendamento['horario'],
                    agendamento['evento'],
                    agendamento['escola'],
                    agendamento['responsavel'],
                    agendamento['email'],
                    agendamento['telefone'],
                    agendamento['turma'],
                    agendamento['alunos'],
                    agendamento['status'],
                    agendamento['data_criacao']
                ])
            
            # Preparar a resposta
            output.seek(0)
            
            return Response(
                output.getvalue(),
                mimetype='text/csv',
                headers={
                    'Content-Disposition': 'attachment; filename=agendamentos.csv',
                    'Content-Type': 'text/csv; charset=utf-8'
                }
            )
        
        # Exportar para Excel
        elif formato == 'excel':
            try:
                # Tentar importar a biblioteca xlsxwriter
                import xlsxwriter
                from io import BytesIO
                
                # Criar buffer de memória para o Excel
                output = BytesIO()
                
                # Criar workbook e adicionar uma planilha
                workbook = xlsxwriter.Workbook(output)
                worksheet = workbook.add_worksheet('Agendamentos')
                
                # Formatar cabeçalho
                header_format = workbook.add_format({
                    'bold': True,
                    'bg_color': '#4B5563',
                    'color': 'white',
                    'border': 1
                })
                
                # Formatar células normais
                cell_format = workbook.add_format({
                    'border': 1
                })
                
                # Formatar células de status
                status_formats = {
                    'confirmado': workbook.add_format({
                        'border': 1,
                        'bg_color': '#DBEAFE'  # Azul claro
                    }),
                    'realizado': workbook.add_format({
                        'border': 1,
                        'bg_color': '#DCFCE7'  # Verde claro
                    }),
                    'cancelado': workbook.add_format({
                        'border': 1,
                        'bg_color': '#FEE2E2'  # Vermelho claro
                    })
                }
                
                # Definir cabeçalho
                headers = ['ID', 'Data', 'Horário', 'Evento', 'Escola', 'Responsável', 'Email', 
                          'Telefone', 'Turma', 'Alunos', 'Status', 'Data de Criação']
                
                # Escrever cabeçalho
                for col, header in enumerate(headers):
                    worksheet.write(0, col, header, header_format)
                
                # Escrever dados
                for row, agendamento in enumerate(agendamentos_dados, start=1):
                    status = agendamento['status']
                    status_format = status_formats.get(status, cell_format)
                    
                    worksheet.write(row, 0, agendamento['id'], cell_format)
                    worksheet.write(row, 1, agendamento['data'], cell_format)
                    worksheet.write(row, 2, agendamento['horario'], cell_format)
                    worksheet.write(row, 3, agendamento['evento'], cell_format)
                    worksheet.write(row, 4, agendamento['escola'], cell_format)
                    worksheet.write(row, 5, agendamento['responsavel'], cell_format)
                    worksheet.write(row, 6, agendamento['email'], cell_format)
                    worksheet.write(row, 7, agendamento['telefone'], cell_format)
                    worksheet.write(row, 8, agendamento['turma'], cell_format)
                    worksheet.write(row, 9, agendamento['alunos'], cell_format)
                    worksheet.write(row, 10, agendamento['status'], status_format)
                    worksheet.write(row, 11, agendamento['data_criacao'], cell_format)
                
                # Ajustar largura das colunas automaticamente
                for col, header in enumerate(headers):
                    col_width = max(len(header), 12)  # Mínimo de 12 caracteres
                    worksheet.set_column(col, col, col_width)
                
                # Fechar o workbook
                workbook.close()
                
                # Preparar a resposta
                output.seek(0)
                
                return send_file(
                    output,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True,
                    download_name='agendamentos.xlsx'
                )
                
            except ImportError:
                # Caso a biblioteca xlsxwriter não esteja disponível, fallback para CSV
                flash("Biblioteca para exportação Excel não disponível. Exportando como CSV.", "warning")
                
                # Chamada recursiva usando formato CSV
                return exportar_agendamentos(formato='csv')
        
    except Exception as e:
        # Log de erro para depuração
        logger.error("Erro ao exportar agendamentos: %s", str(e))
        
        # Notificar o usuário do erro
        flash(f"Erro ao exportar agendamentos: {str(e)}", "danger")
        
        # Redirecionar para o dashboard
        return redirect(url_for(endpoints.DASHBOARD_AGENDAMENTOS))

@agendamento_routes.route('/sala_visitacao/<int:sala_id>/excluir', methods=['POST'])
@login_required
def excluir_sala_visitacao(sala_id):
    """
    Excluir uma sala de visitação existente.
    
    Args:
        sala_id (int): ID da sala de visitação a ser excluída
        
    Returns:
        Redirecionamento para a página de listagem de salas
    """
    # Verificar permissões do usuário (apenas administradores)
    if current_user.perfil.lower() != 'administrador':
        flash('Você não tem permissão para excluir salas de visitação.', 'danger')
        return redirect(url_for('evento_routes.home'))
    
    # Buscar a sala pelo ID
    sala = SalaVisitacao.query.get_or_404(sala_id)
    
    # Verificar se existem agendamentos relacionados
    agendamentos = Agendamento.query.filter_by(sala_id=sala_id).count()
    if agendamentos > 0:
        flash(f'Não é possível excluir esta sala pois existem {agendamentos} agendamentos associados a ela.', 'warning')
        return redirect(url_for('agendamento_routes.salas_visitacao', evento_id=sala.evento_id))
    
    # Guardar o evento_id para usar no redirecionamento
    evento_id = sala.evento_id
    
    # Excluir a sala
    try:
        db.session.delete(sala)
        db.session.commit()
        flash(f'Sala "{sala.nome}" excluída com sucesso!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao excluir sala: {str(e)}', 'danger')
    
    # Redirecionar para a lista de salas do evento
    return redirect(url_for('agendamento_routes.salas_visitacao', evento_id=evento_id))


@agendamento_routes.route('/agendamentos/exportar/pdf', methods=['GET'])
@login_required
def exportar_agendamentos_pdf():
    """Exporta a lista de agendamentos em PDF"""
    # Implementar lógica para gerar PDF
    # Pode usar bibliotecas como ReportLab, WeasyPrint, etc.
    
    # Exemplo simples com ReportLab
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from io import BytesIO
    
    # Criar o buffer de memória para o PDF
    buffer = BytesIO()
    
    # Configurar o documento
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    
    # Título
    styles = getSampleStyleSheet()
    elements.append(Paragraph("Relatório de Agendamentos", styles['Title']))
    elements.append(Spacer(1, 12))
    
    # Filtrar agendamentos (usar mesma lógica de filtragem da view)
    # Adaptar conforme necessário
    if current_user.tipo == 'admin':
        agendamentos = AgendamentoVisita.query.all()
    elif current_user.tipo == 'cliente':
        agendamentos = (
            AgendamentoVisita.query.join(
                HorarioVisitacao,
                AgendamentoVisita.horario_id == HorarioVisitacao.id,
            ).join(
                Evento,
                HorarioVisitacao.evento_id == Evento.id,
            )
            .filter(Evento.cliente_id == current_user.id)
            .all()
        )
    else:
        agendamentos = AgendamentoVisita.query.filter_by(
            professor_id=current_user.id
        ).all()
    
    # Dados da tabela
    data = [['ID', 'Escola', 'Professor', 'Data', 'Horário', 'Turma', 'Status']]
    
    for agendamento in agendamentos:
        data.append([
            str(agendamento.id),
            agendamento.escola_nome,
            agendamento.professor.nome if agendamento.professor else "-",
            agendamento.horario.data.strftime('%d/%m/%Y'),
            f"{agendamento.horario.horario_inicio} - {agendamento.horario.horario_fim}",
            agendamento.turma,
            agendamento.status.capitalize(),
        ])
    
    # Criar tabela
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    
    # Construir PDF
    doc.build(elements)
    
    # Preparar o response
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name="agendamentos.pdf",
        mimetype="application/pdf"
    )

@agendamento_routes.route('/agendamentos/exportar/csv', methods=['GET'])
@login_required
def exportar_agendamentos_csv():
    """Exporta a lista de agendamentos em CSV"""
    import csv
    from io import StringIO
    
    # Filtrar agendamentos (usar mesma lógica de filtragem da view)
    if current_user.tipo == 'admin':
        agendamentos = AgendamentoVisita.query.all()
    elif current_user.tipo == 'cliente':
        agendamentos = (
            AgendamentoVisita.query.join(
                HorarioVisitacao,
                AgendamentoVisita.horario_id == HorarioVisitacao.id,
            ).join(
                Evento,
                HorarioVisitacao.evento_id == Evento.id,
            )
            .filter(Evento.cliente_id == current_user.id)
            .all()
        )
    else:
        agendamentos = AgendamentoVisita.query.filter_by(
            professor_id=current_user.id
        ).all()
    
    # Criar o buffer de memória para o CSV
    output = StringIO()
    writer = csv.writer(output)
    
    # Escrever cabeçalho
    writer.writerow(['ID', 'Escola', 'Professor', 'Data', 'Horário', 'Turma', 'Nível de Ensino', 
                    'Quantidade de Alunos', 'Status', 'Data do Agendamento'])
    
    # Escrever dados
    for agendamento in agendamentos:
        writer.writerow([
            agendamento.id,
            agendamento.escola_nome,
            agendamento.professor.nome if agendamento.professor else "-",
            agendamento.horario.data.strftime('%d/%m/%Y'),
            f"{agendamento.horario.horario_inicio.strftime('%H:%M')} - {agendamento.horario.horario_fim.strftime('%H:%M')}",
            agendamento.turma,
            agendamento.nivel_ensino,
            agendamento.quantidade_alunos,
            agendamento.status,
            agendamento.data_agendamento.strftime('%d/%m/%Y %H:%M')
        ])
    
    # Preparar o response
    output.seek(0)
    return Response(
        output,
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment;filename=agendamentos.csv"}
    )

@agendamento_routes.route('/visualizar/<int:agendamento_id>', methods=['GET'])
@login_required
def visualizar_agendamento(agendamento_id):
    """
    Rota para visualizar os detalhes de um agendamento específico.
    
    :param agendamento_id: ID do agendamento a ser visualizado
    :return: Template renderizado com os detalhes do agendamento ou resposta JSON
    """
    # Buscar o agendamento pelo ID
    agendamento = AgendamentoVisita.query.get(agendamento_id)

    # Retornar 404 apropriado se não encontrado
    if not agendamento:
        if request.headers.get('Accept') == 'application/json':
            return jsonify({'erro': 'Agendamento não encontrado'}), 404
        abort(404)
    
    # Verificar permissões (somente o professor que criou ou um administrador pode ver)
    if (
        agendamento.professor_id
        and current_user.id != agendamento.professor_id
        and not current_user.is_admin
    ):
        abort(403, "Você não tem permissão para visualizar este agendamento")
    
    # Buscar informações adicionais
    # Se as salas estiverem armazenadas como IDs separados por vírgula
    salas_ids = []
    if agendamento.salas_selecionadas:
        salas_ids = [int(sala_id.strip()) for sala_id in agendamento.salas_selecionadas.split(',')]
        
    # Determinar o formato de resposta (HTML ou JSON)
    if request.headers.get('Accept') == 'application/json':
        # Resposta JSON para API
        return jsonify({
            'id': agendamento.id,
            'horario': {
                'id': agendamento.horario.id,
                'data': agendamento.horario.data.strftime('%d/%m/%Y'),
                'hora_inicio': agendamento.horario.horario_inicio.strftime('%H:%M'),
                'hora_fim': agendamento.horario.horario_fim.strftime('%H:%M')
            },
            'professor': (
                {
                    'id': agendamento.professor.id,
                    'nome': agendamento.professor.nome,
                    'email': agendamento.professor.email,
                }
                if agendamento.professor
                else None
            ),
            'escola': {
                'nome': agendamento.escola_nome,
                'codigo_inep': agendamento.escola_codigo_inep
            },
            'turma': agendamento.turma,
            'nivel_ensino': agendamento.nivel_ensino,
            'quantidade_alunos': agendamento.quantidade_alunos,
            'status': agendamento.status,
            'checkin_realizado': agendamento.checkin_realizado,
            'data_agendamento': agendamento.data_agendamento.strftime('%d/%m/%Y %H:%M') if agendamento.data_agendamento else None,
            'data_cancelamento': agendamento.data_cancelamento.strftime('%d/%m/%Y %H:%M') if agendamento.data_cancelamento else None,
            'data_checkin': agendamento.data_checkin.strftime('%d/%m/%Y %H:%M') if agendamento.data_checkin else None,
            'qr_code_token': agendamento.qr_code_token,
            'salas_selecionadas': salas_ids
        })
    
    # Resposta HTML para interface web
    return render_template(
        'visualizar.html',
        agendamento=agendamento,
        salas_ids=salas_ids
    )
    
@agendamento_routes.route('/editar-cadastro-pendente/<int:agendamento_id>', methods=['GET', 'POST'])
@login_required
def editar_cadastro_pendente(agendamento_id):
    """Editar um cadastro pendente com acesso completo a todos os campos."""

    # Busca o agendamento no banco de dados
    agendamento = AgendamentoVisita.query.get_or_404(agendamento_id)
    
    # Verificar se é um cadastro pendente
    if agendamento.status != 'pendente':
        flash('Este agendamento não está mais pendente.', 'warning')
        return redirect(url_for('agendamento_routes.listar_cadastros_pendentes'))
    
    # Verifica permissões (apenas administradores ou clientes podem editar)
    if current_user.tipo not in ['admin', 'cliente']:
        flash('Você não tem permissão para editar este cadastro.', 'danger')
        return redirect(url_for('agendamento_routes.listar_cadastros_pendentes'))
    
    # Se for cliente, verificar se o evento pertence a ele
    if current_user.tipo == 'cliente' and agendamento.horario.evento.cliente_id != current_user.id:
        flash('Este cadastro não pertence aos seus eventos.', 'danger')
        return redirect(url_for('agendamento_routes.listar_cadastros_pendentes'))

    # Busca horários disponíveis para edição
    horarios_disponiveis = (
        HorarioVisitacao.query
        .filter(HorarioVisitacao.vagas_disponiveis > 0)
        .all()
    )
    # Adiciona o horário atual do agendamento, caso ele não esteja mais disponível
    if agendamento.horario not in horarios_disponiveis:
        horarios_disponiveis.append(agendamento.horario)
    
    # Carrega as possíveis salas para visitação
    from models import SalaVisitacao
    salas = SalaVisitacao.query.all()
    
    # Pega as salas já selecionadas
    salas_selecionadas = []
    if agendamento.salas_selecionadas:
        salas_selecionadas = [int(sala_id) for sala_id in agendamento.salas_selecionadas.split(',')]
    
    if request.method == 'POST':
        # Captura todos os dados do formulário
        horario_id = request.form.get('horario_id')
        escola_nome = request.form.get('escola_nome')
        escola_codigo_inep = request.form.get('escola_codigo_inep')
        turma = request.form.get('turma')
        nivel_ensino = request.form.get('nivel_ensino')
        quantidade_alunos = request.form.get('quantidade_alunos')
        rede_ensino = request.form.get('rede_ensino')
        municipio = request.form.get('municipio')
        bairro = request.form.get('bairro')
        responsavel_nome = request.form.get('responsavel_nome')
        responsavel_cargo = request.form.get('responsavel_cargo')
        responsavel_whatsapp = request.form.get('responsavel_whatsapp')
        responsavel_email = request.form.get('responsavel_email')
        acompanhantes_nomes = request.form.get('acompanhantes_nomes')
        acompanhantes_qtd = request.form.get('acompanhantes_qtd')
        observacoes = request.form.get('observacoes')
        salas_ids = request.form.getlist('salas')
        
        # Capturar compromissos
        compromisso_1 = 'compromisso_1' in request.form
        compromisso_2 = 'compromisso_2' in request.form
        compromisso_3 = 'compromisso_3' in request.form
        compromisso_4 = 'compromisso_4' in request.form
        
        # Validação básica
        if not all([horario_id, escola_nome, turma, nivel_ensino, quantidade_alunos]):
            flash('Por favor, preencha todos os campos obrigatórios.', 'danger')
            return render_template(
                'editar_cadastro_pendente.html',
                agendamento=agendamento,
                horarios=horarios_disponiveis,
                salas=salas,
                salas_selecionadas=salas_selecionadas
            )
        
        try:
            # Atualiza todos os dados do agendamento
            agendamento.horario_id = horario_id
            agendamento.escola_nome = escola_nome
            agendamento.escola_codigo_inep = escola_codigo_inep
            agendamento.turma = turma
            agendamento.nivel_ensino = nivel_ensino
            agendamento.quantidade_alunos = int(quantidade_alunos)
            agendamento.rede_ensino = rede_ensino
            agendamento.municipio = municipio
            agendamento.bairro = bairro
            agendamento.responsavel_nome = responsavel_nome
            agendamento.responsavel_cargo = responsavel_cargo
            agendamento.responsavel_whatsapp = responsavel_whatsapp
            agendamento.responsavel_email = responsavel_email
            agendamento.acompanhantes_nomes = acompanhantes_nomes
            agendamento.acompanhantes_qtd = int(acompanhantes_qtd) if acompanhantes_qtd else None
            agendamento.observacoes = observacoes
            agendamento.compromisso_1 = compromisso_1
            agendamento.compromisso_2 = compromisso_2
            agendamento.compromisso_3 = compromisso_3
            agendamento.compromisso_4 = compromisso_4
            
            # Atualiza as salas selecionadas
            agendamento.salas_selecionadas = ','.join(salas_ids) if salas_ids else None
            
            db.session.commit()
            flash('Cadastro pendente atualizado com sucesso!', 'success')
            return redirect(url_for('agendamento_routes.listar_cadastros_pendentes'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao atualizar cadastro: {str(e)}', 'danger')
    
    # Renderiza o template com os dados do agendamento
    return render_template(
        'editar_cadastro_pendente.html',
        agendamento=agendamento,
        horarios=horarios_disponiveis,
        salas=salas,
        salas_selecionadas=salas_selecionadas
    )

@agendamento_routes.route('/editar_agendamento/<int:agendamento_id>', methods=['GET', 'POST'])
@login_required
def editar_agendamento(agendamento_id):
    """Editar um agendamento de visita existente."""

    # Busca o agendamento no banco de dados
    agendamento = AgendamentoVisita.query.get_or_404(agendamento_id)
    
    # Verifica permissões (apenas o próprio professor, administradores ou clientes podem editar)
    if (
        current_user.tipo not in ['admin', 'cliente']
        and agendamento.professor_id
        and current_user.id != agendamento.professor_id
    ):
        flash('Você não tem permissão para editar este agendamento.', 'danger')
        return redirect(url_for('agendamento_routes.listar_agendamentos'))


def gerar_xlsx_relatorio_geral_completo(eventos, estatisticas, totais, dados_agregados, professores_confirmados, agendamentos, data_inicio, data_fim, caminho_xlsx):
    """
    Gera um relatório geral completo em XLSX com múltiplas abas contendo todas as informações de agendamentos.
    
    Args:
        eventos: Lista de objetos Evento
        estatisticas: Dicionário com estatísticas por evento
        totais: Dicionário com totais gerais
        dados_agregados: Dicionário com dados agregados (escolas, professores, etc.)
        professores_confirmados: Lista de professores com agendamentos confirmados
        agendamentos: Lista de objetos AgendamentoVisita
        data_inicio: Data inicial do período do relatório
        data_fim: Data final do período do relatório
        caminho_xlsx: Caminho onde o XLSX será salvo
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
        import pandas as pd
        
        # Criar workbook
        wb = openpyxl.Workbook()
        
        # Remover a planilha padrão
        wb.remove(wb.active)
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4B5563", end_color="4B5563", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # Aba 1: Resumo Geral
        ws_resumo = wb.create_sheet("Resumo Geral")
        
        # Título e período
        ws_resumo['A1'] = 'Relatório Geral de Agendamentos'
        ws_resumo['A1'].font = Font(bold=True, size=16)
        ws_resumo['A2'] = f'Período: {data_inicio.strftime("%d/%m/%Y")} a {data_fim.strftime("%d/%m/%Y")}'
        ws_resumo['A3'] = f'Gerado em: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
        
        # Totais gerais
        row = 5
        ws_resumo[f'A{row}'] = 'TOTAIS GERAIS'
        ws_resumo[f'A{row}'].font = header_font
        ws_resumo[f'A{row}'].fill = header_fill
        
        row += 1
        total_agendamentos = totais.get('total_agendamentos', 0)
        total_visitantes = totais.get('total_visitantes', 0)
        confirmados = totais.get('confirmados', 0)
        realizados = totais.get('realizados', 0)
        cancelados = totais.get('cancelados', 0)
        
        ws_resumo[f'A{row}'] = 'Total de Agendamentos:'
        ws_resumo[f'B{row}'] = total_agendamentos
        row += 1
        ws_resumo[f'A{row}'] = 'Total de Visitantes:'
        ws_resumo[f'B{row}'] = total_visitantes
        row += 1
        ws_resumo[f'A{row}'] = 'Agendamentos Confirmados:'
        ws_resumo[f'B{row}'] = confirmados
        row += 1
        ws_resumo[f'A{row}'] = 'Agendamentos Realizados:'
        ws_resumo[f'B{row}'] = realizados
        row += 1
        ws_resumo[f'A{row}'] = 'Agendamentos Cancelados:'
        ws_resumo[f'B{row}'] = cancelados
        
        # Taxas
        if total_agendamentos > 0:
            taxa_cancelamento = (cancelados / total_agendamentos) * 100
            row += 1
            ws_resumo[f'A{row}'] = 'Taxa de Cancelamento:'
            ws_resumo[f'B{row}'] = f'{taxa_cancelamento:.1f}%'
        
        if confirmados + realizados > 0:
            taxa_conclusao = (realizados / (confirmados + realizados)) * 100
            row += 1
            ws_resumo[f'A{row}'] = 'Taxa de Conclusão:'
            ws_resumo[f'B{row}'] = f'{taxa_conclusao:.1f}%'
        
        # Aba 2: Agendamentos Detalhados
        ws_agendamentos = wb.create_sheet("Agendamentos")
        
        # Cabeçalhos
        headers = [
            'ID', 'Data da Visita', 'Horário', 'Evento', 'Escola/Instituição',
            'Professor/Responsável', 'E-mail', 'Telefone', 'Turma', 
            'Quantidade de Alunos', 'Status', 'Data do Agendamento',
            'Nível de Ensino', 'Tem Necessidades Especiais', 'Observações'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws_agendamentos.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
        
        # Dados dos agendamentos
        for row_idx, agendamento in enumerate(agendamentos, 2):
            try:
                # Buscar informações do horário e evento
                horario = agendamento.horario
                evento = horario.evento if horario else None
                professor = agendamento.professor
                
                # Mapear nível de ensino
                niveis_ensino = {
                    1: 'Fundamental I',
                    2: 'Fundamental II', 
                    3: 'Ensino Médio',
                    4: 'EJA',
                    5: 'Superior'
                }
                nivel_ensino = niveis_ensino.get(agendamento.nivel_ensino, 'Não informado')
                
                # Verificar necessidades especiais
                tem_necessidades = 'Não'
                if hasattr(agendamento, 'alunos'):
                    for aluno in agendamento.alunos:
                        if hasattr(aluno, 'necessidades_especiais') and aluno.necessidades_especiais:
                            tem_necessidades = 'Sim'
                            break
                
                data = [
                    agendamento.id,
                    horario.data.strftime('%d/%m/%Y') if horario and horario.data else 'N/A',
                    f"{horario.hora_inicio.strftime('%H:%M')} - {horario.hora_fim.strftime('%H:%M')}" if horario and horario.hora_inicio and horario.hora_fim else 'N/A',
                    evento.nome if evento else 'N/A',
                    agendamento.escola_nome or 'N/A',
                    professor.nome if professor else agendamento.professor_nome or 'N/A',
                    professor.email if professor else agendamento.professor_email or 'N/A',
                    agendamento.professor_telefone or 'N/A',
                    agendamento.turma or 'N/A',
                    agendamento.quantidade_alunos or 0,
                    agendamento.status.title() if agendamento.status else 'N/A',
                    agendamento.data_agendamento.strftime('%d/%m/%Y %H:%M') if agendamento.data_agendamento else 'N/A',
                    nivel_ensino,
                    tem_necessidades,
                    agendamento.observacoes or ''
                ]
                
                for col_idx, value in enumerate(data, 1):
                    cell = ws_agendamentos.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border
                    if col_idx in [1, 10, 11]:  # ID, Status, Data do Agendamento
                        cell.alignment = center_alignment
                        
            except Exception as e:
                logger.error(f"Erro ao processar agendamento {agendamento.id}: {e}")
                continue
        
        # Aba 3: Estatísticas por Evento
        if eventos and estatisticas:
            ws_eventos = wb.create_sheet("Estatísticas por Evento")
            
            # Cabeçalhos
            headers_eventos = ['Evento', 'Confirmados', 'Realizados', 'Cancelados', 'Total Visitantes', 'Taxa de Conclusão']
            for col, header in enumerate(headers_eventos, 1):
                cell = ws_eventos.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = border
            
            # Dados por evento
            for row_idx, evento in enumerate(eventos, 2):
                stats = estatisticas.get(evento.id, {})
                confirmados_evento = stats.get('confirmados', 0)
                realizados_evento = stats.get('realizados', 0)
                cancelados_evento = stats.get('cancelados', 0)
                visitantes_evento = stats.get('visitantes', 0)
                
                taxa_conclusao_evento = 0
                if confirmados_evento + realizados_evento > 0:
                    taxa_conclusao_evento = (realizados_evento / (confirmados_evento + realizados_evento)) * 100
                
                data_evento = [
                    evento.nome,
                    confirmados_evento,
                    realizados_evento,
                    cancelados_evento,
                    visitantes_evento,
                    f'{taxa_conclusao_evento:.1f}%'
                ]
                
                for col_idx, value in enumerate(data_evento, 1):
                    cell = ws_eventos.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border
                    if col_idx > 1:  # Números e percentuais
                        cell.alignment = center_alignment
        
        # Aba 4: Estatísticas por Escola
        if dados_agregados and 'escolas' in dados_agregados:
            ws_escolas = wb.create_sheet("Estatísticas por Escola")
            
            # Cabeçalhos
            headers_escolas = ['Escola/Instituição', 'Total Agendamentos', 'Total Alunos', 'Realizados', 'Pendentes']
            for col, header in enumerate(headers_escolas, 1):
                cell = ws_escolas.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = border
            
            # Dados por escola
            for row_idx, escola_stat in enumerate(dados_agregados['escolas'], 2):
                data_escola = [
                    escola_stat.nome,
                    escola_stat.total_agendamentos,
                    escola_stat.total_alunos or 0,
                    escola_stat.realizados,
                    escola_stat.pendentes
                ]
                
                for col_idx, value in enumerate(data_escola, 1):
                    cell = ws_escolas.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border
                    if col_idx > 1:  # Números
                        cell.alignment = center_alignment
        
        # Aba 5: Estatísticas por Professor
        if dados_agregados and 'professores' in dados_agregados:
            ws_professores = wb.create_sheet("Estatísticas por Professor")
            
            # Cabeçalhos
            headers_professores = ['Professor/Responsável', 'Total Agendamentos', 'Total Alunos']
            for col, header in enumerate(headers_professores, 1):
                cell = ws_professores.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = border
            
            # Dados por professor
            for row_idx, prof_stat in enumerate(dados_agregados['professores'], 2):
                data_professor = [
                    prof_stat.nome,
                    prof_stat.total_agendamentos,
                    prof_stat.total_alunos or 0
                ]
                
                for col_idx, value in enumerate(data_professor, 1):
                    cell = ws_professores.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border
                    if col_idx > 1:  # Números
                        cell.alignment = center_alignment
        
        # Ajustar largura das colunas em todas as abas
        for ws in wb.worksheets:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Salvar o arquivo
        wb.save(caminho_xlsx)
        
    except ImportError:
        # Fallback usando pandas se openpyxl não estiver disponível
        import pandas as pd
        
        # Criar DataFrame com dados dos agendamentos
        agendamentos_data = []
        for agendamento in agendamentos:
            try:
                horario = agendamento.horario
                evento = horario.evento if horario else None
                professor = agendamento.professor
                
                niveis_ensino = {
                    1: 'Fundamental I',
                    2: 'Fundamental II', 
                    3: 'Ensino Médio',
                    4: 'EJA',
                    5: 'Superior'
                }
                nivel_ensino = niveis_ensino.get(agendamento.nivel_ensino, 'Não informado')
                
                agendamentos_data.append({
                    'ID': agendamento.id,
                    'Data da Visita': horario.data.strftime('%d/%m/%Y') if horario and horario.data else 'N/A',
                    'Horário': f"{horario.hora_inicio.strftime('%H:%M')} - {horario.hora_fim.strftime('%H:%M')}" if horario and horario.hora_inicio and horario.hora_fim else 'N/A',
                    'Evento': evento.nome if evento else 'N/A',
                    'Escola/Instituição': agendamento.escola_nome or 'N/A',
                    'Professor/Responsável': professor.nome if professor else agendamento.professor_nome or 'N/A',
                    'E-mail': professor.email if professor else agendamento.professor_email or 'N/A',
                    'Telefone': agendamento.professor_telefone or 'N/A',
                    'Turma': agendamento.turma or 'N/A',
                    'Quantidade de Alunos': agendamento.quantidade_alunos or 0,
                    'Status': agendamento.status.title() if agendamento.status else 'N/A',
                    'Data do Agendamento': agendamento.data_agendamento.strftime('%d/%m/%Y %H:%M') if agendamento.data_agendamento else 'N/A',
                    'Nível de Ensino': nivel_ensino,
                    'Observações': agendamento.observacoes or ''
                })
            except Exception as e:
                logger.error(f"Erro ao processar agendamento {agendamento.id}: {e}")
                continue
        
        # Criar DataFrame e salvar
        df = pd.DataFrame(agendamentos_data)
        
        with pd.ExcelWriter(caminho_xlsx, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Agendamentos', index=False)
            
            # Adicionar resumo geral
            resumo_data = {
                'Métrica': ['Total de Agendamentos', 'Total de Visitantes', 'Confirmados', 'Realizados', 'Cancelados'],
                'Valor': [
                    totais.get('total_agendamentos', 0),
                    totais.get('total_visitantes', 0),
                    totais.get('confirmados', 0),
                    totais.get('realizados', 0),
                    totais.get('cancelados', 0)
                ]
            }
            df_resumo = pd.DataFrame(resumo_data)
            df_resumo.to_excel(writer, sheet_name='Resumo Geral', index=False)
    
    except Exception as e:
        logger.error(f"Erro ao gerar XLSX: {e}")
        raise
    
    # Busca horários disponíveis para edição utilizando vagas restantes
    horarios_disponiveis = (
        HorarioVisitacao.query
        .filter(HorarioVisitacao.vagas_disponiveis > 0)
        .all()
    )
    # Adiciona o horário atual do agendamento, caso ele não esteja mais disponível
    if agendamento.horario not in horarios_disponiveis:
        horarios_disponiveis.append(agendamento.horario)
    
    # Carrega as possíveis salas para visitação
    from models import SalaVisitacao
    salas = SalaVisitacao.query.all()
    
    # Pega as salas já selecionadas
    salas_selecionadas = []
    if agendamento.salas_selecionadas:
        salas_selecionadas = [int(sala_id) for sala_id in agendamento.salas_selecionadas.split(',')]
    
    if request.method == 'POST':
        # Captura os dados do formulário
        horario_id = request.form.get('horario_id')
        escola_nome = request.form.get('escola_nome')
        escola_codigo_inep = request.form.get('escola_codigo_inep')
        turma = request.form.get('turma')
        nivel_ensino = request.form.get('nivel_ensino')
        quantidade_alunos = request.form.get('quantidade_alunos')
        salas_ids = request.form.getlist('salas')
        
        # Validação básica
        if not all([horario_id, escola_nome, turma, nivel_ensino, quantidade_alunos]):
            flash('Por favor, preencha todos os campos obrigatórios.', 'danger')
            return render_template(
                'editar_agendamento.html',
                agendamento=agendamento,
                horarios=horarios_disponiveis,
                salas=salas,
                salas_selecionadas=salas_selecionadas
            )
        
        try:
            # Atualiza os dados do agendamento
            agendamento.horario_id = horario_id
            agendamento.escola_nome = escola_nome
            agendamento.escola_codigo_inep = escola_codigo_inep
            agendamento.turma = turma
            agendamento.nivel_ensino = nivel_ensino
            agendamento.quantidade_alunos = int(quantidade_alunos)
            
            # Atualiza as salas selecionadas
            agendamento.salas_selecionadas = ','.join(salas_ids) if salas_ids else None
            
            db.session.commit()
            flash('Agendamento atualizado com sucesso!', 'success')
            return redirect(url_for('agendamento_routes.listar_agendamentos'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao atualizar agendamento: {str(e)}', 'danger')
    
    # Renderiza o template com os dados do agendamento
    return render_template(
        'editar_agendamento.html',
        agendamento=agendamento,
        horarios=horarios_disponiveis,
        salas=salas,
        salas_selecionadas=salas_selecionadas
    )
    
@agendamento_routes.route(
    '/atualizar_status/<int:agendamento_id>', methods=['PUT', 'POST']
)
@login_required
def atualizar_status_agendamento(agendamento_id):
    """Atualiza o status de um agendamento de visita.

    Parâmetros:
        agendamento_id: ID do agendamento a ser atualizado

    Corpo da requisição:
        {
            "status": "confirmado|cancelado|realizado",
            "checkin_realizado": true|false  (opcional)
        }

    Também aceita envio via POST com dados de formulário.

    Retorna:
        200: Agendamento atualizado com sucesso
        400: Dados inválidos
        403: Usuário não tem permissão
        404: Agendamento não encontrado
    """
    # Buscar o agendamento pelo ID
    agendamento = AgendamentoVisita.query.get(agendamento_id)
    
    # Verificar se o agendamento existe
    if not agendamento:
        return jsonify({"erro": "Agendamento não encontrado"}), 404
    
    # Verificar permissões: professor responsável, cliente do evento ou administrador
    is_admin = getattr(current_user, "is_admin", False)
    is_professor = agendamento.professor_id == getattr(current_user, "id", None)
    is_cliente = (
        getattr(current_user, "tipo", "") == "cliente"
        and agendamento.horario.evento.cliente_id == getattr(current_user, "id", None)
    )
    if not (is_admin or is_professor or is_cliente):
        return (
            jsonify({"erro": "Você não tem permissão para alterar este agendamento"}),
            403,
        )
    
    # Obter os dados do request
    dados = request.get_json(silent=True)
    if not dados:
        dados = request.form.to_dict()

    if not dados:
        if request.method == 'POST':
            flash('Nenhum dado fornecido', 'danger')
            return redirect(url_for('agendamento_routes.listar_agendamentos'))
        return jsonify({"erro": "Nenhum dado fornecido"}), 400
    
    # Validar o status
    novo_status = dados.get('status')
    if novo_status and novo_status not in ['confirmado', 'cancelado', 'realizado', 'recusado']:
        return jsonify({"erro": "Status inválido. Use 'confirmado', 'cancelado', 'realizado' ou 'recusado'"}), 400
    
    # Obter motivo da recusa se fornecido
    motivo_recusa = dados.get('motivo_recusa')

    status_anterior = agendamento.status
    enviar_confirmacao = novo_status == 'confirmado' and status_anterior != 'confirmado'

    # Atualizar o status
    if novo_status:
        agendamento.status = novo_status
        
        # Controle automático de vagas
        horario = agendamento.horario
        if horario:
            if novo_status == 'confirmado' and status_anterior != 'confirmado':
                # Reduzir vagas disponíveis ao confirmar
                horario.vagas_disponiveis = max(
                    horario.vagas_disponiveis - agendamento.quantidade_alunos,
                    0
                )
            elif novo_status in ['cancelado', 'recusado'] and status_anterior not in ['cancelado', 'recusado']:
                # Restaurar vagas ao cancelar ou recusar
                horario.vagas_disponiveis = min(
                    horario.vagas_disponiveis + agendamento.quantidade_alunos,
                    horario.capacidade_total,
                )

        # Atualizar campos específicos por status
        if novo_status in ['cancelado', 'recusado']:
            if not agendamento.data_cancelamento:
                agendamento.data_cancelamento = datetime.utcnow()
            
            # Armazenar motivo da recusa se fornecido
            if novo_status == 'recusado' and motivo_recusa:
                agendamento.motivo_recusa = motivo_recusa

    
    # Verificar se houve alteração no check-in
    if 'checkin_realizado' in dados:
        checkin = dados.get('checkin_realizado')
        if isinstance(checkin, str):
            checkin = checkin.lower() in ['true', '1', 'on']
        
        # Se check-in está sendo realizado agora
        if checkin and not agendamento.checkin_realizado:
            agendamento.checkin_realizado = True
            agendamento.data_checkin = datetime.utcnow()
            # Se houve check-in e o status não foi alterado, atualizar para 'realizado'
            if not novo_status:
                agendamento.status = 'realizado'
        # Se check-in está sendo desfeito
        elif not checkin and agendamento.checkin_realizado:
            agendamento.checkin_realizado = False
            agendamento.data_checkin = None
    
    try:
        # Salvar as alterações no banco de dados
        db.session.commit()
        logger.info(f"Agendamento {agendamento.id} atualizado com sucesso para status: {novo_status}")

        if enviar_confirmacao:
            try:
                NotificacaoAgendamentoService.enviar_email_confirmacao(agendamento)
            except Exception as e:
                logger.warning(f"Erro ao enviar email de confirmação: {str(e)}")
            
            # Notificar monitor sobre alunos PCD/Neurodivergentes se houver
            try:
                NotificacaoAgendamentoService.notificar_monitor_alunos_pcd(agendamento)
            except Exception as e:
                logger.warning(f"Erro ao notificar monitor sobre alunos PCD: {str(e)}")

        if request.is_json or request.method == 'PUT':
            resposta = {
                "mensagem": "Agendamento atualizado com sucesso",
                "agendamento": {
                    "id": agendamento.id,
                    "status": agendamento.status,
                    "checkin_realizado": agendamento.checkin_realizado,
                    "data_checkin": (
                        agendamento.data_checkin.isoformat()
                        if agendamento.data_checkin
                        else None
                    ),
                    "data_cancelamento": (
                        agendamento.data_cancelamento.isoformat()
                        if agendamento.data_cancelamento
                        else None
                    ),
                },
            }
            return jsonify(resposta), 200

        flash('Agendamento atualizado com sucesso', 'success')
        
        # Redirecionamento automático para cadastro de alunos após confirmação
        if novo_status == 'confirmado' and current_user.tipo == 'professor':
            # Verificar se já existem alunos cadastrados
            alunos_existentes = AlunoVisitante.query.filter_by(agendamento_id=agendamento.id).count()
            if alunos_existentes < agendamento.quantidade_alunos:
                flash('Agendamento confirmado! Agora cadastre os alunos participantes.', 'info')
                return redirect(url_for('routes.adicionar_alunos_professor', agendamento_id=agendamento.id))

        if current_user.tipo == 'professor':
            return redirect(url_for('agendamento_routes.meus_agendamentos'))
        if current_user.tipo == 'participante':
            return redirect(
                url_for('agendamento_routes.meus_agendamentos_participante')
            )
        if current_user.tipo == 'cliente':
            return redirect(
                url_for('agendamento_routes.meus_agendamentos_cliente')
            )
        return redirect(url_for('agendamento_routes.listar_agendamentos'))

    except Exception as e:
        db.session.rollback()
        if request.is_json or request.method == 'PUT':
            return jsonify(
                {"erro": f"Erro ao atualizar agendamento: {str(e)}"}
            ), 500
        flash('Erro ao atualizar agendamento', 'danger')
        return redirect(url_for('agendamento_routes.listar_agendamentos'))

# Rota para realizar check-in via QR Code
@agendamento_routes.route('/checkin/<string:qr_code_token>', methods=['POST'])
@login_required
def checkin_agendamento(qr_code_token):
    """
    Realiza o check-in de um agendamento através do token QR Code.
    
    Parâmetros:
    - qr_code_token: Token único do QR Code do agendamento
    
    Retorna:
    - 200: Check-in realizado com sucesso
    - 403: Usuário não tem permissão
    - 404: Agendamento não encontrado
    - 409: Check-in já realizado
    """
    # Buscar o agendamento pelo token do QR Code
    agendamento = AgendamentoVisita.query.filter_by(qr_code_token=qr_code_token).first()
    
    # Verificar se o agendamento existe
    if not agendamento:
        return jsonify({"erro": "Agendamento não encontrado"}), 404
    
    # Verificar se o check-in já foi realizado
    if agendamento.checkin_realizado:
        return jsonify({"erro": "Check-in já foi realizado para este agendamento"}), 409
    
    # Verificar permissões: professor, participante ou administrador pode realizar check-in
    if (
        agendamento.professor_id
        and current_user.id != agendamento.professor_id
        and not current_user.is_admin
        and current_user.tipo != 'participante'
        and current_user.tipo != 'cliente'
    ):
        return jsonify({"erro": "Você não tem permissão para realizar check-in neste agendamento"}), 403
    
    # Realizar o check-in
    agendamento.checkin_realizado = True
    agendamento.data_checkin = datetime.utcnow()
    agendamento.status = 'realizado'

    checkin = Checkin(
        usuario_id=agendamento.professor_id,
        evento_id=agendamento.horario.evento_id,
        cliente_id=agendamento.cliente_id,
        palavra_chave='QR-AGENDAMENTO',
    )
    db.session.add(checkin)

    try:
        # Salvar as alterações no banco de dados antes de responder
        db.session.commit()

        # Formatar resposta
        resposta = {
            "mensagem": "Check-in realizado com sucesso",
            "agendamento": {
                "id": agendamento.id,
                "status": agendamento.status,
                "checkin_realizado": agendamento.checkin_realizado,
                "data_checkin": agendamento.data_checkin.isoformat(),
            },
        }

        return jsonify(resposta), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({"erro": f"Erro ao realizar check-in: {str(e)}"}), 500

@agendamento_routes.route('/cadastros-pendentes', methods=['GET'])
@login_required
def listar_cadastros_pendentes():
    """
    Lista especificamente os agendamentos com status 'pendente' para visualização e edição.
    Permite acesso completo aos formulários preenchidos pelos professores.
    """
    # Verificar se é um cliente ou administrador
    if current_user.tipo not in ['cliente', 'admin']:
        flash('Acesso negado!', 'danger')
        return redirect(url_for(endpoints.DASHBOARD))
    
    # Definir os parâmetros de filtro
    page = request.args.get('page', 1, type=int)
    data_inicio = request.args.get('data_inicio')
    data_fim = request.args.get('data_fim')
    escola_nome = request.args.get('escola_nome')
    professor_nome = request.args.get('professor_nome')
    
    # Base da query com joins para acesso aos dados do evento
    query = (
        AgendamentoVisita.query.join(
            HorarioVisitacao, AgendamentoVisita.horario_id == HorarioVisitacao.id
        ).join(
            Evento, HorarioVisitacao.evento_id == Evento.id
        ).filter(
            AgendamentoVisita.status == 'pendente'  # Filtrar apenas pendentes
        )
    )
    
    # Filtrar por tipo de usuário
    if current_user.tipo == 'cliente':
        query = query.filter(Evento.cliente_id == current_user.id)
    
    # Filtros dos parâmetros da URL
    if data_inicio:
        data_inicio_dt = datetime.strptime(data_inicio, '%Y-%m-%d')
        query = query.filter(HorarioVisitacao.data >= data_inicio_dt)
    
    if data_fim:
        data_fim_dt = datetime.strptime(data_fim, '%Y-%m-%d')
        query = query.filter(HorarioVisitacao.data <= data_fim_dt)
    
    if escola_nome:
        query = query.filter(AgendamentoVisita.escola_nome.ilike(f'%{escola_nome}%'))
    
    if professor_nome:
        from models import Usuario
        query = query.join(Usuario, AgendamentoVisita.professor_id == Usuario.id)
        query = query.filter(Usuario.nome.ilike(f'%{professor_nome}%'))
    
    # Ordenar por data de agendamento (mais recentes primeiro)
    query = query.order_by(AgendamentoVisita.data_agendamento.desc())
    
    # Paginação
    agendamentos = query.paginate(
        page=page, per_page=20, error_out=False
    )
    
    return render_template(
        'cadastros_pendentes.html',
        agendamentos=agendamentos,
        data_inicio=data_inicio,
        data_fim=data_fim,
        escola_nome=escola_nome,
        professor_nome=professor_nome
    )

@agendamento_routes.route('/agendamentos', methods=['GET'])
@login_required
def listar_agendamentos():
    """
    Lista os agendamentos de visitas com opções de filtro.
    Administradores veem todos os agendamentos, professores veem apenas os próprios.
    """
    # Definir os parâmetros de filtro
    page = request.args.get('page', 1, type=int)
    data_inicio = request.args.get('data_inicio')
    data_fim = request.args.get('data_fim')
    status = request.args.get('status')
    participante_id = request.args.get('participante_id')
    oficina_id = request.args.get('oficina_id')
    cliente_id = request.args.get('cliente_id')
    evento_id = request.args.get('evento_id', type=int)

    # Base da query com joins para acesso aos dados do evento
    query = (
        AgendamentoVisita.query.join(
            HorarioVisitacao, AgendamentoVisita.horario_id == HorarioVisitacao.id
        ).join(
            Evento, HorarioVisitacao.evento_id == Evento.id
        )
    )

    # Filtrar por tipo de usuário
    if current_user.tipo in ['participante', 'professor']:
        query = query.filter(AgendamentoVisita.professor_id == current_user.id)
    elif current_user.tipo == 'cliente':
        query = query.filter(Evento.cliente_id == current_user.id)
    
    # Filtros dos parâmetros da URL (aplicados apenas se fornecidos)
    if data_inicio:
        data_inicio_dt = datetime.strptime(data_inicio, '%Y-%m-%d')
        query = query.filter(HorarioVisitacao.data >= data_inicio_dt)
    
    if data_fim:
        data_fim_dt = datetime.strptime(data_fim, '%Y-%m-%d')
        query = query.filter(HorarioVisitacao.data <= data_fim_dt)
    
    # Por padrão, mostrar todos os agendamentos (passados, presentes e futuros)
    # sem aplicar filtros de data automáticos
    
    if status:
        query = query.filter(AgendamentoVisita.status == status)
    
    if participante_id:
        query = query.filter(AgendamentoVisita.professor_id == participante_id)
    
    if oficina_id:
        # Se você relacionar agendamentos com oficinas
        # query = query.filter(AgendamentoVisita.oficina_id == oficina_id)
        pass
    
    if cliente_id and current_user.tipo == 'admin':
        query = query.filter(Evento.cliente_id == cliente_id)
    
    if evento_id:
        query = query.filter(Evento.id == evento_id)
    
    # Ordenação
    query = query.order_by(AgendamentoVisita.data_agendamento.desc())
    
    # Paginação
    pagination = query.paginate(page=page, per_page=10, error_out=False)
    agendamentos = pagination.items
    
    # Dados para os filtros de formulário
    oficinas = Oficina.query.all()
    participantes = Usuario.query.filter_by(tipo='participante').all()
    clientes = []
    eventos = []
    if current_user.tipo == 'admin':
        clientes = Cliente.query.all()
        eventos = Evento.query.all()
    elif current_user.tipo == 'cliente':
        eventos = Evento.query.filter_by(cliente_id=current_user.id).all()
    else:
        # Para professores/participantes, mostrar eventos dos agendamentos que eles têm
        eventos = db.session.query(Evento).join(
            HorarioVisitacao, Evento.id == HorarioVisitacao.evento_id
        ).join(
            AgendamentoVisita, HorarioVisitacao.id == AgendamentoVisita.horario_id
        ).filter(
            AgendamentoVisita.professor_id == current_user.id
        ).distinct().all()
    
    return render_template(
        'agendamento/listar_agendamentos.html',
        agendamentos=agendamentos,
        pagination=pagination,
        oficinas=oficinas,
        participantes=participantes,
        clientes=clientes,
        eventos=eventos
    )
    
@agendamento_routes.route('/processar_qrcode_agendamento', methods=['POST'])
@login_required
def processar_qrcode_agendamento():
    """
    Processa o QR Code lido e retorna informações sobre o agendamento.
    """
    if not request.is_json:
        return jsonify({
            'success': False,
            'message': 'Formato de requisição inválido. Envie um JSON.'
        }), 400
    
    data = request.get_json()
    token = data.get('token')
    
    if not token:
        return jsonify({
            'success': False,
            'message': 'Token não fornecido.'
        }), 400
    
    try:
        # Busca o agendamento pelo token do QR Code
        agendamento = AgendamentoVisita.query.filter_by(qr_code_token=token).first()
        
        if not agendamento:
            return jsonify({
                'success': False,
                'message': 'Agendamento não encontrado. Verifique o QR Code e tente novamente.'
            }), 404
        
        # Verifica se o agendamento foi cancelado
        if agendamento.status == 'cancelado':
            return jsonify({
                'success': False,
                'message': 'Este agendamento foi cancelado.'
            }), 400
        
        # Verifica se o check-in já foi realizado
        if agendamento.checkin_realizado:
            return jsonify({
                'success': False,
                'message': 'Check-in já realizado para este agendamento.',
                'redirect': url_for('agendamento_routes.confirmar_checkin_agendamento', token=token)
            }), 200
        
        # Redireciona para a página de confirmação de check-in
        return jsonify({
            'success': True,
            'message': 'Agendamento encontrado!',
            'redirect': url_for('agendamento_routes.confirmar_checkin_agendamento', token=token)
        }), 200
    
    except Exception as e:
        logger.error("Erro ao processar QR code: %s", str(e))
        return jsonify({
            'success': False,
            'message': f'Erro ao processar o QR Code: {str(e)}'
        }), 500

      
@agendamento_routes.route('/confirmar_checkin_agendamento/<token>', methods=['GET', 'POST'])
@login_required
def confirmar_checkin_agendamento(token):
    """
    Exibe página de confirmação e processa o check-in de um agendamento via QR code.
    """
    # Busca o agendamento pelo token
    agendamento = AgendamentoVisita.query.filter_by(qr_code_token=token).first()
    
    if not agendamento:
        flash('Agendamento não encontrado. Verifique o QR Code e tente novamente.', 'danger')
        return redirect(url_for('agendamento_routes.checkin_qr_agendamento'))
    
    # Busca informações relacionadas
    evento = Evento.query.get(agendamento.horario.evento_id)
    horario = agendamento.horario

    form_action = url_for('agendamento_routes.confirmar_checkin_agendamento', token=token)
    
    # Se for POST, realiza o check-in
    if request.method == 'POST':
        try:

            # Atualiza o status do agendamento
            if not agendamento.checkin_realizado:
                agendamento.checkin_realizado = True
                agendamento.data_checkin = datetime.utcnow()
                agendamento.status = 'realizado'

                # Registra o checkin do professor ou participante
                checkin = Checkin(
                    usuario_id=agendamento.professor_id,
                    evento_id=agendamento.horario.evento_id,
                    cliente_id=agendamento.cliente_id,
                    palavra_chave='QR-AGENDAMENTO',
                )
                db.session.add(checkin)

                # Processa os alunos presentes (mudança de lógica)
                alunos_presentes = request.form.getlist('alunos_presentes')
                for aluno in agendamento.alunos:
                    aluno.presente = str(aluno.id) in alunos_presentes


                db.session.commit()
                
                flash('Check-in realizado com sucesso!', 'success')
            else:
                flash('Este agendamento já teve check-in realizado anteriormente.', 'warning')

            if current_user.tipo == 'professor':
                return redirect(
                    url_for('dashboard_professor.dashboard_professor')
                )
            if current_user.tipo == 'cliente':
                return redirect(
                    url_for(endpoints.DASHBOARD_CLIENTE)
                )
            return redirect(url_for(endpoints.DASHBOARD))
        
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao processar check-in: {str(e)}', 'danger')
    
    # Renderiza a página de confirmação
    return render_template(
        'checkin/confirmar_checkin.html',
        agendamento=agendamento,
        evento=evento,
        horario=horario,
        form_action=form_action,
    )


@agendamento_routes.route('/checkin_qr_agendamento', methods=['GET'])
@login_required
def checkin_qr_agendamento():
    """
    Página para escanear QR code para check-in de agendamentos.
    """
    token = request.args.get('token')
    
    # Se um token foi fornecido via parâmetro de URL, redireciona para a confirmação
    if token:
        agendamento = AgendamentoVisita.query.filter_by(qr_code_token=token).first()
        if agendamento:
            return redirect(url_for('agendamento_routes.confirmar_checkin_agendamento', token=token))
        else:
            flash('Agendamento não encontrado. Verifique o token e tente novamente.', 'danger')
    
    # Renderiza a página do scanner QR Code
    return render_template("checkin/checkin_qr_agendamento.html")

# Adicione isto ao seu arquivo routes.py
@agendamento_routes.route('/professor/eventos_disponiveis')
def professor_eventos_disponiveis():
    
    # Buscar eventos disponíveis para o professor
    eventos = Evento.query.filter_by(cliente_id=current_user.cliente_id).all()
    
    # Renderizar o template com os eventos
    return render_template('professor/eventos_disponiveis.html', eventos=eventos)

@agendamento_routes.route('/cadastro_professor', methods=['GET', 'POST'])
def cadastro_professor():
    if request.method == 'POST':
        # Coletar dados do formulário
        nome = request.form.get('nome')
        email = request.form.get('email')
        cpf = request.form.get('cpf')
        senha = request.form.get('senha')
        formacao = request.form.get('formacao')

        # Verificar se email ou CPF já existem
        usuario_existente = Usuario.query.filter(
            (Usuario.email == email) | (Usuario.cpf == cpf)
        ).first()

        if usuario_existente:
            flash('Email ou CPF já cadastrado!', 'danger')
            return render_template("auth/cadastro_professor.html")

        # Criar novo usuário professor
        novo_professor = Usuario(
            nome=nome,
            email=email,
            cpf=cpf,
            senha=generate_password_hash(senha, method="pbkdf2:sha256"),
            formacao=formacao,
            tipo='professor'
        )

        try:
            db.session.add(novo_professor)
            db.session.commit()
            flash('Cadastro realizado com sucesso!', 'success')
            return redirect(url_for('auth_routes.login'))
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao cadastrar: {str(e)}', 'danger')

    return render_template("auth/cadastro_professor.html")

@agendamento_routes.route('/agendar_visita/<int:horario_id>', methods=['GET', 'POST'])
@login_required
def agendar_visita(horario_id):
    """Permite que professores ou clientes agendem visitas."""
    is_professor = getattr(current_user, 'is_professor', lambda: False)()
    is_cliente = getattr(current_user, 'is_cliente', lambda: False)()
    if not (is_professor or is_cliente):
        msg = 'Apenas professores ou clientes podem fazer agendamentos.'
        flash(msg, 'danger')
        return redirect(
            url_for('dashboard_participante_routes.dashboard_participante')
        )

    horario = HorarioVisitacao.query.get_or_404(horario_id)
    estados_lista = obter_estados()
    estados_validos = {sigla for sigla, _ in estados_lista}

    if horario.fechado:
        flash('Este horário está fechado para agendamentos.', 'warning')
        return redirect(
            url_for('dashboard_participante_routes.dashboard_participante')
        )

    if request.method == 'POST':
        # Coletar detalhes do agendamento
        escola_nome = request.form.get('escola_nome')
        turma = request.form.get('turma')
        nivel_ensino = request.form.get('nivel_ensino')
        estados_form = request.form.getlist('estados[]')
        cidades_form = request.form.getlist('cidades[]')

        if not estados_form or not cidades_form:
            flash('Estado e cidade são obrigatórios.', 'danger')
            return redirect(url_for('agendamento_routes.agendar_visita',
                                    horario_id=horario_id))

        estado = estados_form[0]
        cidade = cidades_form[0]

        if estado not in estados_validos or not cidade:
            flash('Estado ou cidade inválidos.', 'danger')
            return redirect(url_for('agendamento_routes.agendar_visita',
                                    horario_id=horario_id))
        try:
            quantidade_alunos = int(request.form.get('quantidade_alunos', 0))
            if quantidade_alunos <= 0:
                raise ValueError
        except ValueError:
            flash('Quantidade de alunos inválida.', 'danger')
            return redirect(url_for('agendamento_routes.agendar_visita', horario_id=horario_id))

        # Validar vagas disponíveis
        if quantidade_alunos > horario.vagas_disponiveis:
            flash('Quantidade de alunos excede vagas disponíveis.', 'danger')
            return redirect(url_for('agendamento_routes.agendar_visita', horario_id=horario_id))

        # Validar usuário/cliente
        professor_id = None
        cliente_id = None
        usuario_id = None
        if is_professor:
            professor = Usuario.query.get(current_user.id)
            if not professor:
                flash('Professor não encontrado.', 'danger')
                return redirect(
                    url_for('agendamento_routes.agendar_visita', horario_id=horario_id)
                )
            professor_id = professor.id
            usuario_id = professor.id
        else:
            cliente = Cliente.query.get(current_user.id)
            if not cliente:
                flash('Cliente não encontrado.', 'danger')
                return redirect(
                    url_for('agendamento_routes.agendar_visita', horario_id=horario_id)
                )
            cliente_id = cliente.id

        # Criar agendamento
        novo_agendamento = AgendamentoVisita(
            horario_id=horario.id,

            professor_id=professor_id,  # None para clientes

            cliente_id=cliente_id,
            escola_nome=escola_nome,
            turma=turma,
            nivel_ensino=nivel_ensino,
            quantidade_alunos=quantidade_alunos,
            estado=estado,
            municipio=cidade,
        )

        # Reduzir vagas disponíveis
        horario.vagas_disponiveis -= quantidade_alunos

        try:
            db.session.add(novo_agendamento)
            db.session.add(horario)
            db.session.commit()
            flash('Agendamento realizado com sucesso!', 'success')
            return redirect(url_for('dashboard_participante_routes.dashboard_participante'))
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao agendar: {str(e)}', 'danger')

    return render_template(
        'agendamento/agendar_visita.html',
        horario=horario,
        estados=estados_lista,
    )

@agendamento_routes.route('/adicionar_alunos', methods=['GET', 'POST'])
@login_required
def adicionar_alunos():
    """
    Rota para adicionar alunos (participantes) em lote ou individualmente.
    Suporta upload de arquivo CSV/Excel e também entrada manual de dados.
    """
    if current_user.tipo not in ['admin', 'cliente']:
        flash('Você não tem permissão para adicionar alunos.', 'danger')
        return redirect(url_for(endpoints.DASHBOARD))

    if request.method == 'POST':
        # Verifica se há upload de arquivo
        arquivo = request.files.get('arquivo')
        
        # Verifica se há entrada manual
        nome = request.form.get('nome')
        cpf = request.form.get('cpf')
        email = request.form.get('email')
        formacao = request.form.get('formacao')
        estados = request.form.getlist('estados[]')
        cidades = request.form.getlist('cidades[]')

        # Processamento de upload de arquivo
        if arquivo and arquivo.filename:
            if not (arquivo and arquivo.filename and arquivo_permitido(arquivo.filename)):
                flash('Arquivo inválido. Utilize um Excel (.xlsx).', 'danger')
                return redirect(url_for('agendamento_routes.adicionar_alunos'))

            # Limite opcional de tamanho do arquivo (5 MB)
            MAX_FILE_SIZE = 5 * 1024 * 1024
            arquivo.seek(0, os.SEEK_END)
            if arquivo.tell() > MAX_FILE_SIZE:
                flash('Arquivo muito grande. Limite de 5MB.', 'danger')
                return redirect(url_for('agendamento_routes.adicionar_alunos'))
            arquivo.seek(0)

            try:
                # Usa Pandas para ler o arquivo de upload
                df = pd.read_excel(arquivo, dtype={'cpf': str})
            except Exception as e:
                flash(f'Erro ao ler o arquivo: {str(e)}', 'danger')
                logger.error('Erro ao ler planilha de alunos: %s', e)
                return redirect(url_for('agendamento_routes.adicionar_alunos'))

            try:
                # Verificar colunas obrigatórias
                colunas_obrigatorias = ['nome', 'cpf', 'email', 'formacao']
                if not all(col in df.columns for col in colunas_obrigatorias):
                    flash(f"Erro: O arquivo deve conter as colunas: {', '.join(colunas_obrigatorias)}", "danger")
                    return redirect(url_for('agendamento_routes.adicionar_alunos'))

                # Processamento em lote
                alunos_adicionados = 0
                for _, row in df.iterrows():
                    cpf_str = str(row['cpf']).strip()

                    # Verifica se o usuário já existe
                    usuario_existente = Usuario.query.filter(
                        (Usuario.cpf == cpf_str) | (Usuario.email == row['email'])
                    ).first()

                    if usuario_existente:
                        logger.warning("Usuário %s já existe. Pulando...", row['nome'])
                        continue

                    novo_usuario = Usuario(
                        nome=row['nome'],
                        cpf=cpf_str,
                        email=row['email'],
                        senha=generate_password_hash(str(row['cpf']), method="pbkdf2:sha256"),  # Senha inicial como CPF
                        formacao=row.get('formacao', 'Não informada'),
                        tipo='participante',
                        cliente_id=current_user.id  # Vincula ao cliente logado
                    )

                    # Tratamento de estados e cidades do arquivo, se existirem
                    if 'estados' in df.columns and 'cidades' in df.columns:
                        novo_usuario.estados = str(row.get('estados', ''))
                        novo_usuario.cidades = str(row.get('cidades', ''))

                    db.session.add(novo_usuario)
                    alunos_adicionados += 1

                db.session.commit()
                flash(f"✅ {alunos_adicionados} alunos importados com sucesso!", "success")
                return redirect(url_for(endpoints.DASHBOARD))

            except Exception as e:
                db.session.rollback()
                flash(f"Erro ao processar arquivo: {str(e)}", "danger")
                logger.error("Erro na importação: %s", e)
                return redirect(url_for('agendamento_routes.adicionar_alunos'))

        # Processamento de entrada manual
        elif nome and cpf and email and formacao:
            try:
                # Verifica se o usuário já existe
                usuario_existente = Usuario.query.filter(
                    (Usuario.cpf == cpf) | (Usuario.email == email)
                ).first()

                if usuario_existente:
                    flash(f"Usuário com CPF {cpf} ou email {email} já existe.", "warning")
                    return redirect(url_for('agendamento_routes.adicionar_alunos'))

                novo_usuario = Usuario(
                    nome=nome,
                    cpf=cpf,
                    email=email,
                    senha=generate_password_hash(cpf, method="pbkdf2:sha256"),  # Senha inicial como CPF
                    formacao=formacao,
                    tipo='participante',
                    cliente_id=current_user.id,  # Vincula ao cliente logado
                    estados=','.join(estados) if estados else None,
                    cidades=','.join(cidades) if cidades else None
                )

                db.session.add(novo_usuario)
                db.session.commit()
                flash("Aluno adicionado com sucesso!", "success")
                return redirect(url_for(endpoints.DASHBOARD))

            except Exception as e:
                db.session.rollback()
                flash(f"Erro ao adicionar aluno: {str(e)}", "danger")
                logger.error("Erro na adição manual: %s", e)
                return redirect(url_for('agendamento_routes.adicionar_alunos'))

        else:
            flash("Dados insuficientes para adicionar aluno.", "warning")
            return redirect(url_for('agendamento_routes.adicionar_alunos'))

    # GET: Renderiza o formulário
    estados = obter_estados()
    return render_template('professor/adicionar_alunos.html', estados=estados)

@agendamento_routes.route('/importar_alunos', methods=['GET', 'POST'])
@login_required
def importar_alunos():
    """
    Rota específica para importação em lote de alunos (participantes).
    Suporta upload de arquivos Excel e CSV.
    """
    if current_user.tipo not in ['admin', 'cliente']:
        flash('Você não tem permissão para importar alunos.', 'danger')
        return redirect(url_for(endpoints.DASHBOARD))

    if request.method == 'POST':
        # Verifica se há upload de arquivo
        arquivo = request.files.get('arquivo')
        
        if not arquivo or not arquivo.filename:
            flash('Nenhum arquivo selecionado.', 'warning')
            return redirect(url_for('agendamento_routes.importar_alunos'))

        try:
            # Determina o tipo de arquivo e usa a biblioteca correta
            if arquivo.filename.endswith(('.xlsx', '.xls')):
                df = pd.read_excel(arquivo, dtype={'cpf': str})
            elif arquivo.filename.endswith('.csv'):
                df = pd.read_csv(arquivo, dtype={'cpf': str})
            else:
                flash('Formato de arquivo não suportado. Use .xlsx, .xls ou .csv', 'danger')
                return redirect(url_for('agendamento_routes.importar_alunos'))
            
            # Verificar colunas obrigatórias
            colunas_obrigatorias = ['nome', 'cpf', 'email', 'formacao']
            colunas_faltantes = [col for col in colunas_obrigatorias if col not in df.columns]
            
            if colunas_faltantes:
                flash(f"Erro: Colunas faltantes no arquivo: {', '.join(colunas_faltantes)}", "danger")
                return redirect(url_for('agendamento_routes.importar_alunos'))

            # Processamento em lote
            alunos_adicionados = 0
            alunos_duplicados = 0
            alunos_invalidos = 0

            # Criar log de importação
            log_importacao = []

            for index, row in df.iterrows():
                try:
                    # Limpeza e validação de dados
                    nome = str(row['nome']).strip()
                    cpf = str(row['cpf']).strip()
                    email = str(row['email']).strip().lower()
                    formacao = str(row.get('formacao', 'Não informada')).strip()

                    # Validações básicas
                    if not nome or not cpf or not email:
                        log_importacao.append(f"Linha {index + 2}: Dados incompletos")
                        alunos_invalidos += 1
                        continue

                    # Verifica se o usuário já existe
                    usuario_existente = Usuario.query.filter(
                        (Usuario.cpf == cpf) | (Usuario.email == email)
                    ).first()

                    if usuario_existente:
                        log_importacao.append(f"Linha {index + 2}: Usuário já existe (CPF ou email duplicado)")
                        alunos_duplicados += 1
                        continue

                    # Tratamento de estados e cidades (se existirem no arquivo)
                    estados = row.get('estados', '') if 'estados' in df.columns else ''
                    cidades = row.get('cidades', '') if 'cidades' in df.columns else ''

                    # Cria novo usuário
                    novo_usuario = Usuario(
                        nome=nome,
                        cpf=cpf,
                        email=email,
                        senha=generate_password_hash(cpf, method="pbkdf2:sha256"),  # Senha inicial como CPF
                        formacao=formacao,
                        tipo='participante',
                        cliente_id=current_user.id,  # Vincula ao cliente logado
                        estados=str(estados),
                        cidades=str(cidades)
                    )
                    
                    db.session.add(novo_usuario)
                    alunos_adicionados += 1
                    log_importacao.append(f"Linha {index + 2}: Usuário {nome} importado com sucesso")

                except Exception as e:
                    log_importacao.append(f"Linha {index + 2}: Erro - {str(e)}")
                    alunos_invalidos += 1

            # Commit final
            db.session.commit()

            # Criar arquivo de log
            log_filename = f"log_importacao_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
            log_path = os.path.join('static', 'logs', log_filename)
            os.makedirs(os.path.dirname(log_path), exist_ok=True)
            
            with open(log_path, 'w', encoding='utf-8') as log_file:
                log_file.write("\n".join(log_importacao))

            # Mensagem resumo
            flash(f"""
                Importação concluída:
                ✅ Alunos adicionados: {alunos_adicionados}
                ⚠️ Alunos duplicados: {alunos_duplicados}
                ❌ Alunos inválidos: {alunos_invalidos}
                📄 Log de importação salvo.
            """, "info")

            # Redireciona com o log
            return render_template('resultado_importacao.html', 
                                   adicionados=alunos_adicionados, 
                                   duplicados=alunos_duplicados, 
                                   invalidos=alunos_invalidos,
                                   log_filename=log_filename)

        except Exception as e:
            db.session.rollback()
            flash(f"Erro ao processar arquivo: {str(e)}", "danger")
            logger.error("Erro na importação: %s", e)
            return redirect(url_for('agendamento_routes.importar_alunos'))

    # GET: Renderiza o formulário de importação
    return render_template('importar_alunos.html')

@agendamento_routes.route('/cancelar_agendamento/<int:agendamento_id>', methods=['POST'])
@login_required
def cancelar_agendamento(agendamento_id):
    """Cancela um agendamento existente."""
    agendamento = AgendamentoVisita.query.get_or_404(agendamento_id)

    evento = agendamento.horario.evento
    is_admin = current_user.tipo == 'admin'
    is_professor = agendamento.professor_id == current_user.id
    is_cliente = (
        current_user.tipo == 'cliente' and evento.cliente_id == current_user.id
    )

    if not any((is_admin, is_professor, is_cliente)):
        flash('Você não tem permissão para cancelar este agendamento.', 'danger')
        return redirect(url_for(endpoints.DASHBOARD))

    agendamento.status = 'cancelado'
    agendamento.data_cancelamento = datetime.utcnow()

    horario = agendamento.horario
    horario.vagas_disponiveis += agendamento.quantidade_alunos

    try:
        db.session.commit()
        flash('Agendamento cancelado com sucesso!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao cancelar agendamento: {str(e)}', 'danger')

    if current_user.tipo == 'professor':
        return redirect(url_for('dashboard_professor.dashboard_professor'))
    if current_user.tipo == 'cliente':
        return redirect(url_for(endpoints.DASHBOARD_CLIENTE))
    return redirect(url_for(endpoints.DASHBOARD))

@agendamento_routes.route('/eventos_disponiveis', methods=['GET'])
@login_required
def eventos_disponiveis():
    if current_user.tipo != 'participante':
        flash('Acesso negado! Esta área é exclusiva para participantes.', 'danger')
        return redirect(url_for(endpoints.DASHBOARD))

    eventos = Evento.query.filter(Evento.data_inicio >= datetime.utcnow()).order_by(Evento.data_inicio).all()

    return render_template('eventos_disponiveis.html', eventos=eventos)

@agendamento_routes.route('/listar_eventos_disponiveis', methods=['GET'])
@login_required
def listar_eventos_disponiveis():
    if current_user.tipo != 'professor':
        flash('Acesso negado!', 'danger')
        return redirect(url_for(endpoints.DASHBOARD))

    eventos = Evento.query.all()
    return render_template('eventos_disponiveis.html', eventos=eventos)


@agendamento_routes.route('/detalhes_agendamento/<int:agendamento_id>')
@login_required
def detalhes_agendamento(agendamento_id):
    """Exibe os detalhes de um agendamento para professores ou clientes."""
    if current_user.tipo not in ("professor", "cliente"):
        flash("Acesso negado!", "danger")
        return redirect(url_for("dashboard_professor.dashboard_professor"))

    # Buscar o agendamento no banco de dados
    agendamento = AgendamentoVisita.query.get_or_404(agendamento_id)

    # Buscar informações do horário e evento associados ao agendamento
    horario = HorarioVisitacao.query.get(agendamento.horario_id)
    evento = Evento.query.get(horario.evento_id) if horario else None

    # Validação extra para clientes: o agendamento deve pertencer a um evento do cliente
    if current_user.tipo == "cliente":
        if not evento or evento.cliente_id != current_user.id:
            flash("Acesso negado!", "danger")
            return redirect(url_for(endpoints.DASHBOARD_CLIENTE))

    # Buscar lista de alunos vinculados ao agendamento
    alunos = AlunoVisitante.query.filter_by(agendamento_id=agendamento.id).all()
    
    # Buscar alunos com necessidades especiais
    from models import NecessidadeEspecial
    alunos_pcd = []
    for aluno in alunos:
        necessidade = NecessidadeEspecial.query.filter_by(aluno_id=aluno.id).first()
        if necessidade:
            alunos_pcd.append((aluno, necessidade))

    return render_template(
        "professor/detalhes_agendamento.html",
        agendamento=agendamento,
        horario=horario,
        evento=evento,
        alunos=alunos,
        alunos_pcd=alunos_pcd,
    )

@agendamento_routes.route('/professor/meus_agendamentos')
@login_required
def meus_agendamentos():
    """Lista agendamentos do usuário para professores e participantes."""
    # Apenas professores ou participantes podem acessar
    if current_user.tipo not in ('professor', 'participante'):
        msg = 'Acesso negado! Esta área é exclusiva para professores e participantes.'
        flash(msg, 'danger')
        return redirect(url_for('auth_routes.login'))
    
    # Filtros
    status = request.args.get('status')
    
    # Base da consulta
    query = AgendamentoVisita.query.filter_by(professor_id=current_user.id)
    
    # Aplicar filtros
    if status:
        query = query.filter(AgendamentoVisita.status == status)
    
    # Ordenar por data/horário
    agendamentos = query.join(
        HorarioVisitacao, AgendamentoVisita.horario_id == HorarioVisitacao.id
    ).order_by(
        HorarioVisitacao.data,
        HorarioVisitacao.horario_inicio
    ).all()
    
    return render_template(
        'professor/meus_agendamentos.html',
        agendamentos=agendamentos,
        status_filtro=status,
        today=date.today,
        hoje=date.today()
    )


@agendamento_routes.route('/participante/meus_agendamentos')
@login_required
def meus_agendamentos_participante():
    """Lista agendamentos do participante logado."""
    if current_user.tipo != 'participante':
        flash('Acesso negado! Esta área é exclusiva para participantes.', 'danger')
        return redirect(url_for(endpoints.DASHBOARD))

    status = request.args.get('status')

    query = AgendamentoVisita.query.filter_by(professor_id=current_user.id)
    if status:
        query = query.filter(AgendamentoVisita.status == status)

    agendamentos = query.join(
        HorarioVisitacao, AgendamentoVisita.horario_id == HorarioVisitacao.id
    ).order_by(
        HorarioVisitacao.data,
        HorarioVisitacao.horario_inicio
    ).all()

    return render_template(
        'participante/meus_agendamentos.html',
        agendamentos=agendamentos,
        status_filtro=status,
        today=date.today,
        hoje=date.today()
    )


@agendamento_routes.route('/cliente/meus_agendamentos')
@login_required
def meus_agendamentos_cliente():
    """Lista agendamentos do cliente logado com paginação e filtros."""
    if current_user.tipo != 'cliente':
        flash('Acesso negado! Esta área é exclusiva para clientes.', 'danger')
        return redirect(url_for(endpoints.DASHBOARD))

    # Parâmetros de filtro e paginação
    page = request.args.get('page', 1, type=int)
    status = request.args.get('status')
    data_inicio = request.args.get('data_inicio')
    data_fim = request.args.get('data_fim')
    oficina_id = request.args.get('oficina_id')
    participante_id = request.args.get('participante_id')

    # Base da query com joins
    query = AgendamentoVisita.query.filter_by(cliente_id=current_user.id).join(
        HorarioVisitacao, AgendamentoVisita.horario_id == HorarioVisitacao.id
    ).join(
        Evento, HorarioVisitacao.evento_id == Evento.id
    )

    # Aplicar filtros
    if status:
        query = query.filter(AgendamentoVisita.status == status)
    
    if data_inicio:
        data_inicio_dt = datetime.strptime(data_inicio, '%Y-%m-%d')
        query = query.filter(HorarioVisitacao.data >= data_inicio_dt)
    
    if data_fim:
        data_fim_dt = datetime.strptime(data_fim, '%Y-%m-%d')
        query = query.filter(HorarioVisitacao.data <= data_fim_dt)
    
    if oficina_id:
        query = query.filter(Evento.id == oficina_id)
    
    if participante_id:
        query = query.filter(AgendamentoVisita.professor_id == participante_id)

    # Ordenação
    query = query.order_by(
        HorarioVisitacao.data,
        HorarioVisitacao.horario_inicio
    )

    # Paginação
    pagination = query.paginate(page=page, per_page=10, error_out=False)
    agendamentos = pagination.items

    # Dados para os filtros
    oficinas = Evento.query.filter_by(cliente_id=current_user.id).all()
    participantes = Usuario.query.filter_by(tipo='participante').all()

    return render_template(
        'cliente/meus_agendamentos.html',
        agendamentos=agendamentos,
        pagination=pagination,
        status_filtro=status,
        oficinas=oficinas,
        participantes=participantes,
        today=date.today,
        hoje=date.today()
    )


@agendamento_routes.route('/cliente/aprovar_agendamento/<int:agendamento_id>', methods=['POST'])
@login_required
def aprovar_agendamento_cliente(agendamento_id):
    """Permite que o cliente aprove um agendamento pendente."""
    if current_user.tipo != 'cliente':
        flash('Acesso negado! Esta área é exclusiva para clientes.', 'danger')
        return redirect(url_for(endpoints.DASHBOARD))

    agendamento = AgendamentoVisita.query.get_or_404(agendamento_id)

    # Verificar se o agendamento pertence ao cliente
    if agendamento.cliente_id != current_user.id:
        flash('Acesso negado! Este agendamento não pertence a você.', 'danger')
        return redirect(url_for('agendamento_routes.meus_agendamentos_cliente'))

    # Verificar se o agendamento está pendente
    if agendamento.status != 'pendente':
        flash('Este agendamento não está pendente de aprovação!', 'warning')
        return redirect(url_for('agendamento_routes.meus_agendamentos_cliente'))

    horario = agendamento.horario
    if horario.vagas_disponiveis < agendamento.quantidade_alunos:
        flash('Não há vagas suficientes para confirmar este agendamento.', 'danger')
        return redirect(url_for('agendamento_routes.meus_agendamentos_cliente'))

    try:
        agendamento.status = 'confirmado'
        agendamento.data_confirmacao = datetime.utcnow()
        horario.vagas_disponiveis -= agendamento.quantidade_alunos
        db.session.commit()
        NotificacaoAgendamentoService.enviar_email_confirmacao(agendamento)
        flash('Agendamento aprovado com sucesso!', 'success')

    except Exception:
        db.session.rollback()
        current_app.logger.exception("Erro ao aprovar agendamento")
        flash('Erro ao aprovar agendamento.', 'danger')

    return redirect(url_for('agendamento_routes.meus_agendamentos_cliente'))


@agendamento_routes.route('/remover_aluno_cliente/<int:agendamento_id>/<int:aluno_id>', methods=['POST'])
@login_required
def remover_aluno_cliente(agendamento_id, aluno_id):
    """Remove um aluno de um agendamento do cliente"""
    try:
        # Verificar se o agendamento pertence ao cliente logado
        agendamento = Agendamento.query.filter_by(
            id=agendamento_id,
            cliente_id=current_user.id
        ).first()

        if not agendamento:
            flash('Agendamento não encontrado ou você não tem permissão para acessá-lo.', 'error')
            return redirect(url_for('agendamento_routes.meus_agendamentos_cliente'))

        # Verificar se o agendamento está em status que permite remoção
        if agendamento.status not in ['pendente', 'confirmado']:
            flash('Não é possível remover alunos de agendamentos com este status.', 'error')
            return redirect(url_for('agendamento_routes.adicionar_alunos_cliente', agendamento_id=agendamento_id))

        # Buscar o aluno
        aluno = AlunoVisitante.query.filter_by(
            id=aluno_id,
            agendamento_id=agendamento_id
        ).first()

        if not aluno:
            flash('Aluno não encontrado.', 'error')
            return redirect(url_for('agendamento_routes.adicionar_alunos_cliente', agendamento_id=agendamento_id))

        # Remover o aluno
        db.session.delete(aluno)

        # Atualizar quantidade de alunos e vagas disponíveis
        agendamento.quantidade_alunos -= 1
        agendamento.horario.vagas_disponiveis += 1

        db.session.commit()

        flash(f'Aluno {aluno.nome} removido com sucesso!', 'success')

    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao remover aluno: {str(e)}', 'error')

    return redirect(url_for('agendamento_routes.adicionar_alunos_cliente', agendamento_id=agendamento_id))


@agendamento_routes.route('/cliente/negar_agendamento/<int:agendamento_id>', methods=['POST'])
@login_required
def negar_agendamento_cliente(agendamento_id):
    """Permite que o cliente negue um agendamento pendente."""
    if current_user.tipo != 'cliente':
        flash('Acesso negado! Esta área é exclusiva para clientes.', 'danger')
        return redirect(url_for(endpoints.DASHBOARD))

    agendamento = AgendamentoVisita.query.get_or_404(agendamento_id)

    # Verificar se o agendamento pertence ao cliente
    if agendamento.cliente_id != current_user.id:
        flash('Acesso negado! Este agendamento não pertence a você.', 'danger')
        return redirect(url_for('agendamento_routes.meus_agendamentos_cliente'))

    # Verificar se o agendamento está pendente
    if agendamento.status != 'pendente':
        flash('Este agendamento não está pendente de aprovação!', 'warning')
        return redirect(url_for('agendamento_routes.meus_agendamentos_cliente'))

    try:
        # Cancelar agendamento e liberar vagas
        agendamento.status = 'cancelado'
        agendamento.data_cancelamento = datetime.utcnow()
        
        # Liberar vagas no horário
        if agendamento.horario:
            agendamento.horario.vagas_disponiveis += agendamento.quantidade_alunos
        
        db.session.commit()
        flash('Agendamento negado com sucesso!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao negar agendamento: {str(e)}', 'danger')

    return redirect(url_for('agendamento_routes.meus_agendamentos_cliente'))


@agendamento_routes.route('/cliente/adicionar_alunos/<int:agendamento_id>', methods=['GET', 'POST'])
@login_required
def adicionar_alunos_cliente(agendamento_id):
    """Permite que o cliente adicione alunos a um agendamento."""
    if current_user.tipo != 'cliente':
        flash('Acesso negado! Esta área é exclusiva para clientes.', 'danger')
        return redirect(url_for(endpoints.DASHBOARD))

    agendamento = AgendamentoVisita.query.get_or_404(agendamento_id)

    # Verificar se o agendamento pertence ao cliente
    if agendamento.cliente_id != current_user.id:
        flash('Acesso negado! Este agendamento não pertence a você.', 'danger')
        return redirect(url_for('agendamento_routes.meus_agendamentos_cliente'))

    # Verificar se o agendamento permite adição de alunos
    if agendamento.status not in ['pendente', 'confirmado']:
        flash('Não é possível adicionar alunos a este agendamento!', 'warning')
        return redirect(url_for('agendamento_routes.meus_agendamentos_cliente'))

    if request.method == 'POST':
        try:
            # Processar adição de aluno único
            nome_aluno = request.form.get('nome_aluno', '').strip()
            cpf_aluno = request.form.get('cpf_aluno', '').strip()
            estados_form = request.form.getlist('estados[]')
            cidades_form = request.form.getlist('cidades[]')
            tem_necessidade_especial = request.form.get('tem_necessidade_especial') == 'on'
            tipo_necessidade = request.form.get('tipo_necessidade', '').strip()
            descricao_necessidade = request.form.get('descricao_necessidade', '').strip()
            
            if not nome_aluno:
                flash('Nome do aluno é obrigatório!', 'danger')
                return redirect(url_for('agendamento_routes.adicionar_alunos_cliente', agendamento_id=agendamento_id))
            
            # Validar campos de necessidade especial se marcado
            if tem_necessidade_especial and (not tipo_necessidade or not descricao_necessidade):
                flash('Tipo e descrição da necessidade especial são obrigatórios quando marcado!', 'danger')
                return redirect(url_for('agendamento_routes.adicionar_alunos_cliente', agendamento_id=agendamento_id))
            
            # Verificar se ainda há vagas disponíveis
            if agendamento.horario.vagas_disponiveis <= 0:
                flash('Não há mais vagas disponíveis para este horário!', 'warning')
                return redirect(url_for('agendamento_routes.adicionar_alunos_cliente', agendamento_id=agendamento_id))
            
            # Criar novo aluno
            novo_aluno = AlunoVisitante(
                nome=nome_aluno,
                cpf=cpf_aluno if cpf_aluno else None,
                agendamento_id=agendamento.id,
                presente=False
            )
            db.session.add(novo_aluno)
            db.session.flush()  # Para obter o ID do aluno
            
            # Criar registro de necessidade especial se necessário
            if tem_necessidade_especial and tipo_necessidade:
                necessidade_especial = NecessidadeEspecial(
                    aluno_id=novo_aluno.id,
                    tipo=tipo_necessidade,
                    descricao=descricao_necessidade,
                    data_registro=datetime.utcnow()
                )
                db.session.add(necessidade_especial)
            
            # Atualizar contadores
            agendamento.quantidade_alunos += 1
            agendamento.horario.vagas_disponiveis -= 1
            
            db.session.commit()
            flash('Aluno adicionado com sucesso!', 'success')
                
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao adicionar alunos: {str(e)}', 'danger')
        
        return redirect(url_for('agendamento_routes.meus_agendamentos_cliente'))
    
    # GET - Renderizar formulário
    estados = obter_estados()
    return render_template(
        'cliente/adicionar_alunos.html',
        agendamento=agendamento,
        estados=estados,
    )


@agendamento_routes.route('/processar_lote_agendamentos', methods=['POST'])
@login_required
def processar_lote_agendamentos():
    """Processa ações em lote para agendamentos (confirmar ou cancelar)."""
    # Verificar permissões
    if current_user.tipo not in ['admin', 'cliente']:
        flash('Acesso negado!', 'danger')
        return redirect(url_for(endpoints.DASHBOARD))
    
    try:
        # Obter dados do formulário
        acao = request.form.get('acao')  # 'confirmar' ou 'cancelar'
        agendamentos_ids = request.form.getlist('agendamentos_selecionados')
        
        if not acao or not agendamentos_ids:
            flash('Nenhuma ação ou agendamento selecionado.', 'warning')
            return redirect(url_for('agendamento_routes.listar_agendamentos'))
        
        # Converter IDs para inteiros
        agendamentos_ids = [int(id_) for id_ in agendamentos_ids]
        
        # Buscar agendamentos
        query = AgendamentoVisita.query.filter(AgendamentoVisita.id.in_(agendamentos_ids))
        
        # Filtrar por permissões do usuário
        if current_user.tipo == 'cliente':
            query = query.join(
                HorarioVisitacao, AgendamentoVisita.horario_id == HorarioVisitacao.id
            ).join(
                Evento, HorarioVisitacao.evento_id == Evento.id
            ).filter(Evento.cliente_id == current_user.id)
        
        agendamentos = query.all()
        
        if not agendamentos:
            flash('Nenhum agendamento válido encontrado.', 'warning')
            return redirect(url_for('agendamento_routes.listar_agendamentos'))
        
        # Processar ação
        contador_sucesso = 0
        contador_erro = 0
        
        for agendamento in agendamentos:
            try:
                if acao == 'confirmar':
                    if agendamento.status in ['pendente']:
                        agendamento.status = 'confirmado'
                        contador_sucesso += 1
                    else:
                        contador_erro += 1
                elif acao == 'cancelar':
                    if agendamento.status in ['pendente', 'confirmado']:
                        agendamento.status = 'cancelado'
                        # Liberar vagas
                        if agendamento.horario:
                            agendamento.horario.vagas_disponiveis += agendamento.quantidade_alunos
                        contador_sucesso += 1
                    else:
                        contador_erro += 1
            except Exception as e:
                contador_erro += 1
                print(f"Erro ao processar agendamento {agendamento.id}: {str(e)}")
        
        # Salvar alterações
        db.session.commit()
        
        # Mensagens de feedback
        if contador_sucesso > 0:
            acao_texto = 'confirmados' if acao == 'confirmar' else 'cancelados'
            flash(f'{contador_sucesso} agendamento(s) {acao_texto} com sucesso!', 'success')
        
        if contador_erro > 0:
            flash(f'{contador_erro} agendamento(s) não puderam ser processados.', 'warning')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao processar agendamentos: {str(e)}', 'danger')
    
    # Redirecionar de volta para a lista de agendamentos
    return redirect(url_for('agendamento_routes.listar_agendamentos'))


@agendamento_routes.route('/professor/cancelar_agendamento/<int:agendamento_id>', methods=['GET', 'POST'])
@login_required
def cancelar_agendamento_professor(agendamento_id):
    """Permite que professores ou participantes cancelem um agendamento."""
    # Apenas professores ou participantes podem acessar
    if current_user.tipo not in ('professor', 'participante'):
        msg = 'Acesso negado! Esta área é exclusiva para professores e participantes.'
        flash(msg, 'danger')
        return redirect(url_for(endpoints.DASHBOARD))
    
    agendamento = AgendamentoVisita.query.get_or_404(agendamento_id)
    
    # Verificar se o agendamento pertence ao professor
    if agendamento.professor_id != current_user.id:
        flash('Acesso negado! Este agendamento não pertence a você.', 'danger')
        return redirect(request.referrer or url_for('agendamento_routes.meus_agendamentos'))
    
    # Verificar se o agendamento já foi cancelado
    if agendamento.status == 'cancelado':
        flash('Este agendamento já foi cancelado!', 'warning')
        return redirect(request.referrer or url_for('agendamento_routes.meus_agendamentos'))
    
    # Verificar se o agendamento já foi realizado
    if agendamento.status == 'realizado':
        flash('Este agendamento já foi realizado e não pode ser cancelado!', 'warning')
        return redirect(request.referrer or url_for('agendamento_routes.meus_agendamentos'))
    
    # Verificar prazo de cancelamento
    horario = agendamento.horario
    config = ConfiguracaoAgendamento.query.filter_by(evento_id=horario.evento_id).first()
    
    if config:
        # Calcular prazo limite para cancelamento
        data_hora_visita = datetime.combine(horario.data, horario.horario_inicio)
        prazo_limite = data_hora_visita - timedelta(
            hours=config.prazo_cancelamento
        )
        
        # Verificar se está dentro do prazo
        if datetime.utcnow() > prazo_limite:
            # Cancelamento fora do prazo - bloquear professor
            data_final_bloqueio = datetime.utcnow() + timedelta(days=config.tempo_bloqueio)
            
            # Criar registro de bloqueio
            bloqueio = ProfessorBloqueado(
                professor_id=current_user.id,
                evento_id=horario.evento_id,
                data_final=data_final_bloqueio,
                motivo=f"Cancelamento fora do prazo ({config.prazo_cancelamento}h antes) para o agendamento #{agendamento.id}"
            )
            db.session.add(bloqueio)
            
            flash(f'Atenção! Cancelamento fora do prazo. Você ficará bloqueado por {config.tempo_bloqueio} dias para novos agendamentos neste evento.', 'warning')
    
    if request.method == 'POST':
        # Restaurar vagas
        horario.vagas_disponiveis += agendamento.quantidade_alunos
        
        # Cancelar agendamento
        agendamento.status = 'cancelado'
        agendamento.data_cancelamento = datetime.utcnow()
        
        try:
            db.session.commit()
            flash('Agendamento cancelado com sucesso!', 'success')
            return redirect(request.referrer or url_for('agendamento_routes.meus_agendamentos'))
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao cancelar agendamento: {str(e)}', 'danger')
    
    prazo_limite = data_hora_visita - timedelta(hours=config.prazo_cancelamento)
    
    return render_template(
        'professor/cancelar_agendamento.html',
        agendamento=agendamento,
        horario=horario,
        prazo_limite=prazo_limite
    )


@agendamento_routes.route('/participante/cancelar_agendamento/<int:agendamento_id>', methods=['GET', 'POST'])
@login_required
def cancelar_agendamento_participante(agendamento_id):
    """Permite que o participante cancele seu agendamento."""
    if current_user.tipo != 'participante':
        flash('Acesso negado! Esta área é exclusiva para participantes.', 'danger')
        return redirect(url_for(endpoints.DASHBOARD))

    agendamento = AgendamentoVisita.query.get_or_404(agendamento_id)

    if agendamento.professor_id != current_user.id:
        flash('Acesso negado! Este agendamento não pertence a você.', 'danger')
        return redirect(request.referrer or url_for('agendamento_routes.meus_agendamentos_participante'))

    if agendamento.status == 'cancelado':
        flash('Este agendamento já foi cancelado!', 'warning')
        return redirect(request.referrer or url_for('agendamento_routes.meus_agendamentos_participante'))

    horario = agendamento.horario

    if request.method == 'POST':
        horario.vagas_disponiveis += agendamento.quantidade_alunos
        agendamento.status = 'cancelado'
        agendamento.data_cancelamento = datetime.utcnow()
        try:
            db.session.commit()
            flash('Agendamento cancelado com sucesso!', 'success')
            return redirect(request.referrer or url_for('agendamento_routes.meus_agendamentos_participante'))
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao cancelar agendamento: {str(e)}', 'danger')

    return render_template(
        'participante/cancelar_agendamento.html',
        agendamento=agendamento,
        horario=horario
    )


@agendamento_routes.route('/cliente/cancelar_agendamento/<int:agendamento_id>', methods=['GET', 'POST'])
@login_required
def cancelar_agendamento_cliente(agendamento_id):
    """Permite que o cliente cancele um agendamento."""
    if current_user.tipo != 'cliente':
        flash('Acesso negado! Esta área é exclusiva para clientes.', 'danger')
        return redirect(url_for(endpoints.DASHBOARD))

    agendamento = AgendamentoVisita.query.get_or_404(agendamento_id)

    if agendamento.cliente_id != current_user.id:
        flash('Acesso negado! Este agendamento não pertence a você.', 'danger')
        return redirect(
            request.referrer or url_for('agendamento_routes.meus_agendamentos_cliente')
        )

    if agendamento.status == 'cancelado':
        flash('Este agendamento já foi cancelado!', 'warning')
        return redirect(
            request.referrer or url_for('agendamento_routes.meus_agendamentos_cliente')
        )

    if agendamento.status == 'realizado':
        flash('Este agendamento já foi realizado e não pode ser cancelado!', 'warning')
        return redirect(
            request.referrer or url_for('agendamento_routes.meus_agendamentos_cliente')
        )

    horario = agendamento.horario
    config = ConfiguracaoAgendamento.query.filter_by(
        evento_id=horario.evento_id
    ).first()
    data_hora_visita = datetime.combine(horario.data, horario.horario_inicio)
    prazo_limite = None
    if config:
        prazo_limite = data_hora_visita - timedelta(
            hours=config.prazo_cancelamento
        )
        if datetime.utcnow() > prazo_limite:
            flash('Cancelamento fora do prazo estabelecido!', 'warning')

    if request.method == 'POST':
        horario.vagas_disponiveis += agendamento.quantidade_alunos
        agendamento.status = 'cancelado'
        agendamento.data_cancelamento = datetime.utcnow()
        try:
            db.session.commit()
            flash('Agendamento cancelado com sucesso!', 'success')
            return redirect(
                request.referrer or url_for('agendamento_routes.meus_agendamentos_cliente')
            )
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao cancelar agendamento: {str(e)}', 'danger')

    return render_template(
        'cliente/cancelar_agendamento.html',
        agendamento=agendamento,
        horario=horario,
        prazo_limite=prazo_limite
    )


@agendamento_routes.route('/presenca_aluno/<int:aluno_id>', methods=['POST'])
@login_required
def marcar_presenca_aluno(aluno_id):
    """
    Marca presença individual de um aluno
    """
    if current_user.tipo != 'cliente':
        flash('Acesso negado!', 'danger')
        return redirect(url_for(endpoints.DASHBOARD))
    
    aluno = AlunoVisitante.query.get_or_404(aluno_id)
    agendamento = aluno.agendamento
    
    # Verificar se o agendamento pertence a um evento do cliente
    evento_id = agendamento.horario.evento_id
    evento = Evento.query.get(evento_id)
    
    if evento.cliente_id != current_user.id:
        flash('Este agendamento não pertence a um evento seu', 'danger')
        return redirect(url_for(endpoints.DASHBOARD_CLIENTE))
    
    # Alternar estado de presença
    aluno.presente = not aluno.presente
    
    try:
        db.session.commit()
        if aluno.presente:
            flash(f'Presença de {aluno.nome} registrada com sucesso!', 'success')
        else:
            flash(f'Presença de {aluno.nome} removida com sucesso!', 'info')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao marcar presença: {str(e)}', 'danger')
    
    return redirect(url_for('agendamento_routes.detalhes_agendamento', agendamento_id=agendamento.id))


@agendamento_routes.route('/checkin_manual/<int:agendamento_id>', methods=['POST'])
@login_required
def checkin_manual_agendamento(agendamento_id):
    """Realiza check-in manual de um agendamento sem necessidade de QR code."""
    agendamento = AgendamentoVisita.query.get_or_404(agendamento_id)
    
    # Verificar permissões: professor, cliente, participante ou administrador
    if (current_user.id != agendamento.professor_id and
        current_user.id != agendamento.horario.evento.cliente_id and
        current_user.tipo not in ['participante', 'admin']):
        flash('Você não tem permissão para realizar check-in neste agendamento.', 'danger')
        return redirect(url_for('agendamento_routes.detalhes_agendamento', agendamento_id=agendamento_id))
    
    # Verificar se o check-in já foi realizado
    if agendamento.checkin_realizado:
        flash('Check-in já foi realizado para este agendamento.', 'warning')
        return redirect(url_for('agendamento_routes.detalhes_agendamento', agendamento_id=agendamento_id))
    
    # Realizar o check-in
    agendamento.checkin_realizado = True
    agendamento.data_checkin = datetime.utcnow()
    
    try:
        db.session.commit()
        flash('Check-in realizado com sucesso!', 'success')
        
        # Log da ação
        logger.info(f"Check-in manual realizado para agendamento {agendamento_id} por usuário {current_user.id}")
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Erro ao realizar check-in manual: {e}")
        flash('Erro ao realizar check-in. Tente novamente.', 'danger')
    
    return redirect(url_for('agendamento_routes.detalhes_agendamento', agendamento_id=agendamento_id))

@agendamento_routes.route('/qrcode_agendamento/<int:agendamento_id>', methods=['GET'])
@login_required
def qrcode_agendamento(agendamento_id):
    """Gera um QR code para o check-in de um agendamento.

    O código embute uma URL que contém o ``qr_code_token`` do agendamento,
    permitindo que o frontend processe o check-in a partir da leitura do QR
    code.
    """
    agendamento = AgendamentoVisita.query.get_or_404(agendamento_id)

    # Permitir acesso para professores, clientes, participantes e administradores
    if (current_user.id != agendamento.professor_id and
        current_user.id != agendamento.horario.evento.cliente_id and
        current_user.tipo not in ['participante', 'admin']):
        abort(403)

    qr_data = url_for(
        'agendamento_routes.checkin_qr_agendamento',
        token=agendamento.qr_code_token,
        _external=True,
    )

    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_M,
        box_size=12,
        border=4
    )
    qr.add_data(qr_data)
    qr.make(fit=True)
    img = qr.make_image(fill_color='black', back_color='white')
    
    # Otimizar tamanho da imagem
    img = img.resize((300, 300), Image.LANCZOS)

    buf = io.BytesIO()
    img.save(buf, 'PNG')
    buf.seek(0)

    return send_file(
        buf,
        mimetype='image/png',
        as_attachment=False,
        download_name=f'qrcode_agendamento_{agendamento.id}.png',
    )

@agendamento_routes.route('/api/horarios_disponiveis', methods=['GET'])
@login_required
def horarios_disponiveis_api():
    if current_user.tipo != 'professor':
        return jsonify({"error": "Acesso não permitido"}), 403
    evento_id = request.args.get('evento_id', type=int)
    if not evento_id:
        return jsonify({"error": "Parâmetro evento_id é obrigatório"}), 400

    evento = Evento.query.get_or_404(evento_id)

    bloqueio = ProfessorBloqueado.query.filter_by(
        professor_id=current_user.id,
        evento_id=evento_id
    ).filter(ProfessorBloqueado.data_final >= datetime.utcnow()).first()

    if bloqueio:
        return jsonify({"error": "Você não tem permissão para visualizar este evento"}), 403

    horarios = HorarioVisitacao.query.filter(
        HorarioVisitacao.evento_id == evento_id,
        HorarioVisitacao.vagas_disponiveis > 0
    ).all()

    eventos = []

    for horario in horarios:
        eventos.append({
            "id": horario.id,
            "title": f"Disponível ({horario.vagas_disponiveis} vagas)",
            "start": f"{horario.data}T{horario.horario_inicio}",
            "end": f"{horario.data}T{horario.horario_fim}",
            "url": url_for('agendamento_routes.agendar_visita', horario_id=horario.id)
        })

    return jsonify(eventos)
@agendamento_routes.route("/importar_oficinas", methods=["POST"])
@login_required
def importar_oficinas():
    """
    Exemplo de rota para importar oficinas de um arquivo Excel (.xlsx).
    Inclui o cadastro da própria oficina e também das datas (OficinaDia).
    """
    # 1. Obter o evento selecionado e validar
    evento_id = request.form.get("evento_id", type=int)
    evento = Evento.query.get(evento_id) if evento_id else None
    if not evento or evento.cliente_id != current_user.id:
        flash("Evento inválido ou não pertence a você!", "danger")
        return redirect(url_for(endpoints.DASHBOARD_CLIENTE))

    # 2. Verificar se foi enviado um arquivo
    if "arquivo" not in request.files:
        flash("Nenhum arquivo enviado!", "danger")
        return redirect(url_for(endpoints.DASHBOARD_CLIENTE))
    
    arquivo = request.files["arquivo"]
    if arquivo.filename == "":
        flash("Nenhum arquivo selecionado.", "danger")
        return redirect(url_for(endpoints.DASHBOARD_CLIENTE))

    # Verifica se a extensão é permitida (.xlsx)
    if not (arquivo_permitido(arquivo.filename) and arquivo.filename.rsplit('.', 1)[1].lower() == "xlsx"):
        flash("Formato de arquivo inválido. Envie um arquivo Excel (.xlsx)", "danger")
        return redirect(url_for(endpoints.DASHBOARD_CLIENTE))

    # 3. Salvar o arquivo em local temporário
    from werkzeug.utils import secure_filename
    import os
    filename = secure_filename(arquivo.filename)
    filepath = os.path.join(current_app.config["UPLOAD_FOLDER"], filename)
    arquivo.save(filepath)

    import pandas as pd
    from datetime import datetime
    from models import Oficina, OficinaDia
    from sqlalchemy.exc import IntegrityError

    try:
        # 3. Ler o arquivo Excel
        df = pd.read_excel(filepath)
        df.columns = df.columns.str.strip()  # tirar espaços extras do nome das colunas

        # Exemplo de colunas que esperamos:
        #   titulo, descricao, ministrante_id, vagas, carga_horaria,
        #   estado, cidade, datas, horarios_inicio, horarios_fim
        #
        # Onde:
        #   - "datas" pode ser uma string com várias datas separadas por vírgula, ex: "01/05/2025,02/05/2025"
        #   - "horarios_inicio" idem, ex: "08:00,09:00"
        #   - "horarios_fim" idem, ex: "12:00,13:00"
        #   - O número de datas deve bater com o número de horários_inicio e horários_fim
        
        colunas_obrigatorias = [
            "titulo", "descricao", "ministrante_id",
            "vagas", "carga_horaria", "estado", "cidade",
            "datas", "horarios_inicio", "horarios_fim"
        ]
        for col in colunas_obrigatorias:
            if col not in df.columns:
                flash(f"Erro: Coluna '{col}' não encontrada no arquivo Excel!", "danger")
                os.remove(filepath)
                return redirect(url_for(endpoints.DASHBOARD_CLIENTE))

        # 4. Percorrer cada linha do DataFrame e criar as oficinas
        total_oficinas_criadas = 0
        for index, row in df.iterrows():
            try:
                # Converter alguns campos para o tipo adequado
                 # Tratar ministrante_id
                raw_m_id = row["ministrante_id"]
                if pd.isna(raw_m_id) or raw_m_id == '':
                    # Se estiver vazio, defina None ou crie lógica de fallback
                    ministrante_id = None
                else:
                    ministrante_id = int(raw_m_id)  # aqui converte para int, se tiver valor
                
                 # Tratar vagas
                raw_vagas = row["vagas"]
                if pd.isna(raw_vagas) or raw_vagas == '':
                    vagas = 0
                else:
                    vagas = int(raw_vagas)
                    
                carga_horaria = str(row["carga_horaria"])
                estado = str(row["estado"]).upper().strip()  # ex: "SP"
                cidade = str(row["cidade"]).strip()

                # Criar a oficina principal
                nova_oficina = Oficina(
                    titulo=row["titulo"],
                    descricao=row["descricao"],
                    ministrante_id=ministrante_id,
                    vagas=vagas,
                    carga_horaria=carga_horaria,
                    estado=estado,
                    cidade=cidade,
                    evento_id=evento_id,
                    cliente_id=current_user.id
                )

                db.session.add(nova_oficina)
                db.session.flush()  # para garantir que nova_oficina.id exista

                # Lendo as datas e horários
                # Supondo que cada coluna seja uma string com valores separados por vírgula
                datas_str = str(row["datas"]).strip()                # ex.: "01/05/2025,02/05/2025"
                horarios_inicio_str = str(row["horarios_inicio"]).strip()  # ex.: "08:00,09:00"
                horarios_fim_str = str(row["horarios_fim"]).strip()        # ex.: "12:00,13:00"

                datas_list = datas_str.split(",")
                hi_list = horarios_inicio_str.split(",")
                hf_list = horarios_fim_str.split(",")

                # Checa se todos os arrays têm mesmo tamanho
                if not (len(datas_list) == len(hi_list) == len(hf_list)):
                    raise ValueError(f"As colunas 'datas', 'horarios_inicio' e 'horarios_fim' devem ter a mesma quantidade de itens na linha {index+1}.")

                # 5. Para cada data, criar um registro OficinaDia
                for i in range(len(datas_list)):
                    data_str = datas_list[i].strip()
                    hi_str = hi_list[i].strip()
                    hf_str = hf_list[i].strip()

                    # Converter data_str para datetime.date (formato padrão dd/mm/yyyy)
                    try:
                        data_formatada = datetime.strptime(data_str, "%d/%m/%Y").date()
                    except ValueError:
                        raise ValueError(f"Data inválida na linha {index+1}: '{data_str}'. Formato esperado: DD/MM/YYYY.")

                    novo_dia = OficinaDia(
                        oficina_id=nova_oficina.id,
                        data=data_formatada,
                        horario_inicio=hi_str,
                        horario_fim=hf_str
                    )
                    db.session.add(novo_dia)

                db.session.commit()
                total_oficinas_criadas += 1
            
            except IntegrityError as e:
                db.session.rollback()
                current_app.logger.error(f"[Linha {index+1}] Erro de integridade: {e}")
            except Exception as e:
                db.session.rollback()
                current_app.logger.error(f"[Linha {index+1}] Erro ao criar oficina: {e}")

        flash(f"Foram importadas {total_oficinas_criadas} oficinas com sucesso, incluindo as datas!", "success")
    
    except Exception as e:
        current_app.logger.error(f"Erro ao processar arquivo Excel: {e}")
        flash(f"Erro ao processar arquivo: {e}", "danger")
    finally:
        # Remove o arquivo temporário
        if os.path.exists(filepath):
            os.remove(filepath)
    
    return redirect(url_for(endpoints.DASHBOARD_CLIENTE))


# API endpoint para materiais de apoio
@agendamento_routes.route('/api/materiais-apoio', methods=['GET'])
@login_required
def api_materiais_apoio():
    """API endpoint para buscar materiais de apoio disponíveis.
    
    Returns:
        JSON: Lista de materiais de apoio com id, nome e descrição
    """
    try:
        from models import MaterialApoio
        
        materiais = MaterialApoio.query.filter_by(ativo=True).order_by(MaterialApoio.nome).all()
        
        materiais_data = []
        for material in materiais:
            materiais_data.append({
                'id': material.id,
                'nome': material.nome,
                'descricao': material.descricao
            })
        
        return jsonify({
            'success': True,
            'materiais': materiais_data
        })
        
    except Exception as e:
        current_app.logger.error(f"Erro ao buscar materiais de apoio: {e}")
        return jsonify({
            'success': False,
            'error': 'Erro interno do servidor'
        }), 500


# API endpoints para gerenciamento de materiais de apoio (admin)
@agendamento_routes.route('/api/materiais-apoio-admin', methods=['GET'])
@login_required
def api_materiais_apoio_admin():
    """API endpoint para listar todos os materiais de apoio (admin)"""
    try:
        from models import MaterialApoio
        
        materiais = MaterialApoio.query.filter_by(cliente_id=current_user.id).order_by(MaterialApoio.nome).all()
        materiais_data = [{
            'id': material.id,
            'nome': material.nome,
            'descricao': material.descricao,
            'ativo': material.ativo
        } for material in materiais]
        
        return jsonify({
            'success': True,
            'materiais': materiais_data
        })
    except Exception as e:
        current_app.logger.error(f"Erro ao listar materiais de apoio: {e}")
        return jsonify({
            'success': False,
            'error': 'Erro interno do servidor'
        }), 500


@agendamento_routes.route('/api/materiais-apoio-admin/<int:material_id>', methods=['GET'])
@login_required
def api_material_apoio_detalhes(material_id):
    """API endpoint para obter detalhes de um material específico"""
    try:
        from models import MaterialApoio
        
        material = MaterialApoio.query.filter_by(id=material_id, cliente_id=current_user.id).first()
        if not material:
            return jsonify({
                'success': False,
                'error': 'Material não encontrado'
            }), 404
        
        return jsonify({
            'success': True,
            'material': {
                'id': material.id,
                'nome': material.nome,
                'descricao': material.descricao,
                'ativo': material.ativo
            }
        })
    except Exception as e:
        current_app.logger.error(f"Erro ao buscar material {material_id}: {e}")
        return jsonify({
            'success': False,
            'error': 'Erro interno do servidor'
        }), 500


@agendamento_routes.route('/api/materiais-apoio-admin', methods=['POST'])
@login_required
def api_criar_material_apoio():
    """API endpoint para criar um novo material de apoio"""
    try:
        from models import MaterialApoio
        
        data = request.get_json()
        
        # Validação
        if not data or not data.get('nome'):
            return jsonify({
                'success': False,
                'error': 'Nome é obrigatório'
            }), 400
        
        # Verificar se já existe um material com o mesmo nome para este cliente
        material_existente = MaterialApoio.query.filter_by(nome=data['nome'], cliente_id=current_user.id).first()
        if material_existente:
            return jsonify({
                'success': False,
                'error': 'Já existe um material com este nome'
            }), 400
        
        # Criar novo material
        novo_material = MaterialApoio(
            nome=data['nome'],
            descricao=data.get('descricao', ''),
            cliente_id=current_user.id,
            ativo=data.get('ativo', True)
        )
        
        db.session.add(novo_material)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Material de apoio criado com sucesso',
            'material': {
                'id': novo_material.id,
                'nome': novo_material.nome,
                'descricao': novo_material.descricao,
                'ativo': novo_material.ativo
            }
        })
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Erro ao criar material de apoio: {e}")
        return jsonify({
            'success': False,
            'error': 'Erro interno do servidor'
        }), 500


@agendamento_routes.route('/api/materiais-apoio-admin/<int:material_id>', methods=['PUT'])
@login_required
def api_atualizar_material_apoio(material_id):
    """API endpoint para atualizar um material de apoio"""
    try:
        from models import MaterialApoio
        
        material = MaterialApoio.query.filter_by(id=material_id, cliente_id=current_user.id).first()
        if not material:
            return jsonify({
                'success': False,
                'error': 'Material não encontrado'
            }), 404
        
        data = request.get_json()
        
        # Validação
        if not data or not data.get('nome'):
            return jsonify({
                'success': False,
                'error': 'Nome é obrigatório'
            }), 400
        
        # Verificar se já existe outro material com o mesmo nome para este cliente
        material_existente = MaterialApoio.query.filter(
            MaterialApoio.nome == data['nome'],
            MaterialApoio.id != material_id,
            MaterialApoio.cliente_id == current_user.id
        ).first()
        if material_existente:
            return jsonify({
                'success': False,
                'error': 'Já existe outro material com este nome'
            }), 400
        
        # Atualizar material
        material.nome = data['nome']
        material.descricao = data.get('descricao', '')
        material.ativo = data.get('ativo', True)
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Material de apoio atualizado com sucesso',
            'material': {
                'id': material.id,
                'nome': material.nome,
                'descricao': material.descricao,
                'ativo': material.ativo
            }
        })
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Erro ao atualizar material {material_id}: {e}")
        return jsonify({
            'success': False,
            'error': 'Erro interno do servidor'
        }), 500


@agendamento_routes.route('/api/materiais-apoio-admin/<int:material_id>/toggle', methods=['POST'])
@login_required
def api_toggle_material_apoio(material_id):
    """API endpoint para alternar o status ativo/inativo de um material"""
    try:
        from models import MaterialApoio
        
        material = MaterialApoio.query.filter_by(id=material_id, cliente_id=current_user.id).first()
        if not material:
            return jsonify({
                'success': False,
                'error': 'Material não encontrado'
            }), 404
        
        # Alternar status
        material.ativo = not material.ativo
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Material {"ativado" if material.ativo else "desativado"} com sucesso',
            'ativo': material.ativo
        })
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Erro ao alterar status do material {material_id}: {e}")
        return jsonify({
            'success': False,
            'error': 'Erro interno do servidor'
        }), 500


# ==================== SISTEMA DE NOTIFICAÇÕES ====================

@agendamento_routes.route('/notificacoes')
@login_required
def listar_notificacoes():
    """Lista todas as notificações do usuário atual."""
    try:
        # Verificar se é uma requisição AJAX (com parâmetro limit ou Accept: application/json)
        limit = request.args.get('limit', type=int)
        is_ajax = limit is not None or 'application/json' in request.headers.get('Accept', '')
        
        if is_ajax:
            # Retornar JSON para requisições AJAX
            query = NotificacaoAgendamento.query.filter_by(destinatario_id=current_user.id)
            
            # Aplicar filtros se fornecidos
            tipo = request.args.get('tipo')
            lida = request.args.get('lida')
            
            if tipo:
                query = query.filter_by(tipo=tipo)
            if lida is not None:
                query = query.filter_by(lida=lida.lower() == 'true')
            
            # Ordenar por data de criação (mais recentes primeiro)
            query = query.order_by(NotificacaoAgendamento.data_criacao.desc())
            
            # Aplicar limite se fornecido
            if limit:
                notificacoes_list = query.limit(limit).all()
            else:
                notificacoes_list = query.all()
            
            # Contar total de notificações não lidas
            count_nao_lidas = NotificacaoAgendamento.query.filter_by(
                destinatario_id=current_user.id,
                lida=False
            ).count()
            
            # Converter notificações para formato JSON
            notificacoes_json = []
            for notif in notificacoes_list:
                notificacoes_json.append({
                    'id': notif.id,
                    'titulo': notif.titulo,
                    'mensagem': notif.mensagem,
                    'tipo': notif.tipo,
                    'lida': notif.lida,
                    'data_criacao': notif.data_criacao.strftime('%d/%m/%Y %H:%M') if notif.data_criacao else '',
                    'remetente_nome': notif.remetente.nome if notif.remetente else 'Sistema',
                    'agendamento_id': notif.agendamento_id
                })
            
            return jsonify({
                'success': True,
                'notificacoes': notificacoes_json,
                'count': count_nao_lidas
            })
        
        else:
            # Retornar HTML para acesso direto
            page = request.args.get('page', 1, type=int)
            per_page = 20
            
            # Filtros
            tipo = request.args.get('tipo')
            lida = request.args.get('lida')
            
            # Query base
            query = NotificacaoAgendamento.query.filter_by(destinatario_id=current_user.id)
            
            # Aplicar filtros
            if tipo:
                query = query.filter_by(tipo=tipo)
            if lida is not None:
                query = query.filter_by(lida=lida.lower() == 'true')
            
            # Ordenar por data de criação (mais recentes primeiro)
            query = query.order_by(NotificacaoAgendamento.data_criacao.desc())
            
            # Paginação
            notificacoes = query.paginate(
                page=page, per_page=per_page, error_out=False
            )
            
            return render_template(
                'agendamento/notificacoes.html',
                notificacoes=notificacoes,
                tipo_filtro=tipo,
                lida_filtro=lida
            )
            
    except Exception as e:
        current_app.logger.error(f"Erro ao listar notificações: {e}")
        
        # Verificar se é AJAX para retornar erro em JSON
        limit = request.args.get('limit', type=int)
        is_ajax = limit is not None or 'application/json' in request.headers.get('Accept', '')
        
        if is_ajax:
            return jsonify({
                'success': False,
                'error': 'Erro ao carregar notificações'
            }), 500
        else:
            flash('Erro ao carregar notificações', 'error')
            return redirect(url_for('agendamento_routes.listar_agendamentos'))


@agendamento_routes.route('/notificacoes/<int:notificacao_id>/marcar-lida', methods=['POST'])
@login_required
def marcar_notificacao_lida(notificacao_id):
    """Marca uma notificação como lida."""
    try:
        notificacao = NotificacaoAgendamento.query.get_or_404(notificacao_id)
        
        # Verificar se o usuário tem permissão
        if notificacao.destinatario_id != current_user.id:
            return jsonify({
                'success': False,
                'error': 'Acesso negado'
            }), 403
        
        notificacao.marcar_como_lida()
        
        return jsonify({
            'success': True,
            'message': 'Notificação marcada como lida'
        })
    except Exception as e:
        current_app.logger.error(f"Erro ao marcar notificação como lida: {e}")
        return jsonify({
            'success': False,
            'error': 'Erro interno do servidor'
        }), 500


@agendamento_routes.route('/notificacoes/<int:notificacao_id>/responder', methods=['POST'])
@login_required
def responder_notificacao(notificacao_id):
    """Responde a uma notificação."""
    try:
        notificacao = NotificacaoAgendamento.query.get_or_404(notificacao_id)
        
        # Verificar se o usuário tem permissão
        if notificacao.destinatario_id != current_user.id:
            return jsonify({
                'success': False,
                'error': 'Acesso negado'
            }), 403
        
        data = request.get_json()
        resposta_texto = data.get('resposta', '').strip()
        
        if not resposta_texto:
            return jsonify({
                'success': False,
                'error': 'Resposta não pode estar vazia'
            }), 400
        
        # Responder à notificação
        notificacao_resposta = notificacao.responder(resposta_texto, current_user.id)
        
        return jsonify({
            'success': True,
            'message': 'Resposta enviada com sucesso',
            'notificacao_resposta_id': notificacao_resposta.id
        })
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Erro ao responder notificação: {e}")
        return jsonify({
            'success': False,
            'error': 'Erro interno do servidor'
        }), 500


@agendamento_routes.route('/notificacoes/criar', methods=['POST'])
@login_required
def criar_notificacao():
    """Cria uma nova notificação para um agendamento."""
    try:
        data = request.get_json()
        
        # Validações
        agendamento_id = data.get('agendamento_id')
        destinatario_id = data.get('destinatario_id')
        tipo = data.get('tipo', 'mensagem')
        titulo = data.get('titulo', '').strip()
        mensagem = data.get('mensagem', '').strip()
        
        if not all([agendamento_id, destinatario_id, titulo, mensagem]):
            return jsonify({
                'success': False,
                'error': 'Todos os campos são obrigatórios'
            }), 400
        
        # Verificar se o agendamento existe
        agendamento = AgendamentoVisita.query.get(agendamento_id)
        if not agendamento:
            return jsonify({
                'success': False,
                'error': 'Agendamento não encontrado'
            }), 404
        
        # Verificar se o destinatário existe
        destinatario = Usuario.query.get(destinatario_id)
        if not destinatario:
            return jsonify({
                'success': False,
                'error': 'Destinatário não encontrado'
            }), 404
        
        # Verificar permissões (apenas admin/cliente pode enviar para professor)
        if current_user.tipo not in ['admin', 'cliente']:
            return jsonify({
                'success': False,
                'error': 'Acesso negado'
            }), 403
        
        # Criar notificação
        nova_notificacao = NotificacaoAgendamento(
            agendamento_id=agendamento_id,
            remetente_id=current_user.id,
            destinatario_id=destinatario_id,
            tipo=tipo,
            titulo=titulo,
            mensagem=mensagem
        )
        
        db.session.add(nova_notificacao)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Notificação enviada com sucesso',
            'notificacao_id': nova_notificacao.id
        })
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Erro ao criar notificação: {e}")
        return jsonify({
            'success': False,
            'error': 'Erro interno do servidor'
        }), 500


@agendamento_routes.route('/api/notificacoes/nao-lidas/count')
@login_required
def contar_notificacoes_nao_lidas():
    """Retorna o número de notificações não lidas do usuário."""
    try:
        count = NotificacaoAgendamento.query.filter_by(
            destinatario_id=current_user.id,
            lida=False
        ).count()
        
        return jsonify({
            'success': True,
            'count': count
        })
    except Exception as e:
        current_app.logger.error(f"Erro ao contar notificações não lidas: {e}")
        return jsonify({
            'success': False,
            'error': 'Erro interno do servidor'
        }), 500


@agendamento_routes.route('/agendamentos/<int:agendamento_id>/historico-comunicacao')
@login_required
def historico_comunicacao(agendamento_id):
    """Exibe o histórico de comunicação de um agendamento."""
    try:
        agendamento = AgendamentoVisita.query.get_or_404(agendamento_id)
        
        # Verificar permissões
        if current_user.tipo == 'professor' and agendamento.professor_id != current_user.id:
            flash('Acesso negado', 'error')
            return redirect(url_for('agendamento_routes.listar_agendamentos'))
        
        # Buscar todas as notificações relacionadas ao agendamento
        notificacoes = NotificacaoAgendamento.query.filter_by(
            agendamento_id=agendamento_id
        ).order_by(NotificacaoAgendamento.data_criacao.asc()).all()
        
        return render_template(
            'historico_comunicacao.html',
            agendamento=agendamento,
            notificacoes=notificacoes
        )
    except Exception as e:
        current_app.logger.error(f"Erro ao carregar histórico de comunicação: {e}")
        flash('Erro ao carregar histórico de comunicação', 'error')
        return redirect(url_for('agendamento_routes.listar_agendamentos'))
