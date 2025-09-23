from flask import Blueprint, render_template, request, jsonify, send_file
from flask_login import login_required, current_user
from datetime import datetime, timedelta
import io
import pandas as pd

try:  # Alguns ambientes de execução não contam com openpyxl completo
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:  # pragma: no cover - fallback
    Workbook = None
    Font = PatternFill = Alignment = None
    OPENPYXL_AVAILABLE = False

from models.orcamento import Orcamento, HistoricoOrcamento
from models.user import Usuario
from models.material import Polo
from services.historico_orcamento_service import HistoricoOrcamentoService
from services.export_service import ExportService
from extensions import db

orcamento_bp = Blueprint('orcamento_routes', __name__, url_prefix='/orcamento')

@orcamento_bp.route('/historico')
@login_required
def historico_orcamento():
    """Página de visualização do histórico de alterações orçamentárias"""
    return render_template('compra/historico_orcamento.html')

@orcamento_bp.route('/api/historico-orcamento')
@login_required
def api_historico_orcamento():
    """API para obter histórico de alterações orçamentárias"""
    try:
        # Parâmetros de filtro
        data_inicio = request.args.get('data_inicio')
        data_fim = request.args.get('data_fim')
        tipo_alteracao = request.args.get('tipo_alteracao')
        polo_id = request.args.get('polo_id')
        pagina = int(request.args.get('pagina', 1))
        limite = int(request.args.get('limite', 20))
        
        # Construir query
        query = db.session.query(
            HistoricoOrcamento,
            Usuario.nome.label('usuario_nome')
        ).outerjoin(
            Usuario, HistoricoOrcamento.usuario_id == Usuario.id
        )
        
        # Aplicar filtros
        if data_inicio:
            query = query.filter(HistoricoOrcamento.data_alteracao >= datetime.strptime(data_inicio, '%Y-%m-%d'))
        
        if data_fim:
            data_fim_dt = datetime.strptime(data_fim, '%Y-%m-%d') + timedelta(days=1)
            query = query.filter(HistoricoOrcamento.data_alteracao < data_fim_dt)
        
        if tipo_alteracao:
            query = query.filter(HistoricoOrcamento.tipo_alteracao == tipo_alteracao)
        
        if polo_id:
            # Filtrar por polo através do orçamento
            query = query.join(Orcamento).filter(Orcamento.polo_id == polo_id)
        
        # Ordenar por data (mais recente primeiro)
        query = query.order_by(HistoricoOrcamento.data_alteracao.desc())
        
        # Paginação
        total = query.count()
        total_paginas = (total + limite - 1) // limite
        
        offset = (pagina - 1) * limite
        resultados = query.offset(offset).limit(limite).all()
        
        # Formatar dados
        historico = []
        for item, usuario_nome in resultados:
            historico.append({
                'id': item.id,
                'orcamento_id': item.orcamento_id,
                'tipo_alteracao': item.tipo_alteracao,
                'data_alteracao': item.data_alteracao.isoformat(),
                'usuario_nome': usuario_nome,
                'motivo': item.motivo,
                'observacoes': item.observacoes,
                'valor_total_anterior': float(item.valor_total_anterior or 0),
                'valor_total_novo': float(item.valor_total_novo),
                'valor_custeio_anterior': float(item.valor_custeio_anterior or 0),
                'valor_custeio_novo': float(item.valor_custeio_novo),
                'valor_capital_anterior': float(item.valor_capital_anterior or 0),
                'valor_capital_novo': float(item.valor_capital_novo),
                'variacao_total': float(item.variacao_total),
                'variacao_custeio': float(item.variacao_custeio),
                'variacao_capital': float(item.variacao_capital),
                'variacao_percentual_total': float(item.variacao_percentual_total),
                'variacao_percentual_custeio': float(item.variacao_percentual_custeio),
                'variacao_percentual_capital': float(item.variacao_percentual_capital)
            })
        
        return jsonify({
            'success': True,
            'historico': historico,
            'total': total,
            'pagina': pagina,
            'total_paginas': total_paginas
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Erro ao carregar histórico: {str(e)}'
        }), 500

@orcamento_bp.route('/api/historico-orcamento/<int:historico_id>')
@login_required
def api_detalhes_historico(historico_id):
    """API para obter detalhes de uma alteração específica"""
    try:
        item = db.session.query(
            HistoricoOrcamento,
            Usuario.nome.label('usuario_nome')
        ).outerjoin(
            Usuario, HistoricoOrcamento.usuario_id == Usuario.id
        ).filter(HistoricoOrcamento.id == historico_id).first()
        
        if not item:
            return jsonify({
                'success': False,
                'message': 'Registro não encontrado'
            }), 404
        
        historico_item, usuario_nome = item
        
        return jsonify({
            'success': True,
            'historico': {
                'id': historico_item.id,
                'orcamento_id': historico_item.orcamento_id,
                'tipo_alteracao': historico_item.tipo_alteracao,
                'data_alteracao': historico_item.data_alteracao.isoformat(),
                'usuario_nome': usuario_nome,
                'motivo': historico_item.motivo,
                'observacoes': historico_item.observacoes,
                'valor_total_anterior': float(historico_item.valor_total_anterior or 0),
                'valor_total_novo': float(historico_item.valor_total_novo),
                'valor_custeio_anterior': float(historico_item.valor_custeio_anterior or 0),
                'valor_custeio_novo': float(historico_item.valor_custeio_novo),
                'valor_capital_anterior': float(historico_item.valor_capital_anterior or 0),
                'valor_capital_novo': float(historico_item.valor_capital_novo),
                'variacao_total': float(historico_item.variacao_total),
                'variacao_custeio': float(historico_item.variacao_custeio),
                'variacao_capital': float(historico_item.variacao_capital),
                'variacao_percentual_total': float(historico_item.variacao_percentual_total),
                'variacao_percentual_custeio': float(historico_item.variacao_percentual_custeio),
                'variacao_percentual_capital': float(historico_item.variacao_percentual_capital)
            }
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Erro ao carregar detalhes: {str(e)}'
        }), 500

@orcamento_bp.route('/api/historico-orcamento/estatisticas')
@login_required
def api_estatisticas_historico():
    """API para obter estatísticas do histórico"""
    try:
        service = HistoricoOrcamentoService()
        
        # Parâmetros de filtro
        data_inicio = request.args.get('data_inicio')
        data_fim = request.args.get('data_fim')
        tipo_alteracao = request.args.get('tipo_alteracao')
        polo_id = request.args.get('polo_id')
        
        # Converter datas
        data_inicio_dt = None
        data_fim_dt = None
        
        if data_inicio:
            data_inicio_dt = datetime.strptime(data_inicio, '%Y-%m-%d')
        
        if data_fim:
            data_fim_dt = datetime.strptime(data_fim, '%Y-%m-%d') + timedelta(days=1)
        
        estatisticas = service.obter_estatisticas(
            data_inicio=data_inicio_dt,
            data_fim=data_fim_dt,
            tipo_alteracao=tipo_alteracao,
            polo_id=int(polo_id) if polo_id else None
        )
        
        return jsonify({
            'success': True,
            'estatisticas': estatisticas
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Erro ao carregar estatísticas: {str(e)}'
        }), 500

@orcamento_bp.route('/api/historico-orcamento/exportar')
@login_required
def api_exportar_historico():
    """API para exportar histórico em Excel"""
    try:
        if not OPENPYXL_AVAILABLE:
            return jsonify({
                'success': False,
                'message': 'Exportação XLSX indisponível: biblioteca openpyxl não instalada.'
            }), 503

        # Parâmetros de filtro
        data_inicio = request.args.get('data_inicio')
        data_fim = request.args.get('data_fim')
        tipo_alteracao = request.args.get('tipo_alteracao')
        polo_id = request.args.get('polo_id')
        
        # Construir query
        query = db.session.query(
            HistoricoOrcamento,
            Usuario.nome.label('usuario_nome'),
            Orcamento.nome.label('orcamento_nome'),
            Polo.nome.label('polo_nome')
        ).outerjoin(
            Usuario, HistoricoOrcamento.usuario_id == Usuario.id
        ).join(
            Orcamento, HistoricoOrcamento.orcamento_id == Orcamento.id
        ).outerjoin(
            Polo, Orcamento.polo_id == Polo.id
        )
        
        # Aplicar filtros
        if data_inicio:
            query = query.filter(HistoricoOrcamento.data_alteracao >= datetime.strptime(data_inicio, '%Y-%m-%d'))
        
        if data_fim:
            data_fim_dt = datetime.strptime(data_fim, '%Y-%m-%d') + timedelta(days=1)
            query = query.filter(HistoricoOrcamento.data_alteracao < data_fim_dt)
        
        if tipo_alteracao:
            query = query.filter(HistoricoOrcamento.tipo_alteracao == tipo_alteracao)
        
        if polo_id:
            query = query.filter(Orcamento.polo_id == polo_id)
        
        # Ordenar por data
        query = query.order_by(HistoricoOrcamento.data_alteracao.desc())
        
        resultados = query.all()
        
        # Criar workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Histórico de Alterações"
        
        # Cabeçalhos
        headers = [
            'ID', 'Data/Hora', 'Tipo', 'Orçamento', 'Polo', 'Usuário',
            'Valor Total Anterior', 'Valor Total Novo', 'Variação Total',
            'Valor Custeio Anterior', 'Valor Custeio Novo', 'Variação Custeio',
            'Valor Capital Anterior', 'Valor Capital Novo', 'Variação Capital',
            'Motivo', 'Observações'
        ]
        
        # Estilo do cabeçalho
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Dados
        for row, (item, usuario_nome, orcamento_nome, polo_nome) in enumerate(resultados, 2):
            ws.cell(row=row, column=1, value=item.id)
            ws.cell(row=row, column=2, value=item.data_alteracao.strftime('%d/%m/%Y %H:%M:%S'))
            ws.cell(row=row, column=3, value=item.tipo_alteracao.upper())
            ws.cell(row=row, column=4, value=orcamento_nome or f'ID {item.orcamento_id}')
            ws.cell(row=row, column=5, value=polo_nome or 'N/A')
            ws.cell(row=row, column=6, value=usuario_nome or 'Sistema')
            ws.cell(row=row, column=7, value=float(item.valor_total_anterior or 0))
            ws.cell(row=row, column=8, value=float(item.valor_total_novo))
            ws.cell(row=row, column=9, value=float(item.variacao_total))
            ws.cell(row=row, column=10, value=float(item.valor_custeio_anterior or 0))
            ws.cell(row=row, column=11, value=float(item.valor_custeio_novo))
            ws.cell(row=row, column=12, value=float(item.variacao_custeio))
            ws.cell(row=row, column=13, value=float(item.valor_capital_anterior or 0))
            ws.cell(row=row, column=14, value=float(item.valor_capital_novo))
            ws.cell(row=row, column=15, value=float(item.variacao_capital))
            ws.cell(row=row, column=16, value=item.motivo or '')
            ws.cell(row=row, column=17, value=item.observacoes or '')
        
        # Ajustar largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Salvar em buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Gerar nome do arquivo
        data_atual = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'historico_orcamento_{data_atual}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Erro ao exportar histórico: {str(e)}'
        }), 500

@orcamento_bp.route('/api/polos')
@login_required
def api_polos():
    """API para obter lista de polos"""
    try:
        polos = Polo.query.filter_by(ativo=True).order_by(Polo.nome).all()
        
        return jsonify({
            'success': True,
            'polos': [{'id': polo.id, 'nome': polo.nome} for polo in polos]
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Erro ao carregar polos: {str(e)}'
        }), 500


@orcamento_bp.route('/exportar/excel')
@login_required
def exportar_orcamentos_excel():
    """Exporta relatório de orçamentos em Excel"""
    try:
        # Obter filtros da query string
        data_inicio = request.args.get('data_inicio')
        data_fim = request.args.get('data_fim')
        polo_id = request.args.get('polo_id')
        
        # Construir query
        query = Orcamento.query
        
        if data_inicio:
            try:
                data_inicio_obj = datetime.strptime(data_inicio, '%Y-%m-%d')
                query = query.filter(Orcamento.data_criacao >= data_inicio_obj)
            except ValueError:
                pass
        
        if data_fim:
            try:
                data_fim_obj = datetime.strptime(data_fim, '%Y-%m-%d')
                query = query.filter(Orcamento.data_criacao <= data_fim_obj)
            except ValueError:
                pass
        
        if polo_id:
            query = query.filter(Orcamento.polo_id == polo_id)
        
        # Executar query
        orcamentos = query.order_by(Orcamento.data_criacao.desc()).all()
        
        # Exportar para Excel
        filepath = ExportService.export_orcamentos_excel(orcamentos)
        
        return send_file(filepath, as_attachment=True, download_name=os.path.basename(filepath))
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Erro ao exportar orçamentos para Excel: {str(e)}'
        }), 500


@orcamento_bp.route('/exportar/pdf')
@login_required
def exportar_orcamentos_pdf():
    """Exporta relatório de orçamentos em PDF"""
    try:
        # Obter filtros da query string
        data_inicio = request.args.get('data_inicio')
        data_fim = request.args.get('data_fim')
        polo_id = request.args.get('polo_id')
        
        # Construir query
        query = Orcamento.query
        
        if data_inicio:
            try:
                data_inicio_obj = datetime.strptime(data_inicio, '%Y-%m-%d')
                query = query.filter(Orcamento.data_criacao >= data_inicio_obj)
            except ValueError:
                pass
        
        if data_fim:
            try:
                data_fim_obj = datetime.strptime(data_fim, '%Y-%m-%d')
                query = query.filter(Orcamento.data_criacao <= data_fim_obj)
            except ValueError:
                pass
        
        if polo_id:
            query = query.filter(Orcamento.polo_id == polo_id)
        
        # Executar query
        orcamentos = query.order_by(Orcamento.data_criacao.desc()).all()
        
        # Exportar para PDF
        filepath = ExportService.export_orcamentos_pdf(orcamentos)
        
        return send_file(filepath, as_attachment=True, download_name=os.path.basename(filepath))
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Erro ao exportar orçamentos para PDF: {str(e)}'
        }), 500


@orcamento_bp.route('/historico/exportar/excel')
@login_required
def exportar_historico_excel():
    """Exporta histórico de orçamentos em Excel"""
    try:
        # Obter filtros da query string
        data_inicio = request.args.get('data_inicio')
        data_fim = request.args.get('data_fim')
        tipo_alteracao = request.args.get('tipo_alteracao')
        polo_id = request.args.get('polo_id')
        
        # Construir query
        query = HistoricoOrcamento.query
        
        if data_inicio:
            try:
                data_inicio_obj = datetime.strptime(data_inicio, '%Y-%m-%d')
                query = query.filter(HistoricoOrcamento.data_alteracao >= data_inicio_obj)
            except ValueError:
                pass
        
        if data_fim:
            try:
                data_fim_obj = datetime.strptime(data_fim, '%Y-%m-%d') + timedelta(days=1)
                query = query.filter(HistoricoOrcamento.data_alteracao < data_fim_obj)
            except ValueError:
                pass
        
        if tipo_alteracao:
            query = query.filter(HistoricoOrcamento.tipo_alteracao == tipo_alteracao)
        
        if polo_id:
            query = query.filter(HistoricoOrcamento.polo_id == polo_id)
        
        # Executar query
        historico = query.order_by(HistoricoOrcamento.data_alteracao.desc()).all()
        
        # Exportar para Excel
        filepath = ExportService.export_historico_orcamento_excel(historico)
        
        return send_file(filepath, as_attachment=True, download_name=os.path.basename(filepath))
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Erro ao exportar histórico para Excel: {str(e)}'
        }), 500


@orcamento_bp.route('/historico/exportar/pdf')
@login_required
def exportar_historico_pdf():
    """Exporta histórico de orçamentos em PDF"""
    try:
        # Obter filtros da query string
        data_inicio = request.args.get('data_inicio')
        data_fim = request.args.get('data_fim')
        tipo_alteracao = request.args.get('tipo_alteracao')
        polo_id = request.args.get('polo_id')
        
        # Construir query
        query = HistoricoOrcamento.query
        
        if data_inicio:
            try:
                data_inicio_obj = datetime.strptime(data_inicio, '%Y-%m-%d')
                query = query.filter(HistoricoOrcamento.data_alteracao >= data_inicio_obj)
            except ValueError:
                pass
        
        if data_fim:
            try:
                data_fim_obj = datetime.strptime(data_fim, '%Y-%m-%d') + timedelta(days=1)
                query = query.filter(HistoricoOrcamento.data_alteracao < data_fim_obj)
            except ValueError:
                pass
        
        if tipo_alteracao:
            query = query.filter(HistoricoOrcamento.tipo_alteracao == tipo_alteracao)
        
        if polo_id:
            query = query.filter(HistoricoOrcamento.polo_id == polo_id)
        
        # Executar query
        historico = query.order_by(HistoricoOrcamento.data_alteracao.desc()).all()
        
        # Exportar para PDF
        filepath = ExportService.export_historico_orcamento_pdf(historico)
        
        return send_file(filepath, as_attachment=True, download_name=os.path.basename(filepath))
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Erro ao exportar histórico para PDF: {str(e)}'
        }), 500
