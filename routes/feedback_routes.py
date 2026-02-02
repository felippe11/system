"""
Rotas para gerenciar perguntas e respostas de feedback personalizadas.
"""

from flask import Blueprint, request, jsonify, render_template, flash, redirect, url_for
from flask_login import login_required, current_user
from models import (
    PerguntaFeedback,
    RespostaFeedback,
    FeedbackSession,
    Oficina,
    Cliente,
    Feedback,
    Usuario,
    Evento,
    FeedbackTemplate,
    FeedbackTemplateOficina,
)
from services.feedback_service import FeedbackService
from extensions import db
import json

feedback_routes = Blueprint('feedback_routes', __name__)


@feedback_routes.route('/feedback/perguntas/<int:oficina_id>')
@login_required
def gerenciar_perguntas_oficina(oficina_id):
    """Página para gerenciar perguntas de feedback de uma oficina específica."""
    oficina = Oficina.query.get_or_404(oficina_id)
    
    # Verificar permissão
    if oficina.cliente_id != current_user.cliente_id:
        flash("Você não tem permissão para acessar esta oficina.", "error")
        return redirect(url_for('feedback_routes.gerenciar_perguntas'))
    
    perguntas = FeedbackService.listar_perguntas(current_user.cliente_id, oficina_id)
    
    return render_template('feedback/gerenciar_perguntas.html', 
                         oficina=oficina, 
                         perguntas=perguntas)


@feedback_routes.route('/feedback/perguntas/<int:oficina_id>/criar', methods=['GET', 'POST'])
@login_required
def criar_pergunta(oficina_id):
    """Criar nova pergunta de feedback."""
    oficina = Oficina.query.get_or_404(oficina_id)
    
    # Verificar permissão
    if oficina.cliente_id != current_user.cliente_id:
        flash("Você não tem permissão para acessar esta oficina.", "error")
        return redirect(url_for('feedback_routes.gerenciar_perguntas'))
    
    if request.method == 'POST':
        try:
            data = request.get_json()

            # Se a oficina estiver vinculada a um template, desfaz o vínculo e copia as perguntas
            FeedbackService.desvincular_oficina(oficina_id)
            
            pergunta = FeedbackService.criar_pergunta(
                cliente_id=current_user.cliente_id,
                oficina_id=oficina_id,
                titulo=data['titulo'],
                descricao=data.get('descricao', ''),
                tipo=data['tipo'],
                opcoes=data.get('opcoes', []),
                obrigatoria=data.get('obrigatoria', True),
                ordem=data.get('ordem', 0)
            )
            
            return jsonify({
                'success': True,
                'message': 'Pergunta criada com sucesso!',
                'pergunta': pergunta.to_dict()
            })
            
        except Exception as e:
            return jsonify({
                'success': False,
                'message': f'Erro ao criar pergunta: {str(e)}'
            }), 400
    
    return render_template('feedback/criar_pergunta.html', oficina=oficina)


@feedback_routes.route('/feedback/perguntas/<int:pergunta_id>/editar', methods=['GET', 'POST'])
@login_required
def editar_pergunta(pergunta_id):
    """Editar pergunta de feedback."""
    pergunta = PerguntaFeedback.query.get_or_404(pergunta_id)
    
    # Verificar permissão
    if pergunta.cliente_id != current_user.cliente_id:
        flash("Você não tem permissão para editar esta pergunta.", "error")
        return redirect(url_for('feedback_routes.gerenciar_perguntas'))
    
    if request.method == 'POST':
        try:
            data = request.get_json()
            
            pergunta_atualizada = FeedbackService.atualizar_pergunta(
                pergunta_id,
                titulo=data.get('titulo'),
                descricao=data.get('descricao'),
                tipo=data.get('tipo'),
                opcoes=data.get('opcoes'),
                obrigatoria=data.get('obrigatoria'),
                ordem=data.get('ordem')
            )
            
            return jsonify({
                'success': True,
                'message': 'Pergunta atualizada com sucesso!',
                'pergunta': pergunta_atualizada.to_dict()
            })
            
        except Exception as e:
            return jsonify({
                'success': False,
                'message': f'Erro ao atualizar pergunta: {str(e)}'
            }), 400
    
    return render_template('feedback/editar_pergunta.html', pergunta=pergunta)


@feedback_routes.route('/feedback/perguntas/detalhes/<int:pergunta_id>')
@login_required
def detalhes_pergunta(pergunta_id):
    pergunta = PerguntaFeedback.query.get_or_404(pergunta_id)
    data = pergunta.to_dict()
    try:
        data["opcoes"] = json.loads(pergunta.opcoes) if pergunta.opcoes else []
    except Exception:
        data["opcoes"] = []
    data["oficina_id"] = pergunta.oficina_id
    data["atividade_id"] = pergunta.atividade_id
    return jsonify(data)


@feedback_routes.route('/feedback/perguntas/<int:pergunta_id>/deletar', methods=['POST'])
@login_required
def deletar_pergunta(pergunta_id):
    """Deletar pergunta de feedback."""
    pergunta = PerguntaFeedback.query.get_or_404(pergunta_id)
    
    # Verificar permissão
    if pergunta.cliente_id != current_user.cliente_id:
        return jsonify({
            'success': False,
            'message': 'Você não tem permissão para deletar esta pergunta.'
        }), 403
    
    try:
        FeedbackService.deletar_pergunta(pergunta_id)
        
        return jsonify({
            'success': True,
            'message': 'Pergunta deletada com sucesso!'
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Erro ao deletar pergunta: {str(e)}'
        }), 400


@feedback_routes.route('/feedback/perguntas/reordenar', methods=['POST'])
@login_required
def reordenar_perguntas():
    """Reordenar perguntas de feedback."""
    try:
        data = request.get_json()
        pergunta_ids = data.get('pergunta_ids', [])
        
        FeedbackService.reordenar_perguntas(pergunta_ids)
        
        return jsonify({
            'success': True,
            'message': 'Ordem das perguntas atualizada com sucesso!'
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Erro ao reordenar perguntas: {str(e)}'
        }), 400


@feedback_routes.route('/feedback/responder/<token>')
def responder_feedback(token):
    """Página para responder feedback via QR Code."""
    sessao = FeedbackService.validar_sessao_feedback(token)
    
    if not sessao:
        return render_template('feedback/sessao_invalida.html')
    
    # Obter perguntas da oficina
    perguntas = FeedbackService.listar_perguntas(sessao.oficina.cliente_id, sessao.oficina_id)
    
    return render_template('feedback/responder_feedback.html', 
                         sessao=sessao, 
                         perguntas=perguntas)


@feedback_routes.route('/feedback/responder/<token>', methods=['POST'])
def salvar_feedback(token):
    """Salvar respostas de feedback."""
    try:
        data = request.get_json()
        respostas = data.get('respostas', {})
        
        FeedbackService.salvar_respostas(token, respostas)
        
        return jsonify({
            'success': True,
            'message': 'Feedback enviado com sucesso! Obrigado pela sua participação.'
        })
        
    except ValueError as e:
        return jsonify({
            'success': False,
            'message': str(e)
        }), 400
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Erro ao salvar feedback: {str(e)}'
        }), 500


@feedback_routes.route('/feedback/responder/oficina/<int:oficina_id>')
def responder_feedback_publico(oficina_id):
    """Página pública para responder feedback por oficina."""
    oficina = Oficina.query.get_or_404(oficina_id)
    perguntas = FeedbackService.listar_perguntas(oficina.cliente_id, oficina.id)
    return render_template(
        'feedback/responder_feedback_publico.html',
        oficina=oficina,
        perguntas=perguntas,
    )


@feedback_routes.route('/feedback/responder/oficina/<int:oficina_id>', methods=['POST'])
def salvar_feedback_publico(oficina_id):
    """Salvar respostas públicas de feedback."""
    try:
        data = request.get_json()
        respostas = data.get('respostas', {})
        FeedbackService.salvar_respostas_publico(oficina_id, respostas)
        return jsonify({
            'success': True,
            'message': 'Feedback enviado com sucesso! Obrigado pela sua participação.'
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Erro ao salvar feedback: {str(e)}'
        }), 500


@feedback_routes.route('/feedback/estatisticas/<int:oficina_id>')
@login_required
def estatisticas_feedback(oficina_id):
    """Página de estatísticas de feedback de uma oficina."""
    oficina = Oficina.query.get_or_404(oficina_id)
    
    # Verificar permissão
    if oficina.cliente_id != current_user.cliente_id:
        flash("Você não tem permissão para acessar esta oficina.", "error")
        return redirect(url_for('feedback_routes.gerenciar_perguntas'))
    
    estatisticas = FeedbackService.obter_estatisticas_feedback(oficina_id)
    respostas = FeedbackService.obter_respostas_oficina(oficina_id)
    
    return render_template('feedback/estatisticas.html', 
                         oficina=oficina, 
                         estatisticas=estatisticas,
                         respostas=respostas)


@feedback_routes.route('/feedback/gerar-qr/<int:oficina_id>')
@login_required
def gerar_qr_feedback(oficina_id):
    """Gerar QR Code para feedback de uma oficina."""
    oficina = Oficina.query.get_or_404(oficina_id)
    
    # Verificar permissão
    if oficina.cliente_id != current_user.cliente_id:
        flash("Você não tem permissão para acessar esta oficina.", "error")
        return redirect(url_for('feedback_routes.gerenciar_perguntas'))
    
    # Obter inscritos da oficina
    inscritos = db.session.query(Usuario).join(Usuario.inscricoes).filter(
        Usuario.inscricoes.any(oficina_id=oficina_id)
    ).all()
    
    return render_template('feedback/gerar_qr.html', 
                         oficina=oficina, 
                         inscritos=inscritos)


@feedback_routes.route('/feedback/qr-code-geral/<int:oficina_id>')
@login_required
def gerar_qr_code_geral(oficina_id):
    """Gera QR Code geral para feedback de uma oficina."""
    oficina = Oficina.query.get_or_404(oficina_id)
    
    # Verificar permissão
    if oficina.cliente_id != current_user.cliente_id:
        flash("Você não tem permissão para acessar esta oficina.", "error")
        return redirect(url_for('feedback_routes.gerenciar_perguntas'))
    
    try:
        import qrcode
        from io import BytesIO
        from flask import Response
        
        # URL pública para responder feedback da oficina
        feedback_url = url_for(
            'feedback_routes.responder_feedback_publico',
            oficina_id=oficina_id,
            _external=True
        )
        
        # Criar QR Code
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )
        qr.add_data(feedback_url)
        qr.make(fit=True)
        
        # Gerar imagem
        img = qr.make_image(fill_color="black", back_color="white")
        
        # Converter para bytes
        img_io = BytesIO()
        img.save(img_io, format='PNG')
        img_io.seek(0)
        
        return Response(img_io.getvalue(), mimetype='image/png')
        
    except Exception as e:
        flash(f"Erro ao gerar QR Code: {str(e)}", "error")
        return redirect(url_for('feedback_routes.gerar_qr_feedback', oficina_id=oficina_id))


@feedback_routes.route('/feedback/qr-code-individual/<int:usuario_id>/<int:oficina_id>')
@login_required
def gerar_qr_code_individual(usuario_id, oficina_id):
    """Gera QR Code individual para feedback de um usuário específico."""
    oficina = Oficina.query.get_or_404(oficina_id)
    
    # Verificar permissão
    if oficina.cliente_id != current_user.cliente_id:
        flash("Você não tem permissão para acessar esta oficina.", "error")
        return redirect(url_for('feedback_routes.gerenciar_perguntas'))
    
    try:
        import qrcode
        from io import BytesIO
        from flask import Response
        
        # Criar sessão de feedback para o usuário
        sessao = FeedbackService.criar_sessao_feedback(usuario_id, oficina_id)
        
        # URL para o feedback individual
        feedback_url = url_for('feedback_routes.responder_feedback', token=sessao.token, _external=True)
        
        # Criar QR Code
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=8,
            border=4,
        )
        qr.add_data(feedback_url)
        qr.make(fit=True)
        
        # Gerar imagem
        img = qr.make_image(fill_color="black", back_color="white")
        
        # Converter para bytes
        img_io = BytesIO()
        img.save(img_io, format='PNG')
        img_io.seek(0)
        
        return Response(img_io.getvalue(), mimetype='image/png')
        
    except Exception as e:
        flash(f"Erro ao gerar QR Code individual: {str(e)}", "error")
        return redirect(url_for('feedback_routes.gerar_qr_feedback', oficina_id=oficina_id))


@feedback_routes.route('/feedback/gerar-sessao/<int:usuario_id>/<int:oficina_id>', methods=['POST'])
@login_required
def gerar_sessao_feedback(usuario_id, oficina_id):
    """Gerar sessão de feedback para um usuário específico."""
    oficina = Oficina.query.get_or_404(oficina_id)
    
    # Verificar permissão
    if oficina.cliente_id != current_user.cliente_id:
        return jsonify({
            'success': False,
            'message': 'Você não tem permissão para acessar esta oficina.'
        }), 403
    
    try:
        sessao = FeedbackService.criar_sessao_feedback(usuario_id, oficina_id)
        
        return jsonify({
            'success': True,
            'message': 'Sessão de feedback criada com sucesso!',
            'token': sessao.token,
            'url': url_for('feedback_routes.responder_feedback', token=sessao.token, _external=True)
        })
        
    except ValueError as e:
        return jsonify({
            'success': False,
            'message': str(e)
        }), 400
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Erro ao gerar sessão: {str(e)}'
        }), 500


@feedback_routes.route('/feedback/gerenciar')
@login_required
def gerenciar_perguntas():
    """Página principal para gerenciar perguntas de feedback."""
    # Debug: Informações do usuário atual
    print(f"DEBUG: Usuário atual: {current_user}")
    print(f"DEBUG: ID do usuário: {current_user.id}")
    print(f"DEBUG: Tipo do usuário: {current_user.tipo}")
    print(f"DEBUG: Cliente ID: {current_user.cliente_id}")
    
    cliente_id = getattr(current_user, 'cliente_id', None)

    # Identificar todos os clientes associados ao usuário
    cliente_ids = set()

    if cliente_id:
        cliente_ids.add(cliente_id)

    # Usuários do tipo Cliente (modelo Cliente) devem enxergar suas próprias oficinas
    if isinstance(current_user, Cliente):
        cliente_ids.add(getattr(current_user, 'id', None))

    # Considerar associações via relacionamento muitos-para-muitos (usuario_clientes)
    if hasattr(current_user, 'clientes'):
        try:
            associados = current_user.clientes
            # InstrumentedList é iterável; usamos list() para garantir avaliação imediata
            for cliente in list(associados):
                cliente_ids.add(cliente.id)
        except Exception as exc:
            print(f"DEBUG: Erro ao carregar clientes associados: {exc}")

    cliente_ids.discard(None)

    print(f"DEBUG: Cliente IDs associados ao usuário: {cliente_ids}")

    if not cliente_ids:
        oficinas = []
        eventos = []
        templates = []
        template_padrao = None
        print("DEBUG: Usuário sem clientes associados. Nenhuma oficina carregada.")
    else:
        oficinas = (
            Oficina.query
            .filter(Oficina.cliente_id.in_(list(cliente_ids)))
            .order_by(Oficina.titulo.asc())
            .all()
        )
        
        eventos = (
            Evento.query
            .filter(Evento.cliente_id.in_(list(cliente_ids)))
            .order_by(Evento.nome.asc())
            .all()
        )
        
        print(f"DEBUG: Total de oficinas encontradas: {len(oficinas)}")
        print(f"DEBUG: Total de eventos encontrados: {len(eventos)}")
        for oficina in oficinas:
            print(
                "DEBUG: Oficina - ID: {oficina.id}, "
                f"Título: {oficina.titulo}, Cliente ID: {oficina.cliente_id}"
            )

        cliente_base_id = next(iter(cliente_ids)) if cliente_ids else None
        templates = (
            FeedbackTemplate.query
            .filter(FeedbackTemplate.cliente_id == cliente_base_id)
            .order_by(FeedbackTemplate.nome.asc())
            .all()
        )
        template_padrao = next((t for t in templates if t.is_default), None)
    
    return render_template(
        'feedback/gerenciar_perguntas.html',
        oficinas=oficinas,
        eventos=eventos,
        templates=templates,
        template_padrao=template_padrao,
    )


@feedback_routes.route('/feedback/templates')
@login_required
def gerenciar_templates():
    cliente_id = getattr(current_user, "cliente_id", None)
    if isinstance(current_user, Cliente) and not cliente_id:
        cliente_id = current_user.id
    templates = FeedbackTemplate.query.filter_by(cliente_id=cliente_id).order_by(
        FeedbackTemplate.nome.asc()
    ).all()
    template_padrao = next((t for t in templates if t.is_default), None)
    return render_template(
        'feedback/gerenciar_templates.html',
        templates=templates,
        template_padrao=template_padrao,
    )


@feedback_routes.route('/feedback/templates/criar', methods=['POST'])
@login_required
def criar_template():
    data = request.get_json()
    cliente_id = getattr(current_user, "cliente_id", None)
    if isinstance(current_user, Cliente) and not cliente_id:
        cliente_id = current_user.id
    template = FeedbackService.criar_template(
        cliente_id=cliente_id,
        nome=data.get("nome"),
        descricao=data.get("descricao"),
        is_default=bool(data.get("is_default")),
    )
    return jsonify({"success": True, "template": {"id": template.id, "nome": template.nome}})


@feedback_routes.route('/feedback/templates/<int:template_id>/definir-padrao', methods=['POST'])
@login_required
def definir_template_padrao(template_id):
    cliente_id = getattr(current_user, "cliente_id", None)
    if isinstance(current_user, Cliente) and not cliente_id:
        cliente_id = current_user.id
    FeedbackService.definir_template_padrao(cliente_id, template_id)
    return jsonify({"success": True})


@feedback_routes.route('/feedback/templates/<int:template_id>/aplicar', methods=['POST'])
@login_required
def aplicar_template(template_id):
    data = request.get_json()
    oficina_ids = data.get("oficina_ids", [])
    cliente_id = getattr(current_user, "cliente_id", None)
    if isinstance(current_user, Cliente) and not cliente_id:
        cliente_id = current_user.id
    oficinas = Oficina.query.filter(
        Oficina.id.in_(oficina_ids),
        Oficina.cliente_id == cliente_id,
    ).all()
    oficina_ids = [oficina.id for oficina in oficinas]
    FeedbackService.aplicar_template(template_id, oficina_ids)
    return jsonify({"success": True})


@feedback_routes.route('/feedback/templates/<int:template_id>/perguntas')
@login_required
def gerenciar_perguntas_template(template_id):
    template = FeedbackTemplate.query.get_or_404(template_id)
    perguntas = PerguntaFeedback.query.filter_by(
        template_id=template_id, ativa=True
    ).order_by(PerguntaFeedback.ordem, PerguntaFeedback.id).all()
    return render_template(
        "feedback/gerenciar_template.html",
        template=template,
        perguntas=perguntas,
    )


@feedback_routes.route('/feedback/templates/<int:template_id>/perguntas/criar', methods=['POST'])
@login_required
def criar_pergunta_template(template_id):
    data = request.get_json()
    template = FeedbackTemplate.query.get_or_404(template_id)
    pergunta = FeedbackService.criar_pergunta_template(
        template_id=template_id,
        cliente_id=template.cliente_id,
        titulo=data.get("titulo"),
        descricao=data.get("descricao", ""),
        tipo=data.get("tipo"),
        opcoes=data.get("opcoes", []),
        obrigatoria=data.get("obrigatoria", True),
        ordem=data.get("ordem", 0),
    )
    return jsonify(
        {"success": True, "message": "Pergunta criada com sucesso!", "pergunta": pergunta.to_dict()}
    )


@feedback_routes.route('/feedback/oficinas-por-evento/<int:evento_id>')
@login_required
def oficinas_por_evento(evento_id):
    """Retorna oficinas de um evento específico."""
    evento = Evento.query.get_or_404(evento_id)

    user_role = getattr(current_user, "tipo", None)
    is_admin = user_role in ("admin", "superadmin")
    cliente_id = getattr(current_user, "cliente_id", None)
    if isinstance(current_user, Cliente) and not cliente_id:
        cliente_id = current_user.id

    # Verificar permissão
    if not is_admin and evento.cliente_id != cliente_id:
        return jsonify({"error": "Acesso negado"}), 403
    
    # Buscar oficinas do evento
    oficinas = Oficina.query.filter_by(evento_id=evento_id).order_by(Oficina.titulo.asc()).all()
    
    oficinas_data = []
    for oficina in oficinas:
        primeira_data = None
        if getattr(oficina, "dias", None):
            try:
                primeira_data = min(
                    (dia.data for dia in oficina.dias if getattr(dia, "data", None)),
                    default=None,
                )
            except TypeError:
                primeira_data = None
        local = getattr(oficina, "local", None)
        if not local:
            cidade = getattr(oficina, "cidade", None)
            estado = getattr(oficina, "estado", None)
            if cidade and estado:
                local = f"{cidade}/{estado}"
            elif cidade:
                local = cidade
        if not local and getattr(oficina, "evento", None):
            local = getattr(oficina.evento, "localizacao", None)

        oficinas_data.append({
            'id': oficina.id,
            'titulo': oficina.titulo,
            'data_inicio': primeira_data.isoformat() if primeira_data else None,
            'local': local
        })
    
    return jsonify({'oficinas': oficinas_data})


@feedback_routes.route('/feedback/debug-oficinas')
@login_required
def debug_oficinas():
    """Rota de debug para verificar oficinas do cliente."""
    cliente_id = current_user.cliente_id
    
    # Buscar todas as oficinas do cliente
    oficinas = Oficina.query.filter_by(cliente_id=cliente_id).all()
    
    # Buscar todas as oficinas (para comparação)
    todas_oficinas = Oficina.query.all()
    
    # Informações de debug
    debug_info = {
        'cliente_id': cliente_id,
        'usuario_tipo': current_user.tipo,
        'total_oficinas': len(oficinas),
        'oficinas_cliente': len(oficinas),
        'total_geral': len(todas_oficinas),
        'oficinas': [
            {
                'id': oficina.id,
                'titulo': oficina.titulo,
                'cliente_id': oficina.cliente_id,
                'data_inicio': oficina.data_inicio.isoformat() if oficina.data_inicio else None
            }
            for oficina in oficinas
        ]
    }
    
    return jsonify(debug_info)


@feedback_routes.route('/feedback/oficina/<int:oficina_id>')
@login_required
def feedback_oficina(oficina_id):
    """Exibe os feedbacks e estatísticas de uma oficina específica."""
    oficina = Oficina.query.get_or_404(oficina_id)
    user_role = getattr(current_user, 'tipo', None)
    is_admin = user_role in ('admin', 'superadmin')

    if not is_admin:
        if user_role == 'cliente':
            if oficina.cliente_id != current_user.id:
                flash('Você não tem permissão para acessar esta oficina.', 'error')
                return redirect(url_for('dashboard_routes.dashboard_cliente'))
        elif user_role == 'ministrante':
            ministrante_id = getattr(current_user, 'id', None)
            associado = bool(
                oficina.ministrantes_associados.filter_by(id=ministrante_id).first()
            ) if hasattr(oficina.ministrantes_associados, 'filter_by') else False
            autorizado = any((
                oficina.ministrante_id == ministrante_id,
                oficina.formador_id == ministrante_id,
                associado,
            ))
            if not autorizado:
                flash('Você não tem permissão para acessar esta oficina.', 'error')
                return redirect(url_for('formador_routes.dashboard_formador'))
        else:
            flash('Você não tem permissão para acessar esta página.', 'error')
            return redirect(url_for('evento_routes.home'))

    cliente_filter = request.args.get('cliente_id', '').strip()
    tipo_filter = request.args.get('tipo', '').strip()
    estrelas_filter = request.args.get('estrelas', '').strip()

    if not cliente_filter:
        cliente_filter = str(oficina.cliente_id) if oficina.cliente_id else ''

    feedback_query = Feedback.query.filter_by(oficina_id=oficina.id)

    if tipo_filter == 'usuario':
        feedback_query = feedback_query.filter(Feedback.usuario_id.isnot(None))
    elif tipo_filter == 'ministrante':
        feedback_query = feedback_query.filter(Feedback.ministrante_id.isnot(None))

    if estrelas_filter:
        try:
            estrelas = int(estrelas_filter)
        except ValueError:
            estrelas = None
        if estrelas:
            feedback_query = feedback_query.filter(Feedback.rating == estrelas)

    feedbacks = feedback_query.order_by(Feedback.created_at.desc()).all()

    def _media(feedback_list):
        notas = [fb.rating for fb in feedback_list if fb.rating is not None]
        return sum(notas) / len(notas) if notas else 0

    total_count = len(feedbacks)
    total_avg = _media(feedbacks)

    feedbacks_usuarios = [fb for fb in feedbacks if fb.usuario_id]
    count_usuarios = len(feedbacks_usuarios)
    avg_usuarios = _media(feedbacks_usuarios)

    feedbacks_ministrantes = [fb for fb in feedbacks if fb.ministrante_id]
    count_ministrantes = len(feedbacks_ministrantes)
    avg_ministrantes = _media(feedbacks_ministrantes)

    clientes = Cliente.query.order_by(Cliente.nome).all() if is_admin else []

    return render_template(
        'oficina/feedback_oficina.html',
        oficina=oficina,
        feedbacks=feedbacks,
        total_count=total_count,
        total_avg=total_avg,
        count_usuarios=count_usuarios,
        avg_usuarios=avg_usuarios,
        count_ministrantes=count_ministrantes,
        avg_ministrantes=avg_ministrantes,
        is_admin=is_admin,
        clientes=clientes,
        cliente_filter=cliente_filter,
        tipo_filter=tipo_filter,
        estrelas_filter=estrelas_filter,
    )


@feedback_routes.route('/feedback/exportar')
@login_required
def exportar_feedback():
    """Página para exportar relatórios de feedback"""
    cliente_id = current_user.cliente_id
    oficinas = Oficina.query.filter_by(cliente_id=cliente_id).all()
    return render_template('feedback/exportar_feedback.html', oficinas=oficinas)


@feedback_routes.route('/feedback/exportar/pdf/<int:oficina_id>')
@login_required
def exportar_feedback_pdf(oficina_id):
    """Exporta feedback de uma oficina específica em PDF"""
    from services.pdf_service import gerar_pdf_feedback_route
    return gerar_pdf_feedback_route(oficina_id)


@feedback_routes.route('/feedback/exportar/xlsx/<int:oficina_id>')
@login_required
def exportar_feedback_xlsx(oficina_id):
    """Exporta feedback de uma oficina específica em XLSX"""
    from services.feedback_service import FeedbackService
    from flask import send_file
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from datetime import datetime
    
    oficina = Oficina.query.get_or_404(oficina_id)
    if oficina.cliente_id != current_user.cliente_id:
        flash("Acesso negado!", "danger")
        return redirect(url_for('feedback_routes.exportar_feedback'))
    
    # Buscar feedbacks
    from models import Feedback
    feedbacks = Feedback.query.filter_by(oficina_id=oficina_id).all()
    
    # Buscar perguntas personalizadas
    perguntas = FeedbackService.listar_perguntas(current_user.cliente_id, oficina_id)
    
    # Criar workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"Feedback - {oficina.titulo[:30]}"
    
    # Cabeçalho
    ws['A1'] = f"Relatório de Feedback - {oficina.titulo}"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A2'] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M')}"
    ws['A2'].font = Font(size=12)
    
    # Dados dos feedbacks
    row = 4
    ws[f'A{row}'] = "Usuário"
    ws[f'B{row}'] = "Avaliação"
    ws[f'C{row}'] = "Comentário"
    ws[f'D{row}'] = "Data"
    
    # Estilizar cabeçalho
    for col in ['A', 'B', 'C', 'D']:
        cell = ws[f'{col}{row}']
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    row += 1
    
    for fb in feedbacks:
        nome = fb.usuario.nome if fb.usuario else "N/A"
        ws[f'A{row}'] = nome
        ws[f'B{row}'] = fb.rating
        ws[f'C{row}'] = fb.comentario or ""
        ws[f'D{row}'] = fb.created_at.strftime('%d/%m/%Y %H:%M')
        row += 1
    
    # Adicionar seção de perguntas personalizadas
    if perguntas:
        row += 2
        ws[f'A{row}'] = "PERGUNTAS PERSONALIZADAS"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 2
        
        for pergunta in perguntas:
            ws[f'A{row}'] = f"Pergunta: {pergunta.titulo}"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            ws[f'A{row}'] = f"Tipo: {pergunta.tipo.value if hasattr(pergunta.tipo, 'value') else pergunta.tipo}"
            row += 1
            
            # Buscar respostas para esta pergunta
            respostas = []
            for fb in feedbacks:
                if hasattr(fb, 'respostas_personalizadas'):
                    for resposta in fb.respostas_personalizadas:
                        if resposta.pergunta_id == pergunta.id:
                            respostas.append({
                                'usuario': fb.usuario.nome if fb.usuario else "N/A",
                                'resposta': resposta.resposta,
                                'data': fb.created_at.strftime('%d/%m/%Y %H:%M')
                            })
            
            if respostas:
                ws[f'A{row}'] = "Usuário"
                ws[f'B{row}'] = "Resposta"
                ws[f'C{row}'] = "Data"
                
                # Estilizar cabeçalho das respostas
                for col in ['A', 'B', 'C']:
                    cell = ws[f'{col}{row}']
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
                
                row += 1
                
                for resposta in respostas:
                    ws[f'A{row}'] = resposta['usuario']
                    ws[f'B{row}'] = resposta['resposta']
                    ws[f'C{row}'] = resposta['data']
                    row += 1
            else:
                ws[f'A{row}'] = "Nenhuma resposta encontrada"
                row += 1
            
            row += 1
    
    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 20
    
    # Salvar em buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"feedback_{oficina.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
