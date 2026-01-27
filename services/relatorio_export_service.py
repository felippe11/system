"""
Serviço de Exportação de Relatórios
Fornece funcionalidades para exportar relatórios em PDF, XLSX, CSV e JSON
"""

import os
import json
from datetime import datetime
from io import BytesIO
from typing import Dict, List, Any, Optional
import logging

# Bibliotecas para exportação
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.linecharts import HorizontalLineChart
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.charts.barcharts import VerticalBarChart

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows

from extensions import db
from models.relatorio_bi import RelatorioBI, ExportacaoRelatorio
from services.bi_analytics_service import BIAnalyticsService

logger = logging.getLogger(__name__)

class RelatorioExportService:
    """Serviço para exportação de relatórios em diferentes formatos"""
    
    def __init__(self):
        self.bi_service = BIAnalyticsService()
        self.upload_folder = 'uploads/relatorios'
        self._ensure_upload_folder()
    
    def _ensure_upload_folder(self):
        """Garante que a pasta de upload existe"""
        if not os.path.exists(self.upload_folder):
            os.makedirs(self.upload_folder, exist_ok=True)
    
    def exportar_relatorio_pdf(self, relatorio_id: int, configuracao: Dict = None) -> str:
        """Exporta relatório para PDF"""
        try:
            relatorio = RelatorioBI.query.get_or_404(relatorio_id)
            
            # Gerar dados do relatório
            dados = self.bi_service.gerar_relatorio_personalizado(relatorio_id)
            
            # Configurar nome do arquivo
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"relatorio_{relatorio.nome}_{timestamp}.pdf"
            filepath = os.path.join(self.upload_folder, filename)
            
            # Criar PDF
            doc = SimpleDocTemplate(filepath, pagesize=A4)
            story = []
            
            # Estilos
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                spaceAfter=30,
                alignment=TA_CENTER,
                textColor=colors.darkblue
            )
            
            heading_style = ParagraphStyle(
                'CustomHeading',
                parent=styles['Heading2'],
                fontSize=14,
                spaceAfter=12,
                textColor=colors.darkblue
            )
            
            # Título do relatório
            story.append(Paragraph(relatorio.nome, title_style))
            story.append(Spacer(1, 12))
            
            # Informações do relatório
            info_data = [
                ['Tipo:', relatorio.tipo_relatorio.title()],
                ['Criado em:', relatorio.data_criacao.strftime('%d/%m/%Y %H:%M')],
                ['Período:', f"{relatorio.periodo_inicio or 'N/A'} a {relatorio.periodo_fim or 'N/A'}"],
                ['Status:', relatorio.status.title()]
            ]
            
            info_table = Table(info_data, colWidths=[2*inch, 4*inch])
            info_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                ('BACKGROUND', (1, 0), (1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(info_table)
            story.append(Spacer(1, 20))
            
            # Conteúdo baseado no tipo de relatório
            if relatorio.tipo_relatorio == 'executivo':
                story.extend(self._gerar_conteudo_executivo_pdf(dados, heading_style))
            elif relatorio.tipo_relatorio == 'operacional':
                story.extend(self._gerar_conteudo_operacional_pdf(dados, heading_style))
            elif relatorio.tipo_relatorio == 'financeiro':
                story.extend(self._gerar_conteudo_financeiro_pdf(dados, heading_style))
            elif relatorio.tipo_relatorio == 'qualidade':
                story.extend(self._gerar_conteudo_qualidade_pdf(dados, heading_style))
            
            # Construir PDF
            doc.build(story)
            
            # Registrar exportação
            self._registrar_exportacao(relatorio_id, 'pdf', filepath, len(story))
            
            return filepath
            
        except Exception as e:
            logger.error(f"Erro ao exportar relatório PDF: {str(e)}")
            raise
    
    def exportar_relatorio_xlsx(self, relatorio_id: int, configuracao: Dict = None) -> str:
        """Exporta relatório para XLSX"""
        try:
            relatorio = RelatorioBI.query.get_or_404(relatorio_id)
            
            # Gerar dados do relatório
            dados = self.bi_service.gerar_relatorio_personalizado(relatorio_id)
            
            # Configurar nome do arquivo
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"relatorio_{relatorio.nome}_{timestamp}.xlsx"
            filepath = os.path.join(self.upload_folder, filename)
            
            # Criar workbook
            wb = Workbook()
            
            # Remover planilha padrão
            wb.remove(wb.active)
            
            # Criar planilhas baseadas no tipo de relatório
            if relatorio.tipo_relatorio == 'executivo':
                self._criar_planilha_executivo_xlsx(wb, dados, relatorio)
            elif relatorio.tipo_relatorio == 'operacional':
                self._criar_planilha_operacional_xlsx(wb, dados, relatorio)
            elif relatorio.tipo_relatorio == 'financeiro':
                self._criar_planilha_financeiro_xlsx(wb, dados, relatorio)
            elif relatorio.tipo_relatorio == 'qualidade':
                self._criar_planilha_qualidade_xlsx(wb, dados, relatorio)
            
            # Salvar arquivo
            wb.save(filepath)
            
            # Registrar exportação
            self._registrar_exportacao(relatorio_id, 'xlsx', filepath, 0)
            
            return filepath
            
        except Exception as e:
            logger.error(f"Erro ao exportar relatório XLSX: {str(e)}")
            raise
    
    def exportar_relatorio_csv(self, relatorio_id: int, configuracao: Dict = None) -> str:
        """Exporta relatório para CSV"""
        try:
            relatorio = RelatorioBI.query.get_or_404(relatorio_id)
            
            # Gerar dados do relatório
            dados = self.bi_service.gerar_relatorio_personalizado(relatorio_id)
            
            # Configurar nome do arquivo
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"relatorio_{relatorio.nome}_{timestamp}.csv"
            filepath = os.path.join(self.upload_folder, filename)
            
            # Converter dados para DataFrame
            df = self._dados_para_dataframe(dados, relatorio.tipo_relatorio)
            
            # Salvar CSV
            df.to_csv(filepath, index=False, encoding='utf-8-sig')
            
            # Registrar exportação
            self._registrar_exportacao(relatorio_id, 'csv', filepath, len(df))
            
            return filepath
            
        except Exception as e:
            logger.error(f"Erro ao exportar relatório CSV: {str(e)}")
            raise
    
    def exportar_relatorio_json(self, relatorio_id: int, configuracao: Dict = None) -> str:
        """Exporta relatório para JSON"""
        try:
            relatorio = RelatorioBI.query.get_or_404(relatorio_id)
            
            # Gerar dados do relatório
            dados = self.bi_service.gerar_relatorio_personalizado(relatorio_id)
            
            # Configurar nome do arquivo
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"relatorio_{relatorio.nome}_{timestamp}.json"
            filepath = os.path.join(self.upload_folder, filename)
            
            # Preparar dados para exportação
            export_data = {
                'relatorio': {
                    'id': relatorio.id,
                    'nome': relatorio.nome,
                    'tipo': relatorio.tipo_relatorio,
                    'data_criacao': relatorio.data_criacao.isoformat(),
                    'periodo': {
                        'inicio': relatorio.periodo_inicio.isoformat() if relatorio.periodo_inicio else None,
                        'fim': relatorio.periodo_fim.isoformat() if relatorio.periodo_fim else None
                    }
                },
                'dados': dados,
                'exportado_em': datetime.now().isoformat()
            }
            
            # Salvar JSON
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(export_data, f, indent=2, ensure_ascii=False, default=str)
            
            # Registrar exportação
            self._registrar_exportacao(relatorio_id, 'json', filepath, 0)
            
            return filepath
            
        except Exception as e:
            logger.error(f"Erro ao exportar relatório JSON: {str(e)}")
            raise
    
    def exportar_dashboard_pdf(self, dashboard_id: int, configuracao: Dict = None) -> str:
        """Exporta dashboard para PDF"""
        try:
            from models.relatorio_bi import DashboardBI
            dashboard = DashboardBI.query.get_or_404(dashboard_id)
            
            # Configurar nome do arquivo
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"dashboard_{dashboard.nome}_{timestamp}.pdf"
            filepath = os.path.join(self.upload_folder, filename)
            
            # Criar PDF do dashboard
            doc = SimpleDocTemplate(filepath, pagesize=A4)
            story = []
            
            # Estilos
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'DashboardTitle',
                parent=styles['Heading1'],
                fontSize=20,
                spaceAfter=30,
                alignment=TA_CENTER,
                textColor=colors.darkblue
            )
            
            # Título
            story.append(Paragraph(f"Dashboard: {dashboard.nome}", title_style))
            story.append(Spacer(1, 20))
            
            # Conteúdo do dashboard será implementado baseado nos widgets
            story.append(Paragraph("Dashboard em desenvolvimento", styles['Normal']))
            
            # Construir PDF
            doc.build(story)
            
            return filepath
            
        except Exception as e:
            logger.error(f"Erro ao exportar dashboard PDF: {str(e)}")
            raise
    
    # Métodos auxiliares para PDF
    
    def _gerar_conteudo_executivo_pdf(self, dados: Dict, heading_style) -> List:
        """Gera conteúdo executivo para PDF"""
        story = []
        
        # KPIs principais
        story.append(Paragraph("KPIs Executivos", heading_style))
        
        if 'inscricoes_totais' in dados:
            kpi_data = [
                ['Métrica', 'Valor'],
                ['Inscrições Totais', str(dados.get('inscricoes_totais', 0))],
                ['Usuários Únicos', str(dados.get('usuarios_unicos', 0))],
                ['Taxa de Conversão', f"{dados.get('taxa_conversao', 0):.2f}%"],
                ['Receita Total', f"R$ {dados.get('receita_total', 0):,.2f}"],
                ['Ticket Médio', f"R$ {dados.get('ticket_medio', 0):,.2f}"],
                ['Taxa de Presença', f"{dados.get('taxa_presenca', 0):.2f}%"],
                ['Satisfação Média', f"{dados.get('satisfacao_media', 0):.2f}/5.0"],
                ['NPS', str(dados.get('nps', 0))]
            ]
            
            kpi_table = Table(kpi_data, colWidths=[3*inch, 2*inch])
            kpi_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(kpi_table)
            story.append(Spacer(1, 20))
        
        return story
    
    def _gerar_conteudo_operacional_pdf(self, dados: Dict, heading_style) -> List:
        """Gera conteúdo operacional para PDF"""
        story = []
        story.append(Paragraph("Análise Operacional", heading_style))
        story.append(Paragraph("Conteúdo operacional em desenvolvimento...", getSampleStyleSheet()['Normal']))
        return story
    
    def _gerar_conteudo_financeiro_pdf(self, dados: Dict, heading_style) -> List:
        """Gera conteúdo financeiro para PDF"""
        story = []
        story.append(Paragraph("Análise Financeira", heading_style))
        story.append(Paragraph("Conteúdo financeiro em desenvolvimento...", getSampleStyleSheet()['Normal']))
        return story
    
    def _gerar_conteudo_qualidade_pdf(self, dados: Dict, heading_style) -> List:
        """Gera conteúdo de qualidade para PDF"""
        story = []
        story.append(Paragraph("Análise de Qualidade", heading_style))
        story.append(Paragraph("Conteúdo de qualidade em desenvolvimento...", getSampleStyleSheet()['Normal']))
        return story
    
    # Métodos auxiliares para XLSX
    
    def _criar_planilha_executivo_xlsx(self, wb, dados: Dict, relatorio: RelatorioBI):
        """Cria planilha executiva no XLSX"""
        ws = wb.create_sheet("KPIs Executivos")
        
        # Título
        ws['A1'] = f"Relatório Executivo - {relatorio.nome}"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')
        
        # Cabeçalho
        ws['A3'] = "Métrica"
        ws['B3'] = "Valor"
        ws['C3'] = "Unidade"
        ws['D3'] = "Observações"
        
        # Estilo do cabeçalho
        for cell in ws['A3:D3']:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF")
        
        # Dados dos KPIs
        kpis = [
            ("Inscrições Totais", dados.get('inscricoes_totais', 0), "unidades", ""),
            ("Usuários Únicos", dados.get('usuarios_unicos', 0), "pessoas", ""),
            ("Taxa de Conversão", dados.get('taxa_conversao', 0), "%", ""),
            ("Receita Total", dados.get('receita_total', 0), "R$", ""),
            ("Ticket Médio", dados.get('ticket_medio', 0), "R$", ""),
            ("Taxa de Presença", dados.get('taxa_presenca', 0), "%", ""),
            ("Satisfação Média", dados.get('satisfacao_media', 0), "/5.0", ""),
            ("NPS", dados.get('nps', 0), "pontos", "")
        ]
        
        for i, (metrica, valor, unidade, obs) in enumerate(kpis, start=4):
            ws[f'A{i}'] = metrica
            ws[f'B{i}'] = valor
            ws[f'C{i}'] = unidade
            ws[f'D{i}'] = obs
        
        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 30
    
    def _criar_planilha_operacional_xlsx(self, wb, dados: Dict, relatorio: RelatorioBI):
        """Cria planilha operacional no XLSX"""
        ws = wb.create_sheet("Análise Operacional")
        ws['A1'] = f"Análise Operacional - {relatorio.nome}"
        ws['A1'].font = Font(size=16, bold=True)
        # Implementar conteúdo operacional
        pass
    
    def _criar_planilha_financeiro_xlsx(self, wb, dados: Dict, relatorio: RelatorioBI):
        """Cria planilha financeira no XLSX"""
        ws = wb.create_sheet("Análise Financeira")
        ws['A1'] = f"Análise Financeira - {relatorio.nome}"
        ws['A1'].font = Font(size=16, bold=True)
        # Implementar conteúdo financeiro
        pass
    
    def _criar_planilha_qualidade_xlsx(self, wb, dados: Dict, relatorio: RelatorioBI):
        """Cria planilha de qualidade no XLSX"""
        ws = wb.create_sheet("Análise de Qualidade")
        ws['A1'] = f"Análise de Qualidade - {relatorio.nome}"
        ws['A1'].font = Font(size=16, bold=True)
        # Implementar conteúdo de qualidade
        pass
    
    def _dados_para_dataframe(self, dados: Dict, tipo_relatorio: str) -> pd.DataFrame:
        """Converte dados para DataFrame do pandas"""
        if tipo_relatorio == 'executivo':
            kpis = [
                {"Métrica": "Inscrições Totais", "Valor": dados.get('inscricoes_totais', 0)},
                {"Métrica": "Usuários Únicos", "Valor": dados.get('usuarios_unicos', 0)},
                {"Métrica": "Taxa de Conversão", "Valor": dados.get('taxa_conversao', 0)},
                {"Métrica": "Receita Total", "Valor": dados.get('receita_total', 0)},
                {"Métrica": "Ticket Médio", "Valor": dados.get('ticket_medio', 0)},
                {"Métrica": "Taxa de Presença", "Valor": dados.get('taxa_presenca', 0)},
                {"Métrica": "Satisfação Média", "Valor": dados.get('satisfacao_media', 0)},
                {"Métrica": "NPS", "Valor": dados.get('nps', 0)}
            ]
            return pd.DataFrame(kpis)
        else:
            # Para outros tipos, criar DataFrame básico
            return pd.DataFrame([{"Dados": "Em desenvolvimento", "Valor": 0}])
    
    def _registrar_exportacao(self, relatorio_id: int, formato: str, filepath: str, registros: int):
        """Registra exportação no banco de dados"""
        try:
            exportacao = ExportacaoRelatorio(
                relatorio_id=relatorio_id,
                usuario_id=1,  # TODO: Obter do contexto atual
                formato=formato,
                status='concluido',
                arquivo_path=filepath,
                tamanho_arquivo=os.path.getsize(filepath) if os.path.exists(filepath) else 0,
                data_conclusao=datetime.utcnow()
            )
            
            db.session.add(exportacao)
            db.session.commit()
            
        except Exception as e:
            logger.error(f"Erro ao registrar exportação: {str(e)}")
            db.session.rollback()
    
    def obter_historico_exportacoes(self, relatorio_id: int = None, usuario_id: int = None) -> List[Dict]:
        """Obtém histórico de exportações"""
        try:
            query = ExportacaoRelatorio.query
            
            if relatorio_id:
                query = query.filter_by(relatorio_id=relatorio_id)
            if usuario_id:
                query = query.filter_by(usuario_id=usuario_id)
            
            exportacoes = query.order_by(desc(ExportacaoRelatorio.data_inicio)).limit(50).all()
            
            return [{
                'id': exp.id,
                'relatorio_id': exp.relatorio_id,
                'formato': exp.formato,
                'status': exp.status,
                'arquivo_path': exp.arquivo_path,
                'tamanho_arquivo': exp.tamanho_arquivo,
                'data_inicio': exp.data_inicio.isoformat(),
                'data_conclusao': exp.data_conclusao.isoformat() if exp.data_conclusao else None,
                'tempo_processamento': exp.tempo_processamento
            } for exp in exportacoes]
            
        except Exception as e:
            logger.error(f"Erro ao obter histórico de exportações: {str(e)}")
            return []
    
    def limpar_arquivos_antigos(self, dias_antigos: int = 30):
        """Remove arquivos de exportação antigos"""
        try:
            from datetime import timedelta
            data_limite = datetime.now() - timedelta(days=dias_antigos)
            
            # Buscar exportações antigas
            exportacoes_antigas = ExportacaoRelatorio.query.filter(
                ExportacaoRelatorio.data_inicio < data_limite
            ).all()
            
            arquivos_removidos = 0
            for exportacao in exportacoes_antigas:
                if exportacao.arquivo_path and os.path.exists(exportacao.arquivo_path):
                    try:
                        os.remove(exportacao.arquivo_path)
                        arquivos_removidos += 1
                    except Exception as e:
                        logger.warning(f"Erro ao remover arquivo {exportacao.arquivo_path}: {str(e)}")
                
                # Remover registro do banco
                db.session.delete(exportacao)
            
            db.session.commit()
            logger.info(f"Removidos {arquivos_removidos} arquivos antigos")
            
        except Exception as e:
            logger.error(f"Erro ao limpar arquivos antigos: {str(e)}")
            db.session.rollback()
